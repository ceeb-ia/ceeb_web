import json
import importlib.util
import tempfile
import unittest
from pathlib import Path

from calendaritzacions.engine.base import EngineResult
from calendaritzacions.engine.config import EngineConfig
from calendaritzacions.engine.registry import get_engine
from calendaritzacions.engine.variants.resource_solver.service import ResourceSolverEngine

HAS_PANDAS = importlib.util.find_spec("pandas") is not None
HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None

if HAS_PANDAS:
    import pandas as pd

if HAS_OPENPYXL:
    from openpyxl import load_workbook


def _write_input(path: Path) -> None:
    df = pd.DataFrame(
        [
            {
                "Id": "T1",
                "Nom": "Equip 1",
                "Entitat": "Club 1",
                "Nom Lliga": "Lliga",
                "Nivell": "Nivell A",
                "Num. sorteig": "",
                "Dia partit": "Divendres",
                "Horari partit": "18:00",
                "Pista joc": "Pavello",
                "Categoria": "Cat",
            }
        ]
    )
    df.to_excel(path, index=False)


class ResourceSolverServiceTests(unittest.TestCase):
    @unittest.skipUnless(HAS_PANDAS and HAS_OPENPYXL, "pandas/openpyxl not installed")
    def test_run_returns_engine_result_with_audit_paths(self):
        with tempfile.TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            _write_input(input_path)

            result = ResourceSolverEngine().run(str(input_path), EngineConfig(name="resource_solver"))

            self.assertIsInstance(result, EngineResult)
            self.assertTrue(Path(result.output_path).exists())
            self.assertEqual(Path(result.output_path).suffix, ".xlsx")
            self.assertIn("resource_solution", result.audit_paths)
            self.assertIn("solver_explanations", result.audit_paths)
            self.assertIn("resource_solver_result", result.audit_paths)
            self.assertIn("resource_solver_final_plots", result.audit_paths)
            self.assertIn("input_demand", result.audit_paths)
            self.assertIn("input_validation", result.audit_paths)
            self.assertTrue(Path(result.audit_paths["solver_explanations"]).exists())
            final_plots = json.loads(Path(result.audit_paths["resource_solver_final_plots"]).read_text(encoding="utf-8"))
            self.assertEqual(final_plots["artifact_type"], "resource_solver_final_plots")
            self.assertTrue(final_plots["plots"])
            payload = json.loads(Path(result.audit_paths["resource_solver_result"]).read_text(encoding="utf-8"))
            self.assertIn(payload["status"], {"OPTIMAL", "FEASIBLE"})
            workbook = load_workbook(result.output_path, read_only=True)
            self.assertIn("Resum", workbook.sheetnames)
            assignment_sheet = workbook["Lliga"]
            headers = [cell.value for cell in assignment_sheet[2]]
            self.assertIn("Diferències jornades", headers)
            self.assertNotIn("J1", headers)
            group_col = headers.index("Grup") + 1
            self.assertEqual(assignment_sheet.cell(row=3, column=group_col).value, "G1")
            self.assertTrue(
                any(log.startswith("pre-solver competicio: Nom Lliga: Lliga |") for log in result.logs)
            )
            self.assertTrue(
                any(log.startswith("post-solver competicio: Nom Lliga: Lliga |") for log in result.logs)
            )
            self.assertTrue(any(log.startswith("resource_solver: status=") for log in result.logs))

    def test_registry_exposes_resource_solver_and_legacy(self):
        self.assertIsInstance(get_engine("resource_solver"), ResourceSolverEngine)
        self.assertIsInstance(get_engine("resource_solver_linkage"), ResourceSolverEngine)
        self.assertTrue(callable(get_engine("legacy")))

    def test_process_calendarization_can_route_to_resource_solver(self):
        if importlib.util.find_spec("asgiref") is None:
            self.skipTest("asgiref is not installed in this environment")
        if not (HAS_PANDAS and HAS_OPENPYXL):
            self.skipTest("pandas/openpyxl not installed")

        from calendaritzacions.application.use_cases import process_calendarization

        with tempfile.TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            _write_input(input_path)

            output_path, logs = process_calendarization(
                str(input_path),
                return_logs=True,
                engine_name="resource_solver",
            )

            self.assertTrue(Path(output_path).exists())
            self.assertEqual(Path(output_path).suffix, ".xlsx")
            self.assertTrue(any(log.startswith("resource_solver: status=") for log in logs))

    def test_process_calendarization_can_return_engine_artifacts(self):
        if importlib.util.find_spec("asgiref") is None:
            self.skipTest("asgiref is not installed in this environment")
        if not (HAS_PANDAS and HAS_OPENPYXL):
            self.skipTest("pandas/openpyxl not installed")

        from calendaritzacions.application.use_cases import process_calendarization

        with tempfile.TemporaryDirectory() as directory:
            input_path = Path(directory) / "input.xlsx"
            _write_input(input_path)

            output_path, logs, audit_paths, kpis_path = process_calendarization(
                str(input_path),
                return_logs=True,
                return_artifacts=True,
                engine_name="resource_solver",
            )

            self.assertTrue(Path(output_path).exists())
            self.assertEqual(Path(output_path).suffix, ".xlsx")
            self.assertIn("resource_solution", audit_paths)
            self.assertIn("solver_explanations", audit_paths)
            self.assertIn("resource_solver_result", audit_paths)
            self.assertIn("resource_solver_final_plots", audit_paths)
            self.assertIn("input_demand", audit_paths)
            self.assertIn("input_validation", audit_paths)
            self.assertEqual(kpis_path, "")
            self.assertTrue(any(log.startswith("input: teams=") for log in logs))
            self.assertTrue(any(log.startswith("resource_solver: status=") for log in logs))


if __name__ == "__main__":
    unittest.main()
