import json
import re
from io import BytesIO
from datetime import date, timedelta
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.sessions.backends.db import SessionStore
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from ceeb_web.auth_groups import GLOBAL_AUTH_GROUPS

from . import live_cache
from .access import user_has_competicio_capability
from .forms import CompeticioAparellForm
from .models import (
    Competicio,
    Equip,
    EquipContext,
    GrupCompeticio,
    Inscripcio,
    InscripcioEquipAssignacio,
    InscripcioMedia,
)
from .models_judging import (
    JudgeConversation,
    JudgeConversationMessage,
    JudgeDeviceToken,
    PublicLiveToken,
)
from .models_classificacions import ClassificacioConfig, ClassificacioTemplateGlobal
from .models_rotacions import (
    RotacioAssignacio,
    RotacioAssignacioGrup,
    RotacioEstacio,
    RotacioFranja,
)
from .models_scoring import (
    ScoringSchema,
    ScoreEntry,
    ScoreEntryVideo,
    ScoreEntryVideoEvent,
    TeamScoreEntry,
    TeamScoreEntryVideo,
    TeamScoreEntryVideoEvent,
)
from .models_trampoli import (
    Aparell,
    CompeticioAparell,
    InscripcioAparellExclusio,
)
from .models import CompeticioMembership
from .scoring_engine import ScoringEngine
from .views import (
    apply_inscripcions_history_snapshot,
    _split_custom_sort_tokens,
    capture_inscripcions_history_snapshot,
    renumber_groups_for_competicio,
    get_competicio_custom_sort_rank_map,
    sort_records_by_field_stable,
)
from .views_classificacions import (
    _build_scoreable_meta_for_schema,
    _normalize_particions_schema,
    _scoreable_codes_by_app_id,
    _schema_to_template_schema,
    _template_schema_to_competicio_schema,
    _validate_schema_for_competicio,
    _validate_particions_schema,
)
from .services.services_classificacions_2 import DEFAULT_SCHEMA, compute_classificacio
from .templatetags.competicio_extras import (
    DEFAULT_COMPETITION_BACKGROUND,
    get_competicio_background_url_from_request,
)


class _BaseTrampoliDataMixin:
    def _create_competicio(self, nom="Comp"):
        return Competicio.objects.create(
            nom=nom,
            tipus=Competicio.Tipus.TRAMPOLI,
        )

    def _ensure_default_aparell_owner(self):
        owner = getattr(self, "_default_aparell_owner", None)
        if owner is not None:
            return owner
        User = get_user_model()
        owner = User.objects.create_user(
            username=f"ap_owner_{self.__class__.__name__.lower()}",
            password="testpass123",
            email=f"ap-owner-{self.__class__.__name__.lower()}@example.com",
        )
        self._default_aparell_owner = owner
        return owner

    def _create_aparell(self, codi, nom, owner=None):
        owner = owner or getattr(self, "user", None) or self._ensure_default_aparell_owner()
        return Aparell.objects.create(codi=codi, nom=nom, actiu=True, created_by=owner)

    def _create_comp_aparell(self, competicio, aparell, ordre=1, actiu=True):
        return CompeticioAparell.objects.create(
            competicio=competicio,
            aparell=aparell,
            ordre=ordre,
            actiu=actiu,
        )

    def _create_inscripcio(self, competicio, nom, ordre=1, grup=1):
        return Inscripcio.objects.create(
            competicio=competicio,
            nom_i_cognoms=nom,
            ordre_sortida=ordre,
            grup=grup,
        )


class CompeticioBackgroundTemplateTagTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def _request_with_kwargs(self, kwargs, path="/competicio/1/inscripcions/"):
        request = self.factory.get(path)
        request.resolver_match = SimpleNamespace(kwargs=kwargs)
        return request

    def test_returns_mapped_background_for_existing_type_image(self):
        comp = Competicio.objects.create(
            nom="Comp fons natacio",
            tipus=Competicio.Tipus.NATACIO,
        )

        result = get_competicio_background_url_from_request(
            self._request_with_kwargs({"pk": comp.id})
        )

        self.assertTrue(result.endswith("/static/images/natacio.jpg"))

    def test_falls_back_to_default_when_mapped_image_file_is_missing(self):
        comp = Competicio.objects.create(
            nom="Comp fons artistica",
            tipus=Competicio.Tipus.ARTISTICA,
        )

        result = get_competicio_background_url_from_request(
            self._request_with_kwargs({"pk": comp.id})
        )

        self.assertTrue(result.endswith(f"/static/{DEFAULT_COMPETITION_BACKGROUND}"))

    def test_falls_back_to_default_when_no_active_competicio_exists(self):
        result = get_competicio_background_url_from_request(
            self._request_with_kwargs({}, path="/competicions/created/")
        )

        self.assertTrue(result.endswith(f"/static/{DEFAULT_COMPETITION_BACKGROUND}"))

    def test_falls_back_to_default_for_non_competition_route_even_with_pk(self):
        comp = Competicio.objects.create(
            nom="Comp no relacionada",
            tipus=Competicio.Tipus.NATACIO,
        )

        result = get_competicio_background_url_from_request(
            self._request_with_kwargs({"pk": comp.id}, path="/altres/modul/1/")
        )

        self.assertTrue(result.endswith(f"/static/{DEFAULT_COMPETITION_BACKGROUND}"))

    def test_base_template_injects_competicio_background_for_competition_route(self):
        comp = Competicio.objects.create(
            nom="Comp render natacio",
            tipus=Competicio.Tipus.NATACIO,
        )
        User = get_user_model()
        user = User.objects.create_user(
            username="bg_route_user",
            password="testpass123",
            email="bg-route@example.com",
        )
        CompeticioMembership.objects.create(
            user=user,
            competicio=comp,
            role=CompeticioMembership.Role.OWNER,
            is_active=True,
        )
        self.client.force_login(user)

        response = self.client.get(reverse("inscripcions_list", kwargs={"pk": comp.id}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "--ceeb-page-background-image: url('/static/images/natacio.jpg');",
        )

    def test_base_template_uses_default_background_outside_competition_routes(self):
        User = get_user_model()
        user = User.objects.create_user(
            username="bg_default_user",
            password="testpass123",
            email="bg-default@example.com",
        )
        self.client.force_login(user)

        response = self.client.get(reverse("created"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f"--ceeb-page-background-image: url('/static/{DEFAULT_COMPETITION_BACKGROUND}');",
        )


class ScoringEngineAliasResolutionTests(TestCase):
    def test_computed_var_alias_can_be_used_before_definition_order(self):
        schema = {
            "fields": [
                {"code": "A", "var": "a", "type": "number"},
            ],
            "computed": [
                # Deliberadament abans que C1 per provar ordre topo + alias var.
                {"code": "C2", "formula": "u + 2"},
                {"code": "C1", "var": "u", "formula": "a + 1"},
                {"code": "TOTAL", "formula": "C2"},
            ],
        }

        result = ScoringEngine(schema).compute({"A": 3})

        self.assertEqual(result.outputs.get("C1"), 4)
        self.assertEqual(result.outputs.get("C2"), 6)
        self.assertEqual(float(result.total), 6.0)


class CustomSortOrderFallbackTests(TestCase):
    def _make_row(self, rid, categoria):
        return SimpleNamespace(id=rid, categoria=categoria, extra={})

    def test_custom_order_is_applied_before_default_fallback(self):
        comp = Competicio.objects.create(
            nom="Comp custom sort",
            tipus=Competicio.Tipus.TRAMPOLI,
            inscripcions_view={
                "custom_sort_orders": {
                    "categoria": ["alevi", "infantil", "cadet"],
                }
            },
        )

        rows = [
            self._make_row(1, "cadet"),
            self._make_row(2, "junior"),
            self._make_row(3, "infantil"),
            self._make_row(4, "alevi"),
            self._make_row(5, "benjami"),
        ]

        rank = get_competicio_custom_sort_rank_map(comp, "categoria")
        ordered = sort_records_by_field_stable(
            rows,
            "categoria",
            descending=False,
            custom_rank_map=rank,
        )

        self.assertEqual(
            [r.categoria for r in ordered],
            ["alevi", "infantil", "cadet", "benjami", "junior"],
        )

    def test_without_custom_order_fallback_matches_default_sort(self):
        comp = Competicio.objects.create(
            nom="Comp no custom sort",
            tipus=Competicio.Tipus.TRAMPOLI,
        )

        rows = [
            self._make_row(1, "zeta"),
            self._make_row(2, "beta"),
            self._make_row(3, "alfa"),
        ]

        rank = get_competicio_custom_sort_rank_map(comp, "categoria")
        ordered = sort_records_by_field_stable(
            rows,
            "categoria",
            descending=False,
            custom_rank_map=rank,
        )

        self.assertEqual([r.categoria for r in ordered], ["alfa", "beta", "zeta"])

    def test_split_custom_sort_tokens_separates_active_and_stale(self):
        active, stale = _split_custom_sort_tokens(
            ["alevi", "fantasma", "CADET", "fantasma"],
            {"alevi", "cadet"},
        )
        self.assertEqual(active, ["alevi", "CADET"])
        self.assertEqual(stale, ["fantasma"])


class InscripcionsSortFlowTests(TestCase):
    def setUp(self):
        self.comp = Competicio.objects.create(
            nom="Comp sort flow",
            tipus=Competicio.Tipus.TRAMPOLI,
        )
        User = get_user_model()
        self.user = User.objects.create_user(
            username="sort_editor_user",
            password="testpass123",
            email="sort-editor@example.com",
        )
        self.readonly_user = User.objects.create_user(
            username="sort_readonly_user",
            password="testpass123",
            email="sort-readonly@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.readonly_user,
            competicio=self.comp,
            role=CompeticioMembership.Role.READONLY,
            is_active=True,
        )
        self.client.force_login(self.user)

    def _post_json(self, url_name, payload):
        url = reverse(url_name, kwargs={"pk": self.comp.id})
        return self.client.post(
            url,
            data=json.dumps(payload),
            content_type="application/json",
        )

    def _post_history(self, direction):
        url_name = "inscripcions_history_undo" if direction == "undo" else "inscripcions_history_redo"
        url = reverse(url_name, kwargs={"pk": self.comp.id})
        return self.client.post(url, data="{}", content_type="application/json")

    def _toggle_competition_tail(self, enabled, **overrides):
        payload = {
            "enabled": enabled,
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }
        payload.update(overrides)
        return self._post_json("inscripcions_sort_competition_tail_toggle", payload)

    def _groups_payload(self, **overrides):
        payload = {
            "resolution_mode": "auto",
            "strategy": "count",
            "group_count": 1,
            "preview_only": True,
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }
        payload.update(overrides)
        return payload

    def test_inscripcions_list_group_column_uses_group_label_with_fallbacks(self):
        named_group = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Participant Final",
            grup=2,
            ordre_sortida=1,
        )
        fallback_group = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Participant Grup 3",
            grup=3,
            ordre_sortida=2,
        )
        no_group = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Participant Sense Grup",
            ordre_sortida=3,
        )

        group = GrupCompeticio.objects.get(competicio=self.comp, display_num=2)
        group.nom = "Final"
        group.save(update_fields=["nom"])

        self.comp.inscripcions_view = {
            "table_columns": ["nom_i_cognoms", "grup"],
        }
        self.comp.save(update_fields=["inscripcions_view"])

        response = self.client.get(reverse("inscripcions_list", kwargs={"pk": self.comp.id}))
        self.assertEqual(response.status_code, 200)

        body = response.content.decode("utf-8")
        self.assertRegex(
            body,
            re.compile(
                r"Participant Final.*?<td class=\"[^\"]*\">\s*Final\s*</td>",
                re.S,
            ),
        )
        self.assertRegex(
            body,
            re.compile(
                r"Participant Grup 3.*?<td class=\"[^\"]*\">\s*Grup 3\s*</td>",
                re.S,
            ),
        )
        self.assertRegex(
            body,
            re.compile(
                r"Participant Sense Grup.*?<td class=\"[^\"]*\">\s*-\s*</td>",
                re.S,
            ),
        )

    def test_inscripcions_list_renders_explicit_sort_scope_labels(self):
        response = self.client.get(reverse("inscripcions_list", kwargs={"pk": self.comp.id}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Totes les inscripcions filtrades")
        self.assertContains(response, "Dins de cada pestanya activa")
        self.assertContains(response, "Dins de cada grup")
        self.assertContains(response, "Nomes un grup numeric concret")
        self.assertContains(response, "incloent membres fora del filtre actual")

    def test_inscripcions_list_hides_show_real_order_button_but_keeps_other_group_actions(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Participant Grup 1",
            grup=1,
            ordre_sortida=1,
        )

        response = self.client.get(reverse("inscripcions_list", kwargs={"pk": self.comp.id}))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Mostrar ordre real")
        self.assertContains(response, "Veure ordre competició")
        self.assertContains(response, "Desar ordre competició")

    def test_inscripcions_list_renders_competition_order_tail_toggle_state(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha",
            grup=1,
            ordre_sortida=1,
            ordre_competicio=2,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta",
            grup=1,
            ordre_sortida=2,
            ordre_competicio=1,
        )

        apply_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "nom_i_cognoms",
                "sort_dir": "asc",
                "scope": "all_groups",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(apply_resp.status_code, 200)
        self.assertTrue(apply_resp.json().get("ok"))

        toggle_resp = self._toggle_competition_tail(True)
        self.assertEqual(toggle_resp.status_code, 200)
        self.assertTrue(toggle_resp.json().get("ok"))

        response = self.client.get(reverse("inscripcions_list", kwargs={"pk": self.comp.id}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ordre de competici")
        self.assertContains(response, "Criteri final actiu")

    def test_inscripcions_list_renders_unified_groups_panel(self):
        response = self.client.get(reverse("inscripcions_list", kwargs={"pk": self.comp.id}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Crear grups amb resolucio automatica")
        self.assertContains(response, "Divisions actives:")
        self.assertContains(response, "Ordenacions detectades:")
        self.assertContains(response, "Resolucio final:")
        self.assertNotContains(response, "Base de creacio:")
        self.assertNotContains(response, "Crear grups per pestanyes")
        self.assertNotContains(response, "Crear grups per ordenacio")

    def test_sort_apply_tab_preserves_tab_block_order(self):
        beta_1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="David",
            categoria="Beta",
            ordre_sortida=1,
        )
        beta_2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Carla",
            categoria="Beta",
            ordre_sortida=2,
        )
        alpha_1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Biel",
            categoria="Alpha",
            ordre_sortida=3,
        )
        alpha_2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Anna",
            categoria="Alpha",
            ordre_sortida=4,
        )

        response = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "nom_i_cognoms",
                "sort_dir": "asc",
                "scope": "tab",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": ["categoria"],
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json().get("ok"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [beta_2.id, beta_1.id, alpha_2.id, alpha_1.id])

    def test_sort_apply_all_can_reorder_tab_blocks(self):
        beta_1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="David",
            categoria="Beta",
            ordre_sortida=1,
        )
        beta_2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Carla",
            categoria="Beta",
            ordre_sortida=2,
        )
        alpha_1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Biel",
            categoria="Alpha",
            ordre_sortida=3,
        )
        alpha_2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Anna",
            categoria="Alpha",
            ordre_sortida=4,
        )

        response = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "nom_i_cognoms",
                "sort_dir": "asc",
                "scope": "all",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": ["categoria"],
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json().get("ok"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [alpha_2.id, alpha_1.id, beta_2.id, beta_1.id])

    def test_sort_apply_group_reorders_full_group_even_outside_filter(self):
        visible = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Zulu",
            entitat="Club A",
            ordre_sortida=1,
            grup=1,
        )
        hidden = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha",
            entitat="Club B",
            ordre_sortida=2,
            grup=1,
        )
        other = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Resta",
            ordre_sortida=3,
            grup=2,
        )

        response = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "nom_i_cognoms",
                "sort_dir": "asc",
                "scope": "group",
                "group_num": 1,
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": "Club A"},
                "group_by": [],
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json().get("ok"))

        visible.refresh_from_db()
        hidden.refresh_from_db()
        other.refresh_from_db()
        self.assertEqual(hidden.ordre_sortida, 1)
        self.assertEqual(visible.ordre_sortida, 2)
        self.assertEqual(other.ordre_sortida, 3)

    def test_sort_apply_all_groups_reorders_full_visible_groups_only(self):
        g1_visible = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Zulu",
            entitat="Club A",
            ordre_sortida=1,
            grup=1,
        )
        g1_hidden = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha",
            entitat="Club B",
            ordre_sortida=2,
            grup=1,
        )
        g2_visible = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Yara",
            entitat="Club A",
            ordre_sortida=3,
            grup=2,
        )
        g2_hidden = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta",
            entitat="Club B",
            ordre_sortida=4,
            grup=2,
        )
        g3_first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Omega",
            entitat="Club C",
            ordre_sortida=5,
            grup=3,
        )
        g3_second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Gamma",
            entitat="Club D",
            ordre_sortida=6,
            grup=3,
        )

        response = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "nom_i_cognoms",
                "sort_dir": "asc",
                "scope": "all_groups",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": "Club A"},
                "group_by": [],
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json().get("ok"))

        g1_visible.refresh_from_db()
        g1_hidden.refresh_from_db()
        g2_visible.refresh_from_db()
        g2_hidden.refresh_from_db()
        g3_first.refresh_from_db()
        g3_second.refresh_from_db()

        self.assertEqual(g1_hidden.ordre_sortida, 1)
        self.assertEqual(g1_visible.ordre_sortida, 2)
        self.assertEqual(g2_hidden.ordre_sortida, 3)
        self.assertEqual(g2_visible.ordre_sortida, 4)
        self.assertEqual(g3_first.ordre_sortida, 5)
        self.assertEqual(g3_second.ordre_sortida, 6)

    def test_group_competition_order_preview_returns_saved_order_for_full_group(self):
        first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Participant 1",
            entitat="Club A",
            grup=2,
            ordre_sortida=1,
        )
        second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Participant 2",
            entitat="Club B",
            grup=2,
            ordre_sortida=2,
        )
        third = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Participant 3",
            entitat="Club C",
            grup=2,
            ordre_sortida=3,
        )
        group = GrupCompeticio.objects.get(competicio=self.comp, display_num=2)
        group.nom = "Final"
        group.save(update_fields=["nom"])
        Inscripcio.objects.filter(pk=first.pk).update(ordre_competicio=2)
        Inscripcio.objects.filter(pk=second.pk).update(ordre_competicio=1)
        Inscripcio.objects.filter(pk=third.pk).update(ordre_competicio=3)

        response = self._post_json(
            "inscripcions_group_competition_order_preview",
            {"group_num": 2},
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data.get("ok"))
        self.assertEqual(data.get("group_label"), "Final")
        self.assertEqual(data.get("total_count"), 3)
        self.assertTrue(data.get("can_edit"))
        self.assertEqual(
            data.get("rows"),
            [
                {
                    "id": second.id,
                    "label": "Participant 2",
                    "secondary_label": "Club B",
                    "saved_order": 1,
                },
                {
                    "id": first.id,
                    "label": "Participant 1",
                    "secondary_label": "Club A",
                    "saved_order": 2,
                },
                {
                    "id": third.id,
                    "label": "Participant 3",
                    "secondary_label": "Club C",
                    "saved_order": 3,
                },
            ],
        )

    def test_group_competition_order_preview_falls_back_to_group_number_when_unnamed(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Sense nom de grup",
            grup=3,
            ordre_sortida=1,
        )

        response = self._post_json(
            "inscripcions_group_competition_order_preview",
            {"group_num": 3},
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data.get("ok"))
        self.assertEqual(data.get("group_label"), "Grup 3")

    def test_group_competition_order_preview_allows_readonly_and_save_remains_protected(self):
        first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Readonly 1",
            grup=1,
            ordre_sortida=1,
        )
        second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Readonly 2",
            grup=1,
            ordre_sortida=2,
        )

        self.client.force_login(self.readonly_user)

        preview_res = self._post_json(
            "inscripcions_group_competition_order_preview",
            {"group_num": 1},
        )
        self.assertEqual(preview_res.status_code, 200)
        self.assertFalse(preview_res.json().get("can_edit"))

        save_url = reverse("inscripcions_save_group_competition_order", kwargs={"pk": self.comp.id})
        save_res = self.client.post(
            save_url,
            data=json.dumps({"group_num": 1, "ids": [second.id, first.id]}),
            content_type="application/json",
        )
        self.assertEqual(save_res.status_code, 403)

    def test_sort_apply_reapplying_existing_criterion_keeps_priority(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            entitat="B",
            categoria="beta",
            ordre_sortida=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            entitat="A",
            categoria="alpha",
            ordre_sortida=2,
        )
        i3 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I3",
            entitat="A",
            categoria="beta",
            ordre_sortida=3,
        )
        i4 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I4",
            entitat="B",
            categoria="alpha",
            ordre_sortida=4,
        )

        base_payload = {
            "scope": "all",
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }

        r1 = self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "entitat", "sort_dir": "asc"},
        )
        self.assertEqual(r1.status_code, 200)
        self.assertTrue(r1.json().get("ok"))

        r2 = self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "categoria", "sort_dir": "asc"},
        )
        self.assertEqual(r2.status_code, 200)
        self.assertTrue(r2.json().get("ok"))

        # Reapliquem el primer criteri amb una direccio diferent.
        # Ha de mantenir la prioritat original (primer criteri = mes important).
        r3 = self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "entitat", "sort_dir": "desc"},
        )
        self.assertEqual(r3.status_code, 200)
        self.assertTrue(r3.json().get("ok"))
        self.assertEqual(r3.json().get("stack_count"), 2)

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [i4.id, i1.id, i2.id, i3.id])

    def test_custom_sort_save_reapplies_active_stack_immediately(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            categoria="B",
            ordre_sortida=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            categoria="C",
            ordre_sortida=2,
        )
        i3 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I3",
            categoria="A",
            ordre_sortida=3,
        )

        base_payload = {
            "scope": "all",
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }

        r_apply = self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "categoria", "sort_dir": "custom"},
        )
        self.assertEqual(r_apply.status_code, 200)
        self.assertTrue(r_apply.json().get("ok"))

        r_custom = self._post_json(
            "inscripcions_sort_custom_save",
            {
                "sort_key": "categoria",
                "order": ["C", "B", "A"],
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
                "preserve_missing_context": True,
            },
        )
        self.assertEqual(r_custom.status_code, 200)
        data = r_custom.json()
        self.assertTrue(data.get("ok"))
        self.assertTrue(data.get("reapplied"))
        self.assertEqual(data.get("stack_count"), 1)

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [i2.id, i1.id, i3.id])

    def test_competition_order_tail_toggle_requires_active_stack(self):
        first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha",
            grup=1,
            ordre_sortida=1,
            ordre_competicio=2,
        )
        second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta",
            grup=1,
            ordre_sortida=2,
            ordre_competicio=1,
        )

        response = self._toggle_competition_tail(True)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data.get("ok"))
        self.assertFalse(data.get("applied"))
        self.assertEqual(data.get("reason"), "no_stack")
        self.assertFalse(data.get("competition_order_tail"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [first.id, second.id])

    def test_competition_order_tail_is_less_prioritary_than_active_stack(self):
        g1_first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha",
            grup=1,
            ordre_sortida=1,
            ordre_competicio=2,
        )
        g1_second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta",
            grup=1,
            ordre_sortida=2,
            ordre_competicio=1,
        )
        g2_first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Charlie",
            grup=2,
            ordre_sortida=3,
            ordre_competicio=2,
        )
        g2_second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Delta",
            grup=2,
            ordre_sortida=4,
            ordre_competicio=1,
        )

        apply_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "nom_i_cognoms",
                "sort_dir": "asc",
                "scope": "all_groups",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(apply_resp.status_code, 200)
        self.assertTrue(apply_resp.json().get("ok"))

        toggle_resp = self._toggle_competition_tail(True)
        self.assertEqual(toggle_resp.status_code, 200)
        data = toggle_resp.json()
        self.assertTrue(data.get("ok"))
        self.assertTrue(data.get("applied"))
        self.assertTrue(data.get("competition_order_tail"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [g1_first.id, g1_second.id, g2_first.id, g2_second.id])

    def test_competition_order_tail_toggle_applies_and_persists_to_ordre_sortida(self):
        g1_first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=1,
            ordre_competicio=2,
        )
        g1_second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=2,
            ordre_competicio=1,
        )
        g2_first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Charlie",
            categoria="Mateixa",
            grup=2,
            ordre_sortida=3,
            ordre_competicio=2,
        )
        g2_second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Delta",
            categoria="Mateixa",
            grup=2,
            ordre_sortida=4,
            ordre_competicio=1,
        )

        apply_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "categoria",
                "sort_dir": "asc",
                "scope": "all_groups",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(apply_resp.status_code, 200)
        self.assertTrue(apply_resp.json().get("ok"))

        toggle_resp = self._toggle_competition_tail(True)
        self.assertEqual(toggle_resp.status_code, 200)
        data = toggle_resp.json()
        self.assertTrue(data.get("ok"))
        self.assertTrue(data.get("applied"))
        self.assertTrue(data.get("competition_order_tail"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [g1_second.id, g1_first.id, g2_second.id, g2_first.id])

    def test_competition_order_tail_toggle_disable_restores_stack_only_order(self):
        g1_first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=1,
            ordre_competicio=2,
        )
        g1_second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=2,
            ordre_competicio=1,
        )
        g2_first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Charlie",
            categoria="Mateixa",
            grup=2,
            ordre_sortida=3,
            ordre_competicio=2,
        )
        g2_second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Delta",
            categoria="Mateixa",
            grup=2,
            ordre_sortida=4,
            ordre_competicio=1,
        )

        apply_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "categoria",
                "sort_dir": "asc",
                "scope": "all_groups",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(apply_resp.status_code, 200)
        self.assertTrue(apply_resp.json().get("ok"))

        toggle_on_resp = self._toggle_competition_tail(True)
        self.assertEqual(toggle_on_resp.status_code, 200)
        self.assertTrue(toggle_on_resp.json().get("competition_order_tail"))

        disable_resp = self._toggle_competition_tail(False)
        self.assertEqual(disable_resp.status_code, 200)
        data = disable_resp.json()
        self.assertTrue(data.get("ok"))
        self.assertFalse(data.get("competition_order_tail"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [g1_first.id, g1_second.id, g2_first.id, g2_second.id])

    def test_competition_order_tail_places_missing_saved_order_last_and_stable(self):
        first_missing = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=1,
        )
        second_with = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=2,
            ordre_competicio=2,
        )
        third_with = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Charlie",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=3,
            ordre_competicio=1,
        )
        second_missing = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Delta",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=4,
        )
        Inscripcio.objects.filter(pk__in=[first_missing.pk, second_missing.pk]).update(
            ordre_competicio=None
        )

        apply_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "categoria",
                "sort_dir": "asc",
                "scope": "all_groups",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(apply_resp.status_code, 200)
        self.assertTrue(apply_resp.json().get("ok"))

        toggle_resp = self._toggle_competition_tail(True)
        self.assertEqual(toggle_resp.status_code, 200)
        self.assertTrue(toggle_resp.json().get("competition_order_tail"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(
            actual_ids,
            [third_with.id, second_with.id, first_missing.id, second_missing.id],
        )

    def test_sort_clear_resets_competition_order_tail_flag(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=1,
            ordre_competicio=2,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=2,
            ordre_competicio=1,
        )

        apply_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "categoria",
                "sort_dir": "asc",
                "scope": "all_groups",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(apply_resp.status_code, 200)
        self.assertTrue(apply_resp.json().get("ok"))
        self.assertTrue(self._toggle_competition_tail(True).json().get("competition_order_tail"))

        clear_resp = self._post_json(
            "inscripcions_sort_clear",
            {
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(clear_resp.status_code, 200)
        data = clear_resp.json()
        self.assertTrue(data.get("ok"))
        self.assertTrue(data.get("cleared"))
        self.assertFalse(data.get("competition_order_tail"))

    def test_competition_order_tail_applies_independently_per_group(self):
        g1_first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Anna",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=1,
            ordre_competicio=2,
        )
        g2_first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Biel",
            categoria="Mateixa",
            grup=2,
            ordre_sortida=2,
            ordre_competicio=2,
        )
        g1_second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Carla",
            categoria="Mateixa",
            grup=1,
            ordre_sortida=3,
            ordre_competicio=1,
        )
        g2_second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="David",
            categoria="Mateixa",
            grup=2,
            ordre_sortida=4,
            ordre_competicio=1,
        )

        apply_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "categoria",
                "sort_dir": "asc",
                "scope": "all",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(apply_resp.status_code, 200)
        self.assertTrue(apply_resp.json().get("ok"))

        toggle_resp = self._toggle_competition_tail(True)
        self.assertEqual(toggle_resp.status_code, 200)
        self.assertTrue(toggle_resp.json().get("competition_order_tail"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [g1_second.id, g2_second.id, g1_first.id, g2_first.id])

    def test_custom_sort_save_does_not_reapply_if_mode_is_not_custom(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            categoria="B",
            ordre_sortida=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            categoria="C",
            ordre_sortida=2,
        )
        i3 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I3",
            categoria="A",
            ordre_sortida=3,
        )

        base_payload = {
            "scope": "all",
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }

        r_apply = self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "categoria", "sort_dir": "asc"},
        )
        self.assertEqual(r_apply.status_code, 200)
        self.assertTrue(r_apply.json().get("ok"))

        after_apply_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(after_apply_ids, [i3.id, i1.id, i2.id])

        r_custom = self._post_json(
            "inscripcions_sort_custom_save",
            {
                "sort_key": "categoria",
                "order": ["C", "B", "A"],
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
                "preserve_missing_context": True,
            },
        )
        self.assertEqual(r_custom.status_code, 200)
        data = r_custom.json()
        self.assertTrue(data.get("ok"))
        self.assertFalse(data.get("reapplied"))
        self.assertEqual(data.get("reapplied_updated"), 0)

        after_custom_save_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(after_custom_save_ids, [i3.id, i1.id, i2.id])

    def test_custom_sort_values_for_group_use_named_labels(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="G1",
            grup=2,
            ordre_sortida=1,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="G2",
            grup=2,
            ordre_sortida=2,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="G3",
            grup=3,
            ordre_sortida=3,
        )
        group = GrupCompeticio.objects.get(competicio=self.comp, display_num=2)
        group.nom = "Final"
        group.save(update_fields=["nom"])

        response = self._post_json(
            "inscripcions_sort_custom_values",
            {
                "sort_key": "grup",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data.get("ok"))

        values_by_token = {str(item["value"]): item for item in data.get("values", [])}
        self.assertEqual(values_by_token["2"]["label"], "Final")
        self.assertEqual(values_by_token["2"]["count"], 2)
        self.assertEqual(values_by_token["3"]["label"], "Grup 3")

    def test_custom_sort_values_for_equip_use_team_names_and_skip_empty(self):
        equip = Equip.objects.create(competicio=self.comp, nom="Equip Alpha")
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="E1",
            equip=equip,
            ordre_sortida=1,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="E2",
            equip=equip,
            ordre_sortida=2,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Sense equip",
            ordre_sortida=3,
        )

        response = self._post_json(
            "inscripcions_sort_custom_values",
            {
                "sort_key": "equip",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data.get("ok"))
        self.assertEqual(
            data.get("values"),
            [
                {
                    "value": str(equip.id),
                    "label": "Equip Alpha",
                    "count": 2,
                    "detected": True,
                    "in_custom": False,
                }
            ],
        )

    def test_equip_sort_apply_uses_team_name_for_fallback(self):
        equip_beta = Equip.objects.create(competicio=self.comp, nom="Beta")
        equip_alpha = Equip.objects.create(competicio=self.comp, nom="Alpha")
        ins_beta = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta 1",
            equip=equip_beta,
            ordre_sortida=1,
        )
        ins_alpha = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha 1",
            equip=equip_alpha,
            ordre_sortida=2,
        )

        response = self._post_json(
            "inscripcions_sort_apply",
            {
                "sort_key": "equip",
                "sort_dir": "asc",
                "scope": "all",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json().get("ok"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [ins_alpha.id, ins_beta.id])

    def test_equip_custom_sort_save_reapplies_active_stack_immediately(self):
        equip_alpha = Equip.objects.create(competicio=self.comp, nom="Alpha")
        equip_beta = Equip.objects.create(competicio=self.comp, nom="Beta")
        ins_alpha = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha 1",
            equip=equip_alpha,
            ordre_sortida=1,
        )
        ins_beta = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta 1",
            equip=equip_beta,
            ordre_sortida=2,
        )

        base_payload = {
            "scope": "all",
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }

        r_apply = self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "equip", "sort_dir": "custom"},
        )
        self.assertEqual(r_apply.status_code, 200)
        self.assertTrue(r_apply.json().get("ok"))

        r_custom = self._post_json(
            "inscripcions_sort_custom_save",
            {
                "sort_key": "equip",
                "order": [str(equip_beta.id), str(equip_alpha.id)],
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
                "preserve_missing_context": True,
            },
        )
        self.assertEqual(r_custom.status_code, 200)
        data = r_custom.json()
        self.assertTrue(data.get("ok"))
        self.assertTrue(data.get("reapplied"))
        self.assertEqual(data.get("stack_count"), 1)

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [ins_beta.id, ins_alpha.id])

    def test_equip_custom_order_survives_team_rename(self):
        equip_alpha = Equip.objects.create(competicio=self.comp, nom="Alpha")
        equip_beta = Equip.objects.create(competicio=self.comp, nom="Beta")
        ins_alpha = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alpha 1",
            equip=equip_alpha,
            ordre_sortida=1,
        )
        ins_beta = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Beta 1",
            equip=equip_beta,
            ordre_sortida=2,
        )

        base_payload = {
            "scope": "all",
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }

        self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "equip", "sort_dir": "custom"},
        )
        save_resp = self._post_json(
            "inscripcions_sort_custom_save",
            {
                "sort_key": "equip",
                "order": [str(equip_beta.id), str(equip_alpha.id)],
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
                "preserve_missing_context": True,
            },
        )
        self.assertEqual(save_resp.status_code, 200)
        self.assertTrue(save_resp.json().get("ok"))

        equip_alpha.nom = "Alfa"
        equip_alpha.save(update_fields=["nom"])
        equip_beta.nom = "Zeta"
        equip_beta.save(update_fields=["nom"])

        asc_resp = self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "equip", "sort_dir": "asc"},
        )
        self.assertEqual(asc_resp.status_code, 200)
        self.assertTrue(asc_resp.json().get("ok"))

        custom_resp = self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "equip", "sort_dir": "custom"},
        )
        self.assertEqual(custom_resp.status_code, 200)
        self.assertTrue(custom_resp.json().get("ok"))

        actual_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(actual_ids, [ins_beta.id, ins_alpha.id])

    def test_history_undo_redo_restores_sort_apply(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            entitat="B",
            ordre_sortida=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            entitat="A",
            ordre_sortida=2,
        )
        base_payload = {
            "scope": "all",
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }

        r_apply = self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "entitat", "sort_dir": "asc"},
        )
        self.assertEqual(r_apply.status_code, 200)
        self.assertTrue(r_apply.json().get("ok"))

        ordered_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(ordered_ids, [i2.id, i1.id])

        r_undo = self._post_history("undo")
        self.assertEqual(r_undo.status_code, 200)
        self.assertTrue(r_undo.json().get("ok"))
        self.assertTrue(r_undo.json().get("applied"))

        undone_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(undone_ids, [i1.id, i2.id])

        r_redo = self._post_history("redo")
        self.assertEqual(r_redo.status_code, 200)
        self.assertTrue(r_redo.json().get("ok"))
        self.assertTrue(r_redo.json().get("applied"))

        redone_ids = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(redone_ids, [i2.id, i1.id])

    def test_history_new_action_clears_redo_branch(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            entitat="B",
            ordre_sortida=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            entitat="A",
            ordre_sortida=2,
        )
        base_payload = {
            "scope": "all",
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }
        self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "entitat", "sort_dir": "asc"},
        )
        self._post_history("undo")

        r_reorder = self._post_json(
            "inscripcions_reorder",
            {
                "ids": [i2.id, i1.id],
                "moved_id": i2.id,
                "new_index": 0,
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(r_reorder.status_code, 200)
        self.assertTrue(r_reorder.json().get("ok"))
        self.assertFalse(r_reorder.json().get("history", {}).get("can_redo"))

        r_redo = self._post_history("redo")
        self.assertEqual(r_redo.status_code, 200)
        self.assertTrue(r_redo.json().get("ok"))
        self.assertFalse(r_redo.json().get("applied"))

    def test_sort_undo_compat_wrapper_uses_global_history(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            entitat="B",
            ordre_sortida=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            entitat="A",
            ordre_sortida=2,
        )
        base_payload = {
            "scope": "all",
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
        }
        self._post_json(
            "inscripcions_sort_apply",
            {**base_payload, "sort_key": "entitat", "sort_dir": "asc"},
        )

        r_compat = self._post_json(
            "inscripcions_sort_undo",
            {"filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""}, "group_by": []},
        )
        self.assertEqual(r_compat.status_code, 200)
        data = r_compat.json()
        self.assertTrue(data.get("ok"))
        self.assertEqual(data.get("restored"), 1)

        ids_after = list(
            Inscripcio.objects.filter(competicio=self.comp)
            .order_by("ordre_sortida", "id")
            .values_list("id", flat=True)
        )
        self.assertEqual(ids_after, [i1.id, i2.id])

    def test_reorder_cleans_orphan_group_labels(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            ordre_sortida=1,
            grup=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            ordre_sortida=2,
            grup=2,
        )
        self.comp.inscripcions_view = {"group_names": {"1": "Un", "2": "Dos"}}
        self.comp.save(update_fields=["inscripcions_view"])

        resp = self._post_json(
            "inscripcions_reorder",
            {
                "ids": [i2.id, i1.id],
                "moved_id": i1.id,
                "new_index": 1,
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json().get("ok"))

        i1.refresh_from_db()
        self.assertEqual(i1.grup, 1)

        self.comp.refresh_from_db()
        self.assertEqual(self.comp.inscripcions_view.get("group_names"), {"1": "Un", "2": "Dos"})

    def test_reorder_prefers_target_group_over_previous_row_group(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            ordre_sortida=1,
            grup=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            ordre_sortida=2,
            grup=1,
        )
        i3 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I3",
            ordre_sortida=3,
            grup=2,
        )
        i4 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I4",
            ordre_sortida=4,
            grup=2,
        )

        resp = self._post_json(
            "inscripcions_reorder",
            {
                "ids": [i2.id, i1.id, i3.id, i4.id],
                "moved_id": i1.id,
                "new_index": 1,
                "target_group": 2,
                "mode": "group_edit",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json().get("ok"))

        i1.refresh_from_db()
        self.assertEqual(i1.grup, 2)
        self.assertEqual(i1.ordre_competicio, 3)

    def test_reorder_without_header_target_keeps_legacy_edge_behavior(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            ordre_sortida=1,
            grup=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            ordre_sortida=2,
            grup=2,
        )

        resp = self._post_json(
            "inscripcions_reorder",
            {
                "ids": [i2.id, i1.id],
                "moved_id": i2.id,
                "new_index": 0,
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json().get("ok"))

        i2.refresh_from_db()
        self.assertEqual(i2.grup, 2)

    def test_save_group_competition_order_updates_only_real_order(self):
        i1 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I1",
            ordre_sortida=1,
            grup=1,
        )
        i2 = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="I2",
            ordre_sortida=2,
            grup=1,
        )

        resp = self._post_json(
            "inscripcions_save_group_competition_order",
            {
                "group_num": 1,
                "ids": [i2.id, i1.id],
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json().get("ok"))

        i1.refresh_from_db()
        i2.refresh_from_db()
        self.assertEqual(i1.ordre_sortida, 1)
        self.assertEqual(i2.ordre_sortida, 2)
        self.assertEqual(i2.ordre_competicio, 1)
        self.assertEqual(i1.ordre_competicio, 2)

    def test_groups_preview_marks_existing_group_as_reduced_when_members_remain_outside_filter(self):
        moving = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Visible",
            entitat="Club A",
            ordre_sortida=1,
            grup=1,
        )
        staying = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Resta",
            entitat="Club B",
            ordre_sortida=2,
            grup=1,
        )

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(filters={"q": "", "categoria": "", "subcategoria": "", "entitat": "Club A"}),
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get("ok"))

        preview = data.get("preview") or {}
        existing_groups = preview.get("existing_groups") or []
        self.assertEqual(len(existing_groups), 1)
        existing = existing_groups[0]
        self.assertEqual(existing.get("group_num"), 1)
        self.assertEqual(existing.get("impact_kind"), "reduced")
        self.assertEqual(existing.get("members_count"), 2)
        self.assertEqual(existing.get("moving_members_count"), 1)
        self.assertEqual(existing.get("remaining_members_count"), 1)
        self.assertEqual(existing.get("moving_member_names_preview"), [moving.nom_i_cognoms])

        source_map = {row.get("label"): row for row in existing.get("sources") or []}
        self.assertEqual(source_map["Totes les inscripcions filtrades"]["moving_count"], 1)
        self.assertEqual(source_map["Totes les inscripcions filtrades"]["remaining_count"], 0)
        self.assertEqual(source_map["Fora del filtre actual"]["moving_count"], 0)
        self.assertEqual(source_map["Fora del filtre actual"]["remaining_count"], 1)
        self.assertEqual(source_map["Fora del filtre actual"]["count"], 1)
        self.assertEqual(preview.get("existing_members_total"), 2)

        self.assertEqual(staying.grup, 1)

    def test_groups_preview_marks_existing_group_as_removed_when_all_members_move(self):
        first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Primer",
            ordre_sortida=1,
            grup=1,
        )
        second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Segon",
            ordre_sortida=2,
            grup=1,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Sense grup",
            ordre_sortida=3,
            grup=None,
        )

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(),
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get("ok"))

        preview = data.get("preview") or {}
        existing_groups = preview.get("existing_groups") or []
        self.assertEqual(len(existing_groups), 1)
        existing = existing_groups[0]
        self.assertEqual(existing.get("group_num"), 1)
        self.assertEqual(existing.get("impact_kind"), "removed")
        self.assertEqual(existing.get("members_count"), 2)
        self.assertEqual(existing.get("moving_members_count"), 2)
        self.assertEqual(existing.get("remaining_members_count"), 0)
        self.assertEqual(existing.get("moving_member_names_preview"), [first.nom_i_cognoms, second.nom_i_cognoms])
        self.assertEqual(preview.get("existing_groups_total"), 1)

    def test_groups_preview_existing_group_exposes_existing_name_label(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Primer",
            ordre_sortida=1,
            grup=1,
        )
        renumber_groups_for_competicio(self.comp)
        group = GrupCompeticio.objects.get(competicio=self.comp, display_num=1)
        group.nom = "Final"
        group.save(update_fields=["nom"])

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(),
        )
        self.assertEqual(resp.status_code, 200)

        preview = resp.json().get("preview") or {}
        existing_groups = preview.get("existing_groups") or []
        self.assertEqual(len(existing_groups), 1)
        self.assertEqual(existing_groups[0].get("group_label"), "Final")

    def test_groups_preview_suggests_name_from_single_sort_bucket(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A1",
            entitat="Club A",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A2",
            entitat="Club A",
            ordre_sortida=2,
            grup=None,
        )

        sort_payload = {
            "scope": "all",
            "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
            "group_by": [],
            "sort_key": "entitat",
            "sort_dir": "asc",
        }
        sort_resp = self._post_json("inscripcions_sort_apply", sort_payload)
        self.assertEqual(sort_resp.status_code, 200)
        self.assertTrue(sort_resp.json().get("ok"))

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(strategy="per_bucket"),
        )
        self.assertEqual(resp.status_code, 200)

        groups = (resp.json().get("preview") or {}).get("groups") or []
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0].get("suggested_name"), "Club A")

    def test_groups_preview_suggests_name_from_single_tab_bucket(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A1",
            categoria="Alevi",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A2",
            categoria="Alevi",
            ordre_sortida=2,
            grup=None,
        )

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(strategy="per_bucket", group_by=["categoria"]),
        )
        self.assertEqual(resp.status_code, 200)

        groups = (resp.json().get("preview") or {}).get("groups") or []
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0].get("suggested_name"), "Alevi")

    def test_groups_preview_auto_resolution_combines_group_by_and_sort(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi A",
            categoria="Alevi",
            entitat="Club A",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi B",
            categoria="Alevi",
            entitat="Club B",
            ordre_sortida=2,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Infantil A",
            categoria="Infantil",
            entitat="Club A",
            ordre_sortida=3,
            grup=None,
        )

        sort_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "scope": "all",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": ["categoria"],
                "sort_key": "entitat",
                "sort_dir": "asc",
            },
        )
        self.assertEqual(sort_resp.status_code, 200)
        self.assertTrue(sort_resp.json().get("ok"))

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(strategy="per_bucket", group_by=["categoria"]),
        )
        self.assertEqual(resp.status_code, 200)

        data = resp.json()
        self.assertEqual(data.get("resolution_mode"), "auto")
        self.assertEqual(data.get("layers_used"), ["tabs", "sort"])
        self.assertEqual(data.get("effective_bucket_count"), 3)

        preview = data.get("preview") or {}
        groups = preview.get("groups") or []
        self.assertEqual(len(groups), 3)
        self.assertEqual(preview.get("layers_used"), ["tabs", "sort"])
        self.assertEqual(preview.get("effective_bucket_count"), 3)
        self.assertEqual(
            [group.get("suggested_name") for group in groups],
            ["Alevi + Club A", "Infantil + Club A", "Alevi + Club B"],
        )

    def test_groups_preview_auto_resolution_uses_tabs_only_without_sort(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi 1",
            categoria="Alevi",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi 2",
            categoria="Alevi",
            ordre_sortida=2,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Infantil 1",
            categoria="Infantil",
            ordre_sortida=3,
            grup=None,
        )

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(strategy="per_bucket", group_by=["categoria"]),
        )
        self.assertEqual(resp.status_code, 200)

        data = resp.json()
        self.assertEqual(data.get("layers_used"), ["tabs"])
        self.assertEqual(data.get("effective_bucket_count"), 2)
        groups = (data.get("preview") or {}).get("groups") or []
        self.assertEqual([group.get("suggested_name") for group in groups], ["Alevi", "Infantil"])

    def test_groups_preview_auto_resolution_uses_sort_only_without_group_by(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Club A 1",
            entitat="Club A",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Club A 2",
            entitat="Club A",
            ordre_sortida=2,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Club B 1",
            entitat="Club B",
            ordre_sortida=3,
            grup=None,
        )

        sort_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "scope": "all",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": [],
                "sort_key": "entitat",
                "sort_dir": "asc",
            },
        )
        self.assertEqual(sort_resp.status_code, 200)
        self.assertTrue(sort_resp.json().get("ok"))

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(strategy="per_bucket"),
        )
        self.assertEqual(resp.status_code, 200)

        data = resp.json()
        self.assertEqual(data.get("layers_used"), ["sort"])
        self.assertEqual(data.get("effective_bucket_count"), 2)
        groups = (data.get("preview") or {}).get("groups") or []
        self.assertEqual([group.get("suggested_name") for group in groups], ["Club A", "Club B"])

    def test_groups_preview_auto_resolution_deduplicates_redundant_sort_partition(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi 1",
            categoria="Alevi",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Infantil 1",
            categoria="Infantil",
            ordre_sortida=2,
            grup=None,
        )

        sort_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "scope": "all",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": ["categoria"],
                "sort_key": "categoria",
                "sort_dir": "asc",
            },
        )
        self.assertEqual(sort_resp.status_code, 200)
        self.assertTrue(sort_resp.json().get("ok"))

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(strategy="per_bucket", group_by=["categoria"]),
        )
        self.assertEqual(resp.status_code, 200)

        data = resp.json()
        self.assertEqual(data.get("layers_used"), ["tabs"])
        self.assertEqual(data.get("effective_bucket_count"), 2)

    def test_groups_preview_selected_keys_accept_combined_bucket_key(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi A",
            categoria="Alevi",
            entitat="Club A",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi B",
            categoria="Alevi",
            entitat="Club B",
            ordre_sortida=2,
            grup=None,
        )

        sort_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "scope": "all",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": ["categoria"],
                "sort_key": "entitat",
                "sort_dir": "asc",
            },
        )
        self.assertEqual(sort_resp.status_code, 200)
        self.assertTrue(sort_resp.json().get("ok"))

        tab_key = json.dumps(["Alevi"], ensure_ascii=False)
        sort_key = json.dumps(["Club A"], ensure_ascii=False)
        combined_key = json.dumps(
            [
                {"kind": "tabs", "key": tab_key},
                {"kind": "sort", "key": sort_key},
            ],
            ensure_ascii=False,
        )

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(
                strategy="per_bucket",
                group_by=["categoria"],
                selected_keys=[combined_key],
            ),
        )
        self.assertEqual(resp.status_code, 200)

        data = resp.json()
        self.assertEqual(data.get("buckets_applied"), 1)
        preview = data.get("preview") or {}
        self.assertEqual(preview.get("members_total"), 1)
        groups = preview.get("groups") or []
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0].get("suggested_name"), "Alevi + Club A")

    def test_groups_preview_tab_merges_apply_before_combining_with_sort(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi A",
            categoria="Alevi",
            entitat="Club A",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Infantil A",
            categoria="Infantil",
            entitat="Club A",
            ordre_sortida=2,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi B",
            categoria="Alevi",
            entitat="Club B",
            ordre_sortida=3,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Infantil B",
            categoria="Infantil",
            entitat="Club B",
            ordre_sortida=4,
            grup=None,
        )
        self.comp.tab_merges = {
            "categoria": [[json.dumps(["Alevi"], ensure_ascii=False), json.dumps(["Infantil"], ensure_ascii=False)]]
        }
        self.comp.save(update_fields=["tab_merges"])

        sort_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "scope": "all",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": ["categoria"],
                "sort_key": "entitat",
                "sort_dir": "asc",
            },
        )
        self.assertEqual(sort_resp.status_code, 200)
        self.assertTrue(sort_resp.json().get("ok"))

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(strategy="per_bucket", group_by=["categoria"]),
        )
        self.assertEqual(resp.status_code, 200)

        data = resp.json()
        self.assertEqual(data.get("layers_used"), ["tabs", "sort"])
        self.assertEqual(data.get("effective_bucket_count"), 2)
        groups = (data.get("preview") or {}).get("groups") or []
        self.assertEqual(
            [group.get("suggested_name") for group in groups],
            ["Alevi + Infantil + Club A", "Alevi + Infantil + Club B"],
        )

    def test_groups_preview_existing_groups_use_combined_sources(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi A",
            categoria="Alevi",
            entitat="Club A",
            ordre_sortida=1,
            grup=1,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Alevi B",
            categoria="Alevi",
            entitat="Club B",
            ordre_sortida=2,
            grup=1,
        )

        sort_resp = self._post_json(
            "inscripcions_sort_apply",
            {
                "scope": "all",
                "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                "group_by": ["categoria"],
                "sort_key": "entitat",
                "sort_dir": "asc",
            },
        )
        self.assertEqual(sort_resp.status_code, 200)
        self.assertTrue(sort_resp.json().get("ok"))

        tab_key = json.dumps(["Alevi"], ensure_ascii=False)
        sort_key = json.dumps(["Club A"], ensure_ascii=False)
        combined_key = json.dumps(
            [
                {"kind": "tabs", "key": tab_key},
                {"kind": "sort", "key": sort_key},
            ],
            ensure_ascii=False,
        )

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(
                strategy="per_bucket",
                group_by=["categoria"],
                selected_keys=[combined_key],
            ),
        )
        self.assertEqual(resp.status_code, 200)

        preview = resp.json().get("preview") or {}
        existing_groups = preview.get("existing_groups") or []
        self.assertEqual(len(existing_groups), 1)
        existing = existing_groups[0]
        self.assertEqual(existing.get("impact_kind"), "reduced")
        source_map = {row.get("label"): row for row in existing.get("sources") or []}
        self.assertEqual(source_map["Alevi / Club A"]["moving_count"], 1)
        self.assertEqual(source_map["Alevi / Club A"]["remaining_count"], 0)
        self.assertEqual(source_map["Alevi / Club B"]["moving_count"], 0)
        self.assertEqual(source_map["Alevi / Club B"]["remaining_count"], 1)

    def test_groups_preview_builds_composed_name_from_weighted_sources(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A1",
            entitat="Club A",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A2",
            entitat="Club A",
            ordre_sortida=2,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="B1",
            entitat="Club B",
            ordre_sortida=3,
            grup=None,
        )

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(
                source="tabs",
                strategy="count",
                group_count=1,
                group_by=["entitat"],
            ),
        )
        self.assertEqual(resp.status_code, 200)

        groups = (resp.json().get("preview") or {}).get("groups") or []
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0].get("suggested_name"), "Club A + Club B")

    def test_groups_preview_disambiguates_duplicate_suggested_names(self):
        for idx in range(1, 5):
            Inscripcio.objects.create(
                competicio=self.comp,
                nom_i_cognoms=f"A{idx}",
                entitat="Club A",
                ordre_sortida=idx,
                grup=None,
            )

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(
                source="tabs",
                strategy="count",
                group_count=2,
                group_by=["entitat"],
            ),
        )
        self.assertEqual(resp.status_code, 200)

        groups = (resp.json().get("preview") or {}).get("groups") or []
        self.assertEqual(
            [group.get("suggested_name") for group in groups],
            ["Club A (1)", "Club A (2)"],
        )

    def test_groups_preview_ignores_generic_fallback_labels(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Sense Nom 1",
            ordre_sortida=1,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Sense Nom 2",
            ordre_sortida=2,
            grup=None,
        )

        resp = self._post_json("inscripcions_groups_from_sort", self._groups_payload())
        self.assertEqual(resp.status_code, 200)

        groups = (resp.json().get("preview") or {}).get("groups") or []
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0].get("suggested_name"), "")

    def test_groups_apply_persists_suggested_name_without_overwriting_existing_manual_names(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Manual",
            entitat="Club X",
            ordre_sortida=1,
            grup=1,
        )
        renumber_groups_for_competicio(self.comp)
        existing_group = GrupCompeticio.objects.get(competicio=self.comp, display_num=1)
        existing_group.nom = "Manual"
        existing_group.save(update_fields=["nom"])
        self.comp.inscripcions_view = {"group_names": {"1": "Manual"}}
        self.comp.save(update_fields=["inscripcions_view"])

        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Nou 1",
            entitat="Club A",
            ordre_sortida=2,
            grup=None,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Nou 2",
            entitat="Club A",
            ordre_sortida=3,
            grup=None,
        )

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(
                source="tabs",
                strategy="per_bucket",
                preview_only=False,
                group_by=["entitat"],
                filters={"q": "", "categoria": "", "subcategoria": "", "entitat": "Club A"},
            ),
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json().get("ok"))

        existing_group.refresh_from_db()
        self.assertEqual(existing_group.nom, "Manual")

        new_group = GrupCompeticio.objects.get(competicio=self.comp, display_num=2)
        self.assertEqual(new_group.nom, "Club A")

        self.comp.refresh_from_db()
        self.assertEqual(
            self.comp.inscripcions_view.get("group_names"),
            {"1": "Manual", "2": "Club A"},
        )

    def test_groups_apply_deactivates_group_that_becomes_empty(self):
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Primer",
            ordre_sortida=1,
            grup=1,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Segon",
            ordre_sortida=2,
            grup=1,
        )
        renumber_groups_for_competicio(self.comp)
        old_group = GrupCompeticio.objects.get(competicio=self.comp, display_num=1)
        self.assertTrue(old_group.actiu)

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(preview_only=False),
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get("ok"))

        old_group.refresh_from_db()
        self.assertFalse(old_group.actiu)
        self.assertTrue(
            GrupCompeticio.objects.filter(competicio=self.comp, display_num=2, actiu=True).exists()
        )

    def test_groups_apply_with_rotations_keeps_programmed_group_when_members_remain(self):
        first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Primer",
            entitat="Club A",
            ordre_sortida=1,
            grup=1,
        )
        second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Segon",
            entitat="Club B",
            ordre_sortida=2,
            grup=1,
        )
        group = first.grup_competicio
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="09:00",
            hora_fi="09:30",
            ordre=1,
            titol="Franja 1",
        )
        estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="descans",
            ordre=1,
            actiu=True,
        )
        assignacio = RotacioAssignacio.objects.create(
            competicio=self.comp,
            franja=franja,
            estacio=estacio,
        )
        RotacioAssignacioGrup.objects.create(assignacio=assignacio, grup=group, ordre=1)

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(
                preview_only=False,
                filters={"q": "", "categoria": "", "subcategoria": "", "entitat": "Club A"},
            ),
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get("ok"))

        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.grup, 2)
        self.assertEqual(second.grup, 1)
        self.assertTrue(
            GrupCompeticio.objects.filter(competicio=self.comp, display_num=2, actiu=True).exists()
        )

    def test_groups_apply_with_rotations_rejects_emptying_programmed_group(self):
        first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Primer",
            ordre_sortida=1,
            grup=1,
        )
        second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Segon",
            ordre_sortida=2,
            grup=1,
        )
        group = first.grup_competicio
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="09:00",
            hora_fi="09:30",
            ordre=1,
            titol="Franja 1",
        )
        estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="descans",
            ordre=1,
            actiu=True,
        )
        assignacio = RotacioAssignacio.objects.create(
            competicio=self.comp,
            franja=franja,
            estacio=estacio,
        )
        RotacioAssignacioGrup.objects.create(assignacio=assignacio, grup=group, ordre=1)

        resp = self._post_json(
            "inscripcions_groups_from_sort",
            self._groups_payload(preview_only=False),
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn("No es pot deixar buit un grup inclos al programa de rotacions", resp.content.decode("utf-8"))

        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.grup, 1)
        self.assertEqual(second.grup, 1)


class GroupNameSyncTests(TestCase):
    def test_renumber_remaps_group_labels_and_drops_stale(self):
        comp = Competicio.objects.create(
            nom="Comp labels remap",
            tipus=Competicio.Tipus.TRAMPOLI,
            inscripcions_view={"group_names": {"1": "Antic", "2": "Beta", "4": "Gamma"}},
        )
        Inscripcio.objects.create(
            competicio=comp,
            nom_i_cognoms="I1",
            ordre_sortida=1,
            grup=2,
        )
        Inscripcio.objects.create(
            competicio=comp,
            nom_i_cognoms="I2",
            ordre_sortida=2,
            grup=4,
        )

        renumber_groups_for_competicio(comp)

        comp.refresh_from_db()
        self.assertEqual(comp.inscripcions_view.get("group_names"), {"2": "Beta", "4": "Gamma"})

    def test_renumber_without_groups_clears_group_labels(self):
        comp = Competicio.objects.create(
            nom="Comp labels clear",
            tipus=Competicio.Tipus.TRAMPOLI,
            inscripcions_view={"group_names": {"1": "Orfe", "9": "Fantasma"}},
        )

        renumber_groups_for_competicio(comp)

        comp.refresh_from_db()
        self.assertNotIn("group_names", comp.inscripcions_view or {})


class InscripcioManualFormViewTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp manual form")
        self.other_comp = self._create_competicio("Comp altre")
        User = get_user_model()
        self.user = User.objects.create_user(
            username="manual_form_editor",
            password="testpass123",
            email="manual-form@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        self.client.force_login(self.user)

    def _add_url(self):
        return reverse("inscripcio_add", kwargs={"pk": self.comp.id})

    def _edit_url(self, inscripcio):
        return reverse("inscripcio_edit", kwargs={"pk": self.comp.id, "ins_id": inscripcio.id})

    def test_create_form_uses_schema_fields_and_competition_scoped_choices(self):
        self.comp.inscripcions_schema = {
            "columns": [
                {"code": "categoria", "label": "Categoria", "kind": "builtin"},
                {"code": "subcategoria", "label": "Subcategoria", "kind": "builtin"},
                {"code": "modalitat", "label": "Modalitat", "kind": "extra"},
            ],
            "value_aliases": {
                "entitat": {"club a": "Club A"},
                "categoria": {"open": "Open"},
                "subcategoria": {"nivell 4": "Nivell 4"},
            },
        }
        self.comp.save(update_fields=["inscripcions_schema"])

        GrupCompeticio.objects.create(
            competicio=self.comp,
            legacy_num=1,
            display_num=1,
            nom="Matinal",
            actiu=True,
        )
        GrupCompeticio.objects.create(
            competicio=self.comp,
            legacy_num=2,
            display_num=2,
            nom="",
            actiu=True,
        )
        equip = Equip.objects.create(competicio=self.comp, nom="Equip A")
        Equip.objects.create(competicio=self.other_comp, nom="Equip Extern")
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A",
            entitat="club a",
            categoria="open",
            subcategoria="nivell 4",
            equip=equip,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="B",
            entitat="Club B",
            categoria="Base",
            subcategoria="Nivell 3",
        )
        Inscripcio.objects.create(
            competicio=self.other_comp,
            nom_i_cognoms="C",
            entitat="Club Extern",
            categoria="Externa",
            subcategoria="Externa",
        )

        response = self.client.get(self._add_url())
        self.assertEqual(response.status_code, 200)

        form = response.context["form"]
        self.assertNotIn("categoria", form.fields)
        self.assertNotIn("subcategoria", form.fields)
        self.assertNotIn("document", form.fields)
        self.assertIn("grup_competicio_choice", form.fields)
        self.assertIn("entitat_choice", form.fields)
        self.assertIn("entitat_altres", form.fields)
        self.assertIn("categoria_choice", form.fields)
        self.assertIn("categoria_altres", form.fields)
        self.assertIn("subcategoria_choice", form.fields)
        self.assertIn("subcategoria_altres", form.fields)
        self.assertIn("equip_choice", form.fields)
        self.assertIn("equip_altres", form.fields)
        self.assertIn("extra__modalitat", form.fields)

        group_choices = dict(form.fields["grup_competicio_choice"].choices)
        self.assertEqual(group_choices[""], "Sense grup")
        self.assertIn("Matinal", group_choices.values())
        self.assertIn("Grup 2", group_choices.values())

        entity_choices = dict(form.fields["entitat_choice"].choices)
        self.assertIn("Club A", entity_choices)
        self.assertIn("Club B", entity_choices)
        self.assertIn("__other__", entity_choices)
        self.assertNotIn("club a", entity_choices)
        self.assertNotIn("Club Extern", entity_choices)

        categoria_choices = dict(form.fields["categoria_choice"].choices)
        self.assertIn("Open", categoria_choices)
        self.assertIn("Base", categoria_choices)
        self.assertNotIn("open", categoria_choices)
        self.assertNotIn("Externa", categoria_choices)

        subcategoria_choices = dict(form.fields["subcategoria_choice"].choices)
        self.assertIn("Nivell 4", subcategoria_choices)
        self.assertIn("Nivell 3", subcategoria_choices)
        self.assertNotIn("nivell 4", subcategoria_choices)
        self.assertNotIn("Externa", subcategoria_choices)

        equip_choices = dict(form.fields["equip_choice"].choices)
        self.assertEqual(equip_choices[str(equip.id)], "Equip A")
        self.assertIn("__other__", equip_choices)
        self.assertNotIn("Equip Extern", equip_choices.values())

        self.assertEqual(response.context["show_altres_fields"], {
            "entitat_altres": False,
            "categoria_altres": False,
            "subcategoria_altres": False,
            "equip_altres": False,
        })

    def test_create_form_saves_group_text_other_equip_other_and_extra_fields(self):
        self.comp.inscripcions_schema = {
            "columns": [
                {"code": "categoria", "label": "Categoria", "kind": "builtin"},
                {"code": "subcategoria", "label": "Subcategoria", "kind": "builtin"},
                {"code": "modalitat", "label": "Modalitat", "kind": "extra"},
            ]
        }
        self.comp.save(update_fields=["inscripcions_schema"])
        group = GrupCompeticio.objects.create(
            competicio=self.comp,
            legacy_num=4,
            display_num=4,
            nom="Tarda",
            actiu=True,
        )

        response = self.client.post(
            self._add_url(),
            data={
                "nom_i_cognoms": "Nova gimnasta",
                "entitat_choice": "__other__",
                "entitat_altres": "Club Nou",
                "categoria_choice": "__other__",
                "categoria_altres": "Open",
                "subcategoria_choice": "__other__",
                "subcategoria_altres": "Nivell 4",
                "grup_competicio_choice": str(group.id),
                "equip_choice": "__other__",
                "equip_altres": "Equip Nou",
                "extra__modalitat": "Sincronitzada",
            },
        )

        self.assertEqual(response.status_code, 302)

        inscripcio = Inscripcio.objects.get(competicio=self.comp, nom_i_cognoms="Nova gimnasta")
        self.assertEqual(inscripcio.entitat, "Club Nou")
        self.assertEqual(inscripcio.categoria, "Open")
        self.assertEqual(inscripcio.subcategoria, "Nivell 4")
        self.assertEqual(inscripcio.grup_competicio_id, group.id)
        self.assertEqual(inscripcio.grup, 4)
        self.assertEqual(inscripcio.ordre_sortida, 1)
        self.assertEqual(inscripcio.ordre_competicio, 1)
        self.assertIsNotNone(inscripcio.equip_id)
        self.assertEqual(inscripcio.equip.nom, "Equip Nou")
        self.assertEqual(inscripcio.equip.origen, Equip.Origen.MANUAL)
        self.assertEqual(inscripcio.equip.criteri, {})
        self.assertEqual(inscripcio.extra, {"modalitat": "Sincronitzada"})

    def test_create_form_requires_other_values_when_other_is_selected(self):
        response = self.client.post(
            self._add_url(),
            data={
                "nom_i_cognoms": "Sense valors",
                "entitat_choice": "__other__",
                "entitat_altres": "",
                "categoria_choice": "__other__",
                "categoria_altres": "",
                "subcategoria_choice": "__other__",
                "subcategoria_altres": "",
                "grup_competicio_choice": "",
                "equip_choice": "__other__",
                "equip_altres": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFormError(response.context["form"], "entitat_altres", "Cal indicar l'entitat si tries Altres.")
        self.assertFormError(response.context["form"], "categoria_altres", "Cal indicar la categoria si tries Altres.")
        self.assertFormError(response.context["form"], "subcategoria_altres", "Cal indicar la subcategoria si tries Altres.")
        self.assertFormError(response.context["form"], "equip_altres", "Cal indicar l'equip si tries Altres.")
        self.assertEqual(response.context["show_altres_fields"], {
            "entitat_altres": True,
            "categoria_altres": True,
            "subcategoria_altres": True,
            "equip_altres": True,
        })
        self.assertFalse(Inscripcio.objects.filter(competicio=self.comp, nom_i_cognoms="Sense valors").exists())

    def test_create_form_reuses_existing_team_when_other_is_selected(self):
        existing = Equip.objects.create(competicio=self.comp, nom="Equip Reutilitzat")

        response = self.client.post(
            self._add_url(),
            data={
                "nom_i_cognoms": "Amb equip existent",
                "entitat_choice": "",
                "categoria_choice": "",
                "subcategoria_choice": "",
                "grup_competicio_choice": "",
                "equip_choice": "__other__",
                "equip_altres": "Equip Reutilitzat",
            },
        )

        self.assertEqual(response.status_code, 302)
        inscripcio = Inscripcio.objects.get(competicio=self.comp, nom_i_cognoms="Amb equip existent")
        self.assertEqual(inscripcio.equip_id, existing.id)
        self.assertEqual(Equip.objects.filter(competicio=self.comp, nom="Equip Reutilitzat").count(), 1)

    def test_edit_form_keeps_inactive_group_available_and_can_clear_it(self):
        self.comp.inscripcions_schema = {
            "columns": [
                {"code": "modalitat", "label": "Modalitat", "kind": "extra"},
            ]
        }
        self.comp.save(update_fields=["inscripcions_schema"])

        inactive_group = GrupCompeticio.objects.create(
            competicio=self.comp,
            legacy_num=7,
            display_num=7,
            nom="Grup ocult",
            actiu=False,
        )
        equip = Equip.objects.create(competicio=self.comp, nom="Equip Fix")
        inscripcio = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Edita'm",
            entitat="Club A",
            categoria="Open",
            subcategoria="Nivell 4",
            grup=7,
            grup_competicio=inactive_group,
            equip=equip,
            ordre_sortida=1,
            ordre_competicio=1,
            extra={"modalitat": "Dobles"},
        )

        response = self.client.get(self._edit_url(inscripcio))
        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertIn((str(inactive_group.id), "Grup ocult"), form.fields["grup_competicio_choice"].choices)
        self.assertEqual(form["extra__modalitat"].value(), "Dobles")
        self.assertEqual(form["entitat_choice"].value(), "Club A")
        self.assertEqual(form["categoria_choice"].value(), "Open")
        self.assertEqual(form["subcategoria_choice"].value(), "Nivell 4")
        self.assertEqual(form["equip_choice"].value(), str(equip.id))

        response = self.client.post(
            self._edit_url(inscripcio),
            data={
                "nom_i_cognoms": "Edita'm",
                "entitat_choice": "Club A",
                "categoria_choice": "Open",
                "subcategoria_choice": "Nivell 4",
                "grup_competicio_choice": "",
                "equip_choice": str(equip.id),
                "extra__modalitat": "Dobles",
            },
        )

        self.assertEqual(response.status_code, 302)
        inscripcio.refresh_from_db()
        self.assertEqual(inscripcio.entitat, "Club A")
        self.assertIsNone(inscripcio.grup_competicio_id)
        self.assertIsNone(inscripcio.grup)
        self.assertIsNone(inscripcio.ordre_competicio)
        self.assertEqual(inscripcio.equip_id, equip.id)

    def test_form_fallback_without_schema_only_uses_basic_fields(self):
        response = self.client.get(self._add_url())
        self.assertEqual(response.status_code, 200)

        form = response.context["form"]
        self.assertEqual(
            set(form.fields.keys()),
            {
                "nom_i_cognoms",
                "entitat_choice",
                "entitat_altres",
                "categoria_choice",
                "categoria_altres",
                "subcategoria_choice",
                "subcategoria_altres",
                "grup_competicio_choice",
                "equip_choice",
                "equip_altres",
            },
        )


class InscripcioAparellExclusioModelTests(_BaseTrampoliDataMixin, TestCase):
    def test_clean_rejects_cross_competition_pair(self):
        comp_a = self._create_competicio("Comp A")
        comp_b = self._create_competicio("Comp B")
        ins = self._create_inscripcio(comp_a, "Participant A")

        app_b = self._create_aparell("DMT_B", "DMT B")
        comp_app_b = self._create_comp_aparell(comp_b, app_b)

        ex = InscripcioAparellExclusio(inscripcio=ins, comp_aparell=comp_app_b)
        with self.assertRaises(ValidationError):
            ex.full_clean()


class InscripcionsSetAparellsViewTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio()
        app1 = self._create_aparell("TRAMP_I", "Tramp I")
        app2 = self._create_aparell("TRAMP_II", "Tramp II")

        self.comp_app_1 = self._create_comp_aparell(self.comp, app1, ordre=1, actiu=True)
        self.comp_app_2 = self._create_comp_aparell(self.comp, app2, ordre=2, actiu=True)

        self.ins = self._create_inscripcio(self.comp, "Ginmasta 1")

    def test_set_aparells_creates_and_replaces_exclusions(self):
        url = reverse("inscripcions_set_aparells", kwargs={"pk": self.comp.id})

        r1 = self.client.post(
            url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins.id,
                    "selected_comp_aparell_ids": [self.comp_app_1.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(r1.status_code, 200)
        self.assertEqual(
            list(
                InscripcioAparellExclusio.objects.filter(inscripcio=self.ins)
                .values_list("comp_aparell_id", flat=True)
            ),
            [self.comp_app_2.id],
        )

        r2 = self.client.post(
            url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins.id,
                    "selected_comp_aparell_ids": [self.comp_app_1.id, self.comp_app_2.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(r2.status_code, 200)
        self.assertFalse(
            InscripcioAparellExclusio.objects.filter(inscripcio=self.ins).exists()
        )

    def test_set_aparells_rejects_ids_outside_competition(self):
        other_comp = self._create_competicio("Comp Altre")
        other_app = self._create_aparell("TUMB_X", "Tumbling X")
        other_comp_app = self._create_comp_aparell(other_comp, other_app, ordre=1, actiu=True)

        url = reverse("inscripcions_set_aparells", kwargs={"pk": self.comp.id})
        r = self.client.post(
            url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins.id,
                    "selected_comp_aparell_ids": [self.comp_app_1.id, other_comp_app.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 400)


class InscripcionsMediaFlowTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Media")
        self.ins = self._create_inscripcio(self.comp, "LUCIA POZO SANCHEZ")
        self.ins.entitat = "Collegi Sagrat Cor Diputacio"
        self.ins.subcategoria = "GEN"
        self.ins.sexe = "F"
        self.ins.save(update_fields=["entitat", "subcategoria", "sexe"])

        self.ins_2 = self._create_inscripcio(self.comp, "MARTA LOPEZ", ordre=2, grup=1)
        self.ins_2.entitat = "Club Prova"
        self.ins_2.subcategoria = "GEN"
        self.ins_2.sexe = "F"
        self.ins_2.save(update_fields=["entitat", "subcategoria", "sexe"])

        User = get_user_model()
        self.user = User.objects.create_user(
            username="media_editor_user",
            password="testpass123",
            email="media-editor@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        self.client.force_login(self.user)

    def _upload_media(self, inscripcio_id, filename="track.mp3", content_type="audio/mpeg"):
        url = reverse("inscripcions_media_upload", kwargs={"pk": self.comp.id})
        f = SimpleUploadedFile(filename, b"abc123", content_type=content_type)
        return self.client.post(
            url,
            data={
                "inscripcio_id": inscripcio_id,
                "media_file": f,
            },
        )

    def test_manual_upload_creates_primary_media(self):
        res = self._upload_media(self.ins.id, filename="routine.mp3", content_type="audio/mpeg")
        self.assertEqual(res.status_code, 200)
        payload = res.json()
        self.assertTrue(payload.get("ok"))
        self.assertTrue(InscripcioMedia.objects.filter(inscripcio=self.ins).exists())
        item = InscripcioMedia.objects.get(inscripcio=self.ins)
        self.assertEqual(item.source, InscripcioMedia.Source.MANUAL)
        self.assertEqual(item.tipus, InscripcioMedia.Tipus.AUDIO)
        self.assertTrue(item.is_primary)

    def test_set_primary_and_delete_promotes_next_item(self):
        r1 = self._upload_media(self.ins.id, filename="first.mp3")
        self.assertEqual(r1.status_code, 200)
        r2 = self._upload_media(self.ins.id, filename="second.mp3")
        self.assertEqual(r2.status_code, 200)

        first = InscripcioMedia.objects.get(original_filename="first.mp3")
        second = InscripcioMedia.objects.get(original_filename="second.mp3")
        self.assertTrue(first.is_primary)
        self.assertFalse(second.is_primary)

        set_primary_url = reverse("inscripcions_media_set_primary", kwargs={"pk": self.comp.id})
        set_res = self.client.post(
            set_primary_url,
            data=json.dumps({"media_id": second.id}),
            content_type="application/json",
        )
        self.assertEqual(set_res.status_code, 200)

        first.refresh_from_db()
        second.refresh_from_db()
        self.assertFalse(first.is_primary)
        self.assertTrue(second.is_primary)

        delete_url = reverse("inscripcions_media_delete", kwargs={"pk": self.comp.id})
        del_res = self.client.post(
            delete_url,
            data=json.dumps({"media_id": second.id}),
            content_type="application/json",
        )
        self.assertEqual(del_res.status_code, 200)

        self.assertFalse(InscripcioMedia.objects.filter(id=second.id).exists())
        first.refresh_from_db()
        self.assertTrue(first.is_primary)

    def test_assisted_preview_and_apply_creates_assisted_media(self):
        preview_url = reverse("inscripcions_media_match_preview", kwargs={"pk": self.comp.id})
        preview_res = self.client.post(
            preview_url,
            data=json.dumps(
                {
                    "files": [
                        {
                            "key": "0",
                            "filename": "1 - -LUCIA POZO SANCHEZ-Collegi-Sagrat-Cor-Diputacio-GEN-F.mp3",
                            "relative_path": "music/1 - -LUCIA POZO SANCHEZ-Collegi-Sagrat-Cor-Diputacio-GEN-F.mp3",
                            "size": 1234,
                        }
                    ]
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(preview_res.status_code, 200)
        rows = preview_res.json().get("rows", [])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].get("suggested_inscripcio_id"), self.ins.id)

        apply_url = reverse("inscripcions_media_match_apply", kwargs={"pk": self.comp.id})
        media_file = SimpleUploadedFile(
            "1 - -LUCIA POZO SANCHEZ-Collegi-Sagrat-Cor-Diputacio-GEN-F.mp3",
            b"abc123",
            content_type="audio/mpeg",
        )
        apply_res = self.client.post(
            apply_url,
            data={
                "mapping_json": json.dumps(
                    [
                        {
                            "key": "0",
                            "inscripcio_id": self.ins.id,
                            "score": rows[0].get("score"),
                        }
                    ]
                ),
                "file_0": media_file,
            },
        )
        self.assertEqual(apply_res.status_code, 200)
        payload = apply_res.json()
        self.assertTrue(payload.get("ok"))
        self.assertEqual(payload.get("created_count"), 1)

        item = InscripcioMedia.objects.get(inscripcio=self.ins)
        self.assertEqual(item.source, InscripcioMedia.Source.ASSISTED)


class ScoringMediaPlaybackContextTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Media Notes")
        self.ins = self._create_inscripcio(self.comp, "Participant Reproductor")

        self.app = self._create_aparell("TRAMP_MEDIA", "Tramp Media")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        User = get_user_model()
        self.user = User.objects.create_user(
            username="scoring_media_owner",
            password="testpass123",
            email="scoring-media-owner@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.OWNER,
            is_active=True,
        )
        self.client.force_login(self.user)

    def test_media_context_orders_tracks_and_keeps_judge_video_separate(self):
        InscripcioMedia.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            fitxer=SimpleUploadedFile("a-main.mp3", b"aaa", content_type="audio/mpeg"),
            tipus=InscripcioMedia.Tipus.AUDIO,
            mime_type="audio/mpeg",
            original_filename="a-main.mp3",
            file_size_bytes=3,
            is_primary=True,
        )
        InscripcioMedia.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            fitxer=SimpleUploadedFile("a-alt.mp3", b"bbb", content_type="audio/mpeg"),
            tipus=InscripcioMedia.Tipus.AUDIO,
            mime_type="audio/mpeg",
            original_filename="a-alt.mp3",
            file_size_bytes=3,
            is_primary=False,
        )
        InscripcioMedia.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            fitxer=SimpleUploadedFile("v-main.mp4", b"1111", content_type="video/mp4"),
            tipus=InscripcioMedia.Tipus.VIDEO,
            mime_type="video/mp4",
            original_filename="v-main.mp4",
            file_size_bytes=4,
            is_primary=True,
        )
        InscripcioMedia.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            fitxer=SimpleUploadedFile("v-alt.mp4", b"2222", content_type="video/mp4"),
            tipus=InscripcioMedia.Tipus.VIDEO,
            mime_type="video/mp4",
            original_filename="v-alt.mp4",
            file_size_bytes=4,
            is_primary=False,
        )

        entry = ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=0,
        )
        ScoreEntryVideo.objects.create(
            score_entry=entry,
            video_file=SimpleUploadedFile("judge.mp4", b"3333", content_type="video/mp4"),
            status=ScoreEntryVideo.Status.READY,
            file_size_bytes=4,
            mime_type="video/mp4",
            original_filename="judge.mp4",
        )

        url = reverse("scoring_media_context", kwargs={"pk": self.comp.id})
        res = self.client.get(
            url,
            {
                "inscripcio_id": self.ins.id,
                "comp_aparell_id": self.comp_app.id,
                "exercici": 1,
            },
        )
        self.assertEqual(res.status_code, 200)
        payload = res.json()
        self.assertTrue(payload.get("ok"))

        media = payload.get("media", {})
        self.assertEqual((media.get("audio_primary") or {}).get("original_filename"), "a-main.mp3")
        self.assertEqual(
            [x.get("original_filename") for x in media.get("audio_others", [])],
            ["a-alt.mp3"],
        )
        self.assertEqual((media.get("video_primary") or {}).get("original_filename"), "v-main.mp4")
        self.assertEqual(
            [x.get("original_filename") for x in media.get("video_others", [])],
            ["v-alt.mp4"],
        )
        self.assertEqual((payload.get("judge_video") or {}).get("original_filename"), "judge.mp4")

    def test_media_context_rejects_foreign_comp_aparell(self):
        other_comp = self._create_competicio("Comp Altre Media")
        other_app = self._create_aparell("TRAMP_MEDIA_X", "Tramp Media X")
        other_comp_app = self._create_comp_aparell(other_comp, other_app, ordre=1, actiu=True)

        url = reverse("scoring_media_context", kwargs={"pk": self.comp.id})
        res = self.client.get(
            url,
            {
                "inscripcio_id": self.ins.id,
                "comp_aparell_id": other_comp_app.id,
                "exercici": 1,
            },
        )
        self.assertEqual(res.status_code, 400)

    def test_scoring_notes_home_context_includes_media_counts_and_judge_presence(self):
        ins_without_media = self._create_inscripcio(self.comp, "Participant Sense Media", ordre=2, grup=1)

        InscripcioMedia.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            fitxer=SimpleUploadedFile("ctx-audio.mp3", b"aaa", content_type="audio/mpeg"),
            tipus=InscripcioMedia.Tipus.AUDIO,
            mime_type="audio/mpeg",
            original_filename="ctx-audio.mp3",
            file_size_bytes=3,
            is_primary=True,
        )
        InscripcioMedia.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            fitxer=SimpleUploadedFile("ctx-video.mp4", b"bbbb", content_type="video/mp4"),
            tipus=InscripcioMedia.Tipus.VIDEO,
            mime_type="video/mp4",
            original_filename="ctx-video.mp4",
            file_size_bytes=4,
            is_primary=True,
        )

        entry = ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=0,
        )
        ScoreEntryVideo.objects.create(
            score_entry=entry,
            video_file=SimpleUploadedFile("ctx-judge.mp4", b"cccc", content_type="video/mp4"),
            status=ScoreEntryVideo.Status.READY,
            file_size_bytes=4,
            mime_type="video/mp4",
            original_filename="ctx-judge.mp4",
        )

        url = reverse("scoring_notes_home", kwargs={"pk": self.comp.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)

        media_counts = res.context["media_counts_by_inscripcio"]
        self.assertEqual(media_counts[str(self.ins.id)]["audio"], 1)
        self.assertEqual(media_counts[str(self.ins.id)]["video"], 1)
        self.assertEqual(media_counts[str(ins_without_media.id)]["audio"], 0)
        self.assertEqual(media_counts[str(ins_without_media.id)]["video"], 0)

        judge_map = res.context["judge_video_presence_by_key"]
        self.assertEqual(int(judge_map.get(f"{self.ins.id}|1|{self.comp_app.id}") or 0), 1)
        self.assertEqual(int(judge_map.get(f"{ins_without_media.id}|1|{self.comp_app.id}") or 0), 0)


class ScoringAndJudgeExclusionFlowTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Flux")
        self.app = self._create_aparell("TRAMP_FLOW", "Tramp Flow")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)
        self.comp_app.nombre_exercicis = 3
        self.comp_app.save(update_fields=["nombre_exercicis"])
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "fields": [
                    {
                        "label": "Execucio",
                        "code": "E",
                        "type": "matrix",
                        "shape": "judge_x_item",
                        "judges": {"count": 2},
                        "items": {"count": 5},
                        "decimals": 1,
                        "crash": {"enabled": True},
                    },
                    {
                        "label": "Altre",
                        "code": "X",
                        "type": "matrix",
                        "shape": "judge_x_item",
                        "judges": {"count": 2},
                        "items": {"count": 5},
                        "decimals": 1,
                        "crash": {"enabled": True},
                    },
                ],
                "computed": [],
            },
        )

        self.ins_allowed = self._create_inscripcio(self.comp, "Allowed", ordre=1)
        self.ins_blocked = self._create_inscripcio(self.comp, "Blocked", ordre=2)
        InscripcioAparellExclusio.objects.create(
            inscripcio=self.ins_blocked,
            comp_aparell=self.comp_app,
        )

        self.token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Judge A",
            permissions=[{"field_code": "E", "judge_index": 1}],
            can_record_video=True,
            is_active=True,
        )
        User = get_user_model()
        self.user = User.objects.create_user(
            username="scoring_exclusion_owner",
            password="testpass123",
            email="scoring-exclusion-owner@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.OWNER,
            is_active=True,
        )
        self.client.force_login(self.user)

    def test_scoring_save_partial_returns_403_for_excluded_inscripcio(self):
        url = reverse("scoring_save_partial", kwargs={"pk": self.comp.id})
        r = self.client.post(
            url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins_blocked.id,
                    "comp_aparell_id": self.comp_app.id,
                    "exercici": 1,
                    "inputs_patch": {},
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 403)

    def test_judge_portal_hides_excluded_and_save_returns_403(self):
        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url)
        self.assertEqual(portal_res.status_code, 200)
        body = portal_res.content.decode("utf-8")
        self.assertIn("Allowed", body)
        self.assertNotIn("Blocked", body)
        self.assertIn(reverse("judge_video_upload", kwargs={"token": self.token.id}), body)
        self.assertIn(reverse("judge_video_status", kwargs={"token": self.token.id}), body)
        self.assertIn(reverse("judge_video_delete", kwargs={"token": self.token.id}), body)

        save_url = reverse("judge_save_partial", kwargs={"token": self.token.id})
        save_res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins_blocked.id,
                    "exercici": 1,
                    "inputs_patch": {"E": 1},
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(save_res.status_code, 403)

    def test_judge_portal_supports_ex_query_and_multiex_payload(self):
        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url, {"ex": 2})
        self.assertEqual(portal_res.status_code, 200)

        payload = portal_res.context["scores_payload_json"][str(self.ins_allowed.id)]
        self.assertEqual(sorted(payload["exercises"].keys()), ["1", "2", "3"])

        body = portal_res.content.decode("utf-8")
        self.assertIn("const EXERCICI_HINT = 2;", body)
        self.assertIn('data-exercise-chip="1"', body)
        self.assertIn(f'id="editor-inner-{self.ins_allowed.id}-1"', body)
        self.assertIn(f'id="editor-inner-{self.ins_allowed.id}-2"', body)
        self.assertIn(f'id="editor-inner-{self.ins_allowed.id}-3"', body)
        self.assertIn("function getExerciseVisualState(insId, exercici)", body)
        self.assertIn("function getInsVisualState(insId)", body)
        self.assertIn("judge-nav-status", body)
        self.assertIn("judge-nav-ex-chip", body)
        self.assertIn(".judge-nav-link.is-complete", body)
        self.assertIn(".judge-nav-link.is-partial", body)
        self.assertNotIn(".judge-nav-link.is-saved", body)
        self.assertNotIn('href="?ex=1"', body)
        self.assertNotIn('href="?ex=3"', body)

    def test_judge_portal_renders_keyboard_navigation_helpers_for_score_inputs(self):
        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url, {"ex": 1})
        self.assertEqual(portal_res.status_code, 200)

        body = portal_res.content.decode("utf-8")
        self.assertIn("function getEditableInputs(insId, exercici)", body)
        self.assertIn("function focusNextEditableInput(currentInput)", body)
        self.assertIn("function focusPrevEditableInput(currentInput)", body)
        self.assertIn("function selectInputContents(input)", body)
        self.assertIn("function bindScoreInputNavigation(input, insId, exercici)", body)
        self.assertIn('input.dataset.insId = String(insId);', body)
        self.assertIn('input.dataset.exercici = exerciseKey(exercici);', body)
        self.assertIn('input.dataset.navScope = "score-input";', body)
        self.assertIn('input.addEventListener("keydown", (evt) => {', body)
        self.assertIn('if(evt.key !== "Enter" && evt.key !== "Tab") return;', body)
        self.assertIn('if(evt.key === "Tab" && evt.shiftKey){', body)
        self.assertIn("focusNextEditableInput(input);", body)
        self.assertIn("focusPrevEditableInput(input);", body)
        self.assertIn(".filter((input) => !input.disabled);", body)
        self.assertIn('[data-exercise-panel="1"][data-ins-id="${String(insId)}"][data-exercici="${exerciseKey(exercici)}"]', body)
        self.assertIn("selectInputContents(input);", body)
        self.assertIn("bindScoreInputNavigation(inp, insId, exercici);", body)
        self.assertNotIn("value.length", body)

    def test_judge_portal_nav_visual_status_uses_per_exercise_semantics(self):
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_allowed,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={"E": [0.2, 0.3, 0.4, 0.5, 0.6]},
            outputs={},
            total=1,
        )
        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url)
        self.assertEqual(portal_res.status_code, 200)

        payload = portal_res.context["scores_payload_json"][str(self.ins_allowed.id)]
        self.assertEqual(payload["exercises"]["1"]["inputs"]["E"], [0.2, 0.3, 0.4, 0.5, 0.6])
        self.assertEqual(payload["exercises"]["2"]["inputs"], {})
        self.assertEqual(payload["exercises"]["3"]["inputs"], {})

        body = portal_res.content.decode("utf-8")
        self.assertIn('data-nav-status="1"', body)
        self.assertIn('data-nav-ex-chip="1"', body)
        self.assertIn("Complet", body)
        self.assertIn("Parcial", body)
        self.assertIn("Pendent", body)

    def test_judge_portal_hides_video_controls_when_video_disabled(self):
        self.token.can_record_video = False
        self.token.save(update_fields=["can_record_video"])
        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url, {"ex": 1})
        self.assertEqual(portal_res.status_code, 200)
        body = portal_res.content.decode("utf-8")
        self.assertNotIn('id="video-rec-btn-', body)
        self.assertIn("Gravacio desactivada per aquest QR.", body)

    def test_judge_save_partial_accepts_crash_for_authorized_field(self):
        save_url = reverse("judge_save_partial", kwargs={"token": self.token.id})
        save_res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins_allowed.id,
                    "exercici": 1,
                    "inputs_patch": {
                        "E": [0.2, 0.3, 0.4, 0.5, 0.6],
                        "__crash__E": [3, 2],
                    },
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(save_res.status_code, 200)
        payload = save_res.json()
        self.assertTrue(payload.get("ok"))
        self.assertEqual(payload.get("inputs", {}).get("__crash__E", [None])[0], 3)

        entry = ScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins_allowed,
            exercici=1,
        )
        self.assertEqual(entry.inputs.get("__crash__E", [None, None])[0], 3)
        self.assertEqual(entry.inputs.get("__crash__E", [None, None])[1], 0)

    def test_judge_save_partial_rejects_crash_for_unauthorized_field(self):
        save_url = reverse("judge_save_partial", kwargs={"token": self.token.id})
        save_res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins_allowed.id,
                    "exercici": 1,
                    "inputs_patch": {"__crash__X": [2, 0]},
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(save_res.status_code, 403)

    def test_judge_save_partial_clamps_exercici_to_aparell_max(self):
        save_url = reverse("judge_save_partial", kwargs={"token": self.token.id})
        save_res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins_allowed.id,
                    "exercici": 99,
                    "inputs_patch": {
                        "E": [0.2, 0.3, 0.4, 0.5, 0.6],
                    },
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(save_res.status_code, 200)
        entry = ScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins_allowed,
            exercici=3,
        )
        self.assertIsNotNone(entry)
        self.assertFalse(
            ScoreEntry.objects.filter(
                competicio=self.comp,
                comp_aparell=self.comp_app,
                inscripcio=self.ins_allowed,
                exercici=99,
            ).exists()
        )

    def test_scoring_updates_omits_excluded_entries(self):
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_allowed,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=1,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_blocked,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=2,
        )

        since = (timezone.now() - timedelta(minutes=10)).isoformat()
        url = reverse("scoring_updates", kwargs={"pk": self.comp.id})
        r = self.client.get(url, {"since": since})

        self.assertEqual(r.status_code, 200)
        payload = r.json()
        updated_ids = {u["inscripcio_id"] for u in payload.get("updates", [])}

        self.assertIn(self.ins_allowed.id, updated_ids)
        self.assertNotIn(self.ins_blocked.id, updated_ids)

    def test_judge_updates_accept_multiple_exercicis(self):
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_allowed,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={"E": [0.2, 0, 0, 0, 0]},
            outputs={},
            total=1,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_allowed,
            exercici=2,
            comp_aparell=self.comp_app,
            inputs={"E": [0.4, 0, 0, 0, 0]},
            outputs={},
            total=2,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_blocked,
            exercici=2,
            comp_aparell=self.comp_app,
            inputs={"E": [0.6, 0, 0, 0, 0]},
            outputs={},
            total=3,
        )

        since = (timezone.now() - timedelta(minutes=10)).isoformat()
        url = reverse("judge_updates", kwargs={"token": self.token.id})
        res = self.client.get(url, {"since": since, "exercici": [1, 2]})

        self.assertEqual(res.status_code, 200)
        payload = res.json()
        updates = payload.get("updates", [])
        self.assertEqual({u["exercici"] for u in updates}, {1, 2})
        self.assertEqual({u["inscripcio_id"] for u in updates}, {self.ins_allowed.id})


class ProgrammedGroupReconfigurationTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Programmed Group Reconfig")
        self.comp.group_by_default = ["entitat"]
        self.comp.save(update_fields=["group_by_default"])
        User = get_user_model()
        self.user = User.objects.create_user(
            username="programmed_group_editor",
            password="testpass123",
            email="programmed-group-editor@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        self.client.force_login(self.user)

    def _attach_rotation_to_group(self, group):
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="09:00",
            hora_fi="09:30",
            ordre=1,
            titol="Franja 1",
        )
        estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="descans",
            ordre=1,
            actiu=True,
        )
        assignacio = RotacioAssignacio.objects.create(
            competicio=self.comp,
            franja=franja,
            estacio=estacio,
        )
        RotacioAssignacioGrup.objects.create(assignacio=assignacio, grup=group, ordre=1)

    def test_make_independent_group_with_rotations_creates_new_group_when_origin_keeps_members(self):
        first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A1",
            entitat="Club A",
            ordre_sortida=1,
            grup=1,
        )
        second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A2",
            entitat="Club A",
            ordre_sortida=2,
            grup=1,
        )
        third = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="B1",
            entitat="Club B",
            ordre_sortida=3,
            grup=1,
        )
        self._attach_rotation_to_group(first.grup_competicio)

        url = reverse("inscripcions_list", kwargs={"pk": self.comp.id})
        res = self.client.get(
            url,
            {
                "make_independent_group": "1",
                "lvl": "g1",
                "v1": "Club A",
            },
            follow=True,
        )
        self.assertEqual(res.status_code, 200)

        first.refresh_from_db()
        second.refresh_from_db()
        third.refresh_from_db()
        self.assertEqual(first.grup, 2)
        self.assertEqual(second.grup, 2)
        self.assertEqual(third.grup, 1)
        self.assertContains(res, "Creat el grup 2")

    def test_make_independent_group_with_rotations_rejects_emptying_programmed_origin(self):
        first = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A1",
            entitat="Club A",
            ordre_sortida=1,
            grup=1,
        )
        second = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="A2",
            entitat="Club A",
            ordre_sortida=2,
            grup=1,
        )
        self._attach_rotation_to_group(first.grup_competicio)

        url = reverse("inscripcions_list", kwargs={"pk": self.comp.id})
        res = self.client.get(
            url,
            {
                "make_independent_group": "1",
                "lvl": "g1",
                "v1": "Club A",
            },
            follow=True,
        )
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, "No es pot deixar buit un grup inclos al programa de rotacions")

        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.grup, 1)
        self.assertEqual(second.grup, 1)


class RotationOrderingDisplayTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Rotations Display")
        self.app_prior = self._create_aparell("ROT_PRIOR", "Rotation Prior")
        self.app = self._create_aparell("ROT_APP", "Rotation App")
        self.comp_app_prior = self._create_comp_aparell(self.comp, self.app_prior, ordre=1, actiu=True)
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        self.ins_1 = self._create_inscripcio(self.comp, "Participant 1", ordre=1, grup=1)
        self.ins_2 = self._create_inscripcio(self.comp, "Participant 2", ordre=2, grup=1)
        self.ins_3 = self._create_inscripcio(self.comp, "Participant 3", ordre=3, grup=1)

        group = self.ins_1.grup_competicio
        Inscripcio.objects.filter(pk=self.ins_1.pk).update(ordre_competicio=2)
        Inscripcio.objects.filter(pk=self.ins_2.pk).update(ordre_competicio=1)
        Inscripcio.objects.filter(pk=self.ins_3.pk).update(ordre_competicio=3)
        self.ins_1.refresh_from_db()
        self.ins_2.refresh_from_db()
        self.ins_3.refresh_from_db()

        self.franja_1 = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="09:00",
            hora_fi="09:30",
            ordre=1,
            titol="Franja 1",
        )
        self.franja_2 = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="09:30",
            hora_fi="10:00",
            ordre=2,
            titol="Franja 2",
        )
        self.franja_3 = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="10:00",
            hora_fi="10:30",
            ordre=3,
            titol="Franja 3",
        )
        self.estacio_prior = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app_prior,
            ordre=1,
            actiu=True,
        )
        self.estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=2,
            actiu=True,
        )
        assignacio_0 = RotacioAssignacio.objects.create(
            competicio=self.comp,
            franja=self.franja_1,
            estacio=self.estacio_prior,
        )
        RotacioAssignacioGrup.objects.create(assignacio=assignacio_0, grup=group, ordre=1)
        assignacio = RotacioAssignacio.objects.create(
            competicio=self.comp,
            franja=self.franja_2,
            estacio=self.estacio,
        )
        RotacioAssignacioGrup.objects.create(assignacio=assignacio, grup=group, ordre=1)
        assignacio_2 = RotacioAssignacio.objects.create(
            competicio=self.comp,
            franja=self.franja_3,
            estacio=self.estacio,
        )
        RotacioAssignacioGrup.objects.create(assignacio=assignacio_2, grup=group, ordre=1)

        self.comp.inscripcions_view = {
            "rotacions_order_modes": {
                str(self.franja_2.id): "rotate",
                str(self.franja_3.id): "rotate",
            }
        }
        self.comp.save(update_fields=["inscripcions_view"])

        self.token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Judge Rotation",
            permissions=[{"field_code": "E", "judge_index": 1}],
            is_active=True,
        )
        User = get_user_model()
        self.user = User.objects.create_user(
            username="rotation_display_owner",
            password="testpass123",
            email="rotation-display-owner@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.OWNER,
            is_active=True,
        )
        self.client.force_login(self.user)

    def _assign_group_to_app_franja(self, group, franja):
        assignacio, _created = RotacioAssignacio.objects.get_or_create(
            competicio=self.comp,
            franja=franja,
            estacio=self.estacio,
        )
        next_ordre = (
            RotacioAssignacioGrup.objects
            .filter(assignacio=assignacio)
            .count() + 1
        )
        RotacioAssignacioGrup.objects.create(assignacio=assignacio, grup=group, ordre=next_ordre)
        return assignacio

    def test_judge_portal_uses_first_app_franja_order_by_default_and_allows_override(self):
        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url)
        self.assertEqual(portal_res.status_code, 200)
        self.assertIsNone(portal_res.context["franja_override_id"])

        block = portal_res.context["group_blocks"][0]
        self.assertEqual(block["franja_id"], self.franja_2.id)
        self.assertEqual(
            [ins.nom_i_cognoms for ins in block["list"]],
            ["Participant 1", "Participant 3", "Participant 2"],
        )
        self.assertEqual(
            [ins.rotation_order_display for ins in block["list"]],
            [1, 2, 3],
        )

        body = portal_res.content.decode("utf-8")
        self.assertIn("Participant 1", body)
        self.assertIn("Ordre 1", body)
        self.assertIn("Ordre 2", body)
        self.assertIn("Base 3", body)

        third_res = self.client.get(portal_url, {"franja": self.franja_3.id})
        self.assertEqual(third_res.status_code, 200)
        self.assertEqual(third_res.context["franja_override_id"], self.franja_3.id)

        third_block = third_res.context["group_blocks"][0]
        self.assertEqual(third_block["franja_id"], self.franja_3.id)
        self.assertEqual(
            [ins.nom_i_cognoms for ins in third_block["list"]],
            ["Participant 3", "Participant 2", "Participant 1"],
        )
        self.assertEqual(
            [ins.rotation_order_display for ins in third_block["list"]],
            [1, 2, 3],
        )

    def test_judge_portal_shows_all_programmed_groups_with_each_group_ordered_by_own_franja(self):
        extra_1 = self._create_inscripcio(self.comp, "Participant 4", ordre=4, grup=2)
        extra_2 = self._create_inscripcio(self.comp, "Participant 5", ordre=5, grup=2)
        Inscripcio.objects.filter(pk=extra_1.pk).update(ordre_competicio=1)
        Inscripcio.objects.filter(pk=extra_2.pk).update(ordre_competicio=2)
        extra_1.refresh_from_db()
        extra_2.refresh_from_db()
        self._assign_group_to_app_franja(extra_1.grup_competicio, self.franja_3)

        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url)
        self.assertEqual(portal_res.status_code, 200)

        blocks = portal_res.context["group_blocks"]
        self.assertEqual(
            [block["key"] for block in blocks],
            [self.ins_1.grup_competicio_id, extra_1.grup_competicio_id],
        )
        self.assertEqual(blocks[0]["franja_id"], self.franja_2.id)
        self.assertEqual(
            [ins.nom_i_cognoms for ins in blocks[0]["list"]],
            ["Participant 1", "Participant 3", "Participant 2"],
        )
        self.assertEqual(blocks[1]["franja_id"], self.franja_3.id)
        self.assertEqual(
            [ins.nom_i_cognoms for ins in blocks[1]["list"]],
            ["Participant 5", "Participant 4"],
        )

    def test_judge_portal_franja_override_only_affects_groups_present_in_that_franja(self):
        extra_1 = self._create_inscripcio(self.comp, "Participant 6", ordre=6, grup=3)
        extra_2 = self._create_inscripcio(self.comp, "Participant 7", ordre=7, grup=3)
        Inscripcio.objects.filter(pk=extra_1.pk).update(ordre_competicio=1)
        Inscripcio.objects.filter(pk=extra_2.pk).update(ordre_competicio=2)
        extra_1.refresh_from_db()
        extra_2.refresh_from_db()
        self._assign_group_to_app_franja(extra_1.grup_competicio, self.franja_2)

        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url, {"franja": self.franja_3.id})
        self.assertEqual(portal_res.status_code, 200)

        blocks = portal_res.context["group_blocks"]
        self.assertEqual(
            [block["key"] for block in blocks],
            [self.ins_1.grup_competicio_id, extra_1.grup_competicio_id],
        )
        self.assertEqual(blocks[0]["franja_id"], self.franja_3.id)
        self.assertEqual(
            [ins.nom_i_cognoms for ins in blocks[0]["list"]],
            ["Participant 3", "Participant 2", "Participant 1"],
        )
        self.assertEqual(blocks[1]["franja_id"], self.franja_2.id)
        self.assertEqual(
            [ins.nom_i_cognoms for ins in blocks[1]["list"]],
            ["Participant 7", "Participant 6"],
        )

    def test_judge_portal_uses_group_query_for_active_group(self):
        extra_1 = self._create_inscripcio(self.comp, "Participant 4", ordre=4, grup=2)
        extra_2 = self._create_inscripcio(self.comp, "Participant 5", ordre=5, grup=2)
        Inscripcio.objects.filter(pk=extra_1.pk).update(ordre_competicio=1)
        Inscripcio.objects.filter(pk=extra_2.pk).update(ordre_competicio=2)
        extra_1.refresh_from_db()
        self._assign_group_to_app_franja(extra_1.grup_competicio, self.franja_3)

        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url, {"group": extra_1.grup_competicio_id})
        self.assertEqual(portal_res.status_code, 200)
        self.assertEqual(portal_res.context["active_group_key"], extra_1.grup_competicio_id)

        body = portal_res.content.decode("utf-8")
        self.assertRegex(
            body,
            rf'class="nav-link active"\s+data-group-target="group-{extra_1.grup_competicio_id}"',
        )
        self.assertRegex(
            body,
            rf'class="group-pane\s*"\s+data-group-pane="group-{extra_1.grup_competicio_id}"',
        )

    def test_judge_portal_invalid_group_query_falls_back_to_first_visible_group(self):
        extra_1 = self._create_inscripcio(self.comp, "Participant 4", ordre=4, grup=2)
        extra_2 = self._create_inscripcio(self.comp, "Participant 5", ordre=5, grup=2)
        Inscripcio.objects.filter(pk=extra_1.pk).update(ordre_competicio=1)
        Inscripcio.objects.filter(pk=extra_2.pk).update(ordre_competicio=2)
        extra_1.refresh_from_db()
        self._assign_group_to_app_franja(extra_1.grup_competicio, self.franja_3)

        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url, {"group": 999999})
        self.assertEqual(portal_res.status_code, 200)
        self.assertEqual(portal_res.context["active_group_key"], portal_res.context["group_blocks"][0]["key"])

    def test_judge_portal_preserves_group_and_initial_exercise_with_franja(self):
        self.comp_app.nombre_exercicis = 3
        self.comp_app.save(update_fields=["nombre_exercicis"])
        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(
            portal_url,
            {
                "ex": 2,
                "group": self.ins_1.grup_competicio_id,
                "franja": self.franja_3.id,
            },
        )
        self.assertEqual(portal_res.status_code, 200)
        self.assertEqual(portal_res.context["active_group_key"], self.ins_1.grup_competicio_id)
        self.assertEqual(portal_res.context["franja_override_id"], self.franja_3.id)

        body = portal_res.content.decode("utf-8")
        self.assertIn("const EXERCICI_HINT = 2;", body)
        self.assertIn('data-exercise-chip="1"', body)
        self.assertNotIn("data-exercise-link", body)

    def test_scoring_notes_home_supports_franja_selection_for_programmed_groups(self):
        scoring_url = reverse("scoring_notes_home", kwargs={"pk": self.comp.id})

        default_res = self.client.get(scoring_url)
        self.assertEqual(default_res.status_code, 200)
        default_body = default_res.content.decode("utf-8")
        self.assertNotIn('id="franjaSelect"', default_body)
        self.assertNotIn(">Franja</label>", default_body)
        self.assertIsNone(default_res.context["franja_selected_id"])

        scoring_res = self.client.get(scoring_url, {"franja": self.franja_3.id})
        self.assertEqual(scoring_res.status_code, 200)
        self.assertEqual(scoring_res.context["franja_selected_id"], self.franja_3.id)

        app_key = str(self.comp_app.id)
        self.assertEqual(
            scoring_res.context["rotation_groups_by_app"].get(app_key),
            [self.ins_1.grup_competicio_id],
        )

        rank_map = scoring_res.context["rotation_rank_map"]
        self.assertEqual(rank_map.get(f"{self.comp_app.id}|{self.ins_3.id}"), 1)
        self.assertEqual(rank_map.get(f"{self.comp_app.id}|{self.ins_2.id}"), 2)
        self.assertEqual(rank_map.get(f"{self.comp_app.id}|{self.ins_1.id}"), 3)

    def test_out_of_program_visibility_toggle_controls_notes_and_judge_views(self):
        extra_group = GrupCompeticio.objects.create(
            competicio=self.comp,
            display_num=2,
            legacy_num=2,
            nom="Fora Programa",
            actiu=True,
        )
        extra_ins = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Participant Fora Programa",
            ordre_sortida=4,
            grup=2,
            grup_competicio=extra_group,
            ordre_competicio=1,
        )
        ungrouped_ins = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Participant Sense Grup",
            ordre_sortida=5,
            grup=0,
            grup_competicio=None,
            ordre_competicio=1,
        )

        scoring_url = reverse("scoring_notes_home", kwargs={"pk": self.comp.id})
        scoring_res = self.client.get(scoring_url)
        self.assertEqual(scoring_res.status_code, 200)
        self.assertFalse(scoring_res.context["show_out_of_program_in_competition_views"])
        self.assertEqual([g for g, _rows in scoring_res.context["groups"]], [self.ins_1.grup_competicio_id, 0])
        self.assertEqual(scoring_res.context["out_of_program_groups"], [])
        scoring_body = scoring_res.content.decode("utf-8")
        self.assertIn("Sense grup", scoring_body)
        self.assertNotIn("Fora Programa", scoring_body)

        planner_url = reverse("rotacions_planner", kwargs={"pk": self.comp.id})
        planner_res = self.client.get(planner_url)
        self.assertEqual(planner_res.status_code, 200)
        self.assertEqual(planner_res.context["out_of_program_groups_count"], 1)
        self.assertEqual(planner_res.context["out_of_program_members_total"], 1)
        self.assertFalse(planner_res.context["show_out_of_program_in_competition_views"])
        planner_body = planner_res.content.decode("utf-8")
        self.assertIn("Fora de programa", planner_body)
        self.assertIn("Fora Programa", planner_body)

        portal_url = reverse("judge_portal", kwargs={"token": self.token.id})
        portal_res = self.client.get(portal_url)
        self.assertEqual(portal_res.status_code, 200)
        self.assertFalse(portal_res.context["show_out_of_program_in_competition_views"])
        self.assertEqual([block["key"] for block in portal_res.context["group_blocks"]], [self.ins_1.grup_competicio_id, 0])
        self.assertEqual(portal_res.context["out_of_program_group_blocks"], [])
        portal_body = portal_res.content.decode("utf-8")
        self.assertIn(ungrouped_ins.nom_i_cognoms, portal_body)
        self.assertIn("Sense grup", portal_body)
        self.assertNotIn(extra_ins.nom_i_cognoms, portal_body)

        grouped_portal_res = self.client.get(portal_url, {"group": 0})
        self.assertEqual(grouped_portal_res.status_code, 200)
        self.assertEqual(grouped_portal_res.context["active_group_key"], 0)

        toggle_url = reverse("rotacions_out_of_program_visibility_save", kwargs={"pk": self.comp.id})
        toggle_res = self.client.post(
            toggle_url,
            data=json.dumps({"value": True}),
            content_type="application/json",
        )
        self.assertEqual(toggle_res.status_code, 200)
        self.assertJSONEqual(
            toggle_res.content.decode("utf-8"),
            {"ok": True, "value": True},
        )
        self.comp.refresh_from_db()
        self.assertTrue(
            self.comp.inscripcions_view.get("show_out_of_program_in_competition_views")
        )

        scoring_res = self.client.get(scoring_url)
        self.assertEqual(scoring_res.status_code, 200)
        self.assertTrue(scoring_res.context["show_out_of_program_in_competition_views"])
        self.assertEqual([g for g, _rows in scoring_res.context["groups"]], [self.ins_1.grup_competicio_id, 0])
        self.assertEqual([g for g, _rows in scoring_res.context["out_of_program_groups"]], [extra_group.id])
        scoring_body = scoring_res.content.decode("utf-8")
        self.assertIn("Fora de programa", scoring_body)
        self.assertIn("Fora Programa", scoring_body)

        portal_res = self.client.get(portal_url)
        self.assertEqual(portal_res.status_code, 200)
        self.assertTrue(portal_res.context["show_out_of_program_in_competition_views"])
        self.assertEqual([block["key"] for block in portal_res.context["group_blocks"]], [self.ins_1.grup_competicio_id, 0])
        self.assertEqual([block["key"] for block in portal_res.context["out_of_program_group_blocks"]], [extra_group.id])
        portal_body = portal_res.content.decode("utf-8")
        self.assertIn("Fora de programa", portal_body)
        self.assertIn(extra_ins.nom_i_cognoms, portal_body)


class ClassificacioMatrixScalarTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Classificacio")
        self.app_a = self._create_aparell("APP_A", "Aparell A")
        self.app_b = self._create_aparell("APP_B", "Aparell B")
        self.comp_app_a = self._create_comp_aparell(self.comp, self.app_a, ordre=1, actiu=True)
        self.comp_app_b = self._create_comp_aparell(self.comp, self.app_b, ordre=2, actiu=True)

        self.ins_a = self._create_inscripcio(self.comp, "Participant A", ordre=1)
        self.ins_b = self._create_inscripcio(self.comp, "Participant B", ordre=2)

    def _base_cfg_schema(self):
        return {
            "particions": [],
            "filtres": {},
            "puntuacio": {
                "aparells": {"mode": "tots", "ids": []},
                "camps_per_aparell": {
                    str(self.comp_app_a.id): ["X"],
                    str(self.comp_app_b.id): ["Y"],
                },
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "exercicis_best_n": 1,
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
                "camp": "total",
                "agregacio": "sum",
                "best_n": 1,
            },
            "desempat": [],
            "presentacio": {"top_n": 0, "mostrar_empats": True},
        }

    def _valid_partition_schema(self):
        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["total"]}
        return schema

    def _selected_total_schema(self, app_ids=None, mode_resultat="score"):
        selected = list(app_ids or [self.comp_app_a.id, self.comp_app_b.id])
        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": selected}
        schema["puntuacio"]["camps_per_aparell"] = {str(app_id): ["total"] for app_id in selected}
        schema["puntuacio"]["mode_resultat_aparells"] = mode_resultat
        schema["puntuacio"]["victories"] = {
            "punts_victoria": 1,
            "punts_empat": 0.5,
            "sense_nota_mode": "skip",
            "mode_camps": "agregat",
            "mode_exercicis": "agregat",
            "mode_seleccio_exercicis_camps_separats": "per_camp",
            "agregacio_victories_camps": "sum",
            "agregacio_victories_exercicis": "sum",
            "desempat_comparacio": [],
        }
        return schema

    def test_compute_classificacio_uses_input_matrix_1x1_as_scalar(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 1}}
                ],
                "computed": [],
            },
        )
        ScoringSchema.objects.create(
            aparell=self.app_b,
            schema={
                "fields": [
                    {"code": "Y", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 1}}
                ],
                "computed": [],
            },
        )

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"X": [[7.5]]},
            outputs={},
            total=0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_b,
            inputs={"Y": [[8.2]]},
            outputs={},
            total=0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Global A/B",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=self._base_cfg_schema(),
        )
        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        points_by_name = {r["participant"]: r["punts"] for r in rows}

        self.assertEqual(points_by_name.get("Participant A"), 7.5)
        self.assertEqual(points_by_name.get("Participant B"), 8.2)

    def test_compute_classificacio_keeps_zero_for_non_1x1_matrix(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 2}}
                ],
                "computed": [],
            },
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"X": [[7.0, 8.0]]},
            outputs={},
            total=0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["X"]}
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="No 1x1",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )
        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["participant"], "Participant A")
        self.assertEqual(rows[0]["punts"], 0.0)

    def test_classificacio_save_rejects_non_1x1_matrix_field(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 2}, "items": {"count": 1}}
                ],
                "computed": [],
            },
        )

        payload = {
            "nom": "Cfg invalida",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["X"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("1x1" in e for e in body.get("errors", [])))

    def test_compute_classificacio_uses_input_list_1_as_scalar(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "L", "type": "list", "shape": "judge", "judges": {"count": 1}}
                ],
                "computed": [],
            },
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"L": [7.2]},
            outputs={},
            total=0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["L"]}
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Llista 1x1",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )
        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["participant"], "Participant A")
        self.assertEqual(rows[0]["punts"], 7.2)

    def test_classificacio_save_rejects_non_scalar_computed_main_field(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 2}}
                ],
                "computed": [
                    {"code": "X_copy", "label": "X copy", "formula": "X"},
                ],
            },
        )

        payload = {
            "nom": "Cfg computed no escalar",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["X_copy"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("X_copy" in e and "no es puntuable directament" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_non_scalar_computed_tie_field(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 2}}
                ],
                "computed": [
                    {"code": "X_copy", "label": "X copy", "formula": "X"},
                ],
            },
        )

        payload = {
            "nom": "Cfg tie computed no escalar",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["X_copy"],
                        "ordre": "desc",
                        "scope": {
                            "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                            "exercicis": {"mode": "hereta"},
                        },
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("desempat[0]" in e and "X_copy" in e and "no es puntuable directament" in e for e in body.get("errors", [])))

    def test_classificacio_save_requires_real_camps_for_selected_apps(self):
        payload = {
            "nom": "Cfg sense camps reals",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("camps_per_aparell" in e and "camp real" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_invalid_raw_column_field(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "E_total", "label": "Execucio", "type": "number"},
                ],
                "computed": [],
            },
        )

        payload = {
            "nom": "Cfg raw invalid",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["E_total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {
                    "top_n": 0,
                    "mostrar_empats": True,
                    "columnes": [
                        {
                            "type": "raw",
                            "key": "raw_exec",
                            "label": "Exec",
                            "align": "right",
                            "decimals": 3,
                            "source": {
                                "aparell_id": self.comp_app_a.id,
                                "exercici": 1,
                                "camp": "EX",
                                "jutges": {"ids": []},
                            },
                        }
                    ],
                },
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("presentacio.columnes[0]" in e and "'EX'" in e for e in body.get("errors", [])))

    def test_classificacio_save_keeps_tie_schema_canonical_without_legacy_camp(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "E_total", "label": "Execucio", "type": "number"},
                ],
                "computed": [],
            },
        )

        payload = {
            "nom": "Cfg canonical tie",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["E_total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["E_total"],
                        "agregacio_camps": "hereta",
                        "ordre": "desc",
                        "scope": {
                            "aparells": {"mode": "hereta"},
                            "exercicis": {"mode": "hereta"},
                        },
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=res.json()["id"])
        self.assertNotIn("camp", cfg.schema.get("puntuacio") or {})
        self.assertNotIn("agregacio", cfg.schema.get("puntuacio") or {})
        self.assertNotIn("best_n", cfg.schema.get("puntuacio") or {})
        self.assertEqual((cfg.schema.get("desempat") or [])[0].get("camps"), ["E_total"])
        self.assertNotIn("camp", (cfg.schema.get("desempat") or [])[0])

    def test_classificacio_save_accepts_row_compute_by_judge_when_single_judge(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 3}}
            ],
            "computed": [
                {
                    "code": "E_by_judge",
                    "label": "E by judge",
                    "formula": "row_custom_compute('E', '1 - x', return_mode='by_judge')",
                },
            ],
        }
        meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        self.assertTrue(meta.get("E_by_judge", {}).get("scoreable"))

    def test_classificacio_save_rejects_row_compute_by_judge_when_multiple_judges(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 2}, "items": {"count": 3}}
            ],
            "computed": [
                {
                    "code": "E_by_judge",
                    "label": "E by judge",
                    "formula": "row_custom_compute('E', '1 - x', return_mode='by_judge')",
                },
            ],
        }
        meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        self.assertFalse(meta.get("E_by_judge", {}).get("scoreable"))
        self.assertIn("by_judge", str(meta.get("E_by_judge", {}).get("reason") or ""))

    def test_classificacio_save_accepts_column_compute_by_item_when_count_is_one(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 2}, "items": {"count": 4}}
            ],
            "computed": [
                {
                    "code": "E_by_item",
                    "label": "E by item",
                    "formula": "column_custom_compute('E', '1 - x', return_mode='by_item', count=1)",
                },
            ],
        }
        meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        self.assertTrue(meta.get("E_by_item", {}).get("scoreable"))

    def test_classificacio_save_rejects_column_compute_by_item_when_multiple_items(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 2}, "items": {"count": 4}}
            ],
            "computed": [
                {
                    "code": "E_by_item",
                    "label": "E by item",
                    "formula": "column_custom_compute('E', '1 - x', return_mode='by_item')",
                },
            ],
        }
        meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        self.assertFalse(meta.get("E_by_item", {}).get("scoreable"))
        self.assertIn("by_item", str(meta.get("E_by_item", {}).get("reason") or ""))

    def test_classificacio_save_tie_accepts_row_compute_by_judge_when_single_judge(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 3}}
            ],
            "computed": [
                {
                    "code": "E_by_judge",
                    "label": "E by judge",
                    "formula": "row_custom_compute('E', '1 - x', return_mode='by_judge')",
                },
            ],
        }
        strict_meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        ui_meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=False)
        self.assertTrue(strict_meta.get("E_by_judge", {}).get("scoreable"))
        self.assertTrue(ui_meta.get("E_by_judge", {}).get("scoreable"))

    def test_classificacio_save_rejects_invalid_tie_exercicis_selection_mode(self):
        payload = {
            "nom": "Cfg tie invalida",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["total"],
                        "ordre": "desc",
                        "scope": {
                            "aparells": {"mode": "tots"},
                            "exercicis": {"mode": "hereta"},
                        },
                        "mode_seleccio_exercicis": "invalid_mode",
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("desempat[0].mode_seleccio_exercicis" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_main_aparells_mode_tots(self):
        payload = {
            "nom": "Cfg main tots invalid",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "tots", "ids": []},
                    "camps_per_aparell": {
                        str(self.comp_app_a.id): ["total"],
                        str(self.comp_app_b.id): ["total"],
                    },
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("puntuacio.aparells.mode='tots'" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_negative_exercicis_max_per_participant(self):
        User = get_user_model()
        user = User.objects.create_user(
            username="classif_max_pp_editor",
            password="testpass123",
            email="classif-max-pp-editor@example.com",
        )
        CompeticioMembership.objects.create(
            user=user,
            competicio=self.comp,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        self.client.force_login(user)

        payload = {
            "nom": "Cfg max per participant invalid",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 2, "max_per_participant": -1},
                    "exercicis_best_n": 2,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("puntuacio.exercicis.max_per_participant" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_tie_aparells_mode_tots(self):
        payload = {
            "nom": "Cfg tie tots invalid",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["total"],
                        "ordre": "desc",
                        "scope": {
                            "aparells": {"mode": "tots"},
                            "exercicis": {"mode": "hereta"},
                        },
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("desempat[0].scope.aparells.mode='tots'" in e for e in body.get("errors", [])))

    def test_compute_classificacio_tie_supports_per_app_override_exercicis(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        self.comp_app_b.nombre_exercicis = 2
        self.comp_app_b.save(update_fields=["nombre_exercicis"])

        # Participant A
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=1.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_b,
            inputs={},
            outputs={},
            total=5.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=2,
            comp_aparell=self.comp_app_b,
            inputs={},
            outputs={},
            total=5.0,
        )

        # Participant B
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=6.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=4.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_b,
            inputs={},
            outputs={},
            total=8.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=2,
            comp_aparell=self.comp_app_b,
            inputs={},
            outputs={},
            total=2.0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "tots", "ids": []}
        schema["puntuacio"]["camps_per_aparell"] = {
            str(self.comp_app_a.id): ["total"],
            str(self.comp_app_b.id): ["total"],
        }
        schema["puntuacio"]["exercicis"] = {"mode": "tots"}
        schema["puntuacio"]["agregacio_camps"] = "sum"
        schema["puntuacio"]["agregacio_exercicis"] = "sum"
        schema["puntuacio"]["agregacio_aparells"] = "sum"
        schema["desempat"] = [
            {
                "camps": ["total"],
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "tots"},
                    "exercicis": {"mode": "hereta"},
                },
                "agregacio_camps": "sum",
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "mode_seleccio_exercicis": "per_aparell_override",
                "exercicis_per_aparell": {
                    str(self.comp_app_a.id): {"mode": "millor_1"},
                    str(self.comp_app_b.id): {"mode": "pitjor_1"},
                },
            }
        ]

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Tie per app override",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["participant"], "Participant A")
        self.assertEqual(rows[1]["participant"], "Participant B")

    def test_compute_classificacio_main_exercicis_respects_max_per_participant(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=6.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=5.0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["total"]}
        schema["puntuacio"]["exercicis"] = {
            "mode": "millor_n",
            "best_n": 2,
            "max_per_participant": 1,
        }
        schema["puntuacio"]["agregacio_camps"] = "sum"
        schema["puntuacio"]["agregacio_exercicis"] = "sum"
        schema["puntuacio"]["agregacio_aparells"] = "sum"

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Main max per participant",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        self.assertEqual(len(rows), 2)
        points_by_name = {r["participant"]: r["punts"] for r in rows}
        self.assertEqual(points_by_name.get("Participant A"), 9.0)
        self.assertEqual(points_by_name.get("Participant B"), 6.0)

    def test_compute_classificacio_tie_exercicis_respects_max_per_participant(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=6.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=4.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=7.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=3.0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["total"]}
        schema["puntuacio"]["exercicis"] = {"mode": "tots"}
        schema["puntuacio"]["agregacio_camps"] = "sum"
        schema["puntuacio"]["agregacio_exercicis"] = "sum"
        schema["puntuacio"]["agregacio_aparells"] = "sum"
        schema["desempat"] = [
            {
                "camps": ["total"],
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "exercicis": {"mode": "millor_n", "best_n": 2, "max_per_participant": 1},
                },
                "agregacio_camps": "sum",
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
            }
        ]

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Tie max per participant",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["participant"], "Participant B")
        self.assertEqual(rows[1]["participant"], "Participant A")

    def test_compute_classificacio_supports_custom_partition_groups_for_categoria(self):
        self.ins_a.categoria = "ALEVI"
        self.ins_a.save(update_fields=["categoria"])
        self.ins_b.categoria = "PREBENJAMI"
        self.ins_b.save(update_fields=["categoria"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )

        schema = self._valid_partition_schema()
        schema["particions"] = ["categoria"]
        schema["particions_custom"] = {
            "categoria": {
                "mode": "custom",
                "grups": [
                    {"key": "base", "label": "Base", "values": ["ALEVI", "PREBENJAMI"]},
                ],
            }
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Particio custom categoria",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        self.assertEqual(list(out.keys()), ["categoria:Base"])
        self.assertEqual(len(out["categoria:Base"]), 2)

    def test_compute_classificacio_supports_partition_by_schema_extra_field(self):
        self.comp.inscripcions_schema = {
            "columns": [
                {"code": "nivell", "label": "Nivell", "kind": "extra"},
            ]
        }
        self.comp.save(update_fields=["inscripcions_schema"])

        self.ins_a.extra = {"nivell": "N1"}
        self.ins_a.save(update_fields=["extra"])
        self.ins_b.extra = {"nivell": "N2"}
        self.ins_b.save(update_fields=["extra"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )

        schema = self._valid_partition_schema()
        schema["particions"] = ["nivell"]
        schema["particions_custom"] = {
            "nivell": {
                "mode": "custom",
                "grups": [
                    {"key": "bloc_1", "label": "Bloc 1", "values": ["N1", "N2"]},
                ],
            }
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Particio extra",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        self.assertEqual(list(out.keys()), ["nivell:Bloc 1"])
        self.assertEqual(len(out["nivell:Bloc 1"]), 2)

    def test_compute_classificacio_supports_conditional_partition_on_parent_group(self):
        ins_c = self._create_inscripcio(self.comp, "Participant C", ordre=3)

        self.ins_a.categoria = "ALEVI"
        self.ins_a.subcategoria = "N1"
        self.ins_a.save(update_fields=["categoria", "subcategoria"])
        self.ins_b.categoria = "PREBENJAMI"
        self.ins_b.subcategoria = "N2"
        self.ins_b.save(update_fields=["categoria", "subcategoria"])
        ins_c.categoria = "INFANTIL"
        ins_c.subcategoria = "N3"
        ins_c.save(update_fields=["categoria", "subcategoria"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=ins_c,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=7.0,
        )

        schema = self._valid_partition_schema()
        schema["particions"] = ["categoria", "subcategoria"]
        schema["particions_v2"] = [
            {"code": "categoria", "apply_mode": "all"},
            {"code": "subcategoria", "apply_mode": "some_parents", "parent_values": ["Base"]},
        ]
        schema["particions_custom"] = {
            "categoria": {
                "mode": "custom",
                "grups": [
                    {"key": "base", "label": "Base", "values": ["ALEVI", "PREBENJAMI"]},
                    {"key": "grans", "label": "Grans", "values": ["INFANTIL"]},
                ],
            }
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Particio condicional",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        self.assertEqual(
            set(out.keys()),
            {"categoria:Base|subcategoria:N1", "categoria:Base|subcategoria:N2", "categoria:Grans"},
        )
        self.assertEqual(out["categoria:Grans"][0]["participant"], "Participant C")

    def test_normalize_particions_schema_populates_particions_v2_from_legacy_list(self):
        schema = self._valid_partition_schema()
        schema["particions"] = ["categoria", "subcategoria"]

        normalized = _normalize_particions_schema(schema)
        self.assertEqual(normalized.get("particions"), ["categoria", "subcategoria"])
        self.assertEqual(
            normalized.get("particions_v2"),
            [
                {"code": "categoria", "apply_mode": "all", "parent_values": []},
                {"code": "subcategoria", "apply_mode": "all", "parent_values": []},
            ],
        )

    def test_compute_classificacio_partitions_by_birth_year_ranges(self):
        self.ins_a.data_naixement = date(2008, 5, 4)
        self.ins_b.data_naixement = date(2011, 2, 10)
        self.ins_a.save(update_fields=["data_naixement"])
        self.ins_b.save(update_fields=["data_naixement"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )

        schema = self._valid_partition_schema()
        schema["particions"] = ["any_naixement_forquilla"]
        schema["particions_v2"] = [
            {"code": "any_naixement_forquilla", "apply_mode": "all", "parent_values": []},
        ]
        schema["particions_config"] = {
            "any_naixement_forquilla": {
                "ranges": [
                    {"label": "2007-2009", "from_year": 2007, "to_year": 2009},
                    {"label": "2010-2012", "from_year": 2010, "to_year": 2012},
                ],
                "sense_data_label": "Sense data",
                "fora_rang_label": "Fora de forquilla",
            }
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Forquilles naixement",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        self.assertEqual(
            set(out.keys()),
            {"any_naixement_forquilla:2007-2009", "any_naixement_forquilla:2010-2012"},
        )
        self.assertEqual(out["any_naixement_forquilla:2007-2009"][0]["participant"], "Participant A")
        self.assertEqual(out["any_naixement_forquilla:2010-2012"][0]["participant"], "Participant B")

    def test_classificacio_save_rejects_overlapping_birth_year_ranges(self):
        schema = self._valid_partition_schema()
        schema["particions"] = ["any_naixement_forquilla"]
        schema["particions_v2"] = [
            {"code": "any_naixement_forquilla", "apply_mode": "all", "parent_values": []},
        ]
        schema["particions_config"] = {
            "any_naixement_forquilla": {
                "ranges": [
                    {"label": "2007-2009", "from_year": 2007, "to_year": 2009},
                    {"label": "2009-2011", "from_year": 2009, "to_year": 2011},
                ]
            }
        }

        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertTrue(any("solapament" in e for e in errors))

    def test_classificacio_save_rejects_birth_year_ranges_for_team_rankings(self):
        schema = self._valid_partition_schema()
        schema["particions"] = ["any_naixement_forquilla"]
        schema["particions_v2"] = [
            {"code": "any_naixement_forquilla", "apply_mode": "all", "parent_values": []},
        ]
        schema["particions_config"] = {
            "any_naixement_forquilla": {
                "ranges": [
                    {"label": "2007-2009", "from_year": 2007, "to_year": 2009},
                ]
            }
        }

        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="equips")
        self.assertTrue(any("nomes es valid per classificacions individuals" in e for e in errors))

    def test_validate_particions_schema_rejects_conditional_partition_without_parent_values(self):
        schema = self._valid_partition_schema()
        schema["particions_v2"] = [
            {"code": "categoria", "apply_mode": "all"},
            {"code": "subcategoria", "apply_mode": "some_parents", "parent_values": []},
        ]
        normalized = _normalize_particions_schema(schema)
        errors = _validate_particions_schema(self.comp, normalized)
        self.assertTrue(any("parent_values" in e for e in errors))

    def test_classificacio_save_rejects_unknown_partition_field(self):
        payload = {
            "nom": "Cfg particio desconeguda",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": self._valid_partition_schema(),
        }
        payload["schema"]["particions"] = ["camp_inexistent"]
        payload["schema"]["particions_custom"] = {
            "camp_inexistent": {
                "mode": "custom",
                "grups": [{"label": "X", "values": ["A"]}],
            }
        }

        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("camp no perm" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_duplicate_custom_partition_values(self):
        payload = {
            "nom": "Cfg particio custom duplicada",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": self._valid_partition_schema(),
        }
        payload["schema"]["particions"] = ["categoria"]
        payload["schema"]["particions_custom"] = {
            "categoria": {
                "mode": "custom",
                "grups": [
                    {"label": "Bloc 1", "values": ["ALEVI"]},
                    {"label": "Bloc 2", "values": ["ALEVI"]},
                ],
            }
        }

        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("valor repetit entre grups" in e for e in body.get("errors", [])))

    def test_compute_classificacio_victories_single_app_matches_order(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=7.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories 1 app",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([r["participant"] for r in rows], ["Participant A", "Participant B"])
        self.assertEqual(rows[0]["by_app_base"][self.comp_app_a.id], 9.0)
        self.assertEqual(rows[0]["by_app"][self.comp_app_a.id], 1.0)
        self.assertEqual(rows[0]["punts"], 1.0)
        self.assertEqual(rows[1]["by_app"][self.comp_app_a.id], 0.0)
        self.assertEqual(rows[1]["punts"], 0.0)

    def test_compute_classificacio_victories_multiple_apps_aggregates_after_duels(self):
        ins_c = self._create_inscripcio(self.comp, "Participant C", ordre=3)
        schema = self._selected_total_schema(
            [self.comp_app_a.id, self.comp_app_b.id],
            mode_resultat="victories",
        )

        scores = {
            self.ins_a.id: {self.comp_app_a.id: 100.0, self.comp_app_b.id: 1.0},
            self.ins_b.id: {self.comp_app_a.id: 60.0, self.comp_app_b.id: 60.0},
            ins_c.id: {self.comp_app_a.id: 59.0, self.comp_app_b.id: 59.0},
        }
        for ins_id, per_app in scores.items():
            ins = Inscripcio.objects.get(pk=ins_id)
            for app_id, total in per_app.items():
                comp_app = self.comp_app_a if app_id == self.comp_app_a.id else self.comp_app_b
                ScoreEntry.objects.create(
                    competicio=self.comp,
                    inscripcio=ins,
                    exercici=1,
                    comp_aparell=comp_app,
                    inputs={},
                    outputs={},
                    total=total,
                )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories 2 apps",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([r["participant"] for r in rows], ["Participant B", "Participant A", "Participant C"])
        self.assertEqual(rows[0]["punts"], 3.0)
        self.assertEqual(rows[1]["punts"], 2.0)
        self.assertEqual(rows[2]["punts"], 1.0)

    def test_compute_classificacio_victories_internal_tie_break_uses_compare_tie(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["desempat_comparacio"] = [
            {
                "camp": "E",
                "camps": ["E"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "scope": {"exercicis": {"mode": "hereta"}},
            }
        ]

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"E": 9.0},
            outputs={},
            total=10.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"E": 8.0},
            outputs={},
            total=10.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories compare tie",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([r["participant"] for r in rows], ["Participant A", "Participant B"])
        self.assertEqual(rows[0]["punts"], 1.0)
        self.assertEqual(rows[1]["punts"], 0.0)

    def test_compute_classificacio_victories_unresolved_tie_gives_half_points(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=10.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=10.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories tied",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual(rows[0]["punts"], 0.5)
        self.assertEqual(rows[1]["punts"], 0.5)
        self.assertEqual(rows[0]["by_app"][self.comp_app_a.id], 0.5)
        self.assertEqual(rows[1]["by_app"][self.comp_app_a.id], 0.5)

    def test_compute_classificacio_victories_separated_fields_aggregate_per_app(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["E", "D"]}
        schema["puntuacio"]["victories"]["mode_camps"] = "separat"
        schema["puntuacio"]["victories"]["agregacio_victories_camps"] = "sum"

        for exercici, e_val, d_val in ((1, 10.0, 1.0), (2, 10.0, 1.0)):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={"E": e_val, "D": d_val},
                outputs={},
                total=e_val + d_val,
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={"E": 8.0, "D": 8.0},
                outputs={},
                total=16.0,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories fields separated",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        by_name = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in rows}
        self.assertEqual(by_name.get("Participant A"), 1.0)
        self.assertEqual(by_name.get("Participant B"), 1.0)

    def test_compute_classificacio_victories_separated_exercises_aggregate_per_app(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["mode_exercicis"] = "separat"
        schema["puntuacio"]["victories"]["agregacio_victories_exercicis"] = "sum"

        for exercici, total_a, total_b in ((1, 10.0, 9.0), (2, 1.0, 2.0)):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={},
                outputs={},
                total=total_a,
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={},
                outputs={},
                total=total_b,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories exercises separated",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        by_name = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in rows}
        self.assertEqual(by_name.get("Participant A"), 1.0)
        self.assertEqual(by_name.get("Participant B"), 1.0)

    def test_compute_classificacio_victories_separated_fields_and_exercises_use_intermediate_aggregation(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["E", "D"]}
        schema["puntuacio"]["victories"]["mode_camps"] = "separat"
        schema["puntuacio"]["victories"]["mode_exercicis"] = "separat"
        schema["puntuacio"]["victories"]["agregacio_victories_camps"] = "avg"
        schema["puntuacio"]["victories"]["agregacio_victories_exercicis"] = "sum"

        per_ex = {
            1: {"a": {"E": 10.0, "D": 1.0}, "b": {"E": 9.0, "D": 9.0}},
            2: {"a": {"E": 8.0, "D": 1.0}, "b": {"E": 7.0, "D": 9.0}},
        }
        for exercici, vals in per_ex.items():
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs=vals["a"],
                outputs={},
                total=sum(vals["a"].values()),
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs=vals["b"],
                outputs={},
                total=sum(vals["b"].values()),
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories separated all levels",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        by_name = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in rows}
        self.assertEqual(by_name.get("Participant A"), 1.0)
        self.assertEqual(by_name.get("Participant B"), 1.0)

    def test_compute_classificacio_victories_separated_fields_supports_global_or_per_field_exercise_selection(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["E", "D"]}
        schema["puntuacio"]["exercicis"] = {"mode": "millor_1"}
        schema["puntuacio"]["victories"]["mode_camps"] = "separat"
        schema["puntuacio"]["victories"]["agregacio_victories_camps"] = "sum"

        for exercici, vals_a, vals_b in (
            (1, {"E": 10.0, "D": 1.0}, {"E": 9.0, "D": 9.0}),
            (2, {"E": 1.0, "D": 10.0}, {"E": 8.0, "D": 8.0}),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs=vals_a,
                outputs={},
                total=sum(vals_a.values()),
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs=vals_b,
                outputs={},
                total=sum(vals_b.values()),
            )

        schema_global = json.loads(json.dumps(schema))
        schema_global["puntuacio"]["victories"]["mode_seleccio_exercicis_camps_separats"] = "global"
        cfg_global = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories fields global select",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema_global,
        )
        global_rows = compute_classificacio(self.comp, cfg_global).get("global", [])
        global_points = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in global_rows}

        schema_per_field = json.loads(json.dumps(schema))
        schema_per_field["puntuacio"]["victories"]["mode_seleccio_exercicis_camps_separats"] = "per_camp"
        cfg_per_field = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories fields per field select",
            activa=True,
            ordre=2,
            tipus="individual",
            schema=schema_per_field,
        )
        per_field_rows = compute_classificacio(self.comp, cfg_per_field).get("global", [])
        per_field_points = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in per_field_rows}

        self.assertEqual(global_points.get("Participant A"), 1.0)
        self.assertEqual(global_points.get("Participant B"), 1.0)
        self.assertEqual(per_field_points.get("Participant A"), 2.0)
        self.assertEqual(per_field_points.get("Participant B"), 0.0)

    def test_compute_classificacio_victories_internal_tie_break_respects_separated_exercise_unit(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["mode_exercicis"] = "separat"
        schema["puntuacio"]["victories"]["desempat_comparacio"] = [
            {
                "camp": "E",
                "camps": ["E"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "scope": {"exercicis": {"mode": "hereta"}},
            }
        ]

        for exercici, e_a, e_b in ((1, 9.0, 8.0), (2, 1.0, 9.0)):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={"E": e_a},
                outputs={},
                total=10.0,
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={"E": e_b},
                outputs={},
                total=10.0,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories tie per exercise",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        by_name = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in rows}
        self.assertEqual(by_name.get("Participant A"), 1.0)
        self.assertEqual(by_name.get("Participant B"), 1.0)

    def test_classificacio_save_rejects_invalid_victories_granular_modes(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["mode_camps"] = "invalid"
        schema["puntuacio"]["victories"]["mode_exercicis"] = "broken"
        schema["puntuacio"]["victories"]["mode_seleccio_exercicis_camps_separats"] = "oops"
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertTrue(any("mode_camps invalid" in e for e in errors))
        self.assertTrue(any("mode_exercicis invalid" in e for e in errors))
        self.assertTrue(any("mode_seleccio_exercicis_camps_separats invalid" in e for e in errors))

    def test_classificacio_save_ignores_invalid_victories_granular_modes_in_score_mode(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="score")
        schema["puntuacio"]["victories"]["mode_camps"] = "invalid"
        schema["puntuacio"]["victories"]["mode_exercicis"] = "broken"
        schema["puntuacio"]["victories"]["mode_seleccio_exercicis_camps_separats"] = "oops"
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertFalse(any("mode_camps invalid" in e for e in errors))
        self.assertFalse(any("mode_exercicis invalid" in e for e in errors))
        self.assertFalse(any("mode_seleccio_exercicis_camps_separats invalid" in e for e in errors))

    def test_classificacio_save_rejects_victories_for_non_individual(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="entitat")
        self.assertTrue(any("mode_resultat_aparells='victories'" in e for e in errors))

    def test_classificacio_save_rejects_victories_compare_tie_scope_aparells(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["desempat_comparacio"] = [
            {
                "camps": ["total"],
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "exercicis": {"mode": "hereta"},
                },
            }
        ]
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertTrue(any("scope.aparells no esta permes" in e for e in errors))

    def test_classificacio_save_rejects_victories_compare_tie_scope_participants(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["desempat_comparacio"] = [
            {
                "camps": ["total"],
                "ordre": "desc",
                "scope": {
                    "participants": {"mode": "tots"},
                    "exercicis": {"mode": "hereta"},
                },
            }
        ]
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertTrue(any("scope.participants no esta permes" in e for e in errors))


class ClassificacionsExportExcelTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Export Classificacions")
        self.app = self._create_aparell("TRAMP_EXPORT", "Tramp Export")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        self.ins_a = self._create_inscripcio(self.comp, "Participant A", ordre=1)
        self.ins_b = self._create_inscripcio(self.comp, "Participant B", ordre=2)
        self.ins_c = self._create_inscripcio(self.comp, "Participant C", ordre=3)

        self.ins_a.categoria = "Junior"
        self.ins_a.subcategoria = "Femeni"
        self.ins_a.save(update_fields=["categoria", "subcategoria"])

        self.ins_b.categoria = "Junior"
        self.ins_b.subcategoria = "Femeni"
        self.ins_b.save(update_fields=["categoria", "subcategoria"])

        self.ins_c.categoria = "Senior"
        self.ins_c.subcategoria = "Masculi"
        self.ins_c.save(update_fields=["categoria", "subcategoria"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=9.6,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=8.3,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_c,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=7.1,
        )

        self.cfg_general = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="General",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=self._schema_for_parts([]),
        )
        self.cfg_particions = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Per categories",
            activa=True,
            ordre=2,
            tipus="individual",
            schema=self._schema_for_parts(["categoria", "subcategoria"]),
        )

        User = get_user_model()
        self.user = User.objects.create_user(
            username="classif_export_user",
            password="testpass123",
            email="classif-export@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.READONLY,
            is_active=True,
        )

    def _schema_for_parts(self, parts):
        return {
            "particions": parts,
            "filtres": {},
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "exercicis_best_n": 1,
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
                "camp": "total",
                "agregacio": "sum",
                "best_n": 1,
            },
            "desempat": [],
            "presentacio": {
                "top_n": 0,
                "mostrar_empats": True,
                "columnes": [
                    {"type": "builtin", "key": "posicio", "label": "Pos.", "align": "left"},
                    {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                ],
            },
        }

    def test_export_excel_creates_one_sheet_per_classificacio(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url)

        self.assertEqual(res.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            res["Content-Type"],
        )

        wb = load_workbook(filename=BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), 2)

        ws_general = wb[wb.sheetnames[0]]
        ws_parts = wb[wb.sheetnames[1]]

        self.assertIn("General", str(ws_general["A1"].value))
        self.assertIn("Particio:", str(ws_general["A4"].value))
        self.assertEqual(ws_general.freeze_panes, "A6")

        gold_fill = (ws_general["A6"].fill.fgColor.rgb or "").upper()
        silver_fill = (ws_general["A7"].fill.fgColor.rgb or "").upper()
        header_fill = (ws_general["A5"].fill.fgColor.rgb or "").upper()
        self.assertTrue(gold_fill.endswith("F6E27A"))
        self.assertTrue(silver_fill.endswith("E3E8EF"))
        self.assertTrue(header_fill.endswith("E9EEF7"))

        self.assertIn("Particio:", str(ws_parts["A4"].value))
        self.assertIn("/", str(ws_parts["A4"].value))

    def test_export_excel_can_filter_single_cfg_with_cfg_id(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url, {"cfg_id": self.cfg_particions.id})

        self.assertEqual(res.status_code, 200)
        wb = load_workbook(filename=BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), 1)
        ws = wb[wb.sheetnames[0]]
        self.assertIn("Per categories", str(ws["A1"].value))

    def test_export_excel_rejects_invalid_cfg_id(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url, {"cfg_id": "abc"})
        self.assertEqual(res.status_code, 400)


class ClassificacioTemplateFlowTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp_source = self._create_competicio("Comp Templates Source")
        self.comp_target = self._create_competicio("Comp Templates Target")

        self.app = self._create_aparell("TPL_APP", "Template App")
        self.source_app = self._create_comp_aparell(self.comp_source, self.app, ordre=1, actiu=True)
        self.target_app = self._create_comp_aparell(self.comp_target, self.app, ordre=1, actiu=True)

        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "E_total", "type": "number"},
                ],
                "computed": [],
            },
        )

        self.cfg_source = ClassificacioConfig.objects.create(
            competicio=self.comp_source,
            nom="Cfg Source",
            activa=True,
            ordre=1,
            tipus="individual",
            schema={
                "particions": ["categoria"],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.source_app.id]},
                    "camps_per_aparell": {str(self.source_app.id): ["E_total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "mode_seleccio_exercicis": "per_aparell_global",
                    "exercicis_per_aparell": {},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                    "camp": "total",
                    "agregacio": "sum",
                    "best_n": 1,
                },
                "desempat": [],
                "presentacio": {
                    "top_n": 0,
                    "mostrar_empats": True,
                    "columnes": [
                        {"type": "builtin", "key": "posicio", "label": "#", "align": "left"},
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                    ],
                },
            },
        )

        User = get_user_model()
        self.editor_user = User.objects.create_user(
            username="tpl_editor_user",
            password="testpass123",
            email="tpl-editor@example.com",
        )
        self.other_user = User.objects.create_user(
            username="tpl_other_user",
            password="testpass123",
            email="tpl-other@example.com",
        )

        CompeticioMembership.objects.create(
            user=self.editor_user,
            competicio=self.comp_source,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.editor_user,
            competicio=self.comp_target,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.other_user,
            competicio=self.comp_source,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.other_user,
            competicio=self.comp_target,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )

    def _post_json_as(self, user, url_name, comp_id, payload):
        self.client.force_login(user)
        url = reverse(url_name, kwargs={"pk": comp_id})
        return self.client.post(url, data=json.dumps(payload), content_type="application/json")

    def test_editor_with_classificacions_edit_can_save_global_template(self):
        res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL 1"},
        )
        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.json().get("ok"))

    def test_template_schema_helpers_preserve_victories_config(self):
        schema = json.loads(json.dumps(self.cfg_source.schema or {}))
        schema["puntuacio"]["mode_resultat_aparells"] = "victories"
        schema["puntuacio"]["victories"] = {
            "punts_victoria": 1,
            "punts_empat": 0,
            "sense_nota_mode": "skip",
            "mode_camps": "separat",
            "mode_exercicis": "separat",
            "mode_seleccio_exercicis_camps_separats": "global",
            "agregacio_victories_camps": "avg",
            "agregacio_victories_exercicis": "max",
            "desempat_comparacio": [
                {
                    "camp": "E_total",
                    "camps": ["E_total"],
                    "agregacio_camps": "hereta",
                    "ordre": "desc",
                    "scope": {"exercicis": {"mode": "hereta"}},
                }
            ],
        }

        schema_tpl, warnings = _schema_to_template_schema(self.comp_source, schema)
        self.assertFalse(warnings)
        self.assertEqual(
            ((schema_tpl.get("puntuacio") or {}).get("mode_resultat_aparells")),
            "victories",
        )

        schema_local, mapping_warnings, mapping = _template_schema_to_competicio_schema(self.comp_target, schema_tpl)
        self.assertFalse(mapping_warnings)
        self.assertEqual(mapping.get(self.app.codi), self.target_app.id)
        punt = (schema_local.get("puntuacio") or {})
        self.assertEqual(punt.get("mode_resultat_aparells"), "victories")
        self.assertEqual((punt.get("aparells") or {}).get("ids"), [self.target_app.id])
        self.assertEqual((punt.get("victories") or {}).get("mode_camps"), "separat")
        self.assertEqual((punt.get("victories") or {}).get("mode_exercicis"), "separat")
        self.assertEqual(
            (punt.get("victories") or {}).get("mode_seleccio_exercicis_camps_separats"),
            "global",
        )
        self.assertEqual(
            ((((punt.get("victories") or {}).get("desempat_comparacio")) or [])[0] or {}).get("camps"),
            ["E_total"],
        )

    def test_global_template_can_be_saved_validated_and_applied(self):
        save_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Rapid"},
        )
        self.assertEqual(save_res.status_code, 200)
        save_body = save_res.json()
        self.assertTrue(save_body.get("ok"))
        template_id = save_body.get("template", {}).get("id")
        self.assertTrue(template_id)

        tpl = ClassificacioTemplateGlobal.objects.get(pk=template_id)
        tpl_schema = (tpl.payload or {}).get("schema") or {}
        self.assertEqual(
            ((tpl_schema.get("puntuacio") or {}).get("aparells") or {}).get("ids"),
            [self.app.codi],
        )

        validate_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_validate",
            self.comp_target.id,
            {"template_id": template_id},
        )
        self.assertEqual(validate_res.status_code, 200)
        validate_body = validate_res.json()
        self.assertTrue(validate_body.get("compatible"))

        apply_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {"template_id": template_id, "nom": "Aplicada Target", "activa": False},
        )
        self.assertEqual(apply_res.status_code, 200)
        apply_body = apply_res.json()
        self.assertTrue(apply_body.get("ok"))
        cfg = apply_body.get("cfg") or {}
        self.assertEqual(cfg.get("nom"), "Aplicada Target")

        punt = ((cfg.get("schema") or {}).get("puntuacio") or {})
        self.assertEqual((punt.get("aparells") or {}).get("ids"), [self.target_app.id])
        self.assertEqual((punt.get("camps_per_aparell") or {}).get(str(self.target_app.id)), ["E_total"])

        tpl.refresh_from_db()
        self.assertEqual(tpl.uses_count, 1)
        self.assertIsNotNone(tpl.last_used_at)

    def test_template_apply_requires_ack_for_non_strict_fallback(self):
        save_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Ack"},
        )
        self.assertEqual(save_res.status_code, 200)
        template_id = save_res.json().get("template", {}).get("id")
        self.assertTrue(template_id)

        res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {
                "template_id": template_id,
                "nom": "Aplicada sense ack",
                "activa": False,
                "fallback_mode": "assistit",
            },
        )
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertEqual(body.get("phase"), "assistit")
        self.assertIn("confirmar", str(body.get("error", "")).lower())

    def test_template_apply_fallback_chain_strict_assistit_force(self):
        save_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Fallback Chain"},
        )
        self.assertEqual(save_res.status_code, 200)
        template_id = save_res.json().get("template", {}).get("id")
        self.assertTrue(template_id)

        tpl = ClassificacioTemplateGlobal.objects.get(pk=template_id)
        payload = json.loads(json.dumps(tpl.payload or {}))
        schema = payload.get("schema") or {}
        schema["particions"] = ["categoria"]
        schema["particions_custom"] = {
            "categoria": {
                "mode": "custom",
                "grups": [
                    {"label": "Bloc 1", "values": ["ALEVI"]},
                    {"label": "Bloc 2", "values": ["ALEVI"]},
                ],
            }
        }
        payload["schema"] = schema
        tpl.payload = payload
        tpl.save(update_fields=["payload", "updated_at"])

        strict_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {
                "template_id": template_id,
                "nom": "Aplicada strict",
                "activa": False,
                "fallback_mode": "strict",
            },
        )
        self.assertEqual(strict_res.status_code, 400)
        strict_body = strict_res.json()
        self.assertFalse(strict_body.get("compatible"))
        self.assertEqual(strict_body.get("phase"), "strict")
        self.assertEqual(strict_body.get("next_fallback"), "assistit")
        self.assertTrue(strict_body.get("can_try_next"))

        assistit_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {
                "template_id": template_id,
                "nom": "Aplicada assistit",
                "activa": False,
                "fallback_mode": "assistit",
                "ack_warning": True,
            },
        )
        self.assertEqual(assistit_res.status_code, 400)
        assistit_body = assistit_res.json()
        self.assertFalse(assistit_body.get("compatible"))
        self.assertEqual(assistit_body.get("phase"), "assistit")
        self.assertEqual(assistit_body.get("next_fallback"), "force")
        self.assertTrue(assistit_body.get("can_try_next"))

        force_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {
                "template_id": template_id,
                "nom": "Aplicada force",
                "activa": False,
                "fallback_mode": "force",
                "ack_warning": True,
            },
        )
        self.assertEqual(force_res.status_code, 200)
        force_body = force_res.json()
        self.assertTrue(force_body.get("ok"))
        cfg = force_body.get("cfg") or {}
        self.assertEqual(cfg.get("nom"), "Aplicada force")
        self.assertEqual(
            (((cfg.get("schema") or {}).get("puntuacio") or {}).get("aparells") or {}).get("ids"),
            [self.target_app.id],
        )
        self.assertEqual(
            (((cfg.get("schema") or {}).get("puntuacio") or {}).get("camps_per_aparell") or {}).get(str(self.target_app.id)),
            ["total"],
        )

    def test_template_list_only_shows_owner_templates(self):
        own = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Owner"},
        )
        self.assertEqual(own.status_code, 200)
        own_id = own.json().get("template", {}).get("id")
        self.assertTrue(own_id)

        foreign = self._post_json_as(
            self.other_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Foreign"},
        )
        self.assertEqual(foreign.status_code, 200)
        foreign_id = foreign.json().get("template", {}).get("id")
        self.assertTrue(foreign_id)

        self.client.force_login(self.editor_user)
        list_url = reverse("classificacio_template_list", kwargs={"pk": self.comp_source.id})
        res = self.client.get(list_url)
        self.assertEqual(res.status_code, 200)
        ids = {int(t["id"]) for t in (res.json().get("templates") or [])}
        self.assertIn(int(own_id), ids)
        self.assertNotIn(int(foreign_id), ids)

    def test_cannot_use_or_update_foreign_template_by_id(self):
        foreign = self._post_json_as(
            self.other_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Foreign Locked"},
        )
        self.assertEqual(foreign.status_code, 200)
        foreign_id = foreign.json().get("template", {}).get("id")
        self.assertTrue(foreign_id)

        validate_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_validate",
            self.comp_target.id,
            {"template_id": foreign_id},
        )
        self.assertEqual(validate_res.status_code, 404)

        apply_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {"template_id": foreign_id, "nom": "No hauria d'aplicar"},
        )
        self.assertEqual(apply_res.status_code, 404)

        update_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "template_id": foreign_id, "nom": "No hauria d'editar"},
        )
        self.assertEqual(update_res.status_code, 404)

    def test_classificacions_home_renders_builder_json_contract(self):
        self.client.force_login(self.editor_user)
        url = reverse("classificacions_home", kwargs={"pk": self.comp_source.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, 'id="can-manage-global-templates"')
        self.assertContains(res, 'id="builder-save-url"')
        self.assertContains(res, 'id="builder-delete-url-pattern"')
        self.assertContains(res, 'id="builder-preview-url-pattern"')
        self.assertContains(res, 'id="builder-enable-template-library"')
        self.assertContains(res, 'id="builder-can-preview"')
        self.assertContains(res, 'id="victoryConfigBox"')
        self.assertContains(res, 'id="sVictoryModeCamps"')
        self.assertContains(res, 'id="sVictoryModeExercicis"')
        self.assertContains(res, 'id="classifHelpDrawer"')
        self.assertContains(res, 'id="classif-builder-back-to-top"')
        self.assertContains(res, "classificacions_builder_help.css")
        self.assertContains(res, "classificacions_builder_help.js")
        self.assertContains(res, 'data-help-key="global_overview"')
        self.assertContains(res, 'data-help-key="victories_overview"')
        self.assertNotContains(res, '<option value="entitat">Per entitat</option>', html=True)

    def test_classificacions_home_sanitizes_legacy_field_refs_for_builder(self):
        schema = json.loads(json.dumps(self.cfg_source.schema or {}))
        schema["puntuacio"]["camps_per_aparell"] = {
            str(self.source_app.id): ["E_total", "ex"],
        }
        schema["desempat"] = [
            {
                "camp": "ex",
                "camps": ["ex"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "hereta"},
                    "exercicis": {"mode": "hereta"},
                },
            },
            {
                "camp": "E_total",
                "camps": ["E_total"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "hereta"},
                    "exercicis": {"mode": "hereta"},
                },
            },
        ]
        schema["puntuacio"]["victories"] = {
            "punts_victoria": 1,
            "punts_empat": 0.5,
            "sense_nota_mode": "skip",
            "mode_camps": "agregat",
            "mode_exercicis": "agregat",
            "mode_seleccio_exercicis_camps_separats": "per_camp",
            "agregacio_victories_camps": "sum",
            "agregacio_victories_exercicis": "sum",
            "desempat_comparacio": [
                {
                    "camp": "ex",
                    "camps": ["ex"],
                    "agregacio_camps": "hereta",
                    "ordre": "desc",
                    "scope": {"exercicis": {"mode": "hereta"}},
                },
                {
                    "camp": "E_total",
                    "camps": ["E_total"],
                    "agregacio_camps": "hereta",
                    "ordre": "desc",
                    "scope": {"exercicis": {"mode": "hereta"}},
                },
            ],
        }
        schema["presentacio"]["columnes"] = [
            {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
            {
                "type": "raw",
                "key": "raw_valid",
                "label": "Valid",
                "align": "right",
                "decimals": 3,
                "source": {
                    "aparell_id": self.source_app.id,
                    "exercici": 1,
                    "camp": "E_total",
                    "jutges": {"ids": []},
                },
            },
            {
                "type": "raw",
                "key": "raw_legacy",
                "label": "Legacy",
                "align": "right",
                "decimals": 3,
                "source": {
                    "aparell_id": self.source_app.id,
                    "exercici": 1,
                    "camp": "ex",
                    "jutges": {"ids": []},
                },
            },
        ]
        self.cfg_source.schema = schema
        self.cfg_source.save(update_fields=["schema"])

        self.client.force_login(self.editor_user)
        url = reverse("classificacions_home", kwargs={"pk": self.comp_source.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)

        cfg_payload = next(cfg for cfg in res.context["cfgs"] if cfg["id"] == self.cfg_source.id)
        sanitized = cfg_payload["schema"]
        punt = sanitized.get("puntuacio") or {}
        self.assertNotIn("camp", punt)
        self.assertNotIn("agregacio", punt)
        self.assertNotIn("best_n", punt)
        self.assertEqual((punt.get("camps_per_aparell") or {}).get(str(self.source_app.id)), ["E_total"])

        desempat = sanitized.get("desempat") or []
        self.assertEqual(len(desempat), 1)
        self.assertEqual(desempat[0].get("camps"), ["E_total"])
        self.assertNotIn("camp", desempat[0])

        compare = (((punt.get("victories") or {}).get("desempat_comparacio")) or [])
        self.assertEqual(len(compare), 1)
        self.assertEqual(compare[0].get("camps"), ["E_total"])
        self.assertNotIn("camp", compare[0])

        raw_columns = [c for c in ((sanitized.get("presentacio") or {}).get("columnes") or []) if c.get("type") == "raw"]
        self.assertEqual(len(raw_columns), 1)
        self.assertEqual((((raw_columns[0].get("source") or {}).get("camp"))), "E_total")

        self.cfg_source.refresh_from_db()
        original = self.cfg_source.schema or {}
        self.assertEqual((original.get("puntuacio") or {}).get("camps_per_aparell", {}).get(str(self.source_app.id)), ["E_total", "ex"])
        self.assertEqual(len(original.get("desempat") or []), 2)
        self.assertEqual(len((((original.get("puntuacio") or {}).get("victories") or {}).get("desempat_comparacio")) or []), 2)


class GlobalClassificacioTemplateManagementTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username="global_tpl_owner",
            password="testpass123",
            email="global-tpl-owner@example.com",
        )
        self.other_user = User.objects.create_user(
            username="global_tpl_other",
            password="testpass123",
            email="global-tpl-other@example.com",
        )
        self.admin_user = User.objects.create_superuser(
            username="global_tpl_admin",
            password="testpass123",
            email="global-tpl-admin@example.com",
        )
        self.app = self._create_aparell("TRAMP_GLOB", "Tramp Global", owner=self.user)
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "E", "label": "Execucio", "type": "number"},
                ],
                "computed": [
                    {"code": "TOTAL", "formula": "E"},
                ],
            },
        )
        self.comp = self._create_competicio("Comp Global Templates")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )

    def _build_global_schema_payload(self, app_id):
        schema = json.loads(json.dumps(DEFAULT_SCHEMA))
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [app_id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(app_id): ["total"]}
        schema["presentacio"]["columnes"] = [
            {"type": "builtin", "key": "posicio", "label": "#", "align": "left"},
            {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
            {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
        ]
        return schema

    def test_owner_can_create_list_and_delete_global_template(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        payload = {
            "nom": "Plantilla Global 1",
            "slug": "plantilla-global-1",
            "activa": True,
            "tipus": "individual",
            "schema": self._build_global_schema_payload(self.app.id),
        }
        res = self.client.post(save_url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 200)
        body = res.json()
        self.assertTrue(body.get("ok"))
        cfg = body.get("cfg") or {}
        tpl = ClassificacioTemplateGlobal.objects.get(pk=cfg.get("id"))
        self.assertEqual(((tpl.payload or {}).get("schema") or {}).get("puntuacio", {}).get("aparells", {}).get("ids"), [self.app.codi])
        self.assertEqual(tpl.slug, "plantilla-global-1")

        list_url = reverse("classificacio_template_global_list")
        list_res = self.client.get(list_url)
        self.assertEqual(list_res.status_code, 200)
        self.assertContains(list_res, "Plantilla Global 1")

        delete_url = reverse("classificacio_template_global_delete", kwargs={"pk": tpl.id})
        delete_res = self.client.post(
            delete_url,
            data=json.dumps({}),
            content_type="application/json",
            HTTP_ACCEPT="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(delete_res.status_code, 200)
        self.assertFalse(ClassificacioTemplateGlobal.objects.filter(pk=tpl.id).exists())

    def test_owner_list_hides_foreign_templates_and_admin_sees_both(self):
        own_tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Own",
            slug="tpl-own",
            tipus="individual",
            activa=True,
            payload={"schema": {}},
            requirements={},
            created_by=self.user,
        )
        foreign_tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Foreign",
            slug="tpl-foreign",
            tipus="individual",
            activa=True,
            payload={"schema": {}},
            requirements={},
            created_by=self.other_user,
        )

        list_url = reverse("classificacio_template_global_list")

        self.client.force_login(self.user)
        owner_res = self.client.get(list_url)
        self.assertContains(owner_res, own_tpl.nom)
        self.assertNotContains(owner_res, foreign_tpl.nom)

        self.client.force_login(self.admin_user)
        admin_res = self.client.get(list_url)
        self.assertContains(admin_res, own_tpl.nom)
        self.assertContains(admin_res, foreign_tpl.nom)

    def test_foreign_user_cannot_delete_template(self):
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Locked",
            slug="tpl-locked",
            tipus="individual",
            activa=True,
            payload={"schema": {}},
            requirements={},
            created_by=self.user,
        )
        self.client.force_login(self.other_user)
        delete_url = reverse("classificacio_template_global_delete", kwargs={"pk": tpl.id})
        res = self.client.post(delete_url, data=json.dumps({}), content_type="application/json", HTTP_ACCEPT="application/json")
        self.assertEqual(res.status_code, 404)

    def test_owner_can_update_global_template_and_version_increments(self):
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Editable",
            slug="tpl-editable",
            tipus="individual",
            activa=True,
            payload={"schema": {"puntuacio": {"aparells": {"mode": "seleccionar", "ids": [self.app.codi]}}}},
            requirements={},
            created_by=self.user,
        )
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        payload = {
            "id": tpl.id,
            "nom": "Tpl Editable V2",
            "slug": "tpl-editable-v2",
            "activa": False,
            "tipus": "individual",
            "schema": self._build_global_schema_payload(self.app.id),
        }
        res = self.client.post(save_url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 200)
        tpl.refresh_from_db()
        self.assertEqual(tpl.nom, "Tpl Editable V2")
        self.assertEqual(tpl.slug, "tpl-editable-v2")
        self.assertFalse(tpl.activa)
        self.assertEqual(tpl.version, 2)

    def test_global_validation_rejects_invalid_fields(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_schema_payload(self.app.id)
        schema["particions_v2"] = [{"code": "custom_excel", "apply_mode": "all", "parent_values": []}]
        schema["particions"] = ["custom_excel"]
        schema["puntuacio"]["camps_per_aparell"] = {str(self.app.id): ["NOT_SCOREABLE"]}
        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Tpl Invalid",
                    "slug": "tpl-invalid",
                    "activa": True,
                    "tipus": "individual",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("camp no permes" in err for err in body.get("errors", [])))
        self.assertTrue(any("no es puntuable" in err for err in body.get("errors", [])))

    def test_global_edit_preserves_legacy_extra_fields(self):
        legacy_schema = {
            "particions": ["custom_excel"],
            "particions_v2": [{"code": "custom_excel", "apply_mode": "all", "parent_values": []}],
            "particions_custom": {
                "custom_excel": {
                    "mode": "custom",
                    "fallback_label": "Altres",
                    "grups": [{"key": "grp_1", "label": "Bloc X", "values": ["A"]}],
                }
            },
            "filtres": {"custom_excel_in": ["A"]},
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.app.codi]},
                "camps_per_aparell": {self.app.codi: ["total"]},
            },
            "presentacio": {
                "columnes": [
                    {"type": "builtin", "key": "posicio", "label": "#", "align": "left"},
                    {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                ]
            },
        }
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Legacy",
            slug="tpl-legacy",
            tipus="individual",
            activa=True,
            payload={"schema": legacy_schema},
            requirements={},
            created_by=self.user,
        )
        self.client.force_login(self.user)
        res = self.client.post(
            reverse("classificacio_template_global_save"),
            data=json.dumps(
                {
                    "id": tpl.id,
                    "nom": "Tpl Legacy Updated",
                    "slug": "tpl-legacy-updated",
                    "activa": True,
                    "tipus": "individual",
                    "schema": self._build_global_schema_payload(self.app.id),
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 200)
        tpl.refresh_from_db()
        saved_schema = (tpl.payload or {}).get("schema") or {}
        self.assertEqual(saved_schema.get("filtres", {}).get("custom_excel_in"), ["A"])
        self.assertIn("custom_excel", saved_schema.get("particions", []))
        self.assertIn("custom_excel", saved_schema.get("particions_custom", {}))

    def test_global_template_appears_in_competition_template_list(self):
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl For Competition",
            slug="tpl-for-competition",
            tipus="individual",
            activa=True,
            payload={"schema": {"puntuacio": {"aparells": {"mode": "seleccionar", "ids": [self.app.codi]}}}},
            requirements={},
            created_by=self.user,
        )
        self.client.force_login(self.user)
        url = reverse("classificacio_template_list", kwargs={"pk": self.comp.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        ids = {int(item["id"]) for item in (res.json().get("templates") or [])}
        self.assertIn(tpl.id, ids)

    def test_global_builder_create_renders_builder_json_contract(self):
        self.client.force_login(self.user)
        url = reverse("classificacio_template_global_create")
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, 'id="can-manage-global-templates"')
        self.assertContains(res, 'id="builder-save-url"')
        self.assertContains(res, 'id="builder-delete-url-pattern"')
        self.assertContains(res, 'id="builder-preview-url-pattern"')
        self.assertContains(res, 'id="builder-enable-template-library"')
        self.assertContains(res, 'id="builder-can-preview"')
        self.assertContains(res, 'id="builder-selected-id"')
        self.assertContains(res, 'id="builder-auto-add-new"')
        self.assertContains(res, 'id="victoryConfigBox"')
        self.assertContains(res, 'id="sVictoryModeCamps"')
        self.assertContains(res, 'id="sVictoryModeExercicis"')
        self.assertContains(res, 'id="classifHelpDrawer"')
        self.assertContains(res, 'id="classif-builder-back-to-top"')
        self.assertContains(res, "classificacions_builder_help.css")
        self.assertContains(res, "classificacions_builder_help.js")
        self.assertContains(res, 'data-help-key="global_overview"')
        self.assertContains(res, 'data-help-key="desempat_overview"')
        self.assertNotContains(res, '<option value="entitat">Per entitat</option>', html=True)

    def test_admin_global_builder_edit_is_scoped_to_template_owner_catalog(self):
        own_tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Owner Scope",
            slug="tpl-owner-scope",
            tipus="individual",
            activa=True,
            payload={"schema": {}},
            requirements={},
            created_by=self.user,
        )
        foreign_app = self._create_aparell("TRAMP_OTHER", "Tramp Other", owner=self.other_user)
        ScoringSchema.objects.create(
            aparell=foreign_app,
            schema={
                "fields": [{"code": "E", "label": "Execucio", "type": "number"}],
                "computed": [{"code": "TOTAL", "formula": "E"}],
            },
        )
        foreign_tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Foreign Scope",
            slug="tpl-foreign-scope",
            tipus="individual",
            activa=True,
            payload={"schema": {"puntuacio": {"aparells": {"mode": "seleccionar", "ids": [foreign_app.codi]}}}},
            requirements={},
            created_by=self.other_user,
        )

        self.client.force_login(self.admin_user)
        url = reverse("classificacio_template_global_update", kwargs={"pk": foreign_tpl.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, foreign_tpl.nom)
        self.assertNotContains(res, own_tpl.nom)
        self.assertContains(res, foreign_app.nom)
        self.assertNotContains(res, self.app.nom)


class JudgeVideoApiTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Video")
        self.app = self._create_aparell("TRAMP_VIDEO", "Tramp Video")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        self.ins_allowed = self._create_inscripcio(self.comp, "Allowed", ordre=1)
        self.ins_blocked = self._create_inscripcio(self.comp, "Blocked", ordre=2)
        InscripcioAparellExclusio.objects.create(
            inscripcio=self.ins_blocked,
            comp_aparell=self.comp_app,
        )

        self.token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Judge Video",
            permissions=[{"field_code": "E", "judge_index": 1}],
            can_record_video=True,
            is_active=True,
        )
        self._probe_patcher = patch(
            "competicions_trampoli.views_judge._probe_uploaded_video_metadata",
            side_effect=self._fake_probe_uploaded_video_metadata,
        )
        self._probe_patcher.start()
        self.addCleanup(self._probe_patcher.stop)

    @staticmethod
    def _fake_probe_uploaded_video_metadata(uploaded_file):
        from .views_judge import VideoValidationError

        name = (getattr(uploaded_file, "name", "") or "").lower()
        if name.endswith(".txt"):
            raise VideoValidationError(
                "Tipus MIME no permes: text/plain",
                reason="mime_not_allowed",
                payload={"mime_type": "text/plain", "format_name": "text"},
            )
        return {
            "duration_seconds": 12,
            "mime_type": "video/mp4",
            "format_name": "mp4",
            "video_codec": "h264",
        }

    def _sample_video(self, name="routine.mp4", size=1024):
        return SimpleUploadedFile(name, b"\x00" * size, content_type="video/mp4")

    def test_video_upload_creates_scoreentry_and_video(self):
        upload_url = reverse("judge_video_upload", kwargs={"token": self.token.id})
        r = self.client.post(
            upload_url,
            data={
                "inscripcio_id": self.ins_allowed.id,
                "exercici": 1,
                "duration_seconds": 12,
                "video_file": self._sample_video(),
            },
        )
        self.assertEqual(r.status_code, 200)
        payload = r.json()
        self.assertTrue(payload.get("ok"))
        self.assertTrue(payload.get("created"))

        entry = ScoreEntry.objects.get(
            competicio=self.comp,
            inscripcio=self.ins_allowed,
            exercici=1,
            comp_aparell=self.comp_app,
        )
        video = ScoreEntryVideo.objects.get(score_entry=entry)
        self.assertEqual(video.status, ScoreEntryVideo.Status.READY)
        self.assertEqual(video.mime_type, "video/mp4")
        self.assertEqual(video.duration_seconds, 12)
        self.assertEqual(video.judge_token_id, self.token.id)
        ev = ScoreEntryVideoEvent.objects.filter(
            action=ScoreEntryVideoEvent.Action.UPLOAD,
            score_entry=entry,
            video=video,
            ok=True,
        ).first()
        self.assertIsNotNone(ev)

    def test_video_status_returns_false_when_absent(self):
        status_url = reverse("judge_video_status", kwargs={"token": self.token.id})
        r = self.client.get(status_url, {"inscripcio_id": self.ins_allowed.id, "exercici": 1})
        self.assertEqual(r.status_code, 200)
        payload = r.json()
        self.assertTrue(payload.get("ok"))
        self.assertFalse(payload.get("has_video"))

    def test_video_endpoints_return_403_when_token_video_disabled(self):
        token_no_video = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Judge No Video",
            permissions=[{"field_code": "E", "judge_index": 1}],
            is_active=True,
        )
        status_url = reverse("judge_video_status", kwargs={"token": token_no_video.id})
        upload_url = reverse("judge_video_upload", kwargs={"token": token_no_video.id})
        delete_url = reverse("judge_video_delete", kwargs={"token": token_no_video.id})

        status_res = self.client.get(status_url, {"inscripcio_id": self.ins_allowed.id, "exercici": 1})
        self.assertEqual(status_res.status_code, 403)
        self.assertEqual(status_res.json().get("reason"), "video_disabled")

        upload_res = self.client.post(
            upload_url,
            data={
                "inscripcio_id": self.ins_allowed.id,
                "exercici": 1,
                "video_file": self._sample_video(),
            },
        )
        self.assertEqual(upload_res.status_code, 403)
        self.assertEqual(upload_res.json().get("reason"), "video_disabled")

        delete_res = self.client.post(
            delete_url,
            data={
                "inscripcio_id": self.ins_allowed.id,
                "exercici": 1,
            },
        )
        self.assertEqual(delete_res.status_code, 403)
        self.assertEqual(delete_res.json().get("reason"), "video_disabled")

    def test_video_upload_rejects_excluded_inscripcio(self):
        upload_url = reverse("judge_video_upload", kwargs={"token": self.token.id})
        r = self.client.post(
            upload_url,
            data={
                "inscripcio_id": self.ins_blocked.id,
                "exercici": 1,
                "video_file": self._sample_video(),
            },
        )
        self.assertEqual(r.status_code, 403)
        self.assertTrue(
            ScoreEntryVideoEvent.objects.filter(
                action=ScoreEntryVideoEvent.Action.UPLOAD_REJECTED,
                inscripcio=self.ins_blocked,
                ok=False,
            ).exists()
        )

    def test_video_upload_rejects_invalid_mime(self):
        upload_url = reverse("judge_video_upload", kwargs={"token": self.token.id})
        bad_file = SimpleUploadedFile("routine.txt", b"abc", content_type="text/plain")
        r = self.client.post(
            upload_url,
            data={
                "inscripcio_id": self.ins_allowed.id,
                "exercici": 1,
                "video_file": bad_file,
            },
        )
        self.assertEqual(r.status_code, 400)
        self.assertTrue(
            ScoreEntryVideoEvent.objects.filter(
                action=ScoreEntryVideoEvent.Action.UPLOAD_REJECTED,
                inscripcio=self.ins_allowed,
                ok=False,
            ).exists()
        )

    def test_video_upload_rejects_file_too_large(self):
        old_limit = ScoreEntryVideo.VIDEO_MAX_SIZE_BYTES
        try:
            ScoreEntryVideo.VIDEO_MAX_SIZE_BYTES = 100
            upload_url = reverse("judge_video_upload", kwargs={"token": self.token.id})
            r = self.client.post(
                upload_url,
                data={
                    "inscripcio_id": self.ins_allowed.id,
                    "exercici": 1,
                    "video_file": self._sample_video(size=256),
                },
            )
            self.assertEqual(r.status_code, 400)
            self.assertTrue(
                ScoreEntryVideoEvent.objects.filter(
                    action=ScoreEntryVideoEvent.Action.UPLOAD_REJECTED,
                    inscripcio=self.ins_allowed,
                    ok=False,
                ).exists()
            )
        finally:
            ScoreEntryVideo.VIDEO_MAX_SIZE_BYTES = old_limit

    def test_second_upload_creates_replace_event(self):
        upload_url = reverse("judge_video_upload", kwargs={"token": self.token.id})
        r1 = self.client.post(
            upload_url,
            data={
                "inscripcio_id": self.ins_allowed.id,
                "exercici": 1,
                "video_file": self._sample_video(name="first.mp4", size=1024),
            },
        )
        self.assertEqual(r1.status_code, 200)

        r2 = self.client.post(
            upload_url,
            data={
                "inscripcio_id": self.ins_allowed.id,
                "exercici": 1,
                "video_file": self._sample_video(name="second.mp4", size=1024),
            },
        )
        self.assertEqual(r2.status_code, 200)

        entry = ScoreEntry.objects.get(
            competicio=self.comp,
            inscripcio=self.ins_allowed,
            exercici=1,
            comp_aparell=self.comp_app,
        )
        self.assertTrue(
            ScoreEntryVideoEvent.objects.filter(
                action=ScoreEntryVideoEvent.Action.REPLACE,
                score_entry=entry,
                ok=True,
            ).exists()
        )

    def test_video_delete_removes_existing_capture(self):
        upload_url = reverse("judge_video_upload", kwargs={"token": self.token.id})
        delete_url = reverse("judge_video_delete", kwargs={"token": self.token.id})

        r_upload = self.client.post(
            upload_url,
            data={
                "inscripcio_id": self.ins_allowed.id,
                "exercici": 1,
                "video_file": self._sample_video(),
            },
        )
        self.assertEqual(r_upload.status_code, 200)

        r_delete = self.client.post(
            delete_url,
            data={
                "inscripcio_id": self.ins_allowed.id,
                "exercici": 1,
            },
        )
        self.assertEqual(r_delete.status_code, 200)
        payload = r_delete.json()
        self.assertTrue(payload.get("ok"))
        self.assertTrue(payload.get("deleted"))

        entry = ScoreEntry.objects.get(
            competicio=self.comp,
            inscripcio=self.ins_allowed,
            exercici=1,
            comp_aparell=self.comp_app,
        )
        self.assertFalse(ScoreEntryVideo.objects.filter(score_entry=entry).exists())
        self.assertTrue(
            ScoreEntryVideoEvent.objects.filter(
                action=ScoreEntryVideoEvent.Action.DELETE,
                score_entry=entry,
                ok=True,
            ).exists()
        )


class PublicLiveTokenViewsTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Public Live")
        self.token = PublicLiveToken.objects.create(
            competicio=self.comp,
            label="Pantalla principal",
            is_active=True,
        )

    def test_public_live_portal_renders_public_live_page(self):
        url = reverse("public_live_portal", kwargs={"token": self.token.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        body = res.content.decode("utf-8")
        self.assertIn("Classificacions", body)
        self.assertIn(reverse("public_live_classificacions_data", kwargs={"token": self.token.id}), body)

    def test_public_live_portal_rejects_revoked_token(self):
        self.token.is_active = False
        self.token.revoked_at = timezone.now()
        self.token.save(update_fields=["is_active", "revoked_at"])

        url = reverse("public_live_portal", kwargs={"token": self.token.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 403)

    def test_public_live_qr_png_returns_png(self):
        url = reverse("public_live_qr_png", kwargs={"token": self.token.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["Content-Type"], "image/png")

    def test_public_live_data_requires_valid_token(self):
        url = reverse("public_live_classificacions_data", kwargs={"token": self.token.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.json().get("ok"))
        self.assertEqual(
            res.json().get("permissions", {}).get("can_view_media"),
            False,
        )

    def test_public_live_data_exposes_media_permission_flag(self):
        self.token.can_view_media = True
        self.token.save(update_fields=["can_view_media"])

        url = reverse("public_live_classificacions_data", kwargs={"token": self.token.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.json().get("ok"))
        self.assertEqual(
            res.json().get("permissions", {}).get("can_view_media"),
            True,
        )


class LiveClassificacionsRedisCacheTests(_BaseTrampoliDataMixin, TestCase):
    class FakeRedis:
        def __init__(self):
            self.store = {}

        def get(self, key):
            return self.store.get(key)

        def set(self, key, value, nx=False, ex=None):
            if nx and key in self.store:
                return False
            self.store[key] = value
            return True

        def delete(self, key):
            self.store.pop(key, None)
            return 1

    def setUp(self):
        self.comp = self._create_competicio("Comp Live Cache")
        self.app = self._create_aparell("TRAMP_LIVE_CACHE", "Tramp Live Cache")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)
        self.ins = self._create_inscripcio(self.comp, "Participant Cache", ordre=1)
        self.cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="General",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=self._schema(),
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=9.8,
        )
        self.token = PublicLiveToken.objects.create(
            competicio=self.comp,
            label="Pantalla cache",
            is_active=True,
        )
        User = get_user_model()
        self.user = User.objects.create_user(
            username="live_cache_user",
            password="testpass123",
            email="live-cache@example.com",
        )
        self.editor_user = User.objects.create_user(
            username="live_cache_editor",
            password="testpass123",
            email="live-cache-editor@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.READONLY,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.editor_user,
            competicio=self.comp,
            role=CompeticioMembership.Role.CLASSIFICACIONS,
            is_active=True,
        )

    def _schema(self):
        return {
            "filtres": {},
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "exercicis_best_n": 1,
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
                "camp": "total",
                "agregacio": "sum",
                "best_n": 1,
            },
            "desempat": [],
            "presentacio": {
                "columnes": [
                    {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                ],
            },
        }

    def _public_url(self):
        return reverse("public_live_classificacions_data", kwargs={"token": self.token.id})

    def _internal_url(self):
        return reverse("classificacions_live_data", kwargs={"pk": self.comp.id})

    def _reorder_url(self):
        return reverse("classificacio_reorder", kwargs={"pk": self.comp.id})

    def _snapshot_payload(self):
        return {
            "ok": True,
            "changed": True,
            "stamp": timezone.now().isoformat(),
            "competicio": {"id": self.comp.id, "nom": self.comp.nom},
            "cfgs": [
                {
                    "id": self.cfg.id,
                    "nom": self.cfg.nom,
                    "tipus": self.cfg.tipus,
                    "columns": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                    ],
                    "parts": [
                        {
                            "particio": "global",
                            "rows": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}],
                        }
                    ],
                }
            ],
        }

    def _snapshot_blob(self, generated_at=None):
        payload = self._snapshot_payload()
        payload["generated_at"] = (generated_at or timezone.now()).isoformat()
        return json.dumps(payload)

    def test_first_get_computes_and_second_get_uses_cache(self):
        fake_redis = self.FakeRedis()
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        cache_key = live_cache.live_cache_key(self.comp.id)
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views_classificacions.compute_classificacio", return_value=compute_result) as mocked_compute:
                res_1 = self.client.get(self._public_url())
                res_2 = self.client.get(self._public_url())

        self.assertEqual(res_1.status_code, 200)
        self.assertEqual(res_2.status_code, 200)
        self.assertEqual(mocked_compute.call_count, 1)
        self.assertIn(cache_key, fake_redis.store)
        self.assertEqual(res_1["X-Live-Cache"], "miss")
        self.assertEqual(res_2["X-Live-Cache"], "hit")
        self.assertEqual(
            res_2.json().get("cfgs", [])[0].get("parts", [])[0].get("rows", [])[0].get("participant"),
            "Participant Cache",
        )

    def test_public_and_internal_live_share_same_snapshot(self):
        fake_redis = self.FakeRedis()
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        self.client.force_login(self.user)
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views_classificacions.compute_classificacio", return_value=compute_result) as mocked_compute:
                public_res = self.client.get(self._public_url())
                internal_res = self.client.get(self._internal_url())

        self.assertEqual(public_res.status_code, 200)
        self.assertEqual(internal_res.status_code, 200)
        self.assertEqual(mocked_compute.call_count, 1)
        self.assertEqual(public_res["X-Live-Cache"], "miss")
        self.assertEqual(internal_res["X-Live-Cache"], "hit")
        self.assertIn("permissions", public_res.json())
        self.assertNotIn("permissions", internal_res.json())

    def test_since_is_served_from_cached_stamp_without_recompute(self):
        fake_redis = self.FakeRedis()
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views_classificacions.compute_classificacio", return_value=compute_result) as mocked_compute:
                first_res = self.client.get(self._public_url())
                stamp = first_res.json()["stamp"]
                second_res = self.client.get(self._public_url(), {"since": stamp})

        self.assertEqual(first_res.status_code, 200)
        self.assertEqual(second_res.status_code, 200)
        self.assertEqual(mocked_compute.call_count, 1)
        self.assertFalse(second_res.json()["changed"])
        self.assertEqual(second_res.json()["stamp"], stamp)
        self.assertEqual(second_res.json().get("permissions", {}).get("can_view_media"), False)

    def test_lock_contention_waits_for_snapshot_without_recompute(self):
        fake_redis = self.FakeRedis()
        fake_redis.set(live_cache.live_lock_key(self.comp.id), "busy")
        waited_snapshot = json.loads(self._snapshot_blob())
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.live_cache._wait_for_live_snapshot", return_value=waited_snapshot):
                with patch("competicions_trampoli.views_classificacions.compute_classificacio") as mocked_compute:
                    res = self.client.get(self._public_url())

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["X-Live-Cache"], "wait-hit")
        mocked_compute.assert_not_called()

    def test_stale_snapshot_is_served_when_refresh_lock_is_busy(self):
        fake_redis = self.FakeRedis()
        fake_redis.set(
            live_cache.live_cache_key(self.comp.id),
            self._snapshot_blob(generated_at=timezone.now() - timedelta(seconds=10)),
        )
        fake_redis.set(live_cache.live_lock_key(self.comp.id), "busy")
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views_classificacions.compute_classificacio") as mocked_compute:
                res = self.client.get(self._public_url())

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["X-Live-Cache"], "stale")
        mocked_compute.assert_not_called()

    def test_redis_failure_falls_back_to_direct_compute(self):
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        with patch(
            "competicions_trampoli.live_cache._live_redis_client",
            side_effect=RuntimeError("redis down"),
        ):
            with patch("competicions_trampoli.views_classificacions.compute_classificacio", return_value=compute_result) as mocked_compute:
                res = self.client.get(self._public_url())

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["X-Live-Cache"], "fallback")
        self.assertEqual(mocked_compute.call_count, 1)

    def test_fresh_snapshot_with_dirty_forces_refresh_and_clears_dirty(self):
        fake_redis = self.FakeRedis()
        fake_redis.set(live_cache.live_cache_key(self.comp.id), self._snapshot_blob())
        dirty_key = live_cache.live_dirty_key(self.comp.id)
        fake_redis.set(dirty_key, "dirty-1")
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views_classificacions.compute_classificacio", return_value=compute_result) as mocked_compute:
                res = self.client.get(self._public_url())

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["X-Live-Cache"], "refresh")
        self.assertEqual(mocked_compute.call_count, 1)
        self.assertNotIn(dirty_key, fake_redis.store)

    def test_dirty_marker_changed_during_refresh_is_preserved(self):
        fake_redis = self.FakeRedis()
        fake_redis.set(live_cache.live_cache_key(self.comp.id), self._snapshot_blob())
        dirty_key = live_cache.live_dirty_key(self.comp.id)
        fake_redis.set(dirty_key, "dirty-1")

        def compute_payload(competicio, since_raw=None):
            fake_redis.set(dirty_key, "dirty-2")
            return self._snapshot_payload()

        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            payload, source = live_cache.get_live_payload_cached(
                self.comp,
                compute_payload=compute_payload,
                since_raw=None,
            )

        self.assertEqual(source, "refresh")
        self.assertTrue(payload.get("ok"))
        self.assertEqual(fake_redis.get(dirty_key), "dirty-2")

    def test_scoreentry_signal_marks_dirty_after_commit(self):
        fake_redis = self.FakeRedis()
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with self.captureOnCommitCallbacks(execute=True):
                ScoreEntry.objects.create(
                    competicio=self.comp,
                    inscripcio=self.ins,
                    exercici=2,
                    comp_aparell=self.comp_app,
                    inputs={},
                    outputs={},
                    total=8.4,
                )

        self.assertIsNotNone(fake_redis.get(live_cache.live_dirty_key(self.comp.id)))

    def test_teamscoreentry_signal_marks_dirty_after_commit(self):
        team_ctx = EquipContext.objects.create(
            competicio=self.comp,
            code="parelles-live",
            nom="Parelles live",
        )
        self.comp_app.participant_mode = CompeticioAparell.ParticipantMode.TEAM_CONTEXT
        self.comp_app.team_context = team_ctx
        self.comp_app.expected_team_size = 2
        self.comp_app.team_scoring_mode = CompeticioAparell.TeamScoringMode.MEMBERS_PLUS_SHARED
        self.comp_app.save(
            update_fields=["participant_mode", "team_context", "expected_team_size", "team_scoring_mode"]
        )
        equip = Equip.objects.create(competicio=self.comp, nom="Equip live")
        ins_b = self._create_inscripcio(self.comp, "Participant Cache 2", ordre=2)
        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=team_ctx,
            inscripcio=self.ins,
            equip=equip,
        )
        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=team_ctx,
            inscripcio=ins_b,
            equip=equip,
        )

        fake_redis = self.FakeRedis()
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with self.captureOnCommitCallbacks(execute=True):
                TeamScoreEntry.objects.create(
                    competicio=self.comp,
                    equip=equip,
                    exercici=1,
                    comp_aparell=self.comp_app,
                    inputs={},
                    outputs={},
                    total=8.4,
                )

        self.assertIsNotNone(fake_redis.get(live_cache.live_dirty_key(self.comp.id)))

    def test_classificacioconfig_signal_marks_dirty_after_commit(self):
        fake_redis = self.FakeRedis()
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with self.captureOnCommitCallbacks(execute=True):
                self.cfg.nom = "General Dirty"
                self.cfg.save(update_fields=["nom"])

        self.assertIsNotNone(fake_redis.get(live_cache.live_dirty_key(self.comp.id)))

    def test_classificacio_reorder_marks_dirty_after_bulk_update(self):
        fake_redis = self.FakeRedis()
        cfg_2 = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Segona",
            activa=True,
            ordre=2,
            tipus="individual",
            schema=self._schema(),
        )
        self.client.force_login(self.editor_user)
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with self.captureOnCommitCallbacks(execute=True):
                res = self.client.post(
                    self._reorder_url(),
                    data=json.dumps({"order": [cfg_2.id, self.cfg.id]}),
                    content_type="application/json",
                )

        self.assertEqual(res.status_code, 200)
        self.assertIsNotNone(fake_redis.get(live_cache.live_dirty_key(self.comp.id)))


class CompetitionAccessControlTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Accessos")
        self.other_comp = self._create_competicio("Comp Privada")
        self.app = self._create_aparell("TRAMP_ACCESS", "Tramp Access")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        User = get_user_model()
        self.judge_admin_user = User.objects.create_user(
            username="judge_admin_user",
            password="testpass123",
            email="judge-admin@example.com",
        )
        self.readonly_user = User.objects.create_user(
            username="readonly_user",
            password="testpass123",
            email="readonly@example.com",
        )
        self.manager_user = User.objects.create_user(
            username="manager_user",
            password="testpass123",
            email="manager@example.com",
        )

        for group_name in GLOBAL_AUTH_GROUPS.keys():
            Group.objects.get_or_create(name=group_name)

        competitions_group = Group.objects.get(name="competicions_manager")
        self.manager_user.groups.add(competitions_group)

        CompeticioMembership.objects.create(
            user=self.judge_admin_user,
            competicio=self.comp,
            role=CompeticioMembership.Role.JUDGE_ADMIN,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.readonly_user,
            competicio=self.comp,
            role=CompeticioMembership.Role.READONLY,
            is_active=True,
        )

    def test_judge_admin_membership_can_manage_qr_but_readonly_cannot(self):
        url = reverse("judges_qr_home", kwargs={"competicio_id": self.comp.id})

        self.client.force_login(self.judge_admin_user)
        ok_res = self.client.get(url)
        self.assertEqual(ok_res.status_code, 200)
        self.client.logout()

        self.client.force_login(self.readonly_user)
        denied_res = self.client.get(url)
        self.assertEqual(denied_res.status_code, 403)

        self.assertTrue(
            user_has_competicio_capability(
                self.judge_admin_user,
                self.comp,
                "judge_tokens.manage",
            )
        )
        self.assertFalse(
            user_has_competicio_capability(
                self.readonly_user,
                self.comp,
                "judge_tokens.manage",
            )
        )

    def test_global_competitions_manager_can_access_global_competitions_pages(self):
        url = reverse("competicions_home")
        self.client.force_login(self.manager_user)
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)

    def test_created_list_shows_only_user_membership_competitions(self):
        url = reverse("created")
        self.client.force_login(self.readonly_user)
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, self.comp.nom)
        self.assertNotContains(res, self.other_comp.nom)

    def test_competicions_manager_group_without_membership_cannot_open_foreign_competition(self):
        url = reverse("inscripcions_list", kwargs={"pk": self.other_comp.id})
        self.client.force_login(self.manager_user)
        res = self.client.get(url)
        self.assertEqual(res.status_code, 403)

    def test_create_competition_assigns_owner_membership_to_creator(self):
        url = reverse("create")
        self.client.force_login(self.readonly_user)
        res = self.client.post(
            url,
            data={
                "nom": "Comp Creada Usuari",
                "tipus": Competicio.Tipus.TRAMPOLI,
                "data": "",
            },
        )
        self.assertEqual(res.status_code, 302)

        created_comp = Competicio.objects.get(nom="Comp Creada Usuari")
        membership = CompeticioMembership.objects.get(
            user=self.readonly_user,
            competicio=created_comp,
        )
        self.assertEqual(membership.role, CompeticioMembership.Role.OWNER)
        self.assertTrue(membership.is_active)
        self.assertEqual(membership.granted_by_id, self.readonly_user.id)

    def test_public_live_token_creation_persists_media_permission(self):
        url = reverse("public_live_qr_home", kwargs={"competicio_id": self.comp.id})
        self.client.force_login(self.judge_admin_user)

        res_with_media = self.client.post(
            url,
            data={
                "action": "create",
                "label": "Public A",
                "can_view_media": "1",
            },
        )
        self.assertEqual(res_with_media.status_code, 302)
        token_with_media = PublicLiveToken.objects.get(competicio=self.comp, label="Public A")
        self.assertTrue(token_with_media.can_view_media)

        res_without_media = self.client.post(
            url,
            data={
                "action": "create",
                "label": "Public B",
            },
        )
        self.assertEqual(res_without_media.status_code, 302)
        token_without_media = PublicLiveToken.objects.get(competicio=self.comp, label="Public B")
        self.assertFalse(token_without_media.can_view_media)


class JudgeMessagingFlowTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Missatgeria")
        self.app = self._create_aparell("TRAMP_MSG", "Trampoli Msg")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        self.token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Judge Msg A",
            permissions=[{"field_code": "E", "judge_index": 1}],
            is_active=True,
        )

        User = get_user_model()
        self.manager_user = User.objects.create_user(
            username="judge_msg_manager",
            password="testpass123",
            email="judge-msg-manager@example.com",
        )
        self.readonly_user = User.objects.create_user(
            username="judge_msg_readonly",
            password="testpass123",
            email="judge-msg-readonly@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.manager_user,
            competicio=self.comp,
            role=CompeticioMembership.Role.JUDGE_ADMIN,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.readonly_user,
            competicio=self.comp,
            role=CompeticioMembership.Role.READONLY,
            is_active=True,
        )

    def test_quick_support_creates_requested_conversation_without_explicit_text(self):
        url = reverse("judge_request_support", kwargs={"token": self.token.id})
        res = self.client.post(
            url,
            data=json.dumps({}),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 200)
        payload = res.json()
        self.assertTrue(payload.get("ok"))
        self.assertTrue(payload.get("quick"))

        conv = JudgeConversation.objects.get(judge_token=self.token)
        self.assertEqual(conv.status, JudgeConversation.Status.REQUESTED)
        self.assertGreaterEqual(conv.unread_for_org, 1)

        msg = JudgeConversationMessage.objects.get(conversation=conv)
        self.assertEqual(msg.message_type, JudgeConversationMessage.MessageType.SUPPORT_REQUEST_QUICK)
        self.assertIn("assistencia", (msg.text or "").lower())

    def test_quick_support_has_cooldown(self):
        url = reverse("judge_request_support", kwargs={"token": self.token.id})
        first = self.client.post(url, data=json.dumps({}), content_type="application/json")
        self.assertEqual(first.status_code, 200)

        second = self.client.post(url, data=json.dumps({}), content_type="application/json")
        self.assertEqual(second.status_code, 429)
        self.assertEqual(second.json().get("reason"), "cooldown")

    def test_judge_updates_include_org_instruction(self):
        request_url = reverse("judge_request_support", kwargs={"token": self.token.id})
        self.client.post(request_url, data=json.dumps({}), content_type="application/json")

        send_url = reverse("judge_messages_send_org", kwargs={"competicio_id": self.comp.id})
        self.client.force_login(self.manager_user)
        send_res = self.client.post(
            send_url,
            data=json.dumps(
                {
                    "judge_token_id": str(self.token.id),
                    "message_type": "instruction",
                    "text": "Reinicia tauleta i revisa connexio.",
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(send_res.status_code, 200)
        self.client.logout()

        updates_url = reverse("judge_messages_updates", kwargs={"token": self.token.id})
        updates_res = self.client.get(updates_url)
        self.assertEqual(updates_res.status_code, 200)
        rows = updates_res.json().get("messages", [])
        self.assertTrue(any("Reinicia tauleta" in (x.get("text") or "") for x in rows))

    def test_org_hub_requires_judge_messages_capability(self):
        hub_url = reverse("judge_messages_hub", kwargs={"competicio_id": self.comp.id})

        self.client.force_login(self.manager_user)
        ok_res = self.client.get(hub_url)
        self.assertEqual(ok_res.status_code, 200)
        self.client.logout()

        self.client.force_login(self.readonly_user)
        denied_res = self.client.get(hub_url)
        self.assertEqual(denied_res.status_code, 403)

    def test_org_can_open_conversation_by_token_and_mark_resolved(self):
        send_url = reverse("judge_messages_send_org", kwargs={"competicio_id": self.comp.id})
        self.client.force_login(self.manager_user)
        send_res = self.client.post(
            send_url,
            data=json.dumps(
                {
                    "judge_token_id": str(self.token.id),
                    "message_type": "instruction",
                    "text": "Passa al mode offline.",
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(send_res.status_code, 200)
        conv_id = send_res.json().get("conversation", {}).get("id")
        self.assertTrue(conv_id)

        status_url = reverse("judge_messages_set_status_org", kwargs={"competicio_id": self.comp.id})
        status_res = self.client.post(
            status_url,
            data=json.dumps({"conversation_id": conv_id, "status": "resolved"}),
            content_type="application/json",
        )
        self.assertEqual(status_res.status_code, 200)
        conv = JudgeConversation.objects.get(pk=conv_id)
        self.assertEqual(conv.status, JudgeConversation.Status.RESOLVED)
        self.assertTrue(conv.resolved_at is not None)


class AparellOwnershipIsolationTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        User = get_user_model()
        self.user_a = User.objects.create_user(
            username="ap_owner_a",
            password="testpass123",
            email="ap-owner-a@example.com",
        )
        self.user_b = User.objects.create_user(
            username="ap_owner_b",
            password="testpass123",
            email="ap-owner-b@example.com",
        )

        self.comp = self._create_competicio("Comp Aparell Owners")
        CompeticioMembership.objects.create(
            user=self.user_a,
            competicio=self.comp,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )

    def test_two_users_can_create_same_aparell_code(self):
        ap_a = self._create_aparell("TRA", "Aparell TRA A", owner=self.user_a)
        ap_b = self._create_aparell("TRA", "Aparell TRA B", owner=self.user_b)
        self.assertNotEqual(ap_a.id, ap_b.id)

    def test_cannot_attach_foreign_aparell_to_competition(self):
        self._create_aparell("TRA", "Aparell propi", owner=self.user_a)
        foreign = self._create_aparell("TRA", "Aparell aliè", owner=self.user_b)

        self.client.force_login(self.user_a)
        url = reverse("trampoli_aparell_create", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data={"aparell": foreign.id, "nombre_exercicis": 1})
        self.assertEqual(res.status_code, 200)
        self.assertIn("form", res.context)
        self.assertIn("aparell", res.context["form"].errors)
        self.assertFalse(
            CompeticioAparell.objects.filter(competicio=self.comp, aparell=foreign).exists()
        )

    def test_cannot_edit_foreign_aparell_catalog_entry(self):
        foreign = self._create_aparell("TRA", "Aparell aliè", owner=self.user_b)
        self.client.force_login(self.user_a)
        url = reverse("aparell_update", kwargs={"pk": foreign.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 404)

    def test_can_delete_own_global_aparell_when_unused(self):
        own = self._create_aparell("DEL", "Aparell eliminable", owner=self.user_a)
        self.client.force_login(self.user_a)
        url = reverse("aparell_delete", kwargs={"pk": own.id})
        res = self.client.post(url)
        self.assertEqual(res.status_code, 302)
        self.assertEqual(res.url, reverse("aparells_list"))
        self.assertFalse(Aparell.objects.filter(pk=own.id).exists())

    def test_cannot_delete_own_global_aparell_when_used_in_competition(self):
        own = self._create_aparell("USE", "Aparell en us", owner=self.user_a)
        self._create_comp_aparell(self.comp, own, ordre=1, actiu=True)
        self.client.force_login(self.user_a)
        url = reverse("aparell_delete", kwargs={"pk": own.id})
        res = self.client.post(url)
        self.assertEqual(res.status_code, 302)
        self.assertEqual(res.url, reverse("aparells_list"))
        self.assertTrue(Aparell.objects.filter(pk=own.id).exists())

    def test_cannot_delete_foreign_global_aparell(self):
        foreign = self._create_aparell("ALI", "Aparell alie", owner=self.user_b)
        self.client.force_login(self.user_a)
        url = reverse("aparell_delete", kwargs={"pk": foreign.id})
        res = self.client.post(url)
        self.assertEqual(res.status_code, 404)
        self.assertTrue(Aparell.objects.filter(pk=foreign.id).exists())

    def test_superuser_sees_owner_in_global_aparell_catalog(self):
        self._create_aparell("OWN1", "Aparell owner b", owner=self.user_b)
        User = get_user_model()
        admin = User.objects.create_superuser(
            username="ap_owner_admin_global",
            password="testpass123",
            email="ap-owner-admin-global@example.com",
        )
        self.client.force_login(admin)
        res = self.client.get(reverse("aparells_list"))
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, "Creat per")
        self.assertContains(res, self.user_b.username)

    def test_superuser_sees_owner_in_competicio_aparells_list(self):
        app_b = self._create_aparell("OWN2", "Aparell owner b comp", owner=self.user_b)
        self._create_comp_aparell(self.comp, app_b, ordre=1, actiu=True)
        User = get_user_model()
        admin = User.objects.create_superuser(
            username="ap_owner_admin_comp",
            password="testpass123",
            email="ap-owner-admin-comp@example.com",
        )
        self.client.force_login(admin)
        res = self.client.get(reverse("trampoli_aparells_list", kwargs={"pk": self.comp.id}))
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, "Creat per")
        self.assertContains(res, self.user_b.username)


class EquipContextFlowTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Equip Context")
        self.native_team = Equip.objects.create(competicio=self.comp, nom="Equip Natiu")
        self.custom_team = Equip.objects.create(competicio=self.comp, nom="Equip Context")
        self.ins = self._create_inscripcio(self.comp, "Participant Context", ordre=1)
        self.ins.equip = self.native_team
        self.ins.save(update_fields=["equip"])

        User = get_user_model()
        self.user = User.objects.create_user(
            username="equip_context_user",
            password="testpass123",
            email="equip-context@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.OWNER,
            is_active=True,
        )
        self.client.force_login(self.user)

    def test_custom_context_assignment_does_not_modify_native_team(self):
        create_res = self.client.post(
            reverse("inscripcions_equip_context_create", kwargs={"pk": self.comp.id}),
            data=json.dumps({"name": "Finals"}),
            content_type="application/json",
        )
        self.assertEqual(create_res.status_code, 200)
        context_code = create_res.json()["context"]["code"]

        assign_res = self.client.post(
            reverse("inscripcions_equips_assign", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": context_code,
                    "equip_id": self.custom_team.id,
                    "inscripcio_ids": [self.ins.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(assign_res.status_code, 200)

        self.ins.refresh_from_db()
        self.assertEqual(self.ins.equip_id, self.native_team.id)
        assignacio = InscripcioEquipAssignacio.objects.get(inscripcio=self.ins)
        self.assertEqual(assignacio.context.code, context_code)
        self.assertEqual(assignacio.equip_id, self.custom_team.id)


class EquipPreviewUiTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Team Preview UI")
        self.team_existing = Equip.objects.create(competicio=self.comp, nom="Club A")
        self.team_other = Equip.objects.create(competicio=self.comp, nom="Alt Equip")
        self.ctx = EquipContext.objects.create(competicio=self.comp, code="finals", nom="Finals")

        self.ins_keep = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Anna Keep",
            entitat="Club A",
            ordre_sortida=1,
            equip=self.team_existing,
        )
        self.ins_move = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Berta Move",
            entitat="Club A",
            ordre_sortida=2,
            equip=self.team_other,
        )
        self.ins_new = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Carla New",
            entitat="Club C",
            ordre_sortida=3,
        )
        self.ins_ctx = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Dina Context",
            entitat="Club A",
            ordre_sortida=4,
            equip=self.team_other,
        )
        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=self.ctx,
            inscripcio=self.ins_ctx,
            equip=self.team_existing,
        )

        User = get_user_model()
        self.user = User.objects.create_user(
            username="equip_preview_user",
            password="testpass123",
            email="equip-preview@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.OWNER,
            is_active=True,
        )
        self.client.force_login(self.user)

    def test_inscripcions_list_renders_expandable_team_workbench(self):
        response = self.client.get(reverse("inscripcions_list", kwargs={"pk": self.comp.id}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="teams-main-card"')
        self.assertContains(response, 'data-expand-target="teams-main-card"')
        self.assertContains(response, 'id="team-workspace-shell"')
        self.assertContains(response, 'id="btn-team-workspace-board-mode"')
        self.assertContains(response, 'id="team-filter-q"')
        self.assertContains(response, 'id="btn-team-workspace-preview"')
        self.assertContains(response, 'id="team-preview-status"')
        self.assertContains(response, 'id="team-preview-existing-list"')
        self.assertContains(response, 'id="team-preview-list"')
        self.assertContains(response, 'id="team-context-unassigned-dropzone"')
        self.assertContains(response, "Flux complet d'equips")

    def test_equips_workspace_returns_context_summary_candidates_and_filters(self):
        response = self.client.post(
            reverse("inscripcions_equips_workspace", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": "finals",
                    "filters": {
                        "q": "",
                        "categoria": "",
                        "subcategoria": "",
                        "entitat": "Club A",
                        "assignment_state": "assigned",
                        "equip_id": str(self.team_existing.id),
                    },
                    "page": 1,
                    "page_size": 25,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload.get("ok"))
        self.assertEqual(payload.get("context_code"), "finals")
        self.assertEqual(payload.get("context", {}).get("nom"), "Finals")
        self.assertEqual(payload.get("summary", {}).get("assigned_count"), 1)
        self.assertEqual(payload.get("summary", {}).get("filtered_count"), 1)
        self.assertEqual(payload.get("candidates", {}).get("total"), 1)
        self.assertEqual(payload.get("candidates", {}).get("items", [])[0].get("nom"), "Dina Context")
        self.assertEqual(payload.get("candidates", {}).get("items", [])[0].get("current_team_name"), "Club A")
        self.assertTrue(any(row.get("name") == "Club A" for row in (payload.get("filter_options", {}).get("teams") or [])))
        self.assertTrue(any(ctx.get("code") == "finals" for ctx in (payload.get("contexts") or [])))
        teams_by_name = {row.get("nom"): row for row in (payload.get("teams") or [])}
        self.assertEqual([m.get("nom") for m in teams_by_name["Club A"]["members"]], ["Dina Context"])
        self.assertEqual(teams_by_name["Club A"]["members"][0]["native_team_name"], "Alt Equip")

    def test_equips_workspace_team_members_ignore_candidate_filters_and_keep_stable_order(self):
        ins_early = Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Aina Context",
            entitat="Club Z",
            ordre_sortida=0,
            equip=self.team_other,
        )
        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=self.ctx,
            inscripcio=ins_early,
            equip=self.team_existing,
        )

        response = self.client.post(
            reverse("inscripcions_equips_workspace", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": "finals",
                    "filters": {
                        "q": "",
                        "categoria": "",
                        "subcategoria": "",
                        "entitat": "Club C",
                        "assignment_state": "all",
                        "equip_id": "",
                    },
                    "page": 1,
                    "page_size": 25,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("candidates", {}).get("total"), 1)
        self.assertEqual(payload.get("candidates", {}).get("items", [])[0].get("nom"), "Carla New")
        teams_by_name = {row.get("nom"): row for row in (payload.get("teams") or [])}
        self.assertEqual(
            [m.get("nom") for m in teams_by_name["Club A"]["members"]],
            ["Aina Context", "Dina Context"],
        )

    def test_equips_workspace_returns_native_team_members_ordered(self):
        self.ins_ctx.equip = self.team_existing
        self.ins_ctx.save(update_fields=["equip"])

        response = self.client.post(
            reverse("inscripcions_equips_workspace", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": "native",
                    "filters": {
                        "q": "",
                        "categoria": "",
                        "subcategoria": "",
                        "entitat": "",
                        "assignment_state": "all",
                        "equip_id": "",
                    },
                    "page": 1,
                    "page_size": 25,
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        teams_by_name = {row.get("nom"): row for row in (payload.get("teams") or [])}
        self.assertEqual(
            [m.get("nom") for m in teams_by_name["Club A"]["members"]],
            ["Anna Keep", "Dina Context"],
        )

    def test_equips_workspace_members_refresh_after_assign_and_unassign(self):
        assign_response = self.client.post(
            reverse("inscripcions_equips_assign", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": "finals",
                    "equip_id": self.team_existing.id,
                    "inscripcio_ids": [self.ins_new.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(assign_response.status_code, 200)

        workspace_after_assign = self.client.post(
            reverse("inscripcions_equips_workspace", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": "finals",
                    "filters": {
                        "q": "",
                        "categoria": "",
                        "subcategoria": "",
                        "entitat": "",
                        "assignment_state": "all",
                        "equip_id": "",
                    },
                    "page": 1,
                    "page_size": 25,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(workspace_after_assign.status_code, 200)
        teams_by_name = {row.get("nom"): row for row in (workspace_after_assign.json().get("teams") or [])}
        self.assertEqual(
            [m.get("nom") for m in teams_by_name["Club A"]["members"]],
            ["Carla New", "Dina Context"],
        )

        unassign_response = self.client.post(
            reverse("inscripcions_equips_unassign", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": "finals",
                    "inscripcio_ids": [self.ins_new.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(unassign_response.status_code, 200)

        workspace_after_unassign = self.client.post(
            reverse("inscripcions_equips_workspace", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": "finals",
                    "filters": {
                        "q": "",
                        "categoria": "",
                        "subcategoria": "",
                        "entitat": "",
                        "assignment_state": "all",
                        "equip_id": "",
                    },
                    "page": 1,
                    "page_size": 25,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(workspace_after_unassign.status_code, 200)
        teams_by_name = {row.get("nom"): row for row in (workspace_after_unassign.json().get("teams") or [])}
        self.assertEqual(
            [m.get("nom") for m in teams_by_name["Club A"]["members"]],
            ["Dina Context"],
        )

    def test_equips_preview_returns_rich_contract_for_native_context(self):
        response = self.client.post(
            reverse("inscripcions_equips_preview", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": "native",
                    "fields": ["entitat"],
                    "replace_existing": True,
                    "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                    "selected_ids": [],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload.get("ok"))
        self.assertIn("selection_summary", payload)
        self.assertIn("existing_summary", payload)
        by_name = {row["nom_suggerit"]: row for row in (payload.get("preview") or [])}
        self.assertIn("Club A", by_name)
        self.assertIn("Club C", by_name)
        self.assertEqual(by_name["Club A"]["existing_team_name"], "Club A")
        self.assertTrue(by_name["Club A"]["will_keep"])
        self.assertTrue(by_name["Club A"]["will_reassign"])
        self.assertFalse(by_name["Club A"]["will_create"])
        self.assertEqual(by_name["Club C"]["member_samples"], ["Carla New"])
        self.assertTrue(by_name["Club C"]["will_create"])
        self.assertEqual(payload["selection_summary"]["mode"], "all")
        affected_names = {row["team_name"] for row in payload["existing_summary"]["affected_teams"]}
        self.assertIn("Alt Equip", affected_names)

    def test_equips_preview_supports_custom_context_and_selected_summary(self):
        response = self.client.post(
            reverse("inscripcions_equips_preview", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "context_code": "finals",
                    "fields": ["entitat"],
                    "replace_existing": False,
                    "filters": {"q": "", "categoria": "", "subcategoria": "", "entitat": ""},
                    "selected_ids": [self.ins_ctx.id],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload.get("ok"))
        self.assertEqual(payload["selection_summary"]["mode"], "selected")
        self.assertFalse(payload["selection_summary"]["replace_existing"])
        self.assertEqual(payload["total_inscripcions"], 1)
        self.assertEqual(payload["preview"][0]["existing_team_name"], "Club A")
        self.assertTrue(payload["preview"][0]["will_keep"])
        self.assertFalse(payload["preview"][0]["will_create"])


class EquipContextClassificacioTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Classif Context")
        self.app = self._create_aparell("CTX", "Aparell Context")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        self.team_native = Equip.objects.create(competicio=self.comp, nom="Equip Base")
        self.team_context = Equip.objects.create(competicio=self.comp, nom="Equip Finals")
        self.ctx = EquipContext.objects.create(competicio=self.comp, code="finals", nom="Finals")

        self.ins_a = self._create_inscripcio(self.comp, "Participant A", ordre=1)
        self.ins_b = self._create_inscripcio(self.comp, "Participant B", ordre=2)
        self.ins_a.equip = self.team_native
        self.ins_b.equip = self.team_native
        self.ins_a.save(update_fields=["equip"])
        self.ins_b.save(update_fields=["equip"])

        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=self.ctx,
            inscripcio=self.ins_a,
            equip=self.team_context,
        )

        for ins, total in ((self.ins_a, 9.5), (self.ins_b, 8.0)):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=ins,
                exercici=1,
                comp_aparell=self.comp_app,
                inputs={},
                outputs={},
                total=total,
            )

    def test_compute_classificacio_uses_context_and_fallback_to_native(self):
        schema = {
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
            },
            "equips": {
                "assignment_source": {"mode": "context", "context_code": "finals", "fallback": "native"},
                "incloure_sense_equip": False,
            },
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Per context",
            activa=True,
            ordre=1,
            tipus="equips",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        by_name = {row["participant"]: row for row in rows}
        self.assertIn("Equip Finals", by_name)
        self.assertIn("Equip Base", by_name)
        self.assertEqual(by_name["Equip Finals"]["participants"], 1)
        self.assertEqual(by_name["Equip Base"]["participants"], 1)


class EquipContextHistorySnapshotTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Snapshot Context")
        self.team_native = Equip.objects.create(competicio=self.comp, nom="Equip Base")
        self.team_context = Equip.objects.create(competicio=self.comp, nom="Equip Alt")
        self.ctx = EquipContext.objects.create(competicio=self.comp, code="ctx-alt", nom="Context Alt")
        self.ins = self._create_inscripcio(self.comp, "Participant Snapshot", ordre=1)
        self.ins.equip = self.team_native
        self.ins.save(update_fields=["equip"])
        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=self.ctx,
            inscripcio=self.ins,
            equip=self.team_context,
        )

    def test_snapshot_restores_contexts_and_assignacions(self):
        rf = RequestFactory()
        request = rf.get("/")
        request.session = SessionStore()

        snap = capture_inscripcions_history_snapshot(request, self.comp)
        EquipContext.objects.filter(pk=self.ctx.id).delete()
        self.ins.equip = None
        self.ins.save(update_fields=["equip"])

        apply_inscripcions_history_snapshot(request, self.comp, snap)

        self.ins.refresh_from_db()
        self.assertEqual(self.ins.equip_id, self.team_native.id)
        self.assertTrue(EquipContext.objects.filter(competicio=self.comp, code="ctx-alt").exists())
        self.assertTrue(
            InscripcioEquipAssignacio.objects.filter(
                competicio=self.comp,
                context__code="ctx-alt",
                inscripcio=self.ins,
                equip=self.team_context,
            ).exists()
        )


class TeamContextScoringFlowTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp team scoring")
        User = get_user_model()
        self.user = User.objects.create_user(
            username="team_context_scoring_owner",
            password="testpass123",
            email="team-context-scoring-owner@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.OWNER,
            is_active=True,
        )
        self.client.force_login(self.user)
        self.app = self._create_aparell("SYNC", "Sincronitzat")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1)
        self.ctx = EquipContext.objects.create(
            competicio=self.comp,
            code="parelles",
            nom="Parelles",
        )
        self.other_ctx = EquipContext.objects.create(
            competicio=self.comp,
            code="altre",
            nom="Altre",
        )
        self.equip = Equip.objects.create(competicio=self.comp, nom="Parella 1")
        self.ins1 = self._create_inscripcio(self.comp, "Maria", ordre=1, grup=1)
        self.ins2 = self._create_inscripcio(self.comp, "Laia", ordre=2, grup=1)
        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=self.ctx,
            inscripcio=self.ins1,
            equip=self.equip,
        )
        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=self.ctx,
            inscripcio=self.ins2,
            equip=self.equip,
        )
        self.comp_app.participant_mode = CompeticioAparell.ParticipantMode.TEAM_CONTEXT
        self.comp_app.team_context = self.ctx
        self.comp_app.expected_team_size = 2
        self.comp_app.team_scoring_mode = CompeticioAparell.TeamScoringMode.MEMBERS_PLUS_SHARED
        self.comp_app.save()

    def _create_team_with_members(self, team_name, member_names, *, context=None, start_order=10):
        context = context or self.ctx
        equip = Equip.objects.create(competicio=self.comp, nom=team_name)
        members = []
        for idx, name in enumerate(member_names, start=0):
            ins = self._create_inscripcio(self.comp, name, ordre=start_order + idx, grup=1)
            InscripcioEquipAssignacio.objects.create(
                competicio=self.comp,
                context=context,
                inscripcio=ins,
                equip=equip,
            )
            members.append(ins)
        return equip, members

    def test_scoring_schema_full_clean_accepts_member_scope_for_team_context(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2"},
                ],
            },
        )
        schema.comp_aparell = self.comp_app
        schema.full_clean()

    def test_scoring_save_partial_creates_team_score_entry(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2"},
                ],
            },
        )

        response = self.client.post(
            reverse("scoring_save_partial", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "comp_aparell_id": self.comp_app.id,
                    "subject_kind": "equip",
                    "subject_id": self.equip.id,
                    "exercici": 1,
                    "inputs_patch": {
                        "SYNC": 7.5,
                        "E__m1": 8.1,
                        "E__m2": 8.2,
                    },
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["subject_kind"], "equip")
        self.assertEqual(payload["subject_id"], self.equip.id)

        entry = TeamScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            equip=self.equip,
            exercici=1,
        )
        self.assertEqual(float(entry.total), 23.8)
        self.assertEqual(entry.inputs["E__m1"], 8.1)
        self.assertEqual(entry.inputs["E__m2"], 8.2)

    def test_scoring_save_rejects_individual_payload_for_team_context_app(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [{"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"}],
                "computed": [],
            },
        )

        response = self.client.post(
            reverse("scoring_save", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "comp_aparell_id": self.comp_app.id,
                    "inscripcio_id": self.ins1.id,
                    "exercici": 1,
                    "inputs": {"SYNC": 7.5},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("subject_kind=equip", response.json()["error"])

    def test_judge_save_partial_uses_team_subject_and_runtime_member_permission(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2"},
                ],
            },
        )
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Judge",
            permissions=[
                {"field_code": "SYNC", "runtime_field_code": "SYNC", "scope": "shared", "judge_index": 1},
                {"field_code": "E", "runtime_field_code": "E__m2", "scope": "member", "member_slot": 2, "judge_index": 1},
            ],
            is_active=True,
        )

        response = self.client.post(
            reverse("judge_save_partial", kwargs={"token": token.id}),
            data=json.dumps(
                {
                    "subject_kind": "equip",
                    "subject_id": self.equip.id,
                    "exercici": 1,
                    "inputs_patch": {"SYNC": 6.4, "E__m2": 7.1},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["subject_kind"], "equip")
        self.assertEqual(payload["subject_id"], self.equip.id)

        entry = TeamScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            equip=self.equip,
            exercici=1,
        )
        self.assertEqual(entry.inputs["SYNC"], 6.4)
        self.assertEqual(entry.inputs["E__m2"], 7.1)
        self.assertEqual(entry.inputs["E__m1"], 0.0)

    def test_judge_save_partial_rejects_individual_payload_for_team_context_app(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [{"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"}],
                "computed": [],
            },
        )
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Judge Reject",
            permissions=[{"field_code": "SYNC", "runtime_field_code": "SYNC", "scope": "shared", "judge_index": 1}],
            is_active=True,
        )

        response = self.client.post(
            reverse("judge_save_partial", kwargs={"token": token.id}),
            data=json.dumps(
                {
                    "inscripcio_id": self.ins1.id,
                    "exercici": 1,
                    "inputs_patch": {"SYNC": 6.4},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("subject_kind=equip", response.json()["error"])

    def test_scoring_media_context_accepts_team_subject(self):
        response = self.client.get(
            reverse("scoring_media_context", kwargs={"pk": self.comp.id}),
            {
                "comp_aparell_id": self.comp_app.id,
                "subject_kind": "equip",
                "subject_id": self.equip.id,
                "exercici": 1,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["subject"]["kind"], "equip")
        self.assertEqual(payload["subject"]["id"], self.equip.id)

    def test_judge_video_endpoints_support_team_subjects(self):
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Video",
            permissions=[{"field_code": "SYNC", "scope": "shared", "judge_index": 1}],
            can_record_video=True,
            is_active=True,
        )
        probe_data = {
            "duration_seconds": 9,
            "mime_type": "video/mp4",
            "format_name": "mp4",
            "video_codec": "h264",
        }

        with patch("competicions_trampoli.views_judge._probe_uploaded_video_metadata", return_value=probe_data):
            upload_res = self.client.post(
                reverse("judge_video_upload", kwargs={"token": token.id}),
                data={
                    "subject_kind": "equip",
                    "subject_id": self.equip.id,
                    "exercici": 1,
                    "video_file": SimpleUploadedFile("team.mp4", b"\x00" * 1024, content_type="video/mp4"),
                },
            )

        self.assertEqual(upload_res.status_code, 200)
        upload_payload = upload_res.json()
        self.assertTrue(upload_payload["ok"])
        self.assertEqual(upload_payload["subject_kind"], "equip")
        self.assertEqual(upload_payload["subject_id"], self.equip.id)

        entry = TeamScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            equip=self.equip,
            exercici=1,
        )
        self.assertTrue(TeamScoreEntryVideo.objects.filter(team_score_entry=entry).exists())
        self.assertTrue(
            TeamScoreEntryVideoEvent.objects.filter(
                team_score_entry=entry,
                action=TeamScoreEntryVideoEvent.Action.UPLOAD,
                ok=True,
            ).exists()
        )

        status_res = self.client.get(
            reverse("judge_video_status", kwargs={"token": token.id}),
            {"subject_kind": "equip", "subject_id": self.equip.id, "exercici": 1},
        )
        self.assertEqual(status_res.status_code, 200)
        self.assertTrue(status_res.json()["has_video"])

        delete_res = self.client.post(
            reverse("judge_video_delete", kwargs={"token": token.id}),
            {"subject_kind": "equip", "subject_id": self.equip.id, "exercici": 1},
        )
        self.assertEqual(delete_res.status_code, 200)
        self.assertTrue(delete_res.json()["deleted"])
        self.assertFalse(TeamScoreEntryVideo.objects.filter(team_score_entry=entry).exists())

    def test_compute_classificacio_uses_team_score_entry_for_team_context_app(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [{"code": "TOTAL", "label": "Total", "type": "number", "scope": "shared"}],
                "computed": [],
            },
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            equip=self.equip,
            exercici=1,
            inputs={"TOTAL": 30},
            outputs={},
            total=30,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={},
            outputs={},
            total=2,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins2,
            exercici=1,
            inputs={},
            outputs={},
            total=3,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team direct",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "assignment_source": {"mode": "context", "context_code": "parelles", "fallback": "native"},
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual(rows[0]["participant"], "Parella 1")
        self.assertEqual(rows[0]["score"], 30.0)

    def test_classificacio_validation_rejects_context_mismatch_for_team_context_app(self):
        schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                },
                "equips": {
                    "assignment_source": {"mode": "context", "context_code": "altre", "fallback": "native"},
                },
            },
            tipus="equips",
        )

        self.assertEqual(schema["equips"]["assignment_source"]["context_code"], "altre")
        self.assertTrue(any("requereix context parelles" in err for err in errors))

    def test_competicio_aparell_form_rejects_foreign_context_and_invalid_team_settings(self):
        other_comp = self._create_competicio("Comp externa")
        foreign_ctx = EquipContext.objects.create(competicio=other_comp, code="fora", nom="Fora")
        form = CompeticioAparellForm(
            data={
                "aparell": self.app.id,
                "ordre": 1,
                "actiu": "on",
                "nombre_exercicis": 1,
                "participant_mode": CompeticioAparell.ParticipantMode.TEAM_CONTEXT,
                "team_context": foreign_ctx.id,
                "expected_team_size": 1,
                "team_scoring_mode": "",
            },
            instance=self.comp_app,
            competicio=self.comp,
        )
        self.assertFalse(form.is_valid())
        self.assertIn("team_context", form.errors)
        self.assertIn("expected_team_size", form.errors)
        self.assertIn("team_scoring_mode", form.errors)

    def test_scoring_updates_omits_ineligible_team_entries(self):
        invalid_team, _members = self._create_team_with_members("Parella incompleta", ["Berta"], start_order=20)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            equip=self.equip,
            exercici=1,
            inputs={"SYNC": 5},
            outputs={},
            total=5,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            equip=invalid_team,
            exercici=1,
            inputs={"SYNC": 9},
            outputs={},
            total=9,
        )

        res = self.client.get(
            reverse("scoring_updates", kwargs={"pk": self.comp.id}),
            {
                "since": (timezone.now() - timedelta(minutes=10)).isoformat(),
                "comp_aparell_id": self.comp_app.id,
            },
        )
        self.assertEqual(res.status_code, 200)
        updates = res.json().get("updates", [])
        self.assertEqual({u["subject_id"] for u in updates}, {self.equip.id})

    def test_scoring_media_context_rejects_ineligible_team_subject(self):
        invalid_team, _members = self._create_team_with_members("Parella incompleta", ["Berta"], start_order=20)
        res = self.client.get(
            reverse("scoring_media_context", kwargs={"pk": self.comp.id}),
            {
                "comp_aparell_id": self.comp_app.id,
                "subject_kind": "equip",
                "subject_id": invalid_team.id,
                "exercici": 1,
            },
        )
        self.assertEqual(res.status_code, 403)

    def test_scoring_notes_home_exposes_canonical_score_keys_and_invalid_teams(self):
        invalid_team, _members = self._create_team_with_members("Parella incompleta", ["Berta"], start_order=20)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            equip=self.equip,
            exercici=1,
            inputs={"SYNC": 5},
            outputs={},
            total=5,
        )
        response = self.client.get(reverse("scoring_notes_home", kwargs={"pk": self.comp.id}))
        self.assertEqual(response.status_code, 200)
        scores = response.context["scores"]
        self.assertIn(f"equip:{self.equip.id}|1|{self.comp_app.id}", scores)
        subjects = {str(item["id"]): item for item in response.context["inscripcions"]}
        self.assertIn(f"equip:{invalid_team.id}", subjects)
        self.assertTrue(subjects[f"equip:{invalid_team.id}"]["invalid_reasons"])

    def test_compute_classificacio_team_tie_break_uses_team_score_entries(self):
        equip2, members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            equip=self.equip,
            exercici=1,
            inputs={"TOTAL": 30},
            outputs={"SYNC": 9},
            total=30,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            equip=equip2,
            exercici=1,
            inputs={"TOTAL": 30},
            outputs={"SYNC": 8},
            total=30,
        )
        for ins, raw_total in [(self.ins1, 1), (self.ins2, 1), (members2[0], 50), (members2[1], 50)]:
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=self.comp_app,
                inscripcio=ins,
                exercici=1,
                inputs={},
                outputs={"SYNC": raw_total},
                total=raw_total,
            )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team tiebreak",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camp": "SYNC",
                        "ordre": "desc",
                        "scope": {"aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]}},
                    }
                ],
                "equips": {
                    "assignment_source": {"mode": "context", "context_code": "parelles", "fallback": "native"},
                    "incloure_sense_equip": False,
                },
            },
        )
        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([row["participant"] for row in rows[:2]], ["Parella 1", "Parella 2"])

    def test_scoreable_codes_filter_team_apps_by_tipus_and_context(self):
        individual_app = self._create_aparell("IND", "Individual")
        individual_comp_app = self._create_comp_aparell(self.comp, individual_app, ordre=2)
        ScoringSchema.objects.create(
            aparell=individual_app,
            schema={"fields": [{"code": "TOTAL", "type": "number"}], "computed": []},
        )
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={"fields": [{"code": "TOTAL", "type": "number"}], "computed": []},
        )
        individual_scoreables = _scoreable_codes_by_app_id(self.comp, tipus="individual")
        self.assertIn(individual_comp_app.id, individual_scoreables)
        self.assertNotIn(self.comp_app.id, individual_scoreables)

        team_scoreables = _scoreable_codes_by_app_id(
            self.comp,
            tipus="equips",
            assignment_context_code="altre",
        )
        self.assertNotIn(self.comp_app.id, team_scoreables)
