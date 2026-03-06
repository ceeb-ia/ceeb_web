# views_classificacions.py
import json
import re
from django.shortcuts import get_object_or_404
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.db import transaction
from django.urls import reverse
from django.utils.text import slugify
from .models_scoring import ScoreEntry, ScoringSchema  
from .models import Competicio, Inscripcio, Equip
from .models_trampoli import CompeticioAparell, Aparell
from .models_classificacions import ClassificacioConfig, ClassificacioTemplateGlobal
from .models_judging import PublicLiveToken
from .services.services_classificacions_2 import compute_classificacio, DEFAULT_SCHEMA, get_display_columns
from .views import get_allowed_group_fields, get_inscripcio_value
from .access import user_has_competicio_capability
from .services.scoring_schema_validation import (
    _field_shape,
    _build_alias_map,
    _resolve_name,
    _topo_sort,
    _ast_parse,
    _extract_names,
    DryRunEval,
    TMat,
    Shape,
    RESERVED_NAMES,
    ALLOWED_FUNCTIONS,
)
from django.db import models
# views_classificacions.py
from django.utils.dateparse import parse_datetime
from django.utils.timezone import is_aware
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _active_cfg_values(competicio):
    cfgs = (
        ClassificacioConfig.objects
        .filter(competicio=competicio, activa=True)
        .order_by("ordre", "id")
    )
    return list(cfgs.values("id", "nom", "tipus", "ordre"))


def _live_data_payload(competicio, since_raw=None):
    last_note = (
        ScoreEntry.objects
        .filter(competicio=competicio)
        .order_by("-updated_at")
        .values_list("updated_at", flat=True)
        .first()
    )
    last_cfg = (
        ClassificacioConfig.objects
        .filter(competicio=competicio)
        .order_by("-updated_at")
        .values_list("updated_at", flat=True)
        .first()
    )
    stamp = max([d for d in [last_note, last_cfg] if d is not None], default=timezone.now())

    if since_raw:
        since_dt = parse_datetime(since_raw)
        if since_dt and not is_aware(since_dt):
            since_dt = timezone.make_aware(since_dt, timezone.get_current_timezone())
        if since_dt and stamp <= since_dt:
            return {"ok": True, "changed": False, "stamp": stamp.isoformat()}

    cfgs = (
        ClassificacioConfig.objects
        .filter(competicio=competicio, activa=True)
        .order_by("ordre", "id")
    )

    payload_cfgs = []
    for cfg in cfgs:
        data = compute_classificacio(competicio, cfg)
        parts = []
        for k in sorted(data.keys()):
            parts.append({"particio": k, "rows": data[k]})
        payload_cfgs.append({
            "id": cfg.id,
            "nom": cfg.nom,
            "tipus": cfg.tipus,
            "columns": get_display_columns(cfg.schema or {}),
            "parts": parts,
        })

    return {
        "ok": True,
        "changed": True,
        "stamp": stamp.isoformat(),
        "competicio": {"id": competicio.id, "nom": competicio.nom},
        "cfgs": payload_cfgs,
    }


def _default_live_columns():
    return [
        {"type": "builtin", "key": "posicio", "label": "Pos.", "align": "left"},
        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
        {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
    ]


def _format_partition_title(raw):
    source = "global" if raw in (None, "") else str(raw)
    tokens = []
    for part in source.split("|"):
        token = str(part or "").strip()
        if not token:
            continue
        idx = token.find(":")
        tokens.append(token[idx + 1:].strip() if idx >= 0 else token)
    return " / ".join([x for x in tokens if x]) or "global"


def _fallback_export_value(row: dict, key: str):
    if key in ("participant", "nom"):
        return row.get("participant") or row.get("nom") or row.get("entitat_nom") or ""
    if key == "punts":
        return row.get("punts")
    if key == "posicio":
        return row.get("posicio")
    if key == "entitat_nom":
        return row.get("entitat_nom") or ""
    if key == "participants":
        return row.get("participants") if row.get("participants") is not None else ""
    return row.get(key)


def _extract_export_value(row: dict, col: dict):
    key = str((col or {}).get("key") or "")
    cells = row.get("cells") or row.get("display") or {}
    if isinstance(cells, dict) and key in cells:
        return cells.get(key)
    return _fallback_export_value(row, key)


def _excel_safe_text(raw) -> str:
    txt = "" if raw is None else str(raw)
    if txt[:1] in ("=", "+", "-", "@"):
        return "'" + txt
    return txt


def _try_float(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        txt = value.strip()
        if not txt:
            return None
        try:
            return float(txt)
        except Exception:
            return None
    return None


def _format_scalar_text(value, decimals=None):
    if value is None:
        return ""
    number = _try_float(value)
    if number is not None and decimals is not None:
        return f"{number:.{decimals}f}"
    if number is not None and isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _normalize_excel_cell(value, col):
    decimals = (col or {}).get("decimals")
    try:
        decimals = int(decimals) if decimals is not None else None
    except Exception:
        decimals = None
    if decimals is not None:
        decimals = max(0, min(6, decimals))

    if isinstance(value, dict) and value.get("_kind") == "judge_rows":
        lines = []
        for row in value.get("rows") or []:
            judge_num = row.get("judge")
            try:
                judge_num = int(judge_num)
            except Exception:
                judge_num = None
            judge_label = f"J{judge_num}" if judge_num and judge_num > 0 else "J?"
            items = row.get("items") or []
            if not isinstance(items, list):
                items = [items]
            vals = [_format_scalar_text(it, decimals) for it in items]
            lines.append(f"{judge_label}: " + " | ".join(vals))
        return _excel_safe_text("\n".join([ln for ln in lines if ln])), None, True

    if isinstance(value, (dict, list)):
        return _excel_safe_text(json.dumps(value, ensure_ascii=False)), None, True

    if isinstance(value, bool):
        return value, None, False

    number = _try_float(value)
    if number is not None and not isinstance(value, str):
        if decimals is None:
            if isinstance(value, int):
                return int(value), "0", False
            return float(value), "0.###", False
        rounded = round(number, decimals)
        num_fmt = "0" if decimals == 0 else ("0." + ("0" * decimals))
        return rounded, num_fmt, False

    if isinstance(value, str):
        txt = value.strip()
        if decimals is not None:
            number = _try_float(txt)
            if number is not None:
                rounded = round(number, decimals)
                num_fmt = "0" if decimals == 0 else ("0." + ("0" * decimals))
                return rounded, num_fmt, False
        return _excel_safe_text(value), None, ("\n" in value)

    if value is None:
        return "", None, False
    return _excel_safe_text(value), None, False


def _build_excel_sheet_name(raw_name, used_names):
    base = str(raw_name or "").strip() or "Classificacio"
    base = re.sub(r'[\[\]\*:/\\?]', " ", base)
    base = " ".join(base.split())
    if not base:
        base = "Classificacio"
    base = base[:31]

    candidate = base
    idx = 2
    while candidate.casefold() in used_names:
        suffix = f" ({idx})"
        max_len = 31 - len(suffix)
        truncated = (base[:max_len].rstrip() if max_len > 0 else "") or "Sheet"
        candidate = f"{truncated}{suffix}"
        idx += 1
    used_names.add(candidate.casefold())
    return candidate


def _write_cfg_excel_sheet(ws, competicio, cfg_nom, columns, parts):
    cols = columns if isinstance(columns, list) and columns else _default_live_columns()
    total_cols = max(1, len(cols))

    fill_title = PatternFill("solid", fgColor="1F4E79")
    fill_subtitle = PatternFill("solid", fgColor="D9E1F2")
    fill_partition = PatternFill("solid", fgColor="DDE7FF")
    fill_header = PatternFill("solid", fgColor="E9EEF7")
    fill_zebra = PatternFill("solid", fgColor="F7F9FC")
    fill_first = PatternFill("solid", fgColor="F6E27A")
    fill_second = PatternFill("solid", fgColor="E3E8EF")
    fill_third = PatternFill("solid", fgColor="E8C7A3")

    thin = Side(style="thin", color="9AA7B2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    subtitle_font = Font(bold=True, size=11)
    header_font = Font(bold=True)

    col_widths = [10 for _ in range(total_cols)]

    def update_col_width(col_idx, value):
        txt = "" if value is None else str(value)
        longest = max([len(line) for line in txt.splitlines()] or [0])
        col_widths[col_idx - 1] = min(48, max(col_widths[col_idx - 1], longest + 2))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c_title = ws.cell(row=1, column=1, value=f"{competicio.nom} - {cfg_nom}")
    c_title.fill = fill_title
    c_title.font = title_font
    c_title.alignment = align_center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c_sub = ws.cell(
        row=2,
        column=1,
        value=f"Exportat el {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}",
    )
    c_sub.fill = fill_subtitle
    c_sub.font = subtitle_font
    c_sub.alignment = align_center

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20

    row_idx = 4
    first_data_row = None
    parts_list = parts if isinstance(parts, list) else []
    if not parts_list:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        empty = ws.cell(row=row_idx, column=1, value="Sense resultats.")
        empty.alignment = align_left
        empty.font = Font(italic=True)
        return

    for p in parts_list:
        part_rows = (p or {}).get("rows") or []
        part_name = _format_partition_title((p or {}).get("particio"))

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        pc = ws.cell(row=row_idx, column=1, value=f"Particio: {part_name} ({len(part_rows)} files)")
        pc.fill = fill_partition
        pc.font = Font(bold=True)
        pc.alignment = align_left
        pc.border = border
        row_idx += 1

        header_row = row_idx
        if first_data_row is None:
            first_data_row = header_row + 1
        for col_pos, col in enumerate(cols, start=1):
            label = str(col.get("label") or col.get("key") or "")
            hc = ws.cell(row=header_row, column=col_pos, value=label)
            hc.fill = fill_header
            hc.font = header_font
            hc.alignment = align_center
            hc.border = border
            update_col_width(col_pos, label)
        row_idx += 1

        if not part_rows:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
            rc = ws.cell(row=row_idx, column=1, value="Sense resultats en aquesta particio.")
            rc.alignment = align_left
            rc.font = Font(italic=True)
            row_idx += 2
            continue

        for data_pos, row in enumerate(part_rows):
            try:
                posicio = int((row or {}).get("posicio"))
            except Exception:
                posicio = None

            if posicio == 1:
                podium_fill = fill_first
            elif posicio == 2:
                podium_fill = fill_second
            elif posicio == 3:
                podium_fill = fill_third
            else:
                podium_fill = None

            max_lines = 1
            for col_pos, col in enumerate(cols, start=1):
                raw_value = _extract_export_value(row or {}, col or {})
                value, number_format, wrap_text = _normalize_excel_cell(raw_value, col or {})
                cell = ws.cell(row=row_idx, column=col_pos, value=value)
                cell.border = border
                if number_format:
                    cell.number_format = number_format

                align = str((col or {}).get("align") or "").strip().lower()
                horizontal = "right" if align == "right" else ("center" if align == "center" else "left")
                vertical = "top" if wrap_text else "center"
                cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)

                if podium_fill is not None:
                    cell.fill = podium_fill
                elif data_pos % 2 == 1:
                    cell.fill = fill_zebra

                value_txt = "" if value is None else str(value)
                max_lines = max(max_lines, len(value_txt.splitlines()) if value_txt else 1)
                update_col_width(col_pos, value)

            if max_lines > 1:
                ws.row_dimensions[row_idx].height = min(120, max(18, max_lines * 14))
            row_idx += 1

        row_idx += 1

    if first_data_row:
        ws.freeze_panes = ws[f"A{first_data_row}"]

    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(10, width)


def _sanitize_filename_component(raw):
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(raw or "").strip()).strip("._")
    return cleaned or "classificacions"


def classificacions_live_export_excel(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    cfg_qs = (
        ClassificacioConfig.objects
        .filter(competicio=competicio, activa=True)
        .order_by("ordre", "id")
    )

    cfg_id_raw = (request.GET.get("cfg_id") or "").strip()
    selected_cfg_id = None
    if cfg_id_raw:
        try:
            selected_cfg_id = int(cfg_id_raw)
        except Exception:
            return HttpResponseBadRequest("cfg_id invalid")
        cfg_qs = cfg_qs.filter(id=selected_cfg_id)

    cfgs = list(cfg_qs)
    if not cfgs:
        return HttpResponseBadRequest("No hi ha classificacions actives per exportar.")

    wb = Workbook()
    used_sheet_names = set()

    for idx, cfg in enumerate(cfgs):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = _build_excel_sheet_name(cfg.nom or f"Classificacio {idx + 1}", used_sheet_names)

        data = compute_classificacio(competicio, cfg)
        parts = [{"particio": k, "rows": data[k]} for k in sorted(data.keys())]
        columns = get_display_columns(cfg.schema or {}) or _default_live_columns()
        _write_cfg_excel_sheet(ws, competicio, cfg.nom or f"Classificacio {idx + 1}", columns, parts)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    suffix = f"_cfg_{selected_cfg_id}" if selected_cfg_id else ""
    filename = f"classificacions_{_sanitize_filename_component(competicio.nom)}{suffix}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class ClassificacionsLive(TemplateView):
    template_name = "competicio/classificacions_live.html"

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        competicio = self.competicio
        public_raw = (self.request.GET.get("public") or "").strip().lower()
        is_public = public_raw in {"1", "true", "yes", "on"}

        cfgs = (
            ClassificacioConfig.objects
            .filter(competicio=competicio, activa=True)
            .order_by("ordre", "id")
        )
        ctx.update({
            "competicio": competicio,
            "cfgs": list(cfgs.values("id", "nom", "tipus", "ordre")),
            "is_public": is_public,
            "hide_base_chrome": is_public,
            # interval suggerit (ms) perquè el JS el pugui usar
            "poll_ms": 4000,
        })
        return ctx


class ClassificacionsLoopLive(TemplateView):
    template_name = "competicio/classificacions_loop_live.html"

    @staticmethod
    def _parse_int_param(raw, default: int, min_value: int, max_value: int) -> int:
        try:
            value = int(raw)
        except Exception:
            value = default
        return max(min_value, min(max_value, value))

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        competicio = self.competicio

        public_raw = (self.request.GET.get("public") or "").strip().lower()
        is_public = public_raw in {"1", "true", "yes", "on"}

        cfgs = (
            ClassificacioConfig.objects
            .filter(competicio=competicio, activa=True)
            .order_by("ordre", "id")
        )

        poll_ms = self._parse_int_param(self.request.GET.get("poll_ms"), 4000, 1000, 60000)
        slide_ms = self._parse_int_param(self.request.GET.get("slide_ms"), 8000, 2000, 120000)
        rows_per_page = self._parse_int_param(self.request.GET.get("rows"), 12, 3, 60)

        transition = (self.request.GET.get("transition") or "fade").strip().lower()
        if transition not in {"fade", "none"}:
            transition = "fade"

        ctx.update({
            "competicio": competicio,
            "cfgs": list(cfgs.values("id", "nom", "tipus", "ordre")),
            "is_public": is_public,
            "hide_base_chrome": is_public,
            "poll_ms": poll_ms,
            "slide_ms": slide_ms,
            "rows_per_page": rows_per_page,
            "transition": transition,
        })
        return ctx


def classificacions_live_data(request, pk):
    """
    GET /competicio/<pk>/classificacions/live/data?since=<iso>
    Retorna:
      - changed: False si no hi ha canvis des de `since`
      - changed: True + dades si hi ha canvis
    """
    competicio = get_object_or_404(Competicio, pk=pk)

    # stamp = max(última nota, última cfg)
    last_note = (
        ScoreEntry.objects
            .filter(competicio=competicio)
            .order_by("-updated_at")
            .values_list("updated_at", flat=True)
            .first()    
        )
    last_cfg = (
        ClassificacioConfig.objects
        .filter(competicio=competicio)
        .order_by("-updated_at")
        .values_list("updated_at", flat=True)
        .first()
    )
    stamp = max([d for d in [last_note, last_cfg] if d is not None], default=timezone.now())

    # Si el client passa ?since=..., i no hi ha canvis, no recalculis
    since_raw = request.GET.get("since")
    if since_raw:
        since_dt = parse_datetime(since_raw)
        if since_dt and not is_aware(since_dt):
            since_dt = timezone.make_aware(since_dt, timezone.get_current_timezone())
        if since_dt and stamp <= since_dt:
            return JsonResponse({"ok": True, "changed": False, "stamp": stamp.isoformat()})

    cfgs = (
        ClassificacioConfig.objects
        .filter(competicio=competicio, activa=True)
        .order_by("ordre", "id")
    )

    payload_cfgs = []
    for cfg in cfgs:
        data = compute_classificacio(competicio, cfg)  # {particio_key: [rows]}
        parts = []
        for k in sorted(data.keys()):
            parts.append({"particio": k, "rows": data[k]})
        payload_cfgs.append({
            "id": cfg.id,
            "nom": cfg.nom,
            "tipus": cfg.tipus,
            "columns": get_display_columns(cfg.schema or {}),
            "parts": parts,
        })

    return JsonResponse({
        "ok": True,
        "changed": True,
        "stamp": stamp.isoformat(),
        "competicio": {"id": competicio.id, "nom": competicio.nom},
        "cfgs": payload_cfgs,
    })


class PublicClassificacionsLive(TemplateView):
    template_name = "competicio/classificacions_live.html"

    def dispatch(self, request, *args, **kwargs):
        self.token_obj = get_object_or_404(PublicLiveToken, pk=kwargs["token"])
        if not self.token_obj.is_valid():
            return JsonResponse({"ok": False, "error": "Token invalid o revocat"}, status=403)
        self.token_obj.touch()
        self.competicio = self.token_obj.competicio
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        cfgs = (
            ClassificacioConfig.objects
            .filter(competicio=self.competicio, activa=True)
            .order_by("ordre", "id")
        )
        ctx.update({
            "competicio": self.competicio,
            "cfgs": list(cfgs.values("id", "nom", "tipus", "ordre")),
            "is_public": True,
            "hide_base_chrome": True,
            "poll_ms": 4000,
            "public_token_can_view_media": bool(self.token_obj.can_view_media),
            "data_url": self.request.build_absolute_uri(
                reverse("public_live_classificacions_data", kwargs={"token": self.token_obj.id})
            ),
        })
        return ctx


class PublicClassificacionsLoopLive(TemplateView):
    template_name = "competicio/classificacions_loop_live.html"

    @staticmethod
    def _parse_int_param(raw, default: int, min_value: int, max_value: int) -> int:
        try:
            value = int(raw)
        except Exception:
            value = default
        return max(min_value, min(max_value, value))

    def dispatch(self, request, *args, **kwargs):
        self.token_obj = get_object_or_404(PublicLiveToken, pk=kwargs["token"])
        if not self.token_obj.is_valid():
            return JsonResponse({"ok": False, "error": "Token invalid o revocat"}, status=403)
        self.token_obj.touch()
        self.competicio = self.token_obj.competicio
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        cfgs = (
            ClassificacioConfig.objects
            .filter(competicio=self.competicio, activa=True)
            .order_by("ordre", "id")
        )
        poll_ms = self._parse_int_param(self.request.GET.get("poll_ms"), 4000, 1000, 60000)
        slide_ms = self._parse_int_param(self.request.GET.get("slide_ms"), 8000, 2000, 120000)
        rows_per_page = self._parse_int_param(self.request.GET.get("rows"), 12, 3, 60)
        transition = (self.request.GET.get("transition") or "fade").strip().lower()
        if transition not in {"fade", "none"}:
            transition = "fade"

        ctx.update({
            "competicio": self.competicio,
            "cfgs": list(cfgs.values("id", "nom", "tipus", "ordre")),
            "is_public": True,
            "hide_base_chrome": True,
            "poll_ms": poll_ms,
            "slide_ms": slide_ms,
            "rows_per_page": rows_per_page,
            "transition": transition,
            "public_token_can_view_media": bool(self.token_obj.can_view_media),
            "data_url": self.request.build_absolute_uri(
                reverse("public_live_classificacions_data", kwargs={"token": self.token_obj.id})
            ),
        })
        return ctx


def public_classificacions_live_data(request, token):
    token_obj = get_object_or_404(PublicLiveToken, pk=token)
    if not token_obj.is_valid():
        return JsonResponse({"ok": False, "error": "Token invalid o revocat"}, status=403)

    competicio = token_obj.competicio
    last_note = (
        ScoreEntry.objects
            .filter(competicio=competicio)
            .order_by("-updated_at")
            .values_list("updated_at", flat=True)
            .first()
        )
    last_cfg = (
        ClassificacioConfig.objects
        .filter(competicio=competicio)
        .order_by("-updated_at")
        .values_list("updated_at", flat=True)
        .first()
    )
    stamp = max([d for d in [last_note, last_cfg] if d is not None], default=timezone.now())

    since_raw = request.GET.get("since")
    if since_raw:
        since_dt = parse_datetime(since_raw)
        if since_dt and not is_aware(since_dt):
            since_dt = timezone.make_aware(since_dt, timezone.get_current_timezone())
        if since_dt and stamp <= since_dt:
            return JsonResponse(
                {
                    "ok": True,
                    "changed": False,
                    "stamp": stamp.isoformat(),
                    "permissions": {"can_view_media": bool(token_obj.can_view_media)},
                }
            )

    cfgs = (
        ClassificacioConfig.objects
        .filter(competicio=competicio, activa=True)
        .order_by("ordre", "id")
    )

    payload_cfgs = []
    for cfg in cfgs:
        data = compute_classificacio(competicio, cfg)
        parts = []
        for k in sorted(data.keys()):
            parts.append({"particio": k, "rows": data[k]})
        payload_cfgs.append({
            "id": cfg.id,
            "nom": cfg.nom,
            "tipus": cfg.tipus,
            "columns": get_display_columns(cfg.schema or {}),
            "parts": parts,
        })

    return JsonResponse({
        "ok": True,
        "changed": True,
        "stamp": stamp.isoformat(),
        "competicio": {"id": competicio.id, "nom": competicio.nom},
        "permissions": {"can_view_media": bool(token_obj.can_view_media)},
        "cfgs": payload_cfgs,
    })


class ClassificacionsHome(TemplateView):
    template_name = "competicio/classificacions_builder_v2.html"

    def get(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        competicio = self.competicio

        # aparells actius (com a notes)
        aparells_cfg = CompeticioAparell.objects.filter(
            competicio=competicio,
            actiu=True,
        ).select_related("aparell").order_by("ordre", "id")

        # si no n'hi ha, crea'n un per defecte només si l'usuari pot editar classificacions
        if (
            not aparells_cfg.exists()
            and user_has_competicio_capability(self.request.user, competicio, "classificacions.edit")
        ):
            a, _ = Aparell.objects.get_or_create(
                codi="TRAMP",
                created_by=self.request.user,
                defaults={"nom": "Trampolí", "actiu": True},
            )
            CompeticioAparell.objects.get_or_create(
                competicio=competicio,
                aparell=a,
                defaults={
                    "ordre": 1,
                    "actiu": True,
                },
            )
            aparells_cfg = CompeticioAparell.objects.filter(
                competicio=competicio,
                actiu=True,
            ).select_related("aparell").order_by("ordre", "id")

        aparell_ids = list(aparells_cfg.values_list("aparell_id", flat=True))

        schemas_by_aparell = {
            s.aparell_id: (s.schema or {})
            for s in ScoringSchema.objects.filter(aparell_id__in=aparell_ids).only("aparell_id", "schema")
        }

        aparell_field_options = {}


        for ca in aparells_cfg:
            sch = schemas_by_aparell.get(ca.aparell_id, {}) or {}
            opts = []
            score_meta = _build_scoreable_meta_for_schema(sch)


            for f in (sch.get("fields") or []):
                if isinstance(f, dict) and f.get("code"):
                    code = str(f["code"])
                    info = score_meta.get(code) or {}
                    if not info.get("scoreable", False):
                        continue
                    judges_count = 1
                    j = f.get("judges")
                    if isinstance(j, dict):
                        try:
                            judges_count = int(j.get("count") or 1)
                        except Exception:
                            judges_count = 1
                    else:
                        try:
                            judges_count = int(f.get("judges_count") or 1)
                        except Exception:
                            judges_count = 1
                    judges_count = max(1, judges_count)
                    opts.append({
                        "code": code,
                        "label": str(f.get("label") or code),
                        "kind": "field",
                        "judges_count": judges_count,
                    })

            for c in (sch.get("computed") or []):
                if isinstance(c, dict) and c.get("code"):
                    code = str(c["code"])
                    info = score_meta.get(code) or {}
                    if not info.get("scoreable", False):
                        continue
                    opts.append({
                        "code": code,
                        "label": str(c.get("label") or code),
                        "kind": "computed",
                        "judges_count": 1,
                    })

            # dedup
            seen = set()
            dedup = []
            for o in opts:
                if o["code"] in seen:
                    continue
                seen.add(o["code"])
                dedup.append(o)

            aparell_field_options[str(ca.id)] = dedup

        cfgs = ClassificacioConfig.objects.filter(competicio=competicio).order_by("ordre", "id")

        # si no hi ha cap configuració, crea'n una de base
        if not cfgs.exists():
            ClassificacioConfig.objects.create(
                competicio=competicio,
                nom="General (total)",
                activa=True,
                ordre=1,
                tipus="individual",
                schema={
                    **DEFAULT_SCHEMA,
                    "particions": [],
                    "puntuacio": {**DEFAULT_SCHEMA["puntuacio"], "camp": "total", "agregacio": "sum"},
                    "presentacio": {"top_n": 0, "mostrar_empats": True},
                },
            )
            cfgs = ClassificacioConfig.objects.filter(competicio=competicio).order_by("ordre", "id")


        # Choices per filtres visuals (entitat/categoria/subcategoria/grup)
        ins_qs = Inscripcio.objects.filter(competicio=competicio)

        filter_choices = {}

        # entitat
        if _is_fk(Inscripcio, "entitat"):
            filter_choices["entitats"] = _distinct_fk(ins_qs, "entitat")
        else:
            filter_choices["entitats"] = [{"value": v, "label": str(v)} for v in _distinct_values(ins_qs, "entitat")]

        # categoria
        if _is_fk(Inscripcio, "categoria"):
            filter_choices["categories"] = _distinct_fk(ins_qs, "categoria")
        else:
            filter_choices["categories"] = [{"value": v, "label": str(v)} for v in _distinct_values(ins_qs, "categoria")]

        # subcategoria
        if _is_fk(Inscripcio, "subcategoria"):
            filter_choices["subcategories"] = _distinct_fk(ins_qs, "subcategoria")
        else:
            filter_choices["subcategories"] = [{"value": v, "label": str(v)} for v in _distinct_values(ins_qs, "subcategoria")]

        # grup (normalment text)
        filter_choices["grups"] = [{"value": v, "label": str(v)} for v in _distinct_values(ins_qs, "grup")]

        ctx["filter_choices"] = filter_choices

        particio_fields = []
        for item in get_allowed_group_fields(competicio):
            if not isinstance(item, dict):
                continue
            code = str(item.get("code") or "").strip()
            if not code:
                continue
            particio_fields.append(
                {
                    "code": code,
                    "label": str(item.get("label") or code),
                    "ui_label": str(item.get("ui_label") or item.get("label") or code),
                    "kind": str(item.get("kind") or "builtin"),
                    "source": str(item.get("source") or ""),
                }
            )

        ins_list = list(ins_qs)
        particio_value_choices = _collect_particio_value_choices(
            ins_list,
            [f["code"] for f in particio_fields],
        )

        ctx["particio_fields"] = particio_fields
        ctx["particio_value_choices"] = particio_value_choices


        cfg_payload = []
        for c in cfgs:
            cfg_payload.append({
                "id": c.id,
                "nom": c.nom,
                "activa": c.activa,
                "ordre": c.ordre,
                "tipus": c.tipus,
                "schema": c.schema or {},
            })

        aparell_payload = []
        for ca in aparells_cfg:
            aparell_payload.append({
                "id": ca.id,
                "nom": ca.aparell.nom,
                "codi": ca.aparell.codi,
                "nombre_exercicis": int(getattr(ca, "nombre_exercicis", 1) or 1),
            })

        equips_qs = (
            Equip.objects
            .filter(competicio=competicio)
            .annotate(membres_count=models.Count("membres"))
            .order_by("nom", "id")
        )
        equips_payload = []
        for e in equips_qs:
            equips_payload.append({
                "id": e.id,
                "nom": e.nom,
                "origen": e.origen,
                "membres_count": int(getattr(e, "membres_count", 0) or 0),
            })

        ctx.update({
            "competicio": competicio,
            "cfgs": cfg_payload,
            "aparells": aparell_payload,
            "equips": equips_payload,
            "can_manage_global_templates": bool(
                user_has_competicio_capability(self.request.user, competicio, "classificacions.edit")
            ),
        })

        ctx.update({
            "aparell_field_options": aparell_field_options,
        })

        return ctx


def _safe_int(v, default=0):
    try:
        return int(v)
    except Exception:
        return default


def _normalize_particio_codes(raw):
    out = []
    seen = set()
    if not isinstance(raw, list):
        return out
    for item in raw:
        code = str(item or "").strip()
        if not code or code in seen:
            continue
        seen.add(code)
        out.append(code)
    return out


def _split_particio_custom_values(raw):
    if isinstance(raw, list):
        values = [str(x or "").strip() for x in raw]
    elif isinstance(raw, str):
        values = [x.strip() for x in raw.split(",")]
    else:
        values = []

    out = []
    seen = set()
    for txt in values:
        if not txt:
            continue
        key = " ".join(txt.split()).casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(txt)
    return out


def _normalize_particions_custom(raw):
    out = {}
    if not isinstance(raw, dict):
        return out

    for field_code, cfg in raw.items():
        code = str(field_code or "").strip()
        if not code or not isinstance(cfg, dict):
            continue

        mode = str(cfg.get("mode") or "raw").strip().lower()
        if mode not in {"raw", "custom"}:
            mode = "raw"

        fallback_label = str(cfg.get("fallback_label") or "").strip()
        groups = []
        for idx, group in enumerate(cfg.get("grups") or []):
            if not isinstance(group, dict):
                continue
            label = (
                str(group.get("label") or group.get("key") or f"Grup {idx + 1}").strip()
                or f"Grup {idx + 1}"
            )
            values = _split_particio_custom_values(group.get("values"))
            if not values and not label:
                continue
            groups.append(
                {
                    "key": str(group.get("key") or f"grp_{idx + 1}").strip() or f"grp_{idx + 1}",
                    "label": label,
                    "values": values,
                }
            )

        out[code] = {
            "mode": mode,
            "fallback_label": fallback_label,
            "grups": groups,
        }
    return out


def _normalize_particions_schema(schema):
    if not isinstance(schema, dict):
        return {}
    out = dict(schema)
    out["particions"] = _normalize_particio_codes(schema.get("particions") or [])
    out["particions_custom"] = _normalize_particions_custom(schema.get("particions_custom") or {})
    return out


def _json_clone(value):
    try:
        return json.loads(json.dumps(value, ensure_ascii=False))
    except Exception:
        return {}


def _canon_app_code(raw) -> str:
    return str(raw or "").strip().upper()


def _get_comp_aparell_maps(competicio, active_only=True):
    qs = CompeticioAparell.objects.filter(competicio=competicio).select_related("aparell").order_by("ordre", "id")
    if active_only:
        qs = qs.filter(actiu=True)
    apps = list(qs)
    by_id = {int(ca.id): ca for ca in apps}
    by_code = {}
    for ca in apps:
        code = _canon_app_code(getattr(ca.aparell, "codi", ""))
        if code and code not in by_code:
            by_code[code] = ca
    return apps, by_id, by_code


def _resolve_app_code_for_template(raw, by_id, by_code):
    if raw is None:
        return None
    if isinstance(raw, str):
        txt = raw.strip()
        if not txt:
            return None
        code = _canon_app_code(txt)
        if code in by_code:
            return code
        try:
            app_id = int(txt)
        except Exception:
            return None
        ca = by_id.get(app_id)
        if not ca:
            return None
        return _canon_app_code(getattr(ca.aparell, "codi", ""))
    try:
        app_id = int(raw)
    except Exception:
        return None
    ca = by_id.get(app_id)
    if not ca:
        return None
    return _canon_app_code(getattr(ca.aparell, "codi", ""))


def _resolve_app_id_for_competicio(raw, by_id, by_code):
    if raw is None:
        return None
    if isinstance(raw, str):
        txt = raw.strip()
        if not txt:
            return None
        code = _canon_app_code(txt)
        if code in by_code:
            return int(by_code[code].id)
        try:
            app_id = int(txt)
        except Exception:
            return None
        return app_id if app_id in by_id else None
    try:
        app_id = int(raw)
    except Exception:
        return None
    return app_id if app_id in by_id else None


def _extract_template_schema(payload_or_schema):
    if not isinstance(payload_or_schema, dict):
        return {}
    if isinstance(payload_or_schema.get("schema"), dict):
        return _json_clone(payload_or_schema.get("schema") or {})
    return _json_clone(payload_or_schema)


def _schema_to_template_schema(competicio, schema_local):
    schema = _json_clone(schema_local or {})
    warnings = []

    _apps_all, by_id_all, by_code_all = _get_comp_aparell_maps(competicio, active_only=False)

    punt = schema.get("puntuacio") or {}
    if not isinstance(punt, dict):
        punt = {}
    schema["puntuacio"] = punt

    raw_apps = punt.get("aparells") or {}
    if not isinstance(raw_apps, dict):
        raw_apps = {}
    ids_in = raw_apps.get("ids") or []
    ids_out = []
    seen_codes = set()
    for raw in ids_in if isinstance(ids_in, list) else []:
        code = _resolve_app_code_for_template(raw, by_id_all, by_code_all)
        if not code:
            warnings.append(f"Aparell no exportat (ids): {raw}")
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)
        ids_out.append(code)
    punt["aparells"] = {"mode": "seleccionar", "ids": ids_out}

    def map_keys_to_codes(raw_map, field_label):
        out = {}
        if not isinstance(raw_map, dict):
            return out
        for raw_key, raw_value in raw_map.items():
            code = _resolve_app_code_for_template(raw_key, by_id_all, by_code_all)
            if not code:
                warnings.append(f"{field_label}: aparell no exportat ({raw_key})")
                continue
            out[code] = raw_value
        return out

    punt["camps_per_aparell"] = map_keys_to_codes(punt.get("camps_per_aparell") or {}, "camps_per_aparell")
    punt["exercicis_per_aparell"] = map_keys_to_codes(
        punt.get("exercicis_per_aparell") or {},
        "exercicis_per_aparell",
    )

    desempat = schema.get("desempat") or []
    if not isinstance(desempat, list):
        desempat = []
    des_out = []
    for idx, tie in enumerate(desempat):
        if not isinstance(tie, dict):
            des_out.append(tie)
            continue
        item = dict(tie)

        raw_legacy_app = item.get("aparell_id")
        if raw_legacy_app not in (None, "", 0, "0"):
            code = _resolve_app_code_for_template(raw_legacy_app, by_id_all, by_code_all)
            if code:
                item["aparell_codi"] = code
            else:
                warnings.append(f"desempat[{idx}].aparell_id no exportat: {raw_legacy_app}")
            item.pop("aparell_id", None)

        scope = item.get("scope") or {}
        if isinstance(scope, dict):
            scope2 = dict(scope)
            apps = scope2.get("aparells") or {}
            if isinstance(apps, dict):
                ids_scope = apps.get("ids") or []
                ids_scope_out = []
                seen_scope = set()
                for raw in ids_scope if isinstance(ids_scope, list) else []:
                    code = _resolve_app_code_for_template(raw, by_id_all, by_code_all)
                    if not code:
                        warnings.append(f"desempat[{idx}].scope.aparells.ids no exportat: {raw}")
                        continue
                    if code in seen_scope:
                        continue
                    seen_scope.add(code)
                    ids_scope_out.append(code)
                apps2 = dict(apps)
                apps2["ids"] = ids_scope_out
                scope2["aparells"] = apps2
            item["scope"] = scope2

        item["exercicis_per_aparell"] = map_keys_to_codes(
            item.get("exercicis_per_aparell") or {},
            f"desempat[{idx}].exercicis_per_aparell",
        )

        des_out.append(item)
    schema["desempat"] = des_out

    presentacio = schema.get("presentacio") or {}
    if not isinstance(presentacio, dict):
        presentacio = {}
    cols = presentacio.get("columnes") or []
    if not isinstance(cols, list):
        cols = []
    cols_out = []
    for idx, col in enumerate(cols):
        if not isinstance(col, dict):
            cols_out.append(col)
            continue
        item = dict(col)
        ctype = str(item.get("type") or "builtin").strip().lower()
        if ctype == "raw":
            src = item.get("source") if isinstance(item.get("source"), dict) else {}
            src2 = dict(src)
            code = _resolve_app_code_for_template(src2.get("aparell_codi"), by_id_all, by_code_all)
            if not code:
                code = _resolve_app_code_for_template(src2.get("aparell_id"), by_id_all, by_code_all)
            if code:
                src2["aparell_codi"] = code
            elif src2.get("aparell_id") not in (None, "", 0, "0"):
                warnings.append(f"presentacio.columnes[{idx}] raw: aparell no exportat")
            src2.pop("aparell_id", None)
            item.pop("aparell_id", None)
            item["source"] = src2
        elif ctype == "metric":
            code = _resolve_app_code_for_template(item.get("aparell_codi"), by_id_all, by_code_all)
            if not code:
                code = _resolve_app_code_for_template(item.get("aparell_id"), by_id_all, by_code_all)
            if code:
                item["aparell_codi"] = code
            item.pop("aparell_id", None)
        cols_out.append(item)
    presentacio["columnes"] = cols_out
    schema["presentacio"] = presentacio

    equips_cfg = schema.get("equips") or {}
    if isinstance(equips_cfg, dict):
        manual = equips_cfg.get("particions_manuals") or []
        if isinstance(manual, list):
            team_name_by_id = {
                int(e.id): str(e.nom or "").strip()
                for e in Equip.objects.filter(competicio=competicio).only("id", "nom")
            }
            manual_out = []
            for idx, item in enumerate(manual):
                if not isinstance(item, dict):
                    manual_out.append(item)
                    continue
                row = dict(item)
                names = []
                seen_names = set()
                for raw_id in (row.get("equip_ids") or []):
                    try:
                        eid = int(raw_id)
                    except Exception:
                        continue
                    name = str(team_name_by_id.get(eid) or "").strip()
                    if not name:
                        warnings.append(f"equips.particions_manuals[{idx}]: equip {eid} no exportat")
                        continue
                    key = name.casefold()
                    if key in seen_names:
                        continue
                    seen_names.add(key)
                    names.append(name)
                row.pop("equip_ids", None)
                row["equips_noms"] = names
                manual_out.append(row)
            equips_cfg["particions_manuals"] = manual_out
            schema["equips"] = equips_cfg

    return schema, warnings


def _template_schema_to_competicio_schema(competicio, schema_tpl):
    schema = _json_clone(schema_tpl or {})
    warnings = []

    apps_active, by_id_active, by_code_active = _get_comp_aparell_maps(competicio, active_only=True)
    mapping = {code: int(ca.id) for code, ca in by_code_active.items()}

    punt = schema.get("puntuacio") or {}
    if not isinstance(punt, dict):
        punt = {}
    schema["puntuacio"] = punt

    raw_apps = punt.get("aparells") or {}
    if not isinstance(raw_apps, dict):
        raw_apps = {}
    ids_in = raw_apps.get("ids") or []
    ids_out = []
    seen_ids = set()
    for raw in ids_in if isinstance(ids_in, list) else []:
        app_id = _resolve_app_id_for_competicio(raw, by_id_active, by_code_active)
        if not app_id:
            warnings.append(f"Aparell no disponible a la competicio: {raw}")
            continue
        if app_id in seen_ids:
            continue
        seen_ids.add(app_id)
        ids_out.append(app_id)
    punt["aparells"] = {"mode": "seleccionar", "ids": ids_out}

    def map_keys_to_ids(raw_map, field_label):
        out = {}
        if not isinstance(raw_map, dict):
            return out
        for raw_key, raw_value in raw_map.items():
            app_id = _resolve_app_id_for_competicio(raw_key, by_id_active, by_code_active)
            if not app_id:
                warnings.append(f"{field_label}: aparell no disponible ({raw_key})")
                continue
            out[str(app_id)] = raw_value
        return out

    punt["camps_per_aparell"] = map_keys_to_ids(punt.get("camps_per_aparell") or {}, "camps_per_aparell")
    punt["exercicis_per_aparell"] = map_keys_to_ids(
        punt.get("exercicis_per_aparell") or {},
        "exercicis_per_aparell",
    )

    desempat = schema.get("desempat") or []
    if not isinstance(desempat, list):
        desempat = []
    des_out = []
    for idx, tie in enumerate(desempat):
        if not isinstance(tie, dict):
            des_out.append(tie)
            continue
        item = dict(tie)

        raw_code = item.get("aparell_codi")
        if raw_code not in (None, "", 0, "0"):
            app_id = _resolve_app_id_for_competicio(raw_code, by_id_active, by_code_active)
            if app_id:
                item["aparell_id"] = app_id
            else:
                warnings.append(f"desempat[{idx}].aparell_codi no disponible: {raw_code}")
        item.pop("aparell_codi", None)

        scope = item.get("scope") or {}
        if isinstance(scope, dict):
            scope2 = dict(scope)
            apps = scope2.get("aparells") or {}
            if isinstance(apps, dict):
                ids_scope = apps.get("ids") or []
                ids_scope_out = []
                seen_scope = set()
                for raw in ids_scope if isinstance(ids_scope, list) else []:
                    app_id = _resolve_app_id_for_competicio(raw, by_id_active, by_code_active)
                    if not app_id:
                        warnings.append(f"desempat[{idx}].scope.aparells.ids no disponible: {raw}")
                        continue
                    if app_id in seen_scope:
                        continue
                    seen_scope.add(app_id)
                    ids_scope_out.append(app_id)
                apps2 = dict(apps)
                apps2["ids"] = ids_scope_out
                scope2["aparells"] = apps2
            item["scope"] = scope2

        item["exercicis_per_aparell"] = map_keys_to_ids(
            item.get("exercicis_per_aparell") or {},
            f"desempat[{idx}].exercicis_per_aparell",
        )

        des_out.append(item)
    schema["desempat"] = des_out

    presentacio = schema.get("presentacio") or {}
    if isinstance(presentacio, dict):
        cols = presentacio.get("columnes") or []
        if isinstance(cols, list):
            cols_out = []
            for idx, col in enumerate(cols):
                if not isinstance(col, dict):
                    cols_out.append(col)
                    continue
                item = dict(col)
                ctype = str(item.get("type") or "builtin").strip().lower()
                if ctype == "raw":
                    src = item.get("source") if isinstance(item.get("source"), dict) else {}
                    src2 = dict(src)
                    app_id = _resolve_app_id_for_competicio(src2.get("aparell_codi"), by_id_active, by_code_active)
                    if not app_id:
                        app_id = _resolve_app_id_for_competicio(src2.get("aparell_id"), by_id_active, by_code_active)
                    if app_id:
                        src2["aparell_id"] = app_id
                    elif src2.get("aparell_codi") not in (None, "", 0, "0") or src2.get("aparell_id") not in (None, "", 0, "0"):
                        warnings.append(f"presentacio.columnes[{idx}] raw: aparell no disponible")
                    src2.pop("aparell_codi", None)
                    item["source"] = src2
                    item.pop("aparell_codi", None)
                elif ctype == "metric":
                    app_id = _resolve_app_id_for_competicio(item.get("aparell_codi"), by_id_active, by_code_active)
                    if not app_id:
                        app_id = _resolve_app_id_for_competicio(item.get("aparell_id"), by_id_active, by_code_active)
                    if app_id:
                        item["aparell_id"] = app_id
                    item.pop("aparell_codi", None)
                cols_out.append(item)
            presentacio["columnes"] = cols_out
            schema["presentacio"] = presentacio

    equips_cfg = schema.get("equips") or {}
    if isinstance(equips_cfg, dict):
        manual = equips_cfg.get("particions_manuals") or []
        if isinstance(manual, list):
            teams = list(Equip.objects.filter(competicio=competicio).only("id", "nom"))
            id_by_name = {}
            for t in teams:
                key = str(t.nom or "").strip().casefold()
                if key and key not in id_by_name:
                    id_by_name[key] = int(t.id)
            manual_out = []
            for idx, item in enumerate(manual):
                if not isinstance(item, dict):
                    manual_out.append(item)
                    continue
                row = dict(item)
                names = row.get("equips_noms") or []
                out_ids = []
                seen_eq = set()
                for raw_name in names if isinstance(names, list) else []:
                    key = str(raw_name or "").strip().casefold()
                    if not key:
                        continue
                    eid = id_by_name.get(key)
                    if not eid:
                        warnings.append(
                            f"equips.particions_manuals[{idx}]: equip '{raw_name}' no trobat a la competicio"
                        )
                        continue
                    if eid in seen_eq:
                        continue
                    seen_eq.add(eid)
                    out_ids.append(eid)
                row["equip_ids"] = out_ids
                manual_out.append(row)
            equips_cfg["particions_manuals"] = manual_out
            schema["equips"] = equips_cfg

    return schema, warnings, mapping


def _collect_required_app_codes_from_template(schema_tpl):
    schema = schema_tpl or {}
    out = set()

    punt = schema.get("puntuacio") or {}
    if isinstance(punt, dict):
        apps = punt.get("aparells") or {}
        if isinstance(apps, dict):
            for raw in (apps.get("ids") or []):
                code = _canon_app_code(raw)
                if code:
                    out.add(code)

        for raw_key in (punt.get("camps_per_aparell") or {}).keys() if isinstance(punt.get("camps_per_aparell"), dict) else []:
            code = _canon_app_code(raw_key)
            if code:
                out.add(code)

        for raw_key in (punt.get("exercicis_per_aparell") or {}).keys() if isinstance(punt.get("exercicis_per_aparell"), dict) else []:
            code = _canon_app_code(raw_key)
            if code:
                out.add(code)

    desempat = schema.get("desempat") or []
    if isinstance(desempat, list):
        for tie in desempat:
            if not isinstance(tie, dict):
                continue
            code = _canon_app_code(tie.get("aparell_codi"))
            if code:
                out.add(code)
            scope = tie.get("scope") or {}
            if isinstance(scope, dict):
                apps = scope.get("aparells") or {}
                if isinstance(apps, dict):
                    for raw in (apps.get("ids") or []):
                        code = _canon_app_code(raw)
                        if code:
                            out.add(code)
            raw_map = tie.get("exercicis_per_aparell") or {}
            if isinstance(raw_map, dict):
                for raw_key in raw_map.keys():
                    code = _canon_app_code(raw_key)
                    if code:
                        out.add(code)

    presentacio = schema.get("presentacio") or {}
    if isinstance(presentacio, dict):
        cols = presentacio.get("columnes") or []
        if isinstance(cols, list):
            for col in cols:
                if not isinstance(col, dict):
                    continue
                ctype = str(col.get("type") or "builtin").strip().lower()
                if ctype == "raw":
                    src = col.get("source") if isinstance(col.get("source"), dict) else {}
                    code = _canon_app_code(src.get("aparell_codi"))
                    if code:
                        out.add(code)
                elif ctype == "metric":
                    code = _canon_app_code(col.get("aparell_codi"))
                    if code:
                        out.add(code)

    return out


def _build_template_requirements(schema_tpl):
    schema = schema_tpl or {}
    punt = schema.get("puntuacio") or {}
    if not isinstance(punt, dict):
        punt = {}

    req = {
        "aparells_codis": sorted(_collect_required_app_codes_from_template(schema)),
        "particions": [str(x).strip() for x in (schema.get("particions") or []) if str(x).strip()],
        "camps_per_aparell": {},
        "desempat_camps": [],
    }

    camps_map = punt.get("camps_per_aparell") or {}
    if isinstance(camps_map, dict):
        for raw_key, raw_codes in camps_map.items():
            code = _canon_app_code(raw_key)
            if not code:
                continue
            vals = []
            if isinstance(raw_codes, list):
                vals = [str(x).strip() for x in raw_codes if str(x).strip()]
            elif isinstance(raw_codes, str):
                vals = [x.strip() for x in raw_codes.split(",") if x.strip()]
            req["camps_per_aparell"][code] = vals

    tie_codes = set()
    desempat = schema.get("desempat") or []
    if isinstance(desempat, list):
        for tie in desempat:
            if not isinstance(tie, dict):
                continue
            for code in _normalize_tie_camps_for_validation(tie):
                if code:
                    tie_codes.add(str(code).strip())
    req["desempat_camps"] = sorted([c for c in tie_codes if c])

    return req


def _parse_fallback_mode(raw) -> str:
    mode = str(raw or "strict").strip().lower()
    if mode not in {"strict", "assistit", "force"}:
        return "strict"
    return mode


def _next_fallback_mode(mode: str):
    mode = _parse_fallback_mode(mode)
    if mode == "strict":
        return "assistit"
    if mode == "assistit":
        return "force"
    return None


def _scoreable_codes_by_app_id(competicio):
    active_apps = list(
        CompeticioAparell.objects
        .filter(competicio=competicio, actiu=True)
        .select_related("aparell")
    )
    schemas_by_aparell = {
        s.aparell_id: (s.schema or {})
        for s in ScoringSchema.objects.filter(aparell_id__in=[ca.aparell_id for ca in active_apps]).only("aparell_id", "schema")
    }
    out = {}
    for ca in active_apps:
        sch = schemas_by_aparell.get(ca.aparell_id, {}) or {}
        meta = _build_scoreable_meta_for_schema(sch, strict_unknown=True)
        out[int(ca.id)] = {code for code, info in (meta or {}).items() if (info or {}).get("scoreable")}
    return out


def _validate_schema_for_competicio(competicio, schema_local):
    schema_local = _normalize_particions_schema(schema_local or {})
    errors = []
    errors.extend(_validate_particions_schema(competicio, schema_local))
    errors.extend(_validate_no_tots_mode(schema_local))
    errors.extend(_validate_camps_per_aparell(competicio, schema_local))
    errors.extend(_validate_tie_camps_per_aparell(competicio, schema_local))
    errors.extend(_validate_exercicis_selection(competicio, schema_local))
    errors.extend(_validate_tie_exercicis_selection(competicio, schema_local))
    return schema_local, errors


def _autofix_schema_for_competicio(competicio, schema_local, mode: str):
    mode = _parse_fallback_mode(mode)
    schema = _normalize_particions_schema(_json_clone(schema_local or {}))
    warnings = []
    dropped = []

    if mode == "strict":
        return schema, warnings, dropped

    active_apps, by_id, _ = _get_comp_aparell_maps(competicio, active_only=True)
    active_ids = [int(ca.id) for ca in active_apps]
    active_set = set(active_ids)
    scoreable_by_app = _scoreable_codes_by_app_id(competicio)

    punt = schema.get("puntuacio") or {}
    if not isinstance(punt, dict):
        punt = {}
    schema["puntuacio"] = punt

    apps_cfg = punt.get("aparells") or {}
    if not isinstance(apps_cfg, dict):
        apps_cfg = {}
    ids_raw = apps_cfg.get("ids") or []
    selected_ids = []
    seen = set()
    for raw in ids_raw if isinstance(ids_raw, list) else []:
        try:
            app_id = int(raw)
        except Exception:
            continue
        if app_id in active_set and app_id not in seen:
            seen.add(app_id)
            selected_ids.append(app_id)
    if mode == "force" and not selected_ids and active_ids:
        selected_ids = list(active_ids)
        warnings.append("FORCE: no hi havia aparells seleccionats; s'han seleccionat tots els actius.")
    apps_cfg["mode"] = "seleccionar"
    apps_cfg["ids"] = selected_ids
    punt["aparells"] = apps_cfg

    camps_in = punt.get("camps_per_aparell") or {}
    camps_out = {}
    for app_id in list(selected_ids):
        raw_codes = None
        if isinstance(camps_in, dict):
            raw_codes = camps_in.get(str(app_id))
            if raw_codes is None:
                raw_codes = camps_in.get(app_id)
        if isinstance(raw_codes, str):
            req_codes = [x.strip() for x in raw_codes.split(",") if x.strip()]
        elif isinstance(raw_codes, list):
            req_codes = [str(x).strip() for x in raw_codes if str(x).strip()]
        else:
            req_codes = []

        allowed = scoreable_by_app.get(app_id, {"total", "TOTAL"})
        kept = [code for code in req_codes if code in allowed]
        if not kept and mode == "assistit":
            dropped.append(f"puntuacio.aparells.ids: {app_id} (sense camps compatibles)")
            warnings.append(f"Assistit: s'ha descartat aparell {app_id} per manca de camps compatibles.")
            continue
        if not kept and mode == "force":
            kept = ["total"]
            warnings.append(f"FORCE: aparell {app_id} sense camps compatibles; s'ha aplicat camp 'total'.")
        camps_out[str(app_id)] = kept

    selected_ids_after = [app_id for app_id in selected_ids if str(app_id) in camps_out]
    if mode == "force" and not selected_ids_after and active_ids:
        for app_id in active_ids:
            camps_out[str(app_id)] = ["total"]
        selected_ids_after = list(active_ids)
        warnings.append("FORCE: no quedaven aparells vàlids; s'han activat tots amb camp 'total'.")

    apps_cfg["ids"] = selected_ids_after
    punt["camps_per_aparell"] = camps_out

    ex_per_app_in = punt.get("exercicis_per_aparell") or {}
    ex_per_app_out = {}
    if isinstance(ex_per_app_in, dict):
        for raw_key, cfg in ex_per_app_in.items():
            try:
                app_id = int(raw_key)
            except Exception:
                continue
            if app_id in selected_ids_after:
                ex_per_app_out[str(app_id)] = cfg
    punt["exercicis_per_aparell"] = ex_per_app_out

    allowed_particio_codes = {
        str(item.get("code") or "").strip()
        for item in get_allowed_group_fields(competicio)
        if str(item.get("code") or "").strip()
    }
    parts_in = _normalize_particio_codes(schema.get("particions") or [])
    parts_out = [code for code in parts_in if code in allowed_particio_codes]
    for code in parts_in:
        if code not in parts_out:
            dropped.append(f"particions: {code}")
            warnings.append(f"Assistit/FORCE: s'ha eliminat la particio no compatible '{code}'.")
    schema["particions"] = parts_out

    custom_in = _normalize_particions_custom(schema.get("particions_custom") or {})
    custom_out = {}
    for code, cfg in custom_in.items():
        if code in allowed_particio_codes and code in parts_out:
            custom_out[code] = cfg
        else:
            dropped.append(f"particions_custom: {code}")
            warnings.append(f"Assistit/FORCE: s'ha eliminat la configuracio custom de particio '{code}'.")
    schema["particions_custom"] = custom_out

    des_in = schema.get("desempat") or []
    des_out = []
    for idx, tie in enumerate(des_in if isinstance(des_in, list) else []):
        if not isinstance(tie, dict):
            continue
        item = _json_clone(tie)
        camps = _normalize_tie_camps_for_validation(item)
        if not camps:
            dropped.append(f"desempat[{idx}] (sense camps)")
            warnings.append(f"Assistit/FORCE: desempat[{idx}] eliminat per manca de camps.")
            continue

        scope = item.get("scope") or {}
        if not isinstance(scope, dict):
            scope = {}
        app_scope = scope.get("aparells") or {}
        if not isinstance(app_scope, dict):
            app_scope = {}
        app_mode = str(app_scope.get("mode") or "hereta").strip().lower()

        if app_mode == "seleccionar":
            target_ids = [x for x in _parse_positive_int_list(app_scope.get("ids")) if x in set(selected_ids_after)]
            app_scope["ids"] = target_ids
            scope["aparells"] = app_scope
            item["scope"] = scope
            if not target_ids:
                dropped.append(f"desempat[{idx}] (sense aparells d'abast)")
                warnings.append(f"Assistit/FORCE: desempat[{idx}] eliminat per manca d'aparells compatibles.")
                continue
        else:
            target_ids = list(selected_ids_after)
            if not target_ids:
                dropped.append(f"desempat[{idx}] (sense aparells heretats)")
                warnings.append(f"Assistit/FORCE: desempat[{idx}] eliminat per manca d'aparells heretats.")
                continue

        valid_camps = []
        for code in camps:
            is_valid_for_all = True
            for app_id in target_ids:
                allowed = scoreable_by_app.get(app_id, {"total", "TOTAL"})
                if code not in allowed:
                    is_valid_for_all = False
                    break
            if is_valid_for_all:
                valid_camps.append(code)
        if not valid_camps:
            dropped.append(f"desempat[{idx}] (camps incompatibles)")
            warnings.append(f"Assistit/FORCE: desempat[{idx}] eliminat per camps incompatibles.")
            continue
        item["camps"] = valid_camps
        item["camp"] = valid_camps[0]

        ex_map_in = item.get("exercicis_per_aparell") or {}
        ex_map_out = {}
        if isinstance(ex_map_in, dict):
            for raw_key, cfg in ex_map_in.items():
                try:
                    app_id = int(raw_key)
                except Exception:
                    continue
                if app_id in target_ids:
                    ex_map_out[str(app_id)] = cfg
        item["exercicis_per_aparell"] = ex_map_out
        des_out.append(item)
    schema["desempat"] = des_out

    presentacio = schema.get("presentacio") or {}
    if not isinstance(presentacio, dict):
        presentacio = {}
    cols_in = presentacio.get("columnes") or []
    cols_out = []
    for idx, col in enumerate(cols_in if isinstance(cols_in, list) else []):
        if not isinstance(col, dict):
            continue
        item = _json_clone(col)
        ctype = str(item.get("type") or "builtin").strip().lower()
        if ctype != "raw":
            cols_out.append(item)
            continue
        src = item.get("source") if isinstance(item.get("source"), dict) else {}
        try:
            app_id = int(src.get("aparell_id"))
        except Exception:
            app_id = None
        camp = str(src.get("camp") or "total").strip() or "total"
        if not app_id or app_id not in set(selected_ids_after):
            dropped.append(f"presentacio.columnes[{idx}] raw (aparell invalid)")
            warnings.append(f"Assistit/FORCE: s'ha eliminat columna raw {idx + 1} per aparell no compatible.")
            continue
        allowed = scoreable_by_app.get(app_id, {"total", "TOTAL"})
        if camp not in allowed:
            dropped.append(f"presentacio.columnes[{idx}] raw (camp invalid)")
            warnings.append(f"Assistit/FORCE: s'ha eliminat columna raw {idx + 1} per camp no compatible.")
            continue
        cols_out.append(item)

    if not cols_out:
        cols_out = _json_clone((DEFAULT_SCHEMA.get("presentacio") or {}).get("columnes") or [])
        warnings.append("Assistit/FORCE: no quedaven columnes; s'han aplicat columnes per defecte.")
    presentacio["columnes"] = cols_out
    schema["presentacio"] = presentacio

    equips_cfg = schema.get("equips") or {}
    if isinstance(equips_cfg, dict):
        manual_in = equips_cfg.get("particions_manuals") or []
        manual_out = []
        for idx, item in enumerate(manual_in if isinstance(manual_in, list) else []):
            if not isinstance(item, dict):
                continue
            ids = []
            for raw_id in (item.get("equip_ids") or []):
                try:
                    eid = int(raw_id)
                except Exception:
                    continue
                if Equip.objects.filter(competicio=competicio, id=eid).exists():
                    ids.append(eid)
            if ids:
                row = dict(item)
                row["equip_ids"] = ids
                manual_out.append(row)
            elif mode == "force":
                warnings.append(f"FORCE: s'ha eliminat particio manual d'equips {idx + 1} sense equips mapejats.")
        equips_cfg["particions_manuals"] = manual_out
        schema["equips"] = equips_cfg

    return schema, warnings, dropped


def _build_force_minimal_schema(competicio, schema_local):
    schema = _normalize_particions_schema(_json_clone(schema_local or {}))
    active_ids = list(
        CompeticioAparell.objects
        .filter(competicio=competicio, actiu=True)
        .values_list("id", flat=True)
    )
    punt = schema.get("puntuacio") or {}
    if not isinstance(punt, dict):
        punt = {}
    punt["aparells"] = {"mode": "seleccionar", "ids": active_ids}
    punt["camps_per_aparell"] = {str(app_id): ["total"] for app_id in active_ids}
    punt["mode_seleccio_exercicis"] = "per_aparell_global"
    punt["exercicis_per_aparell"] = {}
    punt["agregacio_camps"] = "sum"
    punt["agregacio_exercicis"] = "sum"
    punt["agregacio_aparells"] = "sum"
    punt["ordre"] = "desc"
    punt["camp"] = "total"
    punt["agregacio"] = "sum"
    punt["best_n"] = 1
    schema["puntuacio"] = punt

    schema["particions"] = []
    schema["particions_custom"] = {}
    schema["desempat"] = []

    presentacio = schema.get("presentacio") or {}
    if not isinstance(presentacio, dict):
        presentacio = {}
    if not isinstance(presentacio.get("columnes"), list) or not presentacio.get("columnes"):
        presentacio["columnes"] = _json_clone((DEFAULT_SCHEMA.get("presentacio") or {}).get("columnes") or [])
    presentacio["mostrar_empats"] = bool(presentacio.get("mostrar_empats", True))
    presentacio["top_n"] = int(presentacio.get("top_n") or 0)
    schema["presentacio"] = presentacio
    return schema


def _validate_template_for_competicio(competicio, template_obj, fallback_mode="strict"):
    fallback_mode = _parse_fallback_mode(fallback_mode)
    schema_tpl = _extract_template_schema(getattr(template_obj, "payload", {}) or {})
    schema_local, mapping_warnings, mapping = _template_schema_to_competicio_schema(competicio, schema_tpl)

    required_codes = _collect_required_app_codes_from_template(schema_tpl)
    _, _by_id_active, by_code_active = _get_comp_aparell_maps(competicio, active_only=True)

    blocking = []
    warnings = list(mapping_warnings or [])
    dropped = []
    for code in sorted(required_codes):
        if code and code not in by_code_active:
            msg = f"Aparell requerit per la plantilla no disponible a la competicio: {code}"
            if fallback_mode == "strict":
                blocking.append(msg)
            else:
                warnings.append(msg)

    schema_local, strict_errors = _validate_schema_for_competicio(competicio, schema_local)
    blocking.extend(strict_errors)

    if fallback_mode in {"assistit", "force"}:
        schema_local, autofix_warnings, autofix_dropped = _autofix_schema_for_competicio(
            competicio,
            schema_local,
            mode=fallback_mode,
        )
        warnings.extend(autofix_warnings)
        dropped.extend(autofix_dropped)
        schema_local, blocking = _validate_schema_for_competicio(competicio, schema_local)

    if fallback_mode == "force" and blocking:
        forced_schema = _build_force_minimal_schema(competicio, schema_local)
        forced_schema, force_errors = _validate_schema_for_competicio(competicio, forced_schema)
        if len(force_errors) <= len(blocking):
            schema_local = forced_schema
            dropped.append("FORCE: aplicat schema minim per garantir compatibilitat.")
            warnings.append("FORCE: s'ha aplicat un schema minim (particions/desempat simplificats).")
            blocking = force_errors

    next_mode = _next_fallback_mode(fallback_mode) if blocking else None

    return {
        "compatible": not blocking,
        "blocking_errors": blocking,
        "warnings": warnings,
        "dropped_rules": dropped,
        "mapping": mapping,
        "resolved_schema": schema_local,
        "phase": fallback_mode,
        "next_fallback": next_mode,
        "can_try_next": bool(next_mode and blocking),
    }


def _template_to_payload_row(obj):
    payload = getattr(obj, "payload", {}) or {}
    schema = _extract_template_schema(payload)
    req = getattr(obj, "requirements", {}) or {}
    if not isinstance(req, dict) or not req:
        req = _build_template_requirements(schema)
    return {
        "id": obj.id,
        "nom": obj.nom,
        "slug": obj.slug,
        "descripcio": obj.descripcio or "",
        "tipus": obj.tipus,
        "activa": bool(obj.activa),
        "version": int(obj.version or 1),
        "uses_count": int(obj.uses_count or 0),
        "requirements": req,
        "updated_at": obj.updated_at.isoformat() if getattr(obj, "updated_at", None) else None,
    }


def _next_template_slug(base_name: str, exclude_template_id=None):
    base = slugify(base_name or "") or "classificacio-template"
    slug = base
    idx = 2
    qs = ClassificacioTemplateGlobal.objects.all()
    if exclude_template_id:
        qs = qs.exclude(id=exclude_template_id)
    existing = set(qs.values_list("slug", flat=True))
    while slug in existing:
        slug = f"{base}-{idx}"
        idx += 1
    return slug


def _next_cfg_ordre_for_competicio(competicio):
    last = (
        ClassificacioConfig.objects
        .filter(competicio=competicio)
        .order_by("-ordre", "-id")
        .values_list("ordre", flat=True)
        .first()
    )
    try:
        return int(last or 0) + 1
    except Exception:
        return 1


def _validate_particions_schema(competicio, schema: dict):
    schema = schema or {}
    errors = []

    allowed_fields = get_allowed_group_fields(competicio)
    allowed_codes = {str(f.get("code") or "").strip() for f in allowed_fields if str(f.get("code") or "").strip()}

    parts = _normalize_particio_codes(schema.get("particions") or [])
    for code in parts:
        if code not in allowed_codes:
            errors.append(f"particions: camp no permès per aquesta competicio: '{code}'")

    custom_map = _normalize_particions_custom(schema.get("particions_custom") or {})
    for code, cfg in custom_map.items():
        if code not in allowed_codes:
            errors.append(f"particions_custom['{code}']: camp no permès per aquesta competicio.")
            continue
        if code not in parts:
            errors.append(f"particions_custom['{code}']: cal incloure el camp a particions.")

        mode = str(cfg.get("mode") or "raw").strip().lower()
        if mode not in {"raw", "custom"}:
            errors.append(f"particions_custom['{code}'].mode invalid: {mode}")
        if mode != "custom":
            continue

        values_owner = {}
        groups = cfg.get("grups") or []
        if not isinstance(groups, list):
            errors.append(f"particions_custom['{code}'].grups ha de ser una llista.")
            continue
        for gidx, group in enumerate(groups):
            if not isinstance(group, dict):
                errors.append(f"particions_custom['{code}'].grups[{gidx}] ha de ser un objecte.")
                continue
            values = _split_particio_custom_values(group.get("values"))
            for val in values:
                key = " ".join(str(val).split()).casefold()
                owner = values_owner.get(key)
                if owner is not None:
                    errors.append(
                        f"particions_custom['{code}']: valor repetit entre grups ({val}) a indexos {owner} i {gidx}."
                    )
                    continue
                values_owner[key] = gidx

    return errors


def _field_is_direct_scoreable(field_cfg: dict):
    """
    Regla robusta per camps d'input:
    - number -> ok
    - list judge -> només si 1 jutge (equivalent 1x1)
    - matrix judge_x_* -> només 1 jutge i 1 item (1x1)
    """
    if not isinstance(field_cfg, dict):
        return False, "config de camp no valida"

    ftype = str(field_cfg.get("type") or "").strip().lower()
    shape = str(field_cfg.get("shape") or "").strip().lower()

    judges_cfg = field_cfg.get("judges") if isinstance(field_cfg.get("judges"), dict) else {}
    items_cfg = field_cfg.get("items") if isinstance(field_cfg.get("items"), dict) else {}

    n_judges = _safe_int(judges_cfg.get("count") or field_cfg.get("judges_count") or 1, 1)
    n_judges = max(1, min(10, n_judges))
    n_items = _safe_int(items_cfg.get("count") or 0, 0)
    n_items = max(0, min(50, n_items))

    if ftype == "number":
        return True, ""

    if ftype == "list" and shape == "judge":
        if n_judges == 1:
            return True, ""
        return False, "camp tipus llista amb mes d'un jutge"

    if ftype == "matrix" and shape in ("judge_x_item", "judge_x_element"):
        if n_judges == 1 and n_items == 1:
            return True, ""
        return False, "camp tipus matriu; per puntuacio directa nomes s'admet 1x1"

    return False, "tipus de camp no puntuable directament"


def _infer_schema_code_shapes(schema_obj: dict):
    schema_obj = schema_obj or {}

    params = schema_obj.get("params", {})
    if params is None or not isinstance(params, dict):
        params = {}

    fields = schema_obj.get("fields", [])
    computed = schema_obj.get("computed", [])
    if not isinstance(fields, list):
        fields = []
    if not isinstance(computed, list):
        computed = []

    field_codes = []
    comp_codes = []
    for f in fields:
        if isinstance(f, dict) and isinstance(f.get("code"), str):
            field_codes.append(f["code"])
    for c in computed:
        if isinstance(c, dict) and isinstance(c.get("code"), str):
            comp_codes.append(c["code"])

    aliases = _build_alias_map(fields, computed, params)
    allowed_names = set(field_codes) | set(comp_codes) | set(aliases.keys()) | RESERVED_NAMES | ALLOWED_FUNCTIONS

    comp_deps = {cc: set() for cc in comp_codes}
    for i, c in enumerate(computed):
        if not isinstance(c, dict):
            continue
        code = str(c.get("code") or "").strip()
        if not code:
            continue
        formula = c.get("formula")
        if not isinstance(formula, str) or not formula.strip():
            continue

        loc = f"computed[{i}]({code})"
        try:
            tree = _ast_parse(formula, loc)
            names = _extract_names(tree)
        except Exception:
            continue

        # best-effort: ignorem desconeguts i extraiem deps conegudes
        resolved = {_resolve_name(n, aliases) for n in names if n in allowed_names}
        for r in resolved:
            if r in comp_deps and r != code:
                comp_deps[code].add(r)

    try:
        order = _topo_sort(comp_codes, comp_deps)
    except Exception:
        order = list(comp_codes)

    ctx = {}
    for f in fields:
        if not isinstance(f, dict):
            continue
        code = str(f.get("code") or "").strip()
        if not code:
            continue
        ctx[code] = TMat(_field_shape(f), name=code)

    ctx["params"] = TMat(Shape(1, 1), name="params")
    for short, code in aliases.items():
        if code in ctx:
            ctx[short] = ctx[code]

    for code in order:
        cobj = next((x for x in computed if isinstance(x, dict) and x.get("code") == code), None)
        if not cobj:
            continue
        formula = str(cobj.get("formula") or "")
        if not formula:
            continue
        loc = f"computed({code})"
        try:
            tree = _ast_parse(formula, loc)
            val = DryRunEval(ctx).visit(tree)
        except Exception:
            continue
        ctx[code] = val
        var = cobj.get("var")
        if isinstance(var, str) and var in aliases:
            ctx[var] = val

    out = {}
    for code in field_codes + comp_codes:
        tm = ctx.get(code)
        if not isinstance(tm, TMat):
            continue
        out[code] = {"rows": tm.shape.rows, "cols": tm.shape.cols}
    return out


def _shape_desc(shape_info) -> str:
    if not isinstance(shape_info, dict):
        return "?"
    r = shape_info.get("rows")
    c = shape_info.get("cols")
    r_txt = "?" if r is None else str(r)
    c_txt = "?" if c is None else str(c)
    return f"{r_txt}x{c_txt}"


def _is_scalar_shape_info(shape_info) -> bool:
    if not isinstance(shape_info, dict):
        return False
    return shape_info.get("rows") == 1 and shape_info.get("cols") == 1


def _normalize_tie_camps_for_validation(tie_obj) -> list:
    if not isinstance(tie_obj, dict):
        return []
    out = []
    raw = tie_obj.get("camps")
    if isinstance(raw, list):
        out = [str(x).strip() for x in raw if str(x).strip()]
    elif isinstance(raw, str):
        out = [x.strip() for x in raw.split(",") if x and x.strip()]

    if not out:
        legacy = str(tie_obj.get("camp") or "").strip()
        if legacy:
            out = [legacy]

    dedup = []
    seen = set()
    for c in out:
        if c in seen:
            continue
        seen.add(c)
        dedup.append(c)
    return dedup


def _computed_mode_hint(comp_cfg: dict) -> str:
    if not isinstance(comp_cfg, dict):
        return ""
    builder = comp_cfg.get("builder") if isinstance(comp_cfg.get("builder"), dict) else {}
    preset = str(builder.get("preset") or "").strip().lower()
    if preset in {"row_compute", "column_compute", "row_agregation", "exec_trampoli", "select_sum_guided"}:
        return preset

    formula = str(comp_cfg.get("formula") or "").strip().lower()
    if re.match(r"^row_custom_compute\s*\(", formula):
        return "row_compute"
    if re.match(r"^column_custom_compute\s*\(", formula):
        return "column_compute"
    if re.match(r"^row_custom_agregation\s*\(", formula):
        return "row_agregation"
    if re.match(r"^exec_by_judge\s*\(", formula):
        return "exec_trampoli"
    if re.match(r"^select_sum\s*\(", formula):
        return "select_sum_guided"
    return ""


def _formula_forces_vector_return(formula: str) -> bool:
    call_name, args_txt = _parse_formula_root_call(formula)
    fn = str(call_name or "").strip().lower()
    if fn not in {"row_custom_compute", "column_custom_compute"}:
        return False
    rm = _extract_kwarg_str(args_txt, "return_mode").lower()
    return rm in {"by_judge", "by_item"}


def _parse_formula_root_call(formula: str):
    txt = str(formula or "").strip()
    if not txt:
        return "", ""
    m = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\s*\((.*)\)\s*$", txt, flags=re.DOTALL)
    if not m:
        return "", ""
    return str(m.group(1) or "").strip(), str(m.group(2) or "")


def _extract_kwarg_int(args_txt: str, key: str):
    txt = str(args_txt or "")
    pat = rf"\b{re.escape(str(key))}\s*=\s*(-?\d+)\b"
    m = re.search(pat, txt)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _extract_kwarg_str(args_txt: str, key: str):
    txt = str(args_txt or "")
    pat = rf"\b{re.escape(str(key))}\s*=\s*(?:(['\"])(.*?)\1|([A-Za-z_][A-Za-z0-9_]*))"
    m = re.search(pat, txt, flags=re.DOTALL)
    if not m:
        return ""
    return str(m.group(2) or m.group(3) or "").strip()


def _extract_first_arg(args_txt: str):
    txt = str(args_txt or "")
    if not txt.strip():
        return ""

    depth = 0
    quote = ""
    escaped = False
    for i, ch in enumerate(txt):
        if quote:
            if escaped:
                escaped = False
                continue
            if ch == "\\":
                escaped = True
                continue
            if ch == quote:
                quote = ""
            continue

        if ch in {"'", '"'}:
            quote = ch
            continue
        if ch in {"(", "[", "{"}:
            depth += 1
            continue
        if ch in {")", "]", "}"}:
            depth = max(0, depth - 1)
            continue
        if ch == "," and depth == 0:
            return txt[:i].strip()
    return txt.strip()


def _extract_source_from_call(comp_cfg: dict, call_name: str, args_txt: str):
    if isinstance(comp_cfg, dict):
        b = comp_cfg.get("builder")
        if isinstance(b, dict):
            src = str(b.get("source") or "").strip()
            if src:
                return src

    fn = str(call_name or "").strip().lower()
    if fn not in {
        "row_custom_compute", "column_custom_compute", "row_custom_agregation",
        "exec_by_judge", "items_reduce", "crash",
    }:
        return ""

    tok = _extract_first_arg(args_txt)
    if not tok:
        return ""

    m = re.match(r"""^\s*field\s*\(\s*(?:'([^']+)'|"([^"]+)"|([A-Za-z_][A-Za-z0-9_]*))\s*\)\s*$""", tok)
    if m:
        return str(m.group(1) or m.group(2) or m.group(3) or "").strip()

    m = re.match(r"""^\s*(?:'([^']+)'|"([^"]+)"|([A-Za-z_][A-Za-z0-9_]*))\s*$""", tok)
    if not m:
        return ""
    return str(m.group(1) or m.group(2) or m.group(3) or "").strip()


def _extract_best_n_value(args_txt: str):
    n_kw = _extract_kwarg_int(args_txt, "n")
    if n_kw is not None:
        return n_kw
    txt = str(args_txt or "")
    # best_n(scores, 1)
    m = re.match(r"""^\s*[^,]+,\s*(-?\d+)\b""", txt)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _schema_field_dims_map(schema_obj: dict):
    out = {}
    schema_obj = schema_obj or {}
    for f in (schema_obj.get("fields") or []):
        if not isinstance(f, dict) or not f.get("code"):
            continue
        code = str(f.get("code"))
        sh = _field_shape(f)
        out[code] = {"rows": sh.rows, "cols": sh.cols}
    return out


def _resolve_source_dims(source_code: str, inferred_shapes: dict, field_dims: dict):
    src = str(source_code or "").strip()
    if not src:
        return {"rows": None, "cols": None}
    info = inferred_shapes.get(src)
    if isinstance(info, dict):
        return {"rows": info.get("rows"), "cols": info.get("cols")}
    info = field_dims.get(src)
    if isinstance(info, dict):
        return {"rows": info.get("rows"), "cols": info.get("cols")}
    return {"rows": None, "cols": None}


def _scoreable_from_conditional_vector(comp_cfg: dict, mode_hint: str, call_name: str, args_txt: str, inferred_shapes: dict, field_dims: dict, strict_unknown: bool):
    mode = str(mode_hint or "").strip().lower()
    fn = str(call_name or "").strip().lower()
    args = str(args_txt or "")

    source_code = _extract_source_from_call(comp_cfg, fn, args)
    dims = _resolve_source_dims(source_code, inferred_shapes, field_dims)
    src_rows = dims.get("rows")
    src_cols = dims.get("cols")

    # row/column compute only become vectors with explicit return_mode
    if mode in {"row_compute", "column_compute"} or fn in {"row_custom_compute", "column_custom_compute"}:
        is_row_compute = mode == "row_compute" or fn == "row_custom_compute"
        is_col_compute = mode == "column_compute" or fn == "column_custom_compute"
        rm = _extract_kwarg_str(args, "return_mode").lower()

        if not _formula_forces_vector_return(comp_cfg.get("formula")) or rm in {"", "final"}:
            return True, "", True

        if is_row_compute:
            if rm != "by_judge":
                return True, "", True
            if src_rows == 1:
                return True, "return_mode by_judge amb 1 jutge (llista d'1)", True
            if src_rows is None:
                if strict_unknown:
                    return False, "no es pot inferir longitud de vector by_judge", True
                return True, "longitud by_judge no inferible (UI tolerant)", True
            return False, f"return_mode by_judge amb {src_rows} jutges", True

        if is_col_compute:
            if rm != "by_item":
                return True, "", True
            # by_item: use cols and/or explicit count=1
            cnt = _extract_kwarg_int(args, "count")
            if cnt is not None and cnt == 1:
                return True, "return_mode by_item amb count=1", True
            if src_cols == 1:
                return True, "return_mode by_item amb 1 item", True
            if src_cols is None:
                if strict_unknown:
                    return False, "no es pot inferir longitud de vector by_item", True
                return True, "longitud by_item no inferible (UI tolerant)", True
            return False, f"return_mode by_item amb {src_cols} items", True

    # Vector by judge families
    if mode in {"row_agregation", "exec_trampoli"} or fn in {"row_custom_agregation", "exec_by_judge", "items_reduce", "crash"}:
        if src_rows == 1:
            return True, "vector per jutge de longitud 1", True
        if src_rows is None:
            if strict_unknown:
                return False, "no es pot inferir longitud de vector per jutge", True
            return True, "longitud per jutge no inferible (UI tolerant)", True
        return False, f"vector per jutge de longitud {src_rows}", True

    # best_n -> vector de mida n
    if fn == "best_n":
        n_val = _extract_best_n_value(args)
        if n_val == 1:
            return True, "best_n amb n=1", True
        if n_val is None:
            if strict_unknown:
                return False, "best_n sense n constant no es pot garantir mida 1", True
            return True, "best_n sense n constant (UI tolerant)", True
        return False, f"best_n amb n={n_val}", True

    return False, "", False


def _build_scoreable_meta_for_schema(schema_obj: dict, strict_unknown=False):
    schema_obj = schema_obj or {}
    meta = {
        "total": {"scoreable": True, "reason": ""},
        "TOTAL": {"scoreable": True, "reason": ""},
    }

    inferred_shapes = {}
    infer_error = ""
    field_dims = _schema_field_dims_map(schema_obj)
    try:
        inferred_shapes = _infer_schema_code_shapes(schema_obj)
    except Exception as exc:
        infer_error = str(exc)

    for f in (schema_obj.get("fields") or []):
        if not isinstance(f, dict) or not f.get("code"):
            continue
        code = str(f["code"])
        ok, reason = _field_is_direct_scoreable(f)
        shape_info = inferred_shapes.get(code)
        if shape_info is not None:
            if _is_scalar_shape_info(shape_info):
                ok, reason = True, ""
            else:
                ok, reason = False, f"shape no escalar {_shape_desc(shape_info)}"
        meta[code] = {"scoreable": bool(ok), "reason": str(reason or "")}

    for c in (schema_obj.get("computed") or []):
        if not isinstance(c, dict) or not c.get("code"):
            continue
        code = str(c["code"])
        mode_hint = _computed_mode_hint(c)
        formula_txt = str(c.get("formula") or "")
        call_name, call_args = _parse_formula_root_call(formula_txt)

        cond_ok, cond_reason, cond_handled = _scoreable_from_conditional_vector(
            c,
            mode_hint,
            call_name,
            call_args,
            inferred_shapes,
            field_dims,
            strict_unknown=bool(strict_unknown),
        )
        if cond_handled:
            meta[code] = {"scoreable": bool(cond_ok), "reason": "" if cond_ok else cond_reason}
            continue

        shape_info = inferred_shapes.get(code)
        if shape_info is None:
            if mode_hint in {"row_compute", "column_compute", "select_sum_guided"}:
                meta[code] = {"scoreable": True, "reason": ""}
                continue

            if strict_unknown:
                if infer_error:
                    reason = f"no es pot inferir shape (schema invalid): {infer_error}"
                else:
                    reason = "no es pot inferir shape del computed"
                meta[code] = {"scoreable": False, "reason": reason}
            else:
                meta[code] = {"scoreable": True, "reason": ""}
            continue
        if mode_hint in {"row_compute", "column_compute", "select_sum_guided"} and _is_scalar_shape_info(shape_info):
            meta[code] = {"scoreable": True, "reason": ""}
            continue
        if _is_scalar_shape_info(shape_info):
            meta[code] = {"scoreable": True, "reason": ""}
        else:
            meta[code] = {
                "scoreable": False,
                "reason": f"computed amb shape no escalar {_shape_desc(shape_info)}",
            }

    return meta


def _validate_camps_per_aparell(competicio, schema: dict):
    schema = schema or {}
    punt = (schema.get("puntuacio") or {})
    camps_per_aparell = punt.get("camps_per_aparell") or {}
    if not camps_per_aparell:
        return []
    if not isinstance(camps_per_aparell, dict):
        return ["puntuacio.camps_per_aparell ha de ser un objecte {app_id:[camps]}."]

    active_apps = list(
        CompeticioAparell.objects
        .filter(competicio=competicio, actiu=True)
        .select_related("aparell")
    )
    app_by_id = {ca.id: ca for ca in active_apps}
    if not app_by_id:
        return []

    aparell_ids = [ca.aparell_id for ca in active_apps]
    schemas_by_aparell = {
        s.aparell_id: (s.schema or {})
        for s in ScoringSchema.objects.filter(aparell_id__in=aparell_ids).only("aparell_id", "schema")
    }

    app_mode = ((punt.get("aparells") or {}).get("mode") or "seleccionar").strip().lower()
    app_ids_raw = ((punt.get("aparells") or {}).get("ids") or [])
    selected_app_ids = set()
    if app_mode == "seleccionar":
        for x in app_ids_raw:
            try:
                selected_app_ids.add(int(x))
            except Exception:
                continue
    else:
        selected_app_ids = set(app_by_id.keys())

    errors = []
    for app_key, raw_codes in camps_per_aparell.items():
        try:
            app_id = int(app_key)
        except Exception:
            errors.append(f"app_id invalid a camps_per_aparell: {app_key}")
            continue

        if app_id not in app_by_id:
            errors.append(f"aparell {app_id} no valid o no actiu a la competicio.")
            continue
        if app_id not in selected_app_ids:
            continue

        if isinstance(raw_codes, str):
            codes = [x.strip() for x in raw_codes.split(",") if x and x.strip()]
        elif isinstance(raw_codes, list):
            codes = [str(x).strip() for x in raw_codes if str(x).strip()]
        else:
            errors.append(f"camps_per_aparell[{app_id}] ha de ser llista o string.")
            continue

        sch = schemas_by_aparell.get(app_by_id[app_id].aparell_id, {}) or {}
        meta = _build_scoreable_meta_for_schema(sch, strict_unknown=True)

        for code in codes:
            info = meta.get(code)
            if not info:
                errors.append(f"aparell {app_id}: camp '{code}' no existeix al schema.")
                continue
            if not info.get("scoreable", False):
                errors.append(f"aparell {app_id}: camp '{code}' no es puntuable directament ({info.get('reason')}).")

    return errors


def _parse_positive_int_list(raw):
    if raw is None:
        return []
    if isinstance(raw, str):
        out = []
        for part in raw.split(","):
            p = str(part or "").strip()
            if not p:
                continue
            try:
                iv = int(p)
            except Exception:
                continue
            if iv > 0:
                out.append(iv)
        return out
    if isinstance(raw, (list, tuple)):
        out = []
        for x in raw:
            try:
                iv = int(x)
            except Exception:
                continue
            if iv > 0:
                out.append(iv)
        return out
    return []


def _validate_exercicis_cfg_obj(cfg, prefix: str):
    errors = []
    if not isinstance(cfg, dict):
        errors.append(f"{prefix} ha de ser un objecte.")
        return errors

    allowed_modes = {"tots", "millor_1", "millor_n", "pitjor_1", "pitjor_n", "primer", "ultim", "index", "llista"}
    mode = str(cfg.get("mode") or "tots").strip().lower()
    if mode not in allowed_modes:
        errors.append(f"{prefix}.mode invalid: {mode}")
        return errors

    if mode in {"millor_n", "pitjor_n"}:
        try:
            n = int(cfg.get("best_n") or 1)
        except Exception:
            n = 0
        if n < 1:
            errors.append(f"{prefix}.best_n ha de ser >= 1.")

    if mode == "index":
        try:
            idx = int(cfg.get("index") or 1)
        except Exception:
            idx = 0
        if idx < 1:
            errors.append(f"{prefix}.index ha de ser >= 1.")

    if mode == "llista":
        ids = _parse_positive_int_list(cfg.get("ids"))
        if not ids:
            errors.append(f"{prefix}.ids ha de contenir almenys un index valid (>0).")

    return errors


def _get_active_and_selected_app_ids(competicio, punt: dict):
    active_app_ids = set(
        CompeticioAparell.objects.filter(competicio=competicio, actiu=True).values_list("id", flat=True)
    )
    app_mode = str(((punt or {}).get("aparells") or {}).get("mode") or "seleccionar").strip().lower()
    app_ids_raw = ((punt or {}).get("aparells") or {}).get("ids") or []
    if app_mode != "seleccionar":
        return active_app_ids, set(active_app_ids)

    selected_ids = set()
    for x in app_ids_raw:
        try:
            selected_ids.add(int(x))
        except Exception:
            continue
    return active_app_ids, selected_ids


def _validate_exercicis_selection(competicio, schema: dict):
    schema = schema or {}
    punt = (schema.get("puntuacio") or {})

    mode_sel = str(punt.get("mode_seleccio_exercicis") or "per_aparell_global").strip().lower()
    allowed_sel = {"per_aparell_global", "per_aparell_override", "global_pool"}
    if mode_sel not in allowed_sel:
        return [f"puntuacio.mode_seleccio_exercicis invalid: {mode_sel}"]

    errors = []
    ex_global = punt.get("exercicis") or {}
    errors.extend(_validate_exercicis_cfg_obj(ex_global, "puntuacio.exercicis"))

    if mode_sel != "per_aparell_override":
        return errors

    raw_map = punt.get("exercicis_per_aparell") or {}
    if not isinstance(raw_map, dict):
        errors.append("puntuacio.exercicis_per_aparell ha de ser un objecte {app_id: cfg}.")
        return errors

    active_app_ids, selected_ids = _get_active_and_selected_app_ids(competicio, punt)

    for app_key, ex_cfg in raw_map.items():
        try:
            app_id = int(app_key)
        except Exception:
            errors.append(f"puntuacio.exercicis_per_aparell: app_id invalid {app_key}")
            continue
        if app_id not in active_app_ids:
            errors.append(f"puntuacio.exercicis_per_aparell: aparell {app_id} no valid o no actiu.")
            continue
        if app_id not in selected_ids:
            continue
        errors.extend(_validate_exercicis_cfg_obj(ex_cfg, f"puntuacio.exercicis_per_aparell[{app_id}]"))

    return errors


def _validate_no_tots_mode(schema: dict):
    schema = schema or {}
    errors = []

    punt = schema.get("puntuacio") or {}
    if not isinstance(punt, dict):
        punt = {}
    app_mode = str(((punt.get("aparells") or {}).get("mode") or "seleccionar")).strip().lower()
    if app_mode == "tots":
        errors.append("puntuacio.aparells.mode='tots' no esta permès; cal seleccionar aparells explicitament.")

    desempat = schema.get("desempat") or []
    if not isinstance(desempat, list):
        return errors

    for idx, tie in enumerate(desempat):
        if not isinstance(tie, dict):
            continue
        scope = tie.get("scope") or {}
        if not isinstance(scope, dict):
            continue
        app_scope = scope.get("aparells") or {}
        if not isinstance(app_scope, dict):
            continue
        tie_mode = str(app_scope.get("mode") or "hereta").strip().lower()
        if tie_mode == "tots":
            errors.append(
                f"desempat[{idx}].scope.aparells.mode='tots' no esta permès; usa 'hereta' o seleccio explicita."
            )

    return errors


def _validate_tie_camps_per_aparell(competicio, schema: dict):
    schema = schema or {}
    punt = (schema.get("puntuacio") or {})
    desempat = schema.get("desempat") or []
    if not isinstance(desempat, list):
        return ["desempat ha de ser una llista."]

    active_apps = list(
        CompeticioAparell.objects
        .filter(competicio=competicio, actiu=True)
        .select_related("aparell")
    )
    app_by_id = {ca.id: ca for ca in active_apps}
    if not app_by_id:
        return []

    schemas_by_aparell = {
        s.aparell_id: (s.schema or {})
        for s in ScoringSchema.objects.filter(aparell_id__in=[ca.aparell_id for ca in active_apps]).only("aparell_id", "schema")
    }

    _, selected_ids_main = _get_active_and_selected_app_ids(competicio, punt)
    active_app_ids = set(app_by_id.keys())
    meta_cache = {}
    errors = []

    for idx, tie in enumerate(desempat):
        if not isinstance(tie, dict):
            continue

        camps = _normalize_tie_camps_for_validation(tie)
        if not camps:
            continue

        scope = tie.get("scope") or {}
        if not isinstance(scope, dict):
            scope = {}
        app_scope = scope.get("aparells") or {}
        if not isinstance(app_scope, dict):
            app_scope = {}
        app_mode = str(app_scope.get("mode") or "hereta").strip().lower()

        if app_mode == "seleccionar":
            target_ids = set(_parse_positive_int_list(app_scope.get("ids")))
        elif app_mode == "tots":
            target_ids = set(active_app_ids)
        else:
            target_ids = set(selected_ids_main)

        for app_id in sorted(target_ids):
            if app_id not in active_app_ids:
                errors.append(f"desempat[{idx}]: aparell {app_id} no valid o no actiu.")
                continue

            if app_id not in meta_cache:
                sch = schemas_by_aparell.get(app_by_id[app_id].aparell_id, {}) or {}
                meta_cache[app_id] = _build_scoreable_meta_for_schema(sch, strict_unknown=True)
            meta = meta_cache[app_id]

            for code in camps:
                info = meta.get(code)
                if not info:
                    errors.append(f"desempat[{idx}]: aparell {app_id}: camp '{code}' no existeix al schema.")
                    continue
                if not info.get("scoreable", False):
                    errors.append(
                        f"desempat[{idx}]: aparell {app_id}: camp '{code}' no es puntuable directament "
                        f"({info.get('reason')})."
                    )

    return errors


def _validate_tie_exercicis_selection(competicio, schema: dict):
    schema = schema or {}
    punt = (schema.get("puntuacio") or {})
    desempat = schema.get("desempat") or []
    if not isinstance(desempat, list):
        return ["desempat ha de ser una llista."]

    errors = []
    active_app_ids, selected_ids_main = _get_active_and_selected_app_ids(competicio, punt)

    for idx, tie in enumerate(desempat):
        if not isinstance(tie, dict):
            errors.append(f"desempat[{idx}] ha de ser un objecte.")
            continue

        scope = tie.get("scope") or {}
        if not isinstance(scope, dict):
            errors.append(f"desempat[{idx}].scope ha de ser un objecte.")
            scope = {}
        ex_scope = scope.get("exercicis") or {}
        if scope.get("exercicis") is not None and not isinstance(scope.get("exercicis"), dict):
            errors.append(f"desempat[{idx}].scope.exercicis ha de ser un objecte.")
        if not isinstance(ex_scope, dict):
            ex_scope = {}

        ex_mode = str(ex_scope.get("mode") or "hereta").strip().lower()
        if ex_mode != "hereta":
            errors.extend(
                _validate_exercicis_cfg_obj(ex_scope, f"desempat[{idx}].scope.exercicis")
            )

        mode_sel_raw = (
            tie.get("mode_seleccio_exercicis")
            or ex_scope.get("mode_seleccio_exercicis")
            or "hereta"
        )
        mode_sel = str(mode_sel_raw).strip().lower()
        allowed_sel = {"hereta", "per_aparell_global", "per_aparell_override", "global_pool"}
        if mode_sel not in allowed_sel:
            errors.append(f"desempat[{idx}].mode_seleccio_exercicis invalid: {mode_sel}")
            continue
        if mode_sel != "per_aparell_override":
            continue

        app_scope = scope.get("aparells") or {}
        if not isinstance(app_scope, dict):
            app_scope = {}
        app_mode = str(app_scope.get("mode") or "hereta").strip().lower()

        if app_mode == "seleccionar":
            target_ids = set(_parse_positive_int_list(app_scope.get("ids")))
        elif app_mode == "tots":
            target_ids = set(active_app_ids)
        else:
            target_ids = set(selected_ids_main)

        raw_map = tie.get("exercicis_per_aparell")
        if raw_map is None:
            raw_map = ex_scope.get("exercicis_per_aparell")
        if raw_map is None:
            raw_map = {}
        if not isinstance(raw_map, dict):
            errors.append(f"desempat[{idx}].exercicis_per_aparell ha de ser un objecte {{app_id: cfg}}.")
            continue

        for app_key, ex_cfg in raw_map.items():
            try:
                app_id = int(app_key)
            except Exception:
                errors.append(f"desempat[{idx}].exercicis_per_aparell: app_id invalid {app_key}")
                continue
            if app_id not in active_app_ids:
                errors.append(
                    f"desempat[{idx}].exercicis_per_aparell: aparell {app_id} no valid o no actiu."
                )
                continue
            if target_ids and app_id not in target_ids:
                continue
            errors.extend(
                _validate_exercicis_cfg_obj(
                    ex_cfg,
                    f"desempat[{idx}].exercicis_per_aparell[{app_id}]",
                )
            )

    return errors


@require_POST
@transaction.atomic
def classificacio_save(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    cid = payload.get("id")
    nom = (payload.get("nom") or "Classificació").strip()
    activa = bool(payload.get("activa", True))
    ordre = int(payload.get("ordre") or 1)
    tipus = (payload.get("tipus") or "individual").strip()
    schema = _normalize_particions_schema(payload.get("schema") or {})

    if tipus not in ("individual", "entitat", "equips"):
        tipus = "individual"

    validation_errors = []
    validation_errors.extend(_validate_particions_schema(competicio, schema))
    validation_errors.extend(_validate_no_tots_mode(schema))
    validation_errors.extend(_validate_camps_per_aparell(competicio, schema))
    validation_errors.extend(_validate_tie_camps_per_aparell(competicio, schema))
    validation_errors.extend(_validate_exercicis_selection(competicio, schema))
    validation_errors.extend(_validate_tie_exercicis_selection(competicio, schema))
    if validation_errors:
        return JsonResponse(
            {
                "ok": False,
                "error": "Configuracio de classificacio invalida.",
                "errors": validation_errors,
            },
            status=400,
        )

    if cid:
        obj = get_object_or_404(ClassificacioConfig, pk=cid, competicio=competicio)
        obj.nom = nom
        obj.activa = activa
        obj.ordre = ordre
        obj.tipus = tipus
        obj.schema = schema
        obj.save()
    else:
        obj = ClassificacioConfig.objects.create(
            competicio=competicio,
            nom=nom,
            activa=activa,
            ordre=ordre,
            tipus=tipus,
            schema=schema,
        )

    return JsonResponse({"ok": True, "id": obj.id})


@require_POST
@transaction.atomic
def classificacio_delete(request, pk, cid):
    competicio = get_object_or_404(Competicio, pk=pk)
    obj = get_object_or_404(ClassificacioConfig, pk=cid, competicio=competicio)
    obj.delete()
    return JsonResponse({"ok": True})


@require_POST
@transaction.atomic
def classificacio_reorder(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    order = payload.get("order") or []
    # order = [id1, id2, ...]
    for idx, cid in enumerate(order, start=1):
        ClassificacioConfig.objects.filter(competicio=competicio, id=cid).update(ordre=idx)

    return JsonResponse({"ok": True})


@require_POST
def classificacio_preview(request, pk, cid):
    competicio = get_object_or_404(Competicio, pk=pk)
    cfg = get_object_or_404(ClassificacioConfig, pk=cid, competicio=competicio)

    data = compute_classificacio(competicio, cfg)

    # Retorna una estructura fàcil pel front:
    # [
    #   {"particio": "global", "rows":[...]}
    # ]
    out = []
    for k in sorted(data.keys()):
        out.append({"particio": k, "rows": data[k]})

    return JsonResponse({
        "ok": True,
        "columns": get_display_columns(cfg.schema or {}),
        "data": out,
    })


def classificacio_template_list(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    can_manage = bool(user_has_competicio_capability(request.user, competicio, "classificacions.edit"))
    include_inactive = str(request.GET.get("all") or "").strip().lower() in {"1", "true", "yes", "on"}

    qs = ClassificacioTemplateGlobal.objects.order_by("nom", "id")
    if not (can_manage and include_inactive):
        qs = qs.filter(activa=True)

    data = [_template_to_payload_row(t) for t in qs]
    return JsonResponse({"ok": True, "templates": data, "can_manage": can_manage})


@require_POST
@transaction.atomic
def classificacio_template_save(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invÃ lid")

    cfg_id = payload.get("cfg_id")
    if not cfg_id:
        return JsonResponse({"ok": False, "error": "Falta cfg_id."}, status=400)

    cfg = get_object_or_404(ClassificacioConfig, pk=cfg_id, competicio=competicio)

    template_id = payload.get("template_id")
    nom = str(payload.get("nom") or cfg.nom or "Plantilla classificacio").strip() or "Plantilla classificacio"
    descripcio = str(payload.get("descripcio") or "").strip()
    activa = bool(payload.get("activa", True))

    schema_tpl, export_warnings = _schema_to_template_schema(competicio, cfg.schema or {})
    requirements = _build_template_requirements(schema_tpl)
    payload_obj = {
        "schema": schema_tpl,
        "source": {
            "competicio_id": competicio.id,
            "cfg_id": cfg.id,
            "cfg_nom": cfg.nom,
            "exported_at": timezone.now().isoformat(),
        },
    }

    if template_id:
        tpl = get_object_or_404(ClassificacioTemplateGlobal, pk=template_id)
        requested_slug = str(payload.get("slug") or tpl.slug or "").strip()
        if requested_slug:
            slug_candidate = slugify(requested_slug)
            if not slug_candidate:
                slug_candidate = _next_template_slug(nom, exclude_template_id=tpl.id)
            exists = (
                ClassificacioTemplateGlobal.objects
                .exclude(id=tpl.id)
                .filter(slug=slug_candidate)
                .exists()
            )
            if exists:
                return JsonResponse({"ok": False, "error": "Ja existeix una plantilla amb aquest slug."}, status=400)
            tpl.slug = slug_candidate
        else:
            tpl.slug = _next_template_slug(nom, exclude_template_id=tpl.id)

        tpl.nom = nom
        tpl.descripcio = descripcio
        tpl.tipus = cfg.tipus or "individual"
        tpl.activa = activa
        tpl.payload = payload_obj
        tpl.requirements = requirements
        tpl.version = int(tpl.version or 1) + 1
        tpl.save()
    else:
        requested_slug = str(payload.get("slug") or "").strip()
        if requested_slug:
            slug_candidate = slugify(requested_slug)
            if not slug_candidate:
                slug_candidate = _next_template_slug(nom)
            if ClassificacioTemplateGlobal.objects.filter(slug=slug_candidate).exists():
                return JsonResponse({"ok": False, "error": "Ja existeix una plantilla amb aquest slug."}, status=400)
        else:
            slug_candidate = _next_template_slug(nom)

        tpl = ClassificacioTemplateGlobal.objects.create(
            nom=nom,
            slug=slug_candidate,
            descripcio=descripcio,
            tipus=cfg.tipus or "individual",
            activa=activa,
            payload=payload_obj,
            requirements=requirements,
            created_by=request.user if getattr(request.user, "is_authenticated", False) else None,
        )

    return JsonResponse(
        {
            "ok": True,
            "template": _template_to_payload_row(tpl),
            "warnings": export_warnings,
        }
    )


@require_POST
def classificacio_template_validate(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invÃ lid")

    template_id = payload.get("template_id")
    if not template_id:
        return JsonResponse({"ok": False, "error": "Falta template_id."}, status=400)
    fallback_mode = _parse_fallback_mode(payload.get("fallback_mode"))

    qs = ClassificacioTemplateGlobal.objects.filter(activa=True)
    tpl = get_object_or_404(qs, pk=template_id)

    result = _validate_template_for_competicio(competicio, tpl, fallback_mode=fallback_mode)
    return JsonResponse(
        {
            "ok": True,
            "template": _template_to_payload_row(tpl),
            **result,
        }
    )


@require_POST
@transaction.atomic
def classificacio_template_apply(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invÃ lid")

    template_id = payload.get("template_id")
    if not template_id:
        return JsonResponse({"ok": False, "error": "Falta template_id."}, status=400)
    fallback_mode = _parse_fallback_mode(payload.get("fallback_mode"))
    ack_warning = bool(payload.get("ack_warning"))
    if fallback_mode != "strict" and not ack_warning:
        return JsonResponse(
            {
                "ok": False,
                "error": "Cal confirmar l'avis per aplicar aquest mode de fallback.",
                "phase": fallback_mode,
            },
            status=400,
        )

    qs = ClassificacioTemplateGlobal.objects.filter(activa=True)
    tpl = get_object_or_404(qs, pk=template_id)

    validation = _validate_template_for_competicio(competicio, tpl, fallback_mode=fallback_mode)
    if not validation.get("compatible"):
        return JsonResponse(
            {
                "ok": False,
                "error": "La plantilla no Ã©s compatible amb la competicio actual.",
                **validation,
            },
            status=400,
        )

    schema = validation.get("resolved_schema") or {}
    tipus = str(getattr(tpl, "tipus", "individual") or "individual").strip().lower()
    if tipus not in ("individual", "entitat", "equips"):
        tipus = "individual"

    nom = str(payload.get("nom") or "").strip() or f"{tpl.nom} (tpl)"
    activa = bool(payload.get("activa", False))

    obj = ClassificacioConfig.objects.create(
        competicio=competicio,
        nom=nom,
        activa=activa,
        ordre=_next_cfg_ordre_for_competicio(competicio),
        tipus=tipus,
        schema=schema,
    )

    tpl.uses_count = int(tpl.uses_count or 0) + 1
    tpl.last_used_at = timezone.now()
    tpl.save(update_fields=["uses_count", "last_used_at", "updated_at"])

    return JsonResponse(
        {
            "ok": True,
            "cfg": {
                "id": obj.id,
                "nom": obj.nom,
                "activa": obj.activa,
                "ordre": obj.ordre,
                "tipus": obj.tipus,
                "schema": obj.schema or {},
            },
            "warnings": validation.get("warnings") or [],
            "template": _template_to_payload_row(tpl),
        }
    )


def _is_fk(model_cls, field_name: str) -> bool:
    try:
        f = model_cls._meta.get_field(field_name)
        return isinstance(f, (models.ForeignKey, models.OneToOneField))
    except Exception:
        return False


def _particio_value_to_text(raw) -> str:
    if raw is None:
        return ""
    if isinstance(raw, (list, dict)):
        try:
            return json.dumps(raw, ensure_ascii=False, sort_keys=True)
        except Exception:
            return str(raw)
    return str(raw).strip()


def _collect_particio_value_choices(ins_list, field_codes, max_per_field=200):
    out = {}
    for code in field_codes:
        unique_labels = {}
        unique_counts = {}
        for ins in ins_list:
            txt = _particio_value_to_text(get_inscripcio_value(ins, code))
            if not txt:
                continue
            key = " ".join(txt.split()).casefold()
            if key in unique_labels:
                unique_counts[key] = int(unique_counts.get(key, 0) or 0) + 1
                continue
            if len(unique_labels) >= max_per_field:
                continue
            unique_labels[key] = txt
            unique_counts[key] = 1

        entries = []
        for key, value in unique_labels.items():
            entries.append(
                {
                    "value": value,
                    "label": value,
                    "count": int(unique_counts.get(key, 0) or 0),
                }
            )

        entries.sort(key=lambda x: str(x.get("label") or "").casefold())
        out[code] = entries
    return out


def _distinct_values(qs, field_name: str):
    vals = qs.values_list(field_name, flat=True).distinct()
    out = []
    seen = set()
    for v in vals:
        if v is None:
            continue
        label = str(v).strip()
        if not label:
            continue
        key = " ".join(label.split()).casefold()   # treu dobles espais + case-insensitive
        if key in seen:
            continue
        seen.add(key)
        out.append(label)
    return out


def _distinct_fk(qs, field_name: str):
    # retorna llista d'objectes {id, label}
    rel = qs.select_related(field_name).values_list(f"{field_name}_id", f"{field_name}__nom").distinct()
    out = []
    seen = set()
    for _id, nom in rel:
        if _id is None or _id in seen:
            continue
        seen.add(_id)
        out.append({"value": _id, "label": nom or str(_id)})
    return out




