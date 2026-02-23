import random
from django.shortcuts import render
import math
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import CreateView, FormView, ListView, TemplateView, DeleteView
from django.shortcuts import redirect
from django.db.models import Q
from .forms import ImportInscripcionsExcelForm
from .models import Competicio, Inscripcio
from .forms import CompeticioForm
from .services.import_excel import importar_inscripcions_excel
import json
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from django.db import transaction
from django.views.generic import UpdateView
from django.urls import reverse
from .forms import InscripcioForm
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import Q, Max, Min, Case, When, IntegerField
from collections import defaultdict
from collections import OrderedDict


ALLOWED_SORT_FIELDS = {
    "nom": "nom_i_cognoms",
    "dni": "document",
    "sexe": "sexe",
    "entitat": "entitat",
    "categoria": "categoria",
    "subcategoria": "subcategoria",
    "grup": "grup",
}

# Camps built-in del model que volem permetre agrupar sempre
BUILTIN_GROUP_FIELDS = [
    {"code": "categoria", "label": "Categoria", "kind": "builtin"},
    {"code": "subcategoria", "label": "Subcategoria", "kind": "builtin"},
    {"code": "entitat", "label": "Entitat", "kind": "builtin"},
    {"code": "sexe", "label": "Sexe", "kind": "builtin"},
    {"code": "data_naixement", "label": "Data naixement", "kind": "builtin"},
    {"code": "document", "label": "Document", "kind": "builtin"},
]

# ✅ NOU: Camps built-in del model + UI que volem permetre mostrar com a columnes de taula
# (Això és independent dels camps per agrupar)
BUILTIN_TABLE_FIELDS = [
    {"code": "nom_i_cognoms", "label": "Nom i cognoms", "kind": "builtin"},
    {"code": "document", "label": "DNI/Document", "kind": "builtin"},
    {"code": "sexe", "label": "Sexe", "kind": "builtin"},
    {"code": "data_naixement", "label": "Data naixement", "kind": "builtin"},
    {"code": "entitat", "label": "Entitat", "kind": "builtin"},
    {"code": "categoria", "label": "Categoria", "kind": "builtin"},
    {"code": "subcategoria", "label": "Subcategoria", "kind": "builtin"},
    {"code": "grup", "label": "Grup", "kind": "builtin"},
    {"code": "ordre_sortida", "label": "Ordre", "kind": "builtin"},
]

def get_available_table_columns(competicio):
    """
    Retorna llista [{code,label,kind}, ...] = builtins + extras detectats a schema + accions.
    """
    out = []
    seen = set()

    for f in BUILTIN_TABLE_FIELDS:
        if f["code"] not in seen:
            out.append(f)
            seen.add(f["code"])

    schema = competicio.inscripcions_schema or {}
    cols = schema.get("columns") or []
    if isinstance(cols, list):
        for c in cols:
            if not isinstance(c, dict):
                continue
            code = c.get("code")
            if not code or code in seen:
                continue
            out.append({"code": code, "label": c.get("label") or code, "kind": c.get("kind") or "extra"})
            seen.add(code)

    # Columna UI "Accions"
    out.append({"code": "__actions__", "label": "Accions", "kind": "ui"})
    return out

def get_selected_table_columns(competicio, available_cols):
    """
    Llista de columnes seleccionades (i en ordre) a partir de competicio.inscripcions_view["table_columns"].
    """
    view = competicio.inscripcions_view or {}
    selected_codes = view.get("table_columns")
    if not isinstance(selected_codes, list) or not selected_codes:
        # Default raonable
        selected_codes = [
            "nom_i_cognoms", "document", "sexe", "data_naixement", "entitat",
            "categoria", "subcategoria", "grup", "ordre_sortida", "__actions__"
        ]

    by_code = {c["code"]: c for c in available_cols}
    selected = [by_code[c] for c in selected_codes if c in by_code]
    if not selected and "nom_i_cognoms" in by_code:
        selected = [by_code["nom_i_cognoms"]]
    return selected


# Per mostrar/ordenar values
def _s(v):
    return "(Sense valor)" if v in (None, "") else str(v)

def _norm_val(v):
    return "__NULL__" if v in (None, "") else str(v)

def get_inscripcio_value(obj, code: str):
    """
    Retorna valor per a un camp d'agrupació.
    - built-in: getattr
    - extra: obj.extra.get(code)
    """
    if hasattr(obj, code):
        return getattr(obj, code)
    # extra
    try:
        return (obj.extra or {}).get(code)
    except Exception:
        return None


def get_allowed_group_fields(competicio):
    """
    Retorna llista [{code,label,kind}, ...] = builtins + extras detectats a schema.
    """
    out = []
    seen = set()

    # builtins sempre
    for f in BUILTIN_GROUP_FIELDS:
        if f["code"] not in seen:
            out.append(f)
            seen.add(f["code"])

    schema = competicio.inscripcions_schema or {}
    cols = schema.get("columns") or []
    if isinstance(cols, list):
        for c in cols:
            if not isinstance(c, dict):
                continue
            code = c.get("code")
            if not code or code in seen:
                continue
            out.append({"code": code, "label": c.get("label") or code, "kind": c.get("kind") or "extra"})
            seen.add(code)

    return out


# --- UNDO (sense canvis) ---
UNDO_SESSION_KEY = "inscripcions_undo"

def save_undo_state(request, ids):
    request.session[UNDO_SESSION_KEY] = ids
    request.session.modified = True

def restore_undo_state(request):
    ids = request.session.get(UNDO_SESSION_KEY)
    if not ids:
        return None
    request.session.pop(UNDO_SESSION_KEY, None)
    request.session.modified = True
    return ids


def _recompute_ordre_sortida(competicio_id: int):
    qs = Inscripcio.objects.filter(competicio_id=competicio_id).order_by("ordre_sortida", "id")
    records = list(qs)
    for idx, obj in enumerate(records, start=1):
        if obj.ordre_sortida != idx:
            Inscripcio.objects.filter(id=obj.id).update(ordre_sortida=idx)


class InscripcionsListView(ListView):
    template_name = "competicio/inscripcions_list.html"
    context_object_name = "records"
    paginate_by = 25

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return self.get_queryset_base_filtrada().order_by("ordre_sortida", "id")

    def get_queryset_base_filtrada(self):
        qs = Inscripcio.objects.filter(competicio=self.competicio)

        q = self.request.GET.get("q")
        categoria = self.request.GET.get("categoria")
        subcategoria = self.request.GET.get("subcategoria")
        entitat = self.request.GET.get("entitat")

        if subcategoria:
            qs = qs.filter(subcategoria__iexact=subcategoria)
        if entitat:
            qs = qs.filter(entitat__icontains=entitat)
        if q:
            qs = qs.filter(
                Q(nom_i_cognoms__icontains=q) |
                Q(document__icontains=q) |
                Q(entitat__icontains=q)
            )
        if categoria:
            qs = qs.filter(categoria__iexact=categoria)

        return qs

    def get(self, request, *args, **kwargs):
        allowed = get_allowed_group_fields(self.competicio)
        allowed_codes = {f["code"] for f in allowed}

        def get_active_group_codes():
            selected = [g for g in request.GET.getlist("group_by") if g in allowed_codes]
            if not selected:
                selected = [g for g in (self.competicio.group_by_default or []) if g in allowed_codes]
            return selected

        def simple_key_for_obj(obj, group_codes):
            vals = [_norm_val(get_inscripcio_value(obj, c)) for c in group_codes]
            return json.dumps(vals, ensure_ascii=False)

        def pretty_label_from_simple_key(simple_key):
            try:
                vals = json.loads(simple_key)
                return " · ".join("(Sense valor)" if v in (None, "", "__NULL__") else str(v) for v in vals)
            except Exception:
                return simple_key

        # 0z) UNDO
        if request.GET.get("undo") == "1":
            restored = restore_undo_state(request)
            if restored:
                messages.success(request, f"S'ha desfet l'última acció ({restored} inscripcions).")
            else:
                messages.info(request, "No hi ha cap acció per desfer.")
            query = request.GET.copy()
            query.pop("undo", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        # 0x) EXPORT EXCEL
        if request.GET.get("export_excel") == "1":
            def fmt_date(d):
                return d.strftime("%d/%m/%Y") if d else "-"

            ALL_COLUMNS = OrderedDict([
                ("nom", ("Nom i cognoms", lambda o: o.nom_i_cognoms or "-")),
                ("dni", ("DNI", lambda o: o.document or "-")),
                ("sexe", ("Sexe", lambda o: o.sexe or "-")),
                ("naixement", ("Data naixement", lambda o: fmt_date(o.data_naixement))),
                ("entitat", ("Entitat", lambda o: o.entitat or "-")),
                ("categoria", ("Categoria", lambda o: o.categoria or "-")),
                ("subcategoria", ("Subcategoria", lambda o: o.subcategoria or "-")),
                ("grup", ("Grup", lambda o: o.grup if o.grup is not None else "-")),
                ("ordre", ("Ordre", lambda o: o.ordre_sortida if o.ordre_sortida is not None else "-")),
            ])

            selected_cols = request.GET.getlist("excel_cols")
            if not selected_cols:
                selected_cols = list(ALL_COLUMNS.keys())
            selected_cols = [c for c in selected_cols if c in ALL_COLUMNS]
            columns = [(ALL_COLUMNS[c][0], ALL_COLUMNS[c][1]) for c in selected_cols] or [
                (ALL_COLUMNS["nom"][0], ALL_COLUMNS["nom"][1])
            ]

            group_codes = get_active_group_codes()
            grouping_sig = "|".join(group_codes) if group_codes else ""
            qs_base = self.get_queryset_base_filtrada()

            wb = Workbook()
            ws = wb.active
            ws.title = "Inscripcions"

            title_font = Font(bold=True, size=12)
            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="E9EEF7")
            group_fill = PatternFill("solid", fgColor="DDE7FF")

            def write_table_header(r):
                for col_i, (label, _) in enumerate(columns, start=1):
                    c = ws.cell(row=r, column=col_i, value=label)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = Alignment(vertical="center")

            def write_row(r, obj):
                for col_i, (_, getter) in enumerate(columns, start=1):
                    ws.cell(row=r, column=col_i, value=getter(obj))

            qs_all = qs_base.annotate(
                grup_null=Case(
                    When(grup__isnull=True, then=1),
                    default=0,
                    output_field=IntegerField(),
                )
            ).order_by("grup_null", "grup", "ordre_sortida", "id")

            merges = (self.competicio.tab_merges or {}).get(grouping_sig, [])
            merge_map = {}
            merged_label_map = {}

            for group_keys in merges:
                t = tuple(group_keys)
                for x in group_keys:
                    merge_map[x] = t
                merged_key = json.dumps(list(t), ensure_ascii=False)
                merged_label_map[merged_key] = " + ".join(pretty_label_from_simple_key(x) for x in t)

            def tab_label_for_obj(obj):
                if not group_codes:
                    return "Sense agrupació"
                simple = simple_key_for_obj(obj, group_codes)
                mid = merge_map.get(simple)
                if mid:
                    mk = json.dumps(list(mid), ensure_ascii=False)
                    return merged_label_map.get(mk, mk)
                return pretty_label_from_simple_key(simple)

            row = 1
            ws.cell(row=row, column=1, value=f"Competicio: {self.competicio.nom}").font = title_font
            row += 2

            write_table_header(row)
            row += 1

            current_label = None
            for obj in qs_all:
                label = tab_label_for_obj(obj)
                if label != current_label:
                    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                    for c in range(1, len(columns) + 1):
                        ws.cell(row=row, column=c).fill = group_fill
                    row += 1
                    current_label = label

                write_row(row, obj)
                row += 1

            for i in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 20

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            resp = HttpResponse(
                bio.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = f'attachment; filename="inscripcions_{self.competicio.id}.xlsx"'
            return resp

        return super().get(request, *args, **kwargs)

    def get_paginate_by(self, queryset):
        if self.request.GET.get("per_page"):
            return None

        per_page = self.request.GET.get("per_page")
        if not per_page or per_page == "all":
            return None

        try:
            return int(per_page)
        except ValueError:
            return None

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = self.competicio

        allowed = get_allowed_group_fields(self.competicio)
        allowed_codes = {f["code"] for f in allowed}

        ctx["allowed_group_fields"] = allowed
        ctx["current_query"] = self.request.GET.urlencode()
        ctx["title_fields_selected"] = self.request.GET.getlist("title_fields")

        selected = self.request.GET.getlist("group_by")
        if not selected:
            selected = self.competicio.group_by_default or []
        selected = [g for g in selected if g in allowed_codes]
        ctx["selected_group_fields"] = selected

        records = self.get_queryset_base_filtrada().order_by("ordre_sortida", "id")
        grouping_sig = "|".join(selected) if selected else ""
        ctx["grouping_sig"] = grouping_sig

        def pretty_val(v):
            return "(Sense valor)" if v in (None, "") else str(v)

        if selected:
            grouped = OrderedDict()
            label_map = {}

            for r in records:
                vals = [_norm_val(get_inscripcio_value(r, code)) for code in selected]
                key = json.dumps(vals, ensure_ascii=False)
                grouped.setdefault(key, []).append(r)

                if key not in label_map:
                    parts = [pretty_val(get_inscripcio_value(r, code)) for code in selected]
                    label_map[key] = " · ".join(parts)

            merges = (self.competicio.tab_merges or {}).get(grouping_sig, [])
            merge_map = {}
            for group_keys in merges:
                if not group_keys:
                    continue
                primary = group_keys[0]
                for k in group_keys:
                    merge_map[k] = primary

            grouped_merged = OrderedDict()
            label_map_merged = {}

            for k, rows in grouped.items():
                pk = merge_map.get(k, k)
                grouped_merged.setdefault(pk, []).extend(rows)
                if pk not in label_map_merged:
                    label_map_merged[pk] = label_map.get(k, pk)

            records_grouped = [
                (label_map_merged.get(k, k), rows, k)
                for k, rows in grouped_merged.items()
            ]
            ctx["tabs"] = [
                {"key": group_key, "label": group_label, "count": len(group_records)}
                for (group_label, group_records, group_key) in records_grouped
            ]
            
            ctx["records_grouped"] = records_grouped
        else:
            ctx["records_grouped"] = None

        ALL_EXCEL_COLUMNS = [
            ("nom", "Nom i cognoms"),
            ("dni", "DNI"),
            ("sexe", "Sexe"),
            ("naixement", "Data naixement"),
            ("entitat", "Entitat"),
            ("categoria", "Categoria"),
            ("subcategoria", "Subcategoria"),
            ("grup", "Grup"),
            ("ordre", "Ordre"),
        ]
        ctx["allowed_excel_columns"] = ALL_EXCEL_COLUMNS

        sel_cols = self.request.GET.getlist("excel_cols")
        if not sel_cols:
            sel_cols = [k for k, _ in ALL_EXCEL_COLUMNS]
        ctx["excel_cols_selected"] = sel_cols

        base = self.get_queryset_base_filtrada()
        ctx["categories_distinct"] = list(base.order_by().values_list("categoria", flat=True).distinct())
        ctx["cats_selected"] = self.request.GET.getlist("cats")

        # ✅ NOU: Preferències de vista: columnes de taula + noms de grup
        available_table_columns = get_available_table_columns(self.competicio)
        selected_table_columns = get_selected_table_columns(self.competicio, available_table_columns)

        ctx["available_table_columns"] = available_table_columns
        ctx["selected_table_columns"] = selected_table_columns
        ctx["table_colspan"] = len(selected_table_columns) if selected_table_columns else 1

        view = self.competicio.inscripcions_view or {}
        group_names = view.get("group_names") or {}
        if isinstance(group_names, dict):
            ctx["group_names"] = {str(k): (v or "") for k, v in group_names.items()}
        else:
            ctx["group_names"] = {}

        return ctx

class InscripcioUpdateView(UpdateView):
    model = Inscripcio
    pk_url_kwarg = "ins_id"
    form_class = InscripcioForm
    template_name = "competicio/inscripcio_form.html"

    def get_queryset(self):
        return Inscripcio.objects.filter(competicio_id=self.kwargs["pk"])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = get_object_or_404(Competicio, pk=self.kwargs["pk"])
        ctx["next"] = self.request.GET.get("next", "")
        return ctx

    def get_success_url(self):
        nxt = self.request.GET.get("next")
        if nxt:
            return nxt
        return reverse("inscripcions_list", kwargs={"pk": self.kwargs["pk"]})


class InscripcioCreateView(CreateView):
    model = Inscripcio
    form_class = InscripcioForm
    template_name = "competicio/inscripcio_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw["competicio"] = self.competicio
        return kw

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = self.competicio
        ctx["next"] = self.request.GET.get("next", "")
        return ctx

    def get_success_url(self):
        nxt = self.request.GET.get("next")
        if nxt:
            return nxt
        return reverse("inscripcions_list", kwargs={"pk": self.kwargs["pk"]})


class InscripcioDeleteView(DeleteView):
    model = Inscripcio
    pk_url_kwarg = "ins_id"
    template_name = "competicio/inscripcio_confirm_delete.html"

    def get_queryset(self):
        return Inscripcio.objects.filter(competicio_id=self.kwargs["pk"])

    def get_success_url(self):
        nxt = self.request.GET.get("next")
        if nxt:
            return nxt
        return reverse("inscripcions_list", kwargs={"pk": self.kwargs["pk"]})


@require_POST
@csrf_protect
def inscripcions_reorder(request, pk):
    """
    Rep:
      {
        "ids": [<id1>, <id2>, ...],   # ordre final després del drag
        "moved_id": <id>,            # el registre arrossegat
        "new_index": <int>           # posició nova (0-based) dins ids
      }

    Guarda ordre_sortida = 1..N
    I (NOU) només pel registre mogut: adopta el grup del registre immediatament superior.
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
        ids = payload.get("ids", [])
        moved_id = payload.get("moved_id", None)
        new_index = payload.get("new_index", None)

        if not isinstance(ids, list) or not ids:
            return HttpResponseBadRequest("Payload invàlid")
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    wanted = [int(x) for x in ids if str(x).isdigit()]
    if not wanted:
        return HttpResponseBadRequest("IDs buits")

    if moved_id is not None:
        try:
            moved_id = int(moved_id)
        except Exception:
            return HttpResponseBadRequest("moved_id invàlid")

    if new_index is not None:
        try:
            new_index = int(new_index)
        except Exception:
            return HttpResponseBadRequest("new_index invàlid")

    qs = Inscripcio.objects.filter(competicio_id=pk, id__in=wanted)
    found = set(qs.values_list("id", flat=True))
    if set(wanted) != found:
        return HttpResponseBadRequest("IDs no vàlids per aquesta competició")

    id_to_group = dict(qs.values_list("id", "grup"))

    with transaction.atomic():
        for idx, ins_id in enumerate(wanted, start=1):
            Inscripcio.objects.filter(id=ins_id).update(ordre_sortida=idx)

        if moved_id is not None and new_index is not None and moved_id in wanted:
            if new_index > 0:
                prev_id = wanted[new_index - 1]
                prev_group = id_to_group.get(prev_id)
                Inscripcio.objects.filter(id=moved_id).update(grup=prev_group)

    return JsonResponse({"ok": True})


@require_POST
@csrf_protect
def inscripcions_merge_tabs(request, pk):
    c = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
        group_field = payload.get("group_field")   # IMPORTANT: ha de ser la "signature" (grouping_sig)
        source_key = payload.get("source_key")
        target_key = payload.get("target_key")
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    if not group_field:
        return HttpResponseBadRequest("group_field buit")
    if not source_key or not target_key or source_key == target_key:
        return HttpResponseBadRequest("claus invàlides")

    merges = c.tab_merges or {}
    lst = merges.get(group_field, [])

    def normalize_key_to_simple_list(k: str) -> list:
        """
        Retorna SEMPRE una llista de 'simple keys' (strings JSON del tipus ["A","B"...]).
        - Si k és simple: retorna [k]
        - Si k és merged: retorna [sub1, sub2, ...] on cada subX és un string JSON d'una clau simple
        """
        try:
            v = json.loads(k)
        except Exception:
            return [k]

        if (
            isinstance(v, list) and v
            and all(isinstance(x, str) for x in v)
            and all(x.strip().startswith("[") for x in v)
        ):
            return v

        if isinstance(v, list):
            return [k]

        return [k]

    s_list = normalize_key_to_simple_list(source_key)
    t_list = normalize_key_to_simple_list(target_key)

    desired = []
    for x in (t_list + s_list):
        if x not in desired:
            desired.append(x)

    consumed_idx = []
    merged_all = []
    for i, g in enumerate(lst):
        if any(x in g for x in desired):
            merged_all.extend(g)
            consumed_idx.append(i)

    for i in sorted(consumed_idx, reverse=True):
        lst.pop(i)

    merged_all.extend(desired)

    final = []
    for x in merged_all:
        if x not in final:
            final.append(x)

    lst.append(final)
    merges[group_field] = lst
    c.tab_merges = merges
    c.save(update_fields=["tab_merges"])

    return JsonResponse({"ok": True, "merged": final})


# ✅ NOU: desa columnes seleccionades + ordre de la taula
@require_POST
@csrf_protect
def inscripcions_save_table_columns(request, pk):
    """Desa columnes seleccionades + ordre per la taula d'inscripcions."""
    c = get_object_or_404(Competicio, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
        cols = payload.get("table_columns") or payload.get("columns") or []
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    if not isinstance(cols, list):
        return HttpResponseBadRequest("table_columns ha de ser una llista")

    available = get_available_table_columns(c)
    allowed_codes = {x["code"] for x in available}

    cleaned = []
    for x in cols:
        if not isinstance(x, str):
            continue
        if x in allowed_codes and x not in cleaned:
            cleaned.append(x)

    if not cleaned:
        return HttpResponseBadRequest("No hi ha cap columna vàlida")

    view = c.inscripcions_view or {}
    view["table_columns"] = cleaned
    c.inscripcions_view = view
    c.save(update_fields=["inscripcions_view"])
    return JsonResponse({"ok": True, "table_columns": cleaned})


# ✅ NOU: nom configurable per grup (Inscripcio.grup)
@require_POST
@csrf_protect
def inscripcions_set_group_name(request, pk):
    """Desa el nom d'un grup numèric (Inscripcio.grup)."""
    c = get_object_or_404(Competicio, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
        group = payload.get("group")
        name = (payload.get("name") or "").strip()
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    try:
        group_int = int(group)
    except Exception:
        return HttpResponseBadRequest("group invàlid")

    view = c.inscripcions_view or {}
    group_names = view.get("group_names")
    if not isinstance(group_names, dict):
        group_names = {}

    key = str(group_int)
    if name:
        group_names[key] = name
    else:
        group_names.pop(key, None)

    view["group_names"] = group_names
    c.inscripcions_view = view
    c.save(update_fields=["inscripcions_view"])
    return JsonResponse({"ok": True, "group": group_int, "name": name})