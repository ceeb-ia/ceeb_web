# competicio/services/import_excel.py
import hashlib
import json
import unicodedata
from datetime import datetime, date
from typing import Optional, Dict, Any, Set, Tuple, List

from openpyxl import load_workbook

from ..models import Inscripcio, Competicio


def _norm_header(s: str) -> str:
    """
    Normalitza capçaleres: minúscules, sense accents, separadors -> underscore.
    Ex: 'Data de naixement' -> 'data_de_naixement'
    """
    s = (s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    for ch in (" ", "-", "/", ".", ":", ";", ","):
        s = s.replace(ch, "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def _to_none(v):
    if v is None:
        return None
    if isinstance(v, str) and not v.strip():
        return None
    return v


def _clean_text(v) -> Optional[str]:
    """
    Neteja suau de text:
    - trim
    - col·lapsa espais múltiples
    - None si queda buit
    """
    v = _to_none(v)
    if v is None:
        return None
    txt = str(v).strip()
    txt = " ".join(txt.split())
    return txt or None


def _norm_text_key(v) -> Optional[str]:
    """
    Clau normalitzada per comparar valors textuals:
    - trim + col·lapsa espais
    - sense accents
    - case-insensitive (casefold)
    """
    txt = _clean_text(v)
    if txt is None:
        return None
    txt = "".join(
        c for c in unicodedata.normalize("NFKD", txt)
        if not unicodedata.combining(c)
    )
    txt = txt.casefold()
    return txt or None


def _normalize_document(v) -> Optional[str]:
    """
    Normalització de document per evitar variants trivials:
    - trim
    - uppercase
    - elimina espais i separadors comuns
    """
    txt = _clean_text(v)
    if txt is None:
        return None
    txt = txt.upper()
    for ch in (" ", "-", ".", "/"):
        txt = txt.replace(ch, "")
    return txt or None


def _parse_date(value) -> Optional[date]:
    """
    Accepta:
    - datetime/date d'Excel
    - string amb formats típics: dd/mm/yyyy o yyyy-mm-dd o dd-mm-yyyy
    """
    value = _to_none(value)
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, str):
        txt = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(txt, fmt).date()
            except ValueError:
                pass
    return None


# Camps coneguts (built-in) i sinònims habituals (normalitzats amb _norm_header)
# "code" = codi que farem servir a group_by i schema
BUILTIN = {
    "document": {
        "label": "Document (DNI/NIF/Passaport)",
        "syn": ["dni", "nif", "document", "document_identitat", "num_document", "numero_document", "id", "identificador"],
        "setter": lambda defaults, v: defaults.__setitem__("document", str(v).strip() if _to_none(v) is not None else None),
    },
    "nom_i_cognoms": {
        "label": "Nom i cognoms",
        "syn": ["nom_i_cognoms", "nom_cognoms", "participant", "nom_complet", "nomcomplert", "nom_i_llinatges", "nombre_y_apellidos","nombre_completo"],
        "setter": lambda defaults, v: defaults.__setitem__("nom_i_cognoms", str(v).strip() if _to_none(v) is not None else None),
    },
    # per si ve separat:
    "nom": {
        "label": "Nom",
        "syn": ["nom", "nombre", "name"],
        "setter": None,  # es tracta a banda
    },
    "cognoms": {
        "label": "Cognoms",
        "syn": ["cognoms", "apellidos", "surname", "llinatges"],
        "setter": None,  # es tracta a banda
    },
    "entitat": {
        "label": "Entitat/Club",
        "syn": ["entitat", "club", "organitzacio", "organizacion"],
        "setter": lambda defaults, v: defaults.__setitem__("entitat", str(v).strip() if _to_none(v) is not None else None),
    },
    "categoria": {
        "label": "Categoria",
        "syn": ["categoria", "category", "cat"],
        "setter": lambda defaults, v: defaults.__setitem__("categoria", str(v).strip() if _to_none(v) is not None else None),
    },
    "subcategoria": {
        "label": "Subcategoria",
        "syn": ["subcategoria", "sub_categoria", "subcat", "subcategory", "nivell", "nivel"],
        "setter": lambda defaults, v: defaults.__setitem__("subcategoria", str(v).strip() if _to_none(v) is not None else None),
    },
    "sexe": {
        "label": "Sexe",
        "syn": ["sexe", "sexo", "sex", "genere", "genero", "g"],
        "setter": lambda defaults, v: defaults.__setitem__("sexe", str(v).strip() if _to_none(v) is not None else None),
    },
    "data_naixement": {
        "label": "Data naixement",
        "syn": ["data_naixement", "data_de_naixement", "naixement", "fecha_nacimiento", "birthdate", "data_nasc"],
        "setter": lambda defaults, v: defaults.__setitem__("data_naixement", _parse_date(v)),
    },
}


def _build_syn_map(competicio: Competicio) -> Dict[str, List[str]]:
    """
    Combina sinònims per defecte + configuració guardada a competicio.inscripcions_schema["synonyms"].
    Els sinònims es normalitzen a _norm_header.
    """
    syn_map = {}
    for code, cfg in BUILTIN.items():
        syn_map[code] = [_norm_header(s) for s in cfg.get("syn", [])]

    schema = competicio.inscripcions_schema or {}
    extra_syn = schema.get("synonyms") or {}
    for code, syns in extra_syn.items():
        if not isinstance(syns, list):
            continue
        syn_map.setdefault(code, [])
        syn_map[code].extend(_norm_header(s) for s in syns)

    # dedup preservant ordre
    for code, syns in syn_map.items():
        seen = set()
        out = []
        for s in syns:
            if s and s not in seen:
                seen.add(s)
                out.append(s)
        syn_map[code] = out

    return syn_map


def _build_value_aliases(competicio: Competicio) -> Dict[str, Dict[str, str]]:
    """
    Llegeix aliases de valors des de:
      competicio.inscripcions_schema["value_aliases"] = {
        "categoria": {"prebenjami": "Prebenjamí", ...},
        ...
      }
    Retorna mapa normalitzat per clau de comparació.
    """
    schema = competicio.inscripcions_schema or {}
    raw = schema.get("value_aliases") or {}
    out: Dict[str, Dict[str, str]] = {}
    if not isinstance(raw, dict):
        return out

    for field, mapping in raw.items():
        if not isinstance(mapping, dict):
            continue
        out[field] = {}
        for raw_key, canonical in mapping.items():
            nk = _norm_text_key(raw_key)
            cv = _clean_text(canonical)
            if nk and cv:
                out[field][nk] = cv
    return out


def _build_existing_text_canon(
    competicio: Competicio,
    fields: Tuple[str, ...],
) -> Dict[str, Dict[str, str]]:
    """
    Construeix un mapa de valors canònics ja existents a BD:
      field -> norm_key -> valor guardat
    """
    out: Dict[str, Dict[str, str]] = {f: {} for f in fields}
    qs = Inscripcio.objects.filter(competicio=competicio)
    for field in fields:
        for val in qs.values_list(field, flat=True).distinct():
            cleaned = _clean_text(val)
            nk = _norm_text_key(cleaned)
            if not cleaned or not nk:
                continue
            out[field].setdefault(nk, cleaned)
    return out


def _canonicalize_text_field(
    field: str,
    value,
    aliases: Dict[str, Dict[str, str]],
    canon_map: Dict[str, Dict[str, str]],
) -> Optional[str]:
    """
    Normalitza i canonicalitza un camp textual:
    1) alias explícit de schema (value_aliases)
    2) valor canònic ja existent (BD/fitxer actual)
    3) valor net "tal qual"
    """
    cleaned = _clean_text(value)
    if cleaned is None:
        return None
    nk = _norm_text_key(cleaned)
    if nk is None:
        return None

    field_aliases = aliases.get(field, {})
    if nk in field_aliases:
        canonical = field_aliases[nk]
    else:
        canonical = canon_map.get(field, {}).get(nk, cleaned)

    canon_map.setdefault(field, {})[nk] = canonical
    return canonical


def _get_modalitat_extra(extra: Dict[str, Any]) -> Optional[str]:
    """
    Recupera la modalitat des d'extra amb variants habituals de clau.
    """
    if not isinstance(extra, dict):
        return None
    for key in ("modalitat", "modalidad"):
        v = _clean_text(extra.get(key))
        if v:
            return v
    return None


def _build_no_document_signature(defaults: Dict[str, Any], extra: Dict[str, Any]) -> Tuple[str, ...]:
    """
    Signatura estable per detectar coincidències quan no hi ha document.
    """
    data_naixement = defaults.get("data_naixement")
    data_key = data_naixement.isoformat() if isinstance(data_naixement, date) else "__NULL__"
    modalitat = _get_modalitat_extra(extra)

    def text_key(v) -> str:
        return _norm_text_key(v) or "__NULL__"

    return (
        text_key(defaults.get("nom_i_cognoms")),
        text_key(defaults.get("entitat")),
        text_key(defaults.get("categoria")),
        text_key(defaults.get("subcategoria")),
        text_key(defaults.get("sexe")),
        data_key,
        text_key(modalitat),
    )


def _normalize_for_dedupe(value):
    """
    Normalitza valors per generar una clau estable entre imports.
    """
    if value is None:
        return "__NULL__"
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return _norm_text_key(value) or "__NULL__"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "g")
    if isinstance(value, list):
        return [_normalize_for_dedupe(v) for v in value]
    if isinstance(value, dict):
        out = {}
        for k in sorted(value.keys(), key=lambda x: str(x)):
            out[str(k)] = _normalize_for_dedupe(value[k])
        return out
    return _norm_text_key(str(value)) or str(value)


def _build_dedupe_key(defaults: Dict[str, Any], extra: Dict[str, Any]) -> str:
    """
    Hash estable de la fila per fer imports idempotents quan no hi ha document.
    """
    data_naixement = defaults.get("data_naixement")
    payload = {
        "nom_i_cognoms": _norm_text_key(defaults.get("nom_i_cognoms")) or "__NULL__",
        "entitat": _norm_text_key(defaults.get("entitat")) or "__NULL__",
        "categoria": _norm_text_key(defaults.get("categoria")) or "__NULL__",
        "subcategoria": _norm_text_key(defaults.get("subcategoria")) or "__NULL__",
        "sexe": _norm_text_key(defaults.get("sexe")) or "__NULL__",
        "data_naixement": data_naixement.isoformat() if isinstance(data_naixement, date) else "__NULL__",
        "document": _normalize_document(defaults.get("document")) or "__NULL__",
        "modalitat": _norm_text_key(_get_modalitat_extra(extra)) or "__NULL__",
        "extra": _normalize_for_dedupe(extra or {}),
    }
    raw = json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _can_match_no_document(defaults: Dict[str, Any], extra: Dict[str, Any]) -> bool:
    """
    Evita comparar només pel nom per reduir falsos positius.
    """
    if not _norm_text_key(defaults.get("nom_i_cognoms")):
        return False

    data_naixement = defaults.get("data_naixement")
    has_context = any(
        [
            _norm_text_key(defaults.get("entitat")),
            _norm_text_key(defaults.get("categoria")),
            _norm_text_key(defaults.get("subcategoria")),
            _norm_text_key(defaults.get("sexe")),
            data_naixement.isoformat() if isinstance(data_naixement, date) else None,
            _norm_text_key(_get_modalitat_extra(extra)),
        ]
    )
    return bool(has_context)


def importar_inscripcions_excel(fitxer, competicio: Competicio, sheet: str = "") -> Dict[str, Any]:
    """
    Importa inscripcions des d'Excel de forma adaptable:
    - Detecta headers (fila 1)
    - Mapeja camps built-in amb sinònims (configurable)
    - Normalitza valors textuals per evitar duplicats trivials
      (majúscules/minúscules, accents, espais)
    - Qualsevol columna no mapejada -> Inscripcio.extra[code_columna]

    Duplicats:
    - Si hi ha document: update_or_create(competicio, document)
    - Si no: intenta match per signatura "humana" i, si no és possible,
      fa fallback a dedupe_key estable per evitar duplicats en reimports idèntics
    """
    wb = load_workbook(fitxer, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    # 1) headers: norm -> (label_original, col_idx)
    headers: Dict[str, Tuple[str, int]] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        raw = str(cell.value).strip()
        if not raw:
            continue
        headers[_norm_header(raw)] = (raw, col_idx)

    # 2) sinònims combinats
    syn_map = _build_syn_map(competicio)

    # 3) detecta quina columna correspon a cada builtin code
    detected_builtin_col: Dict[str, str] = {}  # code -> header_norm
    header_norms = set(headers.keys())

    # primer: si hi ha match exacte pel mateix "code"
    for code in BUILTIN.keys():
        if code in header_norms:
            detected_builtin_col[code] = code

    # després: sinònims
    for code, syns in syn_map.items():
        if code in detected_builtin_col:
            continue
        for s in syns:
            if s in header_norms:
                detected_builtin_col[code] = s
                break

    def cell_value(row: int, header_norm: str):
        meta = headers.get(header_norm)
        if not meta:
            return None
        return ws.cell(row=row, column=meta[1]).value

    # 4) construcció de schema columns (builtins + extras)
    # - builtins "útils" (els que existeixen realment al model Inscripcio)
    builtin_model_fields = {"nom_i_cognoms", "categoria", "subcategoria", "entitat", "document", "sexe", "data_naixement"}
    columns_schema: List[Dict[str, Any]] = []

    def add_col(code: str, label: str, kind: str):
        columns_schema.append({"code": code, "label": label, "kind": kind})

    # builtins detectats
    for code, header_norm in detected_builtin_col.items():
        if code not in builtin_model_fields:
            continue
        label = BUILTIN.get(code, {}).get("label") or headers.get(header_norm, (code, 0))[0]
        add_col(code, label, "builtin")

    # extras (tota columna que no sigui builtin detectada)
    used_header_norms = set(detected_builtin_col.values())
    for h_norm, (raw_label, _col) in headers.items():
        if h_norm in used_header_norms:
            continue
        # evita coses típiques que no vols guardar com extra (si vols, pots ampliar)
        if h_norm in ("nom_competicio", "nom_competició", "nom_competicio_"):
            continue
        add_col(h_norm, raw_label, "extra")

    # 5) merge schema a competicio (preservant el que ja hi havia)
    existing_schema = competicio.inscripcions_schema or {}
    existing_cols = existing_schema.get("columns") or []
    if not isinstance(existing_cols, list):
        existing_cols = []

    # index per code
    existing_by_code = {c.get("code"): c for c in existing_cols if isinstance(c, dict) and c.get("code")}
    for c in columns_schema:
        code = c["code"]
        if code in existing_by_code:
            # conserva label existent si existia (per no “ballar”)
            if existing_by_code[code].get("label"):
                c["label"] = existing_by_code[code]["label"]
        existing_by_code[code] = {**existing_by_code.get(code, {}), **c}

    merged_cols = list(existing_by_code.values())
    competicio.inscripcions_schema = {
        **existing_schema,
        "columns": merged_cols,
        "synonyms": existing_schema.get("synonyms", {}),
        "value_aliases": existing_schema.get("value_aliases", {}),
    }
    competicio.save(update_fields=["inscripcions_schema"])

    text_norm_fields = ("categoria", "subcategoria", "sexe", "entitat")
    value_aliases = _build_value_aliases(competicio)
    text_canon_map = _build_existing_text_canon(competicio, text_norm_fields)

    # Índex per matching quan no hi ha document (nom + context + modalitat)
    existing_no_doc_index: Dict[Tuple[str, ...], List[int]] = {}
    existing_no_doc_qs = Inscripcio.objects.filter(competicio=competicio).only(
        "id",
        "nom_i_cognoms",
        "entitat",
        "categoria",
        "subcategoria",
        "sexe",
        "data_naixement",
        "document",
        "extra",
    )
    for ins in existing_no_doc_qs:
        if _normalize_document(ins.document):
            continue
        defaults_existing = {
            "nom_i_cognoms": ins.nom_i_cognoms,
            "entitat": ins.entitat,
            "categoria": ins.categoria,
            "subcategoria": ins.subcategoria,
            "sexe": ins.sexe,
            "data_naixement": ins.data_naixement,
        }
        sig = _build_no_document_signature(defaults_existing, ins.extra or {})
        existing_no_doc_index.setdefault(sig, []).append(ins.id)

    # Índex tècnic per deduplicació estable entre imports
    existing_dedupe_index: Dict[str, List[int]] = {}
    existing_dedupe_qs = Inscripcio.objects.filter(competicio=competicio).only("id", "dedupe_key")
    for ins in existing_dedupe_qs:
        dk = _clean_text(getattr(ins, "dedupe_key", None))
        if not dk:
            continue
        existing_dedupe_index.setdefault(dk, []).append(ins.id)

    # 6) import rows
    creats = 0
    actualitzats = 0
    ignorats = 0
    ambiguos = 0
    errors = 0
    noms_competicio_excel: Set[str] = set()

    for r in range(2, ws.max_row + 1):
        try:
            # captura possible nom de competició en excel (si existeix)
            for possible in ("nom_competició", "nom_competicio", "nom competicio", "nom competició"):
                h = _norm_header(possible)
                if h in headers:
                    v = _to_none(cell_value(r, h))
                    if v:
                        noms_competicio_excel.add(str(v).strip())
                    break

            defaults: Dict[str, Any] = {
                "nom_i_cognoms": None,
                "entitat": None,
                "categoria": None,
                "subcategoria": None,
                "sexe": None,
                "data_naixement": None,
            }

            # 6.1 omple builtins (excepte nom/cognoms separats)
            for code, header_norm in detected_builtin_col.items():
                if code in ("nom", "cognoms"):
                    continue
                if code not in BUILTIN:
                    continue
                setter = BUILTIN[code].get("setter")
                if setter is None:
                    continue
                setter(defaults, cell_value(r, header_norm))

            # 6.2 si no hi ha nom_i_cognoms, intenta construir-ho amb nom + cognoms
            if not defaults.get("nom_i_cognoms"):
                nom = None
                cognoms = None
                if "nom" in detected_builtin_col:
                    nom = _to_none(cell_value(r, detected_builtin_col["nom"]))
                if "cognoms" in detected_builtin_col:
                    cognoms = _to_none(cell_value(r, detected_builtin_col["cognoms"]))

                if nom or cognoms:
                    full = f"{str(nom).strip() if nom else ''} {str(cognoms).strip() if cognoms else ''}".strip()
                    defaults["nom_i_cognoms"] = full if full else None

            # sense nom no importem
            if not defaults.get("nom_i_cognoms"):
                ignorats += 1
                continue

            # 6.3 extras
            extra: Dict[str, Any] = {}
            for h_norm in headers.keys():
                if h_norm in used_header_norms:
                    continue
                v = ws.cell(row=r, column=headers[h_norm][1]).value
                v = _to_none(v)
                if v is None:
                    continue
                # guarda strings netes
                if isinstance(v, str):
                    v = v.strip()
                extra[h_norm] = v

            # 6.4 document (si existeix)
            document = None
            if "document" in detected_builtin_col:
                dv = _to_none(cell_value(r, detected_builtin_col["document"]))
                document = _normalize_document(dv)

            # normalització/canonicalització de valors textuals
            for k in text_norm_fields:
                defaults[k] = _canonicalize_text_field(
                    field=k,
                    value=defaults.get(k),
                    aliases=value_aliases,
                    canon_map=text_canon_map,
                )
            defaults["nom_i_cognoms"] = _clean_text(defaults.get("nom_i_cognoms"))
            defaults["document"] = _normalize_document(defaults.get("document"))
            dedupe_key = _build_dedupe_key(defaults, extra)
            defaults["dedupe_key"] = dedupe_key

            if document:
                obj, created = Inscripcio.objects.update_or_create(
                    competicio=competicio,
                    document=document,
                    defaults={**defaults, "extra": extra},
                )
                if created:
                    creats += 1
                else:
                    actualitzats += 1
                dedupe_ids = existing_dedupe_index.setdefault(dedupe_key, [])
                if obj.id not in dedupe_ids:
                    dedupe_ids.append(obj.id)
            else:
                defaults["document"] = defaults.get("document") or ""   # assegura string, però sense duplicar kwargs
                if _can_match_no_document(defaults, extra):
                    sig = _build_no_document_signature(defaults, extra)
                    candidates = existing_no_doc_index.get(sig, [])
                else:
                    sig = None
                    candidates = []

                dedupe_candidates = existing_dedupe_index.get(dedupe_key, [])
                target_id = None

                if len(candidates) == 1:
                    target_id = candidates[0]
                elif len(candidates) > 1:
                    # Si hi ha ambigüitat "humana", només resolem amb dedupe tècnic unívoc
                    if len(dedupe_candidates) == 1:
                        target_id = dedupe_candidates[0]
                    else:
                        ambiguos += 1
                        continue
                else:
                    # Sense match "humà", usem dedupe tècnic per garantir idempotència
                    if len(dedupe_candidates) == 1:
                        target_id = dedupe_candidates[0]
                    elif len(dedupe_candidates) > 1:
                        ambiguos += 1
                        continue

                if target_id is not None:
                    updated = Inscripcio.objects.filter(
                        competicio=competicio,
                        id=target_id,
                    ).update(extra=extra, **defaults)
                    if updated:
                        actualitzats += 1
                        dedupe_ids = existing_dedupe_index.setdefault(dedupe_key, [])
                        if target_id not in dedupe_ids:
                            dedupe_ids.append(target_id)
                        if sig is not None:
                            sig_ids = existing_no_doc_index.setdefault(sig, [])
                            if target_id not in sig_ids:
                                sig_ids.append(target_id)
                    else:
                        obj = Inscripcio.objects.create(
                            competicio=competicio,
                            extra=extra,
                            **defaults,
                        )
                        creats += 1
                        if sig is not None:
                            existing_no_doc_index.setdefault(sig, []).append(obj.id)
                        existing_dedupe_index.setdefault(dedupe_key, []).append(obj.id)
                else:
                    obj = Inscripcio.objects.create(
                        competicio=competicio,
                        extra=extra,
                        **defaults,
                    )
                    creats += 1
                    if sig is not None:
                        existing_no_doc_index.setdefault(sig, []).append(obj.id)
                    existing_dedupe_index.setdefault(dedupe_key, []).append(obj.id)

        except Exception as e:
            errors += 1
            
            print(f"[IMPORT] fila {r} error: {e}")

    return {
        "full": ws.title,
        "creats": creats,
        "actualitzats": actualitzats,
        "ignorats": ignorats,
        "ambiguos": ambiguos,
        "errors": errors,
        "noms_competicio_excel": sorted(noms_competicio_excel),
    }
