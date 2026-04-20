import json
import re

from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .filters import normalize_team_mode
from .live import default_live_columns, extract_export_value, format_partition_title


def excel_safe_text(raw) -> str:
    txt = "" if raw is None else str(raw)
    if txt[:1] in ("=", "+", "-", "@"):
        return "'" + txt
    return txt


def try_float(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        txt = value.strip()
        if not txt:
            return None
        try:
            return float(txt)
        except Exception:
            return None
    return None


def format_scalar_text(value, decimals=None):
    if value is None:
        return ""
    number = try_float(value)
    if number is not None and decimals is not None:
        return f"{number:.{decimals}f}"
    if number is not None and isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def normalize_excel_cell(value, col):
    decimals = (col or {}).get("decimals")
    try:
        decimals = int(decimals) if decimals is not None else None
    except Exception:
        decimals = None
    if decimals is not None:
        decimals = max(0, min(6, decimals))

    if isinstance(value, dict) and value.get("_kind") == "judge_rows":
        lines = []
        for row in value.get("rows") or []:
            judge_num = row.get("judge")
            try:
                judge_num = int(judge_num)
            except Exception:
                judge_num = None
            judge_label = f"J{judge_num}" if judge_num and judge_num > 0 else "J?"
            items = row.get("items") or []
            if not isinstance(items, list):
                items = [items]
            vals = [format_scalar_text(item, decimals) for item in items]
            lines.append(f"{judge_label}: " + " | ".join(vals))
        return excel_safe_text("\n".join([line for line in lines if line])), None, True

    if isinstance(value, dict) and value.get("_kind") == "team_raw_detail":
        lines = []
        summary = value.get("summary")
        if summary not in (None, ""):
            lines.append(format_scalar_text(summary, decimals))
        for row in value.get("rows") or []:
            if not isinstance(row, dict):
                continue
            label = str(row.get("label") or "").strip()
            if not label:
                continue
            judge_rows = row.get("judge_rows")
            if isinstance(judge_rows, dict) and judge_rows.get("_kind") == "judge_rows":
                lines.append(f"{label}:")
                for judge_row in judge_rows.get("rows") or []:
                    judge_num = judge_row.get("judge")
                    try:
                        judge_num = int(judge_num)
                    except Exception:
                        judge_num = None
                    judge_label = f"J{judge_num}" if judge_num and judge_num > 0 else "J?"
                    items = judge_row.get("items") or []
                    if not isinstance(items, list):
                        items = [items]
                    vals = [format_scalar_text(item, decimals) for item in items]
                    lines.append(f"  {judge_label}: " + " | ".join(vals))
                continue
            lines.append(f"{label}: {format_scalar_text(row.get('value'), decimals)}")
        return excel_safe_text("\n".join([line for line in lines if line])), None, True

    if isinstance(value, (dict, list)):
        return excel_safe_text(json.dumps(value, ensure_ascii=False)), None, True
    if isinstance(value, bool):
        return value, None, False

    number = try_float(value)
    if number is not None and not isinstance(value, str):
        if decimals is None:
            if isinstance(value, int):
                return int(value), "0", False
            return float(value), "0.###", False
        rounded = round(number, decimals)
        num_fmt = "0" if decimals == 0 else ("0." + ("0" * decimals))
        return rounded, num_fmt, False

    if isinstance(value, str):
        txt = value.strip()
        if decimals is not None:
            number = try_float(txt)
            if number is not None:
                rounded = round(number, decimals)
                num_fmt = "0" if decimals == 0 else ("0." + ("0" * decimals))
                return rounded, num_fmt, False
        return excel_safe_text(value), None, ("\n" in value)

    if value is None:
        return "", None, False
    return excel_safe_text(value), None, False


def build_excel_sheet_name(raw_name, used_names):
    base = str(raw_name or "").strip() or "Classificacio"
    base = re.sub(r"[\[\]\*:/\\?]", " ", base)
    base = " ".join(base.split())
    if not base:
        base = "Classificacio"
    base = base[:31]

    candidate = base
    idx = 2
    while candidate.casefold() in used_names:
        suffix = f" ({idx})"
        max_len = 31 - len(suffix)
        truncated = (base[:max_len].rstrip() if max_len > 0 else "") or "Sheet"
        candidate = f"{truncated}{suffix}"
        idx += 1
    used_names.add(candidate.casefold())
    return candidate


def _copy_col(col):
    return dict(col or {})


def _detail_sections_for_row(row):
    detail = row.get("detail") if isinstance(row, dict) else None
    if not isinstance(detail, dict):
        return []
    sections = detail.get("sections")
    return sections if isinstance(sections, list) else []


def _detail_section_at(row, section_index, expected_type=""):
    sections = _detail_sections_for_row(row)
    if section_index < 0 or section_index >= len(sections):
        return None
    section = sections[section_index]
    if not isinstance(section, dict):
        return None
    if expected_type and str(section.get("type") or "").strip().lower() != expected_type:
        return None
    return section


def _detail_fallback_value(row, key: str):
    if key == "participant":
        return row.get("participant") or row.get("nom") or row.get("entitat_nom") or ""
    if key in {"nom", "aparell_nom", "entitat_nom"}:
        return row.get(key) or ""
    if key in {"exercise_index", "posicio", "participants"}:
        return row.get(key)
    if key == "punts":
        return row.get("punts")
    return row.get(key)


def _detail_cell_value(row, col):
    key = str((col or {}).get("key") or "")
    cells = row.get("cells") or row.get("display") or {}
    if isinstance(cells, dict) and key in cells:
        return cells.get(key)
    return _detail_fallback_value(row, key)


def _last_builtin_insert_index(cols):
    last_builtin_idx = None
    for idx, col in enumerate(cols):
        if str((col or {}).get("type") or "builtin").strip().lower() == "builtin":
            last_builtin_idx = idx
    return last_builtin_idx if last_builtin_idx is not None else len(cols)


def _collect_section_templates(parts):
    ordered = []
    for part in parts if isinstance(parts, list) else []:
        for row in (part or {}).get("rows") or []:
            for sidx, section in enumerate(_detail_sections_for_row(row)):
                if not isinstance(section, dict):
                    continue
                stype = str(section.get("type") or "").strip().lower()
                if not stype:
                    continue
                while len(ordered) <= sidx:
                    ordered.append(None)
                if ordered[sidx] is not None:
                    continue
                ordered[sidx] = {
                    "index": sidx,
                    "type": stype,
                    "label": str(section.get("label") or "").strip() or stype,
                    "columns": [_copy_col(col) for col in (section.get("columns") or []) if isinstance(col, dict)],
                }
    return [item for item in ordered if item]


def _exercise_group_label(app_label, exercise_index):
    base = str(app_label or "").strip()
    if base:
        return f"{base} - Ex.{exercise_index}"
    return f"Ex.{exercise_index}"


def _build_exercise_fixed_section_spec(template, parts):
    groups = []
    seen = set()
    for part in parts if isinstance(parts, list) else []:
        for row in (part or {}).get("rows") or []:
            section = _detail_section_at(row, template["index"], "exercise_table")
            if not isinstance(section, dict):
                continue
            for detail_row in section.get("rows") or []:
                if not isinstance(detail_row, dict):
                    continue
                key = (
                    detail_row.get("app_id"),
                    detail_row.get("exercise_index"),
                )
                if key in seen:
                    continue
                seen.add(key)
                groups.append(
                    {
                        "key": key,
                        "label": _exercise_group_label(
                            detail_row.get("aparell_nom"),
                            detail_row.get("exercise_index"),
                        ),
                    }
                )
    if not groups or not template.get("columns"):
        return None
    return {
        "section_index": template["index"],
        "type": "exercise_table",
        "label": template["label"],
        "columns": [_copy_col(col) for col in template.get("columns") or []],
        "groups": groups,
    }


def _build_direct_section_spec(template, *, participant_only=False):
    columns = [_copy_col(col) for col in template.get("columns") or []]
    if participant_only or not columns:
        columns = [
            {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
        ]
    return {
        "section_index": template["index"],
        "type": template["type"],
        "label": template["label"],
        "columns": columns,
    }


def _layout_kind_for_cfg(tipus="individual", schema=None):
    tipus_norm = str(tipus or "individual").strip().lower()
    equips_cfg = (schema or {}).get("equips") if isinstance(schema, dict) else {}
    team_mode = normalize_team_mode((equips_cfg or {}).get("team_mode"))
    if tipus_norm == "individual":
        return "individual_flat"
    if tipus_norm == "entitat":
        return "entity_block"
    if tipus_norm == "equips" and team_mode == "native_team":
        return "native_team_block"
    if tipus_norm == "equips":
        return "derived_team_block"
    return "plain"


def _build_layout(columns, parts, *, tipus="individual", schema=None):
    cols = columns if isinstance(columns, list) and columns else default_live_columns()
    insert_at = _last_builtin_insert_index(cols)
    leading = [_copy_col(col) for col in cols[:insert_at]]
    trailing = [_copy_col(col) for col in cols[insert_at:]]
    templates = _collect_section_templates(parts)
    layout_kind = _layout_kind_for_cfg(tipus=tipus, schema=schema)

    ordered_sections = []

    for template in templates:
        stype = template["type"]
        if layout_kind == "individual_flat":
            if stype == "exercise_table":
                section_spec = _build_exercise_fixed_section_spec(template, parts)
                if section_spec:
                    section_spec["render_mode"] = "fixed"
                    ordered_sections.append(section_spec)
        elif layout_kind == "entity_block":
            if stype == "entity_members_table":
                section_spec = _build_direct_section_spec(template)
                section_spec["render_mode"] = "variable"
                ordered_sections.append(section_spec)
        elif layout_kind == "derived_team_block":
            if stype == "members_table":
                section_spec = _build_direct_section_spec(template)
                section_spec["render_mode"] = "variable"
                ordered_sections.append(section_spec)
            elif stype == "members_list":
                section_spec = _build_direct_section_spec(template, participant_only=True)
                section_spec["render_mode"] = "variable"
                ordered_sections.append(section_spec)
        elif layout_kind == "native_team_block":
            if stype == "team_metrics":
                section_spec = _build_direct_section_spec(template)
                section_spec["render_mode"] = "fixed"
                ordered_sections.append(section_spec)
            elif stype == "team_members_table":
                section_spec = _build_direct_section_spec(template)
                section_spec["render_mode"] = "variable"
                ordered_sections.append(section_spec)
            elif stype == "members_list":
                section_spec = _build_direct_section_spec(template, participant_only=True)
                section_spec["render_mode"] = "variable"
                ordered_sections.append(section_spec)

    table_types = {"members_table", "entity_members_table", "team_members_table"}
    if any(
        section["render_mode"] == "variable" and section["type"] in table_types
        for section in ordered_sections
    ):
        ordered_sections = [
            section
            for section in ordered_sections
            if not (section["render_mode"] == "variable" and section["type"] == "members_list")
        ]

    variable_sections = [section for section in ordered_sections if section["render_mode"] == "variable"]
    advanced = bool(ordered_sections)

    leaf_columns = []
    col_idx = 1
    for col in leading:
        leaf_columns.append({"col_idx": col_idx, "role": "main", "merge": True, "column": col})
        col_idx += 1
    for section in ordered_sections:
        if section["type"] == "exercise_table":
            for group in section.get("groups") or []:
                for col in section.get("columns") or []:
                    leaf_columns.append(
                        {
                            "col_idx": col_idx,
                            "role": section["render_mode"],
                            "merge": section["render_mode"] == "fixed",
                            "section_index": section["section_index"],
                            "section_type": section["type"],
                            "group_key": group["key"],
                            "column": _copy_col(col),
                        }
                    )
                    col_idx += 1
        else:
            for col in section.get("columns") or []:
                leaf_columns.append(
                    {
                        "col_idx": col_idx,
                        "role": section["render_mode"],
                        "merge": section["render_mode"] == "fixed",
                        "section_index": section["section_index"],
                        "section_type": section["type"],
                        "column": _copy_col(col),
                    }
                )
                col_idx += 1
    for col in trailing:
        leaf_columns.append({"col_idx": col_idx, "role": "main", "merge": True, "column": col})
        col_idx += 1

    return {
        "kind": layout_kind,
        "advanced": advanced,
        "leading": leading,
        "ordered_sections": ordered_sections,
        "variable_sections": variable_sections,
        "trailing": trailing,
        "leaf_columns": leaf_columns,
        "total_columns": max(1, len(leaf_columns) if advanced else len(cols)),
    }


def _fixed_leaf_value(row, leaf):
    section = _detail_section_at(row, leaf.get("section_index", -1), leaf.get("section_type", ""))
    if not isinstance(section, dict):
        return ""
    if leaf.get("section_type") == "exercise_table":
        target_key = leaf.get("group_key")
        for detail_row in section.get("rows") or []:
            if not isinstance(detail_row, dict):
                continue
            key = (detail_row.get("app_id"), detail_row.get("exercise_index"))
            if key == target_key:
                return _detail_cell_value(detail_row, leaf.get("column") or {})
        return ""
    rows = section.get("rows") or []
    if not rows:
        return ""
    return _detail_cell_value(rows[0], leaf.get("column") or {})


def _variable_entries_for_section(row, section):
    current = _detail_section_at(row, section.get("section_index", -1), section.get("type", ""))
    if not isinstance(current, dict):
        return []
    if current.get("type") == "members_list":
        return [item for item in (current.get("items") or []) if isinstance(item, dict)]
    return [item for item in (current.get("rows") or []) if isinstance(item, dict)]


def _build_block_rows(row, variable_sections):
    if not variable_sections:
        return [{}]

    ordered_keys = []
    block_rows = {}
    fallback_counter = 0
    for section in variable_sections:
        entries = _variable_entries_for_section(row, section)
        for pos, entry in enumerate(entries):
            member_id = entry.get("member_id")
            if member_id not in (None, ""):
                key = ("member", member_id)
            else:
                key = ("pos", pos, fallback_counter)
            if key not in block_rows:
                ordered_keys.append(key)
                block_rows[key] = {}
            block_rows[key][section["section_index"]] = entry
        fallback_counter += 1

    return [block_rows[key] for key in ordered_keys] or [{}]


def sanitize_filename_component(raw):
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(raw or "").strip()).strip("._")
    return cleaned or "classificacions"


def write_cfg_excel_sheet(ws, competicio, cfg_nom, columns, parts, *, tipus="individual", schema=None):
    cols = columns if isinstance(columns, list) and columns else default_live_columns()
    layout = _build_layout(cols, parts, tipus=tipus, schema=schema)
    total_cols = layout["total_columns"] if layout["advanced"] else max(1, len(cols))

    fill_title = PatternFill("solid", fgColor="1F4E79")
    fill_subtitle = PatternFill("solid", fgColor="D9E1F2")
    fill_partition = PatternFill("solid", fgColor="DDE7FF")
    fill_header = PatternFill("solid", fgColor="E9EEF7")
    fill_header_group = PatternFill("solid", fgColor="D7E3F4")
    fill_header_subgroup = PatternFill("solid", fgColor="EEF3FB")
    fill_zebra = PatternFill("solid", fgColor="F7F9FC")
    fill_first = PatternFill("solid", fgColor="F6E27A")
    fill_second = PatternFill("solid", fgColor="E3E8EF")
    fill_third = PatternFill("solid", fgColor="E8C7A3")

    thin = Side(style="thin", color="9AA7B2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    subtitle_font = Font(bold=True, size=11)
    header_font = Font(bold=True)

    col_widths = [10 for _ in range(total_cols)]

    def update_col_width(col_idx, value):
        txt = "" if value is None else str(value)
        longest = max([len(line) for line in txt.splitlines()] or [0])
        col_widths[col_idx - 1] = min(48, max(col_widths[col_idx - 1], longest + 2))

    def data_fill(row_pos, posicio):
        if posicio == 1:
            return fill_first
        if posicio == 2:
            return fill_second
        if posicio == 3:
            return fill_third
        if row_pos % 2 == 1:
            return fill_zebra
        return None

    def write_header_group_cell(row_num, start_col, end_col, value, fill):
        if start_col > end_col:
            return
        if start_col != end_col:
            ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
        cell = ws.cell(row=row_num, column=start_col, value=value)
        cell.fill = fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border
        update_col_width(start_col, value)

    def write_data_cell(row_num, col_num, raw_value, col_def, fill):
        value, number_format, wrap_text = normalize_excel_cell(raw_value, col_def or {})
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.border = border
        if fill is not None:
            cell.fill = fill
        if number_format:
            cell.number_format = number_format
        align = str((col_def or {}).get("align") or "").strip().lower()
        horizontal = "right" if align == "right" else ("center" if align == "center" else "left")
        cell.alignment = Alignment(horizontal=horizontal, vertical="top" if wrap_text else "center", wrap_text=wrap_text)
        update_col_width(col_num, value)
        text_value = "" if value is None else str(value)
        return max(1, len(text_value.splitlines()) if text_value else 1)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c_title = ws.cell(row=1, column=1, value=f"{competicio.nom} - {cfg_nom}")
    c_title.fill = fill_title
    c_title.font = title_font
    c_title.alignment = align_center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c_sub = ws.cell(
        row=2,
        column=1,
        value=f"Exportat el {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}",
    )
    c_sub.fill = fill_subtitle
    c_sub.font = subtitle_font
    c_sub.alignment = align_center

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20

    row_idx = 4
    first_data_row = None
    parts_list = parts if isinstance(parts, list) else []
    if not parts_list:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        empty = ws.cell(row=row_idx, column=1, value="Sense resultats.")
        empty.alignment = align_left
        empty.font = Font(italic=True)
        return

    for part in parts_list:
        part_rows = (part or {}).get("rows") or []
        part_name = format_partition_title((part or {}).get("particio"))

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        pc = ws.cell(row=row_idx, column=1, value=f"Particio: {part_name} ({len(part_rows)} files)")
        pc.fill = fill_partition
        pc.font = Font(bold=True)
        pc.alignment = align_left
        pc.border = border
        row_idx += 1

        header_row = row_idx
        if not layout["advanced"]:
            if first_data_row is None:
                first_data_row = header_row + 1
            for col_pos, col in enumerate(cols, start=1):
                label = str(col.get("label") or col.get("key") or "")
                hc = ws.cell(row=header_row, column=col_pos, value=label)
                hc.fill = fill_header
                hc.font = header_font
                hc.alignment = align_center
                hc.border = border
                update_col_width(col_pos, label)
            row_idx += 1
        else:
            if first_data_row is None:
                first_data_row = header_row + 3
            cursor = 1
            for col in layout["leading"]:
                ws.merge_cells(start_row=header_row, start_column=cursor, end_row=header_row + 2, end_column=cursor)
                hc = ws.cell(row=header_row, column=cursor, value=str(col.get("label") or col.get("key") or ""))
                hc.fill = fill_header_group
                hc.font = header_font
                hc.alignment = align_center
                hc.border = border
                update_col_width(cursor, hc.value)
                cursor += 1

            for section in layout["ordered_sections"]:
                start_col = cursor
                if section["type"] == "exercise_table":
                    width = len(section.get("columns") or [])
                    end_col = start_col + (len(section.get("groups") or []) * width) - 1
                    write_header_group_cell(header_row, start_col, end_col, section["label"], fill_header_group)
                    for group in section.get("groups") or []:
                        group_start = cursor
                        group_end = cursor + width - 1
                        write_header_group_cell(header_row + 1, group_start, group_end, group["label"], fill_header_subgroup)
                        for col in section.get("columns") or []:
                            label = str(col.get("label") or col.get("key") or "")
                            hc = ws.cell(row=header_row + 2, column=cursor, value=label)
                            hc.fill = fill_header
                            hc.font = header_font
                            hc.alignment = align_center
                            hc.border = border
                            update_col_width(cursor, label)
                            cursor += 1
                else:
                    width = len(section.get("columns") or [])
                    end_col = start_col + width - 1
                    write_header_group_cell(header_row, start_col, end_col, section["label"], fill_header_group)
                    write_header_group_cell(header_row + 1, start_col, end_col, "", fill_header_subgroup)
                    for col in section.get("columns") or []:
                        label = str(col.get("label") or col.get("key") or "")
                        hc = ws.cell(row=header_row + 2, column=cursor, value=label)
                        hc.fill = fill_header
                        hc.font = header_font
                        hc.alignment = align_center
                        hc.border = border
                        update_col_width(cursor, label)
                        cursor += 1

            for col in layout["trailing"]:
                ws.merge_cells(start_row=header_row, start_column=cursor, end_row=header_row + 2, end_column=cursor)
                hc = ws.cell(row=header_row, column=cursor, value=str(col.get("label") or col.get("key") or ""))
                hc.fill = fill_header_group
                hc.font = header_font
                hc.alignment = align_center
                hc.border = border
                update_col_width(cursor, hc.value)
                cursor += 1

            ws.row_dimensions[header_row].height = 22
            ws.row_dimensions[header_row + 1].height = 20
            ws.row_dimensions[header_row + 2].height = 20
            row_idx += 3

        if not part_rows:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
            rc = ws.cell(row=row_idx, column=1, value="Sense resultats en aquesta particio.")
            rc.alignment = align_left
            rc.font = Font(italic=True)
            row_idx += 2
            continue

        if not layout["advanced"]:
            for data_pos, row in enumerate(part_rows):
                try:
                    posicio = int((row or {}).get("posicio"))
                except Exception:
                    posicio = None
                row_fill = data_fill(data_pos, posicio)
                max_lines = 1
                for col_pos, col in enumerate(cols, start=1):
                    raw_value = extract_export_value(row or {}, col or {})
                    max_lines = max(max_lines, write_data_cell(row_idx, col_pos, raw_value, col or {}, row_fill))
                if max_lines > 1:
                    ws.row_dimensions[row_idx].height = min(120, max(18, max_lines * 14))
                row_idx += 1
            row_idx += 1
            continue

        for data_pos, row in enumerate(part_rows):
            try:
                posicio = int((row or {}).get("posicio"))
            except Exception:
                posicio = None
            row_fill = data_fill(data_pos, posicio)
            block_rows = _build_block_rows(row or {}, layout["variable_sections"])
            block_height = max(1, len(block_rows))
            block_start = row_idx
            block_end = row_idx + block_height - 1

            for rel_idx, block_row in enumerate(block_rows):
                current_row = block_start + rel_idx
                max_lines = 1
                for leaf in layout["leaf_columns"]:
                    col_num = leaf["col_idx"]
                    if leaf["merge"]:
                        if rel_idx != 0:
                            continue
                        if leaf["role"] == "main":
                            raw_value = extract_export_value(row or {}, leaf.get("column") or {})
                        else:
                            raw_value = _fixed_leaf_value(row or {}, leaf)
                        max_lines = max(max_lines, write_data_cell(current_row, col_num, raw_value, leaf.get("column") or {}, row_fill))
                        if block_height > 1:
                            ws.merge_cells(
                                start_row=current_row,
                                start_column=col_num,
                                end_row=block_end,
                                end_column=col_num,
                            )
                    else:
                        source_row = block_row.get(leaf.get("section_index"))
                        raw_value = _detail_cell_value(source_row or {}, leaf.get("column") or {}) if source_row else ""
                        max_lines = max(max_lines, write_data_cell(current_row, col_num, raw_value, leaf.get("column") or {}, row_fill))

                if max_lines > 1:
                    ws.row_dimensions[current_row].height = min(120, max(18, max_lines * 14))

            row_idx = block_end + 1
        row_idx += 1

    if first_data_row:
        ws.freeze_panes = ws[f"A{first_data_row}"]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(10, width)


_normalize_excel_cell = normalize_excel_cell


__all__ = [
    "_normalize_excel_cell",
    "build_excel_sheet_name",
    "excel_safe_text",
    "format_scalar_text",
    "normalize_excel_cell",
    "sanitize_filename_component",
    "try_float",
    "write_cfg_excel_sheet",
]
