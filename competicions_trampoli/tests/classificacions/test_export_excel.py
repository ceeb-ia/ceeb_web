import json
import re
from io import BytesIO, StringIO
from datetime import date, timedelta
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.sessions.backends.db import SessionStore
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Max
from django.test import RequestFactory, TestCase
from django.urls import Resolver404, resolve, reverse
from django.utils import timezone
from openpyxl import load_workbook

from ceeb_web.auth_groups import GLOBAL_AUTH_GROUPS

from ... import live_cache
from ...access import user_has_competicio_capability
from ...forms import CompeticioAparellForm
from ...models import (
    Competicio,
    Equip,
    EquipContext,
    GrupCompeticio,
    Inscripcio,
    InscripcioEquipAssignacio,
    InscripcioMedia,
)
from ...models.judging import (
    JudgeConversation,
    JudgeConversationMessage,
    JudgeDeviceToken,
    PublicLiveToken,
)
from ...models.classificacions import ClassificacioConfig, ClassificacioTemplateGlobal
from ...models.rotacions import (
    RotacioAssignacio,
    RotacioAssignacioGrup,
    RotacioAssignacioSerieEquip,
    RotacioEstacio,
    RotacioFranja,
)
from ...models.scoring import (
    ScoringSchema,
    ScoreEntry,
    ScoreEntryVideo,
    ScoreEntryVideoEvent,
    SerieEquip,
    SerieEquipItem,
    TeamScoreEntry,
    TeamCompetitiveSubject,
    TeamScoreEntryVideo,
    TeamScoreEntryVideoEvent,
)
from ...models.competicio import (
    Aparell,
    CompeticioAparell,
    CompeticioAparellEquipContextSource,
    InscripcioAparellExclusio,
)
from ...models import CompeticioMembership
from ...scoring_engine import ScoringEngine
from ...services.inscripcions.groups import renumber_groups_for_competicio
from ...services.inscripcions.sorting import (
    _split_custom_sort_tokens,
    sort_records_by_field_stable,
)
from ...services.inscripcions.history import (
    apply_inscripcions_history_snapshot,
    capture_inscripcions_history_snapshot,
)
from ...services.inscripcions.queries import (
    COLUMN_FILTER_EMPTY_TOKEN,
    _build_inscripcions_filtered_qs,
    build_inscripcions_sort_context_key,
    get_competicio_custom_sort_rank_map,
)
from ...services.classificacions.classificacio_templates import (
    normalize_particions_schema as _normalize_particions_schema,
    schema_to_template_schema as _schema_to_template_schema,
    template_schema_to_competicio_schema as _template_schema_to_competicio_schema_service,
)
from ...services.classificacions.builder import (
    prepare_schema_for_builder_hydration,
    scoreable_codes_by_app_id as _scoreable_codes_by_app_id,
)
from ...services.classificacions.compute import DEFAULT_SCHEMA, compute_classificacio
from ...services.classificacions.export import _normalize_excel_cell
from ...services.classificacions.partitions import normalize_schema_legacy_team_birth_partition
from ...services.classificacions.validation import (
    build_metric_meta_for_comp_aparell as _build_metric_meta_for_comp_aparell,
    build_scoreable_meta_for_schema as _build_scoreable_meta_for_schema,
    validate_particions_schema as _validate_particions_schema,
    validate_schema_for_competicio as _validate_schema_for_competicio,
)
from ...views.classificacions.builder import ClassificacionsHome
from ...services.shared.competition_groups import (
    assign_groups_by_display_num,
    compact_competition_order_for_group,
    ensure_group_for_display_num,
    get_group_maps,
    get_group_participant_counts,
    get_out_of_program_group_ids,
    get_programmed_group_ids,
    group_label,
    move_inscripcio_to_group,
    next_group_display_num,
)
from ...services.scoring.team_scoring import (
    build_permission_label,
    build_team_subjects_for_comp_aparell,
    resolve_permission_runtime_entries,
    runtime_schema_for_comp_aparell,
)
from ...services.teams.team_series import safe_deactivate_empty_serie
from ...views.judge.admin import _member_slot_choices, _validate_permission_row
from ...templatetags.competicio_extras import (
    DEFAULT_COMPETITION_BACKGROUND,
    get_competicio_background_url_from_request,
)

from ..base import _BaseTrampoliDataMixin


def _template_schema_to_competicio_schema(*args, **kwargs):
    schema_local, mapping_warnings, mapping, _compat_meta = _template_schema_to_competicio_schema_service(
        *args,
        **kwargs,
    )
    return schema_local, mapping_warnings, mapping


class ClassificacionsExportExcelTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Export Classificacions")
        self.app = self._create_aparell("TRAMP_EXPORT", "Tramp Export")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        self.ins_a = self._create_inscripcio(self.comp, "Participant A", ordre=1)
        self.ins_b = self._create_inscripcio(self.comp, "Participant B", ordre=2)
        self.ins_c = self._create_inscripcio(self.comp, "Participant C", ordre=3)

        self.ins_a.categoria = "Junior"
        self.ins_a.subcategoria = "Femeni"
        self.ins_a.save(update_fields=["categoria", "subcategoria"])

        self.ins_b.categoria = "Junior"
        self.ins_b.subcategoria = "Femeni"
        self.ins_b.save(update_fields=["categoria", "subcategoria"])

        self.ins_c.categoria = "Senior"
        self.ins_c.subcategoria = "Masculi"
        self.ins_c.save(update_fields=["categoria", "subcategoria"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=9.6,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=8.3,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_c,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=7.1,
        )

        self.cfg_general = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="General",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=self._schema_for_parts([]),
        )
        self.cfg_particions = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Per categories",
            activa=True,
            ordre=2,
            tipus="individual",
            schema=self._schema_for_parts(["categoria", "subcategoria"]),
        )

        User = get_user_model()
        self.user = User.objects.create_user(
            username="classif_export_user",
            password="testpass123",
            email="classif-export@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.READONLY,
            is_active=True,
        )

    def _schema_for_parts(self, parts):
        return {
            "particions": parts,
            "filtres": {},
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "exercicis_best_n": 1,
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
                "camp": "total",
                "agregacio": "sum",
                "best_n": 1,
            },
            "desempat": [],
            "presentacio": {
                "top_n": 0,
                "mostrar_empats": True,
                "columnes": [
                    {"type": "builtin", "key": "posicio", "label": "Pos.", "align": "left"},
                    {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                ],
            },
        }

    def _export_with_runtime(self, cfg, runtime_payload):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        with patch(
            "competicions_trampoli.views.classificacions.export.execute_classificacio_runtime",
            return_value=runtime_payload,
        ):
            return self.client.get(url, {"cfg_id": cfg.id})

    def test_export_excel_creates_one_sheet_per_classificacio(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url)

        self.assertEqual(res.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            res["Content-Type"],
        )

        wb = load_workbook(filename=BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), 2)

        ws_general = wb[wb.sheetnames[0]]
        ws_parts = wb[wb.sheetnames[1]]

        self.assertIn("General", str(ws_general["A1"].value))
        self.assertIn("Particio:", str(ws_general["A4"].value))
        self.assertEqual(ws_general.freeze_panes, "A6")

        gold_fill = (ws_general["A6"].fill.fgColor.rgb or "").upper()
        silver_fill = (ws_general["A7"].fill.fgColor.rgb or "").upper()
        header_fill = (ws_general["A5"].fill.fgColor.rgb or "").upper()
        self.assertTrue(gold_fill.endswith("F6E27A"))
        self.assertTrue(silver_fill.endswith("E3E8EF"))
        self.assertTrue(header_fill.endswith("E9EEF7"))

        self.assertIn("Particio:", str(ws_parts["A4"].value))
        self.assertIn("/", str(ws_parts["A4"].value))

    def test_export_excel_can_filter_single_cfg_with_cfg_id(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url, {"cfg_id": self.cfg_particions.id})

        self.assertEqual(res.status_code, 200)
        wb = load_workbook(filename=BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), 1)
        ws = wb[wb.sheetnames[0]]
        self.assertIn("Per categories", str(ws["A1"].value))

    def test_export_excel_rejects_invalid_cfg_id(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url, {"cfg_id": "abc"})
        self.assertEqual(res.status_code, 400)

    def test_export_excel_returns_consistent_error_when_compute_fails(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        with patch("competicions_trampoli.views.classificacions.export.compute_classificacio", side_effect=RuntimeError("boom export")):
            res = self.client.get(url, {"cfg_id": self.cfg_general.id})

        self.assertEqual(res.status_code, 400)
        body = res.content.decode("utf-8")
        self.assertIn("No s'ha pogut renderitzar la classificacio.", body)
        self.assertIn("boom export", body)

    def test_export_excel_flattens_individual_detail_sections_before_last_builtin(self):
        runtime = {
            "schema": {"presentacio": {}, "equips": {}},
            "columns": [
                {"type": "builtin", "key": "posicio", "label": "Pos.", "align": "left"},
                {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
            ],
            "parts": [
                {
                    "particio": "global",
                    "rows": [
                        {
                            "posicio": 1,
                            "participant": "Participant A",
                            "punts": 18.5,
                            "cells": {"posicio": 1, "participant": "Participant A", "punts": 18.5},
                            "detail": {
                                "sections": [
                                    {
                                        "type": "exercise_table",
                                        "label": "Exercicis",
                                        "columns": [
                                            {"type": "builtin", "key": "aparell_nom", "label": "Aparell", "align": "left"},
                                            {"type": "builtin", "key": "exercise_index", "label": "Ex.", "align": "left"},
                                            {
                                                "type": "raw",
                                                "key": "exercise_total",
                                                "label": "Total",
                                                "align": "right",
                                                "decimals": 3,
                                            },
                                        ],
                                        "rows": [
                                            {
                                                "app_id": self.comp_app.id,
                                                "exercise_index": 1,
                                                "aparell_nom": "Tramp",
                                                "cells": {
                                                    "aparell_nom": "Tramp",
                                                    "exercise_index": 1,
                                                    "exercise_total": 9.25,
                                                },
                                            },
                                            {
                                                "app_id": self.comp_app.id,
                                                "exercise_index": 2,
                                                "aparell_nom": "Tramp",
                                                "cells": {
                                                    "aparell_nom": "Tramp",
                                                    "exercise_index": 2,
                                                    "exercise_total": 9.25,
                                                },
                                            },
                                        ],
                                    }
                                ]
                            },
                        }
                    ],
                }
            ],
            "error": None,
        }

        res = self._export_with_runtime(self.cfg_general, runtime)
        self.assertEqual(res.status_code, 200)

        wb = load_workbook(filename=BytesIO(res.content))
        ws = wb[wb.sheetnames[0]]

        self.assertEqual(ws.freeze_panes, "A8")
        self.assertIn("C5:H5", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertIn("C6:E6", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertIn("F6:H6", {str(rng) for rng in ws.merged_cells.ranges})
        self.assertEqual(ws["C5"].value, "Exercicis")
        self.assertEqual(ws["C6"].value, "Tramp - Ex.1")
        self.assertEqual(ws["F6"].value, "Tramp - Ex.2")
        self.assertEqual(ws["I7"].value, "Punts")
        self.assertEqual(ws["C8"].value, "Tramp")
        self.assertEqual(ws["D8"].value, 1)
        self.assertEqual(ws["E8"].value, 9.25)
        self.assertEqual(ws["H8"].value, 9.25)
        self.assertEqual(ws["I8"].value, 18.5)

    def test_export_excel_renders_derived_team_detail_as_block_rows(self):
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Equips derivats",
            activa=True,
            ordre=10,
            tipus="equips",
            schema={"equips": {"team_mode": "derived_from_individual"}},
        )
        runtime = {
            "schema": {"equips": {"team_mode": "derived_from_individual"}},
            "columns": [
                {"type": "builtin", "key": "posicio", "label": "Pos.", "align": "left"},
                {"type": "builtin", "key": "participant", "label": "Equip", "align": "left"},
                {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
            ],
            "parts": [
                {
                    "particio": "global",
                    "rows": [
                        {
                            "posicio": 1,
                            "participant": "Equip A",
                            "punts": 25.3,
                            "cells": {"posicio": 1, "participant": "Equip A", "punts": 25.3},
                            "detail": {
                                "sections": [
                                    {
                                        "type": "members_table",
                                        "label": "Detall membres",
                                        "columns": [
                                            {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                                            {
                                                "type": "raw",
                                                "key": "member_total",
                                                "label": "Total",
                                                "align": "right",
                                                "decimals": 3,
                                            },
                                        ],
                                        "rows": [
                                            {
                                                "member_id": 101,
                                                "participant": "Anna",
                                                "cells": {"participant": "Anna", "member_total": 12.6},
                                            },
                                            {
                                                "member_id": 102,
                                                "participant": "Berta",
                                                "cells": {"participant": "Berta", "member_total": 12.7},
                                            },
                                        ],
                                    }
                                ]
                            },
                        }
                    ],
                }
            ],
            "error": None,
        }

        res = self._export_with_runtime(cfg, runtime)
        self.assertEqual(res.status_code, 200)

        wb = load_workbook(filename=BytesIO(res.content))
        ws = wb[wb.sheetnames[0]]
        merged = {str(rng) for rng in ws.merged_cells.ranges}

        self.assertEqual(ws.freeze_panes, "A8")
        self.assertIn("A8:A9", merged)
        self.assertIn("B8:B9", merged)
        self.assertIn("E8:E9", merged)
        self.assertNotIn("C8:C9", merged)
        self.assertEqual(ws["C5"].value, "Detall membres")
        self.assertEqual(ws["C8"].value, "Anna")
        self.assertEqual(ws["C9"].value, "Berta")
        self.assertEqual(ws["D8"].value, 12.6)
        self.assertEqual(ws["E8"].value, 25.3)

    def test_export_excel_renders_native_team_metrics_with_member_rows(self):
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Equips natius",
            activa=True,
            ordre=11,
            tipus="equips",
            schema={"equips": {"team_mode": "native_team"}},
        )
        runtime = {
            "schema": {"equips": {"team_mode": "native_team"}},
            "columns": [
                {"type": "builtin", "key": "posicio", "label": "Pos.", "align": "left"},
                {"type": "builtin", "key": "participant", "label": "Equip", "align": "left"},
                {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
            ],
            "parts": [
                {
                    "particio": "global",
                    "rows": [
                        {
                            "posicio": 1,
                            "participant": "Equip Team",
                            "punts": 24.8,
                            "cells": {"posicio": 1, "participant": "Equip Team", "punts": 24.8},
                            "detail": {
                                "sections": [
                                    {
                                        "type": "team_metrics",
                                        "label": "Notes equip",
                                        "columns": [
                                            {
                                                "type": "raw",
                                                "key": "team_total",
                                                "label": "Total equip",
                                                "align": "right",
                                                "decimals": 3,
                                            }
                                        ],
                                        "rows": [{"cells": {"team_total": 24.8}}],
                                    },
                                    {
                                        "type": "team_members_table",
                                        "label": "Notes per membre",
                                        "columns": [
                                            {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                                            {
                                                "type": "raw",
                                                "key": "member_total",
                                                "label": "Total",
                                                "align": "right",
                                                "decimals": 3,
                                            },
                                        ],
                                        "rows": [
                                            {
                                                "member_id": 201,
                                                "participant": "Anna",
                                                "cells": {"participant": "Anna", "member_total": 12.4},
                                            },
                                            {
                                                "member_id": 202,
                                                "participant": "Berta",
                                                "cells": {"participant": "Berta", "member_total": 12.4},
                                            },
                                        ],
                                    },
                                ]
                            },
                        }
                    ],
                }
            ],
            "error": None,
        }

        res = self._export_with_runtime(cfg, runtime)
        self.assertEqual(res.status_code, 200)

        wb = load_workbook(filename=BytesIO(res.content))
        ws = wb[wb.sheetnames[0]]
        merged = {str(rng) for rng in ws.merged_cells.ranges}

        self.assertEqual(ws.freeze_panes, "A8")
        self.assertIn("A8:A9", merged)
        self.assertIn("B8:B9", merged)
        self.assertIn("C8:C9", merged)
        self.assertIn("F8:F9", merged)
        self.assertNotIn("D8:D9", merged)
        self.assertEqual(ws["C5"].value, "Notes equip")
        self.assertEqual(ws["D5"].value, "Notes per membre")
        self.assertEqual(ws["C8"].value, 24.8)
        self.assertEqual(ws["D8"].value, "Anna")
        self.assertEqual(ws["D9"].value, "Berta")
        self.assertEqual(ws["F8"].value, 24.8)

    def test_export_excel_preserves_interleaved_native_team_section_order(self):
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Equips natius ordre",
            activa=True,
            ordre=11,
            tipus="equips",
            schema={"equips": {"team_mode": "native_team"}},
        )
        runtime = {
            "schema": {"equips": {"team_mode": "native_team"}},
            "columns": [
                {"type": "builtin", "key": "posicio", "label": "Pos.", "align": "left"},
                {"type": "builtin", "key": "participant", "label": "Equip", "align": "left"},
                {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
            ],
            "parts": [
                {
                    "particio": "global",
                    "rows": [
                        {
                            "posicio": 1,
                            "participant": "Equip Team",
                            "punts": 40.0,
                            "cells": {"posicio": 1, "participant": "Equip Team", "punts": 40.0},
                            "detail": {
                                "sections": [
                                    {
                                        "type": "team_metrics",
                                        "label": "Metriques 1",
                                        "columns": [
                                            {
                                                "type": "raw",
                                                "key": "team_total_1",
                                                "label": "Total 1",
                                                "align": "right",
                                                "decimals": 3,
                                            }
                                        ],
                                        "rows": [{"cells": {"team_total_1": 10.0}}],
                                    },
                                    {
                                        "type": "team_members_table",
                                        "label": "Notes 1",
                                        "columns": [
                                            {"type": "builtin", "key": "participant", "label": "Participant 1", "align": "left"},
                                        ],
                                        "rows": [
                                            {
                                                "member_id": 201,
                                                "participant": "Anna",
                                                "cells": {"participant": "Anna"},
                                            },
                                            {
                                                "member_id": 202,
                                                "participant": "Berta",
                                                "cells": {"participant": "Berta"},
                                            },
                                        ],
                                    },
                                    {
                                        "type": "team_metrics",
                                        "label": "Metriques 2",
                                        "columns": [
                                            {
                                                "type": "raw",
                                                "key": "team_total_2",
                                                "label": "Total 2",
                                                "align": "right",
                                                "decimals": 3,
                                            }
                                        ],
                                        "rows": [{"cells": {"team_total_2": 20.0}}],
                                    },
                                    {
                                        "type": "team_members_table",
                                        "label": "Notes 2",
                                        "columns": [
                                            {"type": "builtin", "key": "participant", "label": "Participant 2", "align": "left"},
                                        ],
                                        "rows": [
                                            {
                                                "member_id": 201,
                                                "participant": "Anna 2",
                                                "cells": {"participant": "Anna 2"},
                                            },
                                            {
                                                "member_id": 202,
                                                "participant": "Berta 2",
                                                "cells": {"participant": "Berta 2"},
                                            },
                                        ],
                                    },
                                ]
                            },
                        }
                    ],
                }
            ],
            "error": None,
        }

        res = self._export_with_runtime(cfg, runtime)
        self.assertEqual(res.status_code, 200)

        wb = load_workbook(filename=BytesIO(res.content))
        ws = wb[wb.sheetnames[0]]

        self.assertEqual(ws["C5"].value, "Metriques 1")
        self.assertEqual(ws["D5"].value, "Notes 1")
        self.assertEqual(ws["E5"].value, "Metriques 2")
        self.assertEqual(ws["F5"].value, "Notes 2")
        self.assertEqual(ws["C7"].value, "Total 1")
        self.assertEqual(ws["D7"].value, "Participant 1")
        self.assertEqual(ws["E7"].value, "Total 2")
        self.assertEqual(ws["F7"].value, "Participant 2")
        self.assertEqual(ws["C8"].value, 10.0)
        self.assertEqual(ws["D8"].value, "Anna")
        self.assertEqual(ws["E8"].value, 20.0)
        self.assertEqual(ws["F8"].value, "Anna 2")

    def test_export_excel_renders_entity_detail_as_block_rows(self):
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Per entitats",
            activa=True,
            ordre=12,
            tipus="entitat",
            schema={},
        )
        runtime = {
            "schema": {},
            "columns": [
                {"type": "builtin", "key": "posicio", "label": "Pos.", "align": "left"},
                {"type": "builtin", "key": "participant", "label": "Entitat", "align": "left"},
                {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
            ],
            "parts": [
                {
                    "particio": "global",
                    "rows": [
                        {
                            "posicio": 1,
                            "participant": "Club X",
                            "entitat_nom": "Club X",
                            "punts": 20.0,
                            "cells": {
                                "posicio": 1,
                                "participant": "Club X",
                                "entitat_nom": "Club X",
                                "punts": 20.0,
                            },
                            "detail": {
                                "sections": [
                                    {
                                        "type": "entity_members_table",
                                        "label": "Participants",
                                        "columns": [
                                            {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                                        ],
                                        "rows": [
                                            {
                                                "member_id": 301,
                                                "participant": "Anna",
                                                "cells": {"participant": "Anna"},
                                            },
                                            {
                                                "member_id": 302,
                                                "participant": "Berta",
                                                "cells": {"participant": "Berta"},
                                            },
                                        ],
                                    }
                                ]
                            },
                        }
                    ],
                }
            ],
            "error": None,
        }

        res = self._export_with_runtime(cfg, runtime)
        self.assertEqual(res.status_code, 200)

        wb = load_workbook(filename=BytesIO(res.content))
        ws = wb[wb.sheetnames[0]]
        merged = {str(rng) for rng in ws.merged_cells.ranges}

        self.assertEqual(ws.freeze_panes, "A8")
        self.assertIn("A8:A9", merged)
        self.assertIn("B8:B9", merged)
        self.assertIn("D8:D9", merged)
        self.assertEqual(ws["C5"].value, "Participants")
        self.assertEqual(ws["C8"].value, "Anna")
        self.assertEqual(ws["C9"].value, "Berta")

