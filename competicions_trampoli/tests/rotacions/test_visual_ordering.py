import json
from importlib import import_module
import re
from io import BytesIO, StringIO
from datetime import date, timedelta
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.sessions.backends.db import SessionStore
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Max
from django.test import RequestFactory, TestCase
from django.urls import Resolver404, resolve, reverse
from django.utils import timezone
from openpyxl import load_workbook

from ceeb_web.auth_groups import GLOBAL_AUTH_GROUPS

from ... import live_cache
from ...access import user_has_competicio_capability
from ...forms import CompeticioAparellForm
from ...models import (
    Competicio,
    Equip,
    EquipContext,
    GrupCompeticio,
    Inscripcio,
    InscripcioEquipAssignacio,
    InscripcioMedia,
)
from ...models.judging import (
    JudgeConversation,
    JudgeConversationMessage,
    JudgeDeviceToken,
    PublicLiveToken,
)
from ...models.classificacions import ClassificacioConfig, ClassificacioTemplateGlobal
from ...models.rotacions import (
    RotacioAssignacio,
    RotacioAssignacioGrup,
    RotacioAssignacioSerieEquip,
    RotacioEstacio,
    RotacioFranja,
)
from ...models.scoring import (
    ScoringSchema,
    ScoreEntry,
    ScoreEntryVideo,
    ScoreEntryVideoEvent,
    SerieEquip,
    SerieEquipItem,
    TeamScoreEntry,
    TeamCompetitiveSubject,
    TeamScoreEntryVideo,
    TeamScoreEntryVideoEvent,
)
from ...models.competicio import (
    Aparell,
    CompeticioAparell,
    CompeticioAparellEquipContextSource,
    InscripcioAparellExclusio,
)
from ...models import CompeticioMembership
from ...scoring_engine import ScoringEngine
from ...services.inscripcions.groups import renumber_groups_for_competicio
from ...services.inscripcions.sorting import (
    _split_custom_sort_tokens,
    sort_records_by_field_stable,
)
from ...services.inscripcions.history import (
    apply_inscripcions_history_snapshot,
    capture_inscripcions_history_snapshot,
)
from ...services.inscripcions.queries import (
    COLUMN_FILTER_EMPTY_TOKEN,
    _build_inscripcions_filtered_qs,
    build_inscripcions_sort_context_key,
    get_competicio_custom_sort_rank_map,
)
from ...services.classificacions.builder import scoreable_codes_by_app_id as _scoreable_codes_by_app_id
from ...services.classificacions.compute import DEFAULT_SCHEMA, compute_classificacio
from ...services.classificacions.export import _normalize_excel_cell
from ...services.classificacions.partitions import normalize_schema_legacy_team_birth_partition
from ...services.classificacions.validation import (
    build_metric_meta_for_comp_aparell as _build_metric_meta_for_comp_aparell,
    build_scoreable_meta_for_schema as _build_scoreable_meta_for_schema,
    validate_particions_schema as _validate_particions_schema,
    validate_schema_for_competicio as _validate_schema_for_competicio,
)
from ...services.classificacions.classificacio_templates import (
    normalize_particions_schema as _normalize_particions_schema,
    schema_to_template_schema as _schema_to_template_schema,
    template_schema_to_competicio_schema as _template_schema_to_competicio_schema,
)
from ...views.classificacions.builder import ClassificacionsHome
from ...services.shared.competition_groups import (
    assign_groups_by_display_num,
    compact_competition_order_for_group,
    ensure_group_for_display_num,
    get_group_maps,
    get_group_participant_counts,
    get_out_of_program_group_ids,
    get_programmed_group_ids,
    group_label,
    move_inscripcio_to_group,
    next_group_display_num,
)
from ...services.scoring.team_scoring import (
    build_permission_label,
    build_team_subjects_for_comp_aparell,
    resolve_permission_runtime_entries,
    runtime_schema_for_comp_aparell,
)
from ...services.teams.team_series import safe_deactivate_empty_serie
from ...views.judge.admin import _member_slot_choices, _validate_permission_row
from ...templatetags.competicio_extras import (
    DEFAULT_COMPETITION_BACKGROUND,
    get_competicio_background_url_from_request,
)

from ..base import _BaseTrampoliDataMixin


class FranjaVisualOrderingTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Franja Visual")
        User = get_user_model()
        self.user = User.objects.create_user(
            username="franja_visual_owner",
            password="testpass123",
            email="franja-visual-owner@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.OWNER,
            is_active=True,
        )
        self.client.force_login(self.user)

        self.aparell = self._create_aparell("ROT_VIS", "Rotation Visual")
        self.comp_app = self._create_comp_aparell(self.comp, self.aparell, ordre=1, actiu=True)
        self.estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=1,
            actiu=True,
        )

        self.f1 = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="09:00",
            hora_fi="09:30",
            ordre=1,
            ordre_visual=1,
            titol="Franja 1",
            tipus=RotacioFranja.TIPUS_COMPETITION,
        )
        self.break_f = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="09:30",
            hora_fi="09:40",
            ordre=2,
            ordre_visual=2,
            titol="Descans global",
            tipus=RotacioFranja.TIPUS_BREAK,
        )
        self.f2 = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="09:40",
            hora_fi="10:10",
            ordre=3,
            ordre_visual=3,
            titol="Franja 2",
            tipus=RotacioFranja.TIPUS_COMPETITION,
        )
        self.f3 = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici="10:10",
            hora_fi="10:40",
            ordre=4,
            ordre_visual=4,
            titol="Franja 3",
            tipus=RotacioFranja.TIPUS_COMPETITION,
        )

    def _post_json(self, url, payload):
        return self.client.post(
            url,
            data=json.dumps(payload),
            content_type="application/json",
        )

    def test_visual_reorder_of_global_franja_changes_only_visual_order(self):
        url = reverse("rotacions_franges_reorder_visual", kwargs={"pk": self.comp.id})
        res = self._post_json(
            url,
            {
                "dragged_id": self.break_f.id,
                "target_id": self.f1.id,
                "position": "before",
            },
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            list(RotacioFranja.objects.filter(competicio=self.comp).order_by("ordre_visual", "id").values_list("titol", flat=True)),
            ["Descans global", "Franja 1", "Franja 2", "Franja 3"],
        )
        self.assertEqual(
            list(RotacioFranja.objects.filter(competicio=self.comp).order_by("ordre", "id").values_list("titol", flat=True)),
            ["Franja 1", "Descans global", "Franja 2", "Franja 3"],
        )

    def test_competitive_reorder_syncs_visual_order_around_global_rows(self):
        url = reverse("rotacions_franges_reorder", kwargs={"pk": self.comp.id})
        res = self._post_json(
            url,
            {
                "dragged_id": self.f3.id,
                "target_id": self.f2.id,
                "position": "before",
                "confirm_reorder": True,
            },
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            list(RotacioFranja.objects.filter(competicio=self.comp).order_by("ordre_visual", "id").values_list("titol", flat=True)),
            ["Franja 1", "Descans global", "Franja 3", "Franja 2"],
        )
        self.assertEqual(
            list(RotacioFranja.objects.filter(competicio=self.comp).order_by("ordre", "id").values_list("titol", flat=True)),
            ["Franja 1", "Descans global", "Franja 3", "Franja 2"],
        )

    def test_insert_after_places_new_competitive_immediately_after_origin_visually(self):
        url = reverse("rotacions_franja_insert_after", kwargs={"pk": self.comp.id, "franja_id": self.f1.id})
        res = self._post_json(
            url,
            {
                "titol": "Nova",
                "confirm_reorder": True,
            },
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            list(RotacioFranja.objects.filter(competicio=self.comp).order_by("ordre_visual", "id").values_list("titol", flat=True)),
            ["Franja 1", "Nova", "Descans global", "Franja 2", "Franja 3"],
        )

    def test_create_and_update_franja_color_persist_and_allow_reset(self):
        create_url = reverse("rotacions_franja_create", kwargs={"pk": self.comp.id})
        create_res = self._post_json(
            create_url,
            {
                "hora_inici": "11:00",
                "hora_fi": "11:20",
                "titol": "Premis color",
                "tipus": RotacioFranja.TIPUS_AWARDS,
                "color_fons": "#123456",
            },
        )
        self.assertEqual(create_res.status_code, 200)
        created = RotacioFranja.objects.get(pk=create_res.json()["id"])
        self.assertEqual(created.color_fons, "#123456")

        update_url = reverse("rotacions_franja_update_inline", kwargs={"pk": self.comp.id, "franja_id": created.id})
        update_res = self._post_json(
            update_url,
            {
                "hora_inici": "11:00",
                "hora_fi": "11:20",
                "titol": "Premis color",
                "tipus": RotacioFranja.TIPUS_AWARDS,
                "color_fons": "",
            },
        )
        self.assertEqual(update_res.status_code, 200)
        created.refresh_from_db()
        self.assertEqual(created.color_fons, "")
        self.assertEqual(created.resolved_background_color, RotacioFranja.DEFAULT_BACKGROUND_COLORS[RotacioFranja.TIPUS_AWARDS])

    def test_export_uses_visual_order_and_row_fill_color(self):
        self.break_f.ordre_visual = 1
        self.break_f.color_fons = "#123456"
        self.f1.ordre_visual = 2
        self.f2.ordre_visual = 3
        self.f3.ordre_visual = 4
        RotacioFranja.objects.bulk_update(
            [self.break_f, self.f1, self.f2, self.f3],
            ["ordre_visual", "color_fons"],
        )

        url = reverse("rotacions_franges_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url, {"mode": "groups"})
        self.assertEqual(res.status_code, 200)

        wb = load_workbook(filename=BytesIO(res.content))
        ws = wb.active
        franja_rows = []
        for row_idx in range(1, ws.max_row + 1):
            value = str(ws.cell(row=row_idx, column=1).value or "")
            if "Descans global" in value or "Franja 1" in value:
                franja_rows.append((row_idx, value))
        self.assertGreaterEqual(len(franja_rows), 2)
        self.assertIn("Descans global", franja_rows[0][1])
        self.assertIn("Franja 1", franja_rows[1][1])
        self.assertTrue(str(ws.cell(row=franja_rows[0][0], column=1).fill.fgColor.rgb or "").upper().endswith("123456"))

    def test_planner_context_orders_franges_by_ordre_visual(self):
        self.break_f.ordre_visual = 1
        self.f1.ordre_visual = 2
        self.f2.ordre_visual = 3
        self.f3.ordre_visual = 4
        RotacioFranja.objects.bulk_update(
            [self.break_f, self.f1, self.f2, self.f3],
            ["ordre_visual"],
        )

        url = reverse("rotacions_planner", kwargs={"pk": self.comp.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            [fr.id for fr in res.context["franges"]],
            [self.break_f.id, self.f1.id, self.f2.id, self.f3.id],
        )


