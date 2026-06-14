import json
import re
from io import BytesIO, StringIO
from datetime import date, timedelta
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.sessions.backends.db import SessionStore
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Max
from django.test import RequestFactory, TestCase
from django.urls import Resolver404, resolve, reverse
from django.utils import timezone
from openpyxl import load_workbook

from ceeb_web.auth_groups import GLOBAL_AUTH_GROUPS

from ... import live_cache
from ...access import user_has_competicio_capability
from ...forms import CompeticioAparellForm
from ...models import (
    Competicio,
    Equip,
    EquipContext,
    GrupCompeticio,
    Inscripcio,
    InscripcioEquipAssignacio,
    InscripcioMedia,
)
from ...models.judging import (
    JudgeConversation,
    JudgeConversationMessage,
    JudgeDeviceToken,
    PublicLiveToken,
)
from ...models.classificacions import ClassificacioConfig, ClassificacioTemplateGlobal
from ...models.rotacions import (
    RotacioAssignacio,
    RotacioAssignacioGrup,
    RotacioAssignacioSerieEquip,
    RotacioEstacio,
    RotacioFranja,
)
from ...models.scoring import (
    ScoringSchema,
    ScoreEntry,
    ScoreEntryVideo,
    ScoreEntryVideoEvent,
    SerieEquip,
    SerieEquipItem,
    TeamScoreEntry,
    TeamCompetitiveSubject,
    TeamScoreEntryVideo,
    TeamScoreEntryVideoEvent,
)
from ...models.competicio import (
    Aparell,
    CompeticioAparell,
    CompeticioAparellEquipContextSource,
    InscripcioAparellExclusio,
    InscripcioBaixa,
)
from ...models import CompeticioMembership
from ...scoring_engine import ScoringEngine
from ...services.inscripcions.groups import renumber_groups_for_competicio
from ...services.inscripcions.sorting import (
    _split_custom_sort_tokens,
    sort_records_by_field_stable,
)
from ...services.inscripcions.history import (
    apply_inscripcions_history_snapshot,
    capture_inscripcions_history_snapshot,
)
from ...services.inscripcions.admission import load_excluded_app_ids_by_inscripcio
from ...services.scoring.scoring_subjects import inscripcio_exclosa_en_aparell
from ...services.inscripcions.queries import (
    COLUMN_FILTER_EMPTY_TOKEN,
    _build_inscripcions_filtered_qs,
    build_inscripcions_sort_context_key,
    get_competicio_custom_sort_rank_map,
)
from ...services.classificacions.builder import scoreable_codes_by_app_id as _scoreable_codes_by_app_id
from ...services.classificacions.compute import DEFAULT_SCHEMA, compute_classificacio
from ...services.classificacions.export import _normalize_excel_cell
from ...services.classificacions.partitions import normalize_schema_legacy_team_birth_partition
from ...services.classificacions.validation import (
    build_metric_meta_for_comp_aparell as _build_metric_meta_for_comp_aparell,
    build_scoreable_meta_for_schema as _build_scoreable_meta_for_schema,
    validate_particions_schema as _validate_particions_schema,
    validate_schema_for_competicio as _validate_schema_for_competicio,
)
from ...services.classificacions.classificacio_templates import (
    normalize_particions_schema as _normalize_particions_schema,
    schema_to_template_schema as _schema_to_template_schema,
    template_schema_to_competicio_schema as _template_schema_to_competicio_schema,
)
from ...views.classificacions.builder import ClassificacionsHome
from ...services.shared.competition_groups import (
    assign_groups_by_display_num,
    compact_competition_order_for_group,
    ensure_group_for_display_num,
    get_group_maps,
    get_group_participant_counts,
    get_out_of_program_group_ids,
    get_programmed_group_ids,
    group_label,
    move_inscripcio_to_group,
    next_group_display_num,
)
from ...services.scoring.team_scoring import (
    build_permission_label,
    build_team_subjects_for_comp_aparell,
    resolve_permission_runtime_entries,
    runtime_schema_for_comp_aparell,
)
from ...services.teams.team_series import safe_deactivate_empty_serie
from ...views.judge.admin import _member_slot_choices, _validate_permission_row
from ...templatetags.competicio_extras import (
    DEFAULT_COMPETITION_BACKGROUND,
    get_competicio_background_url_from_request,
)

from ..base import _BaseTrampoliDataMixin



class InscripcioAparellExclusioModelTests(_BaseTrampoliDataMixin, TestCase):
    def test_clean_rejects_cross_competition_pair(self):
        comp_a = self._create_competicio("Comp A")
        comp_b = self._create_competicio("Comp B")
        ins = self._create_inscripcio(comp_a, "Participant A")

        app_b = self._create_aparell("DMT_B", "DMT B")
        comp_app_b = self._create_comp_aparell(comp_b, app_b)

        ex = InscripcioAparellExclusio(inscripcio=ins, comp_aparell=comp_app_b)
        with self.assertRaises(ValidationError):
            ex.full_clean()


class InscripcioBaixaTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio()
        self.user = self._login_competicio_user(
            self.comp,
            role=CompeticioMembership.Role.EDITOR,
            username_prefix="baixes_editor",
        )
        app1 = self._create_aparell("TRAMP_BAIXA", "Tramp Baixa")
        app2 = self._create_aparell("DMT_BAIXA", "DMT Baixa")
        self.comp_app_1 = self._create_comp_aparell(self.comp, app1, ordre=1, actiu=True)
        self.comp_app_2 = self._create_comp_aparell(self.comp, app2, ordre=2, actiu=True)
        self.ins = self._create_inscripcio(self.comp, "Gimnasta baixa")

    def test_global_baixa_blocks_all_requested_apps(self):
        InscripcioBaixa.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            motiu="Lesio",
            marcada_per=self.user,
        )

        excluded = load_excluded_app_ids_by_inscripcio(
            self.comp,
            [self.comp_app_1.id, self.comp_app_2.id],
        )

        self.assertEqual(excluded[self.ins.id], {self.comp_app_1.id, self.comp_app_2.id})
        self.assertTrue(inscripcio_exclosa_en_aparell(self.ins.id, self.comp_app_1.id))

    def test_app_baixa_blocks_only_that_app(self):
        InscripcioBaixa.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            comp_aparell=self.comp_app_2,
            motiu="No es presenta",
        )

        excluded = load_excluded_app_ids_by_inscripcio(
            self.comp,
            [self.comp_app_1.id, self.comp_app_2.id],
        )

        self.assertEqual(excluded[self.ins.id], {self.comp_app_2.id})
        self.assertFalse(inscripcio_exclosa_en_aparell(self.ins.id, self.comp_app_1.id))
        self.assertTrue(inscripcio_exclosa_en_aparell(self.ins.id, self.comp_app_2.id))

    def test_set_and_clear_baixa_endpoints_record_state(self):
        set_url = reverse("inscripcions_set_baixa", kwargs={"pk": self.comp.id})
        clear_url = reverse("inscripcions_clear_baixa", kwargs={"pk": self.comp.id})

        response = self.client.post(
            set_url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins.id,
                    "scope": "apps",
                    "comp_aparell_ids": [self.comp_app_1.id],
                    "motiu": "Malaltia",
                    "notes": "Avisat per l'entitat",
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            InscripcioBaixa.objects.filter(
                competicio=self.comp,
                inscripcio=self.ins,
                comp_aparell=self.comp_app_1,
                anul_lada_at__isnull=True,
            ).exists()
        )

        response = self.client.post(
            clear_url,
            data=json.dumps({"inscripcio_id": self.ins.id}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            InscripcioBaixa.objects.filter(
                competicio=self.comp,
                inscripcio=self.ins,
                anul_lada_at__isnull=True,
            ).exists()
        )

    def test_baixes_export_contains_active_baixa(self):
        InscripcioBaixa.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            comp_aparell=self.comp_app_1,
            motiu="Lesio",
            notes="No surt a competir",
            marcada_per=self.user,
        )

        response = self.client.get(reverse("inscripcions_baixes_export", kwargs={"pk": self.comp.id}))

        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        values = [cell.value for cell in ws[2]]
        self.assertIn("Gimnasta baixa", values)
        self.assertIn("Lesio", values)
        self.assertIn("No surt a competir", values)


class InscripcionsSetAparellsViewTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio()
        self.user = self._login_competicio_user(
            self.comp,
            role=CompeticioMembership.Role.EDITOR,
            username_prefix="set_aparells_editor",
        )
        app1 = self._create_aparell("TRAMP_I", "Tramp I")
        app2 = self._create_aparell("TRAMP_II", "Tramp II")

        self.comp_app_1 = self._create_comp_aparell(self.comp, app1, ordre=1, actiu=True)
        self.comp_app_2 = self._create_comp_aparell(self.comp, app2, ordre=2, actiu=True)

        self.ins = self._create_inscripcio(self.comp, "Ginmasta 1")

    def test_table_fragment_shows_only_individual_apps_selected_by_default(self):
        team_app = self._create_aparell("TEAM_APP", "Equip App")
        team_app.competition_unit = Aparell.CompetitionUnit.TEAM
        team_app.save(update_fields=["competition_unit"])
        self._create_comp_aparell(self.comp, team_app, ordre=3, actiu=True)

        response = self.client.get(
            reverse("inscripcions_list", kwargs={"pk": self.comp.id}),
            {"__fragments": "table"},
        )

        self.assertEqual(response.status_code, 200)
        html = response.json()["fragments"]["table"]
        self.assertIn(f'<option value="{self.comp_app_1.id}" selected>', html)
        self.assertIn(f'<option value="{self.comp_app_2.id}" selected>', html)
        self.assertNotIn("Equip App", html)
        self.assertIn("Per defecte competeix a tots els aparells individuals. Desmarca nomes exclusions.", html)

    def test_set_aparells_creates_and_replaces_exclusions(self):
        url = reverse("inscripcions_set_aparells", kwargs={"pk": self.comp.id})

        r1 = self.client.post(
            url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins.id,
                    "selected_comp_aparell_ids": [self.comp_app_1.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(r1.status_code, 200)
        self.assertEqual(
            list(
                InscripcioAparellExclusio.objects.filter(inscripcio=self.ins)
                .values_list("comp_aparell_id", flat=True)
            ),
            [self.comp_app_2.id],
        )

        r2 = self.client.post(
            url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins.id,
                    "selected_comp_aparell_ids": [self.comp_app_1.id, self.comp_app_2.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(r2.status_code, 200)
        self.assertFalse(
            InscripcioAparellExclusio.objects.filter(inscripcio=self.ins).exists()
        )

    def test_set_aparells_rejects_ids_outside_competition(self):
        other_comp = self._create_competicio("Comp Altre")
        other_app = self._create_aparell("TUMB_X", "Tumbling X")
        other_comp_app = self._create_comp_aparell(other_comp, other_app, ordre=1, actiu=True)

        url = reverse("inscripcions_set_aparells", kwargs={"pk": self.comp.id})
        r = self.client.post(
            url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins.id,
                    "selected_comp_aparell_ids": [self.comp_app_1.id, other_comp_app.id],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 400)

    def test_set_aparells_rejects_team_apps_from_general_selector(self):
        team_app = self._create_aparell("TEAM_APP", "Equip App")
        team_app.competition_unit = Aparell.CompetitionUnit.TEAM
        team_app.save(update_fields=["competition_unit"])
        team_comp_app = self._create_comp_aparell(self.comp, team_app, ordre=3, actiu=True)

        url = reverse("inscripcions_set_aparells", kwargs={"pk": self.comp.id})
        r = self.client.post(
            url,
            data=json.dumps(
                {
                    "inscripcio_id": self.ins.id,
                    "selected_comp_aparell_ids": [self.comp_app_1.id, team_comp_app.id],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(r.status_code, 400)


class CompeticioAparellParticipationViewTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio()
        self.user = self._login_competicio_user(
            self.comp,
            role=CompeticioMembership.Role.EDITOR,
            username_prefix="participation_editor",
        )
        self.comp.inscripcions_schema = {
            "columns": [
                {"code": "modalitat", "label": "Modalitat", "kind": "extra"},
            ]
        }
        self.comp.save(update_fields=["inscripcions_schema"])
        app = self._create_aparell("DMT_PART", "DMT Participacio", owner=self.user)
        self.comp_app = self._create_comp_aparell(self.comp, app, ordre=1, actiu=True)
        self.ins_dmt = self._create_inscripcio(self.comp, "DMT 1", ordre=1)
        self.ins_dmt.extra = {"modalitat": "DMT"}
        self.ins_dmt.save(update_fields=["extra"])
        self.ins_tra = self._create_inscripcio(self.comp, "TRA 1", ordre=2)
        self.ins_tra.extra = {"modalitat": "TRA"}
        self.ins_tra.save(update_fields=["extra"])
        self.url = reverse(
            "trampoli_aparell_participation",
            kwargs={"pk": self.comp.id, "app_id": self.comp_app.id},
        )

    def _rule_payload(self, *, intent="preview", confirm=False, value="DMT"):
        return {
            "participation_mode": "include_matching",
            "filter_field": ["modalitat"],
            "filter_operator": ["is_any"],
            "filter_values": [value],
            "intent": intent,
            "confirm_participation_apply": "1" if confirm else "",
        }

    def test_preview_does_not_replace_exclusions(self):
        response = self.client.post(self.url, data=self._rule_payload(intent="preview"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Previsualitzacio")
        self.assertFalse(InscripcioAparellExclusio.objects.filter(comp_aparell=self.comp_app).exists())

    def test_apply_replaces_participation_from_zero_and_persists_rule(self):
        InscripcioAparellExclusio.objects.create(
            inscripcio=self.ins_dmt,
            comp_aparell=self.comp_app,
            motiu="Manual antic",
        )

        response = self.client.post(self.url, data=self._rule_payload(intent="apply", confirm=True))

        self.assertEqual(response.status_code, 302)
        excluded_ids = set(
            InscripcioAparellExclusio.objects.filter(comp_aparell=self.comp_app).values_list("inscripcio_id", flat=True)
        )
        self.assertEqual(excluded_ids, {self.ins_tra.id})
        self.comp_app.refresh_from_db()
        self.assertEqual(self.comp_app.participation_config["mode"], "include_matching")
        self.assertEqual(self.comp_app.participation_config["filters"][0]["field"], "modalitat")
        self.assertEqual(self.comp_app.participation_config["last_summary"]["included_count"], 1)
        self.assertEqual(self.comp_app.participation_config["last_summary"]["excluded_count"], 1)

    def test_apply_requires_confirmation(self):
        response = self.client.post(self.url, data=self._rule_payload(intent="apply", confirm=False))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Confirma la substitucio massiva")
        self.assertFalse(InscripcioAparellExclusio.objects.filter(comp_aparell=self.comp_app).exists())

    def test_planner_and_aparell_list_link_to_participation_panel(self):
        list_response = self.client.get(reverse("trampoli_aparells_list", kwargs={"pk": self.comp.id}))
        planner_response = self.client.get(reverse("trampoli_fases", kwargs={"pk": self.comp.id}))

        self.assertContains(list_response, "Participacio")
        self.assertContains(planner_response, "Participacio")


