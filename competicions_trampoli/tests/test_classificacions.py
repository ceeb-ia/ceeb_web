import json
import re
from io import BytesIO, StringIO
from datetime import date, timedelta
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.sessions.backends.db import SessionStore
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Max
from django.test import RequestFactory, TestCase
from django.urls import Resolver404, resolve, reverse
from django.utils import timezone
from openpyxl import load_workbook

from ceeb_web.auth_groups import GLOBAL_AUTH_GROUPS

from .. import live_cache
from ..access import user_has_competicio_capability
from ..forms import CompeticioAparellForm
from ..models import (
    Competicio,
    Equip,
    EquipContext,
    GrupCompeticio,
    Inscripcio,
    InscripcioEquipAssignacio,
    InscripcioMedia,
)
from ..models.judging import (
    JudgeConversation,
    JudgeConversationMessage,
    JudgeDeviceToken,
    PublicLiveToken,
)
from ..models.classificacions import ClassificacioConfig, ClassificacioTemplateGlobal
from ..models.rotacions import (
    RotacioAssignacio,
    RotacioAssignacioGrup,
    RotacioAssignacioSerieEquip,
    RotacioEstacio,
    RotacioFranja,
)
from ..models.scoring import (
    ScoringSchema,
    ScoreEntry,
    ScoreEntryVideo,
    ScoreEntryVideoEvent,
    SerieEquip,
    SerieEquipItem,
    TeamScoreEntry,
    TeamCompetitiveSubject,
    TeamScoreEntryVideo,
    TeamScoreEntryVideoEvent,
)
from ..models.competicio import (
    Aparell,
    CompeticioAparell,
    CompeticioAparellEquipContextSource,
    InscripcioAparellExclusio,
)
from ..models import CompeticioMembership
from ..scoring_engine import ScoringEngine
from ..services.inscripcions.groups import renumber_groups_for_competicio
from ..services.inscripcions.sorting import (
    _split_custom_sort_tokens,
    sort_records_by_field_stable,
)
from ..services.inscripcions.history import (
    apply_inscripcions_history_snapshot,
    capture_inscripcions_history_snapshot,
)
from ..services.inscripcions.queries import (
    COLUMN_FILTER_EMPTY_TOKEN,
    _build_inscripcions_filtered_qs,
    build_inscripcions_sort_context_key,
    get_competicio_custom_sort_rank_map,
)
from ..services.classificacions.classificacio_templates import (
    normalize_particions_schema as _normalize_particions_schema,
    schema_to_template_schema as _schema_to_template_schema,
    template_schema_to_competicio_schema as _template_schema_to_competicio_schema_service,
)
from ..services.classificacions.builder import (
    prepare_schema_for_builder_hydration,
    scoreable_codes_by_app_id as _scoreable_codes_by_app_id,
)
from ..services.classificacions.compute import DEFAULT_SCHEMA, compute_classificacio
from ..services.classificacions.export import _normalize_excel_cell
from ..services.classificacions.partitions import normalize_schema_legacy_team_birth_partition
from ..services.classificacions.validation import (
    build_metric_meta_for_comp_aparell as _build_metric_meta_for_comp_aparell,
    build_scoreable_meta_for_schema as _build_scoreable_meta_for_schema,
    validate_particions_schema as _validate_particions_schema,
    validate_schema_for_competicio as _validate_schema_for_competicio,
)
from ..views.classificacions.builder import ClassificacionsHome
from ..services.shared.competition_groups import (
    assign_groups_by_display_num,
    compact_competition_order_for_group,
    ensure_group_for_display_num,
    get_group_maps,
    get_group_participant_counts,
    get_out_of_program_group_ids,
    get_programmed_group_ids,
    group_label,
    move_inscripcio_to_group,
    next_group_display_num,
)
from ..services.scoring.team_scoring import (
    build_permission_label,
    build_team_subjects_for_comp_aparell,
    resolve_permission_runtime_entries,
    runtime_schema_for_comp_aparell,
)
from ..services.teams.team_series import safe_deactivate_empty_serie
from ..views.judge.admin import _member_slot_choices, _validate_permission_row
from ..templatetags.competicio_extras import (
    DEFAULT_COMPETITION_BACKGROUND,
    get_competicio_background_url_from_request,
)

from .base import _BaseTrampoliDataMixin


def _template_schema_to_competicio_schema(*args, **kwargs):
    schema_local, mapping_warnings, mapping, _compat_meta = _template_schema_to_competicio_schema_service(
        *args,
        **kwargs,
    )
    return schema_local, mapping_warnings, mapping


class ClassificacioMatrixScalarTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Classificacio")
        self.user = self._login_competicio_user(
            self.comp,
            role=CompeticioMembership.Role.EDITOR,
            username_prefix="classif_editor",
        )
        self.app_a = self._create_aparell("APP_A", "Aparell A")
        self.app_b = self._create_aparell("APP_B", "Aparell B")
        self.comp_app_a = self._create_comp_aparell(self.comp, self.app_a, ordre=1, actiu=True)
        self.comp_app_b = self._create_comp_aparell(self.comp, self.app_b, ordre=2, actiu=True)

        self.ins_a = self._create_inscripcio(self.comp, "Participant A", ordre=1)
        self.ins_b = self._create_inscripcio(self.comp, "Participant B", ordre=2)

    def _base_cfg_schema(self):
        return {
            "particions": [],
            "filtres": {},
            "puntuacio": {
                "aparells": {"mode": "tots", "ids": []},
                "camps_per_aparell": {
                    str(self.comp_app_a.id): ["X"],
                    str(self.comp_app_b.id): ["Y"],
                },
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "exercicis_best_n": 1,
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
                "camp": "total",
                "agregacio": "sum",
                "best_n": 1,
            },
            "desempat": [],
            "presentacio": {"top_n": 0, "mostrar_empats": True},
        }

    def _valid_partition_schema(self):
        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["total"]}
        return schema

    def _selected_total_schema(self, app_ids=None, mode_resultat="score"):
        selected = list(app_ids or [self.comp_app_a.id, self.comp_app_b.id])
        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": selected}
        schema["puntuacio"]["camps_per_aparell"] = {str(app_id): ["total"] for app_id in selected}
        schema["puntuacio"]["mode_resultat_aparells"] = mode_resultat
        schema["puntuacio"]["victories"] = {
            "punts_victoria": 1,
            "punts_empat": 0.5,
            "sense_nota_mode": "skip",
            "mode_camps": "agregat",
            "mode_exercicis": "agregat",
            "mode_seleccio_exercicis_camps_separats": "per_camp",
            "agregacio_victories_camps": "sum",
            "agregacio_victories_exercicis": "sum",
            "desempat_comparacio": [],
        }
        return schema

    def test_compute_classificacio_uses_input_matrix_1x1_as_scalar(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 1}}
                ],
                "computed": [],
            },
        )
        ScoringSchema.objects.create(
            aparell=self.app_b,
            schema={
                "fields": [
                    {"code": "Y", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 1}}
                ],
                "computed": [],
            },
        )

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"X": [[7.5]]},
            outputs={},
            total=0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_b,
            inputs={"Y": [[8.2]]},
            outputs={},
            total=0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Global A/B",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=self._base_cfg_schema(),
        )
        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        points_by_name = {r["participant"]: r["punts"] for r in rows}

        self.assertEqual(points_by_name.get("Participant A"), 7.5)
        self.assertEqual(points_by_name.get("Participant B"), 8.2)

    def test_compute_classificacio_keeps_zero_for_non_1x1_matrix(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 2}}
                ],
                "computed": [],
            },
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"X": [[7.0, 8.0]]},
            outputs={},
            total=0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["X"]}
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="No 1x1",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )
        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        points_by_name = {r["participant"]: r["punts"] for r in rows}
        self.assertEqual(points_by_name.get("Participant A"), 0.0)
        self.assertEqual(points_by_name.get("Participant B"), 0.0)

    def test_compute_classificacio_individual_candidate_source_modes_preserve_expected_scores(self):
        self.comp_app_a.nombre_exercicis = 3
        self.comp_app_a.save(update_fields=["nombre_exercicis"])

        for inscripcio, exercici, total in (
            (self.ins_a, 1, 1.0),
            (self.ins_a, 2, 10.0),
            (self.ins_a, 3, 2.0),
            (self.ins_b, 1, 4.0),
            (self.ins_b, 2, 9.0),
            (self.ins_b, 3, 1.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=inscripcio,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={},
                outputs={},
                total=total,
            )

        raw_schema = self._selected_total_schema([self.comp_app_a.id])
        raw_schema["puntuacio"].update(
            {
                "candidate_source_mode": "raw_exercise",
                "candidate_source_cfg": {
                    "mode": "tots",
                    "best_n": 1,
                    "index": 1,
                    "ids": [],
                    "agregacio_exercicis": "sum",
                },
                "exercicis": {"mode": "tots"},
                "agregacio_exercicis": "sum",
            }
        )
        raw_cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Individual raw source",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=raw_schema,
        )

        aggregate_schema = self._selected_total_schema([self.comp_app_a.id])
        aggregate_schema["puntuacio"].update(
            {
                "candidate_source_mode": "participant_aggregate",
                "candidate_source_cfg": {
                    "mode": "millor_n",
                    "best_n": 1,
                    "index": 1,
                    "ids": [],
                    "agregacio_exercicis": "sum",
                },
                "exercicis": {"mode": "millor_n", "best_n": 1},
                "agregacio_exercicis": "sum",
            }
        )
        aggregate_cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Individual aggregate source",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=aggregate_schema,
        )

        raw_rows = compute_classificacio(self.comp, raw_cfg).get("global", [])
        aggregate_rows = compute_classificacio(self.comp, aggregate_cfg).get("global", [])
        raw_scores = {row["participant"]: row["score"] for row in raw_rows}
        aggregate_scores = {row["participant"]: row["score"] for row in aggregate_rows}

        self.assertEqual(raw_scores.get("Participant A"), 13.0)
        self.assertEqual(raw_scores.get("Participant B"), 14.0)
        self.assertEqual(aggregate_scores.get("Participant A"), 10.0)
        self.assertEqual(aggregate_scores.get("Participant B"), 9.0)

    def test_classificacio_save_rejects_non_1x1_matrix_field(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 2}, "items": {"count": 1}}
                ],
                "computed": [],
            },
        )

        payload = {
            "nom": "Cfg invalida",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["X"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("no es puntuable directament" in e for e in body.get("errors", [])))

    def test_compute_classificacio_uses_input_list_1_as_scalar(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "L", "type": "list", "shape": "judge", "judges": {"count": 1}}
                ],
                "computed": [],
            },
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"L": [7.2]},
            outputs={},
            total=0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["L"]}
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Llista 1x1",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )
        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        points_by_name = {r["participant"]: r["punts"] for r in rows}
        self.assertEqual(points_by_name.get("Participant A"), 7.2)
        self.assertEqual(points_by_name.get("Participant B"), 0.0)

    def test_classificacio_save_accepts_non_scalar_computed_main_field_under_current_validator(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 2}}
                ],
                "computed": [
                    {"code": "X_copy", "label": "X copy", "formula": "X"},
                ],
            },
        )

        payload = {
            "nom": "Cfg computed no escalar",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["X_copy"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=res.json()["id"])
        self.assertEqual(
            (((cfg.schema.get("puntuacio") or {}).get("camps_per_aparell")) or {}).get(str(self.comp_app_a.id)),
            ["X_copy"],
        )

    def test_classificacio_save_accepts_non_scalar_computed_tie_field_under_current_validator(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "X", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 2}}
                ],
                "computed": [
                    {"code": "X_copy", "label": "X copy", "formula": "X"},
                ],
            },
        )

        payload = {
            "nom": "Cfg tie computed no escalar",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["X_copy"],
                        "ordre": "desc",
                        "scope": {
                            "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                            "exercicis": {"mode": "hereta"},
                        },
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=res.json()["id"])
        self.assertEqual((((cfg.schema.get("desempat") or [])[0]) or {}).get("camps"), ["X_copy"])

    def test_classificacio_save_requires_real_camps_for_selected_apps(self):
        payload = {
            "nom": "Cfg sense camps reals",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("camps_per_aparell" in e and "camp real" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_invalid_raw_column_field(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "E_total", "label": "Execucio", "type": "number"},
                ],
                "computed": [],
            },
        )

        payload = {
            "nom": "Cfg raw invalid",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["E_total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {
                    "top_n": 0,
                    "mostrar_empats": True,
                    "columnes": [
                        {
                            "type": "raw",
                            "key": "raw_exec",
                            "label": "Exec",
                            "align": "right",
                            "decimals": 3,
                            "source": {
                                "aparell_id": self.comp_app_a.id,
                                "exercici": 1,
                                "camp": "EX",
                                "jutges": {"ids": []},
                            },
                        }
                    ],
                },
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("presentacio.columnes[0]" in e and "'EX'" in e for e in body.get("errors", [])))

    def test_classificacio_save_keeps_tie_schema_canonical_without_legacy_camp(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "E_total", "label": "Execucio", "type": "number"},
                ],
                "computed": [],
            },
        )

        payload = {
            "nom": "Cfg canonical tie",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["E_total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["E_total"],
                        "agregacio_camps": "hereta",
                        "ordre": "desc",
                        "scope": {
                            "aparells": {"mode": "hereta"},
                            "exercicis": {"mode": "hereta"},
                        },
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=res.json()["id"])
        self.assertEqual((cfg.schema.get("puntuacio") or {}).get("camp"), "total")
        self.assertEqual((cfg.schema.get("puntuacio") or {}).get("agregacio"), "sum")
        self.assertEqual((cfg.schema.get("puntuacio") or {}).get("best_n"), 1)
        self.assertEqual((cfg.schema.get("desempat") or [])[0].get("camps"), ["E_total"])
        self.assertNotIn("camp", (cfg.schema.get("desempat") or [])[0])

    def test_classificacio_save_persists_tie_pipeline_canonical_format(self):
        ScoringSchema.objects.create(
            aparell=self.app_a,
            schema={
                "fields": [
                    {"code": "E_total", "label": "Execucio", "type": "number"},
                ],
                "computed": [],
            },
        )
        app_key = str(self.comp_app_a.id)
        payload = {
            "nom": "Cfg tie pipeline",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {app_key: ["E_total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "id": "tie_exec",
                        "nom": "Millor execucio",
                        "ordre": "desc",
                        "pipeline_version": 1,
                        "pipeline": {
                            "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                            "camps_per_aparell": {app_key: ["E_total"]},
                            "agregacio_camps_per_aparell": {app_key: "sum"},
                            "agregacio_camps": "sum",
                            "candidate_source_mode": "raw_exercise",
                            "candidate_source_cfg": {
                                "mode": "tots",
                                "best_n": 1,
                                "index": 1,
                                "ids": [],
                                "agregacio_exercicis": "sum",
                            },
                            "candidate_source_per_aparell": {app_key: {"mode": "raw_exercise"}},
                            "exercicis": {"mode": "tots", "index": 1, "ids": [], "max_per_participant": 0},
                            "exercise_selection_scope": "per_member",
                            "mode_seleccio_exercicis": "per_aparell_global",
                            "exercicis_per_aparell": {},
                            "agregacio_exercicis": "sum",
                            "agregacio_aparells": "sum",
                            "mode_resultat_aparells": "score",
                            "ordre": "desc",
                        },
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        res = self.client.post(
            reverse("classificacio_save", kwargs={"pk": self.comp.id}),
            data=json.dumps(payload),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=res.json()["id"])
        tie = (cfg.schema.get("desempat") or [])[0]
        self.assertEqual(tie.get("id"), "tie_exec")
        self.assertEqual(tie.get("nom"), "Millor execucio")
        self.assertEqual(tie.get("ordre"), "desc")
        self.assertEqual(tie.get("pipeline_version"), 1)
        self.assertEqual((((tie.get("pipeline") or {}).get("aparells") or {}).get("ids")), [self.comp_app_a.id])
        self.assertEqual((((tie.get("pipeline") or {}).get("camps_per_aparell") or {}).get(app_key)), ["E_total"])

    def test_classificacio_save_rejects_forbidden_tie_pipeline_keys(self):
        payload = {
            "nom": "Cfg tie pipeline invalid",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "id": "tie_invalid",
                        "ordre": "desc",
                        "pipeline_version": 1,
                        "pipeline": {
                            "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                            "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                            "agregacio_camps_per_aparell": {str(self.comp_app_a.id): "sum"},
                            "agregacio_camps": "sum",
                            "exercicis": {"mode": "tots"},
                            "exercise_selection_scope": "per_member",
                            "mode_seleccio_exercicis": "per_aparell_global",
                            "agregacio_exercicis": "sum",
                            "agregacio_aparells": "sum",
                            "mode_resultat_aparells": "score",
                            "victories": {},
                        },
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        res = self.client.post(
            reverse("classificacio_save", kwargs={"pk": self.comp.id}),
            data=json.dumps(payload),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertTrue(any("desempat[0].pipeline.victories" in err for err in body.get("errors", [])))

    def test_prepare_schema_for_builder_hydration_materializes_legacy_tie_pipeline(self):
        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["total"]}
        schema["desempat"] = [
            {
                "camp": "total",
                "camps": ["total"],
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "exercicis": {"mode": "tots"},
                },
            }
        ]
        hydrated = prepare_schema_for_builder_hydration(self.comp, schema, tipus="individual")
        tie = (hydrated.get("desempat") or [])[0]
        self.assertEqual(tie.get("id"), "tie_1")
        self.assertEqual(tie.get("pipeline_version"), 1)
        self.assertEqual((((tie.get("pipeline") or {}).get("aparells") or {}).get("ids")), [self.comp_app_a.id])
        self.assertEqual((((tie.get("pipeline") or {}).get("camps_per_aparell") or {}).get(str(self.comp_app_a.id))), ["total"])

    def test_compute_classificacio_pipeline_tie_supports_per_app_override_exercicis(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        self.comp_app_b.nombre_exercicis = 2
        self.comp_app_b.save(update_fields=["nombre_exercicis"])

        for inscripcio, comp_aparell, exercici, total in (
            (self.ins_a, self.comp_app_a, 1, 9.0),
            (self.ins_a, self.comp_app_a, 2, 1.0),
            (self.ins_a, self.comp_app_b, 1, 5.0),
            (self.ins_a, self.comp_app_b, 2, 5.0),
            (self.ins_b, self.comp_app_a, 1, 6.0),
            (self.ins_b, self.comp_app_a, 2, 4.0),
            (self.ins_b, self.comp_app_b, 1, 8.0),
            (self.ins_b, self.comp_app_b, 2, 2.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=inscripcio,
                exercici=exercici,
                comp_aparell=comp_aparell,
                inputs={},
                outputs={},
                total=total,
            )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id, self.comp_app_b.id]}
        schema["puntuacio"]["camps_per_aparell"] = {
            str(self.comp_app_a.id): ["total"],
            str(self.comp_app_b.id): ["total"],
        }
        schema["puntuacio"]["agregacio_camps_per_aparell"] = {
            str(self.comp_app_a.id): "sum",
            str(self.comp_app_b.id): "sum",
        }
        schema["puntuacio"]["candidate_source_per_aparell"] = {
            str(self.comp_app_a.id): {"mode": "raw_exercise"},
            str(self.comp_app_b.id): {"mode": "raw_exercise"},
        }
        schema["puntuacio"]["exercicis"] = {"mode": "tots"}
        schema["desempat"] = [
            {
                "id": "tie_pipeline_override",
                "ordre": "desc",
                "pipeline_version": 1,
                "pipeline": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id, self.comp_app_b.id]},
                    "camps_per_aparell": {
                        str(self.comp_app_a.id): ["total"],
                        str(self.comp_app_b.id): ["total"],
                    },
                    "agregacio_camps_per_aparell": {
                        str(self.comp_app_a.id): "sum",
                        str(self.comp_app_b.id): "sum",
                    },
                    "agregacio_camps": "sum",
                    "candidate_source_mode": "raw_exercise",
                    "candidate_source_cfg": {
                        "mode": "tots",
                        "best_n": 1,
                        "index": 1,
                        "ids": [],
                        "agregacio_exercicis": "sum",
                    },
                    "candidate_source_per_aparell": {
                        str(self.comp_app_a.id): {"mode": "raw_exercise"},
                        str(self.comp_app_b.id): {"mode": "raw_exercise"},
                    },
                    "exercicis": {"mode": "tots", "index": 1, "ids": [], "max_per_participant": 0},
                    "exercise_selection_scope": "per_member",
                    "mode_seleccio_exercicis": "per_aparell_override",
                    "exercicis_per_aparell": {
                        str(self.comp_app_a.id): {"mode": "millor_1", "index": 1, "ids": [], "max_per_participant": 0},
                        str(self.comp_app_b.id): {"mode": "pitjor_1", "index": 1, "ids": [], "max_per_participant": 0},
                    },
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "mode_resultat_aparells": "score",
                    "ordre": "desc",
                },
            }
        ]

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Tie pipeline per app override",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )
        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([row["participant"] for row in rows], ["Participant A", "Participant B"])

    def test_compute_classificacio_main_score_supports_per_app_exercise_aggregation(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        self.comp_app_b.nombre_exercicis = 2
        self.comp_app_b.save(update_fields=["nombre_exercicis"])

        for inscripcio, comp_aparell, exercici, total in (
            (self.ins_a, self.comp_app_a, 1, 9.0),
            (self.ins_a, self.comp_app_a, 2, 1.0),
            (self.ins_a, self.comp_app_b, 1, 5.0),
            (self.ins_a, self.comp_app_b, 2, 5.0),
            (self.ins_b, self.comp_app_a, 1, 6.0),
            (self.ins_b, self.comp_app_a, 2, 6.0),
            (self.ins_b, self.comp_app_b, 1, 8.0),
            (self.ins_b, self.comp_app_b, 2, 2.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=inscripcio,
                exercici=exercici,
                comp_aparell=comp_aparell,
                inputs={},
                outputs={},
                total=total,
            )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id, self.comp_app_b.id]}
        schema["puntuacio"]["camps_per_aparell"] = {
            str(self.comp_app_a.id): ["total"],
            str(self.comp_app_b.id): ["total"],
        }
        schema["puntuacio"]["agregacio_camps_per_aparell"] = {
            str(self.comp_app_a.id): "sum",
            str(self.comp_app_b.id): "sum",
        }
        schema["puntuacio"]["candidate_source_per_aparell"] = {
            str(self.comp_app_a.id): {"mode": "raw_exercise"},
            str(self.comp_app_b.id): {"mode": "raw_exercise"},
        }
        schema["puntuacio"]["exercicis"] = {"mode": "tots"}
        schema["puntuacio"]["mode_seleccio_exercicis"] = "per_aparell_override"
        schema["puntuacio"]["exercicis_per_aparell"] = {
            str(self.comp_app_a.id): {"mode": "tots"},
            str(self.comp_app_b.id): {"mode": "tots"},
        }
        schema["puntuacio"]["agregacio_exercicis_per_aparell"] = {
            str(self.comp_app_a.id): "max",
            str(self.comp_app_b.id): "sum",
        }
        schema["puntuacio"]["agregacio_exercicis"] = "sum"
        schema["puntuacio"]["agregacio_aparells"] = "sum"

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Puntuacio agg ex per app",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )
        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([row["participant"] for row in rows], ["Participant A", "Participant B"])
        self.assertEqual((rows[0].get("by_app") or {}).get(self.comp_app_a.id), 9.0)
        self.assertEqual((rows[0].get("by_app") or {}).get(self.comp_app_b.id), 10.0)
        self.assertEqual(rows[0].get("score"), 19.0)

    def test_prepare_schema_for_builder_hydration_preserves_per_app_exercise_aggregation_maps(self):
        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id, self.comp_app_b.id]}
        schema["puntuacio"]["camps_per_aparell"] = {
            str(self.comp_app_a.id): ["total"],
            str(self.comp_app_b.id): ["total"],
        }
        schema["puntuacio"]["exercicis_per_aparell"] = {
            str(self.comp_app_a.id): {"mode": "millor_1"},
            str(self.comp_app_b.id): {"mode": "tots"},
        }
        schema["puntuacio"]["agregacio_exercicis_per_aparell"] = {
            str(self.comp_app_a.id): "max",
            str(self.comp_app_b.id): "sum",
        }
        schema["desempat"] = [
            {
                "camp": "total",
                "camps": ["total"],
                "ordre": "desc",
                "mode_seleccio_exercicis": "per_aparell_override",
                "exercicis_per_aparell": {
                    str(self.comp_app_a.id): {"mode": "millor_1"},
                    str(self.comp_app_b.id): {"mode": "pitjor_1"},
                },
                "agregacio_exercicis_per_aparell": {
                    str(self.comp_app_a.id): "max",
                    str(self.comp_app_b.id): "min",
                },
                "scope": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id, self.comp_app_b.id]},
                    "exercicis": {"mode": "tots"},
                },
            }
        ]
        schema["puntuacio"]["victories"] = {
            "punts_victoria": 1,
            "punts_empat": 0.5,
            "sense_nota_mode": "skip",
            "mode_camps": "agregat",
            "mode_exercicis": "agregat",
            "mode_seleccio_exercicis_camps_separats": "per_camp",
            "agregacio_victories_camps": "sum",
            "agregacio_victories_exercicis": "sum",
            "desempat_comparacio": [
                {
                    "camp": "total",
                    "camps": ["total"],
                    "ordre": "desc",
                    "mode_seleccio_exercicis": "per_aparell_override",
                    "exercicis_per_aparell": {
                        str(self.comp_app_a.id): {"mode": "millor_1"},
                        str(self.comp_app_b.id): {"mode": "tots"},
                    },
                    "agregacio_exercicis_per_aparell": {
                        str(self.comp_app_a.id): "max",
                        str(self.comp_app_b.id): "sum",
                    },
                    "scope": {"exercicis": {"mode": "tots"}},
                }
            ],
        }

        hydrated = prepare_schema_for_builder_hydration(self.comp, schema, tipus="individual")
        punt = hydrated.get("puntuacio") or {}
        self.assertEqual(
            punt.get("agregacio_exercicis_per_aparell"),
            {
                str(self.comp_app_a.id): "max",
                str(self.comp_app_b.id): "sum",
            },
        )
        tie = (hydrated.get("desempat") or [])[0] or {}
        self.assertEqual(
            tie.get("agregacio_exercicis_per_aparell"),
            {
                str(self.comp_app_a.id): "max",
                str(self.comp_app_b.id): "min",
            },
        )
        self.assertEqual(
            (((tie.get("pipeline") or {}).get("agregacio_exercicis_per_aparell")) or {}),
            {
                str(self.comp_app_a.id): "max",
                str(self.comp_app_b.id): "min",
            },
        )
        victory_tie = ((((punt.get("victories") or {}).get("desempat_comparacio")) or [])[0] or {})
        self.assertEqual(
            victory_tie.get("agregacio_exercicis_per_aparell"),
            {
                str(self.comp_app_a.id): "max",
                str(self.comp_app_b.id): "sum",
            },
        )

    def test_classificacio_save_returns_persisted_team_schema_with_mode_resolution(self):
        self._ensure_native_equip_context(self.comp)
        payload = {
            "nom": "Cfg equips persisted",
            "activa": True,
            "ordre": 1,
            "tipus": "equips",
            "schema": {
                **self._selected_total_schema([self.comp_app_a.id]),
                "equips": {
                    "context_code": "native",
                    "assignment_source": {"mode": "context", "context_code": "native", "fallback": "native"},
                    "team_mode": "derived_from_individual",
                },
            },
        }

        res = self.client.post(
            reverse("classificacio_save", kwargs={"pk": self.comp.id}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(res.status_code, 200)
        body = res.json()
        cfg_payload = body.get("cfg") or {}
        self.assertEqual(cfg_payload.get("tipus"), "equips")
        mode_resolution = (((cfg_payload.get("schema") or {}).get("equips") or {}).get("mode_resolution")) or {}
        self.assertTrue(mode_resolution.get("resolved_at"))
        self.assertIn("eligible_team_app_ids_at_save", mode_resolution)

        cfg = ClassificacioConfig.objects.get(pk=body["id"])
        saved_resolution = (((cfg.schema.get("equips") or {}).get("mode_resolution")) or {})
        self.assertEqual(saved_resolution.get("resolved_at"), mode_resolution.get("resolved_at"))

    def test_classificacio_save_roundtrip_preserves_legacy_global_pool_and_member_cap(self):
        self._ensure_native_equip_context(self.comp)
        payload = {
            "nom": "Cfg equips legacy exercise selection",
            "activa": True,
            "ordre": 1,
            "tipus": "equips",
            "schema": {
                **self._selected_total_schema([self.comp_app_a.id, self.comp_app_b.id]),
                "puntuacio": {
                    **self._selected_total_schema([self.comp_app_a.id, self.comp_app_b.id])["puntuacio"],
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id, self.comp_app_b.id]},
                    "camps_per_aparell": {
                        str(self.comp_app_a.id): ["total"],
                        str(self.comp_app_b.id): ["total"],
                    },
                    "exercicis": {"mode": "millor_n", "best_n": 4, "max_per_participant": 1},
                    "exercise_selection_scope": "team_pool",
                    "mode_seleccio_exercicis": "per_aparell_global",
                    "exercicis_per_aparell": {
                        str(self.comp_app_a.id): {"mode": "millor_n", "best_n": 4, "max_per_participant": 1},
                        str(self.comp_app_b.id): {"mode": "millor_n", "best_n": 3, "max_per_participant": 1},
                    },
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "native",
                    "assignment_source": {"mode": "context", "context_code": "native", "fallback": "native"},
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        }

        res = self.client.post(
            reverse("classificacio_save", kwargs={"pk": self.comp.id}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(res.status_code, 200)
        body = res.json()
        cfg = ClassificacioConfig.objects.get(pk=body["id"])
        punt = cfg.schema.get("puntuacio") or {}
        self.assertEqual(punt.get("mode_seleccio_exercicis"), "per_aparell_global")
        self.assertEqual(punt.get("exercise_selection_scope"), "team_pool")
        self.assertEqual((punt.get("exercicis") or {}).get("max_per_participant"), 1)
        self.assertEqual(sorted((punt.get("exercicis_per_aparell") or {}).keys()), [str(self.comp_app_a.id), str(self.comp_app_b.id)])

        hydrated = prepare_schema_for_builder_hydration(self.comp, cfg.schema or {}, tipus="equips")
        hydrated_punt = hydrated.get("puntuacio") or {}
        self.assertEqual(hydrated_punt.get("mode_seleccio_exercicis"), "per_aparell_global")
        self.assertEqual(hydrated_punt.get("exercise_selection_scope"), "team_pool")
        self.assertEqual((hydrated_punt.get("exercicis") or {}).get("max_per_participant"), 1)
        self.assertEqual((hydrated_punt.get("aparells") or {}).get("ids"), [self.comp_app_a.id, self.comp_app_b.id])

    def test_classificacio_save_roundtrip_preserves_candidate_source_contract(self):
        payload_schema = self._selected_total_schema([self.comp_app_a.id, self.comp_app_b.id])
        payload_schema["puntuacio"] = {
            **payload_schema["puntuacio"],
            "candidate_source_mode": "participant_aggregate",
            "candidate_source_cfg": {
                "mode": "millor_n",
                "best_n": 2,
                "index": 1,
                "ids": [],
                "agregacio_exercicis": "sum",
            },
        }
        payload = {
            "nom": "Cfg candidate source",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": payload_schema,
        }

        res = self.client.post(
            reverse("classificacio_save", kwargs={"pk": self.comp.id}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(res.status_code, 200)
        body = res.json()
        cfg = ClassificacioConfig.objects.get(pk=body["id"])
        punt = cfg.schema.get("puntuacio") or {}
        self.assertEqual(punt.get("candidate_source_mode"), "participant_aggregate")
        self.assertEqual((punt.get("candidate_source_cfg") or {}).get("mode"), "millor_n")
        self.assertEqual((punt.get("candidate_source_cfg") or {}).get("best_n"), 2)

        hydrated = prepare_schema_for_builder_hydration(self.comp, cfg.schema or {}, tipus="individual")
        hydrated_punt = hydrated.get("puntuacio") or {}
        self.assertEqual(hydrated_punt.get("candidate_source_mode"), "participant_aggregate")
        self.assertEqual((hydrated_punt.get("candidate_source_cfg") or {}).get("mode"), "millor_n")
        self.assertEqual((hydrated_punt.get("candidate_source_cfg") or {}).get("best_n"), 2)

    def test_classificacio_save_roundtrip_applies_candidate_source_per_app_fallback_cfg(self):
        payload_schema = self._selected_total_schema([self.comp_app_a.id, self.comp_app_b.id])
        payload_schema["puntuacio"] = {
            **payload_schema["puntuacio"],
            "candidate_source_mode": "participant_aggregate",
            "candidate_source_cfg": {
                "mode": "millor_n",
                "best_n": 2,
                "index": 3,
                "ids": [1, 2],
                "agregacio_exercicis": "avg",
            },
            "candidate_source_per_aparell": {
                str(self.comp_app_a.id): {
                    "mode": "participant_aggregate",
                },
                str(self.comp_app_b.id): {
                    "mode": "participant_aggregate",
                    "cfg": {
                        "mode": "index",
                    },
                },
            },
        }
        payload = {
            "nom": "Cfg candidate source fallback",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": payload_schema,
        }

        res = self.client.post(
            reverse("classificacio_save", kwargs={"pk": self.comp.id}),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(res.status_code, 200)
        body = res.json()
        cfg = ClassificacioConfig.objects.get(pk=body["id"])
        punt = cfg.schema.get("puntuacio") or {}
        per_app = punt.get("candidate_source_per_aparell") or {}

        self.assertEqual((per_app.get(str(self.comp_app_a.id)) or {}).get("mode"), "participant_aggregate")
        self.assertEqual(((per_app.get(str(self.comp_app_a.id)) or {}).get("cfg") or {}).get("mode"), "millor_n")
        self.assertEqual(((per_app.get(str(self.comp_app_a.id)) or {}).get("cfg") or {}).get("best_n"), 2)
        self.assertEqual(((per_app.get(str(self.comp_app_a.id)) or {}).get("cfg") or {}).get("index"), 3)
        self.assertEqual(((per_app.get(str(self.comp_app_a.id)) or {}).get("cfg") or {}).get("agregacio_exercicis"), "avg")

        self.assertEqual((per_app.get(str(self.comp_app_b.id)) or {}).get("mode"), "participant_aggregate")
        self.assertEqual(((per_app.get(str(self.comp_app_b.id)) or {}).get("cfg") or {}).get("mode"), "index")
        self.assertEqual(((per_app.get(str(self.comp_app_b.id)) or {}).get("cfg") or {}).get("index"), 3)
        self.assertEqual(((per_app.get(str(self.comp_app_b.id)) or {}).get("cfg") or {}).get("agregacio_exercicis"), "avg")

        hydrated = prepare_schema_for_builder_hydration(self.comp, cfg.schema or {}, tipus="individual")
        hydrated_per_app = ((hydrated.get("puntuacio") or {}).get("candidate_source_per_aparell") or {})
        self.assertEqual((((hydrated_per_app.get(str(self.comp_app_a.id)) or {}).get("cfg")) or {}).get("mode"), "millor_n")
        self.assertEqual((((hydrated_per_app.get(str(self.comp_app_b.id)) or {}).get("cfg")) or {}).get("mode"), "index")

    def test_classificacio_preview_returns_consistent_error_when_compute_fails(self):
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Preview runtime error",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=self._selected_total_schema([self.comp_app_a.id]),
        )

        with patch("competicions_trampoli.views.classificacions.builder.compute_classificacio", side_effect=RuntimeError("boom preview")):
            res = self.client.post(
                reverse("classificacio_preview", kwargs={"pk": self.comp.id, "cid": cfg.id}),
                data=json.dumps({}),
                content_type="application/json",
            )

        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertEqual(body.get("error"), "No s'ha pogut previsualitzar la classificacio.")
        self.assertEqual(body.get("errors"), ["boom preview"])
        self.assertTrue(body.get("error_details"))

    def test_classificacio_save_accepts_row_compute_by_judge_when_single_judge(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 3}}
            ],
            "computed": [
                {
                    "code": "E_by_judge",
                    "label": "E by judge",
                    "formula": "row_custom_compute('E', '1 - x', return_mode='by_judge')",
                },
            ],
        }
        meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        self.assertTrue(meta.get("E_by_judge", {}).get("scoreable"))

    def test_classificacio_save_rejects_row_compute_by_judge_when_multiple_judges(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 2}, "items": {"count": 3}}
            ],
            "computed": [
                {
                    "code": "E_by_judge",
                    "label": "E by judge",
                    "formula": "row_custom_compute('E', '1 - x', return_mode='by_judge')",
                },
            ],
        }
        meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        self.assertFalse(meta.get("E_by_judge", {}).get("scoreable"))
        self.assertIn("by_judge", str(meta.get("E_by_judge", {}).get("reason") or ""))

    def test_classificacio_save_accepts_column_compute_by_item_when_count_is_one(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 2}, "items": {"count": 4}}
            ],
            "computed": [
                {
                    "code": "E_by_item",
                    "label": "E by item",
                    "formula": "column_custom_compute('E', '1 - x', return_mode='by_item', count=1)",
                },
            ],
        }
        meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        self.assertTrue(meta.get("E_by_item", {}).get("scoreable"))

    def test_classificacio_save_rejects_column_compute_by_item_when_multiple_items(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 2}, "items": {"count": 4}}
            ],
            "computed": [
                {
                    "code": "E_by_item",
                    "label": "E by item",
                    "formula": "column_custom_compute('E', '1 - x', return_mode='by_item')",
                },
            ],
        }
        meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        self.assertFalse(meta.get("E_by_item", {}).get("scoreable"))
        self.assertIn("by_item", str(meta.get("E_by_item", {}).get("reason") or ""))

    def test_classificacio_save_tie_accepts_row_compute_by_judge_when_single_judge(self):
        schema_obj = {
            "fields": [
                {"code": "E", "type": "matrix", "shape": "judge_x_item", "judges": {"count": 1}, "items": {"count": 3}}
            ],
            "computed": [
                {
                    "code": "E_by_judge",
                    "label": "E by judge",
                    "formula": "row_custom_compute('E', '1 - x', return_mode='by_judge')",
                },
            ],
        }
        strict_meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=True)
        ui_meta = _build_scoreable_meta_for_schema(schema_obj, strict_unknown=False)
        self.assertTrue(strict_meta.get("E_by_judge", {}).get("scoreable"))
        self.assertTrue(ui_meta.get("E_by_judge", {}).get("scoreable"))

    def test_classificacio_save_rejects_invalid_tie_exercicis_selection_mode(self):
        payload = {
            "nom": "Cfg tie invalida",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["total"],
                        "ordre": "desc",
                        "scope": {
                            "aparells": {"mode": "tots"},
                            "exercicis": {"mode": "hereta"},
                        },
                        "mode_seleccio_exercicis": "invalid_mode",
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("desempat[0].mode_seleccio_exercicis" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_main_aparells_mode_tots(self):
        payload = {
            "nom": "Cfg main tots invalid",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "tots", "ids": []},
                    "camps_per_aparell": {
                        str(self.comp_app_a.id): ["total"],
                        str(self.comp_app_b.id): ["total"],
                    },
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("puntuacio.aparells.mode='tots'" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_negative_exercicis_max_per_participant(self):
        User = get_user_model()
        user = User.objects.create_user(
            username="classif_max_pp_editor",
            password="testpass123",
            email="classif-max-pp-editor@example.com",
        )
        CompeticioMembership.objects.create(
            user=user,
            competicio=self.comp,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        self.client.force_login(user)

        payload = {
            "nom": "Cfg max per participant invalid",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 2, "max_per_participant": -1},
                    "exercicis_best_n": 2,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("puntuacio.exercicis.max_per_participant" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_tie_aparells_mode_tots(self):
        payload = {
            "nom": "Cfg tie tots invalid",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": {
                "particions": [],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "camps_per_aparell": {str(self.comp_app_a.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["total"],
                        "ordre": "desc",
                        "scope": {
                            "aparells": {"mode": "tots"},
                            "exercicis": {"mode": "hereta"},
                        },
                    }
                ],
                "presentacio": {"top_n": 0, "mostrar_empats": True},
            },
        }
        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("desempat[0].scope.aparells.mode='tots'" in e for e in body.get("errors", [])))

    def test_compute_classificacio_tie_supports_per_app_override_exercicis(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        self.comp_app_b.nombre_exercicis = 2
        self.comp_app_b.save(update_fields=["nombre_exercicis"])

        # Participant A
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=1.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_b,
            inputs={},
            outputs={},
            total=5.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=2,
            comp_aparell=self.comp_app_b,
            inputs={},
            outputs={},
            total=5.0,
        )

        # Participant B
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=6.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=4.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_b,
            inputs={},
            outputs={},
            total=8.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=2,
            comp_aparell=self.comp_app_b,
            inputs={},
            outputs={},
            total=2.0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "tots", "ids": []}
        schema["puntuacio"]["camps_per_aparell"] = {
            str(self.comp_app_a.id): ["total"],
            str(self.comp_app_b.id): ["total"],
        }
        schema["puntuacio"]["exercicis"] = {"mode": "tots"}
        schema["puntuacio"]["agregacio_camps"] = "sum"
        schema["puntuacio"]["agregacio_exercicis"] = "sum"
        schema["puntuacio"]["agregacio_aparells"] = "sum"
        schema["desempat"] = [
            {
                "camps": ["total"],
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "tots"},
                    "exercicis": {"mode": "hereta"},
                },
                "agregacio_camps": "sum",
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "mode_seleccio_exercicis": "per_aparell_override",
                "exercicis_per_aparell": {
                    str(self.comp_app_a.id): {"mode": "millor_1"},
                    str(self.comp_app_b.id): {"mode": "pitjor_1"},
                },
            }
        ]

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Tie per app override",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["participant"], "Participant A")
        self.assertEqual(rows[1]["participant"], "Participant B")

    def test_compute_classificacio_main_exercicis_respects_max_per_participant(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=6.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=5.0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["total"]}
        schema["puntuacio"]["exercicis"] = {
            "mode": "millor_n",
            "best_n": 2,
            "max_per_participant": 1,
        }
        schema["puntuacio"]["agregacio_camps"] = "sum"
        schema["puntuacio"]["agregacio_exercicis"] = "sum"
        schema["puntuacio"]["agregacio_aparells"] = "sum"

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Main max per participant",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        self.assertEqual(len(rows), 2)
        points_by_name = {r["participant"]: r["punts"] for r in rows}
        self.assertEqual(points_by_name.get("Participant A"), 9.0)
        self.assertEqual(points_by_name.get("Participant B"), 6.0)

    def test_compute_classificacio_tie_exercicis_respects_max_per_participant(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=6.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=4.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=7.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=2,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=3.0,
        )

        schema = self._base_cfg_schema()
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.comp_app_a.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["total"]}
        schema["puntuacio"]["exercicis"] = {"mode": "tots"}
        schema["puntuacio"]["agregacio_camps"] = "sum"
        schema["puntuacio"]["agregacio_exercicis"] = "sum"
        schema["puntuacio"]["agregacio_aparells"] = "sum"
        schema["desempat"] = [
            {
                "camps": ["total"],
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "exercicis": {"mode": "millor_n", "best_n": 2, "max_per_participant": 1},
                },
                "agregacio_camps": "sum",
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
            }
        ]

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Tie max per participant",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        rows = out.get("global", [])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["participant"], "Participant B")
        self.assertEqual(rows[1]["participant"], "Participant A")

    def test_compute_classificacio_supports_custom_partition_groups_for_categoria(self):
        self.ins_a.categoria = "ALEVI"
        self.ins_a.save(update_fields=["categoria"])
        self.ins_b.categoria = "PREBENJAMI"
        self.ins_b.save(update_fields=["categoria"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )

        schema = self._valid_partition_schema()
        schema["particions"] = ["categoria"]
        schema["particions_custom"] = {
            "categoria": {
                "mode": "custom",
                "grups": [
                    {"key": "base", "label": "Base", "values": ["ALEVI", "PREBENJAMI"]},
                ],
            }
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Particio custom categoria",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        self.assertEqual(list(out.keys()), ["categoria:Base"])
        self.assertEqual(len(out["categoria:Base"]), 2)

    def test_compute_classificacio_supports_partition_by_schema_extra_field(self):
        self.comp.inscripcions_schema = {
            "columns": [
                {"code": "nivell", "label": "Nivell", "kind": "extra"},
            ]
        }
        self.comp.save(update_fields=["inscripcions_schema"])

        self.ins_a.extra = {"nivell": "N1"}
        self.ins_a.save(update_fields=["extra"])
        self.ins_b.extra = {"nivell": "N2"}
        self.ins_b.save(update_fields=["extra"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )

        schema = self._valid_partition_schema()
        schema["particions"] = ["nivell"]
        schema["particions_custom"] = {
            "nivell": {
                "mode": "custom",
                "grups": [
                    {"key": "bloc_1", "label": "Bloc 1", "values": ["N1", "N2"]},
                ],
            }
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Particio extra",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        self.assertEqual(list(out.keys()), ["nivell:Bloc 1"])
        self.assertEqual(len(out["nivell:Bloc 1"]), 2)

    def test_compute_classificacio_supports_conditional_partition_on_parent_group(self):
        ins_c = self._create_inscripcio(self.comp, "Participant C", ordre=3)

        self.ins_a.categoria = "ALEVI"
        self.ins_a.subcategoria = "N1"
        self.ins_a.save(update_fields=["categoria", "subcategoria"])
        self.ins_b.categoria = "PREBENJAMI"
        self.ins_b.subcategoria = "N2"
        self.ins_b.save(update_fields=["categoria", "subcategoria"])
        ins_c.categoria = "INFANTIL"
        ins_c.subcategoria = "N3"
        ins_c.save(update_fields=["categoria", "subcategoria"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=ins_c,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=7.0,
        )

        schema = self._valid_partition_schema()
        schema["particions"] = ["categoria", "subcategoria"]
        schema["particions_v2"] = [
            {"code": "categoria", "apply_mode": "all"},
            {"code": "subcategoria", "apply_mode": "some_parents", "parent_values": ["Base"]},
        ]
        schema["particions_custom"] = {
            "categoria": {
                "mode": "custom",
                "grups": [
                    {"key": "base", "label": "Base", "values": ["ALEVI", "PREBENJAMI"]},
                    {"key": "grans", "label": "Grans", "values": ["INFANTIL"]},
                ],
            }
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Particio condicional",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        self.assertEqual(
            set(out.keys()),
            {"categoria:Base|subcategoria:N1", "categoria:Base|subcategoria:N2", "categoria:Grans"},
        )
        self.assertEqual(out["categoria:Grans"][0]["participant"], "Participant C")

    def test_normalize_particions_schema_populates_particions_v2_from_legacy_list(self):
        schema = self._valid_partition_schema()
        schema["particions"] = ["categoria", "subcategoria"]

        normalized = _normalize_particions_schema(schema)
        self.assertEqual(normalized.get("particions"), ["categoria", "subcategoria"])
        self.assertEqual(
            normalized.get("particions_v2"),
            [
                {"code": "categoria", "apply_mode": "all", "parent_values": []},
                {"code": "subcategoria", "apply_mode": "all", "parent_values": []},
            ],
        )

    def test_compute_classificacio_partitions_by_birth_year_ranges(self):
        self.ins_a.data_naixement = date(2008, 5, 4)
        self.ins_b.data_naixement = date(2011, 2, 10)
        self.ins_a.save(update_fields=["data_naixement"])
        self.ins_b.save(update_fields=["data_naixement"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=8.0,
        )

        schema = self._valid_partition_schema()
        schema["particions"] = ["any_naixement_forquilla"]
        schema["particions_v2"] = [
            {"code": "any_naixement_forquilla", "apply_mode": "all", "parent_values": []},
        ]
        schema["particions_config"] = {
            "any_naixement_forquilla": {
                "ranges": [
                    {"label": "2007-2009", "from_year": 2007, "to_year": 2009},
                    {"label": "2010-2012", "from_year": 2010, "to_year": 2012},
                ],
                "sense_data_label": "Sense data",
                "fora_rang_label": "Fora de forquilla",
            }
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Forquilles naixement",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)
        self.assertEqual(
            set(out.keys()),
            {"any_naixement_forquilla:2007-2009", "any_naixement_forquilla:2010-2012"},
        )
        self.assertEqual(out["any_naixement_forquilla:2007-2009"][0]["participant"], "Participant A")
        self.assertEqual(out["any_naixement_forquilla:2010-2012"][0]["participant"], "Participant B")

    def test_classificacio_save_rejects_overlapping_birth_year_ranges(self):
        schema = self._valid_partition_schema()
        schema["particions"] = ["any_naixement_forquilla"]
        schema["particions_v2"] = [
            {"code": "any_naixement_forquilla", "apply_mode": "all", "parent_values": []},
        ]
        schema["particions_config"] = {
            "any_naixement_forquilla": {
                "ranges": [
                    {"label": "2007-2009", "from_year": 2007, "to_year": 2009},
                    {"label": "2009-2011", "from_year": 2009, "to_year": 2011},
                ]
            }
        }

        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertTrue(any("solapament" in e for e in errors))

    def test_classificacio_save_accepts_birth_year_ranges_for_team_rankings(self):
        schema = self._valid_partition_schema()
        schema["particions"] = ["any_naixement_forquilla"]
        schema["particions_v2"] = [
            {"code": "any_naixement_forquilla", "apply_mode": "all", "parent_values": []},
        ]
        schema["particions_config"] = {
            "any_naixement_forquilla": {
                "ranges": [
                    {"label": "2007-2009", "from_year": 2007, "to_year": 2009},
                ],
                "team_rules": {
                    "reference_mode": "oldest_member_birthdate",
                    "compliance_mode": "strict",
                    "max_members_outside_range": 0,
                    "missing_birthdate_policy": "outside_range",
                },
            }
        }

        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="equips")
        self.assertEqual(errors, [])

    def test_validate_particions_schema_rejects_conditional_partition_without_parent_values(self):
        schema = self._valid_partition_schema()
        schema["particions_v2"] = [
            {"code": "categoria", "apply_mode": "all"},
            {"code": "subcategoria", "apply_mode": "some_parents", "parent_values": []},
        ]
        normalized = _normalize_particions_schema(schema)
        errors = _validate_particions_schema(self.comp, normalized)
        self.assertTrue(any("parent_values" in e for e in errors))

    def test_classificacio_save_rejects_unknown_partition_field(self):
        payload = {
            "nom": "Cfg particio desconeguda",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": self._valid_partition_schema(),
        }
        payload["schema"]["particions"] = ["camp_inexistent"]
        payload["schema"]["particions_custom"] = {
            "camp_inexistent": {
                "mode": "custom",
                "grups": [{"label": "X", "values": ["A"]}],
            }
        }

        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("camp no perm" in e for e in body.get("errors", [])))

    def test_classificacio_save_rejects_duplicate_custom_partition_values(self):
        payload = {
            "nom": "Cfg particio custom duplicada",
            "activa": True,
            "ordre": 1,
            "tipus": "individual",
            "schema": self._valid_partition_schema(),
        }
        payload["schema"]["particions"] = ["categoria"]
        payload["schema"]["particions_custom"] = {
            "categoria": {
                "mode": "custom",
                "grups": [
                    {"label": "Bloc 1", "values": ["ALEVI"]},
                    {"label": "Bloc 2", "values": ["ALEVI"]},
                ],
            }
        }

        url = reverse("classificacio_save", kwargs={"pk": self.comp.id})
        res = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("valor repetit entre grups" in e for e in body.get("errors", [])))

    def test_compute_classificacio_victories_single_app_matches_order(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=9.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=7.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories 1 app",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([r["participant"] for r in rows], ["Participant A", "Participant B"])
        self.assertEqual(rows[0]["by_app_base"][self.comp_app_a.id], 9.0)
        self.assertEqual(rows[0]["by_app"][self.comp_app_a.id], 1.0)
        self.assertEqual(rows[0]["punts"], 1.0)
        self.assertEqual(rows[1]["by_app"][self.comp_app_a.id], 0.0)
        self.assertEqual(rows[1]["punts"], 0.0)

    def test_compute_classificacio_victories_multiple_apps_aggregates_after_duels(self):
        ins_c = self._create_inscripcio(self.comp, "Participant C", ordre=3)
        schema = self._selected_total_schema(
            [self.comp_app_a.id, self.comp_app_b.id],
            mode_resultat="victories",
        )

        scores = {
            self.ins_a.id: {self.comp_app_a.id: 100.0, self.comp_app_b.id: 1.0},
            self.ins_b.id: {self.comp_app_a.id: 60.0, self.comp_app_b.id: 60.0},
            ins_c.id: {self.comp_app_a.id: 59.0, self.comp_app_b.id: 59.0},
        }
        for ins_id, per_app in scores.items():
            ins = Inscripcio.objects.get(pk=ins_id)
            for app_id, total in per_app.items():
                comp_app = self.comp_app_a if app_id == self.comp_app_a.id else self.comp_app_b
                ScoreEntry.objects.create(
                    competicio=self.comp,
                    inscripcio=ins,
                    exercici=1,
                    comp_aparell=comp_app,
                    inputs={},
                    outputs={},
                    total=total,
                )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories 2 apps",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([r["participant"] for r in rows], ["Participant B", "Participant A", "Participant C"])
        self.assertEqual(rows[0]["punts"], 3.0)
        self.assertEqual(rows[1]["punts"], 2.0)
        self.assertEqual(rows[2]["punts"], 1.0)

    def test_compute_classificacio_victories_internal_tie_break_uses_compare_tie(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["desempat_comparacio"] = [
            {
                "camp": "E",
                "camps": ["E"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "scope": {"exercicis": {"mode": "hereta"}},
            }
        ]

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"E": 9.0},
            outputs={},
            total=10.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={"E": 8.0},
            outputs={},
            total=10.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories compare tie",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([r["participant"] for r in rows], ["Participant A", "Participant B"])
        self.assertEqual(rows[0]["punts"], 1.0)
        self.assertEqual(rows[1]["punts"], 0.0)

    def test_compute_classificacio_victories_unresolved_tie_gives_half_points(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=10.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app_a,
            inputs={},
            outputs={},
            total=10.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories tied",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual(rows[0]["punts"], 0.5)
        self.assertEqual(rows[1]["punts"], 0.5)
        self.assertEqual(rows[0]["by_app"][self.comp_app_a.id], 0.5)
        self.assertEqual(rows[1]["by_app"][self.comp_app_a.id], 0.5)

    def test_compute_classificacio_victories_separated_fields_aggregate_per_app(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["E", "D"]}
        schema["puntuacio"]["victories"]["mode_camps"] = "separat"
        schema["puntuacio"]["victories"]["agregacio_victories_camps"] = "sum"

        for exercici, e_val, d_val in ((1, 10.0, 1.0), (2, 10.0, 1.0)):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={"E": e_val, "D": d_val},
                outputs={},
                total=e_val + d_val,
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={"E": 8.0, "D": 8.0},
                outputs={},
                total=16.0,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories fields separated",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        by_name = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in rows}
        self.assertEqual(by_name.get("Participant A"), 1.0)
        self.assertEqual(by_name.get("Participant B"), 1.0)

    def test_compute_classificacio_victories_separated_exercises_aggregate_per_app(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["mode_exercicis"] = "separat"
        schema["puntuacio"]["victories"]["agregacio_victories_exercicis"] = "sum"

        for exercici, total_a, total_b in ((1, 10.0, 9.0), (2, 1.0, 2.0)):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={},
                outputs={},
                total=total_a,
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={},
                outputs={},
                total=total_b,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories exercises separated",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        by_name = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in rows}
        self.assertEqual(by_name.get("Participant A"), 1.0)
        self.assertEqual(by_name.get("Participant B"), 1.0)

    def test_compute_classificacio_victories_separated_fields_and_exercises_use_intermediate_aggregation(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["E", "D"]}
        schema["puntuacio"]["victories"]["mode_camps"] = "separat"
        schema["puntuacio"]["victories"]["mode_exercicis"] = "separat"
        schema["puntuacio"]["victories"]["agregacio_victories_camps"] = "avg"
        schema["puntuacio"]["victories"]["agregacio_victories_exercicis"] = "sum"

        per_ex = {
            1: {"a": {"E": 10.0, "D": 1.0}, "b": {"E": 9.0, "D": 9.0}},
            2: {"a": {"E": 8.0, "D": 1.0}, "b": {"E": 7.0, "D": 9.0}},
        }
        for exercici, vals in per_ex.items():
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs=vals["a"],
                outputs={},
                total=sum(vals["a"].values()),
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs=vals["b"],
                outputs={},
                total=sum(vals["b"].values()),
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories separated all levels",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        by_name = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in rows}
        self.assertEqual(by_name.get("Participant A"), 1.0)
        self.assertEqual(by_name.get("Participant B"), 1.0)

    def test_compute_classificacio_victories_separated_fields_supports_global_or_per_field_exercise_selection(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["camps_per_aparell"] = {str(self.comp_app_a.id): ["E", "D"]}
        schema["puntuacio"]["exercicis"] = {"mode": "millor_1"}
        schema["puntuacio"]["victories"]["mode_camps"] = "separat"
        schema["puntuacio"]["victories"]["agregacio_victories_camps"] = "sum"

        for exercici, vals_a, vals_b in (
            (1, {"E": 10.0, "D": 1.0}, {"E": 9.0, "D": 9.0}),
            (2, {"E": 1.0, "D": 10.0}, {"E": 8.0, "D": 8.0}),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs=vals_a,
                outputs={},
                total=sum(vals_a.values()),
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs=vals_b,
                outputs={},
                total=sum(vals_b.values()),
            )

        schema_global = json.loads(json.dumps(schema))
        schema_global["puntuacio"]["victories"]["mode_seleccio_exercicis_camps_separats"] = "global"
        cfg_global = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories fields global select",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema_global,
        )
        global_rows = compute_classificacio(self.comp, cfg_global).get("global", [])
        global_points = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in global_rows}

        schema_per_field = json.loads(json.dumps(schema))
        schema_per_field["puntuacio"]["victories"]["mode_seleccio_exercicis_camps_separats"] = "per_camp"
        cfg_per_field = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories fields per field select",
            activa=True,
            ordre=2,
            tipus="individual",
            schema=schema_per_field,
        )
        per_field_rows = compute_classificacio(self.comp, cfg_per_field).get("global", [])
        per_field_points = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in per_field_rows}

        self.assertEqual(global_points.get("Participant A"), 1.0)
        self.assertEqual(global_points.get("Participant B"), 1.0)
        self.assertEqual(per_field_points.get("Participant A"), 2.0)
        self.assertEqual(per_field_points.get("Participant B"), 0.0)

    def test_compute_classificacio_victories_internal_tie_break_respects_separated_exercise_unit(self):
        self.comp_app_a.nombre_exercicis = 2
        self.comp_app_a.save(update_fields=["nombre_exercicis"])
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["mode_exercicis"] = "separat"
        schema["puntuacio"]["victories"]["desempat_comparacio"] = [
            {
                "camp": "E",
                "camps": ["E"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "scope": {"exercicis": {"mode": "hereta"}},
            }
        ]

        for exercici, e_a, e_b in ((1, 9.0, 8.0), (2, 1.0, 9.0)):
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_a,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={"E": e_a},
                outputs={},
                total=10.0,
            )
            ScoreEntry.objects.create(
                competicio=self.comp,
                inscripcio=self.ins_b,
                exercici=exercici,
                comp_aparell=self.comp_app_a,
                inputs={"E": e_b},
                outputs={},
                total=10.0,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Victories tie per exercise",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=schema,
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        by_name = {row["participant"]: row["by_app"][self.comp_app_a.id] for row in rows}
        self.assertEqual(by_name.get("Participant A"), 1.0)
        self.assertEqual(by_name.get("Participant B"), 1.0)

    def test_classificacio_save_rejects_invalid_victories_granular_modes(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["mode_camps"] = "invalid"
        schema["puntuacio"]["victories"]["mode_exercicis"] = "broken"
        schema["puntuacio"]["victories"]["mode_seleccio_exercicis_camps_separats"] = "oops"
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertTrue(any("mode_camps invalid" in e for e in errors))
        self.assertTrue(any("mode_exercicis invalid" in e for e in errors))
        self.assertTrue(any("mode_seleccio_exercicis_camps_separats invalid" in e for e in errors))

    def test_classificacio_save_ignores_invalid_victories_granular_modes_in_score_mode(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="score")
        schema["puntuacio"]["victories"]["mode_camps"] = "invalid"
        schema["puntuacio"]["victories"]["mode_exercicis"] = "broken"
        schema["puntuacio"]["victories"]["mode_seleccio_exercicis_camps_separats"] = "oops"
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertFalse(any("mode_camps invalid" in e for e in errors))
        self.assertFalse(any("mode_exercicis invalid" in e for e in errors))
        self.assertFalse(any("mode_seleccio_exercicis_camps_separats invalid" in e for e in errors))

    def test_classificacio_save_rejects_victories_for_non_individual(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="entitat")
        self.assertTrue(any("mode_resultat_aparells='victories'" in e for e in errors))

    def test_classificacio_save_rejects_victories_compare_tie_scope_aparells(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["desempat_comparacio"] = [
            {
                "camps": ["total"],
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app_a.id]},
                    "exercicis": {"mode": "hereta"},
                },
            }
        ]
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertTrue(any("scope.aparells no esta permes" in e for e in errors))

    def test_classificacio_save_rejects_victories_compare_tie_scope_participants(self):
        schema = self._selected_total_schema([self.comp_app_a.id], mode_resultat="victories")
        schema["puntuacio"]["victories"]["desempat_comparacio"] = [
            {
                "camps": ["total"],
                "ordre": "desc",
                "scope": {
                    "participants": {"mode": "tots"},
                    "exercicis": {"mode": "hereta"},
                },
            }
        ]
        _, errors = _validate_schema_for_competicio(self.comp, schema, tipus="individual")
        self.assertTrue(any("scope.participants no esta permes" in e for e in errors))


class ClassificacioBuilderHydrationTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Builder Hydration")
        self.user = self._login_competicio_user(
            self.comp,
            role=CompeticioMembership.Role.EDITOR,
            username_prefix="builder_hydration",
        )
        self.native_ctx = self._ensure_native_equip_context(self.comp)
        self.ctx_finals = EquipContext.objects.create(competicio=self.comp, code="ctx-finals", nom="Finals")
        self.ctx_alt = EquipContext.objects.create(competicio=self.comp, code="ctx-alt", nom="Alt")

        self.ind_app = self._create_aparell("BH_IND", "Builder Hydration Individual")
        self.comp_ind_app = self._create_comp_aparell(self.comp, self.ind_app, ordre=1, actiu=True)

        self.team_app_finals = self._create_aparell("BH_TEAM_F", "Builder Hydration Team Finals")
        self.team_app_finals.competition_unit = Aparell.CompetitionUnit.TEAM
        self.team_app_finals.save(update_fields=["competition_unit"])
        self.comp_team_app_finals = self._create_comp_aparell(self.comp, self.team_app_finals, ordre=2, actiu=True)

        self.team_app_alt = self._create_aparell("BH_TEAM_A", "Builder Hydration Team Alt")
        self.team_app_alt.competition_unit = Aparell.CompetitionUnit.TEAM
        self.team_app_alt.save(update_fields=["competition_unit"])
        self.comp_team_app_alt = self._create_comp_aparell(self.comp, self.team_app_alt, ordre=3, actiu=True)

        CompeticioAparellEquipContextSource.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_team_app_finals,
            context=self.ctx_finals,
        )
        CompeticioAparellEquipContextSource.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_team_app_alt,
            context=self.ctx_alt,
        )

    def _builder_schema(self, *, context_code="native", team_mode=""):
        return {
            "particions": [],
            "particions_v2": [],
            "particions_custom": {},
            "particions_config": {},
            "filtres": {},
            "puntuacio": {
                "aparells": {"mode": "tots", "ids": []},
                "camps_per_aparell": {},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "mode_resultat_aparells": "score",
                "victories": {
                    "punts_victoria": 1,
                    "punts_empat": 0.5,
                    "sense_nota_mode": "skip",
                    "mode_camps": "agregat",
                    "mode_exercicis": "agregat",
                    "mode_seleccio_exercicis_camps_separats": "per_camp",
                    "agregacio_victories_camps": "sum",
                    "agregacio_victories_exercicis": "sum",
                    "desempat_comparacio": [],
                },
                "ordre": "desc",
            },
            "desempat": [],
            "presentacio": {
                "top_n": 0,
                "mostrar_empats": True,
                "columnes": [{"type": "builtin", "key": "participant", "label": "Participant", "align": "left"}],
            },
            "equips": {
                "context_code": context_code,
                "assignment_source": {"mode": "context", "context_code": context_code, "fallback": "native"},
                "team_mode": team_mode,
            },
        }

    def test_prepare_schema_for_builder_hydration_mode_tots_filters_individual_tipus(self):
        hydrated = prepare_schema_for_builder_hydration(
            self.comp,
            self._builder_schema(),
            tipus="individual",
        )

        self.assertEqual((hydrated.get("puntuacio") or {}).get("aparells", {}).get("ids"), [self.comp_ind_app.id])

    def test_prepare_schema_for_builder_hydration_mode_tots_filters_derived_team_tipus(self):
        hydrated = prepare_schema_for_builder_hydration(
            self.comp,
            self._builder_schema(context_code="ctx-finals", team_mode="derived_from_individual"),
            tipus="equips",
        )

        self.assertEqual((hydrated.get("puntuacio") or {}).get("aparells", {}).get("ids"), [self.comp_ind_app.id])

    def test_prepare_schema_for_builder_hydration_mode_tots_filters_native_team_by_context(self):
        hydrated = prepare_schema_for_builder_hydration(
            self.comp,
            self._builder_schema(context_code="ctx-finals", team_mode="native_team"),
            tipus="equips",
        )

        self.assertEqual(
            (hydrated.get("puntuacio") or {}).get("aparells", {}).get("ids"),
            [self.comp_team_app_finals.id],
        )

    def test_prepare_schema_for_builder_hydration_prefers_assignment_source_context_code(self):
        schema = self._builder_schema(context_code="native", team_mode="derived_from_individual")
        schema["equips"]["assignment_source"] = {"mode": "context", "context_code": "ctx-finals", "fallback": "native"}

        hydrated = prepare_schema_for_builder_hydration(
            self.comp,
            schema,
            tipus="equips",
        )

        self.assertEqual((hydrated.get("equips") or {}).get("context_code"), "ctx-finals")
        self.assertEqual(
            (((hydrated.get("equips") or {}).get("assignment_source")) or {}).get("context_code"),
            "ctx-finals",
        )

    def test_prepare_schema_for_builder_hydration_preserves_candidate_source_contract(self):
        schema = self._builder_schema(context_code="ctx-finals", team_mode="derived_from_individual")
        schema["puntuacio"]["candidate_source_mode"] = "participant_aggregate"
        schema["puntuacio"]["candidate_source_cfg"] = {
            "mode": "millor_n",
            "best_n": 1,
            "index": 1,
            "ids": [],
            "agregacio_exercicis": "sum",
        }

        hydrated = prepare_schema_for_builder_hydration(
            self.comp,
            schema,
            tipus="equips",
        )

        punt = hydrated.get("puntuacio") or {}
        self.assertEqual(punt.get("candidate_source_mode"), "participant_aggregate")
        self.assertEqual((punt.get("candidate_source_cfg") or {}).get("mode"), "millor_n")
        self.assertEqual((punt.get("candidate_source_cfg") or {}).get("best_n"), 1)

    def test_prepare_schema_for_builder_hydration_preserves_native_team_candidate_source_contract(self):
        schema = self._builder_schema(context_code="ctx-finals", team_mode="native_team")
        schema["puntuacio"]["candidate_source_mode"] = "team_aggregate"
        schema["puntuacio"]["candidate_source_cfg"] = {
            "mode": "millor_n",
            "best_n": 2,
            "index": 1,
            "ids": [],
            "agregacio_exercicis": "sum",
        }
        schema["puntuacio"]["candidate_source_per_aparell"] = {
            str(self.comp_team_app_finals.id): {
                "mode": "team_aggregate",
                "cfg": {
                    "mode": "index",
                    "index": 2,
                    "best_n": 1,
                    "ids": [],
                    "agregacio_exercicis": "max",
                },
            },
        }

        hydrated = prepare_schema_for_builder_hydration(
            self.comp,
            schema,
            tipus="equips",
        )

        punt = hydrated.get("puntuacio") or {}
        self.assertEqual(punt.get("candidate_source_mode"), "team_aggregate")
        self.assertEqual((punt.get("candidate_source_cfg") or {}).get("mode"), "millor_n")
        self.assertEqual((punt.get("candidate_source_cfg") or {}).get("best_n"), 2)
        per_app = punt.get("candidate_source_per_aparell") or {}
        self.assertEqual((per_app.get(str(self.comp_team_app_finals.id)) or {}).get("mode"), "team_aggregate")
        self.assertEqual((((per_app.get(str(self.comp_team_app_finals.id)) or {}).get("cfg")) or {}).get("mode"), "index")
        self.assertEqual((((per_app.get(str(self.comp_team_app_finals.id)) or {}).get("cfg")) or {}).get("index"), 2)


class ClassificacioFilterSemanticsTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Filter Semantics")
        self.app = self._create_aparell("TRAMP_FILTER", "Tramp Filter")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        self.ins_a = self._create_inscripcio(self.comp, "Participant A", ordre=1)
        self.ins_b = self._create_inscripcio(self.comp, "Participant B", ordre=2)
        self.ins_c = self._create_inscripcio(self.comp, "Participant C", ordre=3)

        self.ins_a.categoria = "Base"
        self.ins_a.entitat = "Club A"
        self.ins_a.save(update_fields=["categoria", "entitat"])
        self.ins_b.categoria = "Promo"
        self.ins_b.entitat = "Club A"
        self.ins_b.save(update_fields=["categoria", "entitat"])
        self.ins_c.categoria = "Base"
        self.ins_c.entitat = "Club B"
        self.ins_c.save(update_fields=["categoria", "entitat"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins_a,
            exercici=1,
            inputs={},
            outputs={},
            total=10,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins_b,
            exercici=1,
            inputs={},
            outputs={},
            total=30,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins_c,
            exercici=1,
            inputs={},
            outputs={},
            total=20,
        )

    def _schema(self, *, filters=None):
        schema = {
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
            },
        }
        if filters is not None:
            schema["filtres"] = filters
        return schema

    def test_compute_classificacio_individual_filters_participants_before_ranking(self):
        rows = compute_classificacio(
            self.comp,
            SimpleNamespace(schema=self._schema(filters={"categories_in": ["Base"]}), tipus="individual"),
        ).get("global", [])

        self.assertEqual([row["participant"] for row in rows], ["Participant C", "Participant A"])
        self.assertEqual([row["score"] for row in rows], [20.0, 10.0])

    def test_compute_classificacio_entitat_aggregates_only_filtered_participants(self):
        rows = compute_classificacio(
            self.comp,
            SimpleNamespace(schema=self._schema(filters={"categories_in": ["Base"]}), tipus="entitat"),
        ).get("global", [])

        self.assertEqual([row["entitat_nom"] for row in rows], ["Club B", "Club A"])
        self.assertEqual([row["score"] for row in rows], [20.0, 10.0])


class ClassificacionsExportExcelTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp Export Classificacions")
        self.app = self._create_aparell("TRAMP_EXPORT", "Tramp Export")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)

        self.ins_a = self._create_inscripcio(self.comp, "Participant A", ordre=1)
        self.ins_b = self._create_inscripcio(self.comp, "Participant B", ordre=2)
        self.ins_c = self._create_inscripcio(self.comp, "Participant C", ordre=3)

        self.ins_a.categoria = "Junior"
        self.ins_a.subcategoria = "Femeni"
        self.ins_a.save(update_fields=["categoria", "subcategoria"])

        self.ins_b.categoria = "Junior"
        self.ins_b.subcategoria = "Femeni"
        self.ins_b.save(update_fields=["categoria", "subcategoria"])

        self.ins_c.categoria = "Senior"
        self.ins_c.subcategoria = "Masculi"
        self.ins_c.save(update_fields=["categoria", "subcategoria"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_a,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=9.6,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_b,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=8.3,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins_c,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=7.1,
        )

        self.cfg_general = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="General",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=self._schema_for_parts([]),
        )
        self.cfg_particions = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Per categories",
            activa=True,
            ordre=2,
            tipus="individual",
            schema=self._schema_for_parts(["categoria", "subcategoria"]),
        )

        User = get_user_model()
        self.user = User.objects.create_user(
            username="classif_export_user",
            password="testpass123",
            email="classif-export@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.READONLY,
            is_active=True,
        )

    def _schema_for_parts(self, parts):
        return {
            "particions": parts,
            "filtres": {},
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "exercicis_best_n": 1,
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
                "camp": "total",
                "agregacio": "sum",
                "best_n": 1,
            },
            "desempat": [],
            "presentacio": {
                "top_n": 0,
                "mostrar_empats": True,
                "columnes": [
                    {"type": "builtin", "key": "posicio", "label": "Pos.", "align": "left"},
                    {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                ],
            },
        }

    def test_export_excel_creates_one_sheet_per_classificacio(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url)

        self.assertEqual(res.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            res["Content-Type"],
        )

        wb = load_workbook(filename=BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), 2)

        ws_general = wb[wb.sheetnames[0]]
        ws_parts = wb[wb.sheetnames[1]]

        self.assertIn("General", str(ws_general["A1"].value))
        self.assertIn("Particio:", str(ws_general["A4"].value))
        self.assertEqual(ws_general.freeze_panes, "A6")

        gold_fill = (ws_general["A6"].fill.fgColor.rgb or "").upper()
        silver_fill = (ws_general["A7"].fill.fgColor.rgb or "").upper()
        header_fill = (ws_general["A5"].fill.fgColor.rgb or "").upper()
        self.assertTrue(gold_fill.endswith("F6E27A"))
        self.assertTrue(silver_fill.endswith("E3E8EF"))
        self.assertTrue(header_fill.endswith("E9EEF7"))

        self.assertIn("Particio:", str(ws_parts["A4"].value))
        self.assertIn("/", str(ws_parts["A4"].value))

    def test_export_excel_can_filter_single_cfg_with_cfg_id(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url, {"cfg_id": self.cfg_particions.id})

        self.assertEqual(res.status_code, 200)
        wb = load_workbook(filename=BytesIO(res.content))
        self.assertEqual(len(wb.sheetnames), 1)
        ws = wb[wb.sheetnames[0]]
        self.assertIn("Per categories", str(ws["A1"].value))

    def test_export_excel_rejects_invalid_cfg_id(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        res = self.client.get(url, {"cfg_id": "abc"})
        self.assertEqual(res.status_code, 400)

    def test_export_excel_returns_consistent_error_when_compute_fails(self):
        self.client.force_login(self.user)
        url = reverse("classificacions_live_export_excel", kwargs={"pk": self.comp.id})
        with patch("competicions_trampoli.views.classificacions.export.compute_classificacio", side_effect=RuntimeError("boom export")):
            res = self.client.get(url, {"cfg_id": self.cfg_general.id})

        self.assertEqual(res.status_code, 400)
        body = res.content.decode("utf-8")
        self.assertIn("No s'ha pogut renderitzar la classificacio.", body)
        self.assertIn("boom export", body)


class ClassificacioTemplateFlowTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp_source = self._create_competicio("Comp Templates Source")
        self.comp_target = self._create_competicio("Comp Templates Target")

        self.app = self._create_aparell("TPL_APP", "Template App")
        self.source_app = self._create_comp_aparell(self.comp_source, self.app, ordre=1, actiu=True)
        self.target_app = self._create_comp_aparell(self.comp_target, self.app, ordre=1, actiu=True)

        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "E_total", "type": "number"},
                ],
                "computed": [],
            },
        )

        self.cfg_source = ClassificacioConfig.objects.create(
            competicio=self.comp_source,
            nom="Cfg Source",
            activa=True,
            ordre=1,
            tipus="individual",
            schema={
                "particions": ["categoria"],
                "filtres": {},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.source_app.id]},
                    "camps_per_aparell": {str(self.source_app.id): ["E_total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "exercicis_best_n": 1,
                    "mode_seleccio_exercicis": "per_aparell_global",
                    "exercicis_per_aparell": {},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                    "camp": "total",
                    "agregacio": "sum",
                    "best_n": 1,
                },
                "desempat": [],
                "presentacio": {
                    "top_n": 0,
                    "mostrar_empats": True,
                    "columnes": [
                        {"type": "builtin", "key": "posicio", "label": "#", "align": "left"},
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                    ],
                },
            },
        )

        User = get_user_model()
        self.editor_user = User.objects.create_user(
            username="tpl_editor_user",
            password="testpass123",
            email="tpl-editor@example.com",
        )
        self.other_user = User.objects.create_user(
            username="tpl_other_user",
            password="testpass123",
            email="tpl-other@example.com",
        )

        CompeticioMembership.objects.create(
            user=self.editor_user,
            competicio=self.comp_source,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.editor_user,
            competicio=self.comp_target,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.other_user,
            competicio=self.comp_source,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.other_user,
            competicio=self.comp_target,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )

    def _post_json_as(self, user, url_name, comp_id, payload):
        self.client.force_login(user)
        url = reverse(url_name, kwargs={"pk": comp_id})
        return self.client.post(url, data=json.dumps(payload), content_type="application/json")

    def test_editor_with_classificacions_edit_can_save_global_template(self):
        res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL 1"},
        )
        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.json().get("ok"))

    def test_template_schema_helpers_preserve_victories_config(self):
        schema = json.loads(json.dumps(self.cfg_source.schema or {}))
        schema["equips"] = {
            "context_code": "native",
            "assignment_source": {"mode": "context", "context_code": "native", "fallback": "native"},
            "team_mode": "derived_from_individual",
        }
        schema["puntuacio"]["mode_resultat_aparells"] = "victories"
        schema["puntuacio"]["victories"] = {
            "punts_victoria": 1,
            "punts_empat": 0,
            "sense_nota_mode": "skip",
            "mode_camps": "separat",
            "mode_exercicis": "separat",
            "mode_seleccio_exercicis_camps_separats": "global",
            "agregacio_victories_camps": "avg",
            "agregacio_victories_exercicis": "max",
            "desempat_comparacio": [
                {
                    "camp": "E_total",
                    "camps": ["E_total"],
                    "agregacio_camps": "hereta",
                    "ordre": "desc",
                    "scope": {"exercicis": {"mode": "hereta"}},
                }
            ],
        }

        schema_tpl, warnings = _schema_to_template_schema(self.comp_source, schema)
        self.assertFalse(warnings)
        self.assertEqual(
            ((schema_tpl.get("puntuacio") or {}).get("mode_resultat_aparells")),
            "victories",
        )

        self._ensure_native_equip_context(self.comp_target)
        schema_local, mapping_warnings, mapping = _template_schema_to_competicio_schema(self.comp_target, schema_tpl)
        self.assertFalse(mapping_warnings)
        self.assertEqual(mapping.get(self.app.codi), self.target_app.id)
        punt = (schema_local.get("puntuacio") or {})
        self.assertEqual(punt.get("mode_resultat_aparells"), "victories")
        self.assertEqual((punt.get("aparells") or {}).get("ids"), [self.target_app.id])
        self.assertEqual((punt.get("victories") or {}).get("mode_camps"), "separat")
        self.assertEqual((punt.get("victories") or {}).get("mode_exercicis"), "separat")
        self.assertEqual(
            (punt.get("victories") or {}).get("mode_seleccio_exercicis_camps_separats"),
            "global",
        )

    def test_template_schema_helpers_roundtrip_per_app_exercise_aggregation_maps(self):
        schema = json.loads(json.dumps(self.cfg_source.schema or {}))
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.source_app.id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(self.source_app.id): ["E_total"]}
        schema["puntuacio"]["mode_seleccio_exercicis"] = "per_aparell_override"
        schema["puntuacio"]["exercicis_per_aparell"] = {str(self.source_app.id): {"mode": "millor_1"}}
        schema["puntuacio"]["agregacio_exercicis_per_aparell"] = {str(self.source_app.id): "max"}
        schema["desempat"] = [
            {
                "camps": ["E_total"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "mode_seleccio_exercicis": "per_aparell_override",
                "exercicis_per_aparell": {str(self.source_app.id): {"mode": "millor_1"}},
                "agregacio_exercicis_per_aparell": {str(self.source_app.id): "max"},
                "scope": {"aparells": {"mode": "seleccionar", "ids": [self.source_app.id]}},
            }
        ]
        schema["puntuacio"]["victories"] = {
            "punts_victoria": 1,
            "punts_empat": 0.5,
            "sense_nota_mode": "skip",
            "mode_camps": "agregat",
            "mode_exercicis": "agregat",
            "mode_seleccio_exercicis_camps_separats": "per_camp",
            "agregacio_victories_camps": "sum",
            "agregacio_victories_exercicis": "sum",
            "desempat_comparacio": [
                {
                    "camps": ["E_total"],
                    "agregacio_camps": "hereta",
                    "ordre": "desc",
                    "mode_seleccio_exercicis": "per_aparell_override",
                    "exercicis_per_aparell": {str(self.source_app.id): {"mode": "millor_1"}},
                    "agregacio_exercicis_per_aparell": {str(self.source_app.id): "max"},
                    "scope": {"exercicis": {"mode": "tots"}},
                }
            ],
        }

        schema_tpl, warnings = _schema_to_template_schema(self.comp_source, schema)
        self.assertEqual(
            warnings,
            ["equips.assignment_source.mode='native' detectat; es normalitza al context Base."],
        )
        punt_tpl = schema_tpl.get("puntuacio") or {}
        self.assertEqual(
            punt_tpl.get("agregacio_exercicis_per_aparell"),
            {self.app.codi: "max"},
        )

        self._ensure_native_equip_context(self.comp_target)
        schema_local, mapping_warnings, mapping = _template_schema_to_competicio_schema(self.comp_target, schema_tpl)
        self.assertFalse(mapping_warnings)
        self.assertEqual(mapping.get(self.app.codi), self.target_app.id)
        punt_local = schema_local.get("puntuacio") or {}
        self.assertEqual(
            punt_local.get("agregacio_exercicis_per_aparell"),
            {str(self.target_app.id): "max"},
        )
        tie_local = (schema_local.get("desempat") or [])[0] or {}
        self.assertEqual(
            tie_local.get("agregacio_exercicis_per_aparell"),
            {str(self.target_app.id): "max"},
        )
        victory_tie_local = ((((punt_local.get("victories") or {}).get("desempat_comparacio")) or [])[0] or {})
        self.assertEqual(
            victory_tie_local.get("agregacio_exercicis_per_aparell"),
            {str(self.target_app.id): "max"},
        )
        self.assertEqual(
            ((((punt_local.get("victories") or {}).get("desempat_comparacio")) or [])[0] or {}).get("camps"),
            ["E_total"],
        )

    def test_template_schema_helpers_roundtrip_presentacio_detall_columns(self):
        schema = {
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.source_app.id]},
                "camps_per_aparell": {str(self.source_app.id): ["E_total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
            },
            "equips": {
                "context_code": "ctx-finals",
                "assignment_source": {"mode": "context", "context_code": "ctx-finals", "fallback": "native"},
                "team_mode": "derived_from_individual",
            },
            "presentacio": {
                "columnes": [
                    {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                ],
                "detall": {
                    "enabled": True,
                    "default_open": True,
                    "sections": [
                        {
                            "type": "members_table",
                            "label": "Detall",
                            "columns": [
                                {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                                {
                                    "type": "raw",
                                    "key": "detail_exec",
                                    "label": "Execucio",
                                    "align": "right",
                                    "decimals": 3,
                                    "source": {
                                        "aparell_id": self.source_app.id,
                                        "exercici": 1,
                                        "camp": "E_total",
                                        "jutges": {"ids": []},
                                    },
                                },
                            ],
                        }
                    ],
                },
            },
        }

        schema_tpl, warnings = _schema_to_template_schema(self.comp_source, schema)
        self.assertFalse(warnings)
        detail_section_tpl = (((((schema_tpl.get("presentacio") or {}).get("detall") or {}).get("sections")) or [])[0] or {})
        detail_tpl = detail_section_tpl.get("columns") or []
        self.assertEqual(detail_tpl[1]["source"]["aparell_codi"], self.app.codi)

        EquipContext.objects.create(competicio=self.comp_target, code="ctx-finals", nom="Finals")
        schema_local, mapping_warnings, mapping = _template_schema_to_competicio_schema(self.comp_target, schema_tpl)
        self.assertFalse(mapping_warnings)
        self.assertEqual(mapping.get(self.app.codi), self.target_app.id)
        detail_section_local = (((((schema_local.get("presentacio") or {}).get("detall") or {}).get("sections")) or [])[0] or {})
        detail_local = detail_section_local.get("columns") or []
        self.assertEqual(detail_local[1]["source"]["aparell_id"], self.target_app.id)

    def test_template_schema_helpers_keep_assignment_source_context_authoritative(self):
        self._ensure_native_equip_context(self.comp_source)
        self._ensure_native_equip_context(self.comp_target)
        EquipContext.objects.create(competicio=self.comp_source, code="ctx-finals", nom="Finals")
        EquipContext.objects.create(competicio=self.comp_target, code="ctx-finals", nom="Finals")

        schema = {
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.source_app.id]},
                "camps_per_aparell": {str(self.source_app.id): ["E_total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
            },
            "equips": {
                "context_code": "native",
                "assignment_source": {"mode": "context", "context_code": "ctx-finals", "fallback": "native"},
                "team_mode": "derived_from_individual",
            },
        }

        schema_tpl, warnings = _schema_to_template_schema(self.comp_source, schema)
        self.assertFalse(warnings)
        self.assertEqual((schema_tpl.get("equips") or {}).get("context_code"), "ctx-finals")
        self.assertEqual(
            (((schema_tpl.get("equips") or {}).get("assignment_source")) or {}).get("context_code"),
            "ctx-finals",
        )

        schema_local, mapping_warnings, _mapping = _template_schema_to_competicio_schema(self.comp_target, schema_tpl)
        self.assertFalse(mapping_warnings)
        self.assertEqual((schema_local.get("equips") or {}).get("context_code"), "ctx-finals")
        self.assertEqual(
            (((schema_local.get("equips") or {}).get("assignment_source")) or {}).get("context_code"),
            "ctx-finals",
        )

    def test_global_template_can_be_saved_validated_and_applied(self):
        save_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Rapid"},
        )
        self.assertEqual(save_res.status_code, 200)
        save_body = save_res.json()
        self.assertTrue(save_body.get("ok"))
        template_id = save_body.get("template", {}).get("id")
        self.assertTrue(template_id)

        tpl = ClassificacioTemplateGlobal.objects.get(pk=template_id)
        tpl_schema = (tpl.payload or {}).get("schema") or {}
        self.assertEqual(
            ((tpl_schema.get("puntuacio") or {}).get("aparells") or {}).get("ids"),
            [self.app.codi],
        )

        validate_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_validate",
            self.comp_target.id,
            {"template_id": template_id},
        )
        self.assertEqual(validate_res.status_code, 200)
        validate_body = validate_res.json()
        self.assertTrue(validate_body.get("compatible"))

        apply_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {"template_id": template_id, "nom": "Aplicada Target", "activa": False},
        )
        self.assertEqual(apply_res.status_code, 200)
        apply_body = apply_res.json()
        self.assertTrue(apply_body.get("ok"))
        cfg = apply_body.get("cfg") or {}
        self.assertEqual(cfg.get("nom"), "Aplicada Target")

        punt = ((cfg.get("schema") or {}).get("puntuacio") or {})
        self.assertEqual((punt.get("aparells") or {}).get("ids"), [self.target_app.id])
        self.assertEqual((punt.get("camps_per_aparell") or {}).get(str(self.target_app.id)), ["E_total"])

        tpl.refresh_from_db()
        self.assertEqual(tpl.uses_count, 1)
        self.assertIsNotNone(tpl.last_used_at)

    def test_template_apply_requires_ack_for_non_strict_fallback(self):
        save_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Ack"},
        )
        self.assertEqual(save_res.status_code, 200)
        template_id = save_res.json().get("template", {}).get("id")
        self.assertTrue(template_id)

        res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {
                "template_id": template_id,
                "nom": "Aplicada sense ack",
                "activa": False,
                "fallback_mode": "assistit",
            },
        )
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertEqual(body.get("phase"), "assistit")
        self.assertIn("confirmar", str(body.get("error", "")).lower())

    def test_template_apply_fallback_chain_strict_assistit_force(self):
        save_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Fallback Chain"},
        )
        self.assertEqual(save_res.status_code, 200)
        template_id = save_res.json().get("template", {}).get("id")
        self.assertTrue(template_id)

        tpl = ClassificacioTemplateGlobal.objects.get(pk=template_id)
        payload = json.loads(json.dumps(tpl.payload or {}))
        schema = payload.get("schema") or {}
        schema["particions"] = ["categoria"]
        schema["particions_custom"] = {
            "categoria": {
                "mode": "custom",
                "grups": [
                    {"label": "Bloc 1", "values": ["ALEVI"]},
                    {"label": "Bloc 2", "values": ["ALEVI"]},
                ],
            }
        }
        payload["schema"] = schema
        tpl.payload = payload
        tpl.save(update_fields=["payload", "updated_at"])

        strict_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {
                "template_id": template_id,
                "nom": "Aplicada strict",
                "activa": False,
                "fallback_mode": "strict",
            },
        )
        self.assertEqual(strict_res.status_code, 400)
        strict_body = strict_res.json()
        self.assertFalse(strict_body.get("compatible"))
        self.assertEqual(strict_body.get("phase"), "strict")
        self.assertEqual(strict_body.get("next_fallback"), "assistit")
        self.assertTrue(strict_body.get("can_try_next"))

        assistit_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {
                "template_id": template_id,
                "nom": "Aplicada assistit",
                "activa": False,
                "fallback_mode": "assistit",
                "ack_warning": True,
            },
        )
        self.assertEqual(assistit_res.status_code, 400)
        assistit_body = assistit_res.json()
        self.assertFalse(assistit_body.get("compatible"))
        self.assertEqual(assistit_body.get("phase"), "assistit")
        self.assertEqual(assistit_body.get("next_fallback"), "force")
        self.assertTrue(assistit_body.get("can_try_next"))

        force_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {
                "template_id": template_id,
                "nom": "Aplicada force",
                "activa": False,
                "fallback_mode": "force",
                "ack_warning": True,
            },
        )
        self.assertEqual(force_res.status_code, 200)
        force_body = force_res.json()
        self.assertTrue(force_body.get("ok"))
        cfg = force_body.get("cfg") or {}
        self.assertEqual(cfg.get("nom"), "Aplicada force")
        self.assertEqual(
            (((cfg.get("schema") or {}).get("puntuacio") or {}).get("aparells") or {}).get("ids"),
            [self.target_app.id],
        )
        self.assertEqual(
            (((cfg.get("schema") or {}).get("puntuacio") or {}).get("camps_per_aparell") or {}).get(str(self.target_app.id)),
            ["total"],
        )

    def test_global_template_apply_persists_mode_resolution_for_team_schema(self):
        self._ensure_native_equip_context(self.comp_source)
        self._ensure_native_equip_context(self.comp_target)
        team_cfg = ClassificacioConfig.objects.create(
            competicio=self.comp_source,
            nom="Cfg Source Team",
            activa=True,
            ordre=2,
            tipus="equips",
            schema={
                **json.loads(json.dumps(self.cfg_source.schema or {})),
                "equips": {
                    "context_code": "native",
                    "assignment_source": {"mode": "context", "context_code": "native", "fallback": "native"},
                    "team_mode": "derived_from_individual",
                },
            },
        )

        save_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": team_cfg.id, "nom": "TPL Team Mode Resolution"},
        )
        self.assertEqual(save_res.status_code, 200)
        template_id = save_res.json().get("template", {}).get("id")
        self.assertTrue(template_id)

        apply_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {"template_id": template_id, "nom": "Aplicada Team", "activa": False},
        )
        self.assertEqual(apply_res.status_code, 200)
        cfg = apply_res.json().get("cfg") or {}
        mode_resolution = (((cfg.get("schema") or {}).get("equips") or {}).get("mode_resolution")) or {}
        self.assertTrue(mode_resolution.get("resolved_at"))
        self.assertIn("eligible_team_app_ids_at_save", mode_resolution)

        saved_cfg = ClassificacioConfig.objects.get(pk=cfg.get("id"))
        self.assertTrue((((saved_cfg.schema.get("equips") or {}).get("mode_resolution")) or {}).get("resolved_at"))

    def test_template_validation_distinguishes_strict_vs_assistit_for_missing_team_context(self):
        schema_tpl, warnings = _schema_to_template_schema(self.comp_source, self.cfg_source.schema or {})
        self.assertEqual(
            warnings,
            ["equips.assignment_source.mode='native' detectat; es normalitza al context Base."],
        )
        schema_tpl["equips"] = {
            "context_code": "ctx-finals",
            "assignment_source": {"mode": "context", "context_code": "ctx-finals", "fallback": "native"},
            "team_mode": "derived_from_individual",
            "particions_manuals": [
                {"key": "manual_1", "label": "Bloc A", "equips_noms": ["Equip Finals"]},
            ],
        }
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="TPL Missing Context",
            slug="tpl-missing-context",
            tipus="equips",
            activa=True,
            payload={"schema": schema_tpl},
            requirements={},
            created_by=self.editor_user,
        )

        strict_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_validate",
            self.comp_target.id,
            {"template_id": tpl.id, "fallback_mode": "strict"},
        )
        self.assertEqual(strict_res.status_code, 200)
        strict_body = strict_res.json()
        self.assertFalse(strict_body.get("compatible"))
        self.assertTrue(any("equips.context_code no existeix" in err for err in strict_body.get("blocking_errors", [])))
        self.assertTrue(any("Equip Finals" in err for err in strict_body.get("blocking_errors", [])))

        assistit_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_validate",
            self.comp_target.id,
            {"template_id": tpl.id, "fallback_mode": "assistit"},
        )
        self.assertEqual(assistit_res.status_code, 200)
        assistit_body = assistit_res.json()
        self.assertTrue(assistit_body.get("compatible"))
        self.assertTrue(assistit_body.get("portable"))
        self.assertTrue(assistit_body.get("adaptable"))
        self.assertTrue(any("context Base" in msg for msg in assistit_body.get("warnings", [])))
        self.assertTrue(any("Equip Finals" in msg for msg in assistit_body.get("warnings", [])))
        self.assertTrue(any("equips.context_code" in msg for msg in assistit_body.get("dropped_rules", [])))

    def test_template_list_only_shows_owner_templates(self):
        own = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Owner"},
        )
        self.assertEqual(own.status_code, 200)
        own_id = own.json().get("template", {}).get("id")
        self.assertTrue(own_id)

        foreign = self._post_json_as(
            self.other_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Foreign"},
        )
        self.assertEqual(foreign.status_code, 200)
        foreign_id = foreign.json().get("template", {}).get("id")
        self.assertTrue(foreign_id)

        self.client.force_login(self.editor_user)
        list_url = reverse("classificacio_template_list", kwargs={"pk": self.comp_source.id})
        res = self.client.get(list_url)
        self.assertEqual(res.status_code, 200)
        ids = {int(t["id"]) for t in (res.json().get("templates") or [])}
        self.assertIn(int(own_id), ids)
        self.assertNotIn(int(foreign_id), ids)

    def test_cannot_use_or_update_foreign_template_by_id(self):
        foreign = self._post_json_as(
            self.other_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "nom": "TPL Foreign Locked"},
        )
        self.assertEqual(foreign.status_code, 200)
        foreign_id = foreign.json().get("template", {}).get("id")
        self.assertTrue(foreign_id)

        validate_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_validate",
            self.comp_target.id,
            {"template_id": foreign_id},
        )
        self.assertEqual(validate_res.status_code, 404)

        apply_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_apply",
            self.comp_target.id,
            {"template_id": foreign_id, "nom": "No hauria d'aplicar"},
        )
        self.assertEqual(apply_res.status_code, 404)

        update_res = self._post_json_as(
            self.editor_user,
            "classificacio_template_save",
            self.comp_source.id,
            {"cfg_id": self.cfg_source.id, "template_id": foreign_id, "nom": "No hauria d'editar"},
        )
        self.assertEqual(update_res.status_code, 404)

    def test_classificacions_home_renders_builder_json_contract(self):
        self.client.force_login(self.editor_user)
        url = reverse("classificacions_home", kwargs={"pk": self.comp_source.id})
        res = self.client.get(url)
        content = res.content.decode("utf-8")
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, 'id="can-manage-global-templates"')
        self.assertContains(res, 'id="builder-save-url"')
        self.assertContains(res, 'id="builder-delete-url-pattern"')
        self.assertContains(res, 'id="builder-preview-url-pattern"')
        self.assertContains(res, 'id="builder-enable-template-library"')
        self.assertContains(res, 'id="builder-can-preview"')
        self.assertContains(res, 'id="victoryConfigBox"')
        self.assertContains(res, 'id="sVictoryModeCamps"')
        self.assertContains(res, 'id="sVictoryModeExercicis"')
        self.assertContains(res, 'id="puntuacioSummaryText"')
        self.assertContains(res, 'class="builder-summary-box__text"')
        self.assertNotContains(res, 'id="exSelectionSummary"')
        self.assertContains(res, 'id="candidateScopeHint"')
        self.assertContains(res, 'id="classifHelpDrawer"')
        self.assertContains(res, 'id="classif-builder-back-to-top"')
        self.assertContains(res, "classificacions_builder_help.css")
        self.assertContains(res, "classificacions_builder_help.js")
        self.assertContains(res, 'data-help-key="global_overview"')
        self.assertContains(res, 'data-help-key="victories_overview"')
        self.assertNotContains(res, '<option value="entitat">Per entitat</option>', html=True)
        self.assertContains(res, 'id="appStaleBanner"')
        self.assertContains(res, "function pruneSchemaAppReferences(schema, allowedIds)")
        self.assertContains(res, "function renderAppStaleWarningBanner(schema, selectedIds)")
        self.assertContains(res, "function filterRawColumnsByAllowedApps(rawCols, allowedIds)")
        self.assertContains(res, "function pruneDetailSectionsByAllowedApps(rawSections, allowedIds)")
        self.assertContains(res, "const selectedCompatibleIds = sanitizeCompatibleAppIds(idsBase);")
        self.assertContains(res, 'buildAparellChecks(selectedCompatibleIds, { includeStale: false });')
        self.assertContains(res, 'const selected = getSingleCompatibleAppId(selectedAppId);')
        self.assertContains(res, '<option value="" ${selected ? "" : "selected"}>Selecciona aparell</option>')
        self.assertContains(res, 'const appIds = sanitizeCompatibleAppIds(getCurrentBuilderAppIds(schema));')
        self.assertContains(res, "renderAppStaleWarningBanner(schema, selectedCompatibleIds);")
        self.assertContains(res, "refreshTipusUI({ includeStale: false, dropInvalidSelection: true });")
        self.assertContains(res, 'function runSafeHydrationRender(label, renderFn)')
        self.assertContains(res, 'runSafeHydrationRender("columnes", () => {')
        self.assertContains(res, 'runSafeHydrationRender("desempat", () => {')
        self.assertContains(res, 'runSafeHydrationRender("per aparell", () => {')
        self.assertContains(res, 'state.rehydrationIssues = [];')
        self.assertEqual(content.count("function buildTieAppScopeOptionsHTML("), 1)
        self.assertContains(res, "function _buildPretractamentSegment(punt, perAppEntries)")
        self.assertContains(res, "function _buildScoreSelectionSegment({")
        self.assertContains(res, "function _buildVictoriesComparisonSegment(victoriesCfg)")
        self.assertNotContains(res, "function buildPuntuacioLiveSummary({")
        self.assertContains(res, "Agregació camps: Mateixa que puntuació")
        self.assertContains(res, "Base: Mateixa que puntuació")
        self.assertContains(res, "Selecció ex: Mateixa que puntuació")
        self.assertContains(res, "Tractament: Mateix que puntuació")
        self.assertContains(res, "Mateixos aparells que la puntuació")
        self.assertNotContains(res, "Hereta (puntuació)")
        self.assertNotContains(res, "Hereta (tots)")
        self.assertContains(res, "function previewRenderTeamRawDetailCell(v, col)")
        self.assertContains(res, "team-raw-summary")
        self.assertContains(res, 'v._kind === "team_raw_detail"')
        self.assertContains(res, 'id="previewBox"')
        self.assertContains(res, "builder-preview-shell")
        self.assertContains(res, 'id="detailConfigAlert"')
        self.assertContains(res, 'id="eqAssignmentContextHint"')
        self.assertContains(res, 'id="saveMsg"')
        self.assertContains(res, "En equips derivats, les columnes de camp mostren un resum i el detall per membres de l'equip.")
        self.assertContains(res, "En equips amb nota nativa, les columnes de camp mostren el valor d'equip i només són representables per aparells d'equip.")

    def test_classificacions_home_exposes_contextual_filter_copy_contract(self):
        self.client.force_login(self.editor_user)
        url = reverse("classificacions_home", kwargs={"pk": self.comp_source.id})
        res = self.client.get(url)

        self.assertContains(res, "Entren només els participants que compleixin aquests filtres abans de calcular resultats o mostrar particions.")
        self.assertContains(res, "Primer es filtren participants i després s'agrega el resultat per entitat.")
        self.assertContains(res, "Primer es filtren membres i després es calcula l'equip amb els membres resultants.")
        self.assertContains(res, "Un equip només entra si tots els seus membres compleixen aquests filtres abans de calcular la nota nativa d'equip.")

    def test_classificacions_home_preserves_legacy_field_refs_for_builder_rehydration(self):
        schema = json.loads(json.dumps(self.cfg_source.schema or {}))
        schema["puntuacio"]["camps_per_aparell"] = {
            str(self.source_app.id): ["E_total", "ex"],
        }
        schema["desempat"] = [
            {
                "camp": "ex",
                "camps": ["ex"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "hereta"},
                    "exercicis": {"mode": "hereta"},
                },
            },
            {
                "camp": "E_total",
                "camps": ["E_total"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "scope": {
                    "aparells": {"mode": "hereta"},
                    "exercicis": {"mode": "hereta"},
                },
            },
        ]
        schema["puntuacio"]["victories"] = {
            "punts_victoria": 1,
            "punts_empat": 0.5,
            "sense_nota_mode": "skip",
            "mode_camps": "agregat",
            "mode_exercicis": "agregat",
            "mode_seleccio_exercicis_camps_separats": "per_camp",
            "agregacio_victories_camps": "sum",
            "agregacio_victories_exercicis": "sum",
            "desempat_comparacio": [
                {
                    "camp": "ex",
                    "camps": ["ex"],
                    "agregacio_camps": "hereta",
                    "ordre": "desc",
                    "scope": {"exercicis": {"mode": "hereta"}},
                },
                {
                    "camp": "E_total",
                    "camps": ["E_total"],
                    "agregacio_camps": "hereta",
                    "ordre": "desc",
                    "scope": {"exercicis": {"mode": "hereta"}},
                },
            ],
        }
        schema["presentacio"]["columnes"] = [
            {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
            {
                "type": "raw",
                "key": "raw_valid",
                "label": "Valid",
                "align": "right",
                "decimals": 3,
                "source": {
                    "aparell_id": self.source_app.id,
                    "exercici": 1,
                    "camp": "E_total",
                    "jutges": {"ids": []},
                },
            },
            {
                "type": "raw",
                "key": "raw_legacy",
                "label": "Legacy",
                "align": "right",
                "decimals": 3,
                "source": {
                    "aparell_id": self.source_app.id,
                    "exercici": 1,
                    "camp": "ex",
                    "jutges": {"ids": []},
                },
            },
        ]
        self.cfg_source.schema = schema
        self.cfg_source.save(update_fields=["schema"])

        self.client.force_login(self.editor_user)
        url = reverse("classificacions_home", kwargs={"pk": self.comp_source.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)

        cfg_payload = next(cfg for cfg in res.context["cfgs"] if cfg["id"] == self.cfg_source.id)
        hydrated = cfg_payload["schema"]
        status = (res.context["cfg_status"] or {}).get(str(self.cfg_source.id)) or {}
        punt = hydrated.get("puntuacio") or {}
        self.assertNotIn("camp", punt)
        self.assertNotIn("agregacio", punt)
        self.assertNotIn("best_n", punt)
        self.assertEqual((punt.get("camps_per_aparell") or {}).get(str(self.source_app.id)), ["E_total", "ex"])

        desempat = hydrated.get("desempat") or []
        self.assertEqual(len(desempat), 2)
        self.assertEqual(desempat[0].get("camps"), ["ex"])
        self.assertEqual(desempat[1].get("camps"), ["E_total"])

        compare = (((punt.get("victories") or {}).get("desempat_comparacio")) or [])
        self.assertEqual(len(compare), 2)
        self.assertEqual(compare[0].get("camps"), ["ex"])
        self.assertEqual(compare[1].get("camps"), ["E_total"])

        raw_columns = [c for c in ((hydrated.get("presentacio") or {}).get("columnes") or []) if c.get("type") == "raw"]
        self.assertEqual(len(raw_columns), 2)
        self.assertEqual({((col.get("source") or {}).get("camp")) for col in raw_columns}, {"E_total", "ex"})

        self.assertTrue(status.get("is_stale"))
        errors = status.get("compatibility_errors") or []
        self.assertTrue(any("puntuacio.camps_per_aparell" in err and "'ex' no existeix" in err for err in errors))
        self.assertTrue(any("desempat[0]" in err and "'ex' no existeix" in err for err in errors))
        self.assertTrue(any("presentacio.columnes" in err and "'ex' no existeix" in err for err in errors))

        self.cfg_source.refresh_from_db()
        original = self.cfg_source.schema or {}
        self.assertEqual((original.get("puntuacio") or {}).get("camps_per_aparell", {}).get(str(self.source_app.id)), ["E_total", "ex"])
        self.assertEqual(len(original.get("desempat") or []), 2)
        self.assertEqual(len((((original.get("puntuacio") or {}).get("victories") or {}).get("desempat_comparacio")) or []), 2)

    def test_classificacions_home_preserves_missing_app_refs_for_builder_rehydration(self):
        schema = json.loads(json.dumps(self.cfg_source.schema or {}))
        missing_app_id = self.source_app.id + 9999
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [self.source_app.id, missing_app_id]}
        schema["puntuacio"]["camps_per_aparell"] = {
            str(self.source_app.id): ["E_total"],
            str(missing_app_id): ["E_total"],
        }
        schema["puntuacio"]["exercicis_per_aparell"] = {
            str(self.source_app.id): {"mode": "tots"},
            str(missing_app_id): {"mode": "tots"},
        }
        schema["desempat"] = [
            {
                "camps": ["E_total"],
                "agregacio_camps": "hereta",
                "ordre": "desc",
                "scope": {"aparells": {"mode": "seleccionar", "ids": [missing_app_id]}},
            }
        ]
        schema["presentacio"]["columnes"] = [
            {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
            {
                "type": "raw",
                "key": "raw_missing_app",
                "label": "No App",
                "align": "right",
                "decimals": 3,
                "source": {"aparell_id": missing_app_id, "exercici": 1, "camp": "E_total", "jutges": {"ids": []}},
            },
        ]
        self.cfg_source.schema = schema
        self.cfg_source.save(update_fields=["schema"])

        self.client.force_login(self.editor_user)
        res = self.client.get(reverse("classificacions_home", kwargs={"pk": self.comp_source.id}))
        self.assertEqual(res.status_code, 200)

        cfg_payload = next(cfg for cfg in res.context["cfgs"] if cfg["id"] == self.cfg_source.id)
        hydrated = cfg_payload["schema"]
        status = (res.context["cfg_status"] or {}).get(str(self.cfg_source.id)) or {}
        punt = hydrated.get("puntuacio") or {}
        self.assertEqual((punt.get("aparells") or {}).get("ids"), [self.source_app.id, missing_app_id])
        self.assertEqual(set((punt.get("camps_per_aparell") or {}).keys()), {str(self.source_app.id), str(missing_app_id)})
        self.assertEqual(set((punt.get("exercicis_per_aparell") or {}).keys()), {str(self.source_app.id), str(missing_app_id)})
        self.assertEqual((((hydrated.get("desempat") or [])[0].get("scope") or {}).get("aparells") or {}).get("ids"), [missing_app_id])
        raw_columns = [c for c in ((hydrated.get("presentacio") or {}).get("columnes") or []) if c.get("type") == "raw"]
        self.assertEqual(len(raw_columns), 1)
        self.assertEqual((((raw_columns[0].get("source") or {}).get("aparell_id"))), missing_app_id)

        self.assertTrue(status.get("is_stale"))
        errors = status.get("compatibility_errors") or []
        self.assertTrue(any(f"puntuacio.aparells.ids: l'aparell {missing_app_id} no valid o no actiu." in err for err in errors))
        self.assertTrue(any("puntuacio.camps_per_aparell: aparell" in err and str(missing_app_id) in err for err in errors))
        self.assertTrue(any("presentacio.columnes[1] raw: aparell" in err and str(missing_app_id) in err for err in errors))


class GlobalClassificacioTemplateManagementTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username="global_tpl_owner",
            password="testpass123",
            email="global-tpl-owner@example.com",
        )
        self.other_user = User.objects.create_user(
            username="global_tpl_other",
            password="testpass123",
            email="global-tpl-other@example.com",
        )
        self.admin_user = User.objects.create_superuser(
            username="global_tpl_admin",
            password="testpass123",
            email="global-tpl-admin@example.com",
        )
        manager_group = Group.objects.get_or_create(name="competicions_manager")[0]
        self.user.groups.add(manager_group)
        self.other_user.groups.add(manager_group)
        self.app = self._create_aparell("TRAMP_GLOB", "Tramp Global", owner=self.user)
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "E", "label": "Execucio", "type": "number"},
                ],
                "computed": [
                    {"code": "TOTAL", "formula": "E"},
                ],
            },
        )
        self.team_app = self._create_aparell("SYNC_GLOB", "Sync Global", owner=self.user)
        self.team_app.competition_unit = Aparell.CompetitionUnit.TEAM
        self.team_app.save(update_fields=["competition_unit"])
        ScoringSchema.objects.create(
            aparell=self.team_app,
            schema={
                "meta": {"subject_mode": "team"},
                "fields": [
                    {
                        "code": "E",
                        "label": "Execucio",
                        "scope": "member",
                        "type": "matrix",
                        "shape": "judge_x_item",
                        "judges": {"count": 3},
                        "items": {"count": 5},
                    },
                    {
                        "code": "SYNC",
                        "label": "Sync",
                        "scope": "shared",
                        "type": "number",
                    },
                ],
                "computed": [
                    {
                        "code": "E_mem",
                        "label": "Execucio membre",
                        "formula": "row_custom_compute('E', '1 - x')",
                    },
                    {
                        "code": "E_by_judge",
                        "label": "Execucio by judge",
                        "formula": "row_custom_compute('E', '1 - x', return_mode='by_judge')",
                    },
                ],
            },
        )
        self.comp = self._create_competicio("Comp Global Templates")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.EDITOR,
            is_active=True,
        )

    def _build_global_schema_payload(self, app_id):
        schema = json.loads(json.dumps(DEFAULT_SCHEMA))
        schema["puntuacio"]["aparells"] = {"mode": "seleccionar", "ids": [app_id]}
        schema["puntuacio"]["camps_per_aparell"] = {str(app_id): ["total"]}
        schema["presentacio"]["columnes"] = [
            {"type": "builtin", "key": "posicio", "label": "#", "align": "left"},
            {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
            {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
        ]
        return schema

    def _build_global_native_team_schema_payload(self, app_id):
        schema = self._build_global_schema_payload(app_id)
        schema["equips"] = {
            "context_code": "native",
            "team_mode": "native_team",
        }
        return schema

    def test_owner_can_create_list_and_delete_global_template(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        payload = {
            "nom": "Plantilla Global 1",
            "slug": "plantilla-global-1",
            "activa": True,
            "tipus": "individual",
            "schema": self._build_global_schema_payload(self.app.id),
        }
        res = self.client.post(save_url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 200)
        body = res.json()
        self.assertTrue(body.get("ok"))
        cfg = body.get("cfg") or {}
        tpl = ClassificacioTemplateGlobal.objects.get(pk=cfg.get("id"))
        self.assertEqual(((tpl.payload or {}).get("schema") or {}).get("puntuacio", {}).get("aparells", {}).get("ids"), [self.app.codi])
        self.assertEqual(tpl.slug, "plantilla-global-1")

        list_url = reverse("classificacio_template_global_list")
        list_res = self.client.get(list_url)
        self.assertEqual(list_res.status_code, 200)
        self.assertContains(list_res, "Plantilla Global 1")

        delete_url = reverse("classificacio_template_global_delete", kwargs={"pk": tpl.id})
        delete_res = self.client.post(
            delete_url,
            data=json.dumps({}),
            content_type="application/json",
            HTTP_ACCEPT="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(delete_res.status_code, 200)
        self.assertFalse(ClassificacioTemplateGlobal.objects.filter(pk=tpl.id).exists())

    def test_global_template_save_roundtrips_presentacio_detall_schema(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_schema_payload(self.app.id)
        schema["equips"] = {
            "context_code": "ctx-finals",
            "assignment_source": {"mode": "context", "context_code": "ctx-finals", "fallback": "native"},
            "team_mode": "derived_from_individual",
        }
        schema["presentacio"]["detall"] = {
            "enabled": True,
            "default_open": True,
            "sections": [
                {
                    "type": "members_table",
                    "label": "Detall",
                    "aparell_id": self.app.id,
                    "columns": [
                        {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                        {
                            "type": "raw",
                            "key": "detail_total",
                            "label": "Total",
                            "align": "right",
                            "decimals": 3,
                            "source": {"aparell_id": self.app.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                        },
                    ],
                }
            ],
        }
        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Plantilla Global Detail",
                    "slug": "plantilla-global-detail",
                    "activa": True,
                    "tipus": "equips",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 200)
        body = res.json()
        self.assertTrue(body.get("ok"))

        cfg = body.get("cfg") or {}
        detail_ui_section = (((((cfg.get("schema") or {}).get("presentacio") or {}).get("detall") or {}).get("sections")) or [])[0]
        detail_ui = detail_ui_section["columns"]
        self.assertEqual(detail_ui_section["aparell_id"], self.app.id)
        self.assertEqual(detail_ui[1]["source"]["aparell_id"], self.app.id)

        tpl = ClassificacioTemplateGlobal.objects.get(pk=cfg.get("id"))
        detail_tpl = (((tpl.payload or {}).get("schema") or {}).get("presentacio") or {}).get("detall") or {}
        self.assertTrue(detail_tpl.get("enabled"))
        detail_tpl_section = (detail_tpl.get("sections") or [])[0] or {}
        detail_tpl_cols = detail_tpl_section.get("columns") or []
        self.assertEqual(detail_tpl_section.get("aparell_codi"), self.app.codi)
        self.assertEqual(detail_tpl_cols[1]["source"]["aparell_codi"], self.app.codi)
        self.assertIn("total", tpl.requirements.get("presentacio_raw_camps") or [])

    def test_global_template_save_returns_error_details_for_invalid_detail_section_field(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_schema_payload(self.app.id)
        schema["presentacio"]["detall"] = {
            "enabled": True,
            "sections": [
                {
                    "type": "exercise_table",
                    "label": "Exercicis",
                    "aparell_id": self.app.id,
                    "columns": [
                        {"type": "builtin", "key": "exercise_index", "label": "Ex.", "align": "left"},
                        {
                            "type": "raw",
                            "key": "detail_bad",
                            "label": "Camp invalid",
                            "align": "right",
                            "decimals": 3,
                            "source": {"aparell_id": self.app.id, "exercici": 1, "camp": "NO_EXISTEIX", "jutges": {"ids": []}},
                        },
                    ],
                }
            ],
        }
        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Plantilla Global Error Detail",
                    "slug": "plantilla-global-error-detail",
                    "activa": True,
                    "tipus": "individual",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 400)
        body = res.json()
        details = body.get("error_details") or []
        self.assertTrue(any(item.get("path") == "presentacio.detall.sections[0].columns[1].source.camp" for item in details))

    def test_global_template_save_defers_detail_exercise_range_validation_until_competicio(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_schema_payload(self.app.id)
        schema["presentacio"]["detall"] = {
            "enabled": True,
            "sections": [
                {
                    "type": "exercise_table",
                    "label": "Exercicis",
                    "aparell_id": self.app.id,
                    "columns": [
                        {"type": "builtin", "key": "exercise_index", "label": "Ex.", "align": "left"},
                        {
                            "type": "raw",
                            "key": "detail_total",
                            "label": "Total",
                            "align": "right",
                            "decimals": 3,
                            "source": {"aparell_id": self.app.id, "exercici": 99, "camp": "TOTAL", "jutges": {"ids": []}},
                        },
                    ],
                }
            ],
        }
        save_res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Plantilla Global Exercise Deferred",
                    "slug": "plantilla-global-exercise-deferred",
                    "activa": True,
                    "tipus": "individual",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(save_res.status_code, 200)
        tpl_id = (save_res.json().get("cfg") or {}).get("id")
        self.assertTrue(tpl_id)

        validate_res = self.client.post(
            reverse("classificacio_template_validate", kwargs={"pk": self.comp.id}),
            data=json.dumps({"template_id": tpl_id}),
            content_type="application/json",
        )
        self.assertEqual(validate_res.status_code, 200)
        validate_body = validate_res.json()
        self.assertFalse(validate_body.get("compatible"))
        self.assertTrue(
            any("source.exercici" in str(err or "") or "fora de rang" in str(err or "") for err in validate_body.get("blocking_errors", []))
        )

    def test_global_builder_context_exposes_displayable_member_fields_for_native_team(self):
        self.client.force_login(self.user)
        res = self.client.get(reverse("classificacio_template_global_create"))
        self.assertEqual(res.status_code, 200)

        options = res.context["aparell_field_options"][str(self.team_app.id)]
        by_code = {item["code"]: item for item in options}
        self.assertEqual(
            next(item for item in res.context["aparells"] if item["id"] == self.team_app.id)["competition_unit"],
            "team",
        )
        self.assertIn("E", by_code)
        self.assertFalse(by_code["E"]["scoreable"])
        self.assertTrue(by_code["E"]["member_dependent"])
        self.assertTrue(by_code["E"]["detail_displayable"])
        self.assertEqual(by_code["E"]["detail_display_kind"], "judge_rows")
        self.assertTrue(by_code["E_mem"]["detail_displayable"])
        self.assertEqual(by_code["E_mem"]["detail_display_kind"], "scalar")

    def test_global_template_save_accepts_native_team_team_members_table_display_only_member_field(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_native_team_schema_payload(self.team_app.id)
        schema["presentacio"]["detall"] = {
            "enabled": True,
            "sections": [
                {
                    "type": "team_members_table",
                    "label": "Notes per membre",
                    "aparell_id": self.team_app.id,
                    "columns": [
                        {
                            "type": "raw",
                            "key": "member_exec",
                            "label": "Exec",
                            "align": "right",
                            "decimals": 3,
                            "source": {
                                "aparell_id": self.team_app.id,
                                "exercise_mode": "fixed",
                                "exercici": 1,
                                "camp": "E",
                                "jutges": {"ids": []},
                            },
                        },
                    ],
                }
            ],
        }

        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Plantilla Global Team Member Detail",
                    "slug": "plantilla-global-team-member-detail",
                    "activa": True,
                    "tipus": "equips",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 200)
        body = res.json()
        self.assertTrue(body.get("ok"))
        detail_ui_section = (((((body.get("cfg") or {}).get("schema") or {}).get("presentacio") or {}).get("detall") or {}).get("sections") or [])[0]
        self.assertEqual((((detail_ui_section or {}).get("columns") or [])[0].get("source") or {}).get("exercise_mode"), "fixed")
        tpl = ClassificacioTemplateGlobal.objects.get(pk=(body.get("cfg") or {}).get("id"))
        detail_tpl_section = (((((tpl.payload or {}).get("schema") or {}).get("presentacio") or {}).get("detall") or {}).get("sections") or [])[0]
        self.assertEqual((((detail_tpl_section or {}).get("columns") or [])[0].get("source") or {}).get("exercise_mode"), "fixed")

    def test_global_template_save_infers_team_members_table_section_app_from_single_raw_app(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_native_team_schema_payload(self.team_app.id)
        schema["presentacio"]["detall"] = {
            "enabled": True,
            "sections": [
                {
                    "type": "team_members_table",
                    "label": "Notes per membre",
                    "columns": [
                        {
                            "type": "raw",
                            "key": "member_exec",
                            "label": "Exec",
                            "align": "right",
                            "decimals": 3,
                            "source": {"aparell_id": self.team_app.id, "exercici": 1, "camp": "E", "jutges": {"ids": []}},
                        },
                    ],
                }
            ],
        }

        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Plantilla Global Team Member Detail Inferred App",
                    "slug": "plantilla-global-team-member-detail-inferred-app",
                    "activa": True,
                    "tipus": "equips",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 200)
        body = res.json()
        self.assertTrue(body.get("ok"))
        section = (((((body.get("cfg") or {}).get("schema") or {}).get("presentacio") or {}).get("detall") or {}).get("sections") or [])[0]
        self.assertEqual((section or {}).get("aparell_id"), self.team_app.id)

    def test_global_template_save_rejects_native_team_team_members_table_shared_field(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_native_team_schema_payload(self.team_app.id)
        schema["presentacio"]["detall"] = {
            "enabled": True,
            "sections": [
                {
                    "type": "team_members_table",
                    "label": "Notes per membre",
                    "aparell_id": self.team_app.id,
                    "columns": [
                        {
                            "type": "raw",
                            "key": "team_sync",
                            "label": "Sync",
                            "align": "right",
                            "decimals": 3,
                            "source": {"aparell_id": self.team_app.id, "exercici": 1, "camp": "SYNC", "jutges": {"ids": []}},
                        },
                    ],
                }
            ],
        }

        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Plantilla Global Team Shared Reject",
                    "slug": "plantilla-global-team-shared-reject",
                    "activa": True,
                    "tipus": "equips",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 400)
        self.assertTrue(any("team_members_table" in err for err in (res.json().get("errors") or [])))

    def test_global_template_save_rejects_native_team_team_members_table_non_displayable_field(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_native_team_schema_payload(self.team_app.id)
        schema["presentacio"]["detall"] = {
            "enabled": True,
            "sections": [
                {
                    "type": "team_members_table",
                    "label": "Notes per membre",
                    "aparell_id": self.team_app.id,
                    "columns": [
                        {
                            "type": "raw",
                            "key": "member_exec_bad",
                            "label": "Exec by judge",
                            "align": "right",
                            "decimals": 3,
                            "source": {"aparell_id": self.team_app.id, "exercici": 1, "camp": "E_by_judge", "jutges": {"ids": []}},
                        },
                    ],
                }
            ],
        }

        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Plantilla Global Team Non Displayable Reject",
                    "slug": "plantilla-global-team-non-displayable-reject",
                    "activa": True,
                    "tipus": "equips",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 400)
        self.assertTrue(any("no es visualitzable a team_members_table" in err for err in (res.json().get("errors") or [])))

    def test_global_template_save_rejects_native_team_team_metrics_member_field(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_native_team_schema_payload(self.team_app.id)
        schema["presentacio"]["detall"] = {
            "enabled": True,
            "sections": [
                {
                    "type": "team_metrics",
                    "label": "Notes equip",
                    "aparell_id": self.team_app.id,
                    "columns": [
                        {
                            "type": "raw",
                            "key": "member_exec",
                            "label": "Exec",
                            "align": "right",
                            "decimals": 3,
                            "source": {"aparell_id": self.team_app.id, "exercici": 1, "camp": "E_mem", "jutges": {"ids": []}},
                        },
                    ],
                }
            ],
        }

        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Plantilla Global Team Metrics Reject",
                    "slug": "plantilla-global-team-metrics-reject",
                    "activa": True,
                    "tipus": "equips",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 400)
        self.assertTrue(any("team_metrics" in err for err in (res.json().get("errors") or [])))

    def test_global_template_save_rejects_native_team_main_column_member_field(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_native_team_schema_payload(self.team_app.id)
        schema["presentacio"]["columnes"] = [
            {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
            {
                "type": "raw",
                "key": "member_exec",
                "label": "Exec",
                "align": "right",
                "decimals": 3,
                "source": {"aparell_id": self.team_app.id, "exercici": 1, "camp": "E_mem", "jutges": {"ids": []}},
            },
        ]

        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Plantilla Global Team Main Reject",
                    "slug": "plantilla-global-team-main-reject",
                    "activa": True,
                    "tipus": "equips",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 400)
        self.assertTrue(any("camps individuals per membre" in err for err in (res.json().get("errors") or [])))

    def test_owner_list_hides_foreign_templates_and_admin_sees_both(self):
        own_tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Own",
            slug="tpl-own",
            tipus="individual",
            activa=True,
            payload={"schema": {}},
            requirements={},
            created_by=self.user,
        )
        foreign_tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Foreign",
            slug="tpl-foreign",
            tipus="individual",
            activa=True,
            payload={"schema": {}},
            requirements={},
            created_by=self.other_user,
        )

        list_url = reverse("classificacio_template_global_list")

        self.client.force_login(self.user)
        owner_res = self.client.get(list_url)
        self.assertContains(owner_res, own_tpl.nom)
        self.assertNotContains(owner_res, foreign_tpl.nom)

        self.client.force_login(self.admin_user)
        admin_res = self.client.get(list_url)
        self.assertContains(admin_res, own_tpl.nom)
        self.assertContains(admin_res, foreign_tpl.nom)

    def test_foreign_user_cannot_delete_template(self):
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Locked",
            slug="tpl-locked",
            tipus="individual",
            activa=True,
            payload={"schema": {}},
            requirements={},
            created_by=self.user,
        )
        self.client.force_login(self.other_user)
        delete_url = reverse("classificacio_template_global_delete", kwargs={"pk": tpl.id})
        res = self.client.post(delete_url, data=json.dumps({}), content_type="application/json", HTTP_ACCEPT="application/json")
        self.assertEqual(res.status_code, 404)

    def test_owner_can_update_global_template_and_version_increments(self):
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Editable",
            slug="tpl-editable",
            tipus="individual",
            activa=True,
            payload={"schema": {"puntuacio": {"aparells": {"mode": "seleccionar", "ids": [self.app.codi]}}}},
            requirements={},
            created_by=self.user,
        )
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        payload = {
            "id": tpl.id,
            "nom": "Tpl Editable V2",
            "slug": "tpl-editable-v2",
            "activa": False,
            "tipus": "individual",
            "schema": self._build_global_schema_payload(self.app.id),
        }
        res = self.client.post(save_url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(res.status_code, 200)
        tpl.refresh_from_db()
        self.assertEqual(tpl.nom, "Tpl Editable V2")
        self.assertEqual(tpl.slug, "tpl-editable-v2")
        self.assertFalse(tpl.activa)
        self.assertEqual(tpl.version, 2)

    def test_global_validation_rejects_invalid_fields(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_schema_payload(self.app.id)
        schema["particions_v2"] = [{"code": "custom_excel", "apply_mode": "all", "parent_values": []}]
        schema["particions"] = ["custom_excel"]
        schema["puntuacio"]["camps_per_aparell"] = {str(self.app.id): ["NOT_SCOREABLE"]}
        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Tpl Invalid",
                    "slug": "tpl-invalid",
                    "activa": True,
                    "tipus": "individual",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 400)
        body = res.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("camp no permes" in err for err in body.get("errors", [])))
        self.assertTrue(any("no es puntuable" in err for err in body.get("errors", [])))

    def test_global_template_save_tracks_extended_team_requirements(self):
        self.client.force_login(self.user)
        save_url = reverse("classificacio_template_global_save")
        schema = self._build_global_schema_payload(self.app.id)
        schema["puntuacio"]["exercise_selection_scope"] = "team_pool"
        schema["equips"] = {
            "context_code": "ctx-finals",
            "assignment_source": {"mode": "context", "context_code": "ctx-finals", "fallback": "native"},
            "team_mode": "derived_from_individual",
            "particions_manuals": [
                {"key": "manual_1", "label": "Bloc A", "equips_noms": ["Equip A", "Equip B"]},
            ],
        }
        res = self.client.post(
            save_url,
            data=json.dumps(
                {
                    "nom": "Tpl Equips Portable",
                    "slug": "tpl-equips-portable",
                    "activa": True,
                    "tipus": "equips",
                    "schema": schema,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 200)
        tpl = ClassificacioTemplateGlobal.objects.get(slug="tpl-equips-portable")
        req = tpl.requirements or {}
        self.assertEqual(req.get("tipus"), "equips")
        self.assertEqual(req.get("team_mode"), "derived_from_individual")
        self.assertEqual(req.get("context_code"), "ctx-finals")
        self.assertTrue(req.get("uses_manual_team_partitions"))
        self.assertTrue(req.get("uses_exercise_selection_scope"))
        self.assertEqual(req.get("exercise_selection_scope"), "team_pool")
        self.assertEqual(req.get("exercise_selection_scope_modes"), ["team_pool"])
        self.assertEqual(
            ((((tpl.payload or {}).get("schema") or {}).get("equips") or {}).get("particions_manuals") or [])[0].get("equips_noms"),
            ["Equip A", "Equip B"],
        )

    def test_global_edit_preserves_legacy_extra_fields(self):
        legacy_schema = {
            "particions": ["custom_excel"],
            "particions_v2": [{"code": "custom_excel", "apply_mode": "all", "parent_values": []}],
            "particions_custom": {
                "custom_excel": {
                    "mode": "custom",
                    "fallback_label": "Altres",
                    "grups": [{"key": "grp_1", "label": "Bloc X", "values": ["A"]}],
                }
            },
            "filtres": {"custom_excel_in": ["A"]},
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.app.codi]},
                "camps_per_aparell": {self.app.codi: ["total"]},
                "legacy_score_meta": {"origin": "legacy"},
            },
            "presentacio": {
                "legacy_presentacio_flag": True,
                "columnes": [
                    {"type": "builtin", "key": "posicio", "label": "#", "align": "left"},
                    {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                ]
            },
            "equips": {
                "legacy_equips_flag": "keep-me",
            },
            "legacy_root_blob": {"foo": "bar"},
        }
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Legacy",
            slug="tpl-legacy",
            tipus="individual",
            activa=True,
            payload={"schema": legacy_schema},
            requirements={},
            created_by=self.user,
        )
        self.client.force_login(self.user)
        res = self.client.post(
            reverse("classificacio_template_global_save"),
            data=json.dumps(
                {
                    "id": tpl.id,
                    "nom": "Tpl Legacy Updated",
                    "slug": "tpl-legacy-updated",
                    "activa": True,
                    "tipus": "individual",
                    "schema": self._build_global_schema_payload(self.app.id),
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 200)
        tpl.refresh_from_db()
        saved_schema = (tpl.payload or {}).get("schema") or {}
        self.assertEqual(saved_schema.get("filtres", {}).get("custom_excel_in"), ["A"])
        self.assertIn("custom_excel", saved_schema.get("particions", []))
        self.assertIn("custom_excel", saved_schema.get("particions_custom", {}))
        self.assertEqual(saved_schema.get("legacy_root_blob"), {"foo": "bar"})
        self.assertTrue((saved_schema.get("presentacio") or {}).get("legacy_presentacio_flag"))
        self.assertEqual((saved_schema.get("puntuacio") or {}).get("legacy_score_meta"), {"origin": "legacy"})
        self.assertEqual((saved_schema.get("equips") or {}).get("legacy_equips_flag"), "keep-me")

    def test_global_template_appears_in_competition_template_list(self):
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl For Competition",
            slug="tpl-for-competition",
            tipus="individual",
            activa=True,
            payload={"schema": {"puntuacio": {"aparells": {"mode": "seleccionar", "ids": [self.app.codi]}}}},
            requirements={},
            created_by=self.user,
        )
        self.client.force_login(self.user)
        url = reverse("classificacio_template_list", kwargs={"pk": self.comp.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        ids = {int(item["id"]) for item in (res.json().get("templates") or [])}
        self.assertIn(tpl.id, ids)

    def test_global_builder_create_renders_builder_json_contract(self):
        self.client.force_login(self.user)
        url = reverse("classificacio_template_global_create")
        res = self.client.get(url)
        content = res.content.decode("utf-8")
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, 'id="can-manage-global-templates"')
        self.assertContains(res, 'id="builder-save-url"')
        self.assertContains(res, 'id="builder-delete-url-pattern"')
        self.assertContains(res, 'id="builder-preview-url-pattern"')
        self.assertContains(res, 'id="builder-enable-template-library"')
        self.assertContains(res, 'id="builder-can-preview"')
        self.assertContains(res, 'id="builder-selected-id"')
        self.assertContains(res, 'id="builder-auto-add-new"')
        self.assertContains(res, 'id="victoryConfigBox"')
        self.assertContains(res, 'id="sVictoryModeCamps"')
        self.assertContains(res, 'id="sVictoryModeExercicis"')
        self.assertContains(res, 'id="puntuacioSummaryText"')
        self.assertContains(res, 'class="builder-summary-box__text"')
        self.assertNotContains(res, 'id="exSelectionSummary"')
        self.assertContains(res, 'id="candidateScopeHint"')
        self.assertContains(res, 'id="classifHelpDrawer"')
        self.assertContains(res, 'id="classif-builder-back-to-top"')
        self.assertContains(res, "classificacions_builder_help.css")
        self.assertContains(res, "classificacions_builder_help.js")
        self.assertContains(res, 'data-help-key="global_overview"')
        self.assertContains(res, 'data-help-key="desempat_overview"')
        self.assertNotContains(res, '<option value="entitat">Per entitat</option>', html=True)
        self.assertContains(res, 'id="appStaleBanner"')
        self.assertContains(res, "function pruneSchemaAppReferences(schema, allowedIds)")
        self.assertContains(res, "function renderAppStaleWarningBanner(schema, selectedIds)")
        self.assertContains(res, 'buildAparellChecks(selectedCompatibleIds, { includeStale: false });')
        self.assertContains(res, 'const selected = getSingleCompatibleAppId(selectedAppId);')
        self.assertContains(res, '<option value="" ${selected ? "" : "selected"}>Selecciona aparell</option>')
        self.assertContains(res, "refreshTipusUI({ includeStale: false, dropInvalidSelection: true });")
        self.assertContains(res, 'function runSafeHydrationRender(label, renderFn)')
        self.assertContains(res, 'runSafeHydrationRender("columnes", () => {')
        self.assertContains(res, 'runSafeHydrationRender("desempat", () => {')
        self.assertContains(res, 'runSafeHydrationRender("per aparell", () => {')
        self.assertContains(res, 'state.rehydrationIssues = [];')
        self.assertEqual(content.count("function buildTieAppScopeOptionsHTML("), 1)
        self.assertContains(res, "function _buildPretractamentSegment(punt, perAppEntries)")
        self.assertContains(res, "function _buildScoreSelectionSegment({")
        self.assertContains(res, "function _buildVictoriesComparisonSegment(victoriesCfg)")
        self.assertNotContains(res, "function buildPuntuacioLiveSummary({")
        self.assertContains(res, "function previewRenderTeamRawDetailCell(v, col)")
        self.assertContains(res, "En equips derivats, les columnes de camp mostren un resum i el detall per membres de l'equip.")
        self.assertContains(res, "En equips amb nota nativa, les columnes de camp mostren el valor d'equip i només són representables per aparells d'equip.")

    def test_admin_global_builder_edit_is_scoped_to_template_owner_catalog(self):
        own_tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Owner Scope",
            slug="tpl-owner-scope",
            tipus="individual",
            activa=True,
            payload={"schema": {}},
            requirements={},
            created_by=self.user,
        )
        foreign_app = self._create_aparell("TRAMP_OTHER", "Tramp Other", owner=self.other_user)
        ScoringSchema.objects.create(
            aparell=foreign_app,
            schema={
                "fields": [{"code": "E", "label": "Execucio", "type": "number"}],
                "computed": [{"code": "TOTAL", "formula": "E"}],
            },
        )
        foreign_tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Foreign Scope",
            slug="tpl-foreign-scope",
            tipus="individual",
            activa=True,
            payload={"schema": {"puntuacio": {"aparells": {"mode": "seleccionar", "ids": [foreign_app.codi]}}}},
            requirements={},
            created_by=self.other_user,
        )

        self.client.force_login(self.admin_user)
        url = reverse("classificacio_template_global_update", kwargs={"pk": foreign_tpl.id})
        res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, foreign_tpl.nom)
        self.assertNotContains(res, own_tpl.nom)
        self.assertContains(res, foreign_app.nom)
        self.assertNotContains(res, self.app.nom)

    def test_global_builder_edit_exposes_portable_team_context_choices(self):
        tpl = ClassificacioTemplateGlobal.objects.create(
            nom="Tpl Context Portable",
            slug="tpl-context-portable",
            tipus="equips",
            activa=True,
            payload={
                "schema": {
                    "puntuacio": {
                        "aparells": {"mode": "seleccionar", "ids": [self.app.codi]},
                        "camps_per_aparell": {self.app.codi: ["total"]},
                    },
                    "equips": {
                        "context_code": "ctx-finals",
                        "assignment_source": {"mode": "context", "context_code": "ctx-finals", "fallback": "native"},
                        "team_mode": "derived_from_individual",
                        "particions_manuals": [
                            {"key": "manual_1", "label": "Bloc A", "equips_noms": ["Equip A"]},
                        ],
                    },
                    "presentacio": {"columnes": []},
                }
            },
            requirements={},
            created_by=self.user,
        )
        self.client.force_login(self.user)
        res = self.client.get(reverse("classificacio_template_global_update", kwargs={"pk": tpl.id}))
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, "ctx-finals")
        self.assertContains(res, "Equip A")


class LiveClassificacionsRedisCacheTests(_BaseTrampoliDataMixin, TestCase):
    class FakeRedis:
        def __init__(self):
            self.store = {}

        def get(self, key):
            return self.store.get(key)

        def set(self, key, value, nx=False, ex=None):
            if nx and key in self.store:
                return False
            self.store[key] = value
            return True

        def delete(self, key):
            self.store.pop(key, None)
            return 1

    def setUp(self):
        self.comp = self._create_competicio("Comp Live Cache")
        self.app = self._create_aparell("TRAMP_LIVE_CACHE", "Tramp Live Cache")
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1, actiu=True)
        self.ins = self._create_inscripcio(self.comp, "Participant Cache", ordre=1)
        self.cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="General",
            activa=True,
            ordre=1,
            tipus="individual",
            schema=self._schema(),
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            inscripcio=self.ins,
            exercici=1,
            comp_aparell=self.comp_app,
            inputs={},
            outputs={},
            total=9.8,
        )
        self.token = PublicLiveToken.objects.create(
            competicio=self.comp,
            label="Pantalla cache",
            is_active=True,
        )
        User = get_user_model()
        self.user = User.objects.create_user(
            username="live_cache_user",
            password="testpass123",
            email="live-cache@example.com",
        )
        self.editor_user = User.objects.create_user(
            username="live_cache_editor",
            password="testpass123",
            email="live-cache-editor@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.READONLY,
            is_active=True,
        )
        CompeticioMembership.objects.create(
            user=self.editor_user,
            competicio=self.comp,
            role=CompeticioMembership.Role.CLASSIFICACIONS,
            is_active=True,
        )

    def _schema(self):
        return {
            "filtres": {},
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "exercicis_best_n": 1,
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
                "camp": "total",
                "agregacio": "sum",
                "best_n": 1,
            },
            "desempat": [],
            "presentacio": {
                "columnes": [
                    {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                ],
            },
        }

    def _public_url(self):
        return reverse("public_live_classificacions_data", kwargs={"token": self.token.id})

    def _internal_url(self):
        return reverse("classificacions_live_data", kwargs={"pk": self.comp.id})

    def _reorder_url(self):
        return reverse("classificacio_reorder", kwargs={"pk": self.comp.id})

    def _snapshot_payload(self):
        return {
            "ok": True,
            "changed": True,
            "stamp": timezone.now().isoformat(),
            "competicio": {"id": self.comp.id, "nom": self.comp.nom},
            "cfgs": [
                {
                    "id": self.cfg.id,
                    "nom": self.cfg.nom,
                    "tipus": self.cfg.tipus,
                    "columns": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                    ],
                    "parts": [
                        {
                            "particio": "global",
                            "rows": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}],
                        }
                    ],
                }
            ],
        }

    def _snapshot_blob(self, generated_at=None):
        payload = self._snapshot_payload()
        payload["generated_at"] = (generated_at or timezone.now()).isoformat()
        return json.dumps(payload)

    def test_internal_live_view_exposes_poll_ms_and_internal_data_url_bootstrap(self):
        self.client.force_login(self.user)
        res = self.client.get(reverse("classificacions_live", kwargs={"pk": self.comp.id}))

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.context["poll_ms"], 4000)
        self.assertFalse(res.context["is_public"])
        self.assertContains(res, 'id="poll-ms"', status_code=200)
        self.assertContains(
            res,
            reverse("classificacions_live_data", kwargs={"pk": self.comp.id}),
            status_code=200,
        )

    def test_loop_live_view_clamps_polling_params_and_uses_internal_data_url(self):
        self.client.force_login(self.user)
        res = self.client.get(
            reverse("classificacions_loop_live", kwargs={"pk": self.comp.id}),
            {
                "poll_ms": 5,
                "slide_ms": 999999,
                "rows": 1,
                "transition": "spin",
            },
        )

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res.context["poll_ms"], 1000)
        self.assertEqual(res.context["slide_ms"], 120000)
        self.assertEqual(res.context["rows_per_page"], 3)
        self.assertEqual(res.context["transition"], "fade")
        self.assertContains(res, 'id="loop-poll-ms"', status_code=200)
        self.assertContains(res, 'id="loop-slide-ms"', status_code=200)
        self.assertContains(res, 'id="loop-data-url"', status_code=200)
        self.assertContains(
            res,
            reverse("classificacions_live_data", kwargs={"pk": self.comp.id}),
            status_code=200,
        )

    def test_public_loop_live_exposes_public_data_url_and_media_capability(self):
        self.token.can_view_media = True
        self.token.save(update_fields=["can_view_media"])

        res = self.client.get(reverse("public_live_loop", kwargs={"token": self.token.id}))

        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.context["is_public"])
        self.assertTrue(res.context["public_token_can_view_media"])
        self.assertEqual(
            res.context["data_url"],
            f"http://testserver{reverse('public_live_classificacions_data', kwargs={'token': self.token.id})}",
        )
        self.assertContains(
            res,
            reverse("public_live_classificacions_data", kwargs={"token": self.token.id}),
            status_code=200,
        )

    def test_loop_live_shows_empty_state_when_no_active_classificacions(self):
        self.cfg.activa = False
        self.cfg.save(update_fields=["activa"])
        self.client.force_login(self.user)

        res = self.client.get(reverse("classificacions_loop_live", kwargs={"pk": self.comp.id}))

        self.assertEqual(res.status_code, 200)
        self.assertContains(
            res,
            "No hi ha cap classificacio activa. Quan n'hi hagi, apareixeran automaticament.",
            status_code=200,
        )

    def test_first_get_computes_and_second_get_uses_cache(self):
        fake_redis = self.FakeRedis()
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        cache_key = live_cache.live_cache_key(self.comp.id)
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views.classificacions.live.compute_classificacio", return_value=compute_result) as mocked_compute:
                res_1 = self.client.get(self._public_url())
                res_2 = self.client.get(self._public_url())

        self.assertEqual(res_1.status_code, 200)
        self.assertEqual(res_2.status_code, 200)
        self.assertEqual(mocked_compute.call_count, 1)
        self.assertIn(cache_key, fake_redis.store)
        self.assertEqual(res_1["X-Live-Cache"], "miss")
        self.assertEqual(res_2["X-Live-Cache"], "hit")
        self.assertEqual(
            res_2.json().get("cfgs", [])[0].get("parts", [])[0].get("rows", [])[0].get("participant"),
            "Participant Cache",
        )

    def test_public_and_internal_live_share_same_snapshot(self):
        fake_redis = self.FakeRedis()
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        self.client.force_login(self.user)
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views.classificacions.live.compute_classificacio", return_value=compute_result) as mocked_compute:
                public_res = self.client.get(self._public_url())
                internal_res = self.client.get(self._internal_url())

        self.assertEqual(public_res.status_code, 200)
        self.assertEqual(internal_res.status_code, 200)
        self.assertEqual(mocked_compute.call_count, 1)
        self.assertEqual(public_res["X-Live-Cache"], "miss")
        self.assertEqual(internal_res["X-Live-Cache"], "hit")
        self.assertIn("permissions", public_res.json())
        self.assertNotIn("permissions", internal_res.json())

    def test_since_is_served_from_cached_stamp_without_recompute(self):
        fake_redis = self.FakeRedis()
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views.classificacions.live.compute_classificacio", return_value=compute_result) as mocked_compute:
                first_res = self.client.get(self._public_url())
                stamp = first_res.json()["stamp"]
                second_res = self.client.get(self._public_url(), {"since": stamp})

        self.assertEqual(first_res.status_code, 200)
        self.assertEqual(second_res.status_code, 200)
        self.assertEqual(mocked_compute.call_count, 1)
        self.assertFalse(second_res.json()["changed"])
        self.assertEqual(second_res.json()["stamp"], stamp)
        self.assertEqual(second_res.json().get("permissions", {}).get("can_view_media"), False)

    def test_internal_since_is_served_from_cached_stamp_without_recompute(self):
        fake_redis = self.FakeRedis()
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        self.client.force_login(self.user)
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views.classificacions.live.compute_classificacio", return_value=compute_result) as mocked_compute:
                first_res = self.client.get(self._internal_url())
                stamp = first_res.json()["stamp"]
                second_res = self.client.get(self._internal_url(), {"since": stamp})

        self.assertEqual(first_res.status_code, 200)
        self.assertEqual(second_res.status_code, 200)
        self.assertEqual(mocked_compute.call_count, 1)
        self.assertFalse(second_res.json()["changed"])
        self.assertEqual(second_res.json()["stamp"], stamp)
        self.assertNotIn("permissions", second_res.json())

    def test_dirty_refresh_with_since_returns_changed_true_and_new_snapshot_stamp(self):
        fake_redis = self.FakeRedis()
        old_stamp = "2026-03-29T10:00:00+00:00"
        snapshot = self._snapshot_payload()
        snapshot["stamp"] = old_stamp
        snapshot["generated_at"] = (timezone.now() - timedelta(seconds=1)).isoformat()
        fake_redis.set(live_cache.live_cache_key(self.comp.id), json.dumps(snapshot))
        fake_redis.set(live_cache.live_dirty_key(self.comp.id), "dirty-1")

        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.9, "posicio": 1}]
        }

        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views.classificacions.live.compute_classificacio", return_value=compute_result):
                res = self.client.get(self._public_url(), {"since": old_stamp})

        self.assertEqual(res.status_code, 200)
        payload = res.json()
        self.assertTrue(payload["changed"])
        self.assertNotEqual(payload["stamp"], old_stamp)
        self.assertEqual(payload.get("cfgs", [])[0].get("parts", [])[0].get("rows", [])[0].get("punts"), 9.9)
        self.assertEqual(res["X-Live-Cache"], "refresh")

    def test_lock_contention_waits_for_snapshot_without_recompute(self):
        fake_redis = self.FakeRedis()
        fake_redis.set(live_cache.live_lock_key(self.comp.id), "busy")
        waited_snapshot = json.loads(self._snapshot_blob())
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.live_cache._wait_for_live_snapshot", return_value=waited_snapshot):
                with patch("competicions_trampoli.views.classificacions.live.compute_classificacio") as mocked_compute:
                    res = self.client.get(self._public_url())

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["X-Live-Cache"], "wait-hit")
        mocked_compute.assert_not_called()

    def test_stale_snapshot_is_served_when_refresh_lock_is_busy(self):
        fake_redis = self.FakeRedis()
        fake_redis.set(
            live_cache.live_cache_key(self.comp.id),
            self._snapshot_blob(generated_at=timezone.now() - timedelta(seconds=10)),
        )
        fake_redis.set(live_cache.live_lock_key(self.comp.id), "busy")
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views.classificacions.live.compute_classificacio") as mocked_compute:
                res = self.client.get(self._public_url())

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["X-Live-Cache"], "stale")
        mocked_compute.assert_not_called()

    def test_redis_failure_falls_back_to_direct_compute(self):
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        with patch(
            "competicions_trampoli.live_cache._live_redis_client",
            side_effect=RuntimeError("redis down"),
        ):
            with patch("competicions_trampoli.views.classificacions.live.compute_classificacio", return_value=compute_result) as mocked_compute:
                res = self.client.get(self._public_url())

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["X-Live-Cache"], "fallback")
        self.assertEqual(mocked_compute.call_count, 1)

    def test_fresh_snapshot_with_dirty_forces_refresh_and_clears_dirty(self):
        fake_redis = self.FakeRedis()
        fake_redis.set(live_cache.live_cache_key(self.comp.id), self._snapshot_blob())
        dirty_key = live_cache.live_dirty_key(self.comp.id)
        fake_redis.set(dirty_key, "dirty-1")
        compute_result = {
            "global": [{"participant": "Participant Cache", "punts": 9.8, "posicio": 1}]
        }
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with patch("competicions_trampoli.views.classificacions.live.compute_classificacio", return_value=compute_result) as mocked_compute:
                res = self.client.get(self._public_url())

        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["X-Live-Cache"], "refresh")
        self.assertEqual(mocked_compute.call_count, 1)
        self.assertNotIn(dirty_key, fake_redis.store)

    def test_dirty_marker_changed_during_refresh_is_preserved(self):
        fake_redis = self.FakeRedis()
        fake_redis.set(live_cache.live_cache_key(self.comp.id), self._snapshot_blob())
        dirty_key = live_cache.live_dirty_key(self.comp.id)
        fake_redis.set(dirty_key, "dirty-1")

        def compute_payload(competicio, since_raw=None):
            fake_redis.set(dirty_key, "dirty-2")
            return self._snapshot_payload()

        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            payload, source = live_cache.get_live_payload_cached(
                self.comp,
                compute_payload=compute_payload,
                since_raw=None,
            )

        self.assertEqual(source, "refresh")
        self.assertTrue(payload.get("ok"))
        self.assertEqual(fake_redis.get(dirty_key), "dirty-2")

    def test_scoreentry_signal_marks_dirty_after_commit(self):
        fake_redis = self.FakeRedis()
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with self.captureOnCommitCallbacks(execute=True):
                ScoreEntry.objects.create(
                    competicio=self.comp,
                    inscripcio=self.ins,
                    exercici=2,
                    comp_aparell=self.comp_app,
                    inputs={},
                    outputs={},
                    total=8.4,
                )

        self.assertIsNotNone(fake_redis.get(live_cache.live_dirty_key(self.comp.id)))

    def test_teamscoreentry_signal_marks_dirty_after_commit(self):
        team_ctx = EquipContext.objects.create(
            competicio=self.comp,
            code="parelles-live",
            nom="Parelles live",
        )
        self.app.competition_unit = Aparell.CompetitionUnit.TEAM
        self.app.save(update_fields=["competition_unit"])
        CompeticioAparellEquipContextSource.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            context=team_ctx,
        )
        equip = self._create_equip(self.comp, "Equip live", context=team_ctx)
        ins_b = self._create_inscripcio(self.comp, "Participant Cache 2", ordre=2)
        self._assign_equip(self.comp, self.ins, equip, context=team_ctx)
        self._assign_equip(self.comp, ins_b, equip, context=team_ctx)
        team_subjects, _issues = build_team_subjects_for_comp_aparell(self.comp, self.comp_app)
        team_subject_id = next(
            item["subject_id"]
            for item in team_subjects
            if int(item.get("equip_id") or 0) == equip.id
        )

        fake_redis = self.FakeRedis()
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with self.captureOnCommitCallbacks(execute=True):
                TeamScoreEntry.objects.create(
                    competicio=self.comp,
                    team_subject_id=team_subject_id,
                    exercici=1,
                    comp_aparell=self.comp_app,
                    inputs={},
                    outputs={},
                    total=8.4,
                )

        self.assertIsNotNone(fake_redis.get(live_cache.live_dirty_key(self.comp.id)))

    def test_teamscoreentry_change_refreshes_cached_live_snapshot(self):
        snapshot = json.loads(self._snapshot_blob())
        stamp = snapshot["stamp"]
        fake_redis = self.FakeRedis()
        fake_redis.set(live_cache.live_cache_key(self.comp.id), json.dumps(snapshot))

        team_ctx = EquipContext.objects.create(
            competicio=self.comp,
            code="parelles-live-cache",
            nom="Parelles live cache",
        )
        self.app.competition_unit = Aparell.CompetitionUnit.TEAM
        self.app.save(update_fields=["competition_unit"])
        CompeticioAparellEquipContextSource.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            context=team_ctx,
        )
        equip = self._create_equip(self.comp, "Equip live cache", context=team_ctx)
        ins_b = self._create_inscripcio(self.comp, "Participant Cache 2", ordre=2)
        self._assign_equip(self.comp, self.ins, equip, context=team_ctx)
        self._assign_equip(self.comp, ins_b, equip, context=team_ctx)
        team_subjects, _issues = build_team_subjects_for_comp_aparell(self.comp, self.comp_app)
        team_subject_id = next(
            item["subject_id"]
            for item in team_subjects
            if int(item.get("equip_id") or 0) == equip.id
        )

        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with self.captureOnCommitCallbacks(execute=True):
                TeamScoreEntry.objects.create(
                    competicio=self.comp,
                    team_subject_id=team_subject_id,
                    exercici=1,
                    comp_aparell=self.comp_app,
                    inputs={},
                    outputs={},
                    total=8.4,
                )
            res = self.client.get(self._public_url(), {"since": stamp})

        self.assertEqual(res.status_code, 200)
        self.assertNotEqual(res["X-Live-Cache"], "hit")
        self.assertTrue(res.json()["changed"])
        self.assertNotEqual(res.json()["stamp"], stamp)

    def test_classificacioconfig_signal_marks_dirty_after_commit(self):
        fake_redis = self.FakeRedis()
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with self.captureOnCommitCallbacks(execute=True):
                self.cfg.nom = "General Dirty"
                self.cfg.save(update_fields=["nom"])

        self.assertIsNotNone(fake_redis.get(live_cache.live_dirty_key(self.comp.id)))

    def test_classificacio_reorder_marks_dirty_after_bulk_update(self):
        fake_redis = self.FakeRedis()
        cfg_2 = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Segona",
            activa=True,
            ordre=2,
            tipus="individual",
            schema=self._schema(),
        )
        self.client.force_login(self.editor_user)
        with patch("competicions_trampoli.live_cache._live_redis_client", return_value=fake_redis):
            with self.captureOnCommitCallbacks(execute=True):
                res = self.client.post(
                    self._reorder_url(),
                    data=json.dumps({"order": [cfg_2.id, self.cfg.id]}),
                    content_type="application/json",
                )

        self.assertEqual(res.status_code, 200)
        self.assertIsNotNone(fake_redis.get(live_cache.live_dirty_key(self.comp.id)))
