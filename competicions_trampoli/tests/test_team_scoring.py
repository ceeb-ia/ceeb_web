import json
import re
from io import BytesIO, StringIO
from datetime import date, timedelta
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.sessions.backends.db import SessionStore
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Max
from django.test import RequestFactory, TestCase
from django.urls import Resolver404, resolve, reverse
from django.utils import timezone
from openpyxl import load_workbook

from ceeb_web.auth_groups import GLOBAL_AUTH_GROUPS

from .. import live_cache
from ..access import user_has_competicio_capability
from ..forms import CompeticioAparellForm
from ..models import (
    Competicio,
    Equip,
    EquipContext,
    GrupCompeticio,
    Inscripcio,
    InscripcioEquipAssignacio,
    InscripcioMedia,
)
from ..models.judging import (
    JudgeConversation,
    JudgeConversationMessage,
    JudgeDeviceToken,
    PublicLiveToken,
)
from ..models.classificacions import ClassificacioConfig, ClassificacioTemplateGlobal
from ..models.rotacions import (
    RotacioAssignacio,
    RotacioAssignacioGrup,
    RotacioAssignacioSerieEquip,
    RotacioEstacio,
    RotacioFranja,
)
from ..models.scoring import (
    ScoringSchema,
    ScoreEntry,
    ScoreEntryVideo,
    ScoreEntryVideoEvent,
    SerieEquip,
    SerieEquipItem,
    TeamScoreEntry,
    TeamCompetitiveSubject,
    TeamScoreEntryVideo,
    TeamScoreEntryVideoEvent,
)
from ..models.competicio import (
    Aparell,
    CompeticioAparell,
    CompeticioAparellEquipContextSource,
    InscripcioAparellExclusio,
)
from ..models import CompeticioMembership
from ..scoring_engine import ScoringEngine
from ..services.inscripcions.groups import renumber_groups_for_competicio
from ..services.inscripcions.sorting import (
    _split_custom_sort_tokens,
    sort_records_by_field_stable,
)
from ..services.inscripcions.history import (
    apply_inscripcions_history_snapshot,
    capture_inscripcions_history_snapshot,
)
from ..services.inscripcions.queries import (
    COLUMN_FILTER_EMPTY_TOKEN,
    _build_inscripcions_filtered_qs,
    build_inscripcions_sort_context_key,
    get_competicio_custom_sort_rank_map,
)
from ..services.classificacions.classificacio_templates import (
    normalize_particions_schema as _normalize_particions_schema,
    schema_to_template_schema as _schema_to_template_schema,
    template_schema_to_competicio_schema as _template_schema_to_competicio_schema_service,
)
from ..services.classificacions.builder import scoreable_codes_by_app_id as _scoreable_codes_by_app_id
from ..services.classificacions.compute import DEFAULT_SCHEMA, compute_classificacio
from ..services.classificacions.export import _normalize_excel_cell
from ..services.classificacions.partitions import normalize_schema_legacy_team_birth_partition
from ..services.classificacions.runtime import prepare_schema_for_persistence
from ..services.classificacions.validation import (
    build_metric_meta_for_comp_aparell as _build_metric_meta_for_comp_aparell,
    build_scoreable_meta_for_schema as _build_scoreable_meta_for_schema,
    validate_particions_schema as _validate_particions_schema,
    validate_schema_for_competicio as _validate_schema_for_competicio,
)
from ..views.classificacions.builder import ClassificacionsHome
from ..services.shared.competition_groups import (
    assign_groups_by_display_num,
    compact_competition_order_for_group,
    ensure_group_for_display_num,
    get_group_maps,
    get_group_participant_counts,
    get_out_of_program_group_ids,
    get_programmed_group_ids,
    group_label,
    move_inscripcio_to_group,
    next_group_display_num,
)
from ..services.scoring.schema_resolution import resolve_scoring_schema_for_comp_aparell
from ..services.scoring.team_scoring import (
    build_permission_label,
    build_team_subjects_for_comp_aparell,
    resolve_permission_runtime_entries,
    runtime_schema_for_comp_aparell,
)
from ..services.teams.team_series import safe_deactivate_empty_serie
from ..views.judge.admin import _member_slot_choices, _validate_permission_row
from ..templatetags.competicio_extras import (
    DEFAULT_COMPETITION_BACKGROUND,
    get_competicio_background_url_from_request,
)

from .base import _BaseTrampoliDataMixin


def _template_schema_to_competicio_schema(*args, **kwargs):
    schema_local, mapping_warnings, mapping, _compat_meta = _template_schema_to_competicio_schema_service(
        *args,
        **kwargs,
    )
    return schema_local, mapping_warnings, mapping


class TeamMemberTreatmentSchemaTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp member treatment")
        self.app = self._create_aparell("TEAMSC", "Team Schema")
        self.app.competition_unit = Aparell.CompetitionUnit.TEAM
        self.app.save(update_fields=["competition_unit"])
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1)

    def test_schema_accepts_member_treatment_on_member_number_field(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "member_treatment(E, agg='sum')"},
                ],
            },
        )
        schema.full_clean()

    def test_schema_accepts_member_treatment_on_member_scalar_computed(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "fields": [
                    {
                        "code": "E",
                        "label": "Exec",
                        "type": "matrix",
                        "shape": "judge_x_item",
                        "scope": "member",
                        "judges": {"count": 1},
                        "items": {"count": 2},
                    },
                ],
                "computed": [
                    {
                        "code": "E_MEMBER",
                        "label": "Exec membre",
                        "formula": "row_custom_compute('E', '1 - x', row_select='all', row_agg='sum', col_select='all', col_agg='sum')",
                    },
                    {
                        "code": "TOTAL",
                        "label": "Total",
                        "formula": "member_treatment(E_MEMBER, select='best_n', n=1, agg='sum')",
                    },
                ],
            },
        )
        schema.full_clean()

    def test_schema_rejects_member_treatment_on_unreduced_member_matrix(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "fields": [
                    {
                        "code": "E",
                        "label": "Exec",
                        "type": "matrix",
                        "shape": "judge_x_item",
                        "scope": "member",
                        "judges": {"count": 2},
                        "items": {"count": 3},
                    },
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "member_treatment(E, agg='sum')"},
                ],
            },
        )
        with self.assertRaises(ValidationError) as ctx:
            schema.full_clean()
        self.assertIn("member_scalar", str(ctx.exception))

    def test_schema_rejects_member_treatment_on_shared_field(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "member_treatment(SYNC, agg='sum')"},
                ],
            },
        )
        with self.assertRaises(ValidationError) as ctx:
            schema.full_clean()
        self.assertIn("member_scalar", str(ctx.exception))

    def test_schema_rejects_member_treatment_with_invalid_select(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "member_treatment(E, select='median_band', agg='sum')"},
                ],
            },
        )
        with self.assertRaises(ValidationError) as ctx:
            schema.full_clean()
        self.assertIn("member_treatment.select invalid", str(ctx.exception))

    def test_schema_rejects_member_treatment_with_invalid_agg(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "member_treatment(E, agg='median')"},
                ],
            },
        )
        with self.assertRaises(ValidationError) as ctx:
            schema.full_clean()
        self.assertIn("member_treatment.agg invalid", str(ctx.exception))

    def test_schema_rejects_member_treatment_missing_n_when_selector_requires_it(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "member_treatment(E, select='drop_extremes_until_n', agg='sum')"},
                ],
            },
        )
        with self.assertRaises(ValidationError) as ctx:
            schema.full_clean()
        self.assertIn("requereix n", str(ctx.exception))

    def test_schema_rejects_member_treatment_with_n_when_selector_does_not_use_it(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "member_treatment(E, select='drop_extremes', n=2, agg='sum')"},
                ],
            },
        )
        with self.assertRaises(ValidationError) as ctx:
            schema.full_clean()
        self.assertIn("no admet n", str(ctx.exception))

    def test_individual_app_rejects_member_treatment(self):
        app = self._create_aparell("INDSC", "Individual Schema")
        schema = ScoringSchema(
            aparell=app,
            schema={
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "member_treatment(E, agg='sum')"},
                ],
            },
        )
        with self.assertRaises(ValidationError) as ctx:
            schema.full_clean()
        self.assertIn("nomes es permes", str(ctx.exception))

    def test_runtime_schema_and_engine_support_member_treatment(self):
        schema = {
            "fields": [
                {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
            ],
            "computed": [
                {"code": "BEST_EXEC", "label": "Best exec", "formula": "member_treatment(E, select='best_n', n=1, agg='sum')"},
                {"code": "TOTAL", "label": "Total", "formula": "BEST_EXEC + SYNC"},
            ],
        }
        runtime_schema = runtime_schema_for_comp_aparell(schema, self.comp_app, member_count=2)
        best_exec_formula = next(
            c["formula"] for c in runtime_schema.get("computed", []) if c.get("code") == "BEST_EXEC"
        )
        self.assertIn("member_treatment", best_exec_formula)
        self.assertIn("E__m1", best_exec_formula)
        self.assertIn("E__m2", best_exec_formula)

        result = ScoringEngine(runtime_schema).compute(
            {
                "SYNC": 6.0,
                "E__m1": 8.1,
                "E__m2": 7.9,
            }
        )
        self.assertAlmostEqual(result.outputs["BEST_EXEC"], 8.1)
        self.assertAlmostEqual(result.total, 14.1)

    def test_runtime_member_treatment_wrappers_match_explicit_contract(self):
        schema = {
            "fields": [
                {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
            ],
            "computed": [
                {"code": "SUM_EXPL", "label": "Sum explicit", "formula": "member_treatment(E, agg='sum')"},
                {"code": "SUM_WRAP", "label": "Sum wrap", "formula": "members_sum(E)"},
                {"code": "AVG_EXPL", "label": "Avg explicit", "formula": "member_treatment(E, agg='avg')"},
                {"code": "AVG_WRAP", "label": "Avg wrap", "formula": "members_avg(E)"},
                {"code": "MIN_EXPL", "label": "Min explicit", "formula": "member_treatment(E, agg='min')"},
                {"code": "MIN_WRAP", "label": "Min wrap", "formula": "members_min(E)"},
                {"code": "MAX_EXPL", "label": "Max explicit", "formula": "member_treatment(E, agg='max')"},
                {"code": "MAX_WRAP", "label": "Max wrap", "formula": "members_max(E)"},
                {"code": "COUNT_EXPL", "label": "Count explicit", "formula": "member_treatment(E, agg='count')"},
                {"code": "COUNT_WRAP", "label": "Count wrap", "formula": "members_count(E)"},
            ],
        }
        runtime_schema = runtime_schema_for_comp_aparell(schema, self.comp_app, member_count=3)
        result = ScoringEngine(runtime_schema).compute(
            {
                "E__m1": 8.1,
                "E__m2": 7.4,
                "E__m3": 8.5,
            }
        )
        self.assertAlmostEqual(result.outputs["SUM_EXPL"], result.outputs["SUM_WRAP"])
        self.assertAlmostEqual(result.outputs["AVG_EXPL"], result.outputs["AVG_WRAP"])
        self.assertAlmostEqual(result.outputs["MIN_EXPL"], result.outputs["MIN_WRAP"])
        self.assertAlmostEqual(result.outputs["MAX_EXPL"], result.outputs["MAX_WRAP"])
        self.assertAlmostEqual(result.outputs["COUNT_EXPL"], result.outputs["COUNT_WRAP"])

    def test_runtime_member_treatment_supports_official_advanced_selectors_and_med(self):
        schema = {
            "fields": [
                {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
            ],
            "computed": [
                {"code": "DROP_SUM", "label": "Drop extremes", "formula": "member_treatment(E, select='drop_extremes', agg='sum')"},
                {"code": "ALT_TWO", "label": "Alternating extremes", "formula": "member_treatment(E, select='drop_extremes_until_n', n=2, agg='sum')"},
                {"code": "MEDIAN", "label": "Median", "formula": "member_treatment(E, agg='med')"},
            ],
        }
        runtime_schema = runtime_schema_for_comp_aparell(schema, self.comp_app, member_count=5)
        result = ScoringEngine(runtime_schema).compute(
            {
                "E__m1": 1.0,
                "E__m2": 3.0,
                "E__m3": 5.0,
                "E__m4": 7.0,
                "E__m5": 9.0,
            }
        )
        self.assertAlmostEqual(result.outputs["DROP_SUM"], 15.0)
        self.assertAlmostEqual(result.outputs["ALT_TWO"], 8.0)
        self.assertAlmostEqual(result.outputs["MEDIAN"], 5.0)

    def test_runtime_schema_expands_member_computed_before_member_treatment(self):
        schema = {
            "fields": [
                {
                    "code": "E",
                    "label": "Exec",
                    "type": "matrix",
                    "shape": "judge_x_item",
                    "scope": "member",
                    "judges": {"count": 1},
                    "items": {"count": 2},
                },
            ],
            "computed": [
                {
                    "code": "E_MEMBER",
                    "label": "Exec membre",
                    "formula": "row_custom_compute('E', '1 - x', row_select='all', row_agg='sum', col_select='all', col_agg='sum')",
                },
                {
                    "code": "TOTAL",
                    "label": "Total",
                    "formula": "member_treatment(E_MEMBER, agg='avg')",
                },
            ],
        }
        runtime_schema = runtime_schema_for_comp_aparell(schema, self.comp_app, member_count=2)
        runtime_codes = [c.get("code") for c in runtime_schema.get("computed", [])]
        self.assertIn("E_MEMBER__m1", runtime_codes)
        self.assertIn("E_MEMBER__m2", runtime_codes)

        engine = ScoringEngine(runtime_schema)
        result = engine.compute(
            {
                "E__m1": [[0.1, 0.2]],
                "E__m2": [[0.4, 0.1]],
            }
        )
        self.assertAlmostEqual(result.outputs["E_MEMBER__m1"], 1.7)
        self.assertAlmostEqual(result.outputs["E_MEMBER__m2"], 1.5)
        self.assertAlmostEqual(result.total, 1.6)


class TeamContextScoringFlowTests(_BaseTrampoliDataMixin, TestCase):
    def setUp(self):
        self.comp = self._create_competicio("Comp team scoring")
        User = get_user_model()
        self.user = User.objects.create_user(
            username="team_context_scoring_owner",
            password="testpass123",
            email="team-context-scoring-owner@example.com",
        )
        CompeticioMembership.objects.create(
            user=self.user,
            competicio=self.comp,
            role=CompeticioMembership.Role.OWNER,
            is_active=True,
        )
        self.client.force_login(self.user)
        self.app = self._create_aparell("SYNC", "Sincronitzat")
        self.app.competition_unit = Aparell.CompetitionUnit.TEAM
        self.app.save(update_fields=["competition_unit"])
        self.comp_app = self._create_comp_aparell(self.comp, self.app, ordre=1)
        self.ctx = EquipContext.objects.create(
            competicio=self.comp,
            code="parelles",
            nom="Parelles",
        )
        self.other_ctx = EquipContext.objects.create(
            competicio=self.comp,
            code="altre",
            nom="Altre",
        )
        self.equip = self._create_equip(self.comp, "Parella 1", context=self.ctx)
        self.ins1 = self._create_inscripcio(self.comp, "Maria", ordre=1, grup=1)
        self.ins2 = self._create_inscripcio(self.comp, "Laia", ordre=2, grup=1)
        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=self.ctx,
            inscripcio=self.ins1,
            equip=self.equip,
        )
        InscripcioEquipAssignacio.objects.create(
            competicio=self.comp,
            context=self.ctx,
            inscripcio=self.ins2,
            equip=self.equip,
        )
        CompeticioAparellEquipContextSource.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            context=self.ctx,
        )

    def _create_team_with_members(self, team_name, member_names, *, context=None, start_order=10):
        context = context or self.ctx
        equip = self._create_equip(self.comp, team_name, context=context)
        members = []
        for idx, name in enumerate(member_names, start=0):
            ins = self._create_inscripcio(self.comp, name, ordre=start_order + idx, grup=1)
            InscripcioEquipAssignacio.objects.create(
                competicio=self.comp,
                context=context,
                inscripcio=ins,
                equip=equip,
            )
            members.append(ins)
        return equip, members

    def _team_subject(self, equip=None):
        equip = equip or self.equip
        subjects, _issues = build_team_subjects_for_comp_aparell(self.comp, self.comp_app)
        for subject in subjects:
            if int(subject.get("equip_id") or 0) == int(equip.id):
                subject_obj = TeamCompetitiveSubject.objects.get(pk=int(subject["subject_id"]))
                return subject_obj, subject
        self.fail(f"No s'ha trobat team_subject per a l'equip {equip.id}")

    def _team_payload(self, equip=None, **extra):
        subject_obj, _subject = self._team_subject(equip)
        payload = {
            "subject_kind": "team_unit",
            "subject_id": subject_obj.id,
        }
        payload.update(extra)
        return payload

    def _create_individual_comp_aparell(self, codi="TRIND", nom="Tramp Individual", ordre=9):
        app = self._create_aparell(codi, nom)
        comp_aparell = self._create_comp_aparell(self.comp, app, ordre=ordre)
        return app, comp_aparell

    def _post_json(self, url_name, payload, **kwargs):
        return self.client.post(
            reverse(url_name, kwargs={"pk": self.comp.id, **kwargs}),
            data=json.dumps(payload),
            content_type="application/json",
        )

    def _classificacio_payload(self, *, tipus="equips", app_ids=None, context_code="parelles", team_mode=None):
        ids = list(app_ids or [self.comp_app.id])
        schema = {
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": ids},
                "camps_per_aparell": {str(app_id): ["total"] for app_id in ids},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
            },
        }
        if tipus == "equips":
            equips_cfg = {
                "context_code": context_code,
                "incloure_sense_equip": False,
            }
            if team_mode is not None:
                equips_cfg["team_mode"] = team_mode
            schema["equips"] = equips_cfg
        return {
            "nom": "Cfg test",
            "activa": True,
            "ordre": 1,
            "tipus": tipus,
            "schema": schema,
        }

    def test_build_metric_meta_marks_native_team_displayable_fields(self):
        schema_obj = {
            "meta": {"subject_mode": "team"},
            "fields": [
                {
                    "code": "E",
                    "label": "Execucio",
                    "scope": "member",
                    "type": "matrix",
                    "shape": "judge_x_item",
                    "judges": {"count": 3},
                    "items": {"count": 5},
                },
                {
                    "code": "S",
                    "label": "Sync",
                    "scope": "shared",
                    "type": "number",
                },
            ],
            "computed": [
                {
                    "code": "E_mem",
                    "label": "Execucio membre",
                    "formula": "row_custom_compute('E', '1 - x')",
                },
                {
                    "code": "E_by_judge",
                    "label": "Execucio by judge",
                    "formula": "row_custom_compute('E', '1 - x', return_mode='by_judge')",
                },
            ],
        }

        meta = _build_metric_meta_for_comp_aparell(self.comp_app, schema_obj, strict_unknown=True)

        self.assertFalse(meta["E"]["scoreable"])
        self.assertTrue(meta["E"]["member_dependent"])
        self.assertTrue(meta["E"]["detail_displayable"])
        self.assertEqual(meta["E"]["detail_display_kind"], "judge_rows")
        self.assertTrue(meta["E_mem"]["detail_displayable"])
        self.assertEqual(meta["E_mem"]["detail_display_kind"], "scalar")
        self.assertTrue(meta["E_mem"]["member_dependent"])
        self.assertFalse(meta["E_by_judge"]["detail_displayable"])
        self.assertEqual(meta["E_by_judge"]["detail_display_kind"], "none")

    def test_builder_context_exposes_displayable_member_fields_for_native_team(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team"},
                "fields": [
                    {
                        "code": "E",
                        "label": "Execucio",
                        "scope": "member",
                        "type": "matrix",
                        "shape": "judge_x_item",
                        "judges": {"count": 3},
                        "items": {"count": 5},
                    },
                    {
                        "code": "S",
                        "label": "Sync",
                        "scope": "shared",
                        "type": "number",
                    },
                ],
                "computed": [
                    {
                        "code": "E_mem",
                        "label": "Execucio membre",
                        "formula": "row_custom_compute('E', '1 - x')",
                    },
                ],
            },
        )

        request = RequestFactory().get("/competicio/test/classificacions/")
        request.user = self.user
        view = ClassificacionsHome()
        view.request = request
        view.kwargs = {"pk": self.comp.id}
        view.competicio = self.comp

        ctx = view.get_context_data()
        options = ctx["aparell_field_options"][str(self.comp_app.id)]
        by_code = {item["code"]: item for item in options}

        self.assertIn("E", by_code)
        self.assertFalse(by_code["E"]["scoreable"])
        self.assertTrue(by_code["E"]["member_dependent"])
        self.assertTrue(by_code["E"]["detail_displayable"])
        self.assertEqual(by_code["E"]["detail_display_kind"], "judge_rows")
        self.assertTrue(by_code["E_mem"]["detail_displayable"])
        self.assertEqual(by_code["E_mem"]["detail_display_kind"], "scalar")

    def _native_team_schema_with_tie(self, tie):
        return {
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
            },
            "desempat": [tie],
            "equips": {
                "context_code": "parelles",
                "team_mode": "native_team",
                "incloure_sense_equip": False,
            },
        }

    def _birth_range_partition_cfg(self, *, compliance_mode="strict", max_outside=0):
        return {
            "particions": ["any_naixement_forquilla"],
            "particions_v2": [
                {"code": "any_naixement_forquilla", "apply_mode": "all", "parent_values": []},
            ],
            "particions_config": {
                "any_naixement_forquilla": {
                    "ranges": [
                        {
                            "label": "U13",
                            "from_date": "2012-01-01",
                            "until_date": "2014-12-31",
                        },
                    ],
                    "sense_data_label": "Sense data",
                    "fora_rang_label": "Fora de forquilla",
                    "team_rules": {
                        "reference_mode": "oldest_member_birthdate",
                        "compliance_mode": compliance_mode,
                        "max_members_outside_range": max_outside,
                        "missing_birthdate_policy": "outside_range",
                    },
                }
            },
        }

    def test_team_builder_native_context_ignores_legacy_team_without_base_assignment(self):
        base_ctx = self._ensure_native_equip_context(self.comp)
        legacy_team = Equip.objects.create(competicio=self.comp, nom="Legacy Base")
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Legacy A",
            ordre_sortida=20,
            grup=1,
            equip=legacy_team,
        )
        Inscripcio.objects.create(
            competicio=self.comp,
            nom_i_cognoms="Legacy B",
            ordre_sortida=21,
            grup=1,
            equip=legacy_team,
        )
        CompeticioAparellEquipContextSource.objects.filter(
            competicio=self.comp,
            comp_aparell=self.comp_app,
        ).delete()
        CompeticioAparellEquipContextSource.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            context=base_ctx,
        )

        subjects, issues = build_team_subjects_for_comp_aparell(self.comp, self.comp_app)

        self.assertEqual(subjects, [])
        self.assertTrue(any(item.get("context_code") == "native" for item in issues))

    def test_scoring_schema_full_clean_accepts_member_scope_for_team_context(self):
        schema = ScoringSchema(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2"},
                ],
            },
        )
        schema.comp_aparell = self.comp_app
        schema.full_clean()

    def test_team_builder_shows_official_member_treatment_options(self):
        self.app.competition_unit = Aparell.CompetitionUnit.TEAM
        self.app.save(update_fields=["competition_unit"])

        response = self.client.get(
            reverse("scoring_schema_update", kwargs={"pk": self.comp.id, "ap_id": self.comp_app.id})
        )

        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn("member_treatment(source, select='all', n=None, agg='sum')", body)
        self.assertIn('<option value="drop_extremes">', body)
        self.assertIn('<option value="drop_extremes_until_n">', body)
        self.assertIn('<option value="count">Comptar</option>', body)
        self.assertIn('<option value="med">Mediana</option>', body)

    def test_team_subject_label_is_truncated_to_model_limit(self):
        long_team_name = "Equip " + ("MoltLlarg" * 20)
        long_members = [
            "Participant " + ("Alpha" * 12),
            "Participant " + ("Beta" * 12),
            "Participant " + ("Gamma" * 12),
        ]
        equip, _members = self._create_team_with_members(long_team_name, long_members, start_order=50)

        subject_obj, subject = self._team_subject(equip)

        self.assertLessEqual(len(subject_obj.label), 255)
        self.assertEqual(subject_obj.label, subject_obj.label.strip())
        self.assertIn("Parelles", subject_obj.label)
        self.assertIn("Equip", subject_obj.label)

    def test_scoring_schema_builder_get_exposes_saved_bootstrap_and_draft_key(self):
        schema = ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number"},
                ],
                "computed": [],
            },
        )

        response = self.client.get(
            reverse("scoring_schema_update", kwargs={"pk": self.comp.id, "ap_id": self.comp_app.id})
        )

        self.assertEqual(response.status_code, 200)
        bootstrap = response.context["schema_bootstrap"]
        self.assertEqual(bootstrap["schema_initial_source"], "saved")
        self.assertEqual(bootstrap["schema_initial"], schema.schema)
        self.assertIn(f"comp-aparell:{self.comp_app.id}", bootstrap["schema_draft_storage_key"])

    def test_scoring_schema_builder_rehydrates_last_posted_invalid_schema(self):
        existing = ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "OLD", "label": "Antic", "type": "number"},
                ],
                "computed": [],
            },
        )
        invalid_schema = {
            "fields": [
                {"code": "E", "label": "Exec", "type": "number"},
            ],
            "computed": [
                {"code": "E", "label": "Duplicat", "formula": "1"},
            ],
        }

        response = self.client.post(
            reverse("scoring_schema_update", kwargs={"pk": self.comp.id, "ap_id": self.comp_app.id}),
            data={"schema_json": json.dumps(invalid_schema)},
        )

        self.assertEqual(response.status_code, 200)
        bootstrap = response.context["schema_bootstrap"]
        self.assertEqual(bootstrap["schema_initial_source"], "posted_invalid")
        self.assertEqual(bootstrap["schema_initial"], invalid_schema)
        self.assertEqual(bootstrap["schema_raw_invalid_json"], "")
        existing.refresh_from_db()
        self.assertEqual(existing.schema["fields"][0]["code"], "OLD")

    def test_scoring_schema_builder_preserves_raw_invalid_json(self):
        existing = ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "OLD", "label": "Antic", "type": "number"},
                ],
                "computed": [],
            },
        )
        invalid_raw_json = '{"fields": ['

        response = self.client.post(
            reverse("scoring_schema_update", kwargs={"pk": self.comp.id, "ap_id": self.comp_app.id}),
            data={"schema_json": invalid_raw_json},
        )

        self.assertEqual(response.status_code, 200)
        bootstrap = response.context["schema_bootstrap"]
        self.assertEqual(bootstrap["schema_initial_source"], "raw_invalid_json")
        self.assertEqual(bootstrap["schema_initial"], existing.schema)
        self.assertEqual(bootstrap["schema_raw_invalid_json"], invalid_raw_json)
        existing.refresh_from_db()
        self.assertEqual(existing.schema["fields"][0]["code"], "OLD")

    def test_scoring_save_partial_creates_team_score_entry(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2"},
                ],
            },
        )
        team_subject, _subject_meta = self._team_subject()

        response = self.client.post(
            reverse("scoring_save_partial", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "comp_aparell_id": self.comp_app.id,
                    "subject_kind": "team_unit",
                    "subject_id": team_subject.id,
                    "exercici": 1,
                    "inputs_patch": {
                        "SYNC": 7.5,
                        "E": {
                            str(team_subject.member_ids[0]): 8.1,
                            str(team_subject.member_ids[1]): 8.2,
                        },
                    },
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["subject_kind"], "team_unit")
        self.assertEqual(payload["subject_id"], team_subject.id)
        self.assertEqual(
            payload["inputs"]["E"],
            {
                str(team_subject.member_ids[0]): 8.1,
                str(team_subject.member_ids[1]): 8.2,
            },
        )

        entry = TeamScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
        )
        self.assertEqual(float(entry.total), 23.8)
        self.assertEqual(entry.inputs["SYNC"], 7.5)
        self.assertEqual(
            entry.inputs["E"],
            {
                str(team_subject.member_ids[0]): 8.1,
                str(team_subject.member_ids[1]): 8.2,
            },
        )

    def test_scoring_save_partial_rejects_runtime_member_keys_for_team_app(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2"},
                ],
            },
        )
        team_subject, _subject_meta = self._team_subject()

        response = self.client.post(
            reverse("scoring_save_partial", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "comp_aparell_id": self.comp_app.id,
                    "subject_kind": "team_unit",
                    "subject_id": team_subject.id,
                    "exercici": 1,
                    "inputs_patch": {"SYNC": 7.5, "E__m1": 8.1},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("__mN", response.json()["error"])

    def test_schema_recalc_for_team_preserves_orphan_inputs(self):
        schema = ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2"},
                ],
            },
        )
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={
                "SYNC": 7.5,
                "E": {
                    str(team_subject.member_ids[0]): 8.1,
                    str(team_subject.member_ids[1]): 8.2,
                },
                "OLD_FIELD": 99.0,
            },
            outputs={"TOTAL": 23.8},
            total=23.8,
        )

        response = self.client.post(
            reverse("scoring_schema_update", kwargs={"pk": self.comp.id, "ap_id": self.comp_app.id}),
            data={
                "schema_json": json.dumps(
                    {
                        "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                        "fields": [
                            {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                            {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                            {"code": "BONUS", "label": "Bonus", "type": "number", "scope": "shared"},
                        ],
                        "computed": [
                            {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2 + BONUS"},
                        ],
                    }
                )
            },
        )

        self.assertEqual(response.status_code, 302)
        schema.refresh_from_db()
        entry = TeamScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
        )
        self.assertIn("OLD_FIELD", entry.inputs)
        self.assertEqual(entry.inputs["OLD_FIELD"], 99.0)
        self.assertEqual(entry.inputs["SYNC"], 7.5)
        self.assertEqual(entry.inputs["E"][str(team_subject.member_ids[0])], 8.1)
        self.assertEqual(entry.inputs["E"][str(team_subject.member_ids[1])], 8.2)

    def test_scoring_save_rejects_individual_payload_for_team_context_app(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [{"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"}],
                "computed": [],
            },
        )

        response = self.client.post(
            reverse("scoring_save", kwargs={"pk": self.comp.id}),
            data=json.dumps(
                {
                    "comp_aparell_id": self.comp_app.id,
                    "inscripcio_id": self.ins1.id,
                    "exercici": 1,
                    "inputs": {"SYNC": 7.5},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("subject_kind=team_unit", response.json()["error"])

    def test_judge_save_partial_uses_team_subject_and_runtime_member_permission(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2"},
                ],
            },
        )
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Judge",
            permissions=[
                {"field_code": "SYNC", "runtime_field_code": "SYNC", "scope": "shared", "judge_index": 1},
                {"field_code": "E", "runtime_field_code": "E__m2", "scope": "member", "member_slot": 2, "judge_index": 1},
            ],
            is_active=True,
        )
        team_subject, _subject_meta = self._team_subject()

        response = self.client.post(
            reverse("judge_save_partial", kwargs={"token": token.id}),
            data=json.dumps(
                {
                    "subject_kind": "team_unit",
                    "subject_id": team_subject.id,
                    "exercici": 1,
                    "inputs_patch": {"SYNC": 6.4, "E__m2": 7.1},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["subject_kind"], "team_unit")
        self.assertEqual(payload["subject_id"], team_subject.id)
        self.assertEqual(payload["inputs"]["SYNC"], 6.4)
        self.assertEqual(payload["inputs"]["E__m2"], 7.1)

        entry = TeamScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
        )
        self.assertEqual(entry.inputs["SYNC"], 6.4)
        self.assertEqual(entry.inputs["E"][str(self.ins2.id)], 7.1)
        self.assertEqual(entry.inputs["E"][str(self.ins1.id)], 0.0)

    def test_judge_save_partial_uses_comp_aparell_specific_schema_before_global(self):
        app, comp_aparell = self._create_individual_comp_aparell(codi="TRSAVE", nom="Tramp Save", ordre=10)
        ScoringSchema.objects.create(
            aparell=app,
            schema={
                "fields": [{"code": "SYNC", "label": "Sync", "type": "number"}],
                "computed": [{"code": "TOTAL", "label": "Total", "formula": "SYNC"}],
            },
        )
        ScoringSchema.objects.create(
            comp_aparell=comp_aparell,
            schema={
                "fields": [{"code": "ALT", "label": "Alt", "type": "number"}],
                "computed": [{"code": "TOTAL", "label": "Total", "formula": "ALT"}],
            },
        )
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=comp_aparell,
            label="Individual override",
            permissions=[{"field_code": "ALT", "runtime_field_code": "ALT", "scope": "shared", "judge_index": 1}],
            is_active=True,
        )

        response = self.client.post(
            reverse("judge_save_partial", kwargs={"token": token.id}),
            data=json.dumps(
                {
                    "subject_kind": "inscripcio",
                    "subject_id": self.ins1.id,
                    "exercici": 1,
                    "inputs_patch": {"ALT": 5.5},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["inputs"]["ALT"], 5.5)
        entry = ScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=comp_aparell,
            inscripcio=self.ins1,
            exercici=1,
        )
        self.assertEqual(entry.inputs["ALT"], 5.5)
        self.assertAlmostEqual(float(entry.total), 5.5)

    def test_permission_runtime_resolution_supports_member_modes_and_legacy_permissions(self):
        all_entries = resolve_permission_runtime_entries(
            {"field_code": "E", "scope": "member", "judge_index": 1, "member_mode": "all"},
            self.comp_app,
            member_count=2,
        )
        self.assertEqual([row["runtime_field_code"] for row in all_entries], ["E__m1", "E__m2"])

        subset_entries = resolve_permission_runtime_entries(
            {"field_code": "E", "scope": "member", "judge_index": 1, "member_mode": "subset", "member_slots": [1, 2]},
            self.comp_app,
            member_count=2,
        )
        self.assertEqual([row["runtime_field_code"] for row in subset_entries], ["E__m1", "E__m2"])
        self.assertEqual(build_permission_label({"field_code": "E", "scope": "member", "member_mode": "subset", "member_slots": [1, 2]}), "E · Individual · M1,M2")

        legacy_entries = resolve_permission_runtime_entries(
            {"field_code": "E", "scope": "member", "judge_index": 1, "runtime_field_code": "E__m2"},
            self.comp_app,
            member_count=2,
        )
        self.assertEqual([row["runtime_field_code"] for row in legacy_entries], ["E__m2"])

        missing_slot_entries = resolve_permission_runtime_entries(
            {"field_code": "E", "scope": "member", "judge_index": 1, "member_mode": "subset", "member_slots": [3]},
            self.comp_app,
            member_count=2,
        )
        self.assertEqual(missing_slot_entries, [])

    def test_member_slot_choices_use_real_max_across_all_context_subjects(self):
        trio_ctx = EquipContext.objects.create(
            competicio=self.comp,
            code="trios",
            nom="Trios",
        )
        CompeticioAparellEquipContextSource.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            context=trio_ctx,
        )
        self._create_team_with_members(
            "Trio 1",
            ["Aina", "Noa", "Judit"],
            context=trio_ctx,
            start_order=40,
        )

        slots = _member_slot_choices(self.comp, self.comp_app)

        self.assertEqual(slots, [1, 2, 3])
        self.assertNotIn(4, slots)

    def test_validate_permission_row_rejects_invalid_member_targeting(self):
        schema_by_code = {
            "E": {"code": "E", "type": "number", "scope": "member", "judges": {"count": 1}},
            "SYNC": {"code": "SYNC", "type": "number", "scope": "shared", "judges": {"count": 1}},
        }

        valid = _validate_permission_row(
            schema_by_code,
            {
                "field_code": "E",
                "scope": "member",
                "judge_index": 1,
                "member_mode": "single",
                "member_slots": "2",
            },
            team_context_mode=True,
        )
        self.assertEqual(valid["member_mode"], "single")
        self.assertEqual(valid["member_slots"], [2])

        with self.assertRaisesMessage(ValueError, "exactament un membre"):
            _validate_permission_row(
                schema_by_code,
                {
                    "field_code": "E",
                    "scope": "member",
                    "judge_index": 1,
                    "member_mode": "single",
                    "member_slots": "1,2",
                },
                team_context_mode=True,
            )

        with self.assertRaisesMessage(ValueError, "almenys un membre"):
            _validate_permission_row(
                schema_by_code,
                {
                    "field_code": "E",
                    "scope": "member",
                    "judge_index": 1,
                    "member_mode": "subset",
                    "member_slots": "",
                },
                team_context_mode=True,
            )

        legacy_individual = _validate_permission_row(
            schema_by_code,
            {
                "field_code": "E",
                "scope": "member",
                "judge_index": 1,
                "member_mode": "all",
            },
            team_context_mode=False,
        )
        self.assertEqual(legacy_individual["scope"], "shared")
        self.assertNotIn("member_mode", legacy_individual)

        with self.assertRaisesMessage(ValueError, "abast compartit"):
            _validate_permission_row(
                schema_by_code,
                {
                    "field_code": "SYNC",
                    "scope": "member",
                    "judge_index": 1,
                    "member_mode": "all",
                },
                team_context_mode=True,
            )

        with self.assertRaisesMessage(ValueError, "abast individual"):
            _validate_permission_row(
                schema_by_code,
                {
                    "field_code": "E",
                    "scope": "shared",
                    "judge_index": 1,
                },
                team_context_mode=True,
            )

    def test_resolve_scoring_schema_for_comp_aparell_prefers_specific_before_global(self):
        global_schema = ScoringSchema.objects.create(
            aparell=self.app,
            schema={"fields": [{"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"}], "computed": []},
        )
        specific_schema = ScoringSchema.objects.create(
            comp_aparell=self.comp_app,
            schema={"fields": [{"code": "ALT", "label": "Alt", "type": "number", "scope": "shared"}], "computed": []},
        )

        schema_obj, schema = resolve_scoring_schema_for_comp_aparell(self.comp_app)

        self.assertEqual(schema_obj.pk, specific_schema.pk)
        self.assertEqual(schema["fields"][0]["code"], "ALT")
        global_schema.refresh_from_db()
        self.assertEqual(global_schema.schema["fields"][0]["code"], "SYNC")

    def test_resolve_scoring_schema_for_comp_aparell_falls_back_to_global(self):
        global_schema = ScoringSchema.objects.create(
            aparell=self.app,
            schema={"fields": [{"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"}], "computed": []},
        )

        schema_obj, schema = resolve_scoring_schema_for_comp_aparell(self.comp_app)

        self.assertEqual(schema_obj.pk, global_schema.pk)
        self.assertEqual(schema["fields"][0]["code"], "SYNC")

    def test_judge_admin_create_token_stores_member_targeting(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        response = self.client.post(
            f"{reverse('judges_qr_home', kwargs={'competicio_id': self.comp.id})}?comp_aparell={self.comp_app.id}",
            data={
                "action": "create",
                "label": "Judge subset",
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "15",
                "form-0-field_code": "E",
                "form-0-scope": "member",
                "form-0-member_mode": "subset",
                "form-0-member_slots": "1,2",
                "form-0-judge_index": "1",
                "form-0-item_start": "1",
                "form-0-item_count": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        token = JudgeDeviceToken.objects.get(label="Judge subset")
        self.assertEqual(token.permissions[0]["member_mode"], "subset")
        self.assertEqual(token.permissions[0]["member_slots"], [1, 2])

    def test_judge_admin_uses_comp_aparell_specific_schema_before_global(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                ],
                "computed": [],
            },
        )
        ScoringSchema.objects.create(
            comp_aparell=self.comp_app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "ALT", "label": "Alternatiu", "type": "number", "scope": "shared"},
                ],
                "computed": [],
            },
        )

        response = self.client.get(
            f"{reverse('judges_qr_home', kwargs={'competicio_id': self.comp.id})}?comp_aparell={self.comp_app.id}"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["schema"]["fields"][0]["code"], "ALT")
        self.assertEqual(response.context["schema_field_catalog"][0]["code"], "ALT")

    def test_judge_admin_individual_app_hides_scope_and_tolerates_legacy_member_schema(self):
        app, comp_aparell = self._create_individual_comp_aparell()
        ScoringSchema.objects.create(
            comp_aparell=comp_aparell,
            schema={
                "fields": [
                    {"code": "ALT", "label": "Alternatiu", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "ALT"},
                ],
            },
        )

        response = self.client.get(
            f"{reverse('judges_qr_home', kwargs={'competicio_id': self.comp.id})}?comp_aparell={comp_aparell.id}"
        )

        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertNotIn("<th>Abast</th>", body)
        self.assertIn('type="hidden" name="form-0-scope"', body)

        post_response = self.client.post(
            f"{reverse('judges_qr_home', kwargs={'competicio_id': self.comp.id})}?comp_aparell={comp_aparell.id}",
            data={
                "action": "create",
                "label": "Judge individual legacy",
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "15",
                "form-0-field_code": "ALT",
                "form-0-scope": "shared",
                "form-0-member_mode": "all",
                "form-0-member_slots": "",
                "form-0-judge_index": "1",
                "form-0-item_start": "1",
                "form-0-item_count": "",
            },
        )

        self.assertEqual(post_response.status_code, 302)
        token = JudgeDeviceToken.objects.get(label="Judge individual legacy")
        self.assertEqual(token.comp_aparell_id, comp_aparell.id)
        self.assertEqual(token.permissions[0]["field_code"], "ALT")
        self.assertEqual(token.permissions[0]["scope"], "shared")
        self.assertEqual(token.permissions[0]["runtime_field_code"], "ALT")
        self.assertNotIn("member_mode", token.permissions[0])

    def test_judge_admin_team_app_keeps_scope_column_visible(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                ],
                "computed": [],
            },
        )

        response = self.client.get(
            f"{reverse('judges_qr_home', kwargs={'competicio_id': self.comp.id})}?comp_aparell={self.comp_app.id}"
        )

        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn("<th>Abast</th>", body)
        self.assertIn('id="judge-schema-field-catalog"', body)
        self.assertIn('"code": "SYNC"', body)
        self.assertIn('"scope": "shared"', body)

    def test_judge_admin_team_shared_field_ignores_member_targeting(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        response = self.client.post(
            f"{reverse('judges_qr_home', kwargs={'competicio_id': self.comp.id})}?comp_aparell={self.comp_app.id}",
            data={
                "action": "create",
                "label": "Judge shared team",
                "form-TOTAL_FORMS": "1",
                "form-INITIAL_FORMS": "0",
                "form-MIN_NUM_FORMS": "0",
                "form-MAX_NUM_FORMS": "15",
                "form-0-field_code": "SYNC",
                "form-0-scope": "shared",
                "form-0-member_mode": "subset",
                "form-0-member_slots": "1,2",
                "form-0-judge_index": "1",
                "form-0-item_start": "1",
                "form-0-item_count": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        token = JudgeDeviceToken.objects.get(label="Judge shared team")
        self.assertEqual(token.permissions[0]["scope"], "shared")
        self.assertNotIn("member_mode", token.permissions[0])
        self.assertNotIn("member_slots", token.permissions[0])

    def test_judge_portal_uses_team_dom_keys_and_member_target_metadata(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team portal subset",
            permissions=[
                {"field_code": "E", "scope": "member", "judge_index": 1, "member_mode": "subset", "member_slots": [1, 2]},
            ],
            is_active=True,
        )
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={
                "E": {
                    str(self.ins1.id): 8.1,
                    str(self.ins2.id): 8.2,
                }
            },
            outputs={},
            total=0,
        )

        response = self.client.get(reverse("judge_portal", kwargs={"token": token.id}))

        self.assertEqual(response.status_code, 200)
        dom_key = f"team_unit:{team_subject.id}"
        self.assertIn(dom_key, response.context["scores_payload_json"])
        exercise_payload = response.context["scores_payload_json"][dom_key]["exercises"]["1"]["inputs"]
        self.assertEqual(exercise_payload["E__m1"], 8.1)
        self.assertEqual(exercise_payload["E__m2"], 8.2)
        self.assertEqual(response.context["permissions"][0]["member_mode"], "subset")
        self.assertEqual(response.context["permissions"][0]["member_slots"], [1, 2])

    def test_judge_portal_renders_missing_member_slots_as_disabled(self):
        trio_ctx = EquipContext.objects.create(
            competicio=self.comp,
            code="trios_portal",
            nom="Trios Portal",
        )
        CompeticioAparellEquipContextSource.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            context=trio_ctx,
        )
        self._create_team_with_members(
            "Trio 2",
            ["Berta", "Clara", "Nina"],
            context=trio_ctx,
            start_order=50,
        )
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 3},
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team portal missing slot",
            permissions=[
                {"field_code": "E", "scope": "member", "judge_index": 1, "member_mode": "subset", "member_slots": [1, 3]},
            ],
            is_active=True,
        )

        response = self.client.get(reverse("judge_portal", kwargs={"token": token.id}))

        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn("Membre no disponible per aquest equip. El camp queda desactivat.", body)
        self.assertIn("Membre inexistent", body)

    def test_judge_save_partial_accepts_member_mode_all_permissions(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Judge All Members",
            permissions=[{"field_code": "E", "scope": "member", "judge_index": 1, "member_mode": "all"}],
            is_active=True,
        )
        team_subject, _subject_meta = self._team_subject()

        response = self.client.post(
            reverse("judge_save_partial", kwargs={"token": token.id}),
            data=json.dumps(
                {
                    "subject_kind": "team_unit",
                    "subject_id": team_subject.id,
                    "exercici": 1,
                    "inputs_patch": {"E__m1": 8.4, "E__m2": 8.1},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["inputs"]["E__m1"], 8.4)
        self.assertEqual(payload["inputs"]["E__m2"], 8.1)

        entry = TeamScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
        )
        self.assertEqual(entry.inputs["E"][str(self.ins1.id)], 8.4)
        self.assertEqual(entry.inputs["E"][str(self.ins2.id)], 8.1)

    def test_judge_save_partial_rejects_individual_payload_for_team_context_app(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [{"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"}],
                "computed": [],
            },
        )
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Judge Reject",
            permissions=[{"field_code": "SYNC", "runtime_field_code": "SYNC", "scope": "shared", "judge_index": 1}],
            is_active=True,
        )

        response = self.client.post(
            reverse("judge_save_partial", kwargs={"token": token.id}),
            data=json.dumps(
                {
                    "inscripcio_id": self.ins1.id,
                    "exercici": 1,
                    "inputs_patch": {"SYNC": 6.4},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("subject_kind=team_unit", response.json()["error"])

    def test_scoring_media_context_accepts_team_subject(self):
        team_subject, _subject_meta = self._team_subject()
        response = self.client.get(
            reverse("scoring_media_context", kwargs={"pk": self.comp.id}),
            {
                "comp_aparell_id": self.comp_app.id,
                "subject_kind": "team_unit",
                "subject_id": team_subject.id,
                "exercici": 1,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["subject"]["kind"], "team_unit")
        self.assertEqual(payload["subject"]["id"], team_subject.id)

    def test_judge_video_endpoints_support_team_subjects(self):
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Video",
            permissions=[{"field_code": "SYNC", "scope": "shared", "judge_index": 1}],
            can_record_video=True,
            is_active=True,
        )
        team_subject, _subject_meta = self._team_subject()
        probe_data = {
            "duration_seconds": 9,
            "mime_type": "video/mp4",
            "format_name": "mp4",
            "video_codec": "h264",
        }

        with patch("competicions_trampoli.views.judge.video._probe_uploaded_video_metadata", return_value=probe_data):
            upload_res = self.client.post(
                reverse("judge_video_upload", kwargs={"token": token.id}),
                data={
                    "subject_kind": "team_unit",
                    "subject_id": team_subject.id,
                    "exercici": 1,
                    "video_file": SimpleUploadedFile("team.mp4", b"\x00" * 1024, content_type="video/mp4"),
                },
            )

        self.assertEqual(upload_res.status_code, 200)
        upload_payload = upload_res.json()
        self.assertTrue(upload_payload["ok"])
        self.assertEqual(upload_payload["subject_kind"], "team_unit")
        self.assertEqual(upload_payload["subject_id"], team_subject.id)

        entry = TeamScoreEntry.objects.get(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
        )
        self.assertTrue(TeamScoreEntryVideo.objects.filter(team_score_entry=entry).exists())
        self.assertTrue(
            TeamScoreEntryVideoEvent.objects.filter(
                team_score_entry=entry,
                action=TeamScoreEntryVideoEvent.Action.UPLOAD,
                ok=True,
            ).exists()
        )

        status_res = self.client.get(
            reverse("judge_video_status", kwargs={"token": token.id}),
            {"subject_kind": "team_unit", "subject_id": team_subject.id, "exercici": 1},
        )
        self.assertEqual(status_res.status_code, 200)
        self.assertTrue(status_res.json()["has_video"])

        delete_res = self.client.post(
            reverse("judge_video_delete", kwargs={"token": token.id}),
            {"subject_kind": "team_unit", "subject_id": team_subject.id, "exercici": 1},
        )
        self.assertEqual(delete_res.status_code, 200)
        self.assertTrue(delete_res.json()["deleted"])
        self.assertFalse(TeamScoreEntryVideo.objects.filter(team_score_entry=entry).exists())

    def test_compute_classificacio_uses_team_score_entry_for_team_context_app(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [{"code": "TOTAL", "label": "Total", "type": "number", "scope": "shared"}],
                "computed": [],
            },
        )
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"TOTAL": 30},
            outputs={},
            total=30,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={},
            outputs={},
            total=2,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins2,
            exercici=1,
            inputs={},
            outputs={},
            total=3,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team direct",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "assignment_source": {"mode": "context", "context_code": "parelles", "fallback": "native"},
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual(rows[0]["participant"], "Parella 1")
        self.assertEqual(rows[0]["score"], 30.0)

    def test_compute_classificacio_supports_new_team_mode_contract(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [{"code": "TOTAL", "label": "Total", "type": "number", "scope": "shared"}],
                "computed": [],
            },
        )
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"TOTAL": 31},
            outputs={},
            total=31,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team direct v1",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual(rows[0]["participant"], "Parella 1")
        self.assertEqual(rows[0]["score"], 31.0)

    def test_compute_classificacio_derived_team_raw_column_returns_team_detail_payload(self):
        ind_app = self._create_aparell("TR_RAW", "Tramp raw")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={},
            outputs={},
            total=12.5,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins2,
            exercici=1,
            inputs={},
            outputs={},
            total=11.25,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived raw detail",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "raw_total",
                            "label": "Raw total",
                            "align": "right",
                            "decimals": 2,
                            "source": {"aparell_id": comp_ind_app.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                        },
                    ]
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        payload = rows[0]["cells"]["raw_total"]
        self.assertEqual(payload["_kind"], "team_raw_detail")
        self.assertEqual(payload["summary"], 23.75)
        self.assertEqual([item["label"] for item in payload["rows"]], ["Maria", "Laia"])
        self.assertEqual([item["value"] for item in payload["rows"]], [12.5, 11.25])

    def test_compute_classificacio_derived_team_raw_multijudge_keeps_member_judge_rows(self):
        ind_app = self._create_aparell("TR_J", "Tramp judges")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={"E": [8.1, 8.2]},
            outputs={},
            total=12.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins2,
            exercici=1,
            inputs={"E": [7.4, 7.6]},
            outputs={},
            total=11.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived raw judges",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "raw_exec",
                            "label": "Exec",
                            "align": "right",
                            "decimals": 2,
                            "source": {"aparell_id": comp_ind_app.id, "exercici": 1, "camp": "E", "jutges": {"ids": []}},
                        },
                    ]
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        payload = rows[0]["cells"]["raw_exec"]
        self.assertEqual(payload["_kind"], "team_raw_detail")
        self.assertEqual(payload["summary"], "")
        self.assertEqual(payload["rows"][0]["judge_rows"]["_kind"], "judge_rows")
        self.assertEqual(payload["rows"][1]["judge_rows"]["_kind"], "judge_rows")

    def test_compute_classificacio_derived_team_raw_column_uses_per_member_selected_exercises(self):
        ind_app = self._create_aparell("TR_PM", "Tramp per member")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)
        comp_ind_app.nombre_exercicis = 2
        comp_ind_app.save(update_fields=["nombre_exercicis"])

        for inscripcio, exercici, total in (
            (self.ins1, 1, 1.0),
            (self.ins1, 2, 10.0),
            (self.ins2, 1, 2.0),
            (self.ins2, 2, 9.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=comp_ind_app,
                inscripcio=inscripcio,
                exercici=exercici,
                inputs={},
                outputs={},
                total=total,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived raw per member",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 1},
                    "exercise_selection_scope": "per_member",
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "raw_total",
                            "label": "Raw total",
                            "align": "right",
                            "decimals": 2,
                            "source": {"aparell_id": comp_ind_app.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                        },
                    ]
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        payload = rows[0]["cells"]["raw_total"]
        self.assertEqual(payload["_kind"], "team_raw_detail")
        self.assertEqual(payload["summary"], 19.0)
        self.assertEqual([item["value"] for item in payload["rows"]], [10.0, 9.0])

    def test_compute_classificacio_derived_team_raw_column_uses_team_pool_selected_exercises(self):
        ind_app = self._create_aparell("TR_TP", "Tramp team pool")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)
        comp_ind_app.nombre_exercicis = 2
        comp_ind_app.save(update_fields=["nombre_exercicis"])

        for inscripcio, exercici, total in (
            (self.ins1, 1, 1.0),
            (self.ins1, 2, 10.0),
            (self.ins2, 1, 2.0),
            (self.ins2, 2, 9.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=comp_ind_app,
                inscripcio=inscripcio,
                exercici=exercici,
                inputs={},
                outputs={},
                total=total,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived raw team pool",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 2, "max_per_participant": 1},
                    "exercise_selection_scope": "team_pool",
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "raw_total",
                            "label": "Raw total",
                            "align": "right",
                            "decimals": 2,
                            "source": {"aparell_id": comp_ind_app.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                        },
                    ]
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        payload = rows[0]["cells"]["raw_total"]
        self.assertEqual(payload["_kind"], "team_raw_detail")
        self.assertEqual(payload["summary"], 19.0)
        self.assertEqual([item["value"] for item in payload["rows"]], [10.0, 9.0])

    def test_compute_classificacio_derived_team_raw_column_respects_global_pool_per_app(self):
        app_a = self._create_aparell("TRA_RAW", "Tramp A raw")
        app_b = self._create_aparell("TRB_RAW", "Tramp B raw")
        comp_app_a = self._create_comp_aparell(self.comp, app_a, ordre=2)
        comp_app_b = self._create_comp_aparell(self.comp, app_b, ordre=3)

        for comp_aparell, inscripcio, total in (
            (comp_app_a, self.ins1, 1.0),
            (comp_app_b, self.ins1, 10.0),
            (comp_app_a, self.ins2, 2.0),
            (comp_app_b, self.ins2, 9.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=comp_aparell,
                inscripcio=inscripcio,
                exercici=1,
                inputs={},
                outputs={},
                total=total,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived raw global pool",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_app_a.id, comp_app_b.id]},
                    "camps_per_aparell": {
                        str(comp_app_a.id): ["total"],
                        str(comp_app_b.id): ["total"],
                    },
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 2, "max_per_participant": 1},
                    "exercise_selection_scope": "team_pool",
                    "mode_seleccio_exercicis": "global_pool",
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "raw_a",
                            "label": "Raw A",
                            "align": "right",
                            "decimals": 2,
                            "source": {"aparell_id": comp_app_a.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                        },
                        {
                            "type": "raw",
                            "key": "raw_b",
                            "label": "Raw B",
                            "align": "right",
                            "decimals": 2,
                            "source": {"aparell_id": comp_app_b.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                        },
                    ]
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertEqual(rows[0]["cells"]["raw_a"], "")
        payload = rows[0]["cells"]["raw_b"]
        self.assertEqual(payload["_kind"], "team_raw_detail")
        self.assertEqual(payload["summary"], 19.0)
        self.assertEqual([item["value"] for item in payload["rows"]], [10.0, 9.0])

    def test_compute_classificacio_native_team_raw_column_returns_team_detail_payload(self):
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 7.5},
            outputs={},
            total=31,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native raw detail",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "raw_sync",
                            "label": "Sync",
                            "align": "right",
                            "decimals": 2,
                            "source": {"aparell_id": self.comp_app.id, "exercici": 1, "camp": "SYNC", "jutges": {"ids": []}},
                        },
                    ]
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        payload = rows[0]["cells"]["raw_sync"]
        self.assertEqual(payload["_kind"], "team_raw_detail")
        self.assertEqual(payload["summary"], 7.5)
        self.assertEqual(payload["rows"], [{"label": "Parella 1", "value": 7.5}])

    def test_compute_classificacio_native_team_raw_column_uses_selected_team_exercises(self):
        self.comp_app.nombre_exercicis = 2
        self.comp_app.save(update_fields=["nombre_exercicis"])
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 1.0},
            outputs={},
            total=10.0,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=2,
            inputs={"SYNC": 7.5},
            outputs={},
            total=30.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native raw selected exercises",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 1},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "raw_sync",
                            "label": "Sync",
                            "align": "right",
                            "decimals": 2,
                            "source": {"aparell_id": self.comp_app.id, "exercici": 1, "camp": "SYNC", "jutges": {"ids": []}},
                        },
                    ]
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        payload = rows[0]["cells"]["raw_sync"]
        self.assertEqual(payload["_kind"], "team_raw_detail")
        self.assertEqual(payload["summary"], 7.5)
        self.assertEqual(payload["rows"], [{"label": "Parella 1", "value": 7.5}])

    def test_compute_classificacio_derived_team_detail_payload_exposes_member_rows(self):
        ind_app = self._create_aparell("TR_DETAIL", "Tramp detail")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={},
            outputs={},
            total=12.5,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins2,
            exercici=1,
            inputs={},
            outputs={},
            total=11.25,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived member detail",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "default_open": True,
                        "columnes": [
                            {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                            {
                                "type": "raw",
                                "key": "detail_total",
                                "label": "Total",
                                "align": "right",
                                "decimals": 2,
                                "source": {
                                    "aparell_id": comp_ind_app.id,
                                    "exercici": 1,
                                    "camp": "total",
                                    "jutges": {"ids": []},
                                },
                            },
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertTrue(rows[0]["row_id"].startswith("team:"))
        detail = rows[0]["detail"]
        self.assertTrue(detail["default_open"])
        self.assertEqual([section["type"] for section in detail["sections"]], ["members_table"])
        members_table = detail["sections"][0]
        self.assertEqual([col["key"] for col in members_table["columns"]], ["participant", "detail_total"])
        self.assertEqual([item["participant"] for item in members_table["rows"]], ["Maria", "Laia"])
        self.assertEqual([item["cells"]["detail_total"] for item in members_table["rows"]], [12.5, 11.25])

    def test_compute_classificacio_derived_team_detail_defaults_to_participant_column(self):
        ind_app = self._create_aparell("TR_DETAIL_DEF", "Tramp detail default")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={},
            outputs={},
            total=12.5,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins2,
            exercici=1,
            inputs={},
            outputs={},
            total=11.25,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived member detail default",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "columnes": [],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        detail = rows[0]["detail"]
        self.assertEqual([section["type"] for section in detail["sections"]], ["members_table"])
        members_table = detail["sections"][0]
        self.assertEqual([col["key"] for col in members_table["columns"]], ["participant"])
        self.assertEqual([item["cells"]["participant"] for item in members_table["rows"]], ["Maria", "Laia"])

    def test_compute_classificacio_detail_enabled_without_sections_does_not_invent_defaults(self):
        ind_app = self._create_aparell("TR_DETAIL_NONE", "Tramp detail none")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={},
            outputs={},
            total=12.5,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins2,
            exercici=1,
            inputs={},
            outputs={},
            total=11.25,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived detail no defaults",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertNotIn("detail", rows[0])

    def test_compute_classificacio_native_team_legacy_member_table_detail_is_ignored(self):
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 7.5},
            outputs={},
            total=31,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native detail disabled by mode",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "default_open": True,
                        "columnes": [
                            {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertTrue(rows[0]["row_id"].startswith("team:"))
        self.assertNotIn("detail", rows[0])

    def test_compute_classificacio_native_team_detail_sections_include_members_and_metrics(self):
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 7.5},
            outputs={},
            total=31,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native detail sections",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "default_open": True,
                        "sections": [
                            {"type": "members_list", "label": "Participants"},
                            {
                                "type": "team_metrics",
                                "label": "Notes equip",
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "team_total",
                                        "label": "Total",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercici": 1,
                                            "camp": "total",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        detail = rows[0]["detail"]
        self.assertEqual([section["type"] for section in detail["sections"]], ["members_list", "team_metrics"])
        self.assertEqual([item["participant"] for item in detail["sections"][0]["items"]], ["Maria", "Laia"])
        metrics_section = detail["sections"][1]
        self.assertEqual(metrics_section["aparell_id"], self.comp_app.id)
        self.assertEqual([col["key"] for col in metrics_section["columns"]], ["team_total"])
        metric_cell = metrics_section["rows"][0]["cells"]["team_total"]
        self.assertEqual(metric_cell["_kind"], "team_raw_detail")
        self.assertEqual(metric_cell["summary"], 31.0)

    def test_compute_classificacio_native_team_team_members_table_uses_fixed_exercise_when_configured(self):
        self.comp_app.nombre_exercicis = 2
        self.comp_app.save(update_fields=["nombre_exercicis"])
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 6.5, "E": {str(self.ins1.id): 7.1, str(self.ins2.id): 7.0}},
            outputs={},
            total=20,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=2,
            inputs={"SYNC": 7.4, "E": {str(self.ins1.id): 8.3, str(self.ins2.id): 8.1}},
            outputs={},
            total=30,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native member detail",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 1},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercise_mode": "fixed",
                                            "exercici": 1,
                                            "camp": "E",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        detail = rows[0]["detail"]
        self.assertEqual([section["type"] for section in detail["sections"]], ["team_members_table"])
        members_table = detail["sections"][0]
        self.assertEqual([item["participant"] for item in members_table["rows"]], ["Maria", "Laia"])
        self.assertEqual([item["cells"]["member_exec"] for item in members_table["rows"]], [7.1, 7.0])

    def test_compute_classificacio_native_team_team_members_table_uses_selected_exercises_for_legacy_schema_without_exercise_mode(self):
        self.comp_app.nombre_exercicis = 2
        self.comp_app.save(update_fields=["nombre_exercicis"])
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 6.5, "E": {str(self.ins1.id): 7.1, str(self.ins2.id): 7.0}},
            outputs={},
            total=20,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=2,
            inputs={"SYNC": 7.4, "E": {str(self.ins1.id): 8.3, str(self.ins2.id): 8.1}},
            outputs={},
            total=30,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native member detail selected",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 1},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercici": 1,
                                            "camp": "E",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        members_table = rows[0]["detail"]["sections"][0]
        self.assertEqual([item["cells"]["member_exec"] for item in members_table["rows"]], [8.3, 8.1])

    def test_compute_classificacio_native_team_team_members_table_keeps_member_judge_rows(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {
                        "code": "E",
                        "label": "Exec",
                        "type": "matrix",
                        "shape": "judge_x_item",
                        "scope": "member",
                        "judges": {"count": 2},
                        "items": {"count": 2},
                    },
                ],
                "computed": [],
            },
        )
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={
                "E": {
                    str(self.ins1.id): [[8.1, 8.2], [8.0, 8.3]],
                    str(self.ins2.id): [[7.4, 7.6], [7.5, 7.7]],
                }
            },
            outputs={},
            total=30,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native member judge detail",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercici": 1,
                                            "camp": "E",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        members_table = rows[0]["detail"]["sections"][0]
        self.assertEqual(members_table["rows"][0]["cells"]["member_exec"]["_kind"], "judge_rows")
        self.assertEqual(members_table["rows"][1]["cells"]["member_exec"]["_kind"], "judge_rows")
        self.assertEqual(
            members_table["rows"][0]["cells"]["member_exec"]["rows"],
            [
                {"judge": 1, "items": [8.1, 8.2]},
                {"judge": 2, "items": [8.0, 8.3]},
            ],
        )

    def test_compute_classificacio_native_team_team_members_table_resolves_member_computed_outputs_by_subject_slot(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "E_mem", "label": "Exec neta", "formula": "E__m1 if 1 else E__m2"},
                ],
            },
        )
        team_subject, _subject_meta = self._team_subject()
        team_subject.member_ids = [self.ins2.id, self.ins1.id]
        team_subject.save(update_fields=["member_ids"])
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"E": {str(self.ins1.id): 8.2, str(self.ins2.id): 8.1}},
            outputs={"E_mem__m1": 5.4, "E_mem__m2": 9.7},
            total=30,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native member computed detail",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec_net",
                                        "label": "Exec neta",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercise_mode": "fixed",
                                            "exercici": 1,
                                            "camp": "E_mem",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        members_table = rows[0]["detail"]["sections"][0]
        self.assertEqual([item["participant"] for item in members_table["rows"]], ["Maria", "Laia"])
        self.assertEqual([item["cells"]["member_exec_net"] for item in members_table["rows"]], [9.7, 5.4])

    def test_compute_classificacio_native_team_team_members_table_falls_back_to_row_order_for_inconsistent_subject_slots(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )
        team_subject, _subject_meta = self._team_subject()
        team_subject.member_ids = [self.ins1.id, self.ins2.id, self.ins2.id]
        team_subject.save(update_fields=["member_ids"])
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={},
            outputs={"E__m1": 6.2, "E__m2": 6.4},
            total=30,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native member fallback order detail",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercise_mode": "fixed",
                                            "exercici": 1,
                                            "camp": "E",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        members_table = rows[0]["detail"]["sections"][0]
        self.assertEqual([item["cells"]["member_exec"] for item in members_table["rows"]], [6.2, 6.4])

    def test_compute_classificacio_native_team_team_members_table_keeps_blank_when_member_value_missing(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"E": {str(self.ins1.id): 6.2}},
            outputs={},
            total=30,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native member missing detail",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercise_mode": "fixed",
                                            "exercici": 1,
                                            "camp": "E",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        members_table = rows[0]["detail"]["sections"][0]
        self.assertEqual([item["cells"]["member_exec"] for item in members_table["rows"]], [6.2, ""])

    def test_compute_classificacio_individual_detail_sections_include_exercise_table(self):
        ind_app = self._create_aparell("TR_DETAIL_EX", "Tramp detail exercises")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)
        comp_ind_app.nombre_exercicis = 2
        comp_ind_app.save(update_fields=["nombre_exercicis"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={},
            outputs={},
            total=10.5,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins1,
            exercici=2,
            inputs={},
            outputs={},
            total=11.25,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Individual detail sections",
            activa=True,
            ordre=1,
            tipus="individual",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "exercise_table",
                                "label": "Exercicis",
                                "columns": [
                                    {"type": "builtin", "key": "aparell_nom", "label": "Aparell", "align": "left"},
                                    {"type": "builtin", "key": "exercise_index", "label": "Ex.", "align": "left"},
                                    {
                                        "type": "raw",
                                        "key": "total_ex1",
                                        "label": "Total 1",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": comp_ind_app.id,
                                            "exercici": 1,
                                            "camp": "total",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                    {
                                        "type": "raw",
                                        "key": "total_ex2",
                                        "label": "Total 2",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": comp_ind_app.id,
                                            "exercici": 2,
                                            "camp": "total",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        detail = rows[0]["detail"]
        self.assertEqual([section["type"] for section in detail["sections"]], ["exercise_table"])
        self.assertEqual(detail["sections"][0]["aparell_id"], comp_ind_app.id)
        exercise_rows = detail["sections"][0]["rows"]
        self.assertEqual([item["exercise_index"] for item in exercise_rows], [1, 2])
        self.assertEqual(exercise_rows[0]["cells"]["total_ex1"], 10.5)
        self.assertEqual(exercise_rows[0]["cells"]["total_ex2"], "")
        self.assertEqual(exercise_rows[1]["cells"]["total_ex1"], "")
        self.assertEqual(exercise_rows[1]["cells"]["total_ex2"], 11.25)

    def test_compute_classificacio_entitat_detail_sections_include_member_table(self):
        ind_app = self._create_aparell("TR_DETAIL_ENT", "Tramp detail entitat")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={},
            outputs={},
            total=12.5,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins2,
            exercici=1,
            inputs={},
            outputs={},
            total=11.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Entity detail sections",
            activa=True,
            ordre=1,
            tipus="entitat",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Entitat", "align": "left"},
                    ],
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "entity_members_table",
                                "label": "Participants",
                                "columns": [
                                    {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                                    {
                                        "type": "raw",
                                        "key": "detail_total",
                                        "label": "Total",
                                        "align": "right",
                                        "decimals": 2,
                                        "source": {
                                            "aparell_id": comp_ind_app.id,
                                            "exercici": 1,
                                            "camp": "total",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertTrue(rows[0]["row_id"].startswith("entity:"))
        detail = rows[0]["detail"]
        self.assertEqual([section["type"] for section in detail["sections"]], ["entity_members_table"])
        members_table = detail["sections"][0]
        self.assertEqual([item["participant"] for item in members_table["rows"]], ["Maria", "Laia"])
        self.assertEqual([item["cells"]["detail_total"] for item in members_table["rows"]], [12.5, 11.0])

    def test_compute_classificacio_native_team_raw_column_on_individual_app_returns_blank_for_stale_schema(self):
        ind_app = self._create_aparell("TR_STALE", "Tramp stale")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)
        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 6.0},
            outputs={},
            total=31.0,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native raw stale individual app",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id, comp_ind_app.id]},
                    "camps_per_aparell": {
                        str(self.comp_app.id): ["total"],
                        str(comp_ind_app.id): ["total"],
                    },
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "raw_invalid",
                            "label": "Raw invalid",
                            "align": "right",
                            "decimals": 2,
                            "source": {"aparell_id": comp_ind_app.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                        },
                    ]
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertEqual(rows[0]["cells"]["raw_invalid"], "")

    def test_compute_classificacio_native_team_filters_require_all_members_to_match(self):
        self.ins1.categoria = "Base"
        self.ins2.categoria = "Promo"
        self.ins1.save(update_fields=["categoria"])
        self.ins2.save(update_fields=["categoria"])

        team_subject, _subject_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={},
            outputs={},
            total=31,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team filtered all members",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "filtres": {"categories_in": ["Base"]},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual(rows, [])

    def test_compute_classificacio_native_team_group_filters_prefer_normalized_group_and_fallback_to_legacy(self):
        equip_2, members_2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)
        normalized_group = GrupCompeticio.objects.create(
            competicio=self.comp,
            display_num=2,
            nom="Grup 2",
        )
        self.ins1.grup_competicio = normalized_group
        self.ins2.grup_competicio = normalized_group
        self.ins1.save(update_fields=["grup_competicio"])
        self.ins2.save(update_fields=["grup_competicio"])
        for member in members_2:
            member.grup = 3
            member.save(update_fields=["grup"])

        team_subject_1, _subject_meta_1 = self._team_subject()
        team_subject_2, _subject_meta_2 = self._team_subject(equip_2)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject_1,
            exercici=1,
            inputs={},
            outputs={},
            total=31,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject_2,
            exercici=1,
            inputs={},
            outputs={},
            total=28,
        )

        cfg_normalized = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team normalized group filter",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "filtres": {"grups_in": ["2"]},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )
        cfg_legacy = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team legacy group filter",
            activa=True,
            ordre=2,
            tipus="equips",
            schema={
                "filtres": {"grups_in": [3]},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        normalized_rows = compute_classificacio(self.comp, cfg_normalized).get("global", [])
        legacy_rows = compute_classificacio(self.comp, cfg_legacy).get("global", [])

        self.assertEqual([row["participant"] for row in normalized_rows], ["Parella 1"])
        self.assertEqual([row["participant"] for row in legacy_rows], ["Parella 2"])

    def test_compute_classificacio_derived_team_uses_only_filtered_members(self):
        ind_app = self._create_aparell("TR", "Tramp")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        self.ins1.categoria = "Base"
        self.ins2.categoria = "Promo"
        self.ins1.save(update_fields=["categoria"])
        self.ins2.save(update_fields=["categoria"])

        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={},
            outputs={},
            total=10,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=comp_ind_app,
            inscripcio=self.ins2,
            exercici=1,
            inputs={},
            outputs={},
            total=25,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived team filtered members",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "filtres": {"categories_in": ["Base"]},
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([row["participant"] for row in rows], ["Parella 1"])
        self.assertEqual(rows[0]["score"], 10.0)

    def test_classificacio_save_persists_default_exercise_selection_scope_for_derived_team(self):
        ind_app = self._create_aparell("TR", "Tramp")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        response = self._post_json(
            "classificacio_save",
            self._classificacio_payload(
                tipus="equips",
                app_ids=[comp_ind_app.id],
                context_code="parelles",
                team_mode="derived_from_individual",
            ),
        )

        self.assertEqual(response.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=response.json()["id"])
        self.assertEqual(
            (cfg.schema.get("puntuacio") or {}).get("exercise_selection_scope"),
            "per_member",
        )

    def test_classificacio_save_sanitizes_filter_lists(self):
        payload = self._classificacio_payload(
            tipus="equips",
            app_ids=[self.comp_app.id],
            context_code="parelles",
            team_mode="native_team",
        )
        payload["schema"]["filtres"] = {
            "categories_in": ["Base", "Base", "", None],
            "grups_in": [1, "1", "", None],
        }

        response = self._post_json("classificacio_save", payload)

        self.assertEqual(response.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=response.json()["id"])
        self.assertEqual(
            cfg.schema.get("filtres"),
            {"categories_in": ["Base"], "grups_in": ["1"]},
        )

    def test_classificacio_save_rejects_unknown_filter_key(self):
        payload = self._classificacio_payload(
            tipus="equips",
            app_ids=[self.comp_app.id],
            context_code="parelles",
            team_mode="native_team",
        )
        payload["schema"]["filtres"] = {"unknown_filter": ["X"]}

        response = self._post_json("classificacio_save", payload)

        self.assertEqual(response.status_code, 400)
        self.assertTrue(any("clau no admesa" in err for err in response.json().get("errors", [])))

    def test_classificacio_save_rejects_exercise_selection_scope_for_native_team(self):
        payload = self._classificacio_payload(
            tipus="equips",
            app_ids=[self.comp_app.id],
            context_code="parelles",
            team_mode="native_team",
        )
        payload["schema"]["puntuacio"]["exercise_selection_scope"] = "team_pool"
        payload["schema"]["desempat"] = [
            {
                "camps": ["TOTAL"],
                "scope": {"aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]}},
                "exercise_selection_scope": "team_pool",
            }
        ]

        response = self._post_json("classificacio_save", payload)

        self.assertEqual(response.status_code, 400)
        errors = response.json().get("errors", [])
        self.assertTrue(any("puntuacio.exercise_selection_scope" in err for err in errors))
        self.assertTrue(any("desempat[0].exercise_selection_scope" in err for err in errors))

    def test_compute_classificacio_derived_team_pool_selects_best_n_with_member_cap(self):
        ind_app = self._create_aparell("TR", "Tramp")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)
        comp_ind_app.nombre_exercicis = 2
        comp_ind_app.save(update_fields=["nombre_exercicis"])
        equip_2, members_2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)

        for ins, exercici, total in (
            (self.ins1, 1, 9.0),
            (self.ins1, 2, 8.0),
            (self.ins2, 1, 7.0),
            (self.ins2, 2, 6.0),
            (members_2[0], 1, 9.0),
            (members_2[0], 2, 5.0),
            (members_2[1], 1, 8.0),
            (members_2[1], 2, 7.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=comp_ind_app,
                inscripcio=ins,
                exercici=exercici,
                inputs={},
                outputs={},
                total=total,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived team pool",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {
                        "mode": "millor_n",
                        "best_n": 2,
                        "max_per_participant": 1,
                    },
                    "exercise_selection_scope": "team_pool",
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([row["participant"] for row in rows[:2]], ["Parella 2", "Parella 1"])
        self.assertEqual(rows[0]["punts"], 17.0)
        self.assertEqual(rows[1]["punts"], 16.0)

    def test_compute_classificacio_derived_team_pool_tie_break_uses_team_pool(self):
        ind_app = self._create_aparell("DMT", "Double Mini")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)
        comp_ind_app.nombre_exercicis = 2
        comp_ind_app.save(update_fields=["nombre_exercicis"])
        equip_2, members_2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)

        for ins, exercici, total, d_value in (
            (self.ins1, 1, 10.0, 100.0),
            (self.ins1, 2, 1.0, 0.0),
            (self.ins2, 1, 7.0, 0.0),
            (self.ins2, 2, 0.0, 0.0),
            (members_2[0], 1, 9.0, 50.0),
            (members_2[0], 2, 2.0, 0.0),
            (members_2[1], 1, 8.0, 60.0),
            (members_2[1], 2, 1.0, 0.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=comp_ind_app,
                inscripcio=ins,
                exercici=exercici,
                inputs={"D": d_value},
                outputs={},
                total=total,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived team pool tie",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 1},
                    "exercise_selection_scope": "per_member",
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["D"],
                        "ordre": "desc",
                        "exercise_selection_scope": "team_pool",
                        "scope": {
                            "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                        },
                        "agregacio_camps": "sum",
                        "agregacio_exercicis": "sum",
                        "agregacio_aparells": "sum",
                    }
                ],
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([row["participant"] for row in rows[:2]], ["Parella 2", "Parella 1"])
        self.assertEqual(rows[0]["punts"], 17.0)
        self.assertEqual(rows[1]["punts"], 17.0)

    def test_classificacio_validation_rejects_context_mismatch_for_team_context_app(self):
        schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                },
                "equips": {
                    "assignment_source": {"mode": "context", "context_code": "altre", "fallback": "native"},
                },
            },
            tipus="equips",
        )

        self.assertEqual(schema["equips"]["assignment_source"]["context_code"], "altre")
        self.assertTrue(any("requereix context parelles" in err for err in errors))

    def test_classificacio_validation_rejects_native_team_mode_with_individual_app(self):
        ind_app = self._create_aparell("TR", "Tramp")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                },
            },
            tipus="equips",
        )

        self.assertTrue(any("team_mode=native_team" in err for err in errors))

    def test_classificacio_validation_rejects_native_team_raw_column_on_individual_app(self):
        ind_app = self._create_aparell("TR_RAW_VAL", "Tramp raw validation")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "raw_total",
                            "label": "Raw total",
                            "align": "right",
                            "decimals": 3,
                            "source": {"aparell_id": comp_ind_app.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                        },
                    ]
                },
            },
            tipus="equips",
        )

        self.assertTrue(any("nomes es poden mostrar aparells d'equip" in err for err in errors))

    def test_classificacio_validation_rejects_invalid_detail_builtin_for_derived_team(self):
        ind_app = self._create_aparell("TR_DETAIL_VAL", "Tramp detail validation")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                },
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "columnes": [
                            {"type": "builtin", "key": "punts", "label": "Punts", "align": "right"},
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                },
            },
            tipus="equips",
        )

        self.assertTrue(
            any("presentacio.detall.columnes[0] builtin: clau no permesa" in err for err in errors)
        )

    def test_classificacio_validation_rejects_detail_enabled_without_sections(self):
        ind_app = self._create_aparell("TR_DETAIL_REQ", "Tramp detail required")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                },
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [],
                    },
                },
            },
            tipus="individual",
        )

        self.assertTrue(
            any("presentacio.detall.enabled requereix sections o columnes legacy compatibles" in err for err in errors)
        )

    def test_classificacio_validation_rejects_legacy_detail_columns_for_individual(self):
        ind_app = self._create_aparell("TR_DETAIL_F1_IND", "Tramp detail legacy individual")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                },
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "columnes": [
                            {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                        ],
                    },
                },
            },
            tipus="individual",
        )

        self.assertTrue(
            any("presentacio.detall.columnes nomes es compatible" in err for err in errors)
        )

    def test_classificacio_validation_rejects_legacy_detail_columns_for_native_team(self):
        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "columnes": [
                            {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertTrue(
            any("presentacio.detall.columnes nomes es compatible" in err for err in errors)
        )

    def test_normalize_schema_does_not_inject_legacy_detail_columns_when_absent(self):
        schema, _info = normalize_schema_legacy_team_birth_partition(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                    ],
                    "detall": {
                        "enabled": False,
                        "default_open": False,
                        "sections": [],
                    },
                },
            },
            tipus="equips",
            persist=False,
        )

        self.assertNotIn("columnes", (schema.get("presentacio") or {}).get("detall") or {})

    def test_normalize_schema_preserves_explicit_legacy_detail_columns(self):
        schema, _info = normalize_schema_legacy_team_birth_partition(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "columnes": [
                            {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                        ],
                    },
                },
            },
            tipus="equips",
            persist=False,
        )

        self.assertEqual(
            (((schema.get("presentacio") or {}).get("detall") or {}).get("columnes")) or [],
            [{"type": "builtin", "key": "participant", "label": "Participant", "align": "left"}],
        )

    def test_classificacio_validation_accepts_empty_legacy_detail_columns_for_native_team_sections(self):
        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {"type": "builtin", "key": "punts", "label": "Punts", "align": "right", "decimals": 3},
                    ],
                    "detall": {
                        "enabled": True,
                        "default_open": False,
                        "sections": [
                            {"type": "members_list", "label": "Participants"},
                            {
                                "type": "team_metrics",
                                "label": "Notes equip",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "team_total",
                                        "label": "Total",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercici": 1,
                                            "camp": "total",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                        "columnes": [],
                    },
                },
            },
            tipus="equips",
        )

        self.assertEqual(errors, [])

    def test_classificacio_validation_rejects_multi_app_detail_section(self):
        app_b = self._create_aparell("TR_DETAIL_MULTI", "Tramp detail multi")
        comp_app_b = self._create_comp_aparell(self.comp, app_b, ordre=3)

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id, comp_app_b.id]},
                    "camps_per_aparell": {
                        str(self.comp_app.id): ["total"],
                        str(comp_app_b.id): ["total"],
                    },
                },
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "members_table",
                                "label": "Detall",
                                "columns": [
                                    {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                                    {
                                        "type": "raw",
                                        "key": "detail_total_a",
                                        "label": "Total A",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {"aparell_id": self.comp_app.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                                    },
                                    {
                                        "type": "raw",
                                        "key": "detail_total_b",
                                        "label": "Total B",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {"aparell_id": comp_app_b.id, "exercici": 1, "camp": "total", "jutges": {"ids": []}},
                                    },
                                ],
                            }
                        ],
                    },
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                },
            },
            tipus="equips",
        )

        self.assertTrue(
            any("presentacio.detall.sections[0] barreja aparells multiples" in err for err in errors)
        )

    def test_classificacio_save_returns_error_details_for_invalid_detail_section_field(self):
        ind_app = self._create_aparell("TR_DETAIL_BAD", "Tramp detail bad")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)
        payload = self._classificacio_payload(tipus="individual", app_ids=[comp_ind_app.id])
        payload["schema"]["presentacio"] = {
            "top_n": 0,
            "mostrar_empats": True,
            "columnes": [
                {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
            ],
            "detall": {
                "enabled": True,
                "sections": [
                    {
                        "type": "exercise_table",
                        "label": "Exercicis",
                        "aparell_id": comp_ind_app.id,
                        "columns": [
                            {"type": "builtin", "key": "exercise_index", "label": "Ex.", "align": "left"},
                            {
                                "type": "raw",
                                "key": "detail_bad",
                                "label": "Camp invalid",
                                "align": "right",
                                "decimals": 3,
                                "source": {"aparell_id": comp_ind_app.id, "exercici": 1, "camp": "NO_EXISTEIX", "jutges": {"ids": []}},
                            },
                        ],
                    }
                ],
            },
        }

        response = self._post_json("classificacio_save", payload)
        self.assertEqual(response.status_code, 400)
        body = response.json()
        details = body.get("error_details") or []
        self.assertTrue(any(item.get("path") == "presentacio.detall.sections[0].columns[1].source.camp" for item in details))

    def test_classificacio_save_rejects_detail_exercici_out_of_range_with_precise_path(self):
        ind_app = self._create_aparell("TR_DETAIL_RANGE", "Tramp detail range")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)
        payload = self._classificacio_payload(tipus="individual", app_ids=[comp_ind_app.id])
        payload["schema"]["presentacio"] = {
            "top_n": 0,
            "mostrar_empats": True,
            "columnes": [
                {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
            ],
            "detall": {
                "enabled": True,
                "sections": [
                    {
                        "type": "exercise_table",
                        "label": "Exercicis",
                        "aparell_id": comp_ind_app.id,
                        "columns": [
                            {"type": "builtin", "key": "exercise_index", "label": "Ex.", "align": "left"},
                            {
                                "type": "raw",
                                "key": "detail_total",
                                "label": "Total",
                                "align": "right",
                                "decimals": 3,
                                "source": {"aparell_id": comp_ind_app.id, "exercici": 99, "camp": "total", "jutges": {"ids": []}},
                            },
                        ],
                    }
                ],
            },
        }

        response = self._post_json("classificacio_save", payload)
        self.assertEqual(response.status_code, 400)
        body = response.json()
        self.assertTrue(any("fora de rang" in err for err in (body.get("errors") or [])))
        details = body.get("error_details") or []
        self.assertTrue(any(item.get("path") == "presentacio.detall.sections[0].columns[1].source.exercici" for item in details))

    def test_classificacio_validation_rejects_members_list_columns(self):
        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                },
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "members_list",
                                "label": "Membres",
                                "columns": [
                                    {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                                ],
                            }
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertTrue(any("presentacio.detall.sections[0].columns no es compatible amb members_list" in err for err in errors))

    def test_classificacio_validation_accepts_native_team_team_members_table_member_field(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {"type": "builtin", "key": "participant", "label": "Participant", "align": "left"},
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {"aparell_id": self.comp_app.id, "exercici": 1, "camp": "E", "jutges": {"ids": []}},
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertEqual(errors, [])

    def test_classificacio_validation_accepts_native_team_team_members_table_fixed_exercise_mode(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercise_mode": "fixed",
                                            "exercici": 1,
                                            "camp": "E",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertEqual(errors, [])

    def test_classificacio_validation_legacy_team_members_table_ignores_exercici_when_mode_missing(self):
        self.comp_app.nombre_exercicis = 2
        self.comp_app.save(update_fields=["nombre_exercicis"])
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercici": 99,
                                            "camp": "E",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertEqual(errors, [])

    def test_classificacio_validation_rejects_native_team_team_members_table_invalid_exercise_mode(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {
                                            "aparell_id": self.comp_app.id,
                                            "exercise_mode": "broken",
                                            "camp": "E",
                                            "jutges": {"ids": []},
                                        },
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertTrue(any("exercise_mode invalid" in err for err in errors))

    def test_prepare_schema_for_persistence_preserves_team_members_table_exercise_mode(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )
        schema = {
            **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
            "presentacio": {
                "detall": {
                    "enabled": True,
                    "sections": [
                        {
                            "type": "team_members_table",
                            "label": "Notes per membre",
                            "aparell_id": self.comp_app.id,
                            "columns": [
                                {
                                    "type": "raw",
                                    "key": "member_exec",
                                    "label": "Exec",
                                    "align": "right",
                                    "decimals": 3,
                                    "source": {
                                        "aparell_id": self.comp_app.id,
                                        "exercise_mode": "fixed",
                                        "exercici": 1,
                                        "camp": "E",
                                        "jutges": {"ids": []},
                                    },
                                },
                            ],
                        },
                    ],
                },
            },
            "equips": {
                "context_code": "parelles",
                "assignment_source": {"mode": "context", "context_code": "parelles", "fallback": "native"},
                "team_mode": "native_team",
                "incloure_sense_equip": False,
            },
        }

        prepared = prepare_schema_for_persistence(self.comp, schema, tipus="equips")

        self.assertEqual(prepared["errors"], [])
        col_source = (((((prepared["schema"] or {}).get("presentacio") or {}).get("detall") or {}).get("sections") or [])[0]["columns"][0]["source"])
        self.assertEqual(col_source.get("exercise_mode"), "fixed")
        self.assertNotIn("has_explicit_exercici", col_source)

    def test_competition_template_roundtrip_preserves_team_members_table_exercise_mode(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )
        schema = {
            **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
            "presentacio": {
                "detall": {
                    "enabled": True,
                    "sections": [
                        {
                            "type": "team_members_table",
                            "label": "Notes per membre",
                            "aparell_id": self.comp_app.id,
                            "columns": [
                                {
                                    "type": "raw",
                                    "key": "member_exec",
                                    "label": "Exec",
                                    "align": "right",
                                    "decimals": 3,
                                    "source": {
                                        "aparell_id": self.comp_app.id,
                                        "exercise_mode": "fixed",
                                        "exercici": 1,
                                        "camp": "E",
                                        "jutges": {"ids": []},
                                    },
                                },
                            ],
                        },
                    ],
                },
            },
            "equips": {
                "context_code": "parelles",
                "assignment_source": {"mode": "context", "context_code": "parelles", "fallback": "native"},
                "team_mode": "native_team",
                "incloure_sense_equip": False,
            },
        }

        schema_tpl, warnings = _schema_to_template_schema(self.comp, schema)
        self.assertEqual(warnings, [])
        tpl_source = (((schema_tpl.get("presentacio") or {}).get("detall") or {}).get("sections") or [])[0]["columns"][0]["source"]
        self.assertEqual(tpl_source.get("exercise_mode"), "fixed")
        self.assertNotIn("has_explicit_exercici", tpl_source)

        schema_roundtrip, compat_warnings, mapping, compat_meta = _template_schema_to_competicio_schema_service(
            self.comp,
            schema_tpl,
        )
        self.assertEqual(mapping.get(self.app.codi), self.comp_app.id)
        self.assertEqual(compat_warnings, [])
        self.assertFalse(compat_meta.get("adaptable"))
        roundtrip_source = (((schema_roundtrip.get("presentacio") or {}).get("detall") or {}).get("sections") or [])[0]["columns"][0]["source"]
        self.assertEqual(roundtrip_source.get("exercise_mode"), "fixed")
        self.assertNotIn("has_explicit_exercici", roundtrip_source)

    def test_classificacio_validation_accepts_native_team_team_members_table_display_only_member_field(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {
                        "code": "E",
                        "label": "Exec",
                        "type": "matrix",
                        "shape": "judge_x_item",
                        "scope": "member",
                        "judges": {"count": 3},
                        "items": {"count": 5},
                    },
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {"aparell_id": self.comp_app.id, "exercici": 1, "camp": "E", "jutges": {"ids": []}},
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertEqual(errors, [])

    def test_classificacio_validation_infers_team_members_table_section_app_from_single_raw_app(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {"aparell_id": self.comp_app.id, "exercici": 1, "camp": "E", "jutges": {"ids": []}},
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertEqual(errors, [])

    def test_classificacio_validation_rejects_native_team_team_members_table_shared_field(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_members_table",
                                "label": "Notes per membre",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "team_sync",
                                        "label": "Sync",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {"aparell_id": self.comp_app.id, "exercici": 1, "camp": "SYNC", "jutges": {"ids": []}},
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertTrue(any("team_members_table" in err for err in errors))

    def test_classificacio_validation_rejects_native_team_main_column_member_field(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "columnes": [
                        {"type": "builtin", "key": "participant", "label": "Nom", "align": "left"},
                        {
                            "type": "raw",
                            "key": "member_exec",
                            "label": "Exec",
                            "align": "right",
                            "decimals": 3,
                            "source": {"aparell_id": self.comp_app.id, "exercici": 1, "camp": "E", "jutges": {"ids": []}},
                        },
                    ],
                },
            },
            tipus="equips",
        )

        self.assertTrue(any("camps individuals per membre" in err for err in errors))

    def test_classificacio_validation_rejects_native_team_team_metrics_member_field(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            {
                **self._native_team_schema_with_tie({"camps": ["TOTAL"], "ordre": "desc"}),
                "presentacio": {
                    "detall": {
                        "enabled": True,
                        "sections": [
                            {
                                "type": "team_metrics",
                                "label": "Notes equip",
                                "aparell_id": self.comp_app.id,
                                "columns": [
                                    {
                                        "type": "raw",
                                        "key": "member_exec",
                                        "label": "Exec",
                                        "align": "right",
                                        "decimals": 3,
                                        "source": {"aparell_id": self.comp_app.id, "exercici": 1, "camp": "E", "jutges": {"ids": []}},
                                    },
                                ],
                            },
                        ],
                    },
                },
            },
            tipus="equips",
        )

        self.assertTrue(any("team_metrics" in err for err in errors))

    def test_normalize_excel_cell_supports_team_raw_detail(self):
        value, _fmt, wrap = _normalize_excel_cell(
            {
                "_kind": "team_raw_detail",
                "summary": 23.75,
                "rows": [
                    {"label": "Maria", "value": 12.5},
                    {
                        "label": "Laia",
                        "judge_rows": {
                            "_kind": "judge_rows",
                            "rows": [{"judge": 1, "items": [7.4, 7.6]}],
                        },
                    },
                ],
            },
            {"decimals": 2},
        )

        self.assertEqual(value, "23.75\nMaria: 12.50\nLaia:\n  J1: 7.40 | 7.60")
        self.assertTrue(wrap)

    def test_classificacio_validation_rejects_native_team_tie_with_participant_scope(self):
        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            self._native_team_schema_with_tie(
                {
                    "camps": ["TOTAL"],
                    "scope": {"participants": {"mode": "millor_n", "n": 1}},
                }
            ),
            tipus="equips",
        )

        self.assertTrue(any("scope.participants" in err for err in errors))

    def test_classificacio_validation_rejects_native_team_tie_with_participant_aggregation(self):
        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            self._native_team_schema_with_tie(
                {
                    "camps": ["TOTAL"],
                    "agregacio_participants": "sum",
                }
            ),
            tipus="equips",
        )

        self.assertTrue(any("agregacio_participants" in err for err in errors))

    def test_classificacio_validation_rejects_native_team_tie_with_non_scalar_member_field(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            self._native_team_schema_with_tie({"camps": ["E"]}),
            tipus="equips",
        )

        self.assertTrue(errors)

    def test_classificacio_validation_accepts_native_team_tie_with_shared_computed_scalar(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TEAM_SYNC", "label": "Team sync", "formula": "SYNC"},
                ],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            self._native_team_schema_with_tie({"camps": ["TEAM_SYNC"]}),
            tipus="equips",
        )

        self.assertEqual(errors, [])

    def test_classificacio_validation_accepts_native_team_tie_with_member_derived_scalar_computed(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TEAM_E", "label": "Team exec", "formula": "members_sum(E)"},
                ],
            },
        )

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            self._native_team_schema_with_tie({"camps": ["TEAM_E"]}),
            tipus="equips",
        )

        self.assertEqual(errors, [])

    def test_compute_classificacio_native_team_tie_uses_team_scores_only(self):
        equip_2, _members = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TEAM_SYNC", "label": "Team sync", "formula": "SYNC"},
                ],
            },
        )
        team_subject_1, _subject_meta_1 = self._team_subject(self.equip)
        team_subject_2, _subject_meta_2 = self._team_subject(equip_2)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject_1,
            exercici=1,
            inputs={"SYNC": 9},
            outputs={"TEAM_SYNC": 9},
            total=30,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject_2,
            exercici=1,
            inputs={"SYNC": 7},
            outputs={"TEAM_SYNC": 7},
            total=30,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={"TOTAL": 999},
            outputs={},
            total=999,
        )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team tie strict",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                **self._native_team_schema_with_tie({"camps": ["TEAM_SYNC"], "ordre": "desc"}),
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertEqual([row["participant"] for row in rows], ["Parella 1", "Parella 2"])

    def test_compute_classificacio_derived_team_birth_partition_uses_oldest_member_and_strict_rule(self):
        self.comp.data = date(2025, 5, 10)
        self.comp.save(update_fields=["data"])
        ind_app = self._create_aparell("TRA", "Tramp")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)
        equip_2, members_2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)

        self.ins1.data_naixement = date(2012, 6, 1)
        self.ins2.data_naixement = date(2016, 6, 1)
        self.ins1.save(update_fields=["data_naixement"])
        self.ins2.save(update_fields=["data_naixement"])
        members_2[0].data_naixement = date(2012, 7, 1)
        members_2[1].data_naixement = date(2013, 7, 1)
        members_2[0].save(update_fields=["data_naixement"])
        members_2[1].save(update_fields=["data_naixement"])

        for ins, total in (
            (self.ins1, 8.0),
            (self.ins2, 7.0),
            (members_2[0], 9.0),
            (members_2[1], 6.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=comp_ind_app,
                inscripcio=ins,
                exercici=1,
                inputs={},
                outputs={},
                total=total,
            )

        schema = {
            **self._birth_range_partition_cfg(compliance_mode="strict"),
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "tots"},
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
            },
            "equips": {
                "context_code": "parelles",
                "team_mode": "derived_from_individual",
                "incloure_sense_equip": False,
            },
        }
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Teams by birth range",
            activa=True,
            ordre=1,
            tipus="equips",
            schema=schema,
        )

        out = compute_classificacio(self.comp, cfg)

        self.assertEqual(out["any_naixement_forquilla:U13"][0]["participant"], "Parella 2")
        self.assertEqual(out["any_naixement_forquilla:Fora de forquilla"][0]["participant"], "Parella 1")

    def test_compute_classificacio_derived_team_birth_partition_allow_outside_n(self):
        self.comp.data = date(2025, 5, 10)
        self.comp.save(update_fields=["data"])
        ind_app = self._create_aparell("DMT", "Dmt")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        self.ins1.data_naixement = date(2012, 6, 1)
        self.ins2.data_naixement = date(2016, 6, 1)
        self.ins1.save(update_fields=["data_naixement"])
        self.ins2.save(update_fields=["data_naixement"])

        for ins, total in (
            (self.ins1, 8.0),
            (self.ins2, 7.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=comp_ind_app,
                inscripcio=ins,
                exercici=1,
                inputs={},
                outputs={},
                total=total,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Teams by birth range allow one outside",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                **self._birth_range_partition_cfg(compliance_mode="allow_outside_n", max_outside=1),
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_ind_app.id]},
                    "camps_per_aparell": {str(comp_ind_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        out = compute_classificacio(self.comp, cfg)

        self.assertEqual([row["participant"] for row in out["any_naixement_forquilla:U13"]], ["Parella 1"])
        self.assertNotIn("any_naixement_forquilla:Fora de forquilla", out)

    def test_compute_classificacio_native_team_birth_partition_uses_team_members(self):
        self.comp.data = date(2025, 5, 10)
        self.comp.save(update_fields=["data"])
        equip_2, members_2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)

        self.ins1.data_naixement = date(2012, 6, 1)
        self.ins2.data_naixement = date(2016, 6, 1)
        self.ins1.save(update_fields=["data_naixement"])
        self.ins2.save(update_fields=["data_naixement"])
        members_2[0].data_naixement = date(2012, 7, 1)
        members_2[1].data_naixement = date(2013, 7, 1)
        members_2[0].save(update_fields=["data_naixement"])
        members_2[1].save(update_fields=["data_naixement"])

        team_subject_1, _meta_1 = self._team_subject(self.equip)
        team_subject_2, _meta_2 = self._team_subject(equip_2)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject_1,
            exercici=1,
            inputs={},
            outputs={},
            total=30,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject_2,
            exercici=1,
            inputs={},
            outputs={},
            total=28,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native teams by birth range",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                **self._birth_range_partition_cfg(compliance_mode="strict"),
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        out = compute_classificacio(self.comp, cfg)

        self.assertEqual(out["any_naixement_forquilla:U13"][0]["participant"], "Parella 2")
        self.assertEqual(out["any_naixement_forquilla:Fora de forquilla"][0]["participant"], "Parella 1")

    def test_compute_classificacio_native_team_birth_partition_dedupes_member_ids(self):
        self.comp.data = date(2025, 5, 10)
        self.comp.save(update_fields=["data"])
        equip_2, members_2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)

        self.ins1.data_naixement = date(2012, 6, 1)
        self.ins2.data_naixement = date(2016, 6, 1)
        self.ins1.save(update_fields=["data_naixement"])
        self.ins2.save(update_fields=["data_naixement"])
        members_2[0].data_naixement = date(2012, 7, 1)
        members_2[1].data_naixement = date(2013, 7, 1)
        members_2[0].save(update_fields=["data_naixement"])
        members_2[1].save(update_fields=["data_naixement"])

        team_subject_1, _meta_1 = self._team_subject(self.equip)
        team_subject_2, _meta_2 = self._team_subject(equip_2)
        team_subject_1.member_ids = [self.ins1.id, self.ins2.id, self.ins2.id]
        team_subject_1.save(update_fields=["member_ids"])

        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject_1,
            exercici=1,
            inputs={},
            outputs={},
            total=30,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject_2,
            exercici=1,
            inputs={},
            outputs={},
            total=28,
        )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Native teams by birth range deduped members",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                **self._birth_range_partition_cfg(compliance_mode="allow_outside_n", max_outside=1),
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "incloure_sense_equip": False,
                },
            },
        )

        out = compute_classificacio(self.comp, cfg)

        self.assertEqual(
            [row["participant"] for row in out["any_naixement_forquilla:U13"]],
            ["Parella 1", "Parella 2"],
        )
        self.assertNotIn("any_naixement_forquilla:Fora de forquilla", out)

    def test_classificacio_save_normalizes_legacy_team_age_partition_to_birth_range(self):
        self.comp.data = date(2025, 5, 10)
        self.comp.save(update_fields=["data"])
        ind_app = self._create_aparell("MINI", "Mini")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        payload = self._classificacio_payload(
            tipus="equips",
            app_ids=[comp_ind_app.id],
            context_code="parelles",
            team_mode="derived_from_individual",
        )
        payload["schema"]["equips"]["particio_edat"] = {
            "activa": True,
            "llindars": [12],
            "sense_data_label": "Sense edat",
        }

        response = self._post_json("classificacio_save", payload)

        self.assertEqual(response.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=response.json()["id"])
        part_cfg = ((cfg.schema or {}).get("particions_config") or {}).get("any_naixement_forquilla") or {}
        self.assertIn("any_naixement_forquilla", (cfg.schema or {}).get("particions") or [])
        self.assertTrue(part_cfg.get("ranges"))
        self.assertEqual((part_cfg.get("team_rules") or {}).get("reference_mode"), "oldest_member_birthdate")
        self.assertFalse((((cfg.schema or {}).get("equips") or {}).get("particio_edat") or {}).get("activa", False))

    def test_classificacio_save_rejects_team_app_for_individual_tipus(self):
        response = self._post_json(
            "classificacio_save",
            self._classificacio_payload(tipus="individual", app_ids=[self.comp_app.id]),
        )

        self.assertEqual(response.status_code, 400)
        body = response.json()
        self.assertTrue(any("tipus='individual'" in err for err in body.get("errors", [])))

    def test_classificacio_save_forces_derived_team_mode_when_context_has_no_team_apps(self):
        ind_app = self._create_aparell("DMT", "Doble mini")
        comp_ind_app = self._create_comp_aparell(self.comp, ind_app, ordre=2)

        response = self._post_json(
            "classificacio_save",
            self._classificacio_payload(
                tipus="equips",
                app_ids=[comp_ind_app.id],
                context_code="altre",
                team_mode=None,
            ),
        )

        self.assertEqual(response.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=response.json()["id"])
        equips_cfg = cfg.schema.get("equips") or {}
        self.assertEqual(equips_cfg.get("context_code"), "altre")
        self.assertEqual(equips_cfg.get("team_mode"), "derived_from_individual")
        self.assertEqual((equips_cfg.get("mode_resolution") or {}).get("eligible_team_app_ids_at_save"), [])
        self.assertTrue((equips_cfg.get("mode_resolution") or {}).get("resolved_at"))

    def test_classificacio_preview_rejects_stale_native_team_context(self):
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team preview stale",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "mode_resolution": {
                        "resolved_at": "2026-03-29T10:00:00Z",
                        "eligible_team_app_ids_at_save": [self.comp_app.id],
                    },
                },
            },
        )
        CompeticioAparellEquipContextSource.objects.filter(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            context=self.ctx,
        ).delete()

        response = self.client.post(reverse("classificacio_preview", kwargs={"pk": self.comp.id, "cid": cfg.id}))

        self.assertEqual(response.status_code, 400)
        body = response.json()
        self.assertFalse(body.get("ok"))
        self.assertTrue(any("no admet el context" in err for err in body.get("errors", [])))

    def test_classificacions_home_exposes_team_context_capabilities_and_cfg_statuses(self):
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Cfg status v1",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "native_team",
                    "mode_resolution": {
                        "resolved_at": "2026-03-29T10:00:00Z",
                        "eligible_team_app_ids_at_save": [self.comp_app.id],
                    },
                },
            },
        )

        response = self.client.get(reverse("classificacions_home", kwargs={"pk": self.comp.id}))

        self.assertEqual(response.status_code, 200)
        self.assertIn("cfg_statuses", response.context)
        self.assertIn(str(cfg.id), response.context["cfg_statuses"])
        self.assertIn("resolved_at", response.context["cfg_statuses"][str(cfg.id)])
        self.assertEqual(
            response.context["cfg_statuses"][str(cfg.id)]["resolved_at"],
            "2026-03-29T10:00:00Z",
        )
        self.assertTrue(
            any(item.get("context_code") == "parelles" for item in response.context.get("team_context_capabilities", []))
        )

    def test_competicio_aparell_form_uses_current_team_contract_fields(self):
        form = CompeticioAparellForm(
            data={
                "aparell": self.app.id,
                "nombre_exercicis": 1,
            },
            instance=self.comp_app,
            competicio=self.comp,
        )
        self.assertIn("aparell", form.fields)
        self.assertIn("nombre_exercicis", form.fields)
        self.assertNotIn("team_context", form.fields)
        self.assertNotIn("expected_team_size", form.fields)
        self.assertNotIn("team_scoring_mode", form.fields)
        self.assertTrue(form.is_valid())

    def test_scoring_updates_include_invalid_team_entries_with_series_state(self):
        invalid_team, _members = self._create_team_with_members("Parella incompleta", ["Berta"], start_order=20)
        team_subject, _subject_meta = self._team_subject()
        invalid_team_subject, _invalid_meta = self._team_subject(invalid_team)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 5},
            outputs={},
            total=5,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=invalid_team_subject,
            exercici=1,
            inputs={"SYNC": 9},
            outputs={},
            total=9,
        )

        res = self.client.get(
            reverse("scoring_updates", kwargs={"pk": self.comp.id}),
            {
                "since": (timezone.now() - timedelta(minutes=10)).isoformat(),
                "comp_aparell_id": self.comp_app.id,
            },
        )
        self.assertEqual(res.status_code, 200)
        updates = res.json().get("updates", [])
        self.assertEqual({u["subject_id"] for u in updates}, {team_subject.id, invalid_team_subject.id})
        by_id = {int(row["subject_id"]): row for row in updates}
        self.assertEqual(by_id[team_subject.id]["series_state"], "unassigned")
        self.assertEqual(by_id[invalid_team_subject.id]["series_state"], "invalid")

    def test_scoring_notes_home_renders_team_schema_without_nameerror(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sincronisme", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC + E__m1 + E__m2"},
                ],
            },
        )

        response = self.client.get(reverse("scoring_notes_home", kwargs={"pk": self.comp.id}))

        self.assertEqual(response.status_code, 200)
        schema = response.context["schemas"][str(self.comp_app.id)]
        logical_schema = response.context["logical_schemas"][str(self.comp_app.id)]
        self.assertIn("E__m1", {field["code"] for field in schema["fields"]})
        self.assertIn("E__m2", {field["code"] for field in schema["fields"]})
        self.assertEqual({field["code"] for field in logical_schema["fields"]}, {"SYNC", "E"})

    def test_scoring_notes_home_exposes_scoped_group_apps(self):
        indiv_app = self._create_aparell("IND", "Individual")
        indiv_comp_app = self._create_comp_aparell(self.comp, indiv_app, ordre=2)
        ScoringSchema.objects.create(
            aparell=indiv_app,
            schema={
                "fields": [{"code": "N", "label": "Nota", "type": "number"}],
                "computed": [{"code": "TOTAL", "label": "Total", "formula": "N"}],
            },
        )

        response = self.client.get(reverse("scoring_notes_home", kwargs={"pk": self.comp.id}))

        self.assertEqual(response.status_code, 200)
        groups_render = list(response.context["groups_render"]) + list(response.context["out_of_program_groups_render"])
        individual_groups = [item for item in groups_render if item["kind"] == "individual_group"]
        team_groups = [item for item in groups_render if item["kind"] == "team_bucket"]
        self.assertTrue(any(int(app.id) == indiv_comp_app.id for item in individual_groups for app in item["apps"]))
        self.assertTrue(any(int(app.id) == self.comp_app.id for item in team_groups for app in item["apps"]))
        self.assertFalse(any(int(app.id) == self.comp_app.id for item in individual_groups for app in item["apps"]))

    def test_inscripcions_list_exposes_series_panel_navigation(self):
        response = self.client.get(reverse("inscripcions_list", kwargs={"pk": self.comp.id}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-panel-target="series-equips"')
        self.assertContains(response, 'id="panel-series-equips"')

    def test_series_preview_signature_is_required_for_selection_actions(self):
        equip2, _members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        subject_1, _meta_1 = self._team_subject()
        subject_2, _meta_2 = self._team_subject(equip2)
        serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie A",
        )
        SerieEquipItem.objects.create(serie=serie, team_subject=subject_1, ordre=1)

        preview_res = self._post_json(
            "inscripcions_series_equips_preview",
            {
                "comp_aparell_id": self.comp_app.id,
                "action": "assign",
                "serie_id": serie.id,
                "selected_ids": [subject_2.id],
            },
        )
        self.assertEqual(preview_res.status_code, 200)
        plan_signature = preview_res.json()["preview"]["plan_signature"]

        missing_signature_res = self._post_json(
            "inscripcions_series_equips_assign",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie.id,
                "selected_ids": [subject_2.id],
            },
        )
        self.assertEqual(missing_signature_res.status_code, 400)

        stale_signature_res = self._post_json(
            "inscripcions_series_equips_assign",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie.id,
                "selected_ids": [subject_2.id],
                "plan_signature": "stale-signature",
            },
        )
        self.assertEqual(stale_signature_res.status_code, 400)

        assign_res = self._post_json(
            "inscripcions_series_equips_assign",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie.id,
                "selected_ids": [subject_2.id],
                "plan_signature": plan_signature,
            },
        )
        self.assertEqual(assign_res.status_code, 200)

    def test_series_delete_is_blocked_while_programmed(self):
        serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie Buida",
        )
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici=timezone.datetime(2025, 1, 1, 9, 0).time(),
            hora_fi=timezone.datetime(2025, 1, 1, 10, 0).time(),
            ordre=1,
            titol="Franja 1",
        )
        estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=1,
        )
        assignacio = RotacioAssignacio.objects.create(
            competicio=self.comp,
            franja=franja,
            estacio=estacio,
        )
        RotacioAssignacioSerieEquip.objects.create(assignacio=assignacio, serie=serie, ordre=1)

        preview_res = self._post_json(
            "inscripcions_series_equips_preview",
            {
                "comp_aparell_id": self.comp_app.id,
                "action": "delete",
                "serie_id": serie.id,
            },
        )
        self.assertEqual(preview_res.status_code, 200)
        preview = preview_res.json()["preview"]
        self.assertFalse(preview["can_run"])
        self.assertEqual(preview["reason"], "serie_programmed")

        ok, reason = safe_deactivate_empty_serie(serie)
        self.assertFalse(ok)
        self.assertEqual(reason, "serie_programmed")

    def test_scoring_media_context_rejects_ineligible_team_subject(self):
        invalid_team, _members = self._create_team_with_members("Parella incompleta", ["Berta"], start_order=20)
        invalid_team_subject, _invalid_meta = self._team_subject(invalid_team)
        res = self.client.get(
            reverse("scoring_media_context", kwargs={"pk": self.comp.id}),
            {
                "comp_aparell_id": self.comp_app.id,
                "subject_kind": "team_unit",
                "subject_id": invalid_team_subject.id,
                "exercici": 1,
            },
        )
        self.assertEqual(res.status_code, 403)

    def test_scoring_notes_home_exposes_canonical_score_keys_and_invalid_teams(self):
        invalid_team, _members = self._create_team_with_members("Parella incompleta", ["Berta"], start_order=20)
        team_subject, _subject_meta = self._team_subject()
        invalid_team_subject, _invalid_meta = self._team_subject(invalid_team)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 5},
            outputs={},
            total=5,
        )
        response = self.client.get(reverse("scoring_notes_home", kwargs={"pk": self.comp.id}))
        self.assertEqual(response.status_code, 200)
        scores = response.context["scores"]
        self.assertIn(f"team_unit:{team_subject.id}|1|{self.comp_app.id}", scores)
        subjects = {str(item["id"]): item for item in response.context["inscripcions"]}
        self.assertIn(f"team_unit:{invalid_team_subject.id}", subjects)
        self.assertTrue(subjects[f"team_unit:{invalid_team_subject.id}"]["invalid_reasons"])

    def test_compute_classificacio_team_tie_break_uses_team_score_entries(self):
        equip2, members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        team_subject, _subject_meta = self._team_subject()
        team_subject_2, _subject_meta_2 = self._team_subject(equip2)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"TOTAL": 30},
            outputs={"SYNC": 9},
            total=30,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject_2,
            exercici=1,
            inputs={"TOTAL": 30},
            outputs={"SYNC": 8},
            total=30,
        )
        for ins, raw_total in [(self.ins1, 1), (self.ins2, 1), (members2[0], 50), (members2[1], 50)]:
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=self.comp_app,
                inscripcio=ins,
                exercici=1,
                inputs={},
                outputs={"SYNC": raw_total},
                total=raw_total,
            )
        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Team tiebreak",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]},
                    "camps_per_aparell": {str(self.comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "tots"},
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camp": "SYNC",
                        "ordre": "desc",
                        "scope": {"aparells": {"mode": "seleccionar", "ids": [self.comp_app.id]}},
                    }
                ],
                "equips": {
                    "assignment_source": {"mode": "context", "context_code": "parelles", "fallback": "native"},
                    "incloure_sense_equip": False,
                },
            },
        )
        rows = compute_classificacio(self.comp, cfg).get("global", [])
        self.assertEqual([row["participant"] for row in rows[:2]], ["Parella 1", "Parella 2"])

    def test_series_workspace_respects_persistent_order_and_hides_inactive_series(self):
        equip2, _members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        equip3, _members3 = self._create_team_with_members("Parella 3", ["Jana", "Paula"], start_order=40)
        subject_1, _meta_1 = self._team_subject()
        subject_2, _meta_2 = self._team_subject(equip2)
        subject_3, _meta_3 = self._team_subject(equip3)

        create_preview_res = self._post_json(
            "inscripcions_series_equips_preview",
            {
                "comp_aparell_id": self.comp_app.id,
                "action": "create",
                "name": "Serie Alpha",
                "selected_ids": [subject_1.id, subject_2.id, subject_3.id],
            },
        )
        self.assertEqual(create_preview_res.status_code, 200)
        create_plan_signature = create_preview_res.json()["preview"]["plan_signature"]

        create_res = self._post_json(
            "inscripcions_series_equips_create",
            {
                "comp_aparell_id": self.comp_app.id,
                "name": "Serie Alpha",
                "selected_ids": [subject_1.id, subject_2.id, subject_3.id],
                "plan_signature": create_plan_signature,
            },
        )
        self.assertEqual(create_res.status_code, 200)
        serie = SerieEquip.objects.get(pk=create_res.json()["serie_id"])

        reorder_res = self._post_json(
            "inscripcions_series_equips_reorder",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie.id,
                "subject_ids": [subject_3.id, subject_1.id, subject_2.id],
            },
        )
        self.assertEqual(reorder_res.status_code, 200)

        empty_res = self._post_json(
            "inscripcions_series_equips_create",
            {
                "comp_aparell_id": self.comp_app.id,
                "name": "Serie Buida",
                "selected_ids": [],
            },
        )
        self.assertEqual(empty_res.status_code, 200)
        empty_serie_id = empty_res.json()["serie_id"]
        delete_preview_res = self._post_json(
            "inscripcions_series_equips_preview",
            {
                "comp_aparell_id": self.comp_app.id,
                "action": "delete",
                "serie_id": empty_serie_id,
            },
        )
        self.assertEqual(delete_preview_res.status_code, 200)
        delete_plan_signature = delete_preview_res.json()["preview"]["plan_signature"]
        delete_res = self._post_json(
            "inscripcions_series_equips_delete",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": empty_serie_id,
                "plan_signature": delete_plan_signature,
            },
        )
        self.assertEqual(delete_res.status_code, 200)

        workspace_res = self._post_json(
            "inscripcions_series_equips_workspace",
            {"comp_aparell_id": self.comp_app.id},
        )
        self.assertEqual(workspace_res.status_code, 200)
        workspace = workspace_res.json()["workspace"]
        self.assertEqual([row["id"] for row in workspace["series"]], [serie.id])
        self.assertEqual(
            [row["subject_id"] for row in workspace["series"][0]["subjects"]],
            [subject_3.id, subject_1.id, subject_2.id],
        )
        self.assertFalse(SerieEquip.objects.get(pk=empty_serie_id).actiu)

    def test_series_workspace_and_detail_include_compact_fields(self):
        equip2, _members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        subject_1, _meta_1 = self._team_subject()
        subject_2, _meta_2 = self._team_subject(equip2)
        serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie Compacta",
        )
        SerieEquipItem.objects.create(serie=serie, team_subject=subject_2, ordre=1)

        workspace_res = self._post_json(
            "inscripcions_series_equips_workspace",
            {"comp_aparell_id": self.comp_app.id},
        )
        self.assertEqual(workspace_res.status_code, 200)
        workspace = workspace_res.json()["workspace"]

        candidate = next(row for row in workspace["candidates"]["items"] if row["subject_id"] == subject_2.id)
        self.assertEqual(candidate["members_count"], 2)
        self.assertEqual(candidate["members_preview"], "Nora + Marta")
        self.assertIn("2 membres", candidate["compact_meta"])

        serie_row = workspace["series"][0]
        self.assertEqual(serie_row["summary_label"], "Serie Compacta · 1 unitat · no programada")
        self.assertEqual(serie_row["subjects"][0]["members_preview"], "Nora + Marta")
        self.assertEqual(serie_row["subjects"][0]["members_count"], 2)

        detail_res = self._post_json(
            "inscripcions_series_equips_detail",
            {"comp_aparell_id": self.comp_app.id, "serie_id": serie.id},
        )
        self.assertEqual(detail_res.status_code, 200)
        detail = detail_res.json()["serie"]
        self.assertEqual(detail["summary_label"], "Serie Compacta · 1 unitat · no programada")
        self.assertEqual(detail["subjects"][0]["members_preview"], "Nora + Marta")
        self.assertEqual(detail["subjects"][0]["members_count"], 2)

    def test_series_assignment_moves_subject_between_active_series(self):
        subject_1, _meta_1 = self._team_subject()
        serie_a = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie A",
        )
        serie_b = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=2,
            nom="Serie B",
        )
        SerieEquipItem.objects.create(serie=serie_a, team_subject=subject_1, ordre=1)

        preview_res = self._post_json(
            "inscripcions_series_equips_preview",
            {
                "comp_aparell_id": self.comp_app.id,
                "action": "assign",
                "serie_id": serie_b.id,
                "selected_ids": [subject_1.id],
            },
        )
        self.assertEqual(preview_res.status_code, 200)
        plan_signature = preview_res.json()["preview"]["plan_signature"]

        assign_res = self._post_json(
            "inscripcions_series_equips_assign",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie_b.id,
                "selected_ids": [subject_1.id],
                "plan_signature": plan_signature,
            },
        )
        self.assertEqual(assign_res.status_code, 200)
        self.assertFalse(SerieEquipItem.objects.filter(serie=serie_a, team_subject=subject_1).exists())
        self.assertTrue(SerieEquipItem.objects.filter(serie=serie_b, team_subject=subject_1).exists())
        self.assertEqual(SerieEquipItem.objects.filter(team_subject=subject_1).count(), 1)

    def test_series_preview_updates_and_exports_use_team_unit_contract(self):
        equip2, _members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        subject_1, _meta_1 = self._team_subject()
        subject_2, _meta_2 = self._team_subject(equip2)
        serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie A",
        )
        SerieEquipItem.objects.create(serie=serie, team_subject=subject_1, ordre=1)

        preview_res = self._post_json(
            "inscripcions_series_equips_preview",
            {
                "comp_aparell_id": self.comp_app.id,
                "action": "assign",
                "serie_id": serie.id,
                "selected_ids": [subject_2.id],
            },
        )
        self.assertEqual(preview_res.status_code, 200)
        self.assertTrue(preview_res.json()["preview"]["can_run"])
        self.assertTrue(preview_res.json()["preview"]["plan_signature"])

        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=subject_1,
            exercici=1,
            inputs={"SYNC": 5},
            outputs={},
            total=5,
        )
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=subject_2,
            exercici=1,
            inputs={"SYNC": 7},
            outputs={},
            total=7,
        )

        scoring_res = self.client.get(
            reverse("scoring_updates", kwargs={"pk": self.comp.id}),
            {
                "since": (timezone.now() - timedelta(minutes=10)).isoformat(),
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie.id,
            },
        )
        self.assertEqual(scoring_res.status_code, 200)
        scoring_updates = scoring_res.json()["updates"]
        self.assertEqual({row["subject_kind"] for row in scoring_updates}, {"team_unit"})
        self.assertEqual({row["subject_id"] for row in scoring_updates}, {subject_1.id})
        self.assertEqual(scoring_updates[0]["serie_id"], serie.id)
        self.assertEqual(scoring_updates[0]["series_state"], "assigned")

        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Judge Series",
            permissions=[{"field_code": "SYNC", "runtime_field_code": "SYNC", "scope": "shared", "judge_index": 1}],
            is_active=True,
        )
        judge_res = self.client.get(
            reverse("judge_updates", kwargs={"token": token.id}),
            {
                "since": (timezone.now() - timedelta(minutes=10)).isoformat(),
                "exercici": 1,
                "serie_id": serie.id,
            },
        )
        self.assertEqual(judge_res.status_code, 200)
        judge_updates = judge_res.json()["updates"]
        self.assertEqual({row["subject_kind"] for row in judge_updates}, {"team_unit"})
        self.assertEqual({row["subject_id"] for row in judge_updates}, {subject_1.id})
        self.assertEqual(judge_updates[0]["serie_id"], serie.id)
        self.assertEqual(judge_updates[0]["series_state"], "assigned")

        start_list_res = self.client.get(
            reverse("inscripcions_series_equips_start_list_export", kwargs={"pk": self.comp.id}),
            {"comp_aparell_id": self.comp_app.id},
        )
        self.assertEqual(start_list_res.status_code, 200)
        self.assertIn("series_start_list", start_list_res["Content-Disposition"])

        work_sheet_res = self.client.get(
            reverse("inscripcions_series_equips_work_sheet_export", kwargs={"pk": self.comp.id}),
            {"comp_aparell_id": self.comp_app.id, "serie_id": serie.id},
        )
        self.assertEqual(work_sheet_res.status_code, 200)
        self.assertIn("serie_work_sheet", work_sheet_res["Content-Disposition"])
        workbook = load_workbook(BytesIO(work_sheet_res.content))
        ws = workbook.active
        values = [row[1] for row in ws.iter_rows(min_row=4, values_only=True) if row and row[1]]
        self.assertEqual(values[:1], [subject_1.label])

    def test_series_assign_requires_preview_signature(self):
        subject_1, _meta_1 = self._team_subject()
        equip2, _members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        subject_2, _meta_2 = self._team_subject(equip2)
        serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie A",
        )
        SerieEquipItem.objects.create(serie=serie, team_subject=subject_1, ordre=1)

        missing_preview_res = self._post_json(
            "inscripcions_series_equips_assign",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie.id,
                "selected_ids": [subject_2.id],
            },
        )
        self.assertEqual(missing_preview_res.status_code, 400)
        self.assertContains(missing_preview_res, "preview required", status_code=400)

        preview_res = self._post_json(
            "inscripcions_series_equips_preview",
            {
                "comp_aparell_id": self.comp_app.id,
                "action": "assign",
                "serie_id": serie.id,
                "selected_ids": [subject_2.id],
            },
        )
        self.assertEqual(preview_res.status_code, 200)
        plan_signature = preview_res.json()["preview"]["plan_signature"]

        assign_res = self._post_json(
            "inscripcions_series_equips_assign",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie.id,
                "selected_ids": [subject_2.id],
                "plan_signature": plan_signature,
            },
        )
        self.assertEqual(assign_res.status_code, 200)
        self.assertTrue(SerieEquipItem.objects.filter(serie=serie, team_subject=subject_2).exists())

    def test_scoring_and_judge_updates_include_invalid_team_subjects(self):
        invalid_team, _members = self._create_team_with_members("Parella incompleta", ["Berta"], start_order=20)
        invalid_team_subject, _invalid_meta = self._team_subject(invalid_team)
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=invalid_team_subject,
            exercici=1,
            inputs={"SYNC": 4.0},
            outputs={},
            total=4.0,
        )

        scoring_res = self.client.get(
            reverse("scoring_updates", kwargs={"pk": self.comp.id}),
            {
                "since": (timezone.now() - timedelta(minutes=10)).isoformat(),
                "comp_aparell_id": self.comp_app.id,
            },
        )
        self.assertEqual(scoring_res.status_code, 200)
        scoring_updates = {int(row["subject_id"]): row for row in scoring_res.json()["updates"]}
        self.assertEqual(scoring_updates[invalid_team_subject.id]["series_state"], "invalid")
        self.assertIsNone(scoring_updates[invalid_team_subject.id]["serie_id"])

        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Judge Invalid Bucket",
            permissions=[{"field_code": "SYNC", "runtime_field_code": "SYNC", "scope": "shared", "judge_index": 1}],
            is_active=True,
        )
        judge_res = self.client.get(
            reverse("judge_updates", kwargs={"token": token.id}),
            {
                "since": (timezone.now() - timedelta(minutes=10)).isoformat(),
                "exercici": 1,
            },
        )
        self.assertEqual(judge_res.status_code, 200)
        judge_updates = {int(row["subject_id"]): row for row in judge_res.json()["updates"]}
        self.assertEqual(judge_updates[invalid_team_subject.id]["series_state"], "invalid")
        self.assertIsNone(judge_updates[invalid_team_subject.id]["serie_id"])

    def test_scoring_updates_without_comp_aparell_id_include_individual_and_team_entries(self):
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={
                "meta": {"subject_mode": "team_context", "expected_team_size": 2},
                "fields": [
                    {"code": "SYNC", "label": "Sync", "type": "number", "scope": "shared"},
                    {"code": "E", "label": "Exec", "type": "number", "scope": "member"},
                ],
                "computed": [
                    {"code": "TOTAL", "label": "Total", "formula": "SYNC"},
                ],
            },
        )
        indiv_app = self._create_aparell("IND-UPD", "Individual updates")
        indiv_comp_app = self._create_comp_aparell(self.comp, indiv_app, ordre=2)
        ScoringSchema.objects.create(
            aparell=indiv_app,
            schema={
                "fields": [{"code": "N", "label": "Nota", "type": "number"}],
                "computed": [{"code": "TOTAL", "label": "Total", "formula": "N"}],
            },
        )

        team_subject, _team_meta = self._team_subject()
        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={
                "SYNC": 5.0,
                "E": {
                    str(self.ins1.id): 0.2,
                    str(self.ins2.id): 0.3,
                },
            },
            outputs={"TOTAL": 5.0},
            total=5.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=indiv_comp_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={"N": 7.5},
            outputs={"TOTAL": 7.5},
            total=7.5,
        )

        res = self.client.get(
            reverse("scoring_updates", kwargs={"pk": self.comp.id}),
            {"since": (timezone.now() - timedelta(minutes=10)).isoformat()},
        )

        self.assertEqual(res.status_code, 200)
        updates = res.json()["updates"]
        by_kind = {(row["subject_kind"], int(row["subject_id"])): row for row in updates}
        self.assertIn(("inscripcio", self.ins1.id), by_kind)
        self.assertIn(("team_unit", team_subject.id), by_kind)
        self.assertEqual(by_kind[("inscripcio", self.ins1.id)]["comp_aparell_id"], indiv_comp_app.id)
        self.assertEqual(by_kind[("inscripcio", self.ins1.id)]["inputs"], {"N": 7.5})
        self.assertEqual(
            by_kind[("team_unit", team_subject.id)]["inputs"],
            {
                "SYNC": 5.0,
                "E": {
                    str(self.ins1.id): 0.2,
                    str(self.ins2.id): 0.3,
                },
            },
        )
        self.assertEqual(by_kind[("team_unit", team_subject.id)]["comp_aparell_id"], self.comp_app.id)
        self.assertEqual(by_kind[("team_unit", team_subject.id)]["series_state"], "unassigned")

    def test_judge_updates_team_unit_use_after_id_for_same_timestamp(self):
        equip2, _members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        equip3, _members3 = self._create_team_with_members("Parella 3", ["Jana", "Paula"], start_order=40)
        subject_1, _meta_1 = self._team_subject()
        subject_2, _meta_2 = self._team_subject(equip2)
        subject_3, _meta_3 = self._team_subject(equip3)
        token = JudgeDeviceToken.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            label="Team Judge Cursor",
            permissions=[{"field_code": "SYNC", "runtime_field_code": "SYNC", "scope": "shared", "judge_index": 1}],
            is_active=True,
        )
        base_time = timezone.now()
        e1 = TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=subject_1,
            exercici=1,
            inputs={"SYNC": 5},
            outputs={},
            total=5,
        )
        e2 = TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=subject_2,
            exercici=1,
            inputs={"SYNC": 6},
            outputs={},
            total=6,
        )
        e3 = TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=subject_3,
            exercici=1,
            inputs={"SYNC": 7},
            outputs={},
            total=7,
        )
        TeamScoreEntry.objects.filter(pk__in=[e1.id, e2.id, e3.id]).update(updated_at=base_time)

        url = reverse("judge_updates", kwargs={"token": token.id})
        with patch("competicions_trampoli.views.judge.updates.JUDGE_UPDATES_LIMIT", 2):
            first_res = self.client.get(url, {"since": (base_time - timedelta(seconds=1)).isoformat(), "exercici": 1})
            self.assertEqual(first_res.status_code, 200)
            first_body = first_res.json()
            self.assertTrue(first_body.get("has_more"))
            self.assertEqual(
                [int(row["subject_id"]) for row in first_body.get("updates", [])],
                [subject_1.id, subject_2.id],
            )

            second_res = self.client.get(
                url,
                {
                    "since": first_body.get("next_since"),
                    "after_id": first_body.get("next_after_id"),
                    "exercici": 1,
                },
            )

        self.assertEqual(second_res.status_code, 200)
        self.assertEqual(
            [int(row["subject_id"]) for row in second_res.json().get("updates", [])],
            [subject_3.id],
        )

    def test_scoring_updates_combined_cursor_orders_individual_before_team_for_same_timestamp(self):
        indiv_app = self._create_aparell("IND-CURSOR", "Individual cursor")
        indiv_comp_app = self._create_comp_aparell(self.comp, indiv_app, ordre=2)
        equip2, _members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        subject_1, _meta_1 = self._team_subject()
        subject_2, _meta_2 = self._team_subject(equip2)
        base_time = timezone.now()
        score_entry = ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=indiv_comp_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={"N": 7.5},
            outputs={"TOTAL": 7.5},
            total=7.5,
        )
        team_entry_1 = TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=subject_1,
            exercici=1,
            inputs={"SYNC": 5},
            outputs={},
            total=5,
        )
        team_entry_2 = TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=subject_2,
            exercici=1,
            inputs={"SYNC": 6},
            outputs={},
            total=6,
        )
        ScoreEntry.objects.filter(pk=score_entry.id).update(updated_at=base_time)
        TeamScoreEntry.objects.filter(pk__in=[team_entry_1.id, team_entry_2.id]).update(updated_at=base_time)

        url = reverse("scoring_updates", kwargs={"pk": self.comp.id})
        with patch("competicions_trampoli.views.scoring.updates.SCORING_UPDATES_LIMIT", 2):
            first_res = self.client.get(url, {"since": (base_time - timedelta(seconds=1)).isoformat()})
            self.assertEqual(first_res.status_code, 200)
            first_body = first_res.json()
            self.assertTrue(first_body.get("has_more"))
            self.assertEqual(
                [(row["subject_kind"], int(row["subject_id"])) for row in first_body.get("updates", [])],
                [("inscripcio", self.ins1.id), ("team_unit", subject_1.id)],
            )
            self.assertEqual(first_body.get("next_after_id"), f"team:{team_entry_1.id}")

            second_res = self.client.get(
                url,
                {
                    "since": first_body.get("next_since"),
                    "after_id": first_body.get("next_after_id"),
                },
            )

        self.assertEqual(second_res.status_code, 200)
        self.assertEqual(
            [(row["subject_kind"], int(row["subject_id"])) for row in second_res.json().get("updates", [])],
            [("team_unit", subject_2.id)],
        )

    def test_scoring_updates_group_filter_keeps_team_rows_and_filters_individual_rows(self):
        indiv_app = self._create_aparell("IND-GROUP", "Individual group")
        indiv_comp_app = self._create_comp_aparell(self.comp, indiv_app, ordre=2)
        team_subject, _team_meta = self._team_subject()
        other_group_ins = self._create_inscripcio(self.comp, "Berta grup 2", ordre=50, grup=2)

        TeamScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            team_subject=team_subject,
            exercici=1,
            inputs={"SYNC": 5.0},
            outputs={},
            total=5.0,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=indiv_comp_app,
            inscripcio=self.ins1,
            exercici=1,
            inputs={"N": 7.5},
            outputs={"TOTAL": 7.5},
            total=7.5,
        )
        ScoreEntry.objects.create(
            competicio=self.comp,
            comp_aparell=indiv_comp_app,
            inscripcio=other_group_ins,
            exercici=1,
            inputs={"N": 6.5},
            outputs={"TOTAL": 6.5},
            total=6.5,
        )

        res = self.client.get(
            reverse("scoring_updates", kwargs={"pk": self.comp.id}),
            {
                "since": (timezone.now() - timedelta(minutes=10)).isoformat(),
                "group": self.ins1.grup_competicio_id,
            },
        )

        self.assertEqual(res.status_code, 200)
        updates = {(row["subject_kind"], int(row["subject_id"])): row for row in res.json()["updates"]}
        self.assertIn(("inscripcio", self.ins1.id), updates)
        self.assertNotIn(("inscripcio", other_group_ins.id), updates)
        self.assertIn(("team_unit", team_subject.id), updates)

    def test_series_delete_blocks_programmed_empty_serie(self):
        serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie Programada",
        )
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici=timezone.datetime(2025, 1, 1, 9, 0).time(),
            hora_fi=timezone.datetime(2025, 1, 1, 10, 0).time(),
            ordre=1,
            titol="Franja 1",
        )
        estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=1,
        )
        save_res = self._post_json(
            "rotacions_save",
            {
                "cells": [
                    {
                        "franja": franja.id,
                        "estacio": estacio.id,
                        "items": [f"s:{serie.id}"],
                    }
                ],
            },
        )
        self.assertEqual(save_res.status_code, 200)

        preview_res = self._post_json(
            "inscripcions_series_equips_preview",
            {
                "comp_aparell_id": self.comp_app.id,
                "action": "delete",
                "serie_id": serie.id,
            },
        )
        self.assertEqual(preview_res.status_code, 200)
        preview = preview_res.json()["preview"]
        self.assertFalse(preview["can_run"])
        self.assertEqual(preview["reason"], "serie_programmed")

        delete_res = self._post_json(
            "inscripcions_series_equips_delete",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie.id,
                "plan_signature": preview["plan_signature"],
            },
        )
        self.assertEqual(delete_res.status_code, 400)
        self.assertContains(delete_res, "serie programmed", status_code=400)
        self.assertTrue(SerieEquip.objects.get(pk=serie.id).actiu)

    def test_series_delete_empty_deactivates_only_unprogrammed_empty_series(self):
        equip2, _members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        subject_1, _meta_1 = self._team_subject()
        _subject_2, _meta_2 = self._team_subject(equip2)

        non_empty = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie amb contingut",
        )
        SerieEquipItem.objects.create(serie=non_empty, team_subject=subject_1, ordre=1)

        empty = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=2,
            nom="Serie buida",
        )
        programmed_empty = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=3,
            nom="Serie programada buida",
        )

        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici=timezone.datetime(2025, 1, 1, 9, 0).time(),
            hora_fi=timezone.datetime(2025, 1, 1, 10, 0).time(),
            ordre=1,
            titol="Franja 1",
        )
        estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=1,
        )
        assignacio = RotacioAssignacio.objects.create(
            competicio=self.comp,
            franja=franja,
            estacio=estacio,
        )
        RotacioAssignacioSerieEquip.objects.create(assignacio=assignacio, serie=programmed_empty, ordre=1)

        response = self._post_json(
            "inscripcions_series_equips_delete_empty",
            {"comp_aparell_id": self.comp_app.id},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data.get("ok"))
        self.assertEqual(data.get("deleted"), 1)
        self.assertIn(programmed_empty.id, data.get("skipped_programmed_ids", []))
        self.assertIn(non_empty.id, data.get("skipped_not_empty_ids", []))

        empty.refresh_from_db()
        programmed_empty.refresh_from_db()
        non_empty.refresh_from_db()
        self.assertFalse(empty.actiu)
        self.assertTrue(programmed_empty.actiu)
        self.assertTrue(non_empty.actiu)

    def test_series_start_list_export_includes_unassigned_bucket_and_persistent_order(self):
        equip2, _members2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=30)
        equip3, _members3 = self._create_team_with_members("Parella 3", ["Jana", "Paula"], start_order=40)
        subject_1, meta_1 = self._team_subject()
        subject_2, meta_2 = self._team_subject(equip2)
        subject_3, meta_3 = self._team_subject(equip3)

        create_preview_res = self._post_json(
            "inscripcions_series_equips_preview",
            {
                "comp_aparell_id": self.comp_app.id,
                "action": "create",
                "name": "Serie Export",
                "selected_ids": [subject_1.id, subject_2.id],
            },
        )
        self.assertEqual(create_preview_res.status_code, 200)
        create_plan_signature = create_preview_res.json()["preview"]["plan_signature"]
        create_res = self._post_json(
            "inscripcions_series_equips_create",
            {
                "comp_aparell_id": self.comp_app.id,
                "name": "Serie Export",
                "selected_ids": [subject_1.id, subject_2.id],
                "plan_signature": create_plan_signature,
            },
        )
        self.assertEqual(create_res.status_code, 200)
        serie = SerieEquip.objects.get(pk=create_res.json()["serie_id"])

        reorder_res = self._post_json(
            "inscripcions_series_equips_reorder",
            {
                "comp_aparell_id": self.comp_app.id,
                "serie_id": serie.id,
                "subject_ids": [subject_2.id, subject_1.id],
            },
        )
        self.assertEqual(reorder_res.status_code, 200)

        start_list_res = self.client.get(
            reverse("inscripcions_series_equips_start_list_export", kwargs={"pk": self.comp.id}),
            {"comp_aparell_id": self.comp_app.id},
        )
        self.assertEqual(start_list_res.status_code, 200)
        workbook = load_workbook(BytesIO(start_list_res.content))
        ws = workbook.active
        rows = list(ws.iter_rows(values_only=True))

        serie_title_idx = next(idx for idx, row in enumerate(rows) if row and row[0] == "Serie Export")
        self.assertEqual(rows[serie_title_idx + 2][1], meta_2["name"])
        self.assertEqual(rows[serie_title_idx + 3][1], meta_1["name"])

        unassigned_title_idx = next(idx for idx, row in enumerate(rows) if row and row[0] == "Sense serie")
        self.assertEqual(rows[unassigned_title_idx + 2][1], meta_3["name"])

    def test_rotacions_save_ignores_mixed_program_keys_by_station_mode(self):
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici=timezone.datetime(2025, 1, 1, 9, 0).time(),
            hora_fi=timezone.datetime(2025, 1, 1, 10, 0).time(),
            ordre=1,
            titol="Franja 1",
        )
        team_estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=1,
        )
        individual_app = self._create_aparell("IND2", "Individual 2")
        individual_comp_app = self._create_comp_aparell(self.comp, individual_app, ordre=3)
        individual_estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=individual_comp_app,
            ordre=2,
        )
        serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie Mixta",
        )
        group_id = int(self.ins1.grup_competicio_id)

        save_res = self._post_json(
            "rotacions_save",
            {
                "cells": [
                    {
                        "franja": franja.id,
                        "estacio": team_estacio.id,
                        "items": [f"g:{group_id}", f"s:{serie.id}"],
                    },
                    {
                        "franja": franja.id,
                        "estacio": individual_estacio.id,
                        "items": [f"s:{serie.id}", f"g:{group_id}"],
                    },
                ],
            },
        )
        self.assertEqual(save_res.status_code, 200)

        team_assignacio = RotacioAssignacio.objects.get(competicio=self.comp, franja=franja, estacio=team_estacio)
        individual_assignacio = RotacioAssignacio.objects.get(competicio=self.comp, franja=franja, estacio=individual_estacio)

        self.assertEqual(list(team_assignacio.serie_links.values_list("serie_id", flat=True)), [serie.id])
        self.assertEqual(list(team_assignacio.grup_links.values_list("grup_id", flat=True)), [])
        self.assertEqual(list(individual_assignacio.serie_links.values_list("serie_id", flat=True)), [])
        self.assertEqual(list(individual_assignacio.grup_links.values_list("grup_id", flat=True)), [group_id])

    def test_rotacions_save_filters_team_series_to_matching_comp_aparell(self):
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici=timezone.datetime(2025, 1, 1, 9, 0).time(),
            hora_fi=timezone.datetime(2025, 1, 1, 10, 0).time(),
            ordre=1,
            titol="Franja 1",
        )
        estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=1,
        )
        other_app = self._create_aparell("SYNC2", "Sincronitzat 2")
        other_app.competition_unit = Aparell.CompetitionUnit.TEAM
        other_app.save(update_fields=["competition_unit"])
        other_comp_app = self._create_comp_aparell(self.comp, other_app, ordre=2)
        other_ctx = EquipContext.objects.create(competicio=self.comp, code="trios", nom="Trios")
        CompeticioAparellEquipContextSource.objects.create(
            competicio=self.comp,
            comp_aparell=other_comp_app,
            context=other_ctx,
        )
        valid_serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie OK",
        )
        foreign_serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=other_comp_app,
            display_num=1,
            nom="Serie Fora",
        )

        save_res = self._post_json(
            "rotacions_save",
            {
                "cells": [
                    {
                        "franja": franja.id,
                        "estacio": estacio.id,
                        "items": [f"s:{foreign_serie.id}", f"s:{valid_serie.id}"],
                    }
                ],
            },
        )
        self.assertEqual(save_res.status_code, 200)
        assignacio = RotacioAssignacio.objects.get(competicio=self.comp, franja=franja, estacio=estacio)
        self.assertEqual(list(assignacio.serie_links.values_list("serie_id", flat=True)), [valid_serie.id])

    def test_rotacions_save_keeps_team_and_individual_station_payloads_separated(self):
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici=timezone.datetime(2025, 1, 1, 9, 0).time(),
            hora_fi=timezone.datetime(2025, 1, 1, 10, 0).time(),
            ordre=1,
            titol="Franja 1",
        )
        team_estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=1,
        )
        team_serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie Team",
        )
        individual_app = self._create_aparell("TRA", "Trampolí")
        individual_comp_app = self._create_comp_aparell(self.comp, individual_app, ordre=2)
        individual_estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=individual_comp_app,
            ordre=2,
        )
        group = ensure_group_for_display_num(self.comp, 1, name="Grup 1")

        save_res = self._post_json(
            "rotacions_save",
            {
                "cells": [
                    {
                        "franja": franja.id,
                        "estacio": team_estacio.id,
                        "items": [f"g:{group.id}", f"s:{team_serie.id}"],
                    },
                    {
                        "franja": franja.id,
                        "estacio": individual_estacio.id,
                        "items": [f"g:{group.id}", f"s:{team_serie.id}"],
                    },
                ],
            },
        )
        self.assertEqual(save_res.status_code, 200)

        team_assignacio = RotacioAssignacio.objects.get(competicio=self.comp, franja=franja, estacio=team_estacio)
        individual_assignacio = RotacioAssignacio.objects.get(competicio=self.comp, franja=franja, estacio=individual_estacio)
        self.assertEqual(list(team_assignacio.serie_links.values_list("serie_id", flat=True)), [team_serie.id])
        self.assertEqual(list(team_assignacio.grup_links.values_list("grup_id", flat=True)), [])
        self.assertEqual(list(individual_assignacio.serie_links.values_list("serie_id", flat=True)), [])
        self.assertEqual(list(individual_assignacio.grup_links.values_list("grup_id", flat=True)), [group.id])

    def test_rotacions_save_rejects_duplicate_group_within_same_franja(self):
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici=timezone.datetime(2025, 1, 1, 9, 0).time(),
            hora_fi=timezone.datetime(2025, 1, 1, 10, 0).time(),
            ordre=1,
            titol="Franja 1",
        )
        individual_app = self._create_aparell("TRA_DUP", "Tramp Duplicat")
        comp_app_a = self._create_comp_aparell(self.comp, individual_app, ordre=2)
        comp_app_b = self._create_comp_aparell(self.comp, self._create_aparell("TRA_DUP_2", "Tramp Duplicat 2"), ordre=3)
        estacio_a = RotacioEstacio.objects.create(competicio=self.comp, tipus="aparell", comp_aparell=comp_app_a, ordre=1)
        estacio_b = RotacioEstacio.objects.create(competicio=self.comp, tipus="aparell", comp_aparell=comp_app_b, ordre=2)
        group_id = int(self.ins1.grup_competicio_id)

        save_res = self._post_json(
            "rotacions_save",
            {
                "cells": [
                    {"franja": franja.id, "estacio": estacio_a.id, "items": [f"g:{group_id}"]},
                    {"franja": franja.id, "estacio": estacio_b.id, "items": [f"g:{group_id}"]},
                ],
            },
        )
        self.assertEqual(save_res.status_code, 400)
        self.assertTrue(any("mateixa franja" in err for err in save_res.json().get("errors", [])))

    def test_rotacions_extrapolar_preserves_team_series_links(self):
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici=timezone.datetime(2025, 1, 1, 9, 0).time(),
            hora_fi=timezone.datetime(2025, 1, 1, 10, 0).time(),
            ordre=1,
            titol="Franja 1",
        )
        estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=1,
        )
        serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie A",
        )
        save_res = self._post_json(
            "rotacions_save",
            {
                "cells": [
                    {
                        "franja": franja.id,
                        "estacio": estacio.id,
                        "items": [f"s:{serie.id}"],
                    }
                ],
            },
        )
        self.assertEqual(save_res.status_code, 200)

        extrapolar_res = self.client.post(
            reverse("rotacions_extrapolar", kwargs={"pk": self.comp.id, "franja_id": franja.id}),
            data=json.dumps({"count": 1}),
            content_type="application/json",
        )
        self.assertEqual(extrapolar_res.status_code, 200)
        new_franja = RotacioFranja.objects.exclude(pk=franja.id).get(competicio=self.comp)
        new_assignacio = RotacioAssignacio.objects.get(competicio=self.comp, franja=new_franja, estacio=estacio)
        self.assertEqual(list(new_assignacio.serie_links.values_list("serie_id", flat=True)), [serie.id])

    def test_rotacions_save_ignores_group_keys_for_team_station(self):
        franja = RotacioFranja.objects.create(
            competicio=self.comp,
            hora_inici=timezone.datetime(2025, 1, 1, 9, 0).time(),
            hora_fi=timezone.datetime(2025, 1, 1, 10, 0).time(),
            ordre=1,
            titol="Franja 1",
        )
        estacio = RotacioEstacio.objects.create(
            competicio=self.comp,
            tipus="aparell",
            comp_aparell=self.comp_app,
            ordre=1,
        )
        valid_serie = SerieEquip.objects.create(
            competicio=self.comp,
            comp_aparell=self.comp_app,
            display_num=1,
            nom="Serie OK",
        )

        save_res = self._post_json(
            "rotacions_save",
            {
                "cells": [
                    {
                        "franja": franja.id,
                        "estacio": estacio.id,
                        "items": ["g:1", f"s:{valid_serie.id}"],
                    }
                ],
            },
        )
        self.assertEqual(save_res.status_code, 200)
        assignacio = RotacioAssignacio.objects.get(competicio=self.comp, franja=franja, estacio=estacio)
        self.assertEqual(list(assignacio.serie_links.values_list("serie_id", flat=True)), [valid_serie.id])
        self.assertEqual(assignacio.grup_links.count(), 0)

    def test_scoreable_codes_filter_team_apps_by_tipus_and_context(self):
        individual_app = self._create_aparell("IND", "Individual")
        individual_comp_app = self._create_comp_aparell(self.comp, individual_app, ordre=2)
        ScoringSchema.objects.create(
            aparell=individual_app,
            schema={"fields": [{"code": "TOTAL", "type": "number"}], "computed": []},
        )
        ScoringSchema.objects.create(
            aparell=self.app,
            schema={"fields": [{"code": "TOTAL", "type": "number"}], "computed": []},
        )
        individual_scoreables = _scoreable_codes_by_app_id(self.comp, tipus="individual")
        self.assertIn(individual_comp_app.id, individual_scoreables)
        self.assertNotIn(self.comp_app.id, individual_scoreables)

        team_scoreables = _scoreable_codes_by_app_id(
            self.comp,
            tipus="equips",
            assignment_context_code="altre",
        )
        self.assertNotIn(self.comp_app.id, team_scoreables)

    def test_compute_classificacio_derived_team_pool_selects_best_rows_per_team(self):
        individual_app = self._create_aparell("TR", "Tramp")
        individual_comp_app = self._create_comp_aparell(self.comp, individual_app, ordre=2)
        equip_2, members_2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)

        for inscripcio, exercici, total in (
            (self.ins1, 1, 9.0),
            (self.ins1, 2, 8.0),
            (self.ins2, 1, 7.0),
            (self.ins2, 2, 6.0),
            (members_2[0], 1, 8.5),
            (members_2[0], 2, 7.5),
            (members_2[1], 1, 8.0),
            (members_2[1], 2, 6.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=individual_comp_app,
                inscripcio=inscripcio,
                exercici=exercici,
                inputs={},
                outputs={},
                total=total,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived team pool main",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [individual_comp_app.id]},
                    "camps_per_aparell": {str(individual_comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 2, "max_per_participant": 1},
                    "exercise_selection_scope": "team_pool",
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertEqual([row["participant"] for row in rows[:2]], ["Parella 2", "Parella 1"])
        scores = {row["participant"]: row["score"] for row in rows}
        self.assertEqual(scores["Parella 1"], 16.0)
        self.assertEqual(scores["Parella 2"], 16.5)

    def test_compute_classificacio_derived_team_pool_global_pool_respects_member_cap(self):
        app_a = self._create_aparell("TRA", "Tramp A")
        app_b = self._create_aparell("TRB", "Tramp B")
        comp_app_a = self._create_comp_aparell(self.comp, app_a, ordre=2)
        comp_app_b = self._create_comp_aparell(self.comp, app_b, ordre=3)
        equip_2, members_2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)

        for comp_aparell, inscripcio, total in (
            (comp_app_a, self.ins1, 9.0),
            (comp_app_b, self.ins1, 8.0),
            (comp_app_a, self.ins2, 7.0),
            (comp_app_b, self.ins2, 6.0),
            (comp_app_a, members_2[0], 7.5),
            (comp_app_b, members_2[0], 7.0),
            (comp_app_a, members_2[1], 6.0),
            (comp_app_b, members_2[1], 5.5),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=comp_aparell,
                inscripcio=inscripcio,
                exercici=1,
                inputs={},
                outputs={},
                total=total,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived team pool global",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [comp_app_a.id, comp_app_b.id]},
                    "camps_per_aparell": {
                        str(comp_app_a.id): ["total"],
                        str(comp_app_b.id): ["total"],
                    },
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 3, "max_per_participant": 2},
                    "exercise_selection_scope": "team_pool",
                    "mode_seleccio_exercicis": "global_pool",
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertEqual([row["participant"] for row in rows[:2]], ["Parella 1", "Parella 2"])
        scores = {row["participant"]: row["score"] for row in rows}
        self.assertEqual(scores["Parella 1"], 24.0)
        self.assertEqual(scores["Parella 2"], 20.5)

    def test_compute_classificacio_derived_team_pool_tie_break_reuses_main_selected_rows(self):
        individual_app = self._create_aparell("TRT", "Tramp tie")
        individual_comp_app = self._create_comp_aparell(self.comp, individual_app, ordre=2)
        individual_comp_app.nombre_exercicis = 2
        individual_comp_app.save(update_fields=["nombre_exercicis"])
        equip_2, members_2 = self._create_team_with_members("Parella 2", ["Nora", "Marta"], start_order=20)

        for inscripcio, exercici, total, d_value in (
            (self.ins1, 1, 10.0, 1.0),
            (self.ins1, 2, 1.0, 100.0),
            (self.ins2, 1, 7.0, 7.0),
            (self.ins2, 2, 0.0, 0.0),
            (members_2[0], 1, 9.0, 5.0),
            (members_2[0], 2, 2.0, 6.0),
            (members_2[1], 1, 8.0, 4.0),
            (members_2[1], 2, 1.0, 5.0),
        ):
            ScoreEntry.objects.create(
                competicio=self.comp,
                comp_aparell=individual_comp_app,
                inscripcio=inscripcio,
                exercici=exercici,
                inputs={"D": d_value},
                outputs={},
                total=total,
            )

        cfg = ClassificacioConfig.objects.create(
            competicio=self.comp,
            nom="Derived team pool tie fixed rows",
            activa=True,
            ordre=1,
            tipus="equips",
            schema={
                "puntuacio": {
                    "aparells": {"mode": "seleccionar", "ids": [individual_comp_app.id]},
                    "camps_per_aparell": {str(individual_comp_app.id): ["total"]},
                    "agregacio_camps": "sum",
                    "exercicis": {"mode": "millor_n", "best_n": 2, "max_per_participant": 1},
                    "exercise_selection_scope": "team_pool",
                    "agregacio_exercicis": "sum",
                    "agregacio_aparells": "sum",
                    "ordre": "desc",
                },
                "desempat": [
                    {
                        "camps": ["D"],
                        "ordre": "desc",
                        "exercise_selection_scope": "team_pool",
                        "scope": {
                            "aparells": {"mode": "seleccionar", "ids": [individual_comp_app.id]},
                        },
                        "agregacio_camps": "sum",
                        "agregacio_exercicis": "sum",
                        "agregacio_aparells": "sum",
                    }
                ],
                "equips": {
                    "context_code": "parelles",
                    "team_mode": "derived_from_individual",
                    "incloure_sense_equip": False,
                },
            },
        )

        rows = compute_classificacio(self.comp, cfg).get("global", [])

        self.assertEqual([row["participant"] for row in rows[:2]], ["Parella 2", "Parella 1"])
        self.assertEqual(rows[0]["score"], 17.0)
        self.assertEqual(rows[1]["score"], 17.0)

    def test_classificacio_save_persists_derived_exercise_selection_scope_default(self):
        individual_app = self._create_aparell("TRS", "Tramp save")
        individual_comp_app = self._create_comp_aparell(self.comp, individual_app, ordre=2)
        payload = self._classificacio_payload(
            tipus="equips",
            app_ids=[individual_comp_app.id],
            team_mode="derived_from_individual",
        )

        response = self._post_json("classificacio_save", payload)

        self.assertEqual(response.status_code, 200)
        cfg = ClassificacioConfig.objects.get(pk=response.json()["id"])
        self.assertEqual(
            (cfg.schema.get("puntuacio") or {}).get("exercise_selection_scope"),
            "per_member",
        )

    def test_classificacio_validation_rejects_exercise_selection_scope_for_native_team(self):
        schema = self._native_team_schema_with_tie(
            {
                "camps": ["TOTAL"],
                "ordre": "desc",
                "exercise_selection_scope": "team_pool",
            }
        )
        schema["puntuacio"]["exercise_selection_scope"] = "team_pool"

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            schema,
            tipus="equips",
        )

        self.assertTrue(any("puntuacio.exercise_selection_scope" in err for err in errors))
        self.assertTrue(any("desempat[0].exercise_selection_scope" in err for err in errors))

    def test_classificacio_validation_rejects_team_pool_tie_reselection_fields(self):
        individual_app = self._create_aparell("TRV", "Tramp validation")
        individual_comp_app = self._create_comp_aparell(self.comp, individual_app, ordre=2)
        schema = {
            "puntuacio": {
                "aparells": {"mode": "seleccionar", "ids": [individual_comp_app.id]},
                "camps_per_aparell": {str(individual_comp_app.id): ["total"]},
                "agregacio_camps": "sum",
                "exercicis": {"mode": "millor_n", "best_n": 2, "max_per_participant": 1},
                "exercise_selection_scope": "team_pool",
                "agregacio_exercicis": "sum",
                "agregacio_aparells": "sum",
                "ordre": "desc",
            },
            "desempat": [
                {
                    "camps": ["total"],
                    "ordre": "desc",
                    "exercise_selection_scope": "team_pool",
                    "mode_seleccio_exercicis": "global_pool",
                    "exercicis_per_aparell": {
                        str(individual_comp_app.id): {"mode": "millor_n", "best_n": 1}
                    },
                    "scope": {
                        "aparells": {"mode": "seleccionar", "ids": [individual_comp_app.id]},
                        "exercicis": {"mode": "millor_n", "best_n": 1},
                        "participants": {"mode": "millor_1"},
                    },
                    "agregacio_participants": "sum",
                }
            ],
            "equips": {
                "context_code": "parelles",
                "team_mode": "derived_from_individual",
                "incloure_sense_equip": False,
            },
        }

        _schema, errors = _validate_schema_for_competicio(
            self.comp,
            schema,
            tipus="equips",
        )

        self.assertTrue(any("desempat[0].scope.exercicis" in err for err in errors))
        self.assertTrue(any("desempat[0].mode_seleccio_exercicis" in err for err in errors))
        self.assertTrue(any("desempat[0].exercicis_per_aparell" in err for err in errors))
        self.assertTrue(any("desempat[0].scope.participants" in err for err in errors))
        self.assertTrue(any("desempat[0].agregacio_participants" in err for err in errors))
