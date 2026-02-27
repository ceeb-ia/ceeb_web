import json
import os
import uuid
from django.utils.dateparse import parse_time
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from django.http import JsonResponse, HttpResponseBadRequest
from django.db import transaction
from django.shortcuts import render, get_object_or_404
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db.models import Q, Max, Min, Case, When, IntegerField, F
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from django.http import HttpResponse
from ceeb_web import models
from .models import Competicio, Inscripcio
from .models_trampoli import CompeticioAparell, InscripcioAparellExclusio
from .models_rotacions import RotacioFranja, RotacioAssignacio, RotacioEstacio
from .services.rotacions_ordering import (
    ORDER_MODE_MAINTAIN,
    ORDER_MODE_CHOICES,
    ORDER_MODE_LABELS,
    assignacio_grups,
    assignacio_grups_from_values,
    franja_index_map,
    get_rotacions_order_modes,
    normalize_positive_int_list,
    order_pairs_for_mode,
    set_rotacio_order_mode,
    unique_ordered,
)
from django.db import transaction
from datetime import date, datetime, timedelta
from django.utils.dateparse import parse_time
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def _normalize_grups(value):
    return normalize_positive_int_list(value)

def _assignacio_grups(assignacio):
    return assignacio_grups(assignacio)


ROTACIONS_EXPORT_BUILTIN_FIELDS = [
    {"code": "nom_i_cognoms", "label": "Nom i cognoms", "kind": "builtin"},
    {"code": "document", "label": "DNI/Document", "kind": "builtin"},
    {"code": "sexe", "label": "Sexe", "kind": "builtin"},
    {"code": "data_naixement", "label": "Data naixement", "kind": "builtin"},
    {"code": "entitat", "label": "Entitat", "kind": "builtin"},
    {"code": "categoria", "label": "Categoria", "kind": "builtin"},
    {"code": "subcategoria", "label": "Subcategoria", "kind": "builtin"},
    {"code": "grup", "label": "Grup", "kind": "builtin"},
    {"code": "ordre_sortida", "label": "Ordre", "kind": "builtin"},
]


def _reserved_inscripcio_codes():
    out = set()
    for f in Inscripcio._meta.concrete_fields:
        name = str(getattr(f, "name", "") or "").strip()
        attname = str(getattr(f, "attname", "") or "").strip()
        if name:
            out.add(name)
        if attname:
            out.add(attname)
    return out


def _normalize_schema_extra_code(code: str, reserved_codes):
    code = (code or "").strip()
    if not code:
        return code
    if code.startswith("excel__"):
        return code
    if code in reserved_codes:
        return f"excel__{code}"
    return code


def _rotacions_available_participant_fields(competicio):
    out = []
    seen = set()
    reserved = _reserved_inscripcio_codes()
    schema = competicio.inscripcions_schema or {}
    cols = schema.get("columns") or []
    excel_codes = set()

    if isinstance(cols, list):
        for c in cols:
            if not isinstance(c, dict):
                continue
            code = c.get("code")
            if not code:
                continue
            kind = c.get("kind") or "extra"
            if kind == "extra":
                code = _normalize_schema_extra_code(code, reserved)
            excel_codes.add(code)

    for f in ROTACIONS_EXPORT_BUILTIN_FIELDS:
        code = f["code"]
        if code in seen:
            continue
        source = "excel" if code in excel_codes else "native"
        out.append(
            {
                **f,
                "source": source,
                "ui_label": f'{f["label"]} ({"Excel" if source == "excel" else "Nativa"})',
            }
        )
        seen.add(code)

    if isinstance(cols, list):
        for c in cols:
            if not isinstance(c, dict):
                continue
            code = c.get("code")
            if not code:
                continue
            kind = c.get("kind") or "extra"
            if kind != "extra":
                continue
            code = _normalize_schema_extra_code(code, reserved)
            if code in seen:
                continue
            label = c.get("label") or code
            out.append(
                {
                    "code": code,
                    "label": label,
                    "kind": "extra",
                    "source": "excel",
                    "ui_label": f"{label} (Excel)",
                }
            )
            seen.add(code)

    return out


def _normalize_export_participant_fields(competicio, raw_fields):
    available = _rotacions_available_participant_fields(competicio)
    allowed_codes = {f["code"] for f in available}
    out = []
    seen = set()

    if isinstance(raw_fields, list):
        for raw in raw_fields:
            code = str(raw or "").strip()
            if not code or code not in allowed_codes or code in seen:
                continue
            seen.add(code)
            out.append(code)

    if out:
        return out
    if "nom_i_cognoms" in allowed_codes:
        return ["nom_i_cognoms"]
    if available:
        return [available[0]["code"]]
    return []


def _export_meta_defaults(competicio):
    data_default = ""
    if getattr(competicio, "data", None):
        try:
            data_default = competicio.data.strftime("%Y-%m-%d")
        except Exception:
            data_default = ""
    return {
        "title": getattr(competicio, "nom", "") or "",
        "venue": (getattr(competicio, "seu", "") or ""),
        "date": data_default,
        "logo_path": "",
        "participant_fields": ["nom_i_cognoms"],
    }


def _get_export_meta(competicio):
    defaults = _export_meta_defaults(competicio)
    view_cfg = competicio.inscripcions_view or {}
    raw = view_cfg.get("rotacions_export_meta") or {}
    if not isinstance(raw, dict):
        raw = {}

    out = dict(defaults)
    out["title"] = str(raw.get("title", defaults["title"]) or "").strip()
    out["venue"] = str(raw.get("venue", defaults["venue"]) or "").strip()
    date_val = str(raw.get("date", defaults["date"]) or "").strip()
    out["date"] = date_val
    out["logo_path"] = str(raw.get("logo_path", "") or "").strip()
    out["participant_fields"] = _normalize_export_participant_fields(
        competicio,
        raw.get("participant_fields", defaults["participant_fields"]),
    )
    return out


def _save_export_meta(competicio, meta):
    cfg = competicio.inscripcions_view or {}
    raw = cfg.get("rotacions_export_meta") or {}
    if not isinstance(raw, dict):
        raw = {}
    raw.update(meta or {})
    cfg["rotacions_export_meta"] = raw
    competicio.inscripcions_view = cfg
    competicio.save(update_fields=["inscripcions_view"])


def _logo_url_from_path(logo_path: str) -> str:
    logo_path = str(logo_path or "").strip().lstrip("/").replace("\\", "/")
    if not logo_path:
        return ""
    media_url = str(getattr(settings, "MEDIA_URL", "/media/") or "/media/")
    if not media_url.endswith("/"):
        media_url += "/"
    return f"{media_url}{logo_path}"


def _logo_abs_path(logo_path: str) -> str:
    rel = str(logo_path or "").strip().lstrip("/").replace("\\", os.sep)
    if not rel:
        return ""
    media_root = str(getattr(settings, "MEDIA_ROOT", "") or "")
    if not media_root:
        return ""
    return os.path.normpath(os.path.join(media_root, rel))

@require_POST
@csrf_protect
def franges_auto_create(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    hi = parse_time(payload.get("hora_inici") or "")
    hf = parse_time(payload.get("hora_fi") or "")
    interval = payload.get("interval_min", None)
    clear_existing = bool(payload.get("clear_existing", False))
    titol_base = (payload.get("titol_base") or "Franja").strip()

    try:
        interval = int(interval)
    except Exception:
        return HttpResponseBadRequest("interval_min ha de ser un enter (minuts)")

    if not hi or not hf:
        return HttpResponseBadRequest("Hora inici/fi obligatòries")
    if interval <= 0:
        return HttpResponseBadRequest("interval_min ha de ser > 0")

    # Convertim TimeField a datetime “dummy” per sumar minuts
    day = datetime(2000, 1, 1)
    start = datetime.combine(day.date(), hi)
    end = datetime.combine(day.date(), hf)

    if end <= start:
        return HttpResponseBadRequest("hora_fi ha de ser posterior a hora_inici")

    with transaction.atomic():
        if clear_existing:
            # Esborrem assignacions i franges (mantindrem estacions)
            franges = RotacioFranja.objects.filter(competicio=competicio)
            RotacioAssignacio.objects.filter(competicio=competicio, franja__in=franges).delete()
            franges.delete()

        # ordre a partir del màxim actual
        base_ord = (RotacioFranja.objects.filter(competicio=competicio).aggregate(Max("ordre"))["ordre__max"] or 0)

        to_create = []
        cur = start
        idx = 1
        while cur < end:
            nxt = cur + timedelta(minutes=interval)
            if nxt > end:
                break  # no creem una franja “mig buida” (si vols que la creï igualment, t'ho canvio)
            base_ord += 1
            to_create.append(RotacioFranja(
                competicio=competicio,
                hora_inici=cur.time(),
                hora_fi=nxt.time(),
                ordre=base_ord,
                titol=f"{titol_base} {idx}",
            ))
            cur = nxt
            idx += 1

        RotacioFranja.objects.bulk_create(to_create)

    return JsonResponse({"ok": True, "created": len(to_create)})

def _sync_estacions_aparells(competicio):
    comp_aps = list(CompeticioAparell.objects.filter(competicio=competicio, actiu=True).order_by("ordre", "id"))
    existents = set(
        RotacioEstacio.objects.filter(competicio=competicio, tipus="aparell", comp_aparell__isnull=False)
        .values_list("comp_aparell_id", flat=True)
    )
    to_create = []
    for ca in comp_aps:
        if ca.id not in existents:
            to_create.append(RotacioEstacio(
                competicio=competicio,
                tipus="aparell",
                comp_aparell=ca,
                ordre=int(getattr(ca, "ordre", 1) or 1),
                actiu=True
            ))
    if to_create:
        RotacioEstacio.objects.bulk_create(to_create)


def rotacions_planner(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    _sync_estacions_aparells(competicio)

    grups = list(
        Inscripcio.objects.filter(competicio=competicio, grup__isnull=False)
        .order_by("grup").values_list("grup", flat=True).distinct()
    )

    estacions = list(RotacioEstacio.objects.filter(competicio=competicio, actiu=True).order_by("ordre", "id"))
    franges = list(RotacioFranja.objects.filter(competicio=competicio).order_by("ordre", "id"))
    franja_modes = get_rotacions_order_modes(competicio)
    export_meta = _get_export_meta(competicio)
    export_meta["logo_url"] = _logo_url_from_path(export_meta.get("logo_path", ""))
    export_participant_fields = _rotacions_available_participant_fields(competicio)

    assigns = (
        RotacioAssignacio.objects
        .filter(competicio=competicio)
        .select_related("franja", "estacio")
    )

    grid = {}  # grid[franja_id][estacio_id] = [grups]
    for a in assigns:
        grid.setdefault(a.franja_id, {})[a.estacio_id] = _assignacio_grups(a)

    view_cfg = competicio.inscripcions_view or {}
    group_names = view_cfg.get("group_names") or {}
    if not isinstance(group_names, dict):
        group_names = {}

    grups_display = []
    group_labels_map = {}
    for g in grups:
        label = (group_names.get(str(g)) or "").strip() or f"G{g}"
        grups_display.append({"id": g, "label": label})
        group_labels_map[str(g)] = label

    ctx = {
        "competicio": competicio,
        "grups": grups,
        "grups_display": grups_display,
        "estacions": estacions,
        "franges": franges,
        "order_mode_options": [
            {"value": m, "label": ORDER_MODE_LABELS.get(m, m)}
            for m in ORDER_MODE_CHOICES
        ],
        "grid_json": json.dumps(grid, ensure_ascii=False),
        "group_labels_json": json.dumps(group_labels_map, ensure_ascii=False),
        "franja_order_modes_json": json.dumps(franja_modes, ensure_ascii=False),
        "export_meta_json": json.dumps(export_meta, ensure_ascii=False),
        "export_participant_fields_json": json.dumps(export_participant_fields, ensure_ascii=False),
        "grups_json": json.dumps(grups, ensure_ascii=False),
    }
    return render(request, "competicio/rotacions_planner.html", ctx)


@require_POST
@csrf_protect
def rotacions_save(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    # Esperem:
    # { "cells": [ {"franja":1,"estacio":2,"grups":[3,5]}, ... ] }
    # Compat legacy: {"grup": 3}
    cells = payload.get("cells", [])
    if not isinstance(cells, list):
        return HttpResponseBadRequest("Format incorrecte")

    franja_ids = set(RotacioFranja.objects.filter(competicio=competicio).values_list("id", flat=True))
    estacio_ids = set(RotacioEstacio.objects.filter(competicio=competicio).values_list("id", flat=True))

    with transaction.atomic():
        for c in cells:
            if not isinstance(c, dict):
                continue
            try:
                fr_id = int(c.get("franja"))
                es_id = int(c.get("estacio"))
            except Exception:
                continue

            if fr_id not in franja_ids or es_id not in estacio_ids:
                continue

            if "grups" in c:
                groups = _normalize_grups(c.get("grups"))
            else:
                groups = _normalize_grups(c.get("grup", None))

            RotacioAssignacio.objects.update_or_create(
                competicio=competicio,
                franja_id=fr_id,
                estacio_id=es_id,
                defaults={
                    "grups": groups,
                    "grup": groups[0] if groups else None,  # compat temporal
                },
            )

    return JsonResponse({"ok": True})


@require_POST
@csrf_protect
def franja_create(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    hi = parse_time(payload.get("hora_inici") or "")
    hf = parse_time(payload.get("hora_fi") or "")
    titol = (payload.get("titol") or "").strip()

    if not hi or not hf:
        return HttpResponseBadRequest("Hora inici/fi obligatòries")

    max_ord = (RotacioFranja.objects.filter(competicio=competicio).aggregate(Max("ordre"))["ordre__max"] or 0) + 1

    f = RotacioFranja(competicio=competicio, hora_inici=hi, hora_fi=hf, ordre=max_ord, titol=titol)
    try:
        f.full_clean()
    except Exception as e:
        return HttpResponseBadRequest(str(e))

    f.save()
    return JsonResponse({"ok": True, "id": f.id})


@require_POST
@csrf_protect
def franja_delete(request, pk, franja_id):
    competicio = get_object_or_404(Competicio, pk=pk)
    f = get_object_or_404(RotacioFranja, pk=franja_id, competicio=competicio)

    day = datetime(2000, 1, 1).date()

    with transaction.atomic():
        anchor_ord = f.ordre

        # 🔑 Tanquem el forat: enganxem la cadena a l'hora d'inici de la franja eliminada
        anchor_time = datetime.combine(day, f.hora_inici)

        # 1) Esborrem assignacions i franja
        RotacioAssignacio.objects.filter(competicio=competicio, franja=f).delete()
        f.delete()

        # 2) Recalculem les franges de sota mantenint la durada original de cadascuna
        below = list(
            RotacioFranja.objects
            .filter(competicio=competicio, ordre__gt=anchor_ord)
            .order_by("ordre", "id")
        )

        prev_end = anchor_time
        for fr in below:
            s0 = datetime.combine(day, fr.hora_inici)
            e0 = datetime.combine(day, fr.hora_fi)
            if e0 <= s0:
                raise ValueError("Hi ha una franja amb durada no positiva")

            dur = e0 - s0
            new_start = prev_end
            new_end = new_start + dur

            fr.hora_inici = new_start.time()
            fr.hora_fi = new_end.time()
            prev_end = new_end

        if below:
            RotacioFranja.objects.bulk_update(below, ["hora_inici", "hora_fi"], batch_size=200)

        # 3) Reordenem 'ordre' perquè no quedin forats (recomanat)
        rest = list(RotacioFranja.objects.filter(competicio=competicio).order_by("ordre", "id"))
        for i, fr in enumerate(rest, start=1):
            fr.ordre = i
        if rest:
            RotacioFranja.objects.bulk_update(rest, ["ordre"], batch_size=200)

    return JsonResponse({"ok": True})

@require_POST
@csrf_protect
def estacio_descans_create(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    max_ord = (RotacioEstacio.objects.filter(competicio=competicio).aggregate(Max("ordre"))["ordre__max"] or 0) + 1
    e = RotacioEstacio.objects.create(competicio=competicio, tipus="descans", ordre=max_ord, actiu=True)
    return JsonResponse({"ok": True, "id": e.id})

@require_POST
@csrf_protect
def estacio_delete(request, pk, estacio_id):
    competicio = get_object_or_404(Competicio, pk=pk)
    e = get_object_or_404(RotacioEstacio, pk=estacio_id, competicio=competicio)
    with transaction.atomic():
        RotacioAssignacio.objects.filter(competicio=competicio, estacio=e).delete()
        e.delete()  # o e.actiu=False si prefereixes no perdre històric
    return JsonResponse({"ok": True})


@require_POST
@csrf_protect
def rotacions_extrapolar(request, pk, franja_id):
    competicio = get_object_or_404(Competicio, pk=pk)
    fr_base = get_object_or_404(RotacioFranja, pk=franja_id, competicio=competicio)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    # quantes franges vols omplir a partir de la base
    count = payload.get("count", None)

    estacions = list(
        RotacioEstacio.objects.filter(competicio=competicio, actiu=True)
        .order_by("ordre", "id")
    )
    if not estacions:
        return HttpResponseBadRequest("No hi ha estacions actives.")

    if count is None:
        count = max(0, len(estacions) - 1)

    try:
        count = int(count)
    except Exception:
        return HttpResponseBadRequest("count ha de ser un enter.")
    if count <= 0:
        return JsonResponse({"ok": True, "created_franges": 0, "filled": 0})

    # 1) Llegim assignació de la franja base segons ordre d'estacions
    base_map = {}
    for a in RotacioAssignacio.objects.filter(competicio=competicio, franja=fr_base):
        base_map[a.estacio_id] = _assignacio_grups(a)

    base_groups = []
    for e in estacions:
        base_groups.append(base_map.get(e.id, []))

    # Si no hi ha cap grup definit, no té sentit extrapolar
    if all(len(gs) == 0 for gs in base_groups):
        return HttpResponseBadRequest("La franja base no té cap grup assignat.")

    # 2) Aconseguim/creem franges següents
    franges = list(RotacioFranja.objects.filter(competicio=competicio).order_by("ordre", "id"))
    idx_base = next((i for i, f in enumerate(franges) if f.id == fr_base.id), None)
    if idx_base is None:
        return HttpResponseBadRequest("Franja base no trobada al llistat.")

    # Interval per crear franges: preferim interval de la base; si no es pot, del veí anterior
    def _delta_minutes(f):
        # retorna minuts entre inici i fi, o None
        try:
            day = datetime(2000, 1, 1)
            s = datetime.combine(day.date(), f.hora_inici)
            t = datetime.combine(day.date(), f.hora_fi)
            if t <= s:
                return None
            return int((t - s).total_seconds() // 60)
        except Exception:
            return None

    interval_min = _delta_minutes(fr_base)
    if not interval_min:
        if idx_base > 0:
            interval_min = _delta_minutes(franges[idx_base - 1])
    if not interval_min:
        return HttpResponseBadRequest("No puc deduir l'interval (minuts) per crear franges noves.")

    # Ens assegurem que existeixen idx_base+1 ... idx_base+count
    created = 0
    with transaction.atomic():
        # refresquem per evitar desajustos
        franges = list(RotacioFranja.objects.filter(competicio=competicio).order_by("ordre", "id"))
        idx_base = next((i for i, f in enumerate(franges) if f.id == fr_base.id), None)

        max_ord = (RotacioFranja.objects.filter(competicio=competicio).aggregate(Max("ordre"))["ordre__max"] or 0)

        target_franges = []
        for k in range(1, count + 1):
            idx = idx_base + k
            if idx < len(franges):
                target_franges.append(franges[idx])
            else:
                # crear franja nova a continuació temporalment
                # start = fi de l'última franja existent
                last = franges[-1] if franges else fr_base
                day = datetime(2000, 1, 1).date()
                last_end = datetime.combine(day, last.hora_fi)
                new_start = last_end
                new_end = new_start + timedelta(minutes=interval_min)

                max_ord += 1
                nf = RotacioFranja.objects.create(
                    competicio=competicio,
                    hora_inici=new_start.time(),
                    hora_fi=new_end.time(),
                    ordre=max_ord,
                    titol=f"{fr_base.titol or 'Franja'} +{k}",
                )
                created += 1
                franges.append(nf)
                target_franges.append(nf)

        # 3) Omplim les franges fent shift circular
        filled_cells = 0
        n = len(estacions)

        for k, fr_t in enumerate(target_franges, start=1):
            shifted = [list(base_groups[(i - k) % n]) for i in range(n)]  # shift cap a la dreta

            for e, gs in zip(estacions, shifted):
                RotacioAssignacio.objects.update_or_create(
                    competicio=competicio,
                    franja=fr_t,
                    estacio=e,
                    defaults={
                        "grups": gs,
                        "grup": gs[0] if gs else None,  # compat temporal
                    },
                )
                filled_cells += 1

    return JsonResponse({
        "ok": True,
        "created_franges": created,
        "filled": filled_cells,
    })


@require_POST
@csrf_protect
def estacions_reorder(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    order = payload.get("order", [])
    if not isinstance(order, list) or not order:
        return HttpResponseBadRequest("order ha de ser una llista d'IDs")

    # validació: totes les estacions han de ser de la competició
    estacions = RotacioEstacio.objects.filter(competicio=competicio, id__in=order)
    found = set(estacions.values_list("id", flat=True))
    wanted = []
    for x in order:
        try:
            wanted.append(int(x))
        except Exception:
            pass

    if set(wanted) != found:
        return HttpResponseBadRequest("IDs d'estació invàlids per aquesta competició.")

    with transaction.atomic():
        for idx, estacio_id in enumerate(wanted, start=1):
            RotacioEstacio.objects.filter(competicio=competicio, id=estacio_id).update(ordre=idx)

    return JsonResponse({"ok": True})


@require_POST
@csrf_protect
def rotacions_clear_all(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    with transaction.atomic():
        RotacioAssignacio.objects.filter(competicio=competicio).delete()
        # opcional:
        RotacioFranja.objects.filter(competicio=competicio).delete()
    return JsonResponse({"ok": True})


@require_POST
@csrf_protect
def franja_insert_after(request, pk, franja_id):
    competicio = get_object_or_404(Competicio, pk=pk)
    fr_prev = get_object_or_404(RotacioFranja, pk=franja_id, competicio=competicio)

    # Durada = durada de la franja anterior
    day = datetime(2000, 1, 1).date()
    prev_start = datetime.combine(day, fr_prev.hora_inici)
    prev_end = datetime.combine(day, fr_prev.hora_fi)

    if prev_end <= prev_start:
        return HttpResponseBadRequest("La franja base té hores invàlides.")

    delta = prev_end - prev_start  # timedelta
    new_start = prev_end
    new_end = new_start + delta

    # (Opcional) títol personalitzat
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}
    titol = (payload.get("titol") or "").strip()

    with transaction.atomic():
        # 1) Apliquem un offset temporal gran per evitar col·lisions amb el unique (competicio, ordre)
        qs = RotacioFranja.objects.filter(
            competicio=competicio,
            ordre__gt=fr_prev.ordre
        )

        OFFSET = 1000  # prou gran per sortir de la zona de valors reals
        qs.update(ordre=F("ordre") + OFFSET)

        # 2) Ara ja podem deixar-ho en el valor final (+1) sense col·lisions
        qs.update(ordre=F("ordre") - OFFSET + 1)

        # 3) Crear franja nova
        f_new = RotacioFranja.objects.create(
            competicio=competicio,
            hora_inici=new_start.time(),
            hora_fi=new_end.time(),
            ordre=fr_prev.ordre + 1,
            titol=titol or "",
        )

        # 4) Desplacem horaris de totes les franges que queden per sota (ara tenen ordre > fr_prev.ordre+1)
        to_shift = list(
            RotacioFranja.objects.filter(
                competicio=competicio,
                ordre__gt=f_new.ordre
            ).order_by("ordre", "id")
        )

        for f in to_shift:
            s = datetime.combine(day, f.hora_inici) + delta
            e = datetime.combine(day, f.hora_fi) + delta
            f.hora_inici = s.time()
            f.hora_fi = e.time()

        if to_shift:
            RotacioFranja.objects.bulk_update(to_shift, ["hora_inici", "hora_fi"], batch_size=200)

    return JsonResponse({"ok": True, "id": f_new.id})



@require_POST
@csrf_protect
def franja_update_inline(request, pk, franja_id):
    competicio = get_object_or_404(Competicio, pk=pk)
    fr = get_object_or_404(RotacioFranja, pk=franja_id, competicio=competicio)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    titol = (payload.get("titol") or "").strip()
    hora_inici = payload.get("hora_inici")
    hora_fi = payload.get("hora_fi")

    if not hora_inici or not hora_fi:
        return HttpResponseBadRequest("Falten hores")

    try:
        # esperem "HH:MM" (input type="time")
        hi = datetime.strptime(hora_inici, "%H:%M").time()
        hf = datetime.strptime(hora_fi, "%H:%M").time()
    except ValueError:
        return HttpResponseBadRequest("Format d'hora incorrecte (HH:MM)")

    day = datetime(2000, 1, 1).date()
    d_hi = datetime.combine(day, hi)
    d_hf = datetime.combine(day, hf)
    if d_hf <= d_hi:
        return HttpResponseBadRequest("L'hora de fi ha de ser posterior a l'hora d'inici")

    with transaction.atomic():
        # 1) Actualitza franja editada
        fr.titol = titol
        fr.hora_inici = hi
        fr.hora_fi = hf
        fr.save(update_fields=["titol", "hora_inici", "hora_fi"])

        # 2) Reajusta totes les franges per sota perquè quedin encadenades
        below = list(
            RotacioFranja.objects.filter(
                competicio=competicio,
                ordre__gt=fr.ordre
            ).order_by("ordre", "id")
        )

        prev_end = d_hf
        for f in below:
            # mantenim la durada original de cada franja
            s0 = datetime.combine(day, f.hora_inici)
            e0 = datetime.combine(day, f.hora_fi)
            if e0 <= s0:
                # si alguna té hores dolentes, parem per evitar fer més mal
                raise ValueError("Hi ha una franja amb durada no positiva")

            dur = e0 - s0
            new_start = prev_end
            new_end = new_start + dur

            f.hora_inici = new_start.time()
            f.hora_fi = new_end.time()
            prev_end = new_end

        if below:
            RotacioFranja.objects.bulk_update(below, ["hora_inici", "hora_fi"], batch_size=200)

    return JsonResponse({"ok": True})


@require_POST
@csrf_protect
def franja_order_mode_set(request, pk, franja_id):
    competicio = get_object_or_404(Competicio, pk=pk)
    get_object_or_404(RotacioFranja, pk=franja_id, competicio=competicio)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invalid")

    mode = payload.get("mode")
    clean_mode = set_rotacio_order_mode(competicio, franja_id=franja_id, mode=mode)
    return JsonResponse({"ok": True, "franja_id": int(franja_id), "mode": clean_mode})


@require_POST
@csrf_protect
def rotacions_export_meta_save(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invalid")

    title = str(payload.get("title", "") or "").strip()
    venue = str(payload.get("venue", "") or "").strip()
    date_str = str(payload.get("date", "") or "").strip()
    if date_str:
        if parse_date(date_str) is None:
            return HttpResponseBadRequest("Data invalida. Format esperat: YYYY-MM-DD")
    participant_fields = _normalize_export_participant_fields(
        competicio,
        payload.get("participant_fields"),
    )

    current = _get_export_meta(competicio)
    current["title"] = title or _export_meta_defaults(competicio)["title"]
    current["venue"] = venue
    current["date"] = date_str
    current["participant_fields"] = participant_fields
    _save_export_meta(competicio, current)

    return JsonResponse(
        {
            "ok": True,
            "meta": {
                "title": current["title"],
                "venue": current["venue"],
                "date": current["date"],
                "logo_path": current.get("logo_path", ""),
                "logo_url": _logo_url_from_path(current.get("logo_path", "")),
                "participant_fields": current.get("participant_fields", []),
            },
        }
    )


@require_POST
@csrf_protect
def rotacions_export_logo_upload(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    f = request.FILES.get("logo")
    if not f:
        return HttpResponseBadRequest("Falta fitxer 'logo'")

    max_bytes = 4 * 1024 * 1024
    if int(getattr(f, "size", 0) or 0) > max_bytes:
        return HttpResponseBadRequest("El logo supera el maxim de 4MB")

    ctype = str(getattr(f, "content_type", "") or "").lower()
    if not ctype.startswith("image/"):
        return HttpResponseBadRequest("El fitxer ha de ser una imatge")

    ext = os.path.splitext(getattr(f, "name", "") or "")[1].lower()
    if ext not in {".png", ".jpg", ".jpeg", ".bmp"}:
        ext = ".png"

    rel_dir = f"rotacions/logos/competicio_{competicio.id}"
    rel_path = f"{rel_dir}/{uuid.uuid4().hex}{ext}"

    content = f.read()
    saved_rel = default_storage.save(rel_path, ContentFile(content))

    current = _get_export_meta(competicio)
    old_logo = str(current.get("logo_path", "") or "").strip()
    current["logo_path"] = saved_rel
    _save_export_meta(competicio, current)

    if old_logo and old_logo != saved_rel:
        try:
            if default_storage.exists(old_logo):
                default_storage.delete(old_logo)
        except Exception:
            pass

    return JsonResponse(
        {
            "ok": True,
            "logo_path": saved_rel,
            "logo_url": _logo_url_from_path(saved_rel),
        }
    )


@require_POST
@csrf_protect
def rotacions_export_logo_clear(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    current = _get_export_meta(competicio)
    old_logo = str(current.get("logo_path", "") or "").strip()
    current["logo_path"] = ""
    _save_export_meta(competicio, current)

    if old_logo:
        try:
            if default_storage.exists(old_logo):
                default_storage.delete(old_logo)
        except Exception:
            pass

    return JsonResponse({"ok": True})



def franges_export_excel(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    mode = (request.GET.get("mode") or "participants").strip().lower()
    if mode not in {"participants", "groups"}:
        mode = "participants"

    estacions = list(
        RotacioEstacio.objects.filter(competicio=competicio, actiu=True).order_by("ordre", "id")
    )
    franges = list(RotacioFranja.objects.filter(competicio=competicio).order_by("ordre", "id"))

    franja_modes = get_rotacions_order_modes(competicio)
    franja_pos = franja_index_map(franges)

    assigns = RotacioAssignacio.objects.filter(competicio=competicio).values(
        "franja_id", "estacio_id", "grup", "grups"
    )
    cell_groups = {}
    for a in assigns:
        gs = assignacio_grups_from_values(a.get("grups"), a.get("grup"))
        cell_groups[(a["franja_id"], a["estacio_id"])] = gs

    estacio_comp_aparell = {
        e.id: (e.comp_aparell_id if getattr(e, "tipus", None) == "aparell" else None)
        for e in estacions
    }
    comp_aparell_ids = sorted({x for x in estacio_comp_aparell.values() if x})

    grups = sorted({g for gs in cell_groups.values() for g in gs})
    ins_by_grup = {}
    excluded_pairs = set()
    if grups:
        qs = (
            Inscripcio.objects.filter(competicio=competicio, grup__in=grups)
            .only(
                "id",
                "grup",
                "nom_i_cognoms",
                "document",
                "sexe",
                "data_naixement",
                "entitat",
                "categoria",
                "subcategoria",
                "ordre_sortida",
                "extra",
            )
            .order_by("ordre_sortida", "id")
        )
        ins_ids = []
        for ins in qs:
            ins_by_grup.setdefault(ins.grup, []).append(ins)
            ins_ids.append(ins.id)

        if ins_ids and comp_aparell_ids:
            excluded_pairs = set(
                InscripcioAparellExclusio.objects.filter(
                    inscripcio_id__in=ins_ids,
                    comp_aparell_id__in=comp_aparell_ids,
                ).values_list("inscripcio_id", "comp_aparell_id")
            )

    view_cfg = competicio.inscripcions_view or {}
    group_names = view_cfg.get("group_names") or {}
    if not isinstance(group_names, dict):
        group_names = {}

    def _group_label(g):
        return (group_names.get(str(g)) or "").strip() or f"G{g}"

    export_meta = _get_export_meta(competicio)
    available_participant_fields = _rotacions_available_participant_fields(competicio)
    participant_field_labels = {
        f["code"]: str(f.get("label") or f.get("code") or "").strip()
        for f in available_participant_fields
    }
    participant_fields = _normalize_export_participant_fields(
        competicio,
        export_meta.get("participant_fields"),
    )

    def _inscripcio_field_value(ins, code: str):
        extra = getattr(ins, "extra", None) or {}
        if isinstance(code, str) and code.startswith("excel__") and isinstance(extra, dict):
            if code in extra:
                return extra.get(code)
            legacy_code = code[len("excel__"):]
            if legacy_code in extra:
                return extra.get(legacy_code)
        if hasattr(ins, code):
            return getattr(ins, code)
        if isinstance(extra, dict):
            return extra.get(code)
        return None

    def _format_field_value(value):
        if value in (None, ""):
            return "-"
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, date):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    def _fit_cell_text(value: str, width: int):
        text = str(value or "")
        if len(text) <= width:
            return text.ljust(width)
        if width <= 3:
            return text[:width]
        return (text[: width - 3] + "...")

    def _render_participants_block(inscripcions):
        if not inscripcions:
            return "-"
        if not participant_fields:
            return "\n".join(
                (_format_field_value(getattr(ins, "nom_i_cognoms", None)) or "-")
                for ins in inscripcions
            )

        headers = [participant_field_labels.get(code, code) for code in participant_fields]
        rows = []
        for ins in inscripcions:
            rows.append(
                [
                    _format_field_value(_inscripcio_field_value(ins, code))
                    for code in participant_fields
                ]
            )

        widths = []
        for idx, header in enumerate(headers):
            col_values = [row[idx] for row in rows]
            col_width = max([len(header)] + [len(v) for v in col_values])
            widths.append(min(24, max(6, col_width)))

        lines = [
            " | ".join(_fit_cell_text(headers[idx], widths[idx]) for idx in range(len(widths))),
            "-+-".join("-" * widths[idx] for idx in range(len(widths))),
        ]
        for row in rows:
            lines.append(
                " | ".join(_fit_cell_text(row[idx], widths[idx]) for idx in range(len(widths)))
            )
        return "\n".join(lines)

    titol_competicio = str(export_meta.get("title", "") or "").strip() or getattr(
        competicio, "nom", f"Competicio {competicio.id}"
    )
    seu = str(export_meta.get("venue", "") or "").strip() or (getattr(competicio, "seu", "") or "-")

    data_comp = None
    date_meta = str(export_meta.get("date", "") or "").strip()
    if date_meta:
        data_comp = parse_date(date_meta)
    if not data_comp:
        data_comp = getattr(competicio, "data", None)
    data_txt = data_comp.strftime("%d/%m/%Y") if data_comp else ""
    logo_path = str(export_meta.get("logo_path", "") or "").strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rotacions"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_no_wrap = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)
    mono = Font(name="Consolas", size=10)

    fill_title = PatternFill("solid", fgColor="1F4E79")
    fill_sub = PatternFill("solid", fgColor="D9E1F2")
    fill_hdr = PatternFill("solid", fgColor="E9EEF7")
    fill_zebra = PatternFill("solid", fgColor="F6F8FC")

    thin = Side(style="thin", color="9AA7B2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_cols = 1 + len(estacions)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=titol_competicio)
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = fill_title
    c.alignment = center_no_wrap

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c = ws.cell(row=2, column=1, value=f"Seu: {seu}    {data_txt}")
    c.font = Font(bold=True)
    c.fill = fill_sub
    c.alignment = center_no_wrap

    ws.append([])

    header_row = ws.max_row + 1
    ws.cell(row=header_row, column=1, value="Franja").font = bold
    ws.cell(row=header_row, column=1).fill = fill_hdr
    ws.cell(row=header_row, column=1).alignment = center_no_wrap
    ws.cell(row=header_row, column=1).border = border

    for j, e in enumerate(estacions, start=2):
        cell = ws.cell(row=header_row, column=j, value=e.nom)
        cell.font = bold
        cell.fill = fill_hdr
        cell.alignment = center_no_wrap
        cell.border = border

    row_line_counts = {}

    for i, f in enumerate(franges, start=1):
        r = header_row + i
        label = (f.titol or "").strip() or "Franja"
        fr_txt = f"{label}\n{f.hora_inici.strftime('%H:%M')}-{f.hora_fi.strftime('%H:%M')}"
        row_max_lines = fr_txt.count("\n") + 1

        c0 = ws.cell(row=r, column=1, value=fr_txt)
        c0.alignment = center
        c0.border = border

        if i % 2 == 0:
            for col in range(1, total_cols + 1):
                ws.cell(row=r, column=col).fill = fill_zebra

        for j, e in enumerate(estacions, start=2):
            gs = cell_groups.get((f.id, e.id), [])
            if not gs:
                txt = ""
            elif mode == "groups":
                labels = unique_ordered(_group_label(g) for g in gs)
                txt = "\n".join(labels) if labels else "-"
            else:
                mode_for_franja = franja_modes.get(str(f.id), ORDER_MODE_MAINTAIN)
                rotate_steps = franja_pos.get(f.id, 0)
                comp_aparell_id = estacio_comp_aparell.get(e.id)

                ordered_inscripcions = []
                seen_ins = set()
                for g in gs:
                    base_pairs = []
                    for ins in ins_by_grup.get(g, []):
                        ins_id = ins.id
                        if comp_aparell_id and (ins_id, comp_aparell_id) in excluded_pairs:
                            continue
                        base_pairs.append((ins_id, ins))

                    ordered_pairs = order_pairs_for_mode(
                        base_pairs,
                        mode_for_franja,
                        rotate_steps=rotate_steps,
                        seed_prefix=f"rot-export|{competicio.id}|{f.id}|{e.id}|{g}",
                    )
                    for ins_id, ins in ordered_pairs:
                        if ins_id in seen_ins:
                            continue
                        seen_ins.add(ins_id)
                        ordered_inscripcions.append(ins)

                txt = _render_participants_block(ordered_inscripcions)
            row_max_lines = max(row_max_lines, (txt.count("\n") + 1) if txt else 1)

            cell = ws.cell(row=r, column=j, value=txt)
            if mode == "participants":
                cell.alignment = left_top
                cell.font = mono
            else:
                cell.alignment = center
            cell.border = border
        row_line_counts[r] = row_max_lines

    ws.column_dimensions[get_column_letter(1)].width = 22
    for j in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 36 if mode == "participants" else 24

    for r in range(header_row + 1, header_row + 1 + len(franges)):
        if mode == "participants":
            lines = max(1, int(row_line_counts.get(r, 1)))
            ws.row_dimensions[r].height = max(60, min(220, 13 * lines + 8))
        else:
            ws.row_dimensions[r].height = 30

    logo_added = False
    if logo_path:
        logo_abs = _logo_abs_path(logo_path)
        if logo_abs and os.path.exists(logo_abs):
            try:
                img = XLImage(logo_abs)
                img.height = 52
                img.width = 120
                anchor_col = max(1, total_cols - 1)
                img.anchor = f"{get_column_letter(anchor_col)}1"
                ws.add_image(img)
                logo_added = True
            except Exception:
                logo_added = False

    ws.row_dimensions[1].height = 42 if logo_added else 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[header_row].height = 22

    ws.freeze_panes = ws["B" + str(header_row + 1)]

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    suffix = "participants" if mode == "participants" else "grups"
    response["Content-Disposition"] = (
        f'attachment; filename="rotacions_{competicio.id}_{suffix}.xlsx"'
    )
    wb.save(response)
    return response

