import random
import hashlib
import copy
from django.shortcuts import render
import math
from datetime import date, datetime
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import CreateView, FormView, ListView, TemplateView, DeleteView
from django.shortcuts import redirect
from django.db.models import Q
from .forms import ImportInscripcionsExcelForm
from .models import (
    Competicio,
    CompeticioMembership,
    Equip,
    EquipContext,
    GrupCompeticio,
    Inscripcio,
    InscripcioEquipAssignacio,
)
from .forms import CompeticioForm
from .access import user_has_competicio_capability
from .services.import_excel import importar_inscripcions_excel
import json
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from django.db import transaction
from django.views.generic import UpdateView
from django.urls import reverse
from .forms import InscripcioForm
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models import Q, Max, Min, Case, When, IntegerField
from collections import defaultdict
from collections import OrderedDict

from .models_trampoli import InscripcioAparellExclusio
from .models_rotacions import RotacioAssignacio
from .services.competition_groups import (
    assign_groups_by_display_num,
    compact_competition_order_for_group,
    ensure_group_for_display_num,
    get_group_for_display_num,
    get_group_maps,
    group_label,
    get_programmed_group_ids,
    move_inscripcio_to_group,
    next_group_display_num,
    normalize_positive_int,
    save_group_competition_order,
    sync_competicio_group_names_view,
)


BUILTIN_SORT_FIELDS = [
    {"code": "nom_i_cognoms", "label": "Nom (A-Z)", "kind": "builtin"},
    {"code": "data_naixement", "label": "Edat / Data naixement", "kind": "builtin"},
    {"code": "document", "label": "Document", "kind": "builtin"},
    {"code": "sexe", "label": "Sexe", "kind": "builtin"},
    {"code": "entitat", "label": "Entitat", "kind": "builtin"},
    {"code": "categoria", "label": "Categoria", "kind": "builtin"},
    {"code": "subcategoria", "label": "Subcategoria", "kind": "builtin"},
    {"code": "grup", "label": "Grup", "kind": "builtin"},
    {"code": "equip", "label": "Equip", "kind": "builtin"},
]

# Compatibilitat amb URLs antigues (?sort_key=nom|edat)
LEGACY_SORT_KEY_MAP = {
    "nom": "nom_i_cognoms",
    "edat": "data_naixement",
}

BUILTIN_EXCEL_FIELDS = [
    {"code": "nom_i_cognoms", "label": "Nom i cognoms", "kind": "builtin"},
    {"code": "document", "label": "DNI", "kind": "builtin"},
    {"code": "sexe", "label": "Sexe", "kind": "builtin"},
    {"code": "data_naixement", "label": "Data naixement", "kind": "builtin"},
    {"code": "entitat", "label": "Entitat", "kind": "builtin"},
    {"code": "categoria", "label": "Categoria", "kind": "builtin"},
    {"code": "subcategoria", "label": "Subcategoria", "kind": "builtin"},
    {"code": "grup", "label": "Grup", "kind": "builtin"},
    {"code": "ordre_sortida", "label": "Ordre", "kind": "builtin"},
]

# Compatibilitat amb URLs antigues (?excel_cols=nom|dni|naixement|ordre)
LEGACY_EXCEL_COL_MAP = {
    "nom": "nom_i_cognoms",
    "dni": "document",
    "naixement": "data_naixement",
    "ordre": "ordre_sortida",
}

# Camps built-in del model que volem permetre agrupar sempre
BUILTIN_GROUP_FIELDS = [
    {"code": "categoria", "label": "Categoria", "kind": "builtin"},
    {"code": "subcategoria", "label": "Subcategoria", "kind": "builtin"},
    {"code": "entitat", "label": "Entitat", "kind": "builtin"},
    {"code": "sexe", "label": "Sexe", "kind": "builtin"},
    {"code": "data_naixement", "label": "Data naixement", "kind": "builtin"},
    {"code": "any_naixement_forquilla", "label": "Forquilla any naixement", "kind": "derived"},
    {"code": "document", "label": "Document", "kind": "builtin"},
]


def _reserved_inscripcio_codes():
    out = set()
    for f in Inscripcio._meta.concrete_fields:
        name = str(getattr(f, "name", "") or "").strip()
        attname = str(getattr(f, "attname", "") or "").strip()
        if name:
            out.add(name)
        if attname:
            out.add(attname)
    return out


def _normalize_schema_extra_code(code: str, reserved_codes=None):
    code = (code or "").strip()
    if not code:
        return code
    if code.startswith("excel__"):
        return code
    if reserved_codes is None:
        reserved_codes = _reserved_inscripcio_codes()
    if code in reserved_codes:
        return f"excel__{code}"
    return code


def _excel_schema_codes(competicio):
    schema = competicio.inscripcions_schema or {}
    cols = schema.get("columns") or []
    if not isinstance(cols, list):
        return set()
    reserved = _reserved_inscripcio_codes()
    out = set()
    for c in cols:
        if not isinstance(c, dict):
            continue
        code = c.get("code")
        if not code:
            continue
        kind = c.get("kind") or "extra"
        if kind == "extra":
            code = _normalize_schema_extra_code(code, reserved)
        out.add(code)
    return out


def _label_with_source(label: str, source: str):
    if source == "excel":
        suffix = "Excel"
    elif source == "derived":
        suffix = "Derivada"
    else:
        suffix = "Nativa"
    return f"{label} ({suffix})"


# Per mostrar/ordenar values
def _s(v):
    return "(Sense valor)" if v in (None, "") else str(v)

def _norm_val(v):
    return "__NULL__" if v in (None, "") else str(v)

def get_inscripcio_value(obj, code: str):
    """
    Retorna valor per a un camp d'agrupació.
    - built-in: getattr
    - extra: obj.extra.get(code)
    """
    extra = getattr(obj, "extra", None) or {}
    if isinstance(extra, dict) and isinstance(code, str) and code.startswith("excel__"):
        if code in extra:
            return extra.get(code)
        legacy_code = code[len("excel__"):]
        if legacy_code in extra:
            return extra.get(legacy_code)
    if hasattr(obj, code):
        return getattr(obj, code)
    if isinstance(extra, dict) and code in extra:
        return extra.get(code)
    return extra.get(code)

def get_allowed_group_fields(competicio):
    """
    Retorna llista [{code,label,kind}, ...] = builtins + extras detectats a schema.
    """
    out = []
    seen = set()
    excel_codes = _excel_schema_codes(competicio)
    reserved = _reserved_inscripcio_codes()

    # builtins sempre
    for f in BUILTIN_GROUP_FIELDS:
        if f["code"] not in seen:
            if f.get("kind") == "derived":
                source = "derived"
            else:
                source = "excel" if f["code"] in excel_codes else "native"
            out.append(
                {
                    **f,
                    "source": source,
                    "ui_label": _label_with_source(f["label"], source),
                }
            )
            seen.add(f["code"])

    schema = competicio.inscripcions_schema or {}
    cols = schema.get("columns") or []
    if isinstance(cols, list):
        for c in cols:
            if not isinstance(c, dict):
                continue
            code = c.get("code")
            if not code:
                continue
            kind = c.get("kind") or "extra"
            if kind != "extra":
                continue
            code = _normalize_schema_extra_code(code, reserved)
            label = c.get("label") or code
            if code not in seen:
                out.append(
                    {
                        "code": code,
                        "label": label,
                        "kind": "extra",
                        "source": "excel",
                        "ui_label": _label_with_source(label, "excel"),
                    }
                )
                seen.add(code)

    return out


def get_available_sort_fields(competicio):
    """
    Camps ordenables: builtins + extras detectats a schema.columns.
    """
    out = []
    seen = set()
    excel_codes = _excel_schema_codes(competicio)
    reserved = _reserved_inscripcio_codes()

    for f in BUILTIN_SORT_FIELDS:
        if f["code"] not in seen:
            source = "excel" if f["code"] in excel_codes else "native"
            out.append(
                {
                    **f,
                    "source": source,
                    "ui_label": _label_with_source(f["label"], source),
                }
            )
            seen.add(f["code"])

    schema = competicio.inscripcions_schema or {}
    cols = schema.get("columns") or []
    if isinstance(cols, list):
        for c in cols:
            if not isinstance(c, dict):
                continue
            code = c.get("code")
            if not code or code in seen:
                continue
            kind = c.get("kind") or "extra"
            if kind != "extra":
                continue
            code = _normalize_schema_extra_code(code, reserved)
            if code in seen:
                continue
            label = c.get("label") or code
            out.append(
                {
                    "code": code,
                    "label": label,
                    "kind": "extra",
                    "source": "excel",
                    "ui_label": _label_with_source(label, "excel"),
                }
            )
            seen.add(code)

    return out


def _normalize_custom_sort_token(value):
    if value in (None, ""):
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def _custom_sort_token_key(value):
    token = str(value or "").strip()
    return token.casefold()


def _normalize_custom_sort_order(raw_values):
    out = []
    seen = set()
    if not isinstance(raw_values, list):
        return out

    for raw in raw_values:
        token = _normalize_custom_sort_token(raw)
        if not token:
            continue
        key = _custom_sort_token_key(token)
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(token)
    return out


def _get_custom_sort_orders_map(competicio):
    cfg = competicio.inscripcions_view or {}
    raw = cfg.get("custom_sort_orders")
    if not isinstance(raw, dict):
        return {}

    out = {}
    for raw_code, raw_values in raw.items():
        if not isinstance(raw_code, str):
            continue
        code = LEGACY_SORT_KEY_MAP.get(raw_code.strip(), raw_code.strip())
        if not code:
            continue
        values = _normalize_custom_sort_order(raw_values)
        if values:
            out[code] = values
    return out


def get_competicio_custom_sort_order_values(competicio, sort_code, allowed_sort_codes=None):
    code_raw = str(sort_code or "").strip()
    code = LEGACY_SORT_KEY_MAP.get(code_raw, code_raw)
    if not code:
        return []

    if allowed_sort_codes is not None and code not in set(allowed_sort_codes):
        return []

    data = _get_custom_sort_orders_map(competicio)
    return list(data.get(code) or [])


def get_competicio_custom_sort_rank_map(competicio, sort_code, allowed_sort_codes=None):
    values = get_competicio_custom_sort_order_values(
        competicio,
        sort_code,
        allowed_sort_codes=allowed_sort_codes,
    )
    out = {}
    for idx, token in enumerate(values):
        key = _custom_sort_token_key(token)
        if key and key not in out:
            out[key] = idx
    return out


def get_competicio_custom_sort_codes(competicio, allowed_sort_codes=None):
    data = _get_custom_sort_orders_map(competicio)
    if allowed_sort_codes is None:
        return sorted(data.keys())
    allowed = set(allowed_sort_codes)
    return sorted([code for code in data.keys() if code in allowed])


def set_competicio_custom_sort_order_values(competicio, sort_code, raw_values=None, clear=False, allowed_sort_codes=None):
    code_raw = str(sort_code or "").strip()
    code = LEGACY_SORT_KEY_MAP.get(code_raw, code_raw)
    if not code:
        raise ValueError("sort_key invalid")
    if allowed_sort_codes is not None and code not in set(allowed_sort_codes):
        raise ValueError("sort_key invalid")

    values = [] if clear else _normalize_custom_sort_order(raw_values)

    view_cfg = dict(competicio.inscripcions_view or {})
    custom_map = view_cfg.get("custom_sort_orders")
    if not isinstance(custom_map, dict):
        custom_map = {}
    custom_map = dict(custom_map)

    if values:
        custom_map[code] = values
    else:
        custom_map.pop(code, None)

    if custom_map:
        view_cfg["custom_sort_orders"] = custom_map
    else:
        view_cfg.pop("custom_sort_orders", None)

    competicio.inscripcions_view = view_cfg
    competicio.save(update_fields=["inscripcions_view"])
    return values


def _split_custom_sort_tokens(custom_tokens, available_token_keys):
    """
    Separa l'ordre custom en valors vigents i obsolets segons els tokens disponibles.
    """
    active = []
    stale = []
    available = set(available_token_keys or set())
    seen = set()
    for raw in custom_tokens or []:
        token = _normalize_custom_sort_token(raw)
        if not token:
            continue
        key = _custom_sort_token_key(token)
        if not key or key in seen:
            continue
        seen.add(key)
        if key in available:
            active.append(token)
        else:
            stale.append(token)
    return active, stale


def get_available_excel_columns(competicio):
    """
    Columnes exportables a Excel: builtins + extras detectats al schema.
    """
    out = []
    seen = set()
    excel_codes = _excel_schema_codes(competicio)
    reserved = _reserved_inscripcio_codes()

    for f in BUILTIN_EXCEL_FIELDS:
        if f["code"] not in seen:
            source = "excel" if f["code"] in excel_codes else "native"
            out.append(
                {
                    **f,
                    "source": source,
                    "ui_label": _label_with_source(f["label"], source),
                }
            )
            seen.add(f["code"])

    schema = competicio.inscripcions_schema or {}
    cols = schema.get("columns") or []
    if isinstance(cols, list):
        for c in cols:
            if not isinstance(c, dict):
                continue
            code = c.get("code")
            if not code or code in seen:
                continue
            kind = c.get("kind") or "extra"
            if kind != "extra":
                continue
            code = _normalize_schema_extra_code(code, reserved)
            if code in seen:
                continue
            label = c.get("label") or code
            out.append(
                {
                    "code": code,
                    "label": label,
                    "kind": "extra",
                    "source": "excel",
                    "ui_label": _label_with_source(label, "excel"),
                }
            )
            seen.add(code)

    return out


def get_excel_export_value(obj, code):
    v = get_inscripcio_value(obj, code)
    if code == "data_naixement":
        return v.strftime("%d/%m/%Y") if v else "-"
    if v in (None, ""):
        return "-"
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


def _sort_scalar(v):
    if isinstance(v, (int, float)):
        return (0, float(v))
    if isinstance(v, (date, datetime)):
        return (1, v.isoformat())
    return (2, str(v).casefold())


def _build_sort_field_runtime_context(records, sort_code):
    code = str(sort_code or "").strip()
    if code != "grup":
        return {}

    competicio_id = None
    for obj in records or []:
        competicio_id = getattr(obj, "competicio_id", None)
        if competicio_id:
            break
    if not competicio_id:
        return {}
    return {"group_maps": get_group_maps(competicio_id)}


def _resolve_sort_field_runtime(obj, sort_code, context=None):
    code = str(sort_code or "").strip()
    ctx = context if isinstance(context, dict) else {}
    raw_value = get_inscripcio_value(obj, code)

    if code == "grup":
        group = getattr(obj, "grup_competicio", None)
        display_num = normalize_positive_int(getattr(obj, "grup", None))
        if not display_num and group is not None:
            display_num = normalize_positive_int(getattr(group, "display_num", None))
        if group is None and display_num:
            group = ((ctx.get("group_maps") or {}).get("by_display_num") or {}).get(display_num)

        token = str(display_num) if display_num else ""
        group_name = str(getattr(group, "nom", "") or "").strip() if group is not None else ""
        label = group_name or (f"Grup {display_num}" if display_num else "")
        return {
            "raw_value": display_num,
            "token": token,
            "label": label,
            "sort_scalar": _sort_scalar(display_num if display_num is not None else ""),
        }

    if code == "equip":
        equip_id = getattr(obj, "equip_id", None)
        equip = getattr(obj, "equip", None)
        equip_name = str(getattr(equip, "nom", "") or "").strip() if equip is not None else ""
        token = str(equip_id) if equip_id else ""
        label = equip_name or (f"Equip {equip_id}" if equip_id else "")
        return {
            "raw_value": equip_name,
            "token": token,
            "label": label,
            "sort_scalar": _sort_scalar(equip_name),
        }

    token = _normalize_custom_sort_token(raw_value)
    return {
        "raw_value": raw_value,
        "token": token,
        "label": _s(raw_value) if token else "",
        "sort_scalar": _sort_scalar(raw_value),
    }


def _build_sort_records_queryset(qs, sort_codes=None, include_competition_order=False):
    codes = []
    seen_codes = set()
    for raw_code in sort_codes or []:
        code = str(raw_code or "").strip()
        if not code or code in seen_codes:
            continue
        seen_codes.add(code)
        codes.append(code)

    select_related_fields = []
    only_fields = ["id", "competicio_id", "extra", "ordre_sortida"]
    if include_competition_order:
        only_fields.extend(["grup", "ordre_competicio"])

    for code in codes:
        if code == "grup":
            select_related_fields.append("grup_competicio")
            only_fields.extend([
                "grup",
                "grup_competicio_id",
                "grup_competicio__display_num",
                "grup_competicio__nom",
            ])
            continue
        if code == "equip":
            select_related_fields.append("equip")
            only_fields.extend([
                "equip_id",
                "equip__nom",
            ])
            continue
        if hasattr(Inscripcio, code):
            only_fields.append(code)

    qs_out = qs
    if select_related_fields:
        dedup_related = list(dict.fromkeys(select_related_fields))
        qs_out = qs_out.select_related(*dedup_related)
    if only_fields:
        dedup_only = list(dict.fromkeys(only_fields))
        qs_out = qs_out.only(*dedup_only)
    return qs_out.order_by("ordre_sortida", "id")


def sort_records_by_field(records, sort_code, descending=False):
    """
    Ordenació portable (builtins + extra JSON), amb buits al final.
    """
    filled = []
    empty = []

    for o in records:
        v = get_inscripcio_value(o, sort_code)
        if v in (None, ""):
            empty.append(o)
        else:
            filled.append((o, v))

    # tiebreak estable per id asc, independent de l'ordre asc/desc
    filled.sort(key=lambda t: t[0].id)
    filled.sort(key=lambda t: _sort_scalar(t[1]), reverse=descending)
    empty.sort(key=lambda o: o.id)

    return [o for (o, _v) in filled] + empty


def sort_records_by_field_stable(records, sort_code, descending=False, custom_rank_map=None):
    """
    Ordenacio estable (acumulativa): en empats conserva l'ordre d'entrada.
    Els buits continuen al final, conservant ordre relatiu.
    """
    custom_map = custom_rank_map if isinstance(custom_rank_map, dict) else {}
    custom_enabled = bool(custom_map)
    custom_filled = []
    fallback_filled = []
    empty = []
    context = _build_sort_field_runtime_context(records, sort_code)

    for o in records:
        runtime = _resolve_sort_field_runtime(o, sort_code, context=context)
        token = runtime.get("token") or ""
        if not token:
            empty.append(o)
        else:
            if custom_enabled:
                key = _custom_sort_token_key(token)
                if key in custom_map:
                    custom_filled.append((o, custom_map[key]))
                    continue
            fallback_filled.append((o, runtime.get("sort_scalar")))

    # Python sort es estable: en valors iguals manté l'ordre previ.
    custom_filled.sort(key=lambda t: t[1], reverse=descending)
    fallback_filled.sort(key=lambda t: t[1], reverse=descending)

    return [o for (o, _rank) in custom_filled] + [o for (o, _v) in fallback_filled] + empty


def _s(v):
    return "(Sense valor)" if v in (None, "") else str(v)

def arrow_positions(n: int) -> list[int]:
    """
    Retorna la seqüència de posicions "fletxa" per un grup de mida n.

    Exemple n=8 -> [3,4,2,5,1,6,0,7]
    (index del registre ordenat) -> (posició dins del grup)
    """
    if n <= 0:
        return []

    seq = []
    if n % 2 == 0:
        left = n // 2 - 1
        right = n // 2
        while left >= 0 or right < n:
            if left >= 0:
                seq.append(left)
                left -= 1
            if right < n:
                seq.append(right)
                right += 1
    else:
        center = n // 2
        seq.append(center)
        step = 1
        while center - step >= 0 or center + step < n:
            if center - step >= 0:
                seq.append(center - step)
            if center + step < n:
                seq.append(center + step)
            step += 1

    return seq

def init_ordre_sortida(competicio_id):
    qs = Inscripcio.objects.filter(competicio_id=competicio_id).order_by("id")
    with transaction.atomic():
        for i, obj in enumerate(qs, start=1):
            Inscripcio.objects.filter(id=obj.id).update(ordre_sortida=i)

def shuffle_ordre_sortida(qs):
    ids = list(qs.values_list("id", flat=True))
    random.shuffle(ids)

    with transaction.atomic():
        for idx, ins_id in enumerate(ids, start=1):
            Inscripcio.objects.filter(id=ins_id).update(ordre_sortida=idx)



INSCRIPCIONS_SORT_STACK_SESSION_KEY = "inscripcions_sort_stack_v2"
INSCRIPCIONS_HISTORY_SESSION_KEY = "inscripcions_history_v2"
INSCRIPCIONS_HISTORY_DEPTH = 20

def _normalize_sort_filters(raw_filters):
    data = raw_filters if isinstance(raw_filters, dict) else {}
    return {
        "q": str(data.get("q") or "").strip(),
        "categoria": str(data.get("categoria") or "").strip(),
        "subcategoria": str(data.get("subcategoria") or "").strip(),
        "entitat": str(data.get("entitat") or "").strip(),
    }


def _normalize_sort_group_by(raw_group_by, allowed_group_codes, fallback_group_by=None):
    out = []
    if isinstance(raw_group_by, list):
        for value in raw_group_by:
            if isinstance(value, str) and value in allowed_group_codes and value not in out:
                out.append(value)
    if out:
        return out
    fallback = fallback_group_by if isinstance(fallback_group_by, list) else []
    return [g for g in fallback if g in allowed_group_codes]


def build_inscripcions_sort_context_key(competicio_id, filters=None, group_by=None):
    clean_filters = _normalize_sort_filters(filters)
    clean_group_by = []
    if isinstance(group_by, list):
        for g in group_by:
            if isinstance(g, str) and g not in clean_group_by:
                clean_group_by.append(g)

    parts = [
        str(competicio_id),
        clean_filters["q"],
        clean_filters["categoria"],
        clean_filters["subcategoria"],
        clean_filters["entitat"],
        "|".join(clean_group_by),
    ]
    return "||".join(parts)


def compute_inscripcions_order_signature_from_ids(ids):
    h = hashlib.sha1()
    for ins_id in ids:
        h.update(str(ins_id).encode("utf-8"))
        h.update(b",")
    return h.hexdigest()


def compute_inscripcions_order_signature_for_queryset(qs):
    ids = qs.values_list("id", flat=True)
    return compute_inscripcions_order_signature_from_ids(ids)


def _read_sort_stack_store(request):
    raw = request.session.get(INSCRIPCIONS_SORT_STACK_SESSION_KEY)
    if not isinstance(raw, dict):
        return {}
    return raw


def _write_sort_stack_store(request, store):
    request.session[INSCRIPCIONS_SORT_STACK_SESSION_KEY] = store
    request.session.modified = True


def _history_comp_key(competicio_id):
    return str(int(competicio_id))


def _read_inscripcions_history_store(request):
    raw = request.session.get(INSCRIPCIONS_HISTORY_SESSION_KEY)
    if not isinstance(raw, dict):
        return {}
    return raw


def _write_inscripcions_history_store(request, store):
    request.session[INSCRIPCIONS_HISTORY_SESSION_KEY] = store
    request.session.modified = True


def _json_clone(value):
    try:
        return json.loads(json.dumps(value))
    except Exception:
        return copy.deepcopy(value)


def _capture_sort_stack_state_for_competicio(request, competicio_id):
    prefix = f"{competicio_id}||"
    stack_store = _read_sort_stack_store(request)
    out = {}
    for key, value in stack_store.items():
        if isinstance(key, str) and key.startswith(prefix) and isinstance(value, dict):
            out[key] = _json_clone(value)
    return out


def _restore_sort_stack_state_for_competicio(request, competicio_id, state_snapshot):
    prefix = f"{competicio_id}||"
    stack_store = _read_sort_stack_store(request)
    for key in list(stack_store.keys()):
        if isinstance(key, str) and key.startswith(prefix):
            stack_store.pop(key, None)

    if isinstance(state_snapshot, dict):
        for key, value in state_snapshot.items():
            if isinstance(key, str) and key.startswith(prefix) and isinstance(value, dict):
                stack_store[key] = _json_clone(value)

    _write_sort_stack_store(request, stack_store)


def capture_inscripcions_history_snapshot(request, competicio):
    inscripcions_rows = list(
        Inscripcio.objects
        .filter(competicio=competicio)
        .order_by("id")
        .values(
            "id",
            "ordre_sortida",
            "grup",
            "grup_competicio_id",
            "ordre_competicio",
            "equip_id",
        )
    )
    grups_rows = list(
        GrupCompeticio.objects
        .filter(competicio=competicio)
        .order_by("display_num", "id")
        .values("id", "legacy_num", "display_num", "nom", "actiu")
    )
    exclusions_rows = list(
        InscripcioAparellExclusio.objects
        .filter(inscripcio__competicio=competicio)
        .order_by("inscripcio_id", "comp_aparell_id")
        .values("inscripcio_id", "comp_aparell_id", "motiu")
    )
    equips_rows = list(
        Equip.objects
        .filter(competicio=competicio)
        .order_by("id")
        .values("id", "nom", "origen", "criteri")
    )
    equip_context_rows = list(
        EquipContext.objects
        .filter(competicio=competicio)
        .order_by("id")
        .values("id", "code", "nom", "description")
    )
    equip_assignacio_rows = list(
        InscripcioEquipAssignacio.objects
        .filter(competicio=competicio)
        .order_by("context_id", "inscripcio_id")
        .values("context_id", "inscripcio_id", "equip_id", "origen", "criteri")
    )
    return {
        "inscripcions_fields": _json_clone(inscripcions_rows),
        "grups_competicio": _json_clone(grups_rows),
        "competicio_fields": {
            "tab_merges": _json_clone(competicio.tab_merges or {}),
            "inscripcions_view": _json_clone(competicio.inscripcions_view or {}),
        },
        "aparells_exclusions": _json_clone(exclusions_rows),
        "equips_state": _json_clone(equips_rows),
        "equip_contexts_state": _json_clone(equip_context_rows),
        "equip_assignacions_state": _json_clone(equip_assignacio_rows),
        "sort_stack_state": _capture_sort_stack_state_for_competicio(request, competicio.id),
    }


def _apply_equips_state_snapshot(competicio, equips_state):
    rows = equips_state if isinstance(equips_state, list) else []
    normalized = []
    seen_ids = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        try:
            equip_id = int(row.get("id"))
        except Exception:
            continue
        if equip_id <= 0 or equip_id in seen_ids:
            continue
        seen_ids.add(equip_id)
        normalized.append(
            Equip(
                id=equip_id,
                competicio=competicio,
                nom=str(row.get("nom") or "").strip(),
                origen=str(row.get("origen") or Equip.Origen.MANUAL).strip() or Equip.Origen.MANUAL,
                criteri=_json_clone(row.get("criteri") or {}),
            )
        )

    Equip.objects.filter(competicio=competicio).delete()
    if normalized:
        Equip.objects.bulk_create(normalized)


def _apply_equip_contexts_state_snapshot(competicio, equip_contexts_state):
    rows = equip_contexts_state if isinstance(equip_contexts_state, list) else []
    normalized = []
    seen_ids = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        try:
            ctx_id = int(row.get("id"))
        except Exception:
            continue
        if ctx_id <= 0 or ctx_id in seen_ids:
            continue
        code = str(row.get("code") or "").strip()
        if not code:
            continue
        seen_ids.add(ctx_id)
        normalized.append(
            EquipContext(
                id=ctx_id,
                competicio=competicio,
                code=code,
                nom=str(row.get("nom") or "").strip(),
                description=str(row.get("description") or "").strip(),
            )
        )

    EquipContext.objects.filter(competicio=competicio).delete()
    if normalized:
        EquipContext.objects.bulk_create(normalized)


def _apply_equip_assignacions_state_snapshot(competicio, equip_assignacions_state):
    rows = equip_assignacions_state if isinstance(equip_assignacions_state, list) else []
    normalized = []
    seen_keys = set()
    valid_context_ids = set(
        EquipContext.objects.filter(competicio=competicio).values_list("id", flat=True)
    )
    valid_ins_ids = set(
        Inscripcio.objects.filter(competicio=competicio).values_list("id", flat=True)
    )
    valid_equip_ids = set(
        Equip.objects.filter(competicio=competicio).values_list("id", flat=True)
    )
    for row in rows:
        if not isinstance(row, dict):
            continue
        try:
            context_id = int(row.get("context_id"))
            inscripcio_id = int(row.get("inscripcio_id"))
            equip_id = int(row.get("equip_id"))
        except Exception:
            continue
        key = (context_id, inscripcio_id)
        if (
            key in seen_keys
            or context_id not in valid_context_ids
            or inscripcio_id not in valid_ins_ids
            or equip_id not in valid_equip_ids
        ):
            continue
        seen_keys.add(key)
        normalized.append(
            InscripcioEquipAssignacio(
                competicio=competicio,
                context_id=context_id,
                inscripcio_id=inscripcio_id,
                equip_id=equip_id,
                origen=str(row.get("origen") or InscripcioEquipAssignacio.Origen.MANUAL).strip() or InscripcioEquipAssignacio.Origen.MANUAL,
                criteri=_json_clone(row.get("criteri") or {}),
            )
        )

    InscripcioEquipAssignacio.objects.filter(competicio=competicio).delete()
    if normalized:
        InscripcioEquipAssignacio.objects.bulk_create(normalized, batch_size=500)


def _apply_inscripcions_fields_snapshot(competicio, inscripcions_fields):
    rows = inscripcions_fields if isinstance(inscripcions_fields, list) else []
    by_id = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        try:
            ins_id = int(row.get("id"))
        except Exception:
            continue
        by_id[ins_id] = {
            "ordre_sortida": row.get("ordre_sortida"),
            "grup": row.get("grup"),
            "grup_competicio_id": row.get("grup_competicio_id"),
            "ordre_competicio": row.get("ordre_competicio"),
            "equip_id": row.get("equip_id"),
        }

    if not by_id:
        return

    updates = []
    qs = Inscripcio.objects.filter(competicio=competicio, id__in=list(by_id.keys()))
    for obj in qs.only("id", "ordre_sortida", "grup", "grup_competicio", "ordre_competicio", "equip"):
        row = by_id.get(obj.id)
        if row is None:
            continue
        next_ord = row.get("ordre_sortida")
        next_grp = row.get("grup")
        next_grup_competicio_id = row.get("grup_competicio_id")
        next_ordre_competicio = row.get("ordre_competicio")
        next_equip = row.get("equip_id")
        if (
            obj.ordre_sortida != next_ord
            or obj.grup != next_grp
            or obj.grup_competicio_id != next_grup_competicio_id
            or obj.ordre_competicio != next_ordre_competicio
            or obj.equip_id != next_equip
        ):
            obj.ordre_sortida = next_ord
            obj.grup = next_grp
            obj.grup_competicio_id = next_grup_competicio_id
            obj.ordre_competicio = next_ordre_competicio
            obj.equip_id = next_equip
            updates.append(obj)

    if updates:
        Inscripcio.objects.bulk_update(
            updates,
            ["ordre_sortida", "grup", "grup_competicio", "ordre_competicio", "equip"],
            batch_size=500,
        )


def _apply_grups_competicio_snapshot(competicio, grups_rows):
    rows = grups_rows if isinstance(grups_rows, list) else []
    normalized = []
    seen_ids = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        try:
            group_id = int(row.get("id"))
            display_num = int(row.get("display_num"))
        except Exception:
            continue
        if group_id <= 0 or display_num <= 0 or group_id in seen_ids:
            continue
        seen_ids.add(group_id)
        normalized.append(
            GrupCompeticio(
                id=group_id,
                competicio=competicio,
                legacy_num=row.get("legacy_num"),
                display_num=display_num,
                nom=str(row.get("nom") or "").strip(),
                actiu=bool(row.get("actiu", True)),
            )
        )

    current_ids = set(
        GrupCompeticio.objects.filter(competicio=competicio).values_list("id", flat=True)
    )
    target_ids = {obj.id for obj in normalized}
    stale_ids = list(current_ids - target_ids)
    if stale_ids:
        Inscripcio.objects.filter(competicio=competicio, grup_competicio_id__in=stale_ids).update(
            grup_competicio=None,
            ordre_competicio=None,
        )
        GrupCompeticio.objects.filter(competicio=competicio, id__in=stale_ids).delete()

    existing_ids = current_ids & target_ids
    if existing_ids:
        updates = [obj for obj in normalized if obj.id in existing_ids]
        if updates:
            GrupCompeticio.objects.bulk_update(
                updates,
                ["legacy_num", "display_num", "nom", "actiu"],
                batch_size=200,
            )
    creates = [obj for obj in normalized if obj.id not in existing_ids]
    if creates:
        GrupCompeticio.objects.bulk_create(creates, batch_size=200)


def _apply_aparells_exclusions_snapshot(competicio, exclusions):
    rows = exclusions if isinstance(exclusions, list) else []
    seen_pairs = set()
    normalized = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        try:
            ins_id = int(row.get("inscripcio_id"))
            app_id = int(row.get("comp_aparell_id"))
        except Exception:
            continue
        key = (ins_id, app_id)
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        normalized.append(
            InscripcioAparellExclusio(
                inscripcio_id=ins_id,
                comp_aparell_id=app_id,
                motiu=str(row.get("motiu") or ""),
            )
        )

    InscripcioAparellExclusio.objects.filter(inscripcio__competicio=competicio).delete()
    if normalized:
        InscripcioAparellExclusio.objects.bulk_create(normalized, batch_size=500)


def _apply_competicio_fields_snapshot(competicio, competicio_fields):
    fields = competicio_fields if isinstance(competicio_fields, dict) else {}
    tab_merges = fields.get("tab_merges")
    inscripcions_view = fields.get("inscripcions_view")

    updates = []
    if tab_merges is not None:
        competicio.tab_merges = _json_clone(tab_merges) if isinstance(tab_merges, dict) else {}
        updates.append("tab_merges")
    if inscripcions_view is not None:
        competicio.inscripcions_view = _json_clone(inscripcions_view) if isinstance(inscripcions_view, dict) else {}
        updates.append("inscripcions_view")
    if updates:
        competicio.save(update_fields=updates)


def apply_inscripcions_history_snapshot(request, competicio, snapshot):
    snap = snapshot if isinstance(snapshot, dict) else {}
    with transaction.atomic():
        _apply_equips_state_snapshot(competicio, snap.get("equips_state"))
        _apply_equip_contexts_state_snapshot(competicio, snap.get("equip_contexts_state"))
        _apply_grups_competicio_snapshot(competicio, snap.get("grups_competicio"))
        _apply_inscripcions_fields_snapshot(competicio, snap.get("inscripcions_fields"))
        _apply_equip_assignacions_state_snapshot(competicio, snap.get("equip_assignacions_state"))
        _apply_aparells_exclusions_snapshot(competicio, snap.get("aparells_exclusions"))
        _apply_competicio_fields_snapshot(competicio, snap.get("competicio_fields"))

    sync_competicio_group_names_view(competicio)
    _restore_sort_stack_state_for_competicio(request, competicio.id, snap.get("sort_stack_state"))


def record_inscripcions_history_entry(request, competicio, action_type, action_label, before_snapshot, after_snapshot):
    before = before_snapshot if isinstance(before_snapshot, dict) else {}
    after = after_snapshot if isinstance(after_snapshot, dict) else {}
    if before == after:
        return False

    entry = {
        "action_type": str(action_type or "").strip() or "unknown",
        "action_label": str(action_label or "").strip() or "Canvi",
        "created_at": datetime.utcnow().isoformat(),
        "before": before,
        "after": after,
    }

    store = _read_inscripcions_history_store(request)
    comp_key = _history_comp_key(competicio.id)
    bucket = store.get(comp_key)
    if not isinstance(bucket, dict):
        bucket = {"undo": [], "redo": []}

    undo = bucket.get("undo")
    if not isinstance(undo, list):
        undo = []
    undo.append(entry)
    if len(undo) > INSCRIPCIONS_HISTORY_DEPTH:
        undo = undo[-INSCRIPCIONS_HISTORY_DEPTH:]

    bucket["undo"] = undo
    bucket["redo"] = []
    store[comp_key] = bucket
    _write_inscripcions_history_store(request, store)
    return True


def get_inscripcions_history_state(request, competicio_id):
    store = _read_inscripcions_history_store(request)
    comp_key = _history_comp_key(competicio_id)
    bucket = store.get(comp_key)
    if not isinstance(bucket, dict):
        bucket = {}

    undo = bucket.get("undo")
    redo = bucket.get("redo")
    if not isinstance(undo, list):
        undo = []
    if not isinstance(redo, list):
        redo = []

    undo_top = undo[-1] if undo else {}
    redo_top = redo[-1] if redo else {}
    undo_label = str(undo_top.get("action_label") or "").strip()
    redo_label = str(redo_top.get("action_label") or "").strip()

    return {
        "can_undo": bool(undo),
        "can_redo": bool(redo),
        "undo_label": undo_label,
        "redo_label": redo_label,
        "undo_count": len(undo),
        "redo_count": len(redo),
    }


def with_inscripcions_history_payload(payload, request, competicio_id):
    data = payload if isinstance(payload, dict) else {}
    data["history"] = get_inscripcions_history_state(request, competicio_id)
    return data


def _perform_inscripcions_history_step(request, competicio, direction):
    if direction not in ("undo", "redo"):
        return None

    store = _read_inscripcions_history_store(request)
    comp_key = _history_comp_key(competicio.id)
    bucket = store.get(comp_key)
    if not isinstance(bucket, dict):
        bucket = {"undo": [], "redo": []}

    undo = bucket.get("undo")
    redo = bucket.get("redo")
    if not isinstance(undo, list):
        undo = []
    if not isinstance(redo, list):
        redo = []

    src = undo if direction == "undo" else redo
    dst = redo if direction == "undo" else undo
    if not src:
        return None

    entry = src[-1]
    if not isinstance(entry, dict):
        entry = {}

    target_snapshot = entry.get("before") if direction == "undo" else entry.get("after")
    apply_inscripcions_history_snapshot(request, competicio, target_snapshot)

    src.pop()
    dst.append(entry)
    if len(dst) > INSCRIPCIONS_HISTORY_DEPTH:
        dst[:] = dst[-INSCRIPCIONS_HISTORY_DEPTH:]

    bucket["undo"] = undo
    bucket["redo"] = redo
    store[comp_key] = bucket
    _write_inscripcions_history_store(request, store)
    return entry


def clear_inscripcions_sort_state_for_competicio(request, competicio_id):
    prefix = f"{competicio_id}||"

    stack_store = _read_sort_stack_store(request)
    stack_changed = False
    for key in list(stack_store.keys()):
        if isinstance(key, str) and key.startswith(prefix):
            stack_store.pop(key, None)
            stack_changed = True
    if stack_changed:
        _write_sort_stack_store(request, stack_store)


def get_inscripcions_sort_context_state(request, context_key):
    if not context_key:
        return {
            "stack": [],
            "order_sig": "",
            "base_ids": [],
            "context_ids": [],
            "competition_order_tail": False,
        }
    store = _read_sort_stack_store(request)
    state = store.get(context_key)
    if not isinstance(state, dict):
        return {
            "stack": [],
            "order_sig": "",
            "base_ids": [],
            "context_ids": [],
            "competition_order_tail": False,
        }
    stack = state.get("stack")
    if not isinstance(stack, list):
        stack = []
    order_sig = state.get("order_sig")
    if not isinstance(order_sig, str):
        order_sig = ""
    base_ids = state.get("base_ids")
    if not isinstance(base_ids, list):
        base_ids = []
    context_ids = state.get("context_ids")
    if not isinstance(context_ids, list):
        context_ids = []
    competition_order_tail = bool(state.get("competition_order_tail"))
    return {
        "stack": stack,
        "order_sig": order_sig,
        "base_ids": base_ids,
        "context_ids": context_ids,
        "competition_order_tail": competition_order_tail,
    }


def save_inscripcions_sort_context_state(
    request,
    context_key,
    stack=None,
    order_sig=None,
    base_ids=None,
    context_ids=None,
    competition_order_tail=None,
):
    if not context_key:
        return
    store = _read_sort_stack_store(request)
    state = store.get(context_key)
    if not isinstance(state, dict):
        state = {}
    if stack is not None:
        state["stack"] = stack
    if order_sig is not None:
        state["order_sig"] = order_sig
    if base_ids is not None:
        state["base_ids"] = base_ids
    if context_ids is not None:
        state["context_ids"] = context_ids
    if competition_order_tail is not None:
        state["competition_order_tail"] = bool(competition_order_tail)
    if not state.get("stack"):
        store.pop(context_key, None)
    else:
        store[context_key] = state
    _write_sort_stack_store(request, store)


def append_inscripcions_sort_stack_entry(request, context_key, entry, order_sig="", base_ids=None):
    state = get_inscripcions_sort_context_state(request, context_key)
    stack = list(state.get("stack") or [])
    stack.append(entry)
    # Evita creixements infinits en sessions llargues.
    if len(stack) > 20:
        stack = stack[-20:]
    save_inscripcions_sort_context_state(
        request,
        context_key,
        stack=stack,
        order_sig=order_sig,
        base_ids=base_ids if base_ids is not None else state.get("base_ids"),
    )
    return stack


def pop_inscripcions_sort_stack_entry(request, context_key, order_sig=None):
    state = get_inscripcions_sort_context_state(request, context_key)
    stack = list(state.get("stack") or [])
    popped = None
    if stack:
        popped = stack.pop()
    save_inscripcions_sort_context_state(
        request,
        context_key,
        stack=stack,
        order_sig=order_sig if order_sig is not None else state.get("order_sig", ""),
        base_ids=state.get("base_ids"),
    )
    return popped, stack


def clear_inscripcions_sort_context_state(request, context_key):
    save_inscripcions_sort_context_state(request, context_key, stack=[], order_sig="")


def reconcile_inscripcions_sort_context_state(request, context_key, current_ids, current_base_ids=None):
    """
    Reconciliació de l'stack d'ordenació per un context concret:
      - si l'ordre canvia però el conjunt d'IDs és el mateix -> rebase (manté stack)
      - si el conjunt d'IDs canvia -> clear (stack no fiable)
    """
    state = get_inscripcions_sort_context_state(request, context_key)
    stack = state.get("stack") if isinstance(state.get("stack"), list) else []
    if not stack:
        return {
            "stack": [],
            "order_sig": "",
            "base_ids": [],
            "context_ids": [],
            "competition_order_tail": False,
        }

    current_ids_list = list(current_ids or [])
    current_base_ids_list = list(current_base_ids or current_ids_list)
    current_sig = compute_inscripcions_order_signature_from_ids(current_base_ids_list)
    saved_sig = str(state.get("order_sig") or "")
    context_ids_state = state.get("context_ids") if isinstance(state.get("context_ids"), list) else []
    if not context_ids_state:
        context_ids_state = state.get("base_ids") if isinstance(state.get("base_ids"), list) else []

    # Si no hi ha signatura guardada o ja coincideix, no cal tocar res.
    if (not saved_sig) or saved_sig == current_sig:
        return {
            "stack": stack,
            "order_sig": saved_sig if saved_sig else current_sig,
            "base_ids": state.get("base_ids") if isinstance(state.get("base_ids"), list) else [],
            "context_ids": context_ids_state,
            "competition_order_tail": bool(state.get("competition_order_tail")),
        }

    same_set = (
        len(context_ids_state) == len(current_ids_list)
        and set(context_ids_state) == set(current_ids_list)
    )

    if same_set:
        # Rebase: mantenim stack i actualitzem referència base + signatura.
        save_inscripcions_sort_context_state(
            request,
            context_key,
            stack=stack,
            order_sig=current_sig,
            base_ids=current_base_ids_list,
            context_ids=current_ids_list,
            competition_order_tail=state.get("competition_order_tail"),
        )
        return {
            "stack": stack,
            "order_sig": current_sig,
            "base_ids": current_base_ids_list,
            "context_ids": current_ids_list,
            "competition_order_tail": bool(state.get("competition_order_tail")),
        }

    # Context realment diferent: neteja.
    clear_inscripcions_sort_context_state(request, context_key)
    return {
        "stack": [],
        "order_sig": "",
        "base_ids": [],
        "context_ids": [],
        "competition_order_tail": False,
    }


def _build_inscripcions_filtered_qs(competicio, filters):
    q = (filters.get("q") or "").strip()
    categoria = (filters.get("categoria") or "").strip()
    subcategoria = (filters.get("subcategoria") or "").strip()
    entitat = (filters.get("entitat") or "").strip()

    qs = Inscripcio.objects.filter(competicio=competicio)
    if subcategoria:
        qs = qs.filter(subcategoria__iexact=subcategoria)
    if entitat:
        qs = qs.filter(entitat__icontains=entitat)
    if q:
        qs = qs.filter(
            Q(nom_i_cognoms__icontains=q) |
            Q(document__icontains=q) |
            Q(entitat__icontains=q)
        )
    if categoria:
        qs = qs.filter(categoria__iexact=categoria)
    return qs


def _collect_sort_field_value_stats(records, sort_code):
    out = OrderedDict()
    context = _build_sort_field_runtime_context(records, sort_code)
    for obj in records:
        runtime = _resolve_sort_field_runtime(obj, sort_code, context=context)
        token = runtime.get("token") or ""
        if not token:
            continue
        key = _custom_sort_token_key(token)
        if not key:
            continue
        row = out.get(key)
        if row is None:
            row = {
                "token": token,
                "label": runtime.get("label") or token,
                "count": 0,
                "sort_scalar": runtime.get("sort_scalar"),
            }
            out[key] = row
        row["count"] += 1
    return out


def _normalize_sort_criterion(raw, sort_codes, allowed_group_codes, fallback_group_by=None):
    if not isinstance(raw, dict):
        return None

    c_sort_key_raw = str(raw.get("sort_key") or "").strip()
    c_sort_key = LEGACY_SORT_KEY_MAP.get(c_sort_key_raw, c_sort_key_raw)
    if c_sort_key not in sort_codes:
        return None

    c_sort_dir = str(raw.get("sort_dir") or "asc").strip()
    if c_sort_dir not in ("asc", "desc", "arrow_asc", "arrow_desc", "custom"):
        c_sort_dir = "asc"

    c_scope = str(raw.get("scope") or "all").strip().lower()
    if c_scope not in ("all", "tab", "all_groups", "group"):
        c_scope = "all"

    c_group_num = None
    if c_scope == "group":
        try:
            c_group_num = int(raw.get("group_num"))
        except Exception:
            return None
        if c_group_num <= 0:
            return None

    c_group_by = _normalize_sort_group_by(
        raw.get("group_by"),
        allowed_group_codes,
        fallback_group_by=fallback_group_by or [],
    )

    return {
        "sort_key": c_sort_key,
        "sort_dir": c_sort_dir,
        "scope": c_scope,
        "group_num": c_group_num,
        "group_by": c_group_by,
    }


def _sort_criterion_identity(entry):
    if not isinstance(entry, dict):
        return ("", "all", None)
    sort_key = str(entry.get("sort_key") or "").strip()
    scope = str(entry.get("scope") or "all").strip().lower()
    if scope not in ("all", "tab", "all_groups", "group"):
        scope = "all"
    group_num = None
    if scope == "group":
        try:
            group_num = int(entry.get("group_num"))
        except Exception:
            group_num = None
    return (sort_key, scope, group_num)


def _upsert_sort_stack_entry_preserving_priority(stack, new_entry):
    """
    Afegeix o actualitza un criteri a l'stack mantenint la prioritat existent.
    Si el criteri ja existeix (mateixa identitat), es substitueix a la mateixa posicio.
    """
    if not isinstance(stack, list):
        stack = []
    if not isinstance(new_entry, dict):
        return list(stack)

    target_identity = _sort_criterion_identity(new_entry)
    out = []
    replaced = False
    for entry in stack:
        if _sort_criterion_identity(entry) == target_identity:
            if not replaced:
                out.append(new_entry)
                replaced = True
            continue
        out.append(entry)

    if not replaced:
        out.append(new_entry)
    return out


def _normalize_positive_group_num(value):
    try:
        group_num = int(value)
    except Exception:
        return None
    return group_num if group_num > 0 else None


def _collect_visible_group_nums(records):
    out = []
    seen = set()
    for record in records or []:
        group_num = _normalize_positive_group_num(getattr(record, "grup", None))
        if group_num is None or group_num in seen:
            continue
        seen.add(group_num)
        out.append(group_num)
    return out


def _stack_requires_full_groups(stack):
    for criterion in stack or []:
        scope = str((criterion or {}).get("scope") or "all").strip().lower()
        if scope in ("group", "all_groups"):
            return True
    return False


def _collect_stack_target_group_nums(stack, visible_group_nums):
    target_group_nums = set()
    visible_group_set = set(visible_group_nums or [])
    for criterion in stack or []:
        scope = str((criterion or {}).get("scope") or "all").strip().lower()
        if scope == "all_groups":
            target_group_nums.update(visible_group_set)
            continue
        if scope == "group":
            group_num = _normalize_positive_group_num((criterion or {}).get("group_num"))
            if group_num is not None:
                target_group_nums.add(group_num)
    return target_group_nums


def _build_sort_application_runtime(
    competicio,
    visible_records,
    sort_codes,
    stack,
    include_competition_order=False,
):
    visible_records = list(visible_records or [])
    visible_ids = [record.id for record in visible_records]
    visible_id_set = set(visible_ids)
    visible_group_nums = _collect_visible_group_nums(visible_records)

    sort_codes = [str(code or "").strip() for code in (sort_codes or []) if str(code or "").strip()]
    if "grup" not in sort_codes:
        sort_codes = sort_codes + ["grup"]

    record_by_id = {record.id: record for record in visible_records}
    application_ids = list(visible_ids)

    if _stack_requires_full_groups(stack):
        target_group_nums = _collect_stack_target_group_nums(stack, visible_group_nums)
        if target_group_nums:
            group_records = list(
                _build_sort_records_queryset(
                    Inscripcio.objects.filter(competicio=competicio, grup__in=sorted(target_group_nums)),
                    sort_codes,
                    include_competition_order=include_competition_order,
                )
            )
            for record in group_records:
                record_by_id[record.id] = record

            global_ids = list(
                Inscripcio.objects
                .filter(competicio=competicio)
                .order_by("ordre_sortida", "id")
                .values_list("id", flat=True)
            )
            application_ids = [ins_id for ins_id in global_ids if ins_id in record_by_id]

    return {
        "records": [record_by_id[ins_id] for ins_id in application_ids if ins_id in record_by_id],
        "id_to_record": record_by_id,
        "application_ids": application_ids,
        "visible_ids": visible_id_set,
        "visible_group_nums": visible_group_nums,
    }


def _apply_single_sort_criterion(ids_in_order, id_to_record, criterion, competicio, runtime=None):
    seq_records = [id_to_record[i] for i in ids_in_order if i in id_to_record]
    if not seq_records:
        return list(ids_in_order)

    c_sort_key = criterion["sort_key"]
    c_sort_dir = criterion["sort_dir"]
    c_scope = criterion["scope"]
    c_group_num = criterion["group_num"]
    c_group_by = list(criterion.get("group_by") or [])
    runtime = runtime if isinstance(runtime, dict) else {}
    visible_ids = runtime.get("visible_ids")
    if not isinstance(visible_ids, set):
        visible_ids = set(ids_in_order)
    visible_group_nums = runtime.get("visible_group_nums")
    if not isinstance(visible_group_nums, list):
        visible_group_nums = _collect_visible_group_nums(seq_records)
    visible_group_set = set(visible_group_nums)

    c_desc = c_sort_dir in ("desc", "arrow_desc")
    c_arrow = c_sort_dir in ("arrow_asc", "arrow_desc")
    c_custom = c_sort_dir == "custom"
    custom_rank_map = get_competicio_custom_sort_rank_map(competicio, c_sort_key) if c_custom else {}

    def _ordered_subset(subset_records):
        ordered = sort_records_by_field_stable(
            subset_records,
            c_sort_key,
            descending=c_desc,
            custom_rank_map=custom_rank_map,
        )
        if not c_arrow:
            return ordered
        n = len(ordered)
        pos = arrow_positions(n)
        placed = [None] * n
        for i, obj in enumerate(ordered):
            placed[pos[i]] = obj
        return placed

    def _replace_subset(ids_base, subset_records):
        ids_out = list(ids_base)
        id_to_index = {rid: idx for idx, rid in enumerate(ids_out)}
        target_idxs = [id_to_index[r.id] for r in subset_records if r.id in id_to_index]
        ordered_target = _ordered_subset(subset_records)
        for idx, obj in zip(target_idxs, ordered_target):
            ids_out[idx] = obj.id
        return ids_out

    if c_scope == "all":
        target_records = [r for r in seq_records if r.id in visible_ids]
        return _replace_subset(ids_in_order, target_records)

    if c_scope == "all_groups":
        ids_out = list(ids_in_order)
        id_to_index = {rid: idx for idx, rid in enumerate(ids_out)}
        group_to_records = OrderedDict()
        for record in seq_records:
            group_num = _normalize_positive_group_num(getattr(record, "grup", None))
            if group_num is None or group_num not in visible_group_set:
                continue
            group_to_records.setdefault(group_num, []).append(record)

        for group_records in group_to_records.values():
            target_idxs = [id_to_index[r.id] for r in group_records if r.id in id_to_index]
            ordered_target = _ordered_subset(group_records)
            for idx, obj in zip(target_idxs, ordered_target):
                ids_out[idx] = obj.id
        return ids_out

    if c_scope == "group":
        target_records = [
            r for r in seq_records
            if _normalize_positive_group_num(getattr(r, "grup", None)) == c_group_num
        ]
        return _replace_subset(ids_in_order, target_records)

    # c_scope == "tab"
    visible_seq_records = [r for r in seq_records if r.id in visible_ids]
    if not c_group_by:
        return _replace_subset(ids_in_order, visible_seq_records)

    grouping_sig = "|".join(c_group_by)
    merges = (competicio.tab_merges or {}).get(grouping_sig, [])
    merge_map = {}
    for group_keys in merges:
        t = tuple(group_keys)
        for x in group_keys:
            merge_map[x] = t

    ids_out = list(ids_in_order)
    id_to_index = {rid: idx for idx, rid in enumerate(ids_out)}
    tab_to_records = OrderedDict()
    for r in visible_seq_records:
        vals = [_norm_val(get_inscripcio_value(r, code)) for code in c_group_by]
        simple = json.dumps(vals, ensure_ascii=False)
        mid = merge_map.get(simple)
        tab_key = json.dumps(list(mid), ensure_ascii=False) if mid else simple
        tab_to_records.setdefault(tab_key, []).append(r)

    for _tab_key, tab_records in tab_to_records.items():
        tab_idxs = [id_to_index[r.id] for r in tab_records if r.id in id_to_index]
        ordered_tab = _ordered_subset(tab_records)
        for i, obj in zip(tab_idxs, ordered_tab):
            ids_out[i] = obj.id

    return ids_out


def _apply_sort_stack(ids_base, id_to_record, stack, competicio, runtime=None):
    final_ids = list(ids_base)
    # Primer criteri = mes important, per tant s'aplica l'stack en ordre invers.
    for criterion in reversed(stack):
        final_ids = _apply_single_sort_criterion(final_ids, id_to_record, criterion, competicio, runtime=runtime)
    return final_ids


def _competition_order_group_num(record):
    group_num = _normalize_positive_group_num(getattr(record, "grup", None))
    if group_num is not None:
        return group_num
    group = getattr(record, "grup_competicio", None)
    return normalize_positive_int(getattr(group, "display_num", None))


def _apply_competition_order_tail(ids_in_order, id_to_record, runtime=None):
    seq_records = [id_to_record[i] for i in ids_in_order if i in id_to_record]
    if not seq_records:
        return list(ids_in_order)

    ids_out = list(ids_in_order)
    id_to_index = {rid: idx for idx, rid in enumerate(ids_out)}
    group_to_records = OrderedDict()
    for record in seq_records:
        group_num = _competition_order_group_num(record)
        if group_num is None:
            continue
        group_to_records.setdefault(group_num, []).append(record)

    for group_records in group_to_records.values():
        with_comp_order = []
        without_comp_order = []
        for pos, record in enumerate(group_records):
            comp_order = normalize_positive_int(getattr(record, "ordre_competicio", None))
            if comp_order is None:
                without_comp_order.append(record)
            else:
                with_comp_order.append((comp_order, pos, record))

        with_comp_order.sort(key=lambda item: (item[0], item[1]))
        ordered_records = [record for _comp_order, _pos, record in with_comp_order] + without_comp_order
        target_idxs = [id_to_index[record.id] for record in group_records if record.id in id_to_index]
        for idx, record in zip(target_idxs, ordered_records):
            ids_out[idx] = record.id

    return ids_out


def _apply_sort_pipeline(ids_base, id_to_record, stack, competicio, runtime=None, competition_order_tail=False):
    final_ids = list(ids_base)
    if competition_order_tail:
        final_ids = _apply_competition_order_tail(final_ids, id_to_record, runtime=runtime)
    final_ids = _apply_sort_stack(final_ids, id_to_record, stack, competicio, runtime=runtime)
    return final_ids


def _persist_inscripcions_order_from_ids(id_to_record, final_ids):
    updates = []
    for idx, ins_id in enumerate(final_ids, start=1):
        obj = id_to_record.get(ins_id)
        if not obj:
            continue
        if obj.ordre_sortida != idx:
            obj.ordre_sortida = idx
            updates.append(obj)
    if updates:
        with transaction.atomic():
            Inscripcio.objects.bulk_update(updates, ["ordre_sortida"], batch_size=500)
    return len(updates)


def _normalize_competition_order_tail_flag(raw_value):
    if isinstance(raw_value, bool):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return bool(raw_value)
    token = str(raw_value or "").strip().lower()
    return token in {"1", "true", "yes", "on"}


def _extract_sort_partition_codes(stack):
    out = []
    seen = set()
    for criterion in stack:
        if not isinstance(criterion, dict):
            continue
        scope = str(criterion.get("scope") or "all").strip().lower()
        if scope not in ("all", "tab"):
            continue
        code = str(criterion.get("sort_key") or "").strip()
        if not code or code in seen:
            continue
        seen.add(code)
        out.append(code)
    return out


def _build_sort_partition_buckets(records, partition_codes):
    buckets = OrderedDict()
    for r in records:
        raw_vals = [get_inscripcio_value(r, code) for code in partition_codes]
        norm_vals = [_norm_val(v) for v in raw_vals]
        bucket_key = json.dumps(norm_vals, ensure_ascii=False)
        bucket = buckets.get(bucket_key)
        if bucket is None:
            label_parts = [_s(v) for v in raw_vals]
            bucket = {
                "key": bucket_key,
                "label": " · ".join(label_parts) if label_parts else "(Sense valor)",
                "count": 0,
                "ids": [],
            }
            buckets[bucket_key] = bucket
        bucket["count"] += 1
        bucket["ids"].append(r.id)
    return list(buckets.values())


def _build_tabs_partition_buckets(competicio, records, group_codes):
    if not group_codes:
        return []

    grouping_sig = "|".join(group_codes)
    merges = (competicio.tab_merges or {}).get(grouping_sig, [])
    merge_map = {}
    for group_keys in merges:
        t = tuple(group_keys)
        for x in group_keys:
            merge_map[x] = t

    def _simple_label_from_obj(obj):
        parts = [_s(get_inscripcio_value(obj, code)) for code in group_codes]
        return " · ".join(parts) if parts else "(Sense valor)"

    simple_label_map = {}
    simple_values_by_id = {}
    for r in records:
        vals = [_norm_val(get_inscripcio_value(r, code)) for code in group_codes]
        simple = json.dumps(vals, ensure_ascii=False)
        simple_values_by_id[r.id] = simple
        if simple not in simple_label_map:
            simple_label_map[simple] = _simple_label_from_obj(r)

    tab_to_ids = OrderedDict()
    tab_label_map = {}

    for r in records:
        simple = simple_values_by_id.get(r.id, "")

        merged_tuple = merge_map.get(simple)
        tab_key = json.dumps(list(merged_tuple), ensure_ascii=False) if merged_tuple else simple
        tab_to_ids.setdefault(tab_key, []).append(r.id)

        if tab_key in tab_label_map:
            continue
        if not merged_tuple:
            tab_label_map[tab_key] = simple_label_map.get(simple, simple)
            continue

        parts = []
        for sk in merged_tuple:
            p = simple_label_map.get(sk, sk)
            if p not in parts:
                parts.append(p)
        tab_label_map[tab_key] = " + ".join(parts) if parts else tab_key

    out = []
    for tab_key, ids in tab_to_ids.items():
        out.append(
            {
                "key": tab_key,
                "label": tab_label_map.get(tab_key, tab_key),
                "count": len(ids),
                "ids": ids,
            }
        )
    return out


def _balanced_sizes(n, k):
    if n <= 0 or k <= 0:
        return []
    k = min(k, n)
    base = n // k
    rem = n % k
    return [base + (1 if i < rem else 0) for i in range(k)]


def _fixed_sizes(n, size):
    if n <= 0 or size <= 0:
        return []
    out = []
    remaining = n
    while remaining > 0:
        take = size if remaining >= size else remaining
        out.append(take)
        remaining -= take
    return out


def _assign_group_sizes_in_order(objs, sizes, start_group_num):
    idx = 0
    g = start_group_num
    for sz in sizes:
        if sz <= 0:
            continue
        g += 1
        for _ in range(sz):
            if idx >= len(objs):
                break
            objs[idx].grup = g
            idx += 1
    return g


GROUP_NAME_SUGGESTION_IGNORED_LABELS = {
    "",
    "Sense bloc",
    "Totes les inscripcions filtrades",
    "(Sense valor)",
}


def _clean_group_suggestion_label(raw_label):
    label = str(raw_label or "").strip()
    if not label or label in GROUP_NAME_SUGGESTION_IGNORED_LABELS:
        return ""
    return label


def _build_group_suggested_name(sources):
    ranked_by_label = OrderedDict()
    for idx, row in enumerate(sources or []):
        if not isinstance(row, dict):
            continue
        try:
            count = int(row.get("count") or 0)
        except Exception:
            count = 0
        count = max(0, count)

        labels_by_kind = row.get("labels_by_kind")
        emitted = False
        if isinstance(labels_by_kind, dict):
            for kind_priority, kind in enumerate(("tabs", "sort", "")):
                raw_labels = labels_by_kind.get(kind) or []
                if not isinstance(raw_labels, list):
                    continue
                for label_order, raw_label in enumerate(raw_labels):
                    label = _clean_group_suggestion_label(raw_label)
                    if not label:
                        continue
                    emitted = True
                    key = label.casefold()
                    meta = ranked_by_label.get(key)
                    if meta is None:
                        ranked_by_label[key] = {
                            "label": label,
                            "count": count,
                            "kind_priority": kind_priority,
                            "row_idx": idx,
                            "label_order": label_order,
                        }
                        continue
                    meta["count"] += count
                    meta["kind_priority"] = min(meta["kind_priority"], kind_priority)
                    if (idx, label_order) < (meta["row_idx"], meta["label_order"]):
                        meta["row_idx"] = idx
                        meta["label_order"] = label_order

        if emitted:
            continue

        label = _clean_group_suggestion_label(row.get("label"))
        if not label:
            continue
        key = label.casefold()
        meta = ranked_by_label.get(key)
        if meta is None:
            ranked_by_label[key] = {
                "label": label,
                "count": count,
                "kind_priority": 99,
                "row_idx": idx,
                "label_order": 0,
            }
            continue
        meta["count"] += count
        if idx < meta["row_idx"]:
            meta["row_idx"] = idx

    if not ranked_by_label:
        return ""

    ranked = list(ranked_by_label.values())
    ranked.sort(
        key=lambda item: (
            item["kind_priority"],
            -item["count"],
            item["row_idx"],
            item["label_order"],
            item["label"].casefold(),
        )
    )
    ordered_labels = [item["label"] for item in ranked]
    return " + ".join(ordered_labels)


def _apply_group_suggested_names(preview_groups):
    out = list(preview_groups or [])
    grouped_indexes = OrderedDict()

    for idx, row in enumerate(out):
        suggested_name = _build_group_suggested_name(row.get("sources"))
        row["suggested_name"] = suggested_name
        if not suggested_name:
            continue
        grouped_indexes.setdefault(suggested_name.casefold(), []).append((idx, suggested_name))

    for entries in grouped_indexes.values():
        if len(entries) <= 1:
            continue
        for order_idx, (group_idx, base_name) in enumerate(entries, start=1):
            out[group_idx]["suggested_name"] = f"{base_name} ({order_idx})"

    return out


def _normalize_bucket_source_entries(raw_sources):
    if isinstance(raw_sources, dict):
        raw_sources = [raw_sources]
    out = []
    seen = set()
    for row in raw_sources or []:
        if not isinstance(row, dict):
            continue
        kind = str(row.get("kind") or "").strip().lower()
        key = str(row.get("key") or "").strip()
        label = str(row.get("label") or "").strip()
        if not label:
            continue
        ident = (kind, key, label)
        if ident in seen:
            continue
        seen.add(ident)
        out.append(
            {
                "kind": kind,
                "key": key,
                "label": label,
            }
        )
    return out


def _bucket_labels_by_kind(sources):
    labels_by_kind = OrderedDict()
    for source in _normalize_bucket_source_entries(sources):
        kind = source["kind"]
        label = str(source.get("label") or "").strip()
        if not label:
            continue
        labels = labels_by_kind.setdefault(kind, [])
        if label not in labels:
            labels.append(label)
    return labels_by_kind


def _bucket_source_signature(sources):
    normalized = _normalize_bucket_source_entries(sources)
    return tuple((row["kind"], row["key"], row["label"]) for row in normalized)


def _build_bucket_source_label(sources):
    labels = []
    for source in _normalize_bucket_source_entries(sources):
        label = str(source.get("label") or "").strip()
        if label and label not in labels:
            labels.append(label)
    return " / ".join(labels) if labels else "Sense bloc"


def _build_bucket_source_kinds(sources):
    kinds = []
    for source in _normalize_bucket_source_entries(sources):
        kind = str(source.get("kind") or "").strip().lower()
        if kind and kind not in kinds:
            kinds.append(kind)
    return kinds


def _build_bucket_sources_by_id(buckets):
    out = {}
    for bucket in buckets or []:
        sources = _normalize_bucket_source_entries(bucket.get("sources"))
        for ins_id in bucket.get("ids") or []:
            out[ins_id] = list(sources)
    return out


def _build_partition_signature_set(buckets):
    return {
        tuple(sorted(int(ins_id) for ins_id in (bucket.get("ids") or [])))
        for bucket in (buckets or [])
        if bucket.get("ids")
    }


def _partitions_are_equivalent(buckets_a, buckets_b):
    return _build_partition_signature_set(buckets_a) == _build_partition_signature_set(buckets_b)


def _decorate_buckets_with_sources(buckets, kind):
    out = []
    for bucket in buckets or []:
        label = str(bucket.get("label") or "Sense bloc")
        key = str(bucket.get("key") or "")
        out.append(
            {
                **bucket,
                "label": label,
                "sources": [
                    {
                        "kind": kind,
                        "key": key or label,
                        "label": label,
                    }
                ],
            }
        )
    return out


def _build_combined_partition_buckets(records, layers):
    buckets = OrderedDict()
    for record in records or []:
        raw_sources = []
        for layer in layers or []:
            source_row = layer.get("by_id", {}).get(record.id)
            if not isinstance(source_row, dict):
                continue
            raw_sources.append(
                {
                    "kind": str(source_row.get("kind") or "").strip().lower(),
                    "key": str(source_row.get("key") or "").strip(),
                    "label": str(source_row.get("label") or "").strip(),
                }
            )

        sources = _normalize_bucket_source_entries(raw_sources)
        if not sources:
            continue

        bucket_key = json.dumps(
            [{"kind": row["kind"], "key": row["key"]} for row in sources],
            ensure_ascii=False,
        )
        bucket = buckets.get(bucket_key)
        if bucket is None:
            bucket = {
                "key": bucket_key,
                "label": _build_bucket_source_label(sources),
                "count": 0,
                "ids": [],
                "sources": sources,
            }
            buckets[bucket_key] = bucket
        bucket["count"] += 1
        bucket["ids"].append(record.id)
    return list(buckets.values())


def _resolve_group_creation_buckets(
    competicio,
    records,
    *,
    group_codes=None,
    partition_codes=None,
    fallback_mode="all_filtered",
):
    group_codes = list(group_codes or [])
    partition_codes = list(partition_codes or [])

    tabs_buckets = _build_tabs_partition_buckets(competicio, records, group_codes) if group_codes else []
    sort_buckets = _build_sort_partition_buckets(records, partition_codes) if partition_codes else []

    tabs_buckets = _decorate_buckets_with_sources(tabs_buckets, "tabs")
    sort_buckets = _decorate_buckets_with_sources(sort_buckets, "sort")

    if tabs_buckets and sort_buckets and _partitions_are_equivalent(tabs_buckets, sort_buckets):
        sort_buckets = []

    layers_used = []
    if tabs_buckets:
        layers_used.append("tabs")
    if sort_buckets:
        layers_used.append("sort")

    if not layers_used:
        if fallback_mode == "strict":
            return {
                "ok": False,
                "error": "No hi ha criteris resolubles per construir blocs d'origen",
                "buckets": [],
                "layers_used": [],
                "used_fallback": False,
                "fallback_reason": "no_resolvable_criteria",
            }
        return {
            "ok": True,
            "buckets": [
                {
                    "key": "__ALL_FILTERED__",
                    "label": "Totes les inscripcions filtrades",
                    "count": len(records),
                    "ids": [r.id for r in records],
                    "sources": [
                        {
                            "kind": "fallback",
                            "key": "__ALL_FILTERED__",
                            "label": "Totes les inscripcions filtrades",
                        }
                    ],
                }
            ],
            "layers_used": [],
            "used_fallback": True,
            "fallback_reason": "no_resolvable_criteria_used_all_filtered",
        }

    if len(layers_used) == 1:
        buckets = tabs_buckets if tabs_buckets else sort_buckets
    else:
        layers = []
        if tabs_buckets:
            layers.append(
                {
                    "kind": "tabs",
                    "by_id": {
                        ins_id: bucket["sources"][0]
                        for bucket in tabs_buckets
                        for ins_id in (bucket.get("ids") or [])
                    },
                }
            )
        if sort_buckets:
            layers.append(
                {
                    "kind": "sort",
                    "by_id": {
                        ins_id: bucket["sources"][0]
                        for bucket in sort_buckets
                        for ins_id in (bucket.get("ids") or [])
                    },
                }
            )
        buckets = _build_combined_partition_buckets(records, layers)

    return {
        "ok": True,
        "buckets": buckets,
        "layers_used": layers_used,
        "used_fallback": False,
        "fallback_reason": "",
    }


def _build_group_creation_preview(objs, sizes, start_group_num, bucket_sources_by_id=None):
    bucket_sources_by_id = bucket_sources_by_id or {}
    out = []
    idx = 0
    next_group_num = start_group_num

    for sz in sizes:
        if sz <= 0:
            continue

        members = list(objs[idx:idx + sz])
        idx += sz
        if not members:
            continue

        next_group_num += 1
        source_counts = OrderedDict()
        for obj in members:
            sources = bucket_sources_by_id.get(obj.id) or []
            source_key = _bucket_source_signature(sources)
            row = source_counts.get(source_key)
            if row is None:
                row = {
                    "label": _build_bucket_source_label(sources),
                    "count": 0,
                    "kinds": _build_bucket_source_kinds(sources),
                    "labels_by_kind": dict(_bucket_labels_by_kind(sources)),
                }
                source_counts[source_key] = row
            row["count"] += 1

        member_names = [str(getattr(obj, "nom_i_cognoms", "") or "").strip() for obj in members]
        member_names = [name for name in member_names if name]

        out.append(
            {
                "preview_kind": "created",
                "impact_kind": "created",
                "group_num": next_group_num,
                "members_count": len(members),
                "sources": list(source_counts.values()),
                "member_names_preview": member_names[:4],
                "member_names_remaining": max(0, len(member_names) - 4),
            }
        )

    return _apply_group_suggested_names(out)


def _build_existing_groups_preview(competicio, records, bucket_sources_by_id=None, moving_ids=None):
    bucket_sources_by_id = bucket_sources_by_id or {}
    moving_ids = set(moving_ids or [])
    groups_by_display_num = (get_group_maps(competicio).get("by_display_num") or {})
    visible_group_nums = []
    seen_group_nums = set()
    for record in records:
        group_num = normalize_positive_int(getattr(record, "grup", None))
        if not group_num or group_num in seen_group_nums:
            continue
        seen_group_nums.add(group_num)
        visible_group_nums.append(group_num)

    if not visible_group_nums:
        return []

    grouped = OrderedDict((group_num, []) for group_num in sorted(visible_group_nums))
    all_group_members = (
        Inscripcio.objects
        .filter(competicio=competicio, grup__in=visible_group_nums)
        .order_by("grup", "ordre_sortida", "id")
        .only("id", "grup", "nom_i_cognoms", "ordre_sortida")
    )
    for record in all_group_members:
        group_num = normalize_positive_int(getattr(record, "grup", None))
        if not group_num:
            continue
        grouped.setdefault(group_num, []).append(record)

    out = []
    for group_num in sorted(grouped.keys()):
        members = grouped.get(group_num) or []
        group_obj = groups_by_display_num.get(group_num)
        source_counts = OrderedDict()
        moving_members = []
        for obj in members:
            sources = bucket_sources_by_id.get(obj.id) or []
            source_key = _bucket_source_signature(sources)
            row = source_counts.get(source_key)
            if row is None:
                row = {
                    "label": (
                        _build_bucket_source_label(sources)
                        if sources else "Fora del filtre actual"
                    ),
                    "count": 0,
                    "moving_count": 0,
                    "remaining_count": 0,
                    "kinds": _build_bucket_source_kinds(sources),
                    "labels_by_kind": dict(_bucket_labels_by_kind(sources)),
                }
                source_counts[source_key] = row
            row["count"] += 1
            if obj.id in moving_ids:
                row["moving_count"] += 1
                moving_members.append(obj)
            else:
                row["remaining_count"] += 1

        member_names = [str(getattr(obj, "nom_i_cognoms", "") or "").strip() for obj in members]
        member_names = [name for name in member_names if name]
        moving_member_names = [str(getattr(obj, "nom_i_cognoms", "") or "").strip() for obj in moving_members]
        moving_member_names = [name for name in moving_member_names if name]
        moving_members_count = len(moving_members)
        remaining_members_count = max(0, len(members) - moving_members_count)
        impact_kind = "unchanged"
        if moving_members_count > 0 and remaining_members_count <= 0:
            impact_kind = "removed"
        elif moving_members_count > 0:
            impact_kind = "reduced"

        out.append(
            {
                "preview_kind": "existing",
                "impact_kind": impact_kind,
                "group_num": group_num,
                "group_label": group_label(group_obj),
                "members_count": len(members),
                "moving_members_count": moving_members_count,
                "remaining_members_count": remaining_members_count,
                "sources": list(source_counts.values()),
                "member_names_preview": member_names[:4],
                "member_names_remaining": max(0, len(member_names) - 4),
                "moving_member_names_preview": moving_member_names[:4],
                "moving_member_names_remaining": max(0, len(moving_member_names) - 4),
            }
        )

    return out


def _range_k_bounds(n, min_size, max_size):
    if n <= 0 or min_size <= 0 or max_size <= 0 or min_size > max_size:
        return (1, 0)  # rang invalid
    k_min = math.ceil(n / max_size)
    k_max = math.floor(n / min_size)
    return (k_min, k_max)


def _clamp(v, lo, hi):
    return max(lo, min(hi, v))


def _parse_fallback_mode(raw):
    mode = str(raw or "all_filtered").strip().lower()
    if mode not in ("all_filtered", "strict", "adjust_k", "ignore_range"):
        mode = "all_filtered"
    return mode


def _resolve_k_for_range(n, min_size, max_size, preferred_k=None, fallback_mode="strict"):
    """
    Retorna (k, meta) on meta inclou si s'ha usat fallback.
    """
    k_min, k_max = _range_k_bounds(n, min_size, max_size)
    meta = {"used_fallback": False, "fallback_reason": ""}
    feasible = k_min <= k_max and k_max >= 1

    if feasible:
        if preferred_k is None:
            # Tria equilibrada cap al centre del rang viable.
            k_target = int(round(n / ((min_size + max_size) / 2.0)))
            k = _clamp(k_target, k_min, k_max)
        else:
            k = int(preferred_k)
            if k < k_min or k > k_max:
                if fallback_mode == "strict":
                    return None, {"used_fallback": True, "fallback_reason": "k_out_of_range"}
                if fallback_mode in ("adjust_k", "all_filtered"):
                    k = _clamp(k, k_min, k_max)
                    meta["used_fallback"] = True
                    meta["fallback_reason"] = "k_adjusted_to_feasible_range"
                elif fallback_mode == "ignore_range":
                    k = _clamp(k, 1, max(1, n))
                    meta["used_fallback"] = True
                    meta["fallback_reason"] = "range_ignored_for_k"
        return k, meta

    # Rang min-max inviable per n.
    if fallback_mode == "strict":
        return None, {"used_fallback": True, "fallback_reason": "range_infeasible"}

    if preferred_k is not None:
        k = _clamp(int(preferred_k), 1, max(1, n))
        reason = "range_infeasible_k_kept"
    else:
        # Degrada a una k coherent amb max_size.
        k = _clamp(math.ceil(n / max_size) if max_size > 0 else 1, 1, max(1, n))
        reason = "range_infeasible_auto_k"

    return k, {"used_fallback": True, "fallback_reason": reason}

def assign_groups_balanced(objs, size, start_group_num):
    """
    Reparteix objs en k grups, on k = ceil(n/size),
    i distribueix mides perquè difereixin com a màxim 1.
    Retorna el nou start_group_num (últim grup assignat).
    """
    n = len(objs)
    if n == 0:
        return start_group_num

    k = math.ceil(n / size)           # nombre de grups
    base = n // k                     # mida base de cada grup
    rem = n % k                       # els primers 'rem' grups tindran +1

    idx = 0
    group_num = start_group_num
    for g in range(k):
        group_num += 1
        this_size = base + (1 if g < rem else 0)
        for _ in range(this_size):
            objs[idx].grup = group_num
            idx += 1

    return group_num


def _normalize_group_names_map(raw_group_names):
    """
    Normalitza el mapa de labels de grup:
    - claus numeriques positives
    - valors text no buit
    """
    out = {}
    if not isinstance(raw_group_names, dict):
        return out

    for raw_group, raw_label in raw_group_names.items():
        try:
            group_num = int(str(raw_group).strip())
        except Exception:
            continue
        if group_num <= 0:
            continue

        label = str(raw_label or "").strip()
        if not label:
            continue
        out[group_num] = label
    return out


def _sync_group_names_for_competicio(competicio, mapping=None):
    sync_competicio_group_names_view(competicio)


def _persist_group_suggested_names(competicio, preview_groups):
    suggested_by_display_num = {}
    for row in preview_groups or []:
        group_num = normalize_positive_int(row.get("group_num"))
        suggested_name = str(row.get("suggested_name") or "").strip()
        if not group_num or not suggested_name:
            continue
        suggested_by_display_num[group_num] = suggested_name

    if not suggested_by_display_num:
        sync_competicio_group_names_view(competicio)
        return 0

    groups = list(
        GrupCompeticio.objects.filter(
            competicio=competicio,
            display_num__in=list(suggested_by_display_num.keys()),
        )
    )
    updates = []
    for group in groups:
        suggested_name = suggested_by_display_num.get(group.display_num, "")
        if not suggested_name or group.nom == suggested_name:
            continue
        group.nom = suggested_name
        updates.append(group)

    if updates:
        GrupCompeticio.objects.bulk_update(updates, ["nom"], batch_size=200)

    sync_competicio_group_names_view(competicio)
    return len(updates)


def renumber_groups_for_competicio(competicio):
    """
    Compatibilitat temporal: la numeracio visible dels grups ja no es recalcula automaticament.
    """
    _sync_group_names_for_competicio(competicio)


def competicio_has_rotacions(competicio):
    return RotacioAssignacio.objects.filter(competicio=competicio).exists()


def _resolve_group_id_for_inscripcio(inscripcio, groups_by_display_num):
    group_id = getattr(inscripcio, "grup_competicio_id", None)
    if group_id:
        return int(group_id)
    legacy_group_num = normalize_positive_int(getattr(inscripcio, "grup", None))
    if not legacy_group_num:
        return None
    group = groups_by_display_num.get(legacy_group_num)
    if not group:
        return None
    return int(group.id)


def _programmed_groups_emptied_by_move(competicio, target_ids):
    target_ids = {
        int(ins_id)
        for ins_id in (target_ids or [])
        if normalize_positive_int(ins_id)
    }
    if not target_ids:
        return []

    programmed_group_ids = get_programmed_group_ids(competicio)
    if not programmed_group_ids:
        return []

    group_maps = get_group_maps(competicio)
    groups_by_display_num = group_maps["by_display_num"]
    groups_by_id = group_maps["by_id"]
    selected = list(
        Inscripcio.objects
        .filter(competicio=competicio, id__in=target_ids)
        .select_related("grup_competicio")
        .only("id", "grup", "grup_competicio")
    )
    moving_counts = defaultdict(int)
    for inscripcio in selected:
        group_id = _resolve_group_id_for_inscripcio(inscripcio, groups_by_display_num)
        if group_id:
            moving_counts[group_id] += 1
    if not moving_counts:
        return []

    total_counts = defaultdict(int)
    all_members = (
        Inscripcio.objects
        .filter(competicio=competicio)
        .select_related("grup_competicio")
        .only("id", "grup", "grup_competicio")
    )
    for inscripcio in all_members:
        group_id = _resolve_group_id_for_inscripcio(inscripcio, groups_by_display_num)
        if group_id:
            total_counts[group_id] += 1

    blocked = []
    for group_id, moved_count in moving_counts.items():
        if group_id not in programmed_group_ids:
            continue
        if moved_count >= total_counts.get(group_id, 0):
            group = groups_by_id.get(group_id)
            if group is not None:
                blocked.append(group)
    blocked.sort(key=lambda group: (group.display_num, group.id))
    return blocked


def _message_for_emptied_programmed_groups(groups):
    if not groups:
        return ""
    labels = []
    for group in groups:
        label = str(getattr(group, "nom", "") or "").strip()
        if not label:
            label = f"Grup {getattr(group, 'display_num', '?')}"
        labels.append(label)
    joined = ", ".join(labels)
    return f"No es pot deixar buit un grup inclos al programa de rotacions: {joined}."


def sync_stable_groups_from_legacy(competicio):
    view_cfg = competicio.inscripcions_view or {}
    raw_group_names = view_cfg.get("group_names") or {}
    group_names = raw_group_names if isinstance(raw_group_names, dict) else {}
    live_display_nums = list(
        Inscripcio.objects
        .filter(competicio=competicio, grup__isnull=False)
        .order_by("grup")
        .values_list("grup", flat=True)
        .distinct()
    )
    live_display_nums = [int(num) for num in live_display_nums if isinstance(num, int) and num > 0]

    group_maps = get_group_maps(competicio)
    groups_by_display = group_maps["by_display_num"]
    touched_group_ids = set()
    updates = []
    with transaction.atomic():
        for display_num in live_display_nums:
            group = groups_by_display.get(display_num)
            name = str(group_names.get(str(display_num)) or "").strip()
            if group is None:
                group = ensure_group_for_display_num(competicio, display_num, name=name)
                groups_by_display[display_num] = group
            else:
                fields = []
                if not group.actiu:
                    group.actiu = True
                    fields.append("actiu")
                if name != group.nom:
                    group.nom = name
                    fields.append("nom")
                if fields:
                    group.save(update_fields=fields)
            touched_group_ids.add(group.id)

        stale_groups = [
            group for group in get_group_maps(competicio)["groups"]
            if group.id not in touched_group_ids
        ]
        if stale_groups:
            GrupCompeticio.objects.filter(id__in=[g.id for g in stale_groups]).update(actiu=False)

        counters = defaultdict(int)
        qs = (
            Inscripcio.objects
            .filter(competicio=competicio)
            .order_by("grup", "ordre_sortida", "id")
            .only("id", "grup", "grup_competicio", "ordre_competicio")
        )
        for inscripcio in qs:
            display_num = getattr(inscripcio, "grup", None)
            group = groups_by_display.get(display_num)
            next_group_id = getattr(group, "id", None)
            next_comp_order = None
            if next_group_id:
                counters[next_group_id] += 1
                next_comp_order = counters[next_group_id]
            if (
                inscripcio.grup_competicio_id != next_group_id
                or inscripcio.ordre_competicio != next_comp_order
            ):
                inscripcio.grup_competicio_id = next_group_id
                inscripcio.ordre_competicio = next_comp_order
                updates.append(inscripcio)
        if updates:
            Inscripcio.objects.bulk_update(
                updates,
                ["grup_competicio", "ordre_competicio"],
                batch_size=500,
            )

    sync_competicio_group_names_view(competicio)

def assign_groups_k(objs, k, start_group_num):
    """
    Reparteix objs en k grups (k fix), equilibrant mides (difereixen com a màxim 1).
    """
    n = len(objs)
    if n == 0 or k <= 0:
        return start_group_num

    k = min(k, n)  # no té sentit més grups que persones

    base = n // k
    rem = n % k

    idx = 0
    group_num = start_group_num
    for g in range(k):
        group_num += 1
        this_size = base + (1 if g < rem else 0)
        for _ in range(this_size):
            objs[idx].grup = group_num
            idx += 1

    return group_num

# ------------------------------------------------------------------------------------------------
#
#
#           CREACIÓ COMPETICIONS
#
#
# ------------------------------------------------------------------------------------------------




class CompeticioHomeView(TemplateView):
    template_name = "competicio/home.html"


class CompeticioCreateView(CreateView):
    model = Competicio
    form_class = CompeticioForm
    template_name = "competicio/competicio_form.html"
    success_url = reverse_lazy("created")

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.user.is_authenticated:
            membership, created = CompeticioMembership.objects.get_or_create(
                user=self.request.user,
                competicio=self.object,
                defaults={
                    "role": CompeticioMembership.Role.OWNER,
                    "is_active": True,
                    "granted_by": self.request.user,
                },
            )
            if not created:
                changed = False
                if membership.role != CompeticioMembership.Role.OWNER:
                    membership.role = CompeticioMembership.Role.OWNER
                    changed = True
                if not membership.is_active:
                    membership.is_active = True
                    changed = True
                if membership.granted_by_id is None:
                    membership.granted_by = self.request.user
                    changed = True
                if changed:
                    membership.save(update_fields=["role", "is_active", "granted_by", "updated_at"])
        return response


class CompeticioDeleteView(DeleteView):
    model = Competicio
    template_name = "competicio/competicio_confirm_delete.html"
    success_url = reverse_lazy("created")

class CompeticioListView(ListView):
    model = Competicio
    template_name = "competicio/competicio_created_list.html"
    context_object_name = "competicions"
    paginate_by = 20

    def get_queryset(self):
        user = self.request.user
        qs = Competicio.objects.all().order_by("-created_at", "-id")
        if user.is_superuser or user.groups.filter(name="platform_admin").exists():
            return qs
        return (
            qs.filter(
                memberships__user=user,
                memberships__is_active=True,
            )
            .distinct()
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        actions = {}
        for comp in ctx.get("competicions", []):
            actions[comp.id] = {
                "can_view_inscripcions": user_has_competicio_capability(
                    self.request.user, comp, "inscripcions.view"
                ),
                "can_view_rotacions": user_has_competicio_capability(
                    self.request.user, comp, "rotacions.view"
                ),
                "can_view_notes": user_has_competicio_capability(
                    self.request.user, comp, "scoring.view"
                ),
                "can_view_trampoli_config": user_has_competicio_capability(
                    self.request.user, comp, "scoring.edit"
                ),
                "can_delete_competicio": user_has_competicio_capability(
                    self.request.user, comp, "competition.delete"
                ),
            }
        ctx["comp_actions"] = actions
        return ctx



def notes_home_router(request, pk):
    c = get_object_or_404(Competicio, pk=pk)

    if c.tipus == Competicio.Tipus.TRAMPOLI:
        return redirect("scoring_notes_home", pk=pk)

    # futurs:
    # if c.tipus == Competicio.Tipus.NATACIO: return redirect(...)
    # ...

    return redirect("created")


# ------------------------------------------------------------------------------------------------
#
#
#           TRACTAMENT INSCRIPCIONS
#
#
# ------------------------------------------------------------------------------------------------

class InscripcionsImportExcelView(FormView):
    template_name = "competicio/inscripcions_import.html"
    form_class = ImportInscripcionsExcelForm

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = self.competicio
        return ctx

    def form_valid(self, form):
        fitxer = form.cleaned_data["fitxer"]
        sheet = form.cleaned_data.get("sheet") or ""

        result = importar_inscripcions_excel(fitxer, self.competicio, sheet)
        messages.success(
            self.request,
            f"Importació OK. Full: {result['full']} | Creats: {result['creats']} | "
            f"Actualitzats: {result['actualitzats']} | Ignorats: {result['ignorats']} | "
            f"Ambiguos: {result.get('ambiguos', 0)}"
        )
        warnings = result.get("warnings") or []
        if warnings:
            parts = []
            for w in warnings:
                code = str(w.get("code") or "").strip()
                remapped = str(w.get("remapped_code") or w.get("suggested_code") or "").strip()
                if code and remapped:
                    parts.append(f"{code} -> {remapped}")
                elif code:
                    parts.append(code)
            if parts:
                messages.warning(
                    self.request,
                    "S'han detectat columnes d'Excel amb noms reservats i s'han remapejat automaticament "
                    f"({', '.join(parts)}).",
                )
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("inscripcions_list", kwargs={"pk": self.competicio.pk})


def recalcular_ordre_sortida(qs, group_codes):
    """
    Recalcula ordre_sortida segons camps d'agrupació (codes).
    Compatible amb:
    - builtins (camps del model)
    - extras (JSON: Inscripcio.extra)
    """
    records = list(qs.order_by("ordre_sortida", "id"))

    def sort_key(o):
        gvals = tuple(_norm_val(get_inscripcio_value(o, code)) for code in group_codes)
        prev = o.ordre_sortida if o.ordre_sortida is not None else 10**12
        return (gvals, prev, o.id)

    records.sort(key=sort_key)

    with transaction.atomic():
        for idx, obj in enumerate(records, start=1):
            if obj.ordre_sortida != idx:
                Inscripcio.objects.filter(id=obj.id).update(ordre_sortida=idx)


class InscripcionsListView(ListView):
    template_name = "competicio/inscripcions_list.html"
    context_object_name = "records"
    paginate_by = 25

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return self.get_queryset_base_filtrada().order_by("ordre_sortida", "id")

    def get_queryset_base_filtrada(self):
        qs = Inscripcio.objects.filter(competicio=self.competicio)

        q = self.request.GET.get("q")
        categoria = self.request.GET.get("categoria")
        subcategoria = self.request.GET.get("subcategoria")
        entitat = self.request.GET.get("entitat")

        if subcategoria:
            qs = qs.filter(subcategoria__iexact=subcategoria)
        if entitat:
            qs = qs.filter(entitat__icontains=entitat)
        if q:
            qs = qs.filter(
                Q(nom_i_cognoms__icontains=q) |
                Q(document__icontains=q) |
                Q(entitat__icontains=q)
            )
        if categoria:
            qs = qs.filter(categoria__iexact=categoria)

        return qs

    def get(self, request, *args, **kwargs):
        allowed = get_allowed_group_fields(self.competicio)
        allowed_codes = {f["code"] for f in allowed}
        sort_fields = get_available_sort_fields(self.competicio)
        sort_allowed_codes = {f["code"] for f in sort_fields}
        default_sort_code = "nom_i_cognoms" if "nom_i_cognoms" in sort_allowed_codes else (next(iter(sort_allowed_codes), ""))

        def resolve_sort_key(raw_key):
            k = LEGACY_SORT_KEY_MAP.get((raw_key or "").strip(), (raw_key or "").strip())
            if k in sort_allowed_codes:
                return k
            return default_sort_code

        def get_active_group_codes():
            selected = [g for g in request.GET.getlist("group_by") if g in allowed_codes]
            if not selected:
                selected = [g for g in (self.competicio.group_by_default or []) if g in allowed_codes]
            return selected

        def simple_key_for_obj(obj, group_codes):
            vals = [_norm_val(get_inscripcio_value(obj, c)) for c in group_codes]
            return json.dumps(vals, ensure_ascii=False)

        def pretty_label_from_simple_key(simple_key):
            try:
                vals = json.loads(simple_key)
                return " · ".join("(Sense valor)" if v in (None, "", "__NULL__") else str(v) for v in vals)
            except Exception:
                return simple_key

        def _capture_history_snapshot():
            return capture_inscripcions_history_snapshot(request, self.competicio)

        def _record_history(action_type, action_label, before_snapshot):
            after_snapshot = _capture_history_snapshot()
            record_inscripcions_history_entry(
                request,
                self.competicio,
                action_type=action_type,
                action_label=action_label,
                before_snapshot=before_snapshot,
                after_snapshot=after_snapshot,
            )

        # 0x) EXPORT EXCEL
        if request.GET.get("export_excel") == "1":
            available_excel = get_available_excel_columns(self.competicio)
            by_code = {c["code"]: c for c in available_excel}

            selected_cols_raw = request.GET.getlist("excel_cols")
            selected_codes = []
            for raw in selected_cols_raw:
                code = LEGACY_EXCEL_COL_MAP.get(raw, raw)
                if code in by_code and code not in selected_codes:
                    selected_codes.append(code)

            if not selected_codes:
                selected_codes = list(by_code.keys())

            if not selected_codes and "nom_i_cognoms" in by_code:
                selected_codes = ["nom_i_cognoms"]

            columns = [(by_code[c]["label"], c) for c in selected_codes]

            group_codes = get_active_group_codes()
            grouping_sig = "|".join(group_codes) if group_codes else ""
            qs_base = self.get_queryset_base_filtrada()

            wb = Workbook()
            ws = wb.active
            ws.title = "Inscripcions"

            title_font = Font(bold=True, size=12)
            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="E9EEF7")
            group_fill = PatternFill("solid", fgColor="DDE7FF")

            def write_table_header(r):
                for col_i, (label, _) in enumerate(columns, start=1):
                    c = ws.cell(row=r, column=col_i, value=label)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = Alignment(vertical="center")

            def write_row(r, obj):
                for col_i, (_label, code) in enumerate(columns, start=1):
                    ws.cell(row=r, column=col_i, value=get_excel_export_value(obj, code))

            qs_all = qs_base.annotate(
                grup_null=Case(
                    When(grup__isnull=True, then=1),
                    default=0,
                    output_field=IntegerField(),
                )
            ).order_by("grup_null", "grup", "ordre_sortida", "id")

            merges = (self.competicio.tab_merges or {}).get(grouping_sig, [])
            merge_map = {}
            merged_label_map = {}

            for group_keys in merges:
                t = tuple(group_keys)
                for x in group_keys:
                    merge_map[x] = t
                merged_key = json.dumps(list(t), ensure_ascii=False)
                merged_label_map[merged_key] = " + ".join(pretty_label_from_simple_key(x) for x in t)

            def tab_label_for_obj(obj):
                if not group_codes:
                    return "Sense agrupació"
                simple = simple_key_for_obj(obj, group_codes)
                mid = merge_map.get(simple)
                if mid:
                    mk = json.dumps(list(mid), ensure_ascii=False)
                    return merged_label_map.get(mk, mk)
                return pretty_label_from_simple_key(simple)

            SENTINEL = object()
            current_grp = SENTINEL
            buffer = []
            row = 1

            def flush_group(objs, grp_num):
                nonlocal row
                if not objs:
                    return
                tab_label = tab_label_for_obj(objs[0])
                group_title = f"{tab_label} · Sense grup" if grp_num is None else f"{tab_label} · Grup {grp_num}"

                ws.cell(row=row, column=1, value=group_title).font = title_font
                ws.cell(row=row, column=1).fill = group_fill
                ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
                row += 1

                write_table_header(row)
                row += 1
                for o in objs:
                    write_row(row, o)
                    row += 1
                row += 1

            for obj in qs_all:
                g = obj.grup
                if current_grp is SENTINEL:
                    current_grp = g
                if g != current_grp:
                    flush_group(buffer, current_grp)
                    buffer = []
                    current_grp = g
                buffer.append(obj)

            flush_group(buffer, current_grp)

            for i, (label, _) in enumerate(columns, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, min(35, len(label) + 4))

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            filename = f"inscripcions_competicio_{self.competicio.pk}.xlsx"
            resp = HttpResponse(
                out.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            return resp

        # 0) CLEAR GROUPS
        if request.GET.get("clear_groups") == "1":
            if competicio_has_rotacions(self.competicio):
                messages.error(request, "No es poden esborrar grups mentre hi ha rotacions actives.")
                query = request.GET.copy()
                query.pop("clear_groups", None)
                return redirect(f"{request.path}?{query.urlencode()}")
            qs = self.get_queryset_base_filtrada()
            before_snapshot = _capture_history_snapshot()
            with transaction.atomic():
                qs.update(grup=None, grup_competicio=None, ordre_competicio=None)
                GrupCompeticio.objects.filter(competicio=self.competicio).update(actiu=False)
            clear_inscripcions_sort_state_for_competicio(request, self.competicio.id)
            _record_history("clear_groups", "Esborrar grups", before_snapshot)
            query = request.GET.copy()
            query.pop("clear_groups", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        # 0c) MAKE INDEPENDENT GROUP (subgrup o pestanya)
        if request.GET.get("make_independent_group") == "1":
            has_rotacions = competicio_has_rotacions(self.competicio)
            lvl = request.GET.get("lvl")  # "g1","g2","g3"

            group_codes = get_active_group_codes()
            if not group_codes:
                messages.error(request, "No hi ha agrupació activa per poder crear un grup independent.")
                query = request.GET.copy()
                for k in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                    query.pop(k, None)
                return redirect(f"{request.path}?{query.urlencode()}")

            v1 = request.GET.get("v1")
            v2 = request.GET.get("v2")
            v3 = request.GET.get("v3")

            def build_filter_ids_from_vals(vals):
                qs_base = self.get_queryset_base_filtrada()
                # en memòria per suportar extra
                builtin_fields = [c for c in group_codes if hasattr(Inscripcio, c)]
                recs = list(qs_base.only("id", "extra", *builtin_fields))
                ids = []
                for r in recs:
                    ok = True
                    for code, vv in zip(group_codes, vals):
                        if _norm_val(get_inscripcio_value(r, code)) != vv:
                            ok = False
                            break
                    if ok:
                        ids.append(r.id)
                return ids

            # CAS: nivell g1 i v1 és una key JSON (simple o merged)
            if lvl == "g1" and v1 and v1.strip().startswith("["):
                try:
                    parsed = json.loads(v1)
                except Exception:
                    parsed = None

                base_qs = self.get_queryset_base_filtrada()

                # merged: llista de strings que són JSON arrays
                if isinstance(parsed, list) and parsed and all(isinstance(x, str) and x.strip().startswith("[") for x in parsed):
                    before_snapshot = _capture_history_snapshot()
                    # IDs totals (unió de totes les pestanyes simples)
                    all_ids = []
                    for simple in parsed:
                        try:
                            vals = json.loads(simple)
                        except Exception:
                            continue
                        vals = [_norm_val(x) for x in vals]
                        all_ids.extend(build_filter_ids_from_vals(vals))

                    sub_qs = base_qs.filter(id__in=all_ids)

                # simple: llista de valors
                elif isinstance(parsed, list):
                    vals = [_norm_val(x) for x in parsed]
                    ids = build_filter_ids_from_vals(vals)
                    sub_qs = base_qs.filter(id__in=ids)
                    before_snapshot = _capture_history_snapshot()
                else:
                    sub_qs = None

                if sub_qs is not None:
                    selected_ids = list(sub_qs.values_list("id", flat=True))
                    if has_rotacions:
                        blocked_groups = _programmed_groups_emptied_by_move(self.competicio, selected_ids)
                        if blocked_groups:
                            messages.error(request, _message_for_emptied_programmed_groups(blocked_groups))
                            query = request.GET.copy()
                            for k in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                                query.pop(k, None)
                            return redirect(f"{request.path}?{query.urlencode()}")
                    existing_groups = list(sub_qs.exclude(grup__isnull=True).values_list("grup", flat=True).distinct())
                    if existing_groups and not has_rotacions:
                        new_group_num = min(existing_groups)
                        with transaction.atomic():
                            Inscripcio.objects.filter(competicio=self.competicio, grup=new_group_num).update(grup=None)
                            updated = sub_qs.update(grup=new_group_num)
                            sync_stable_groups_from_legacy(self.competicio)
                    else:
                        new_group_num = next_group_display_num(self.competicio)
                        with transaction.atomic():
                            updated = sub_qs.update(grup=new_group_num)
                            sync_stable_groups_from_legacy(self.competicio)

                    _record_history("make_independent_group", "Fer grup independent", before_snapshot)
                    messages.success(request, f"Creat el grup {new_group_num} amb {updated} inscripcions del subgrup.")
                    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
                    if is_ajax:
                        return JsonResponse(
                            with_inscripcions_history_payload(
                                {"ok": True, "new_group_num": new_group_num, "updated": updated},
                                request,
                                self.competicio.id,
                            )
                        )

                    query = request.GET.copy()
                    for k in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                        query.pop(k, None)
                    return redirect(f"{request.path}?{query.urlencode()}")

            # CAS clàssic: g1/g2/g3 amb v1/v2/v3 (valors normalitzats)
            level = {"g1": 1, "g2": 2, "g3": 3}.get(lvl, 1)
            vals = []
            if level >= 1:
                vals.append(_norm_val(v1))
            if level >= 2:
                vals.append(_norm_val(v2))
            if level >= 3:
                vals.append(_norm_val(v3))

            ids = build_filter_ids_from_vals(vals)
            base_qs = self.get_queryset_base_filtrada()
            sub_qs = base_qs.filter(id__in=ids)

            selected_ids = list(sub_qs.values_list("id", flat=True))
            if has_rotacions:
                blocked_groups = _programmed_groups_emptied_by_move(self.competicio, selected_ids)
                if blocked_groups:
                    messages.error(request, _message_for_emptied_programmed_groups(blocked_groups))
                    query = request.GET.copy()
                    for k in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                        query.pop(k, None)
                    return redirect(f"{request.path}?{query.urlencode()}")

            before_snapshot = _capture_history_snapshot()
            existing_groups = list(sub_qs.exclude(grup__isnull=True).values_list("grup", flat=True).distinct())
            if existing_groups and not has_rotacions:
                new_group_num = min(existing_groups)
                with transaction.atomic():
                    Inscripcio.objects.filter(competicio=self.competicio, grup=new_group_num).update(grup=None)
                    updated = sub_qs.update(grup=new_group_num)
                    sync_stable_groups_from_legacy(self.competicio)
            else:
                new_group_num = next_group_display_num(self.competicio)
                with transaction.atomic():
                    updated = sub_qs.update(grup=new_group_num)
                    sync_stable_groups_from_legacy(self.competicio)

            _record_history("make_independent_group", "Fer grup independent", before_snapshot)
            messages.success(request, f"Creat el grup {new_group_num} amb {updated} inscripcions del subgrup.")
            is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
            if is_ajax:
                return JsonResponse(
                    with_inscripcions_history_payload(
                        {"ok": True, "new_group_num": new_group_num, "updated": updated},
                        request,
                        self.competicio.id,
                    )
                )

            query = request.GET.copy()
            for k in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                query.pop(k, None)
            return redirect(f"{request.path}?{query.urlencode()}")
        # 0b) MAKE GROUPS COUNT (k grups per pestanya seleccionada)
        if request.GET.get("make_groups_count") == "1":
            if competicio_has_rotacions(self.competicio):
                messages.error(request, "No es poden reconfigurar grups mentre hi ha rotacions actives.")
                query = request.GET.copy()
                query.pop("make_groups_count", None)
                query.pop("group_count", None)
                return redirect(f"{request.path}?{query.urlencode()}")
            try:
                k = int(request.GET.get("group_count") or 0)
            except ValueError:
                k = 0
            if k < 1:
                messages.error(request, "El nombre de grups ha de ser com a mínim 1.")
                query = request.GET.copy()
                query.pop("make_groups_count", None)
                return redirect(f"{request.path}?{query.urlencode()}")

            qs_base = self.get_queryset_base_filtrada()
            before_snapshot = _capture_history_snapshot()
            group_codes = get_active_group_codes()
            grouping_sig = "|".join(group_codes) if group_codes else ""
            tab_keys = request.GET.getlist("tab_keys")  # checkboxes de pestanyes

            # Si no hi ha agrupació -> una sola "pestanya" virtual
            if not group_codes:
                selected_tabs = [("__ALL__", list(qs_base.values_list("id", flat=True)))]
            else:
                merges = (self.competicio.tab_merges or {}).get(grouping_sig, [])
                merge_map = {}
                for group_keys in merges:
                    t = tuple(group_keys)
                    for x in group_keys:
                        merge_map[x] = t

                builtin_fields = [c for c in group_codes if hasattr(Inscripcio, c)]
                records = list(qs_base.order_by("ordre_sortida", "id").only("id", "extra", *builtin_fields))

                tab_to_ids = OrderedDict()
                for r in records:
                    simple = simple_key_for_obj(r, group_codes)
                    mid = merge_map.get(simple)
                    tab_key = json.dumps(list(mid), ensure_ascii=False) if mid else simple
                    tab_to_ids.setdefault(tab_key, []).append(r.id)

                # si user ha seleccionat pestanyes, filtrem
                if tab_keys:
                    selected_tabs = [(tk, tab_to_ids.get(tk, [])) for tk in tab_keys if tk in tab_to_ids]
                else:
                    selected_tabs = list(tab_to_ids.items())

            # Assignació: sobre la unió de pestanyes seleccionades
            selected_ids = []
            seen_ids = set()
            for _tab_key, ids in selected_tabs:
                for ins_id in ids:
                    if ins_id not in seen_ids:
                        seen_ids.add(ins_id)
                        selected_ids.append(ins_id)

            if selected_ids:
                sub_qs = qs_base.filter(id__in=selected_ids).order_by("ordre_sortida", "id")
                objs = list(sub_qs.only("id", "grup"))
                if objs:
                    max_grup = (GrupCompeticio.objects.filter(competicio=self.competicio).aggregate(m=Max("display_num"))["m"] or 0)
                    assign_groups_k(objs, k, max_grup)
                    Inscripcio.objects.bulk_update(objs, ["grup"], batch_size=500)

            sync_stable_groups_from_legacy(self.competicio)
            _record_history("make_groups_count", "Crear grups per nombre", before_snapshot)

            query = request.GET.copy()
            query.pop("make_groups_count", None)
            query.pop("group_count", None)
            query.setlist("tab_keys", [])
            return redirect(f"{request.path}?{query.urlencode()}")

        # 0b) MAKE GROUPS SIZE (mida N)
        if request.GET.get("make_groups") == "1":
            if competicio_has_rotacions(self.competicio):
                messages.error(request, "No es poden reconfigurar grups mentre hi ha rotacions actives.")
                query = request.GET.copy()
                query.pop("make_groups", None)
                return redirect(f"{request.path}?{query.urlencode()}")
            try:
                size = int(request.GET.get("group_size") or 0)
            except ValueError:
                size = 0

            if size < 2:
                messages.error(request, "La mida del grup ha de ser com a mínim 2.")
                query = request.GET.copy()
                query.pop("make_groups", None)
                return redirect(f"{request.path}?{query.urlencode()}")

            group_mode = request.GET.get("group_mode") or "fixed"  # fixed / balanced
            group_codes = get_active_group_codes()

            qs = self.get_queryset_base_filtrada().order_by("ordre_sortida", "id")
            before_snapshot = _capture_history_snapshot()
            builtin_fields = [c for c in group_codes if hasattr(Inscripcio, c)]
            objs = list(qs.only("id", "grup", "extra", *builtin_fields))

            global_group_num = 0

            if not group_codes:
                if group_mode == "balanced":
                    global_group_num = assign_groups_balanced(objs, size, global_group_num)
                else:
                    for idx, obj in enumerate(objs, start=1):
                        obj.grup = (idx - 1) // size + 1
            else:
                if group_mode == "balanced":
                    current_key = None
                    buffer = []

                    def flush_buffer():
                        nonlocal global_group_num, buffer
                        if buffer:
                            global_group_num = assign_groups_balanced(buffer, size, global_group_num)
                            buffer = []

                    for obj in objs:
                        key = tuple(_norm_val(get_inscripcio_value(obj, c)) for c in group_codes)
                        if key != current_key:
                            flush_buffer()
                            current_key = key
                        buffer.append(obj)

                    flush_buffer()
                else:
                    current_key = None
                    count_in_chunk = 0
                    for obj in objs:
                        key = tuple(_norm_val(get_inscripcio_value(obj, c)) for c in group_codes)
                        if key != current_key:
                            current_key = key
                            count_in_chunk = 0
                        if count_in_chunk == 0:
                            global_group_num += 1
                        obj.grup = global_group_num
                        count_in_chunk += 1
                        if count_in_chunk >= size:
                            count_in_chunk = 0

            with transaction.atomic():
                Inscripcio.objects.bulk_update(objs, ["grup"], batch_size=500)
                sync_stable_groups_from_legacy(self.competicio)
            _record_history("make_groups_size", "Crear grups per mida", before_snapshot)

            query = request.GET.copy()
            query.pop("make_groups", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        # 1) CLEAR GROUPING SETTINGS
        if request.GET.get("clear_group") == "1":
            before_snapshot = _capture_history_snapshot()
            self.competicio.group_by_default = []
            self.competicio.tab_merges = {}
            self.competicio.save(update_fields=["group_by_default", "tab_merges"])
            clear_inscripcions_sort_state_for_competicio(request, self.competicio.id)
            _record_history("clear_grouping", "Treure agrupacio", before_snapshot)

            query = request.GET.copy()
            query.pop("clear_group", None)
            query.setlist("group_by", [])
            return redirect(f"{request.path}?{query.urlencode()}")

        # 2) RECALC ORDER
        if request.GET.get("recalc_order") == "1":
            group_codes = [g for g in request.GET.getlist("group_by") if g in allowed_codes]

            qs = self.get_queryset_base_filtrada()
            before_snapshot = _capture_history_snapshot()
            # versió portable (builtins+extra) -> si tu ja has substituït la funció global recalcular_ordre_sortida,
            # això ja funciona. Si no, fes-ho servir com a fallback:
            records = list(qs.order_by("ordre_sortida", "id"))

            def sort_key(o):
                gvals = tuple(_norm_val(get_inscripcio_value(o, code)) for code in group_codes)
                prev = o.ordre_sortida if o.ordre_sortida is not None else 10**12
                return (gvals, prev, o.id)

            records.sort(key=sort_key)
            with transaction.atomic():
                for idx, obj in enumerate(records, start=1):
                    if obj.ordre_sortida != idx:
                        Inscripcio.objects.filter(id=obj.id).update(ordre_sortida=idx)
            _record_history("recalc_order", "Aplicar agrupacio", before_snapshot)

            query = request.GET.copy()
            query.pop("recalc_order", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        # 3) SHUFFLE
        if request.GET.get("shuffle_order") == "1":
            qs = self.get_queryset_base_filtrada()
            before_snapshot = _capture_history_snapshot()
            shuffle_ordre_sortida(qs)
            _record_history("shuffle_order", "Barreja aleatoriament", before_snapshot)
            query = request.GET.copy()
            query.pop("shuffle_order", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        # 4) Persistència group_by
        if "group_by" in request.GET:
            selected = [g for g in request.GET.getlist("group_by") if g in allowed_codes]
            if selected != (self.competicio.group_by_default or []):
                self.competicio.group_by_default = selected
                self.competicio.save(update_fields=["group_by_default"])
        else:
            saved = [g for g in (self.competicio.group_by_default or []) if g in allowed_codes]
            if saved:
                query = request.GET.copy()
                for g in saved:
                    query.appendlist("group_by", g)
                return redirect(f"{request.path}?{query.urlencode()}")

        return super().get(request, *args, **kwargs)

    def get_paginate_by(self, queryset):
        if self.request.GET.getlist("group_by"):
            return None

        per_page = self.request.GET.get("per_page")
        if not per_page or per_page == "all":
            return None

        try:
            return int(per_page)
        except ValueError:
            return None

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = self.competicio

        allowed = get_allowed_group_fields(self.competicio)
        allowed_codes = {f["code"] for f in allowed}

        ctx["allowed_group_fields"] = allowed
        ctx["current_query"] = self.request.GET.urlencode()
        ctx["title_fields_selected"] = self.request.GET.getlist("title_fields")
        sort_fields = get_available_sort_fields(self.competicio)
        sort_codes = {f["code"] for f in sort_fields}
        raw_sort_key = self.request.GET.get("sort_key") or "nom_i_cognoms"
        sort_key_selected = LEGACY_SORT_KEY_MAP.get(raw_sort_key, raw_sort_key)
        if sort_key_selected not in sort_codes:
            sort_key_selected = "nom_i_cognoms" if "nom_i_cognoms" in sort_codes else (next(iter(sort_codes), ""))
        ctx["sort_field_options"] = sort_fields
        ctx["sort_key_selected"] = sort_key_selected

        selected = self.request.GET.getlist("group_by")
        if not selected:
            selected = self.competicio.group_by_default or []
        selected = [g for g in selected if g in allowed_codes]
        ctx["selected_group_fields"] = selected

        records = self.get_queryset_base_filtrada().order_by("ordre_sortida", "id")
        grouping_sig = "|".join(selected) if selected else ""
        ctx["grouping_sig"] = grouping_sig

        def pretty_val(v):
            return "(Sense valor)" if v in (None, "") else str(v)

        if selected:
            grouped = OrderedDict()
            label_map = {}

            for r in records:
                vals = [_norm_val(get_inscripcio_value(r, code)) for code in selected]
                key = json.dumps(vals, ensure_ascii=False)
                grouped.setdefault(key, []).append(r)

                if key not in label_map:
                    parts = [pretty_val(get_inscripcio_value(r, code)) for code in selected]
                    label_map[key] = " · ".join(parts)

            merges = (self.competicio.tab_merges or {}).get(grouping_sig, [])
            merge_map = {}
            for group_keys in merges:
                if not group_keys:
                    continue
                t = tuple(group_keys)
                for k in group_keys:
                    merge_map[k] = t

            grouped_merged = OrderedDict()
            label_map_merged = {}

            for k, rows in grouped.items():
                merged_tuple = merge_map.get(k)
                if merged_tuple:
                    tab_key = json.dumps(list(merged_tuple), ensure_ascii=False)
                    grouped_merged.setdefault(tab_key, []).extend(rows)
                    if tab_key not in label_map_merged:
                        parts = []
                        for sk in merged_tuple:
                            p = label_map.get(sk, sk)
                            if p not in parts:
                                parts.append(p)
                        label_map_merged[tab_key] = " + ".join(parts)
                else:
                    tab_key = k
                    grouped_merged.setdefault(tab_key, []).extend(rows)
                    if tab_key not in label_map_merged:
                        label_map_merged[tab_key] = label_map.get(k, k)


            records_grouped = [
                (label_map_merged.get(k, k), rows, k)
                for k, rows in grouped_merged.items()
            ]
            ctx["tabs"] = [
                {"key": group_key, "label": group_label, "count": len(group_records)}
                for (group_label, group_records, group_key) in records_grouped
            ]
            
            ctx["records_grouped"] = records_grouped
        else:
            ctx["records_grouped"] = None

        excel_cols = get_available_excel_columns(self.competicio)
        excel_codes = {c["code"] for c in excel_cols}
        ctx["allowed_excel_columns"] = [(c["code"], c.get("ui_label") or c["label"]) for c in excel_cols]

        sel_cols_raw = self.request.GET.getlist("excel_cols")
        sel_cols = []
        for raw in sel_cols_raw:
            code = LEGACY_EXCEL_COL_MAP.get(raw, raw)
            if code in excel_codes and code not in sel_cols:
                sel_cols.append(code)
        if not sel_cols:
            sel_cols = [c["code"] for c in excel_cols]
        ctx["excel_cols_selected"] = sel_cols

        base = self.get_queryset_base_filtrada()
        ctx["categories_distinct"] = list(base.order_by().values_list("categoria", flat=True).distinct())
        ctx["cats_selected"] = self.request.GET.getlist("cats")

        return ctx

class InscripcioFormViewMixin:
    form_class = InscripcioForm
    template_name = "competicio/inscripcio_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["competicio"] = self.competicio
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        next_url = self.request.GET.get("next", "")
        ctx["competicio"] = self.competicio
        ctx["next"] = next_url
        ctx["cancel_url"] = (
            next_url
            or self.request.META.get("HTTP_REFERER")
            or reverse("inscripcions_list", kwargs={"pk": self.kwargs["pk"]})
        )
        form = ctx.get("form")
        ctx["form_basic_fields"] = []
        ctx["form_extra_fields"] = []
        ctx["show_altres_fields"] = {}
        ctx["full_width_basic_field_names"] = []
        ctx["altres_wrapper_field_names"] = []
        if form is not None:
            ctx["form_basic_fields"] = [
                form[name]
                for name in getattr(form, "basic_field_names", [])
                if name in form.fields
            ]
            ctx["form_extra_fields"] = [
                form[name]
                for name in getattr(form, "extra_field_names", [])
                if name in form.fields
            ]
            ctx["show_altres_fields"] = dict(getattr(form, "show_altres_fields", {}) or {})
            ctx["full_width_basic_field_names"] = list(
                getattr(form, "full_width_basic_field_names", []) or []
            )
            ctx["altres_wrapper_field_names"] = list(
                getattr(form, "other_wrapper_field_names", []) or []
            )
        return ctx

    def get_success_url(self):
        nxt = self.request.GET.get("next")
        if nxt:
            return nxt
        return reverse("inscripcions_list", kwargs={"pk": self.kwargs["pk"]})


class InscripcioUpdateView(InscripcioFormViewMixin, UpdateView):
    model = Inscripcio
    pk_url_kwarg = "ins_id"

    def get_queryset(self):
        # Seguretat: només permet editar inscripcions de la competició del pk
        return Inscripcio.objects.filter(competicio_id=self.competicio.id)

    def form_valid(self, form):
        old_group = None
        if self.object and self.object.pk:
            old_group = (
                Inscripcio.objects
                .select_related("grup_competicio")
                .filter(pk=self.object.pk)
                .first()
            )
        response = super().form_valid(form)
        current = self.object
        new_group = get_group_for_display_num(current.competicio, current.grup)
        if current.grup and new_group is None:
            new_group = ensure_group_for_display_num(current.competicio, current.grup)
        if new_group is not None:
            if old_group and old_group.grup_competicio_id == new_group.id:
                Inscripcio.objects.filter(pk=current.pk).update(
                    grup_competicio=new_group,
                    grup=new_group.display_num,
                )
            else:
                move_inscripcio_to_group(current, new_group)
        else:
            previous_group = old_group.grup_competicio if old_group else None
            Inscripcio.objects.filter(pk=current.pk).update(
                grup=None,
                grup_competicio=None,
                ordre_competicio=None,
            )
            compact_competition_order_for_group(previous_group)
        sync_competicio_group_names_view(current.competicio)
        return response


class InscripcioCreateView(InscripcioFormViewMixin, CreateView):
    model = Inscripcio

    def form_valid(self, form):
        # Assign the competition before saving
        form.instance.competicio = self.competicio
        # If ordre_sortida not provided, append to the end
        if not form.instance.ordre_sortida:
            max_ord = Inscripcio.objects.filter(competicio=self.competicio).aggregate(m=Max("ordre_sortida"))["m"] or 0
            form.instance.ordre_sortida = max_ord + 1
        response = super().form_valid(form)
        group = get_group_for_display_num(self.competicio, self.object.grup)
        if self.object.grup and group is None:
            group = ensure_group_for_display_num(self.competicio, self.object.grup)
        if group is not None:
            move_inscripcio_to_group(self.object, group)
        sync_competicio_group_names_view(self.competicio)
        return response

class InscripcioDeleteView(DeleteView):
    model = Inscripcio
    pk_url_kwarg = "ins_id"
    template_name = "competicio/inscripcio_confirm_delete.html"

    def get_queryset(self):
        return Inscripcio.objects.filter(competicio_id=self.kwargs["pk"])

    def get_success_url(self):
        nxt = self.request.GET.get("next")
        if nxt:
            return nxt
        return reverse("inscripcions_list", kwargs={"pk": self.kwargs["pk"]})

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        old_group = getattr(self.object, "grup_competicio", None)
        response = super().delete(request, *args, **kwargs)
        compact_competition_order_for_group(old_group)
        sync_competicio_group_names_view(get_object_or_404(Competicio, pk=self.kwargs["pk"]))
        return response


@require_POST
@csrf_protect
def inscripcions_sort_apply(request, pk):
    """
    Ordenacio per columna amb ambits:
      - all        : totes les inscripcions filtrades
      - tab        : dins de cada pestanya (group_by/tab_merges actius)
      - all_groups : dins de cada grup numeric complet visible al context
      - group      : nomes un grup numeric complet (Inscripcio.grup)

    IMPORTANT:
      - Empats mantenen l'ordre previ (acumulacio real).
      - S'aplica sobre el queryset filtrat actual.
      - Si es reaplica el mateix camp+ambit, es substitueix mantenint la prioritat.
    """
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invalid")

    sort_fields = get_available_sort_fields(competicio)
    sort_codes = {f["code"] for f in sort_fields}

    raw_sort_key = (payload.get("sort_key") or "").strip()
    sort_key = LEGACY_SORT_KEY_MAP.get(raw_sort_key, raw_sort_key)
    if sort_key not in sort_codes:
        return HttpResponseBadRequest("sort_key invalid")

    sort_dir = (payload.get("sort_dir") or "asc").strip()
    if sort_dir not in ("asc", "desc", "arrow_asc", "arrow_desc", "custom"):
        return HttpResponseBadRequest("sort_dir invalid")

    scope = (payload.get("scope") or "all").strip().lower()
    if scope not in ("all", "tab", "all_groups", "group"):
        return HttpResponseBadRequest("scope invalid")

    allowed_group_fields = get_allowed_group_fields(competicio)
    allowed_group_codes = {f["code"] for f in allowed_group_fields}
    selected_group_codes_context = _normalize_sort_group_by(
        payload.get("group_by"),
        allowed_group_codes,
        fallback_group_by=competicio.group_by_default or [],
    )

    group_num = payload.get("group_num")
    if scope == "group":
        try:
            group_num = int(group_num)
        except Exception:
            return HttpResponseBadRequest("group_num invalid")
        if group_num <= 0:
            return HttpResponseBadRequest("group_num invalid")

    filters = _normalize_sort_filters(payload.get("filters"))
    context_key = build_inscripcions_sort_context_key(
        competicio.id,
        filters=filters,
        group_by=selected_group_codes_context,
    )
    pre_state = get_inscripcions_sort_context_state(request, context_key)
    pre_stack_raw = pre_state.get("stack") if isinstance(pre_state.get("stack"), list) else []
    pre_stack = []
    for it in pre_stack_raw:
        normalized = _normalize_sort_criterion(
            it,
            sort_codes=sort_codes,
            allowed_group_codes=allowed_group_codes,
            fallback_group_by=selected_group_codes_context,
        )
        if normalized is not None:
            pre_stack.append(normalized)

    new_entry = _normalize_sort_criterion(
        {
            "sort_key": sort_key,
            "sort_dir": sort_dir,
            "scope": scope,
            "group_num": group_num if scope == "group" else None,
            "group_by": selected_group_codes_context,
        },
        sort_codes=sort_codes,
        allowed_group_codes=allowed_group_codes,
        fallback_group_by=selected_group_codes_context,
    )
    if new_entry is None:
        return HttpResponseBadRequest("criteri invalid")

    qs = _build_inscripcions_filtered_qs(competicio, filters)
    record_sort_codes = [new_entry.get("sort_key")] + [entry.get("sort_key") for entry in pre_stack]
    records = list(_build_sort_records_queryset(qs, record_sort_codes + ["grup"]))
    if not records:
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "updated": 0,
                    "total": 0,
                    "scope": scope,
                    "stack_count": 0,
                    "competition_order_tail": False,
                },
                request,
                competicio.id,
            )
        )

    current_ids = [r.id for r in records]
    pre_runtime = _build_sort_application_runtime(competicio, records, record_sort_codes, pre_stack)

    state = reconcile_inscripcions_sort_context_state(
        request,
        context_key,
        current_ids,
        current_base_ids=pre_runtime["application_ids"],
    )
    competition_order_tail_active = bool(state.get("competition_order_tail"))
    stack_existing_raw = state.get("stack") if isinstance(state.get("stack"), list) else []

    stack_existing = []
    for it in stack_existing_raw:
        normalized = _normalize_sort_criterion(
            it,
            sort_codes=sort_codes,
            allowed_group_codes=allowed_group_codes,
            fallback_group_by=selected_group_codes_context,
        )
        if normalized is not None:
            stack_existing.append(normalized)

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)

    stack_full = _upsert_sort_stack_entry_preserving_priority(stack_existing, new_entry)
    if len(stack_full) > 20:
        stack_full = stack_full[-20:]

    runtime = _build_sort_application_runtime(
        competicio,
        records,
        [entry.get("sort_key") for entry in stack_full],
        stack_full,
        include_competition_order=competition_order_tail_active,
    )
    id_to_record = runtime["id_to_record"]
    application_ids = runtime["application_ids"]
    base_ids_state = state.get("base_ids") if isinstance(state.get("base_ids"), list) else []
    valid_base = (
        isinstance(base_ids_state, list)
        and len(base_ids_state) == len(application_ids)
        and set(base_ids_state) == set(application_ids)
    )
    if not valid_base or not stack_existing:
        base_ids = list(application_ids)
    else:
        base_ids = list(base_ids_state)

    final_ids = _apply_sort_pipeline(
        base_ids,
        id_to_record,
        stack_full,
        competicio,
        runtime=runtime,
        competition_order_tail=competition_order_tail_active,
    )

    updated_count = _persist_inscripcions_order_from_ids(id_to_record, final_ids)

    order_sig = compute_inscripcions_order_signature_from_ids(final_ids)
    save_inscripcions_sort_context_state(
        request,
        context_key,
        stack=stack_full,
        order_sig=order_sig,
        base_ids=base_ids,
        context_ids=current_ids,
        competition_order_tail=competition_order_tail_active,
    )

    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="sort_apply",
        action_label="Aplicar ordenacio",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )

    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "scope": scope,
                "sort_key": sort_key,
                "sort_dir": sort_dir,
                "updated": updated_count,
                "total": len(runtime["records"]),
                "stack_count": len(stack_full),
                "competition_order_tail": competition_order_tail_active,
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_sort_remove(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    try:
        priority = int(payload.get("priority"))
    except Exception:
        return HttpResponseBadRequest("priority invalid")
    if priority <= 0:
        return HttpResponseBadRequest("priority invalid")

    sort_fields = get_available_sort_fields(competicio)
    sort_codes = {f["code"] for f in sort_fields}
    allowed_group_fields = get_allowed_group_fields(competicio)
    allowed_group_codes = {f["code"] for f in allowed_group_fields}

    selected_group_codes_context = _normalize_sort_group_by(
        payload.get("group_by"),
        allowed_group_codes,
        fallback_group_by=competicio.group_by_default or [],
    )
    filters = _normalize_sort_filters(payload.get("filters"))
    context_key = build_inscripcions_sort_context_key(
        competicio.id,
        filters=filters,
        group_by=selected_group_codes_context,
    )

    state = get_inscripcions_sort_context_state(request, context_key)
    stack_raw = state.get("stack") if isinstance(state.get("stack"), list) else []
    stack = []
    for it in stack_raw:
        norm = _normalize_sort_criterion(
            it,
            sort_codes=sort_codes,
            allowed_group_codes=allowed_group_codes,
            fallback_group_by=selected_group_codes_context,
        )
        if norm is not None:
            stack.append(norm)

    if not stack:
        clear_inscripcions_sort_context_state(request, context_key)
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "removed": False,
                    "stack_count": 0,
                    "competition_order_tail": False,
                },
                request,
                competicio.id,
            )
        )

    qs = _build_inscripcions_filtered_qs(competicio, filters)
    record_sort_codes = [entry.get("sort_key") for entry in stack]
    competition_order_tail_active = bool(state.get("competition_order_tail"))
    records = list(
        _build_sort_records_queryset(
            qs,
            record_sort_codes + ["grup"],
            include_competition_order=competition_order_tail_active,
        )
    )
    if not records:
        clear_inscripcions_sort_context_state(request, context_key)
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "removed": False,
                    "stack_count": 0,
                    "competition_order_tail": False,
                },
                request,
                competicio.id,
            )
        )

    current_ids = [r.id for r in records]
    runtime = _build_sort_application_runtime(
        competicio,
        records,
        record_sort_codes,
        stack,
        include_competition_order=competition_order_tail_active,
    )
    id_to_record = runtime["id_to_record"]
    application_ids = runtime["application_ids"]

    if priority > len(stack):
        return HttpResponseBadRequest("priority out of range")

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)

    removed_idx = priority - 1
    removed = stack.pop(removed_idx)

    base_ids_state = state.get("base_ids") if isinstance(state.get("base_ids"), list) else []
    valid_base = (
        isinstance(base_ids_state, list)
        and len(base_ids_state) == len(application_ids)
        and set(base_ids_state) == set(application_ids)
    )
    base_ids = list(base_ids_state) if valid_base else list(application_ids)

    if stack:
        final_ids = _apply_sort_pipeline(
            base_ids,
            id_to_record,
            stack,
            competicio,
            runtime=runtime,
            competition_order_tail=competition_order_tail_active,
        )
    else:
        final_ids = list(base_ids)

    updated_count = _persist_inscripcions_order_from_ids(id_to_record, final_ids)

    if stack:
        order_sig = compute_inscripcions_order_signature_from_ids(final_ids)
        save_inscripcions_sort_context_state(
            request,
            context_key,
            stack=stack,
            order_sig=order_sig,
            base_ids=base_ids,
            context_ids=current_ids,
            competition_order_tail=competition_order_tail_active,
        )
    else:
        clear_inscripcions_sort_context_state(request, context_key)

    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="sort_remove",
        action_label="Treure criteri d'ordenacio",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )

    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "removed": True,
                "removed_priority": priority,
                "removed_sort_key": removed.get("sort_key"),
                "stack_count": len(stack),
                "updated": updated_count,
                "competition_order_tail": bool(stack) and competition_order_tail_active,
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_sort_clear(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    allowed_group_fields = get_allowed_group_fields(competicio)
    allowed_group_codes = {f["code"] for f in allowed_group_fields}
    selected_group_codes_context = _normalize_sort_group_by(
        payload.get("group_by"),
        allowed_group_codes,
        fallback_group_by=competicio.group_by_default or [],
    )
    filters = _normalize_sort_filters(payload.get("filters"))
    context_key = build_inscripcions_sort_context_key(
        competicio.id,
        filters=filters,
        group_by=selected_group_codes_context,
    )

    sort_fields = get_available_sort_fields(competicio)
    sort_codes = {f["code"] for f in sort_fields}

    qs = _build_inscripcions_filtered_qs(competicio, filters)
    records = list(qs.order_by("ordre_sortida", "id"))
    current_ids = [r.id for r in records]

    state = get_inscripcions_sort_context_state(request, context_key)
    stack_raw = state.get("stack") if isinstance(state.get("stack"), list) else []
    stack = []
    for it in stack_raw:
        norm = _normalize_sort_criterion(
            it,
            sort_codes=sort_codes,
            allowed_group_codes=allowed_group_codes,
            fallback_group_by=selected_group_codes_context,
        )
        if norm is not None:
            stack.append(norm)

    if not stack:
        clear_inscripcions_sort_context_state(request, context_key)
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "cleared": False,
                    "stack_count": 0,
                    "competition_order_tail": False,
                },
                request,
                competicio.id,
            )
        )

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)

    clear_inscripcions_sort_context_state(request, context_key)

    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="sort_clear",
        action_label="Netejar ordenacions",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )

    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "cleared": True,
                "stack_count": 0,
                "competition_order_tail": False,
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_sort_competition_tail_toggle(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    enabled = _normalize_competition_order_tail_flag(payload.get("enabled"))
    sort_fields = get_available_sort_fields(competicio)
    sort_codes = {f["code"] for f in sort_fields}
    allowed_group_fields = get_allowed_group_fields(competicio)
    allowed_group_codes = {f["code"] for f in allowed_group_fields}

    selected_group_codes_context = _normalize_sort_group_by(
        payload.get("group_by"),
        allowed_group_codes,
        fallback_group_by=competicio.group_by_default or [],
    )
    filters = _normalize_sort_filters(payload.get("filters"))
    context_key = build_inscripcions_sort_context_key(
        competicio.id,
        filters=filters,
        group_by=selected_group_codes_context,
    )

    qs = _build_inscripcions_filtered_qs(competicio, filters)
    current_ids = list(qs.order_by("ordre_sortida", "id").values_list("id", flat=True))
    state = reconcile_inscripcions_sort_context_state(request, context_key, current_ids)
    stack_raw = state.get("stack") if isinstance(state.get("stack"), list) else []
    stack = []
    for it in stack_raw:
        norm = _normalize_sort_criterion(
            it,
            sort_codes=sort_codes,
            allowed_group_codes=allowed_group_codes,
            fallback_group_by=selected_group_codes_context,
        )
        if norm is not None:
            stack.append(norm)

    if not stack:
        clear_inscripcions_sort_context_state(request, context_key)
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "applied": False,
                    "updated": 0,
                    "stack_count": 0,
                    "competition_order_tail": False,
                    "reason": "no_stack",
                },
                request,
                competicio.id,
            )
        )

    records = list(
        _build_sort_records_queryset(
            qs,
            [entry.get("sort_key") for entry in stack] + ["grup"],
            include_competition_order=True,
        )
    )

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)

    if not records:
        save_inscripcions_sort_context_state(
            request,
            context_key,
            stack=stack,
            order_sig="",
            base_ids=[],
            context_ids=[],
            competition_order_tail=enabled,
        )
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "applied": True,
                    "updated": 0,
                    "stack_count": len(stack),
                    "competition_order_tail": enabled,
                },
                request,
                competicio.id,
            )
        )

    runtime = _build_sort_application_runtime(
        competicio,
        records,
        [entry.get("sort_key") for entry in stack],
        stack,
        include_competition_order=True,
    )
    id_to_record = runtime["id_to_record"]
    application_ids = runtime["application_ids"]
    base_ids_state = state.get("base_ids") if isinstance(state.get("base_ids"), list) else []
    valid_base = (
        len(base_ids_state) == len(application_ids)
        and set(base_ids_state) == set(application_ids)
    )
    base_ids = list(base_ids_state) if valid_base else list(application_ids)
    final_ids = _apply_sort_pipeline(
        base_ids,
        id_to_record,
        stack,
        competicio,
        runtime=runtime,
        competition_order_tail=enabled,
    )
    updated_count = _persist_inscripcions_order_from_ids(id_to_record, final_ids)
    order_sig = compute_inscripcions_order_signature_from_ids(final_ids)
    save_inscripcions_sort_context_state(
        request,
        context_key,
        stack=stack,
        order_sig=order_sig,
        base_ids=base_ids,
        context_ids=current_ids,
        competition_order_tail=enabled,
    )

    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="sort_competition_tail_toggle",
        action_label="Activar ordre de competicio" if enabled else "Desactivar ordre de competicio",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )

    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "applied": True,
                "updated": updated_count,
                "stack_count": len(stack),
                "competition_order_tail": enabled,
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_sort_custom_values(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    sort_fields = get_available_sort_fields(competicio)
    sort_codes = {f["code"] for f in sort_fields}
    sort_label_by_code = {
        f["code"]: (f.get("ui_label") or f.get("label") or f["code"])
        for f in sort_fields
        if isinstance(f, dict) and f.get("code")
    }

    raw_sort_key = str(payload.get("sort_key") or payload.get("field_code") or "").strip()
    sort_key = LEGACY_SORT_KEY_MAP.get(raw_sort_key, raw_sort_key)
    if sort_key not in sort_codes:
        return HttpResponseBadRequest("sort_key invalid")

    allowed_group_fields = get_allowed_group_fields(competicio)
    allowed_group_codes = {f["code"] for f in allowed_group_fields}
    selected_group_codes_context = _normalize_sort_group_by(
        payload.get("group_by"),
        allowed_group_codes,
        fallback_group_by=competicio.group_by_default or [],
    )
    filters = _normalize_sort_filters(payload.get("filters"))

    qs_context = _build_inscripcions_filtered_qs(competicio, filters)
    context_records = list(_build_sort_records_queryset(qs_context, [sort_key]))
    context_stats = _collect_sort_field_value_stats(context_records, sort_key)

    qs_global = Inscripcio.objects.filter(competicio=competicio)
    global_records = list(_build_sort_records_queryset(qs_global, [sort_key]))
    global_stats = _collect_sort_field_value_stats(global_records, sort_key)

    custom_order_raw = get_competicio_custom_sort_order_values(
        competicio,
        sort_key,
        allowed_sort_codes=sort_codes,
    )
    custom_order, stale_order = _split_custom_sort_tokens(
        custom_order_raw,
        global_stats.keys(),
    )

    values = []
    seen = set()

    for token in custom_order:
        key = _custom_sort_token_key(token)
        if not key or key in seen:
            continue
        row = context_stats.get(key)
        if row is None:
            continue
        seen.add(key)
        values.append(
            {
                "value": token,
                "label": row["label"],
                "count": row["count"],
                "detected": True,
                "in_custom": True,
            }
        )

    remaining = [row for key, row in context_stats.items() if key not in seen]
    remaining.sort(key=lambda r: r["sort_scalar"])
    for row in remaining:
        values.append(
            {
                "value": row["token"],
                "label": row["label"],
                "count": row["count"],
                "detected": True,
                "in_custom": False,
            }
        )

    return JsonResponse(
        {
            "ok": True,
            "sort_key": sort_key,
            "sort_label": sort_label_by_code.get(sort_key, sort_key),
            "custom_order": custom_order,
            "custom_order_raw": custom_order_raw,
            "values": values,
            "stale_values": stale_order,
            "detected_count": len(context_stats),
            "context_group_by": selected_group_codes_context,
            "context_filters": filters,
        }
    )


@require_POST
@csrf_protect
def inscripcions_sort_custom_save(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    sort_fields = get_available_sort_fields(competicio)
    sort_codes = {f["code"] for f in sort_fields}

    raw_sort_key = str(payload.get("sort_key") or payload.get("field_code") or "").strip()
    sort_key = LEGACY_SORT_KEY_MAP.get(raw_sort_key, raw_sort_key)
    if sort_key not in sort_codes:
        return HttpResponseBadRequest("sort_key invalid")

    clear = bool(payload.get("clear"))
    raw_order = payload.get("order")
    if (not clear) and (not isinstance(raw_order, list)):
        return HttpResponseBadRequest("order invalid")

    filters = _normalize_sort_filters(payload.get("filters"))
    allowed_group_fields = get_allowed_group_fields(competicio)
    allowed_group_codes = {f["code"] for f in allowed_group_fields}
    selected_group_codes_context = _normalize_sort_group_by(
        payload.get("group_by"),
        allowed_group_codes,
        fallback_group_by=competicio.group_by_default or [],
    )
    context_key = build_inscripcions_sort_context_key(
        competicio.id,
        filters=filters,
        group_by=selected_group_codes_context,
    )
    preserve_missing_context = bool(payload.get("preserve_missing_context", True))

    context_qs = _build_inscripcions_filtered_qs(competicio, filters)
    context_records = list(_build_sort_records_queryset(context_qs, [sort_key]))
    context_stats = _collect_sort_field_value_stats(context_records, sort_key)
    current_ids = [r.id for r in context_records]
    state = reconcile_inscripcions_sort_context_state(request, context_key, current_ids)
    stack_raw = state.get("stack") if isinstance(state.get("stack"), list) else []
    stack = []
    for it in stack_raw:
        norm = _normalize_sort_criterion(
            it,
            sort_codes=sort_codes,
            allowed_group_codes=allowed_group_codes,
            fallback_group_by=selected_group_codes_context,
        )
        if norm is not None:
            stack.append(norm)
    competition_order_tail_active = bool(stack) and bool(state.get("competition_order_tail"))

    global_qs = Inscripcio.objects.filter(competicio=competicio)
    global_records = list(_build_sort_records_queryset(global_qs, [sort_key]))
    global_stats = _collect_sort_field_value_stats(global_records, sort_key)
    global_keys = set(global_stats.keys())

    existing_raw = get_competicio_custom_sort_order_values(
        competicio,
        sort_key,
        allowed_sort_codes=sort_codes,
    )
    existing_active, existing_stale = _split_custom_sort_tokens(existing_raw, global_keys)
    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)

    try:
        save_order = raw_order
        dropped_outside_competicio = 0
        preserved_outside_context = 0
        stale_removed = len(existing_stale)
        if not clear:
            incoming = _normalize_custom_sort_order(raw_order)
            incoming_active, _incoming_stale = _split_custom_sort_tokens(incoming, global_keys)
            dropped_outside_competicio = max(0, len(incoming) - len(incoming_active))

            incoming_keys = {_custom_sort_token_key(v) for v in incoming_active}
            if preserve_missing_context and existing_active:
                context_keys = set(context_stats.keys())
                for token in existing_active:
                    key = _custom_sort_token_key(token)
                    if not key or key in incoming_keys:
                        continue
                    if key in context_keys:
                        continue
                    incoming_active.append(token)
                    incoming_keys.add(key)
                    preserved_outside_context += 1

            save_order = incoming_active

        saved_values = set_competicio_custom_sort_order_values(
            competicio,
            sort_key,
            raw_values=save_order,
            clear=clear,
            allowed_sort_codes=sort_codes,
        )
    except ValueError:
        return HttpResponseBadRequest("sort_key invalid")

    # Si aquest camp esta actiu en l'stack del context actual, reaplica al moment
    # perque l'usuari vegi l'efecte del custom sense haver de fer un altre "Aplicar".
    reapplied = False
    reapplied_updated = 0
    if stack and current_ids and any(
        entry.get("sort_key") == sort_key and str(entry.get("sort_dir") or "") == "custom"
        for entry in stack
    ):
        runtime = _build_sort_application_runtime(
            competicio,
            context_records,
            [entry.get("sort_key") for entry in stack] + [sort_key],
            stack,
            include_competition_order=competition_order_tail_active,
        )
        id_to_record = runtime["id_to_record"]
        application_ids = runtime["application_ids"]
        base_ids_state = state.get("base_ids") if isinstance(state.get("base_ids"), list) else []
        valid_base = (
            len(base_ids_state) == len(application_ids)
            and set(base_ids_state) == set(application_ids)
        )
        base_ids = list(base_ids_state) if valid_base else list(application_ids)
        final_ids = _apply_sort_pipeline(
            base_ids,
            id_to_record,
            stack,
            competicio,
            runtime=runtime,
            competition_order_tail=competition_order_tail_active,
        )
        reapplied_updated = _persist_inscripcions_order_from_ids(id_to_record, final_ids)

        order_sig = compute_inscripcions_order_signature_from_ids(final_ids)
        save_inscripcions_sort_context_state(
            request,
            context_key,
            stack=stack,
            order_sig=order_sig,
            base_ids=base_ids,
            context_ids=current_ids,
            competition_order_tail=competition_order_tail_active,
        )
        reapplied = True

    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="sort_custom_save",
        action_label="Desar ordre custom",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )

    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "sort_key": sort_key,
                "custom_order": saved_values,
                "custom_active": bool(saved_values),
                "stale_removed": stale_removed if not clear else 0,
                "dropped_outside_competicio": dropped_outside_competicio if not clear else 0,
                "preserved_outside_context": preserved_outside_context if not clear else 0,
                "reapplied": reapplied,
                "reapplied_updated": reapplied_updated,
                "stack_count": len(stack),
                "competition_order_tail": competition_order_tail_active,
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_history_undo(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    entry = _perform_inscripcions_history_step(request, competicio, "undo")
    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "applied": bool(entry),
                "direction": "undo",
                "action_type": str(entry.get("action_type") or "") if isinstance(entry, dict) else "",
                "action_label": str(entry.get("action_label") or "") if isinstance(entry, dict) else "",
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_history_redo(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    entry = _perform_inscripcions_history_step(request, competicio, "redo")
    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "applied": bool(entry),
                "direction": "redo",
                "action_type": str(entry.get("action_type") or "") if isinstance(entry, dict) else "",
                "action_label": str(entry.get("action_label") or "") if isinstance(entry, dict) else "",
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_sort_undo(request, pk):
    """
    Endpoint de compatibilitat amb clients antics.
    Reutilitza el nou motor global d'historial.
    """
    competicio = get_object_or_404(Competicio, pk=pk)
    entry = _perform_inscripcions_history_step(request, competicio, "undo")
    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "restored": 1 if entry else 0,
                "applied": bool(entry),
                "direction": "undo",
                "action_type": str(entry.get("action_type") or "") if isinstance(entry, dict) else "",
                "action_label": str(entry.get("action_label") or "") if isinstance(entry, dict) else "",
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_groups_from_sort(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    has_rotacions = competicio_has_rotacions(competicio)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    strategy = str(payload.get("strategy") or "per_bucket").strip().lower()
    if strategy not in (
        "per_bucket",
        "count",
        "size_fixed",
        "size_balanced",
        "range_balanced",
        "count_with_range",
    ):
        return HttpResponseBadRequest("strategy invalid")

    preview_only = bool(payload.get("preview_only"))
    fallback_mode = _parse_fallback_mode(payload.get("fallback_mode"))

    allowed_group_fields = get_allowed_group_fields(competicio)
    allowed_group_codes = {f["code"] for f in allowed_group_fields}
    selected_group_codes_context = _normalize_sort_group_by(
        payload.get("group_by"),
        allowed_group_codes,
        fallback_group_by=competicio.group_by_default or [],
    )
    filters = _normalize_sort_filters(payload.get("filters"))
    context_key = build_inscripcions_sort_context_key(
        competicio.id,
        filters=filters,
        group_by=selected_group_codes_context,
    )

    sort_fields = get_available_sort_fields(competicio)
    sort_codes = {f["code"] for f in sort_fields}
    state = get_inscripcions_sort_context_state(request, context_key)
    stack_raw = state.get("stack") if isinstance(state.get("stack"), list) else []

    stack = []
    for it in stack_raw:
        normalized = _normalize_sort_criterion(
            it,
            sort_codes=sort_codes,
            allowed_group_codes=allowed_group_codes,
            fallback_group_by=selected_group_codes_context,
        )
        if normalized is not None:
            stack.append(normalized)

    partition_codes = _extract_sort_partition_codes(stack)

    qs = _build_inscripcions_filtered_qs(competicio, filters)
    records = list(qs.order_by("ordre_sortida", "id"))
    if not records:
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "preview_only": preview_only,
                    "updated": 0,
                    "groups_created": 0,
                    "buckets_total": 0,
                    "buckets_applied": 0,
                    "stack_used": partition_codes,
                    "resolution_mode": "auto",
                    "layers_used": [],
                    "effective_bucket_count": 0,
                    "strategy": strategy,
                    "used_fallback": False,
                    "preview": {
                        "groups_total": 0,
                        "members_total": 0,
                        "groups": [],
                        "existing_groups_total": 0,
                        "existing_members_total": 0,
                        "existing_groups": [],
                        "resolution_mode": "auto",
                        "layers_used": [],
                        "effective_bucket_count": 0,
                        "strategy": strategy,
                    } if preview_only else None,
                },
                request,
                competicio.id,
            )
        )

    resolution = _resolve_group_creation_buckets(
        competicio,
        records,
        group_codes=selected_group_codes_context,
        partition_codes=partition_codes,
        fallback_mode=fallback_mode,
    )
    if not resolution.get("ok"):
        return HttpResponseBadRequest(
            resolution.get("error") or "No hi ha criteris resolubles per construir blocs d'origen"
        )

    buckets = list(resolution.get("buckets") or [])
    layers_used = list(resolution.get("layers_used") or [])
    used_fallback = bool(resolution.get("used_fallback"))
    fallback_reason = str(resolution.get("fallback_reason") or "")
    bucket_sources_by_id_all = _build_bucket_sources_by_id(buckets)

    bucket_by_key = {b["key"]: b for b in buckets}
    selected_keys_raw = payload.get("selected_keys")
    if not isinstance(selected_keys_raw, list):
        selected_keys_raw = payload.get("selected_bucket_keys")
    if not isinstance(selected_keys_raw, list):
        selected_keys_raw = payload.get("selected_tab_keys")
    selected_keys = []
    if isinstance(selected_keys_raw, list):
        for value in selected_keys_raw:
            if isinstance(value, str) and value in bucket_by_key and value not in selected_keys:
                selected_keys.append(value)

    if selected_keys:
        buckets_to_apply = [bucket_by_key[k] for k in selected_keys]
    else:
        buckets_to_apply = list(buckets)

    target_ids = []
    seen_ids = set()
    for bucket in buckets_to_apply:
        for ins_id in bucket["ids"]:
            if ins_id in seen_ids:
                continue
            seen_ids.add(ins_id)
            target_ids.append(ins_id)

    existing_groups_preview = _build_existing_groups_preview(
        competicio,
        records,
        bucket_sources_by_id=bucket_sources_by_id_all,
        moving_ids=target_ids,
    )
    existing_members_total = sum(group["members_count"] for group in existing_groups_preview)

    if not target_ids:
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "preview_only": preview_only,
                    "updated": 0,
                    "groups_created": 0,
                    "buckets_total": len(buckets),
                    "buckets_applied": len(buckets_to_apply),
                    "stack_used": partition_codes,
                    "resolution_mode": "auto",
                    "layers_used": layers_used,
                    "effective_bucket_count": len(buckets),
                    "strategy": strategy,
                    "used_fallback": used_fallback,
                    "fallback_reason": fallback_reason,
                    "preview": {
                        "groups_total": 0,
                        "members_total": 0,
                        "groups": [],
                        "existing_groups_total": len(existing_groups_preview),
                        "existing_members_total": existing_members_total,
                        "existing_groups": existing_groups_preview,
                        "resolution_mode": "auto",
                        "layers_used": layers_used,
                        "effective_bucket_count": len(buckets),
                        "strategy": strategy,
                        "used_fallback": used_fallback,
                        "fallback_reason": fallback_reason,
                    } if preview_only else None,
                },
                request,
                competicio.id,
            )
        )

    id_to_record = {r.id: r for r in records}
    objs = [id_to_record[i] for i in target_ids if i in id_to_record]
    n = len(objs)
    if n == 0:
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "preview_only": preview_only,
                    "updated": 0,
                    "groups_created": 0,
                    "resolution_mode": "auto",
                    "layers_used": layers_used,
                    "effective_bucket_count": len(buckets),
                    "strategy": strategy,
                    "preview": {
                        "groups_total": 0,
                        "members_total": 0,
                        "groups": [],
                        "existing_groups_total": len(existing_groups_preview),
                        "existing_members_total": existing_members_total,
                        "existing_groups": existing_groups_preview,
                        "resolution_mode": "auto",
                        "layers_used": layers_used,
                        "effective_bucket_count": len(buckets),
                        "strategy": strategy,
                    } if preview_only else None,
                },
                request,
                competicio.id,
            )
        )

    sizes = []
    strategy_applied = strategy

    if strategy == "per_bucket":
        for bucket in buckets_to_apply:
            bucket_sz = len(bucket["ids"])
            if bucket_sz > 0:
                sizes.append(bucket_sz)

    elif strategy == "count":
        try:
            k = int(payload.get("group_count") or 0)
        except Exception:
            return HttpResponseBadRequest("group_count invalid")
        if k < 1:
            return HttpResponseBadRequest("group_count invalid")
        sizes = _balanced_sizes(n, k)

    elif strategy in ("size_fixed", "size_balanced"):
        try:
            size = int(payload.get("group_size") or 0)
        except Exception:
            return HttpResponseBadRequest("group_size invalid")
        if size < 2:
            return HttpResponseBadRequest("group_size invalid")
        if strategy == "size_fixed":
            sizes = _fixed_sizes(n, size)
        else:
            k = math.ceil(n / size)
            sizes = _balanced_sizes(n, k)

    elif strategy in ("range_balanced", "count_with_range"):
        try:
            min_size = int(payload.get("min_size") or 0)
            max_size = int(payload.get("max_size") or 0)
        except Exception:
            return HttpResponseBadRequest("min_size/max_size invalid")
        if min_size <= 0 or max_size <= 0 or min_size > max_size:
            return HttpResponseBadRequest("min_size/max_size invalid")

        preferred_k = None
        if strategy == "count_with_range":
            try:
                preferred_k = int(payload.get("group_count") or 0)
            except Exception:
                return HttpResponseBadRequest("group_count invalid")
            if preferred_k < 1:
                return HttpResponseBadRequest("group_count invalid")

        k_resolved, meta = _resolve_k_for_range(
            n,
            min_size,
            max_size,
            preferred_k=preferred_k,
            fallback_mode=fallback_mode,
        )
        if k_resolved is None:
            return HttpResponseBadRequest("No es pot resoldre una particio valida amb aquesta forquilla")
        if meta.get("used_fallback"):
            used_fallback = True
            fallback_reason = meta.get("fallback_reason") or fallback_reason
        sizes = _balanced_sizes(n, k_resolved)
        strategy_applied = strategy

    if not sizes:
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "preview_only": preview_only,
                    "updated": 0,
                    "groups_created": 0,
                    "buckets_total": len(buckets),
                    "buckets_applied": len(buckets_to_apply),
                    "stack_used": partition_codes,
                    "resolution_mode": "auto",
                    "layers_used": layers_used,
                    "effective_bucket_count": len(buckets),
                    "strategy": strategy_applied,
                    "used_fallback": used_fallback,
                    "fallback_reason": fallback_reason,
                    "preview": {
                        "groups_total": 0,
                        "members_total": n,
                        "groups": [],
                        "existing_groups_total": len(existing_groups_preview),
                        "existing_members_total": existing_members_total,
                        "existing_groups": existing_groups_preview,
                        "resolution_mode": "auto",
                        "layers_used": layers_used,
                        "effective_bucket_count": len(buckets),
                        "strategy": strategy_applied,
                        "used_fallback": used_fallback,
                        "fallback_reason": fallback_reason,
                    } if preview_only else None,
                },
                request,
                competicio.id,
            )
        )

    max_grup = (GrupCompeticio.objects.filter(competicio=competicio).aggregate(m=Max("display_num"))["m"] or 0)
    bucket_sources_by_id = _build_bucket_sources_by_id(buckets_to_apply)

    preview_groups = _build_group_creation_preview(
        objs,
        sizes,
        start_group_num=max_grup,
        bucket_sources_by_id=bucket_sources_by_id,
    )

    if preview_only:
        return JsonResponse(
            with_inscripcions_history_payload(
                {
                    "ok": True,
                    "preview_only": True,
                    "updated": 0,
                    "groups_created": len(preview_groups),
                    "buckets_total": len(buckets),
                    "buckets_applied": len(buckets_to_apply),
                    "stack_used": partition_codes,
                    "resolution_mode": "auto",
                    "layers_used": layers_used,
                    "effective_bucket_count": len(buckets),
                    "strategy": strategy_applied,
                    "used_fallback": used_fallback,
                    "fallback_reason": fallback_reason,
                    "size_min": min(sizes) if sizes else 0,
                    "size_max": max(sizes) if sizes else 0,
                    "preview": {
                        "groups_total": len(preview_groups),
                        "members_total": n,
                        "groups": preview_groups,
                        "existing_groups_total": len(existing_groups_preview),
                        "existing_members_total": existing_members_total,
                        "existing_groups": existing_groups_preview,
                        "resolution_mode": "auto",
                        "layers_used": layers_used,
                        "effective_bucket_count": len(buckets),
                        "strategy": strategy_applied,
                        "used_fallback": used_fallback,
                        "fallback_reason": fallback_reason,
                        "buckets_total": len(buckets),
                        "buckets_applied": len(buckets_to_apply),
                        "size_min": min(sizes) if sizes else 0,
                        "size_max": max(sizes) if sizes else 0,
                    },
                },
                request,
                competicio.id,
            )
        )

    if has_rotacions:
        group_maps = get_group_maps(competicio)
        programmed_group_ids = get_programmed_group_ids(competicio)
        blocked_groups = []
        for row in existing_groups_preview:
            if str(row.get("impact_kind") or "") != "removed":
                continue
            group_num = normalize_positive_int(row.get("group_num"))
            if not group_num:
                continue
            group = group_maps["by_display_num"].get(group_num)
            if group is None or group.id not in programmed_group_ids:
                continue
            blocked_groups.append(group)
        if blocked_groups:
            return HttpResponseBadRequest(_message_for_emptied_programmed_groups(blocked_groups))

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    updates = list(objs)

    _assign_group_sizes_in_order(objs, sizes, max_grup)

    with transaction.atomic():
        qs.filter(id__in=target_ids).update(grup=None)
        Inscripcio.objects.bulk_update(updates, ["grup"], batch_size=500)
        sync_stable_groups_from_legacy(competicio)
        _persist_group_suggested_names(competicio, preview_groups)

    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="groups_from_sort",
        action_label="Crear grups des del panell",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )

    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "updated": len(updates),
                "groups_created": len(sizes),
                "buckets_total": len(buckets),
                "buckets_applied": len(buckets_to_apply),
                "stack_used": partition_codes,
                "resolution_mode": "auto",
                "layers_used": layers_used,
                "effective_bucket_count": len(buckets),
                "strategy": strategy_applied,
                "used_fallback": used_fallback,
                "fallback_reason": fallback_reason,
                "size_min": min(sizes) if sizes else 0,
                "size_max": max(sizes) if sizes else 0,
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_reorder(request, pk):
    """
    Rep:
      {
        "ids": [<id1>, <id2>, ...],   # ordre final després del drag
        "moved_id": <id>,            # el registre arrossegat
        "new_index": <int>,          # posició nova (0-based) dins ids
        "target_group": <int|null>   # grup inferit per header visual (opcional)
      }

    Guarda ordre_sortida = 1..N
    I (NOU) només pel registre mogut:
      - si arriba target_group, adopta aquest grup (header visual mana)
      - si no, conserva el fallback històric (grup del registre immediatament superior)
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
        ids = payload.get("ids", [])
        moved_id = payload.get("moved_id", None)
        new_index = payload.get("new_index", None)
        target_group = payload.get("target_group", None)
        reorder_mode = str(payload.get("mode") or "visual").strip().lower()
        raw_filters = payload.get("filters")
        raw_group_by = payload.get("group_by")

        if not isinstance(ids, list) or not ids:
            return HttpResponseBadRequest("Payload invàlid")
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    wanted = [int(x) for x in ids if str(x).isdigit()]
    if not wanted:
        return HttpResponseBadRequest("IDs buits")
    if reorder_mode not in ("visual", "group_edit"):
        return HttpResponseBadRequest("mode invàlid")
    competicio = get_object_or_404(Competicio, pk=pk)

    # Validació moved_id / new_index (opcional però recomanada)
    if moved_id is not None:
        try:
            moved_id = int(moved_id)
        except Exception:
            return HttpResponseBadRequest("moved_id invàlid")

    if new_index is not None:
        try:
            new_index = int(new_index)
        except Exception:
            return HttpResponseBadRequest("new_index invàlid")

    if target_group in ("", None):
        target_group = None
    else:
        try:
            target_group = int(target_group)
        except Exception:
            return HttpResponseBadRequest("target_group invàlid")
        if target_group <= 0:
            return HttpResponseBadRequest("target_group invàlid")

    # Ens assegurem que només reordenem inscripcions d'aquesta competició
    qs = Inscripcio.objects.filter(competicio=competicio, id__in=wanted)
    found = set(qs.values_list("id", flat=True))
    if set(wanted) != found:
        return HttpResponseBadRequest("IDs no vàlids per aquesta competició")

    # Diccionari id->grup (estat actual abans de canviar)
    id_to_group = dict(qs.values_list("id", "grup"))
    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    group_maps = get_group_maps(competicio)
    groups_by_display = group_maps["by_display_num"]

    with transaction.atomic():
        # 1) actualitza ordre_sortida (bulk) per evitar N updates i timeouts en llistes grans
        target_order_by_id = {ins_id: idx for idx, ins_id in enumerate(wanted, start=1)}
        order_updates = []
        for obj in qs.only("id", "ordre_sortida"):
            next_ord = target_order_by_id.get(obj.id)
            if next_ord is None:
                continue
            if obj.ordre_sortida != next_ord:
                obj.ordre_sortida = next_ord
                order_updates.append(obj)
        if order_updates:
            Inscripcio.objects.bulk_update(order_updates, ["ordre_sortida"], batch_size=500)

        # 2) en mode explicit de grups, el registre mogut pot canviar de grup.
        if reorder_mode == "group_edit" and moved_id is not None and new_index is not None and moved_id in wanted:
            next_group = None
            should_update_group = False
            if target_group is not None:
                next_group = target_group
                should_update_group = True
            elif new_index > 0:
                prev_id = wanted[new_index - 1]
                next_group = id_to_group.get(prev_id)
                should_update_group = True

            if should_update_group:
                target_group_obj = groups_by_display.get(next_group)
                moved = (
                    Inscripcio.objects
                    .select_related("grup_competicio")
                    .get(id=moved_id, competicio=competicio)
                )
                if target_group_obj is not None:
                    move_inscripcio_to_group(moved, target_group_obj)
                else:
                    old_group = moved.grup_competicio
                    Inscripcio.objects.filter(id=moved_id).update(
                        grup=None,
                        grup_competicio=None,
                        ordre_competicio=None,
                    )
                    compact_competition_order_for_group(old_group)

    # Rebase de l'stack d'ordenacio del context actual quan el conjunt coincideix.
    _sync_group_names_for_competicio(competicio)

    filters = _normalize_sort_filters(raw_filters)
    allowed_group_fields = get_allowed_group_fields(competicio)
    allowed_group_codes = {f["code"] for f in allowed_group_fields}
    selected_group_codes_context = _normalize_sort_group_by(
        raw_group_by,
        allowed_group_codes,
        fallback_group_by=competicio.group_by_default or [],
    )
    context_key = build_inscripcions_sort_context_key(
        competicio.id,
        filters=filters,
        group_by=selected_group_codes_context,
    )
    filtered_ids = list(
        _build_inscripcions_filtered_qs(competicio, filters)
        .order_by("ordre_sortida", "id")
        .values_list("id", flat=True)
    )
    if len(filtered_ids) == len(wanted) and set(filtered_ids) == set(wanted):
        reconcile_inscripcions_sort_context_state(request, context_key, wanted)

    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="reorder" if reorder_mode == "visual" else "move_group_member",
        action_label="Reordenar inscripcions" if reorder_mode == "visual" else "Editar contingut de grups",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )

    return JsonResponse(with_inscripcions_history_payload({"ok": True}, request, competicio.id))


@require_POST
@csrf_protect
def inscripcions_save_group_competition_order(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    try:
        group_num = int(payload.get("group_num"))
    except Exception:
        return HttpResponseBadRequest("group_num invàlid")
    ordered_ids = payload.get("ids")
    if not isinstance(ordered_ids, list) or not ordered_ids:
        return HttpResponseBadRequest("ids invàlids")

    group = get_group_for_display_num(competicio, group_num)
    if group is None:
        return HttpResponseBadRequest("group_num invàlid")

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    try:
        updated = save_group_competition_order(group, ordered_ids)
    except ValueError as exc:
        return HttpResponseBadRequest(str(exc))

    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="set_group_competition_order",
        action_label="Desar ordre de competicio",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )
    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "group_num": group_num,
                "updated": updated,
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def inscripcions_merge_tabs(request, pk):
    c = get_object_or_404(Competicio, pk=pk)
    before_snapshot = capture_inscripcions_history_snapshot(request, c)

    try:
        payload = json.loads(request.body.decode("utf-8"))
        group_field = payload.get("group_field")   # IMPORTANT: ha de ser la "signature" (grouping_sig)
        source_key = payload.get("source_key")
        target_key = payload.get("target_key")
    except Exception:
        return HttpResponseBadRequest("JSON invàlid")

    if not group_field:
        return HttpResponseBadRequest("group_field buit")
    if not source_key or not target_key or source_key == target_key:
        return HttpResponseBadRequest("claus invàlides")

    merges = c.tab_merges or {}
    lst = merges.get(group_field, [])

    def normalize_key_to_simple_list(k: str) -> list[str]:
        """
        Retorna SEMPRE una llista de 'simple keys' (strings JSON del tipus ["A","B"...]).
        - Si k és simple: retorna [k]
        - Si k és merged: retorna [sub1, sub2, ...] on cada subX és un string JSON d'una clau simple
        """
        try:
            v = json.loads(k)
        except Exception:
            return [k]

        # merged: ['["..."]','["..."]', ...]  (cada element és un JSON de llista)
        if (
            isinstance(v, list) and v
            and all(isinstance(x, str) for x in v)
            and all(x.strip().startswith("[") for x in v)
        ):
            return v

        # simple: ["ALEVI","FEMENI", ...]
        if isinstance(v, list):
            return [k]

        return [k]


    s_list = normalize_key_to_simple_list(source_key)
    t_list = normalize_key_to_simple_list(target_key)

    # Unim t + s preservant ordre i dedup
    desired = []
    for x in (t_list + s_list):
        if x not in desired:
            desired.append(x)

    # Si ja hi ha merges que contenen alguna d'aquestes keys, els absorbim
    consumed_idx = []
    merged_all = []
    for i, g in enumerate(lst):
        # g és una llista de simple keys
        if any(x in g for x in desired):
            merged_all.extend(g)
            consumed_idx.append(i)

    for i in sorted(consumed_idx, reverse=True):
        lst.pop(i)

    merged_all.extend(desired)

    final = []
    for x in merged_all:
        if x not in final:
            final.append(x)

    lst.append(final)
    merges[group_field] = lst
    c.tab_merges = merges
    c.save(update_fields=["tab_merges"])
    record_inscripcions_history_entry(
        request,
        c,
        action_type="merge_tabs",
        action_label="Fusionar pestanyes",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, c),
    )

    return JsonResponse(
        with_inscripcions_history_payload(
            {"ok": True, "merged": final},
            request,
            c.id,
        )
    )



