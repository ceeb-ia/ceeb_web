import json
import math
import random
from collections import OrderedDict, defaultdict
from datetime import date, datetime
from io import BytesIO

from django.contrib import messages
from django.db import transaction
from django.db.models import Case, IntegerField, Max, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import FormView, ListView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import ImportInscripcionsExcelForm
from .models import Competicio, GrupCompeticio, Inscripcio
from .models_rotacions import RotacioAssignacio
from .services.competition_groups import (
    ensure_group_for_display_num,
    get_group_for_display_num,
    get_group_maps,
    get_programmed_group_ids,
    next_group_display_num,
    normalize_positive_int,
    sync_competicio_group_names_view,
)
from .services.equip_contexts import NATIVE_EQUIP_CONTEXT_CODE, get_contextual_assignment_map, resolve_inscripcio_equip
from .services.import_excel import importar_inscripcions_excel
from .services.inscripcions.history import (
    capture_inscripcions_history_snapshot,
    get_inscripcions_history_state,
    record_inscripcions_history_entry,
)
from .services.inscripcions.queries import (
    LEGACY_SORT_KEY_MAP,
    _build_inscripcions_filtered_qs,
    _custom_sort_token_key,
    _normalize_custom_sort_order,
    _normalize_custom_sort_token,
    _resolve_sort_field_runtime,
    annotate_inscripcions_queryset_for_group_codes,
    get_allowed_group_fields,
    get_available_sort_fields,
    get_inscripcio_value,
    get_request_inscripcio_filters,
)
from .services.inscripcions.shared import INSCRIPCIONS_SORT_STACK_SESSION_KEY


BUILTIN_EXCEL_FIELDS = [
    {"code": "nom_i_cognoms", "label": "Nom i cognoms", "kind": "builtin"},
    {"code": "document", "label": "DNI", "kind": "builtin"},
    {"code": "sexe", "label": "Sexe", "kind": "builtin"},
    {"code": "data_naixement", "label": "Data naixement", "kind": "builtin"},
    {"code": "entitat", "label": "Entitat", "kind": "builtin"},
    {"code": "categoria", "label": "Categoria", "kind": "builtin"},
    {"code": "subcategoria", "label": "Subcategoria", "kind": "builtin"},
    {"code": "grup", "label": "Grup", "kind": "builtin"},
    {"code": "ordre_sortida", "label": "Ordre", "kind": "builtin"},
]

LEGACY_EXCEL_COL_MAP = {
    "nom": "nom_i_cognoms",
    "dni": "document",
    "naixement": "data_naixement",
    "ordre": "ordre_sortida",
}


def _reserved_inscripcio_codes():
    out = set()
    for field in Inscripcio._meta.concrete_fields:
        name = str(getattr(field, "name", "") or "").strip()
        attname = str(getattr(field, "attname", "") or "").strip()
        if name:
            out.add(name)
        if attname:
            out.add(attname)
    return out


def _normalize_schema_extra_code(code: str, reserved_codes=None):
    code = str(code or "").strip()
    if not code:
        return code
    if code.startswith("excel__"):
        return code
    reserved = reserved_codes if reserved_codes is not None else _reserved_inscripcio_codes()
    if code in reserved:
        return f"excel__{code}"
    return code


def _excel_schema_codes(competicio):
    schema = competicio.inscripcions_schema or {}
    cols = schema.get("columns") or []
    if not isinstance(cols, list):
        return set()
    reserved = _reserved_inscripcio_codes()
    out = set()
    for col in cols:
        if not isinstance(col, dict):
            continue
        code = col.get("code")
        if not code:
            continue
        kind = col.get("kind") or "extra"
        if kind == "extra":
            code = _normalize_schema_extra_code(code, reserved)
        out.add(code)
    return out


def _label_with_source(label: str, source: str):
    if source == "excel":
        suffix = "Excel"
    elif source == "derived":
        suffix = "Derivada"
    else:
        suffix = "Nativa"
    return f"{label} ({suffix})"


def _norm_val(value):
    return "__NULL__" if value in (None, "") else str(value)


def _resolve_base_equip_for_inscripcio(obj, assignment_map=None):
    if obj is None:
        return None
    if assignment_map is None and getattr(obj, "_base_equip_cache_ready", False):
        return getattr(obj, "_base_equip_cache", None)
    if assignment_map is None and getattr(obj, "competicio", None) is None and getattr(obj, "competicio_id", None):
        try:
            obj._competicio_cache = Competicio.objects.filter(id=obj.competicio_id).first()
        except Exception:
            pass
    equip = resolve_inscripcio_equip(
        obj,
        context_code=NATIVE_EQUIP_CONTEXT_CODE,
        fallback=None,
        assignment_map=assignment_map,
    )
    try:
        obj._base_equip_cache = equip
        obj._base_equip_cache_ready = True
        obj._base_equip_id_cache = getattr(equip, "id", None)
        obj._base_equip_name_cache = str(getattr(equip, "nom", "") or "").strip()
    except Exception:
        pass
    return equip


def _attach_base_equip_runtime(records):
    rows = list(records or [])
    if not rows:
        return {}
    competicio = next(
        (getattr(obj, "competicio", None) for obj in rows if getattr(obj, "competicio", None) is not None),
        None,
    )
    if competicio is None:
        competicio_id = next(
            (getattr(obj, "competicio_id", None) for obj in rows if getattr(obj, "competicio_id", None)),
            None,
        )
        if competicio_id:
            competicio = Competicio.objects.filter(id=competicio_id).first()
    if competicio is None:
        return {}
    assignment_map = get_contextual_assignment_map(competicio, rows, NATIVE_EQUIP_CONTEXT_CODE)
    for obj in rows:
        _resolve_base_equip_for_inscripcio(obj, assignment_map=assignment_map)
    return assignment_map


def get_available_excel_columns(competicio):
    out = []
    seen = set()
    excel_codes = _excel_schema_codes(competicio)
    reserved = _reserved_inscripcio_codes()
    for field in BUILTIN_EXCEL_FIELDS:
        if field["code"] in seen:
            continue
        source = "excel" if field["code"] in excel_codes else "native"
        out.append({**field, "source": source, "ui_label": _label_with_source(field["label"], source)})
        seen.add(field["code"])

    schema = competicio.inscripcions_schema or {}
    cols = schema.get("columns") or []
    if isinstance(cols, list):
        for col in cols:
            if not isinstance(col, dict):
                continue
            code = col.get("code")
            if not code or code in seen:
                continue
            kind = col.get("kind") or "extra"
            if kind != "extra":
                continue
            code = _normalize_schema_extra_code(code, reserved)
            if code in seen:
                continue
            label = col.get("label") or code
            out.append(
                {"code": code, "label": label, "kind": "extra", "source": "excel", "ui_label": _label_with_source(label, "excel")}
            )
            seen.add(code)
    return out


def get_excel_export_value(obj, code):
    value = get_inscripcio_value(obj, code)
    if code == "data_naixement":
        return value.strftime("%d/%m/%Y") if value else "-"
    if value in (None, ""):
        return "-"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _build_sort_field_runtime_context(records, sort_code):
    code = str(sort_code or "").strip()
    if code == "equip":
        return {"base_assignment_map": _attach_base_equip_runtime(records)}
    if code != "grup":
        return {}
    competicio_id = next((getattr(obj, "competicio_id", None) for obj in records or [] if getattr(obj, "competicio_id", None)), None)
    if not competicio_id:
        return {}
    competicio = Competicio.objects.filter(id=competicio_id).first()
    if competicio is None:
        return {}
    return {"group_maps": get_group_maps(competicio)}


def sort_records_by_field_stable(records, sort_code, descending=False, custom_rank_map=None):
    custom_map = custom_rank_map if isinstance(custom_rank_map, dict) else {}
    custom_enabled = bool(custom_map)
    custom_filled = []
    fallback_filled = []
    empty = []
    context = _build_sort_field_runtime_context(records, sort_code)
    for obj in records:
        runtime = _resolve_sort_field_runtime(obj, sort_code, context=context)
        token = runtime.get("token") or ""
        if not token:
            empty.append(obj)
            continue
        if custom_enabled:
            key = _custom_sort_token_key(token)
            if key in custom_map:
                custom_filled.append((obj, custom_map[key]))
                continue
        fallback_filled.append((obj, runtime.get("sort_scalar")))
    custom_filled.sort(key=lambda item: item[1], reverse=descending)
    fallback_filled.sort(key=lambda item: item[1], reverse=descending)
    return [obj for (obj, _rank) in custom_filled] + [obj for (obj, _value) in fallback_filled] + empty


def arrow_positions(n: int) -> list[int]:
    if n <= 0:
        return []
    seq = []
    if n % 2 == 0:
        left = n // 2 - 1
        right = n // 2
        while left >= 0 or right < n:
            if left >= 0:
                seq.append(left)
                left -= 1
            if right < n:
                seq.append(right)
                right += 1
    else:
        center = n // 2
        seq.append(center)
        step = 1
        while center - step >= 0 or center + step < n:
            if center - step >= 0:
                seq.append(center - step)
            if center + step < n:
                seq.append(center + step)
            step += 1
    return seq


def shuffle_ordre_sortida(qs):
    ids = list(qs.values_list("id", flat=True))
    random.shuffle(ids)
    with transaction.atomic():
        for idx, ins_id in enumerate(ids, start=1):
            Inscripcio.objects.filter(id=ins_id).update(ordre_sortida=idx)


def assign_groups_balanced(objs, size, start_group_num):
    n = len(objs)
    if n == 0:
        return start_group_num
    k = math.ceil(n / size)
    base = n // k
    rem = n % k
    idx = 0
    group_num = start_group_num
    for group_idx in range(k):
        group_num += 1
        this_size = base + (1 if group_idx < rem else 0)
        for _ in range(this_size):
            objs[idx].grup = group_num
            idx += 1
    return group_num


def assign_groups_k(objs, k, start_group_num):
    n = len(objs)
    if n == 0 or k <= 0:
        return start_group_num
    k = min(k, n)
    base = n // k
    rem = n % k
    idx = 0
    group_num = start_group_num
    for group_idx in range(k):
        group_num += 1
        this_size = base + (1 if group_idx < rem else 0)
        for _ in range(this_size):
            objs[idx].grup = group_num
            idx += 1
    return group_num


def recalcular_ordre_sortida(qs, group_codes):
    records = list(qs.order_by("ordre_sortida", "id"))

    def sort_key(obj):
        group_values = tuple(_norm_val(get_inscripcio_value(obj, code)) for code in group_codes)
        previous = obj.ordre_sortida if obj.ordre_sortida is not None else 10**12
        return (group_values, previous, obj.id)

    records.sort(key=sort_key)
    with transaction.atomic():
        for idx, obj in enumerate(records, start=1):
            if obj.ordre_sortida != idx:
                Inscripcio.objects.filter(id=obj.id).update(ordre_sortida=idx)


def set_competicio_custom_sort_order_values(competicio, sort_code, raw_values=None, clear=False, allowed_sort_codes=None):
    code_raw = str(sort_code or "").strip()
    code = LEGACY_SORT_KEY_MAP.get(code_raw, code_raw)
    if not code:
        raise ValueError("sort_key invalid")
    if allowed_sort_codes is not None and code not in set(allowed_sort_codes):
        raise ValueError("sort_key invalid")

    values = [] if clear else _normalize_custom_sort_order(raw_values)
    view_cfg = dict(competicio.inscripcions_view or {})
    custom_map = view_cfg.get("custom_sort_orders")
    if not isinstance(custom_map, dict):
        custom_map = {}
    custom_map = dict(custom_map)

    if values:
        custom_map[code] = values
    else:
        custom_map.pop(code, None)

    if custom_map:
        view_cfg["custom_sort_orders"] = custom_map
    else:
        view_cfg.pop("custom_sort_orders", None)

    competicio.inscripcions_view = view_cfg
    competicio.save(update_fields=["inscripcions_view"])
    return values


def _split_custom_sort_tokens(custom_tokens, available_token_keys):
    active = []
    stale = []
    available = set(available_token_keys or set())
    seen = set()
    for raw in custom_tokens or []:
        token = _normalize_custom_sort_token(raw)
        if not token:
            continue
        key = _custom_sort_token_key(token)
        if not key or key in seen:
            continue
        seen.add(key)
        if key in available:
            active.append(token)
        else:
            stale.append(token)
    return active, stale


def _normalize_competition_order_tail_flag(raw_value):
    if isinstance(raw_value, bool):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return bool(raw_value)
    token = str(raw_value or "").strip().lower()
    return token in {"1", "true", "yes", "on"}


def clear_inscripcions_sort_state_for_competicio(request, competicio_id):
    prefix = f"{competicio_id}||"
    store = request.session.get(INSCRIPCIONS_SORT_STACK_SESSION_KEY)
    if not isinstance(store, dict):
        return
    changed = False
    for key in list(store.keys()):
        if isinstance(key, str) and key.startswith(prefix):
            store.pop(key, None)
            changed = True
    if changed:
        request.session[INSCRIPCIONS_SORT_STACK_SESSION_KEY] = store
        request.session.modified = True


def _normalize_group_names_map(raw_group_names):
    out = {}
    if not isinstance(raw_group_names, dict):
        return out
    for raw_group, raw_label in raw_group_names.items():
        try:
            group_num = int(str(raw_group).strip())
        except Exception:
            continue
        if group_num <= 0:
            continue
        label = str(raw_label or "").strip()
        if not label:
            continue
        out[group_num] = label
    return out


def _sync_group_names_for_competicio(competicio):
    sync_competicio_group_names_view(competicio)


def _persist_group_suggested_names(competicio, preview_groups):
    suggested_by_display_num = {}
    for row in preview_groups or []:
        group_num = normalize_positive_int(row.get("group_num"))
        suggested_name = str(row.get("suggested_name") or "").strip()
        if not group_num or not suggested_name:
            continue
        suggested_by_display_num[group_num] = suggested_name

    if not suggested_by_display_num:
        sync_competicio_group_names_view(competicio)
        return 0

    groups = list(
        GrupCompeticio.objects.filter(
            competicio=competicio,
            display_num__in=list(suggested_by_display_num.keys()),
        )
    )
    updates = []
    for group in groups:
        suggested_name = suggested_by_display_num.get(group.display_num, "")
        if not suggested_name or group.nom == suggested_name:
            continue
        group.nom = suggested_name
        updates.append(group)

    if updates:
        GrupCompeticio.objects.bulk_update(updates, ["nom"], batch_size=200)

    sync_competicio_group_names_view(competicio)
    return len(updates)


def renumber_groups_for_competicio(competicio):
    _sync_group_names_for_competicio(competicio)


def competicio_has_rotacions(competicio):
    return RotacioAssignacio.objects.filter(competicio=competicio).exists()


def _resolve_group_id_for_inscripcio(inscripcio, groups_by_display_num):
    group_id = getattr(inscripcio, "grup_competicio_id", None)
    if group_id:
        return int(group_id)
    legacy_group_num = normalize_positive_int(getattr(inscripcio, "grup", None))
    if not legacy_group_num:
        return None
    group = groups_by_display_num.get(legacy_group_num)
    if not group:
        return None
    return int(group.id)


def _programmed_groups_emptied_by_move(competicio, target_ids):
    target_ids = {
        int(ins_id)
        for ins_id in (target_ids or [])
        if normalize_positive_int(ins_id)
    }
    if not target_ids:
        return []
    programmed_group_ids = get_programmed_group_ids(competicio)
    if not programmed_group_ids:
        return []

    group_maps = get_group_maps(competicio)
    groups_by_display_num = group_maps["by_display_num"]
    groups_by_id = group_maps["by_id"]
    selected = list(
        Inscripcio.objects.filter(competicio=competicio, id__in=target_ids)
        .select_related("grup_competicio")
        .only("id", "grup", "grup_competicio")
    )
    moving_counts = defaultdict(int)
    for inscripcio in selected:
        group_id = _resolve_group_id_for_inscripcio(inscripcio, groups_by_display_num)
        if group_id:
            moving_counts[group_id] += 1
    if not moving_counts:
        return []

    total_counts = defaultdict(int)
    all_members = (
        Inscripcio.objects.filter(competicio=competicio)
        .select_related("grup_competicio")
        .only("id", "grup", "grup_competicio")
    )
    for inscripcio in all_members:
        group_id = _resolve_group_id_for_inscripcio(inscripcio, groups_by_display_num)
        if group_id:
            total_counts[group_id] += 1

    blocked = []
    for group_id, moved_count in moving_counts.items():
        if group_id not in programmed_group_ids:
            continue
        if moved_count >= total_counts.get(group_id, 0):
            group = groups_by_id.get(group_id)
            if group is not None:
                blocked.append(group)
    blocked.sort(key=lambda group: (group.display_num, group.id))
    return blocked


def _message_for_emptied_programmed_groups(groups):
    if not groups:
        return ""
    labels = []
    for group in groups:
        label = str(getattr(group, "nom", "") or "").strip()
        if not label:
            label = f"Grup {getattr(group, 'display_num', '?')}"
        labels.append(label)
    return f"No es pot deixar buit un grup inclos al programa de rotacions: {', '.join(labels)}."


def sync_stable_groups_from_legacy(competicio):
    view_cfg = competicio.inscripcions_view or {}
    group_names = _normalize_group_names_map(view_cfg.get("group_names") or {})
    live_display_nums = list(
        Inscripcio.objects.filter(competicio=competicio, grup__isnull=False)
        .order_by("grup")
        .values_list("grup", flat=True)
        .distinct()
    )
    live_display_nums = [int(num) for num in live_display_nums if isinstance(num, int) and num > 0]

    group_maps = get_group_maps(competicio)
    groups_by_display = group_maps["by_display_num"]
    touched_group_ids = set()
    updates = []
    with transaction.atomic():
        for display_num in live_display_nums:
            group = groups_by_display.get(display_num)
            name = str(group_names.get(display_num) or "").strip()
            if group is None:
                group = ensure_group_for_display_num(competicio, display_num, name=name)
                groups_by_display[display_num] = group
            else:
                fields = []
                if not group.actiu:
                    group.actiu = True
                    fields.append("actiu")
                if name != group.nom:
                    group.nom = name
                    fields.append("nom")
                if fields:
                    group.save(update_fields=fields)
            touched_group_ids.add(group.id)

        stale_groups = [
            group
            for group in get_group_maps(competicio)["groups"]
            if group.id not in touched_group_ids
        ]
        if stale_groups:
            GrupCompeticio.objects.filter(id__in=[group.id for group in stale_groups]).update(actiu=False)

        counters = defaultdict(int)
        qs = (
            Inscripcio.objects.filter(competicio=competicio)
            .order_by("grup", "ordre_sortida", "id")
            .only("id", "grup", "grup_competicio", "ordre_competicio")
        )
        for inscripcio in qs:
            display_num = getattr(inscripcio, "grup", None)
            group = groups_by_display.get(display_num)
            next_group_id = getattr(group, "id", None)
            next_comp_order = None
            if next_group_id:
                counters[next_group_id] += 1
                next_comp_order = counters[next_group_id]
            if inscripcio.grup_competicio_id != next_group_id or inscripcio.ordre_competicio != next_comp_order:
                inscripcio.grup_competicio_id = next_group_id
                inscripcio.ordre_competicio = next_comp_order
                updates.append(inscripcio)
        if updates:
            Inscripcio.objects.bulk_update(updates, ["grup_competicio", "ordre_competicio"], batch_size=500)

    sync_competicio_group_names_view(competicio)


class InscripcionsImportExcelView(FormView):
    template_name = "competicio/inscripcions_import.html"
    form_class = ImportInscripcionsExcelForm

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = self.competicio
        return ctx

    def form_valid(self, form):
        fitxer = form.cleaned_data["fitxer"]
        sheet = form.cleaned_data.get("sheet") or ""
        result = importar_inscripcions_excel(fitxer, self.competicio, sheet)
        summary = (
            f"Full: {result['full']} | Creats: {result['creats']} | "
            f"Actualitzats: {result['actualitzats']} | Ignorats: {result['ignorats']} | "
            f"Ambiguos: {result.get('ambiguos', 0)} | Errors: {result.get('errors', 0)}"
        )
        if int(result.get("errors", 0) or 0) > 0:
            messages.warning(self.request, f"Importació parcial amb incidències. {summary}")
        else:
            messages.success(self.request, f"Importació OK. {summary}")
        warnings = result.get("warnings") or []
        if warnings:
            parts = []
            for warning in warnings:
                code = str(warning.get("code") or "").strip()
                remapped = str(warning.get("remapped_code") or warning.get("suggested_code") or "").strip()
                if code and remapped:
                    parts.append(f"{code} -> {remapped}")
                elif code:
                    parts.append(code)
            if parts:
                messages.warning(
                    self.request,
                    "S'han detectat columnes d'Excel amb noms reservats i s'han remapejat automaticament "
                    f"({', '.join(parts)}).",
                )
        noms_competicio_excel = result.get("noms_competicio_excel") or []
        if len(noms_competicio_excel) > 1:
            messages.warning(
                self.request,
                "L'Excel conté múltiples noms de competició detectats: " + ", ".join(str(name) for name in noms_competicio_excel),
            )
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("inscripcions_list", kwargs={"pk": self.competicio.pk})


class InscripcionsListView(ListView):
    template_name = "competicio/inscripcions_list.html"
    context_object_name = "records"
    paginate_by = 25

    def dispatch(self, request, *args, **kwargs):
        self.competicio = get_object_or_404(Competicio, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return self.get_queryset_base_filtrada().order_by("ordre_sortida", "id")

    def get_queryset_base_filtrada(self):
        return _build_inscripcions_filtered_qs(
            self.competicio,
            get_request_inscripcio_filters(self.request, competicio=self.competicio),
        )

    def get(self, request, *args, **kwargs):
        allowed = get_allowed_group_fields(self.competicio)
        allowed_codes = {field["code"] for field in allowed}
        sort_fields = get_available_sort_fields(self.competicio)
        sort_allowed_codes = {field["code"] for field in sort_fields}
        default_sort_code = "nom_i_cognoms" if "nom_i_cognoms" in sort_allowed_codes else (next(iter(sort_allowed_codes), ""))

        def resolve_sort_key(raw_key):
            key = LEGACY_SORT_KEY_MAP.get((raw_key or "").strip(), (raw_key or "").strip())
            if key in sort_allowed_codes:
                return key
            return default_sort_code

        def get_active_group_codes():
            selected = [group for group in request.GET.getlist("group_by") if group in allowed_codes]
            if not selected:
                selected = [group for group in (self.competicio.group_by_default or []) if group in allowed_codes]
            return selected

        def simple_key_for_obj(obj, group_codes):
            values = [_norm_val(get_inscripcio_value(obj, code)) for code in group_codes]
            return json.dumps(values, ensure_ascii=False)

        def pretty_label_from_simple_key(simple_key):
            try:
                values = json.loads(simple_key)
                return " · ".join("(Sense valor)" if value in (None, "", "__NULL__") else str(value) for value in values)
            except Exception:
                return simple_key

        def _capture_history_snapshot():
            return capture_inscripcions_history_snapshot(request, self.competicio)

        def _record_history(action_type, action_label, before_snapshot):
            record_inscripcions_history_entry(
                request,
                self.competicio,
                action_type=action_type,
                action_label=action_label,
                before_snapshot=before_snapshot,
                after_snapshot=_capture_history_snapshot(),
            )

        if request.GET.get("export_excel") == "1":
            available_excel = get_available_excel_columns(self.competicio)
            by_code = {column["code"]: column for column in available_excel}
            selected_codes = []
            for raw in request.GET.getlist("excel_cols"):
                code = LEGACY_EXCEL_COL_MAP.get(raw, raw)
                if code in by_code and code not in selected_codes:
                    selected_codes.append(code)
            if not selected_codes:
                selected_codes = list(by_code.keys())
            if not selected_codes and "nom_i_cognoms" in by_code:
                selected_codes = ["nom_i_cognoms"]

            columns = [(by_code[code]["label"], code) for code in selected_codes]
            group_codes = get_active_group_codes()
            grouping_sig = "|".join(group_codes) if group_codes else ""
            qs_base = self.get_queryset_base_filtrada()

            wb = Workbook()
            ws = wb.active
            ws.title = "Inscripcions"
            title_font = Font(bold=True, size=12)
            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="E9EEF7")
            group_fill = PatternFill("solid", fgColor="DDE7FF")

            def write_table_header(row_num):
                for col_idx, (label, _code) in enumerate(columns, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=label)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(vertical="center")

            def write_row(row_num, obj):
                for col_idx, (_label, code) in enumerate(columns, start=1):
                    ws.cell(row=row_num, column=col_idx, value=get_excel_export_value(obj, code))

            qs_all = qs_base.annotate(
                grup_null=Case(
                    When(grup__isnull=True, then=1),
                    default=0,
                    output_field=IntegerField(),
                )
            ).order_by("grup_null", "grup", "ordre_sortida", "id")

            merges = (self.competicio.tab_merges or {}).get(grouping_sig, [])
            merge_map = {}
            merged_label_map = {}
            for group_keys in merges:
                group_tuple = tuple(group_keys)
                for group_key in group_keys:
                    merge_map[group_key] = group_tuple
                merged_key = json.dumps(list(group_tuple), ensure_ascii=False)
                merged_label_map[merged_key] = " + ".join(pretty_label_from_simple_key(group_key) for group_key in group_tuple)

            def tab_label_for_obj(obj):
                if not group_codes:
                    return "Sense agrupació"
                simple = simple_key_for_obj(obj, group_codes)
                merged_tuple = merge_map.get(simple)
                if merged_tuple:
                    merged_key = json.dumps(list(merged_tuple), ensure_ascii=False)
                    return merged_label_map.get(merged_key, merged_key)
                return pretty_label_from_simple_key(simple)

            sentinel = object()
            current_group = sentinel
            buffer = []
            row = 1

            def flush_group(objs, group_num):
                nonlocal row
                if not objs:
                    return
                tab_label = tab_label_for_obj(objs[0])
                group_title = f"{tab_label} · Sense grup" if group_num is None else f"{tab_label} · Grup {group_num}"
                ws.cell(row=row, column=1, value=group_title).font = title_font
                ws.cell(row=row, column=1).fill = group_fill
                ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
                row += 1
                write_table_header(row)
                row += 1
                for obj in objs:
                    write_row(row, obj)
                    row += 1
                row += 1

            for obj in qs_all:
                group_num = obj.grup
                if current_group is sentinel:
                    current_group = group_num
                if group_num != current_group:
                    flush_group(buffer, current_group)
                    buffer = []
                    current_group = group_num
                buffer.append(obj)
            flush_group(buffer, current_group)

            for idx, (label, _code) in enumerate(columns, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = max(12, min(35, len(label) + 4))

            out = BytesIO()
            wb.save(out)
            out.seek(0)
            filename = f"inscripcions_competicio_{self.competicio.pk}.xlsx"
            response = HttpResponse(
                out.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        if request.GET.get("clear_groups") == "1":
            if competicio_has_rotacions(self.competicio):
                messages.error(request, "No es poden esborrar grups mentre hi ha rotacions actives.")
                query = request.GET.copy()
                query.pop("clear_groups", None)
                return redirect(f"{request.path}?{query.urlencode()}")
            qs = self.get_queryset_base_filtrada()
            before_snapshot = _capture_history_snapshot()
            with transaction.atomic():
                qs.update(grup=None, grup_competicio=None, ordre_competicio=None)
                GrupCompeticio.objects.filter(competicio=self.competicio).update(actiu=False)
            clear_inscripcions_sort_state_for_competicio(request, self.competicio.id)
            _record_history("clear_groups", "Esborrar grups", before_snapshot)
            query = request.GET.copy()
            query.pop("clear_groups", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        if request.GET.get("make_independent_group") == "1":
            has_rotacions = competicio_has_rotacions(self.competicio)
            level_token = request.GET.get("lvl")

            group_codes = get_active_group_codes()
            if not group_codes:
                messages.error(request, "No hi ha agrupacio activa per poder crear un grup independent.")
                query = request.GET.copy()
                for key in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                    query.pop(key, None)
                return redirect(f"{request.path}?{query.urlencode()}")

            value_1 = request.GET.get("v1")
            value_2 = request.GET.get("v2")
            value_3 = request.GET.get("v3")

            def build_filter_ids_from_vals(values):
                qs_base = self.get_queryset_base_filtrada()
                builtin_fields = [code for code in group_codes if hasattr(Inscripcio, code)]
                records = list(qs_base.only("id", "extra", *builtin_fields))
                ids = []
                for record in records:
                    ok = True
                    for code, value in zip(group_codes, values):
                        if _norm_val(get_inscripcio_value(record, code)) != value:
                            ok = False
                            break
                    if ok:
                        ids.append(record.id)
                return ids

            if level_token == "g1" and value_1 and value_1.strip().startswith("["):
                try:
                    parsed = json.loads(value_1)
                except Exception:
                    parsed = None

                base_qs = self.get_queryset_base_filtrada()
                if isinstance(parsed, list) and parsed and all(isinstance(item, str) and item.strip().startswith("[") for item in parsed):
                    before_snapshot = _capture_history_snapshot()
                    all_ids = []
                    for simple_key in parsed:
                        try:
                            values = json.loads(simple_key)
                        except Exception:
                            continue
                        values = [_norm_val(item) for item in values]
                        all_ids.extend(build_filter_ids_from_vals(values))
                    sub_qs = base_qs.filter(id__in=all_ids)
                elif isinstance(parsed, list):
                    values = [_norm_val(item) for item in parsed]
                    ids = build_filter_ids_from_vals(values)
                    sub_qs = base_qs.filter(id__in=ids)
                    before_snapshot = _capture_history_snapshot()
                else:
                    sub_qs = None

                if sub_qs is not None:
                    selected_ids = list(sub_qs.values_list("id", flat=True))
                    if has_rotacions:
                        blocked_groups = _programmed_groups_emptied_by_move(self.competicio, selected_ids)
                        if blocked_groups:
                            messages.error(request, _message_for_emptied_programmed_groups(blocked_groups))
                            query = request.GET.copy()
                            for key in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                                query.pop(key, None)
                            return redirect(f"{request.path}?{query.urlencode()}")
                    existing_groups = list(sub_qs.exclude(grup__isnull=True).values_list("grup", flat=True).distinct())
                    if existing_groups and not has_rotacions:
                        new_group_num = min(existing_groups)
                        with transaction.atomic():
                            Inscripcio.objects.filter(competicio=self.competicio, grup=new_group_num).update(grup=None)
                            updated = sub_qs.update(grup=new_group_num)
                            sync_stable_groups_from_legacy(self.competicio)
                    else:
                        new_group_num = next_group_display_num(self.competicio)
                        with transaction.atomic():
                            updated = sub_qs.update(grup=new_group_num)
                            sync_stable_groups_from_legacy(self.competicio)

                    _record_history("make_independent_group", "Fer grup independent", before_snapshot)
                    messages.success(request, f"Creat el grup {new_group_num} amb {updated} inscripcions del subgrup.")
                    if request.headers.get("x-requested-with") == "XMLHttpRequest":
                        return JsonResponse(
                            {
                                "ok": True,
                                "new_group_num": new_group_num,
                                "updated": updated,
                                "history": get_inscripcions_history_state(request, self.competicio.id),
                            }
                        )
                    query = request.GET.copy()
                    for key in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                        query.pop(key, None)
                    return redirect(f"{request.path}?{query.urlencode()}")

            level = {"g1": 1, "g2": 2, "g3": 3}.get(level_token, 1)
            values = []
            if level >= 1:
                values.append(_norm_val(value_1))
            if level >= 2:
                values.append(_norm_val(value_2))
            if level >= 3:
                values.append(_norm_val(value_3))

            ids = build_filter_ids_from_vals(values)
            base_qs = self.get_queryset_base_filtrada()
            sub_qs = base_qs.filter(id__in=ids)
            selected_ids = list(sub_qs.values_list("id", flat=True))
            if has_rotacions:
                blocked_groups = _programmed_groups_emptied_by_move(self.competicio, selected_ids)
                if blocked_groups:
                    messages.error(request, _message_for_emptied_programmed_groups(blocked_groups))
                    query = request.GET.copy()
                    for key in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                        query.pop(key, None)
                    return redirect(f"{request.path}?{query.urlencode()}")

            before_snapshot = _capture_history_snapshot()
            existing_groups = list(sub_qs.exclude(grup__isnull=True).values_list("grup", flat=True).distinct())
            if existing_groups and not has_rotacions:
                new_group_num = min(existing_groups)
                with transaction.atomic():
                    Inscripcio.objects.filter(competicio=self.competicio, grup=new_group_num).update(grup=None)
                    updated = sub_qs.update(grup=new_group_num)
                    sync_stable_groups_from_legacy(self.competicio)
            else:
                new_group_num = next_group_display_num(self.competicio)
                with transaction.atomic():
                    updated = sub_qs.update(grup=new_group_num)
                    sync_stable_groups_from_legacy(self.competicio)

            _record_history("make_independent_group", "Fer grup independent", before_snapshot)
            messages.success(request, f"Creat el grup {new_group_num} amb {updated} inscripcions del subgrup.")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "ok": True,
                        "new_group_num": new_group_num,
                        "updated": updated,
                        "history": get_inscripcions_history_state(request, self.competicio.id),
                    }
                )

            query = request.GET.copy()
            for key in ("make_independent_group", "lvl", "v1", "v2", "v3"):
                query.pop(key, None)
            return redirect(f"{request.path}?{query.urlencode()}")

        if request.GET.get("make_groups_count") == "1":
            if competicio_has_rotacions(self.competicio):
                messages.error(request, "No es poden reconfigurar grups mentre hi ha rotacions actives.")
                query = request.GET.copy()
                query.pop("make_groups_count", None)
                query.pop("group_count", None)
                return redirect(f"{request.path}?{query.urlencode()}")
            try:
                group_count = int(request.GET.get("group_count") or 0)
            except ValueError:
                group_count = 0
            if group_count < 1:
                messages.error(request, "El nombre de grups ha de ser com a minim 1.")
                query = request.GET.copy()
                query.pop("make_groups_count", None)
                return redirect(f"{request.path}?{query.urlencode()}")

            qs_base = self.get_queryset_base_filtrada()
            before_snapshot = _capture_history_snapshot()
            group_codes = get_active_group_codes()
            grouping_sig = "|".join(group_codes) if group_codes else ""
            tab_keys = request.GET.getlist("tab_keys")

            if not group_codes:
                selected_tabs = [("__ALL__", list(qs_base.values_list("id", flat=True)))]
            else:
                merges = (self.competicio.tab_merges or {}).get(grouping_sig, [])
                merge_map = {}
                for group_keys in merges:
                    group_tuple = tuple(group_keys)
                    for group_key in group_keys:
                        merge_map[group_key] = group_tuple
                builtin_fields = [code for code in group_codes if hasattr(Inscripcio, code)]
                records = list(qs_base.order_by("ordre_sortida", "id").only("id", "extra", *builtin_fields))
                tab_to_ids = OrderedDict()
                for record in records:
                    simple = simple_key_for_obj(record, group_codes)
                    merged_tuple = merge_map.get(simple)
                    tab_key = json.dumps(list(merged_tuple), ensure_ascii=False) if merged_tuple else simple
                    tab_to_ids.setdefault(tab_key, []).append(record.id)
                if tab_keys:
                    selected_tabs = [(tab_key, tab_to_ids.get(tab_key, [])) for tab_key in tab_keys if tab_key in tab_to_ids]
                else:
                    selected_tabs = list(tab_to_ids.items())

            selected_ids = []
            seen_ids = set()
            for _tab_key, ids in selected_tabs:
                for ins_id in ids:
                    if ins_id in seen_ids:
                        continue
                    seen_ids.add(ins_id)
                    selected_ids.append(ins_id)

            if selected_ids:
                sub_qs = qs_base.filter(id__in=selected_ids).order_by("ordre_sortida", "id")
                objs = list(sub_qs.only("id", "grup"))
                if objs:
                    max_group = (GrupCompeticio.objects.filter(competicio=self.competicio).aggregate(m=Max("display_num"))["m"] or 0)
                    assign_groups_k(objs, group_count, max_group)
                    Inscripcio.objects.bulk_update(objs, ["grup"], batch_size=500)

            sync_stable_groups_from_legacy(self.competicio)
            _record_history("make_groups_count", "Crear grups per nombre", before_snapshot)
            query = request.GET.copy()
            query.pop("make_groups_count", None)
            query.pop("group_count", None)
            query.setlist("tab_keys", [])
            return redirect(f"{request.path}?{query.urlencode()}")

        if request.GET.get("make_groups") == "1":
            if competicio_has_rotacions(self.competicio):
                messages.error(request, "No es poden reconfigurar grups mentre hi ha rotacions actives.")
                query = request.GET.copy()
                query.pop("make_groups", None)
                return redirect(f"{request.path}?{query.urlencode()}")
            try:
                size = int(request.GET.get("group_size") or 0)
            except ValueError:
                size = 0
            if size < 2:
                messages.error(request, "La mida del grup ha de ser com a minim 2.")
                query = request.GET.copy()
                query.pop("make_groups", None)
                return redirect(f"{request.path}?{query.urlencode()}")

            group_mode = request.GET.get("group_mode") or "fixed"
            group_codes = get_active_group_codes()
            qs = self.get_queryset_base_filtrada().order_by("ordre_sortida", "id")
            before_snapshot = _capture_history_snapshot()
            builtin_fields = [code for code in group_codes if hasattr(Inscripcio, code)]
            objs = list(qs.only("id", "grup", "extra", *builtin_fields))
            global_group_num = 0
            if not group_codes:
                if group_mode == "balanced":
                    global_group_num = assign_groups_balanced(objs, size, global_group_num)
                else:
                    for idx, obj in enumerate(objs, start=1):
                        obj.grup = (idx - 1) // size + 1
            else:
                if group_mode == "balanced":
                    current_key = None
                    buffer = []

                    def flush_buffer():
                        nonlocal global_group_num, buffer
                        if buffer:
                            global_group_num = assign_groups_balanced(buffer, size, global_group_num)
                            buffer = []

                    for obj in objs:
                        key = tuple(_norm_val(get_inscripcio_value(obj, code)) for code in group_codes)
                        if key != current_key:
                            flush_buffer()
                            current_key = key
                        buffer.append(obj)
                    flush_buffer()
                else:
                    current_key = None
                    count_in_chunk = 0
                    for obj in objs:
                        key = tuple(_norm_val(get_inscripcio_value(obj, code)) for code in group_codes)
                        if key != current_key:
                            current_key = key
                            count_in_chunk = 0
                        if count_in_chunk == 0:
                            global_group_num += 1
                        obj.grup = global_group_num
                        count_in_chunk += 1
                        if count_in_chunk >= size:
                            count_in_chunk = 0

            with transaction.atomic():
                Inscripcio.objects.bulk_update(objs, ["grup"], batch_size=500)
                sync_stable_groups_from_legacy(self.competicio)
            _record_history("make_groups_size", "Crear grups per mida", before_snapshot)
            query = request.GET.copy()
            query.pop("make_groups", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        if request.GET.get("clear_group") == "1":
            before_snapshot = _capture_history_snapshot()
            self.competicio.group_by_default = []
            self.competicio.tab_merges = {}
            self.competicio.save(update_fields=["group_by_default", "tab_merges"])
            clear_inscripcions_sort_state_for_competicio(request, self.competicio.id)
            _record_history("clear_grouping", "Treure agrupacio", before_snapshot)
            query = request.GET.copy()
            query.pop("clear_group", None)
            query.setlist("group_by", [])
            return redirect(f"{request.path}?{query.urlencode()}")

        if request.GET.get("recalc_order") == "1":
            group_codes = [group for group in request.GET.getlist("group_by") if group in allowed_codes]
            qs = self.get_queryset_base_filtrada()
            before_snapshot = _capture_history_snapshot()
            records = list(qs.order_by("ordre_sortida", "id"))

            def sort_key(obj):
                group_values = tuple(_norm_val(get_inscripcio_value(obj, code)) for code in group_codes)
                previous = obj.ordre_sortida if obj.ordre_sortida is not None else 10**12
                return (group_values, previous, obj.id)

            records.sort(key=sort_key)
            with transaction.atomic():
                for idx, obj in enumerate(records, start=1):
                    if obj.ordre_sortida != idx:
                        Inscripcio.objects.filter(id=obj.id).update(ordre_sortida=idx)
            _record_history("recalc_order", "Aplicar agrupacio", before_snapshot)
            query = request.GET.copy()
            query.pop("recalc_order", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        if request.GET.get("shuffle_order") == "1":
            qs = self.get_queryset_base_filtrada()
            before_snapshot = _capture_history_snapshot()
            shuffle_ordre_sortida(qs)
            _record_history("shuffle_order", "Barreja aleatoriament", before_snapshot)
            query = request.GET.copy()
            query.pop("shuffle_order", None)
            return redirect(f"{request.path}?{query.urlencode()}")

        if "group_by" in request.GET:
            selected = [group for group in request.GET.getlist("group_by") if group in allowed_codes]
            if selected != (self.competicio.group_by_default or []):
                self.competicio.group_by_default = selected
                self.competicio.save(update_fields=["group_by_default"])
        else:
            saved = [group for group in (self.competicio.group_by_default or []) if group in allowed_codes]
            if saved:
                query = request.GET.copy()
                for group in saved:
                    query.appendlist("group_by", group)
                return redirect(f"{request.path}?{query.urlencode()}")

        return super().get(request, *args, **kwargs)

    def get_paginate_by(self, queryset):
        if self.request.GET.getlist("group_by"):
            return None
        per_page = self.request.GET.get("per_page")
        if not per_page or per_page == "all":
            return None
        try:
            return int(per_page)
        except ValueError:
            return None

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["competicio"] = self.competicio
        allowed = get_allowed_group_fields(self.competicio)
        allowed_codes = {field["code"] for field in allowed}
        ctx["allowed_group_fields"] = allowed
        ctx["current_query"] = self.request.GET.urlencode()
        ctx["title_fields_selected"] = self.request.GET.getlist("title_fields")
        sort_fields = get_available_sort_fields(self.competicio)
        sort_codes = {field["code"] for field in sort_fields}
        raw_sort_key = self.request.GET.get("sort_key") or "nom_i_cognoms"
        sort_key_selected = LEGACY_SORT_KEY_MAP.get(raw_sort_key, raw_sort_key)
        if sort_key_selected not in sort_codes:
            sort_key_selected = "nom_i_cognoms" if "nom_i_cognoms" in sort_codes else (next(iter(sort_codes), ""))
        ctx["sort_field_options"] = sort_fields
        ctx["sort_key_selected"] = sort_key_selected

        selected = self.request.GET.getlist("group_by")
        if not selected:
            selected = self.competicio.group_by_default or []
        selected = [group for group in selected if group in allowed_codes]
        ctx["selected_group_fields"] = selected

        records_qs = self.get_queryset_base_filtrada()
        if selected:
            records_qs = annotate_inscripcions_queryset_for_group_codes(records_qs, self.competicio, selected)
        records = list(records_qs.order_by("ordre_sortida", "id"))
        _attach_base_equip_runtime(records)
        grouping_sig = "|".join(selected) if selected else ""
        ctx["grouping_sig"] = grouping_sig

        def pretty_val(value):
            return "(Sense valor)" if value in (None, "") else str(value)

        if selected:
            grouped = OrderedDict()
            label_map = {}
            for record in records:
                values = [_norm_val(get_inscripcio_value(record, code)) for code in selected]
                key = json.dumps(values, ensure_ascii=False)
                grouped.setdefault(key, []).append(record)
                if key not in label_map:
                    parts = [pretty_val(get_inscripcio_value(record, code)) for code in selected]
                    label_map[key] = " · ".join(parts)

            merges = (self.competicio.tab_merges or {}).get(grouping_sig, [])
            merge_map = {}
            for group_keys in merges:
                if not group_keys:
                    continue
                group_tuple = tuple(group_keys)
                for group_key in group_keys:
                    merge_map[group_key] = group_tuple

            grouped_merged = OrderedDict()
            label_map_merged = {}
            for key, rows in grouped.items():
                merged_tuple = merge_map.get(key)
                if merged_tuple:
                    tab_key = json.dumps(list(merged_tuple), ensure_ascii=False)
                    grouped_merged.setdefault(tab_key, []).extend(rows)
                    if tab_key not in label_map_merged:
                        parts = []
                        for simple_key in merged_tuple:
                            label = label_map.get(simple_key, simple_key)
                            if label not in parts:
                                parts.append(label)
                        label_map_merged[tab_key] = " + ".join(parts)
                else:
                    grouped_merged.setdefault(key, []).extend(rows)
                    label_map_merged.setdefault(key, label_map.get(key, key))

            records_grouped = [(label_map_merged.get(key, key), rows, key) for key, rows in grouped_merged.items()]
            ctx["tabs"] = [
                {"key": group_key, "label": group_label, "count": len(group_records)}
                for (group_label, group_records, group_key) in records_grouped
            ]
            ctx["records_grouped"] = records_grouped
        else:
            ctx["records_grouped"] = None

        ctx_records = ctx.get("records")
        if ctx_records is not None:
            _attach_base_equip_runtime(list(ctx_records))

        excel_cols = get_available_excel_columns(self.competicio)
        excel_codes = {column["code"] for column in excel_cols}
        ctx["allowed_excel_columns"] = [(column["code"], column.get("ui_label") or column["label"]) for column in excel_cols]
        selected_excel_cols = []
        for raw in self.request.GET.getlist("excel_cols"):
            code = LEGACY_EXCEL_COL_MAP.get(raw, raw)
            if code in excel_codes and code not in selected_excel_cols:
                selected_excel_cols.append(code)
        if not selected_excel_cols:
            selected_excel_cols = [column["code"] for column in excel_cols]
        ctx["excel_cols_selected"] = selected_excel_cols
        base = self.get_queryset_base_filtrada()
        ctx["categories_distinct"] = list(base.order_by().values_list("categoria", flat=True).distinct())
        ctx["cats_selected"] = self.request.GET.getlist("cats")
        ctx["history_state"] = get_inscripcions_history_state(self.request, self.competicio.id)
        return ctx


__all__ = [
    "InscripcionsImportExcelView",
    "InscripcionsListView",
    "LEGACY_EXCEL_COL_MAP",
    "_attach_base_equip_runtime",
    "_message_for_emptied_programmed_groups",
    "_normalize_competition_order_tail_flag",
    "_persist_group_suggested_names",
    "_programmed_groups_emptied_by_move",
    "_split_custom_sort_tokens",
    "_sync_group_names_for_competicio",
    "arrow_positions",
    "assign_groups_balanced",
    "assign_groups_k",
    "clear_inscripcions_sort_state_for_competicio",
    "competicio_has_rotacions",
    "get_available_excel_columns",
    "get_excel_export_value",
    "recalcular_ordre_sortida",
    "renumber_groups_for_competicio",
    "set_competicio_custom_sort_order_values",
    "shuffle_ordre_sortida",
    "sort_records_by_field_stable",
    "sync_stable_groups_from_legacy",
]
