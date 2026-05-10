import json
import os
import uuid
from datetime import date, datetime

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ...models import Competicio, Inscripcio
from ...models.competicio import CompeticioAparell, InscripcioAparellExclusio
from ...models.rotacions import RotacioAssignacio, RotacioEstacio, RotacioFranja
from ...models.scoring import SerieEquip
from ...services.shared.competition_groups import (
    get_group_maps,
    get_inscripcio_competition_order,
    get_inscripcio_group_display_num,
    group_label,
)
from ...services.rotacions.rotacions_ordering import (
    ORDER_MODE_MAINTAIN,
    build_rotation_unit_step_map,
    effective_rotate_steps,
    get_rotacions_order_modes,
    order_pairs_for_mode,
    rotation_unit_key,
    unique_ordered,
)
from ...services.scoring.team_scoring import build_team_subjects_for_comp_aparell, is_team_context_app
from ...services.teams.team_series import serie_label
from ._shared import (
    _assignacio_program_keys,
    _export_meta_defaults,
    _get_export_meta,
    _logo_abs_path,
    _logo_url_from_path,
    _normalize_export_participant_fields,
    _rotacions_available_participant_fields,
    _save_export_meta,
)


def _is_competitive_franja(franja):
    return getattr(franja, "tipus", RotacioFranja.TIPUS_COMPETITION) == RotacioFranja.TIPUS_COMPETITION


def _excel_hex(value: str) -> str:
    return str(value or "").strip().lstrip("#").upper()


def _franja_excel_style_parts(franja):
    fill = PatternFill("solid", fgColor=_excel_hex(franja.resolved_background_color))
    border_side = Side(style="thin", color=_excel_hex(franja.resolved_border_color))
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    font_color = _excel_hex(franja.resolved_text_color)
    return fill, font_color, border


@require_POST
@csrf_protect
def rotacions_export_meta_save(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invalid")

    title = str(payload.get("title", "") or "").strip()
    venue = str(payload.get("venue", "") or "").strip()
    date_str = str(payload.get("date", "") or "").strip()
    if date_str:
        if parse_date(date_str) is None:
            return HttpResponseBadRequest("Data invalida. Format esperat: YYYY-MM-DD")
    participant_fields = _normalize_export_participant_fields(
        competicio,
        payload.get("participant_fields"),
    )

    current = _get_export_meta(competicio)
    current["title"] = title or _export_meta_defaults(competicio)["title"]
    current["venue"] = venue
    current["date"] = date_str
    current["participant_fields"] = participant_fields
    _save_export_meta(competicio, current)

    return JsonResponse(
        {
            "ok": True,
            "meta": {
                "title": current["title"],
                "venue": current["venue"],
                "date": current["date"],
                "logo_path": current.get("logo_path", ""),
                "logo_url": _logo_url_from_path(current.get("logo_path", "")),
                "participant_fields": current.get("participant_fields", []),
            },
        }
    )


@require_POST
@csrf_protect
def rotacions_export_logo_upload(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    f = request.FILES.get("logo")
    if not f:
        return HttpResponseBadRequest("Falta fitxer 'logo'")

    max_bytes = 4 * 1024 * 1024
    if int(getattr(f, "size", 0) or 0) > max_bytes:
        return HttpResponseBadRequest("El logo supera el maxim de 4MB")

    ctype = str(getattr(f, "content_type", "") or "").lower()
    if not ctype.startswith("image/"):
        return HttpResponseBadRequest("El fitxer ha de ser una imatge")

    ext = os.path.splitext(getattr(f, "name", "") or "")[1].lower()
    if ext not in {".png", ".jpg", ".jpeg", ".bmp"}:
        ext = ".png"

    rel_dir = f"rotacions/logos/competicio_{competicio.id}"
    rel_path = f"{rel_dir}/{uuid.uuid4().hex}{ext}"

    content = f.read()
    saved_rel = default_storage.save(rel_path, ContentFile(content))

    current = _get_export_meta(competicio)
    old_logo = str(current.get("logo_path", "") or "").strip()
    current["logo_path"] = saved_rel
    _save_export_meta(competicio, current)

    if old_logo and old_logo != saved_rel:
        try:
            if default_storage.exists(old_logo):
                default_storage.delete(old_logo)
        except Exception:
            pass

    return JsonResponse(
        {
            "ok": True,
            "logo_path": saved_rel,
            "logo_url": _logo_url_from_path(saved_rel),
        }
    )


@require_POST
@csrf_protect
def rotacions_export_logo_clear(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    current = _get_export_meta(competicio)
    old_logo = str(current.get("logo_path", "") or "").strip()
    current["logo_path"] = ""
    _save_export_meta(competicio, current)

    if old_logo:
        try:
            if default_storage.exists(old_logo):
                default_storage.delete(old_logo)
        except Exception:
            pass

    return JsonResponse({"ok": True})



def franges_export_excel(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    mode = (request.GET.get("mode") or "participants").strip().lower()
    if mode not in {"participants", "groups"}:
        mode = "participants"

    estacions = list(
        RotacioEstacio.objects
        .filter(competicio=competicio, actiu=True)
        .select_related("comp_aparell__aparell")
        .order_by("ordre", "id")
    )
    franges = list(RotacioFranja.objects.filter(competicio=competicio).order_by("ordre_visual", "id"))

    franja_modes = get_rotacions_order_modes(competicio)

    group_maps = get_group_maps(competicio)
    groups_by_id = group_maps["by_id"]
    group_labels_by_id = group_maps["label_by_id"]
    competition_franja_ids = {fr.id for fr in franges if _is_competitive_franja(fr)}
    assigns = list(
        RotacioAssignacio.objects
        .filter(competicio=competicio, franja_id__in=competition_franja_ids)
        .select_related("franja", "estacio")
        .prefetch_related("grup_links__grup", "serie_links__serie")
        .order_by("franja__ordre", "franja_id", "estacio__ordre", "id")
    )
    rotation_unit_step_map = build_rotation_unit_step_map(
        assigns,
        lambda assignacio: rotation_unit_key(_assignacio_program_keys(assignacio)),
        franja_modes,
    )
    cell_groups = {}
    for a in assigns:
        cell_groups[(a.franja_id, a.estacio_id)] = _assignacio_program_keys(a)

    estacio_comp_aparell = {
        e.id: (e.comp_aparell_id if getattr(e, "tipus", None) == "aparell" else None)
        for e in estacions
    }
    estacio_is_team = {
        e.id: bool(
            getattr(e, "tipus", None) == "aparell"
            and getattr(getattr(e, "comp_aparell", None), "aparell", None)
            and getattr(e.comp_aparell.aparell, "competition_unit", "") == "team"
        )
        for e in estacions
    }
    comp_aparell_ids = sorted({x for x in estacio_comp_aparell.values() if x})

    group_ids = sorted({
        int(key.split(":", 1)[1])
        for items in cell_groups.values()
        for key in items
        if str(key).startswith("g:")
    })
    serie_ids = sorted({
        int(key.split(":", 1)[1])
        for items in cell_groups.values()
        for key in items
        if str(key).startswith("s:")
    })
    ins_by_grup = {}
    excluded_pairs = set()
    if group_ids:
        qs = (
            Inscripcio.objects
            .filter(competicio=competicio, grup_competicio_id__in=group_ids)
            .select_related("grup_competicio")
            .only(
                "id",
                "grup",
                "grup_competicio",
                "ordre_competicio",
                "ordre_sortida",
                "nom_i_cognoms",
                "document",
                "sexe",
                "data_naixement",
                "entitat",
                "categoria",
                "subcategoria",
                "extra",
            )
            .order_by("grup_competicio__display_num", "ordre_competicio", "ordre_sortida", "id")
        )
        ins_ids = []
        for ins in qs:
            ins_by_grup.setdefault(ins.grup_competicio_id, []).append(ins)
            ins_ids.append(ins.id)

        if ins_ids and comp_aparell_ids:
            excluded_pairs = set(
                InscripcioAparellExclusio.objects.filter(
                    inscripcio_id__in=ins_ids,
                    comp_aparell_id__in=comp_aparell_ids,
                ).values_list("inscripcio_id", "comp_aparell_id")
            )

    series_by_id = {
        int(serie.id): serie
        for serie in SerieEquip.objects
        .filter(competicio=competicio, id__in=serie_ids)
        .select_related("comp_aparell__aparell")
    }
    team_subjects_by_serie = {}
    if serie_ids:
        app_ids_for_series = sorted({int(serie.comp_aparell_id) for serie in series_by_id.values()})
        for app_id in app_ids_for_series:
            comp_aparell = CompeticioAparell.objects.filter(pk=app_id, competicio=competicio).select_related("aparell").first()
            if comp_aparell is None or not is_team_context_app(comp_aparell):
                continue
            subjects, _issues = build_team_subjects_for_comp_aparell(competicio, comp_aparell)
            for subject in subjects:
                serie_id = int(subject.get("serie_id") or 0)
                if serie_id > 0:
                    team_subjects_by_serie.setdefault(serie_id, []).append(subject)

    def _group_label(g):
        return group_labels_by_id.get(g) or group_label(groups_by_id.get(g))

    export_meta = _get_export_meta(competicio)
    available_participant_fields = _rotacions_available_participant_fields(competicio)
    participant_field_labels = {
        f["code"]: str(f.get("label") or f.get("code") or "").strip()
        for f in available_participant_fields
    }
    participant_fields = _normalize_export_participant_fields(
        competicio,
        export_meta.get("participant_fields"),
    )

    def _inscripcio_field_value(ins, code: str):
        if code == "grup":
            display_num = get_inscripcio_group_display_num(ins)
            return display_num if display_num is not None else "-"
        if code == "ordre_sortida":
            return get_inscripcio_competition_order(ins)
        extra = getattr(ins, "extra", None) or {}
        if isinstance(code, str) and code.startswith("excel__") and isinstance(extra, dict):
            if code in extra:
                return extra.get(code)
            legacy_code = code[len("excel__"):]
            if legacy_code in extra:
                return extra.get(legacy_code)
        if hasattr(ins, code):
            return getattr(ins, code)
        if isinstance(extra, dict):
            return extra.get(code)
        return None

    def _format_field_value(value):
        if value in (None, ""):
            return "-"
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, date):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    def _ordered_inscripcions_for_cell(franja, estacio, items):
        if not items:
            return []

        mode_for_franja = franja_modes.get(str(franja.id), ORDER_MODE_MAINTAIN)
        comp_aparell_id = estacio_comp_aparell.get(estacio.id)

        base_pairs = []
        seen_ins = set()
        for key in items:
            if not str(key).startswith("g:"):
                continue
            try:
                group_id = int(str(key).split(":", 1)[1])
            except Exception:
                continue
            for ins in ins_by_grup.get(group_id, []):
                ins_id = ins.id
                if comp_aparell_id and (ins_id, comp_aparell_id) in excluded_pairs:
                    continue
                if ins_id in seen_ins:
                    continue
                seen_ins.add(ins_id)
                base_pairs.append((ins_id, ins))

        unit_key = rotation_unit_key(items)
        ordered_pairs = order_pairs_for_mode(
            base_pairs,
            mode_for_franja,
            rotate_steps=effective_rotate_steps(
                mode_for_franja,
                rotation_unit_step_map.get((unit_key, franja.id), 0),
            ),
            seed_prefix=f"rot-export|{competicio.id}|{franja.id}|{estacio.id}|{unit_key}",
        )
        return [ins for _ins_id, ins in ordered_pairs]

    def _ordered_team_subjects_for_cell(franja, estacio, items):
        if not items:
            return []
        mode_for_franja = franja_modes.get(str(franja.id), ORDER_MODE_MAINTAIN)
        base_pairs = []
        seen_subject_ids = set()
        for key in items:
            if not str(key).startswith("s:"):
                continue
            try:
                serie_id = int(str(key).split(":", 1)[1])
            except Exception:
                continue
            rows = list(team_subjects_by_serie.get(serie_id, []))
            for subject in rows:
                subject_id = int(subject.get("subject_id") or 0)
                if subject_id <= 0:
                    continue
                if subject_id in seen_subject_ids:
                    continue
                seen_subject_ids.add(subject_id)
                base_pairs.append((subject_id, subject))

        unit_key = rotation_unit_key(items)
        ordered_pairs = order_pairs_for_mode(
            base_pairs,
            mode_for_franja,
            rotate_steps=effective_rotate_steps(
                mode_for_franja,
                rotation_unit_step_map.get((unit_key, franja.id), 0),
            ),
            seed_prefix=f"rot-export|team|{competicio.id}|{franja.id}|{estacio.id}|{unit_key}",
        )
        return [subject for _subject_id, subject in ordered_pairs]

    titol_competicio = str(export_meta.get("title", "") or "").strip() or getattr(
        competicio, "nom", f"Competicio {competicio.id}"
    )
    seu = str(export_meta.get("venue", "") or "").strip() or (getattr(competicio, "seu", "") or "-")

    data_comp = None
    date_meta = str(export_meta.get("date", "") or "").strip()
    if date_meta:
        data_comp = parse_date(date_meta)
    if not data_comp:
        data_comp = getattr(competicio, "data", None)
    data_txt = data_comp.strftime("%d/%m/%Y") if data_comp else ""
    logo_path = str(export_meta.get("logo_path", "") or "").strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rotacions"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_center = Alignment(horizontal="left", vertical="center", wrap_text=False)
    center_no_wrap = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    fill_title = PatternFill("solid", fgColor="1F4E79")
    fill_sub = PatternFill("solid", fgColor="D9E1F2")
    fill_hdr = PatternFill("solid", fgColor="E9EEF7")
    fill_special = PatternFill("solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="9AA7B2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if mode == "participants":
        if not participant_fields:
            participant_fields = ["nom_i_cognoms"]
            participant_field_labels.setdefault("nom_i_cognoms", "Nom i cognoms")
        participant_cols_per_estacio = max(1, len(participant_fields))
        total_cols = 1 + (len(estacions) * participant_cols_per_estacio)
    else:
        total_cols = 1 + len(estacions)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=titol_competicio)
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = fill_title
    c.alignment = center_no_wrap

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c = ws.cell(row=2, column=1, value=f"Seu: {seu}    {data_txt}")
    c.font = Font(bold=True)
    c.fill = fill_sub
    c.alignment = center_no_wrap

    ws.append([])

    ws.column_dimensions[get_column_letter(1)].width = 22

    if mode == "participants":
        header_row_top = ws.max_row + 1
        header_row_sub = header_row_top + 1
        data_start_row = header_row_sub + 1

        ws.merge_cells(
            start_row=header_row_top,
            start_column=1,
            end_row=header_row_sub,
            end_column=1,
        )
        franja_hdr = ws.cell(row=header_row_top, column=1, value="Franja")
        franja_hdr.font = bold
        franja_hdr.fill = fill_hdr
        franja_hdr.alignment = center_no_wrap
        franja_hdr.border = border
        ws.cell(row=header_row_sub, column=1).fill = fill_hdr
        ws.cell(row=header_row_sub, column=1).border = border

        estacio_col_ranges = {}
        col_cursor = 2
        for e in estacions:
            start_col = col_cursor
            end_col = start_col + participant_cols_per_estacio - 1
            estacio_col_ranges[e.id] = (start_col, end_col)

            ws.merge_cells(
                start_row=header_row_top,
                start_column=start_col,
                end_row=header_row_top,
                end_column=end_col,
            )
            e_cell = ws.cell(row=header_row_top, column=start_col, value=e.nom)
            e_cell.font = bold
            e_cell.fill = fill_hdr
            e_cell.alignment = center_no_wrap

            for col in range(start_col, end_col + 1):
                ws.cell(row=header_row_top, column=col).fill = fill_hdr
                ws.cell(row=header_row_top, column=col).border = border

            for idx, code in enumerate(participant_fields):
                sub_col = start_col + idx
                label = participant_field_labels.get(code, code)
                sub = ws.cell(row=header_row_sub, column=sub_col, value=label)
                sub.font = bold
                sub.fill = fill_hdr
                sub.alignment = center_no_wrap
                sub.border = border

            col_cursor = end_col + 1

        field_width_map = {
            "nom_i_cognoms": 24,
            "sexe": 10,
            "entitat": 20,
            "categoria": 14,
            "subcategoria": 14,
            "document": 16,
            "data_naixement": 14,
            "grup": 8,
            "ordre_sortida": 8,
        }
        for e in estacions:
            start_col, _end_col = estacio_col_ranges[e.id]
            for idx, code in enumerate(participant_fields):
                width = field_width_map.get(code, 14)
                ws.column_dimensions[get_column_letter(start_col + idx)].width = width

        current_row = data_start_row
        for i, f in enumerate(franges, start=1):
            label = getattr(f, "display_label", None) or (f.titol or "").strip() or "Franja"
            fr_txt = f"{label}\n{f.hora_inici.strftime('%H:%M')}-{f.hora_fi.strftime('%H:%M')}"
            is_competitive = _is_competitive_franja(f)
            row_fill, row_font_color, row_border = _franja_excel_style_parts(f)

            if not is_competitive:
                start_row = current_row
                end_row = start_row
                ws.merge_cells(
                    start_row=start_row,
                    start_column=1,
                    end_row=end_row,
                    end_column=total_cols,
                )
                fr_cell = ws.cell(row=start_row, column=1, value=fr_txt)
                fr_cell.alignment = center
                fr_cell.border = row_border
                fr_cell.font = Font(bold=True, color=row_font_color)
                for col in range(1, total_cols + 1):
                    cell = ws.cell(row=start_row, column=col)
                    cell.fill = row_fill
                    cell.border = row_border
                    cell.font = Font(color=row_font_color)
                ws.row_dimensions[start_row].height = 24
                current_row = end_row + 1
                continue

            cell_participants = {}
            max_participants = 0
            for e in estacions:
                gs = cell_groups.get((f.id, e.id), [])
                if estacio_is_team.get(e.id):
                    ordered = _ordered_team_subjects_for_cell(f, e, gs)
                else:
                    ordered = _ordered_inscripcions_for_cell(f, e, gs)
                cell_participants[e.id] = ordered
                if len(ordered) > max_participants:
                    max_participants = len(ordered)

            block_rows = max(1, max_participants)
            start_row = current_row
            end_row = start_row + block_rows - 1

            ws.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=end_row,
                end_column=1,
            )
            fr_cell = ws.cell(row=start_row, column=1, value=fr_txt)
            fr_cell.alignment = center
            fr_cell.border = row_border
            fr_cell.font = Font(bold=True, color=row_font_color)
            for rr in range(start_row, end_row + 1):
                c1 = ws.cell(row=rr, column=1)
                c1.alignment = center
                c1.border = row_border
                c1.fill = row_fill
                c1.font = Font(bold=True, color=row_font_color)
                ws.row_dimensions[rr].height = 22

            for rr in range(start_row, end_row + 1):
                for col in range(1, total_cols + 1):
                    cell = ws.cell(row=rr, column=col)
                    cell.fill = row_fill
                    cell.border = row_border
                    if col != 1:
                        cell.font = Font(color=row_font_color)

            for rr_offset in range(block_rows):
                rr = start_row + rr_offset
                for e in estacions:
                    start_col, _end_col = estacio_col_ranges[e.id]
                    ordered = cell_participants.get(e.id, [])
                    current_ins = ordered[rr_offset] if rr_offset < len(ordered) else None
                    for idx, code in enumerate(participant_fields):
                        value = ""
                        if current_ins is not None:
                            if estacio_is_team.get(e.id):
                                if code == "nom_i_cognoms":
                                    value = str(current_ins.get("name") or current_ins.get("label") or "")
                                elif code == "grup":
                                    value = str(current_ins.get("serie_label") or "")
                                elif code == "ordre_sortida":
                                    value = str(current_ins.get("order") or "")
                                else:
                                    value = str(current_ins.get("members_text") or current_ins.get("meta") or "")
                            else:
                                value = _format_field_value(_inscripcio_field_value(current_ins, code))
                        cell = ws.cell(row=rr, column=start_col + idx, value=value)
                        cell.alignment = left_center
                        cell.border = row_border
                        cell.fill = row_fill
                        cell.font = Font(color=row_font_color)

            current_row = end_row + 1

        ws.row_dimensions[header_row_top].height = 24
        ws.row_dimensions[header_row_sub].height = 22
        ws.freeze_panes = f"B{data_start_row}"
    else:
        header_row = ws.max_row + 1
        ws.cell(row=header_row, column=1, value="Franja").font = bold
        ws.cell(row=header_row, column=1).fill = fill_hdr
        ws.cell(row=header_row, column=1).alignment = center_no_wrap
        ws.cell(row=header_row, column=1).border = border

        for j, e in enumerate(estacions, start=2):
            cell = ws.cell(row=header_row, column=j, value=e.nom)
            cell.font = bold
            cell.fill = fill_hdr
            cell.alignment = center_no_wrap
            cell.border = border

        for i, f in enumerate(franges, start=1):
            r = header_row + i
            label = getattr(f, "display_label", None) or (f.titol or "").strip() or "Franja"
            fr_txt = f"{label}\n{f.hora_inici.strftime('%H:%M')}-{f.hora_fi.strftime('%H:%M')}"
            row_fill, row_font_color, row_border = _franja_excel_style_parts(f)
            if not _is_competitive_franja(f):
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
                c0 = ws.cell(row=r, column=1, value=fr_txt)
                c0.alignment = center
                c0.border = row_border
                c0.font = Font(bold=True, color=row_font_color)
                for col in range(1, total_cols + 1):
                    cell = ws.cell(row=r, column=col)
                    cell.fill = row_fill
                    cell.border = row_border
                    cell.font = Font(color=row_font_color)
                ws.row_dimensions[r].height = 24
                continue

            c0 = ws.cell(row=r, column=1, value=fr_txt)
            c0.alignment = center
            c0.border = row_border
            c0.fill = row_fill
            c0.font = Font(bold=True, color=row_font_color)
            for col in range(1, total_cols + 1):
                cell = ws.cell(row=r, column=col)
                cell.fill = row_fill
                cell.border = row_border
                if col != 1:
                    cell.font = Font(color=row_font_color)

            for j, e in enumerate(estacions, start=2):
                gs = cell_groups.get((f.id, e.id), [])
                if not gs:
                    txt = ""
                else:
                    if estacio_is_team.get(e.id):
                        labels = unique_ordered(
                            f"{getattr(series_by_id.get(int(str(key).split(':', 1)[1])), 'comp_aparell', None).aparell.nom if series_by_id.get(int(str(key).split(':', 1)[1])) and getattr(series_by_id.get(int(str(key).split(':', 1)[1])).comp_aparell, 'aparell', None) else ''} · {serie_label(series_by_id.get(int(str(key).split(':', 1)[1])))}".strip(" ·")
                            for key in gs
                            if str(key).startswith("s:") and int(str(key).split(":", 1)[1]) in series_by_id
                        )
                    else:
                        labels = unique_ordered(
                            _group_label(int(str(key).split(":", 1)[1]))
                            for key in gs
                            if str(key).startswith("g:")
                        )
                    txt = "\n".join(labels) if labels else "-"
                cell = ws.cell(row=r, column=j, value=txt)
                cell.alignment = center
                cell.border = row_border
                cell.fill = row_fill
                cell.font = Font(color=row_font_color)

            ws.row_dimensions[r].height = 30

        for j in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(j)].width = 24

        ws.row_dimensions[header_row].height = 22
        ws.freeze_panes = f"B{header_row + 1}"

    logo_added = False
    if logo_path:
        logo_abs = _logo_abs_path(logo_path)
        if logo_abs and os.path.exists(logo_abs):
            try:
                img = XLImage(logo_abs)
                img.height = 52
                img.width = 120
                anchor_col = max(1, total_cols - 1)
                img.anchor = f"{get_column_letter(anchor_col)}1"
                ws.add_image(img)
                logo_added = True
            except Exception:
                logo_added = False

    ws.row_dimensions[1].height = 42 if logo_added else 28
    ws.row_dimensions[2].height = 20
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    suffix = "participants" if mode == "participants" else "grups"
    response["Content-Disposition"] = (
        f'attachment; filename="rotacions_{competicio.id}_{suffix}.xlsx"'
    )
    wb.save(response)
    return response


__all__ = [
    "franges_export_excel",
    "rotacions_export_logo_clear",
    "rotacions_export_logo_upload",
    "rotacions_export_meta_save",
]
