__all__ = [
    "series_assign",
    "series_create",
    "series_delete",
    "series_delete_empty",
    "series_detail",
    "series_preview",
    "series_rename",
    "series_reorder",
    "series_start_list_export",
    "series_unassign",
    "series_work_sheet_export",
    "series_workspace",
]


import json

from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from ...models import Competicio
from ...models.competicio import CompeticioAparell
from ...models.scoring import SerieEquip
from ...services.scoring.team_scoring import build_team_subjects_for_comp_aparell, is_team_context_app
from ...services.teams.team_series import (
    assign_subjects_to_serie,
    ensure_serie,
    get_programmed_series_ids,
    get_series_cards_payload,
    get_series_summary_payload,
    next_serie_display_num,
    normalize_subject_ids,
    reorder_serie_subjects,
    safe_deactivate_empty_serie,
    series_plan_signature,
    serie_label,
    summarize_subject_selection,
    unassign_subjects_from_series,
    workspace_subject_order,
)
from ...services.inscripcions.history import (
    capture_inscripcions_history_snapshot,
    record_inscripcions_history_entry,
    with_inscripcions_history_payload,
)


def _parse_payload(request):
    try:
        return json.loads(request.body.decode("utf-8"))
    except Exception:
        return {}


def _resolve_team_comp_aparell(competicio, payload):
    app_id = payload.get("comp_aparell_id")
    if not str(app_id).isdigit():
        return None
    comp_aparell = (
        CompeticioAparell.objects
        .filter(pk=int(app_id), competicio=competicio)
        .select_related("aparell")
        .first()
    )
    if comp_aparell is None or not is_team_context_app(comp_aparell):
        return None
    return comp_aparell


def _serialize_candidate(subject):
    members = list(subject.get("members") or [])
    member_names = [
        str(member.get("name") or "").strip()
        for member in members
        if isinstance(member, dict) and str(member.get("name") or "").strip()
    ]
    members_count = len(member_names)
    members_preview = " + ".join(member_names[:2])
    if members_count > 2:
        members_preview = f"{members_preview} +{members_count - 2}"
    elif not members_preview:
        members_preview = str(subject.get("members_text") or "").strip()
    compact_meta_parts = []
    context_name = str(subject.get("context_name") or "").strip()
    meta = str(subject.get("meta") or "").strip()
    if context_name:
        compact_meta_parts.append(context_name)
    if members_count > 0:
        compact_meta_parts.append(f"{members_count} membre{'s' if members_count != 1 else ''}")
    if meta:
        compact_meta_parts.append(meta)
    return {
        "id": int(subject["subject_id"]),
        "subject_id": int(subject["subject_id"]),
        "subject_kind": "team_unit",
        "label": str(subject.get("name") or "").strip(),
        "context_name": str(subject.get("context_name") or "").strip(),
        "members_text": str(subject.get("members_text") or "").strip(),
        "meta": str(subject.get("meta") or "").strip(),
        "series_state": str(subject.get("series_state") or "unassigned"),
        "serie_id": int(subject.get("serie_id") or 0) or None,
        "serie_label": str(subject.get("serie_label") or "").strip(),
        "serie_order": int(subject.get("serie_order") or 0) or None,
        "invalid_reasons": list(subject.get("invalid_reasons") or []),
        "members_count": members_count,
        "members_preview": members_preview,
        "compact_meta": " · ".join(compact_meta_parts),
    }


def _serialize_series_subject(subject):
    data = dict(subject or {})
    data.update(_serialize_candidate(subject))
    data["name"] = str(subject.get("name") or data.get("label") or "").strip()
    data["members"] = list(subject.get("members") or [])
    return data


def _serialize_series_card(row):
    subjects = [_serialize_series_subject(subject) for subject in (row.get("subjects") or [])]
    subjects_count = int(row.get("subjects_count") or len(subjects))
    status_parts = []
    if row.get("is_programmed"):
        status_parts.append("Programada")
    else:
        status_parts.append("No programada")
    if row.get("is_empty"):
        status_parts.append("Buida")
    return {
        **dict(row or {}),
        "subjects": subjects,
        "subjects_count": subjects_count,
        "summary_label": " · ".join(
            part
            for part in [
                str(row.get("label") or "").strip(),
                f"{subjects_count} unitat{'s' if subjects_count != 1 else ''}",
                "programada" if row.get("is_programmed") else "no programada",
            ]
            if part
        ),
        "status_label": " · ".join(status_parts),
    }


def _preview_subject_samples(subjects, limit=4):
    names = []
    for subject in list(subjects or []):
        label = str(subject.get("name") or subject.get("label") or "").strip()
        if label:
            names.append(label)
    return names[:limit], max(0, len(names) - min(len(names), limit))


def _selection_summary(summary):
    summary = summary if isinstance(summary, dict) else {}
    subject_map = summary.get("subject_map") or {}
    valid_subjects = [
        subject_map.get(subject_id)
        for subject_id in list(summary.get("valid_ids") or [])
        if subject_map.get(subject_id)
    ]
    samples, remaining = _preview_subject_samples(valid_subjects)
    valid_count = len(list(summary.get("valid_ids") or []))
    requested_count = len(list(summary.get("requested_ids") or []))
    if valid_count > 0:
        label = f"{valid_count} unitat{'s' if valid_count != 1 else ''} valida{'s' if valid_count != 1 else ''}"
    elif requested_count > 0:
        label = f"{requested_count} seleccionada{'s' if requested_count != 1 else ''}"
    else:
        label = "Sense seleccio"
    return {
        "count": valid_count,
        "requested_count": requested_count,
        "valid_count": valid_count,
        "invalid_selection_count": len(list(summary.get("invalid_ids") or [])),
        "invalid_subject_count": len(list(summary.get("invalid_subject_ids") or [])),
        "assigned_count": len(list(summary.get("assigned_ids") or [])),
        "unassigned_count": len(list(summary.get("unassigned_ids") or [])),
        "label": label,
        "subject_names_preview": samples,
        "subject_names_remaining": remaining,
    }


def _preview_card_from_subjects(
    *,
    label,
    display_num=None,
    subjects=None,
    serie_id=None,
    name="",
    is_programmed=False,
    impact_kind="existing",
    incoming_count=0,
    outgoing_count=0,
    will_create=False,
    will_delete=False,
    status_note="",
):
    serialized_subjects = [_serialize_series_subject(subject) for subject in list(subjects or [])]
    samples, remaining = _preview_subject_samples(serialized_subjects)
    subjects_count = len(serialized_subjects)
    status_parts = []
    if will_delete:
        status_parts.append("Es desactivara")
    elif will_create:
        status_parts.append("Es creara")
    elif str(impact_kind or "").strip() == "removed":
        status_parts.append("Quedara buida")
    elif str(impact_kind or "").strip() == "reduced":
        status_parts.append("Es reduira")
    elif str(impact_kind or "").strip() == "incoming":
        status_parts.append("Rebra unitats")
    if status_note:
        status_parts.append(str(status_note).strip())
    return {
        "id": int(serie_id) if str(serie_id or "").isdigit() else None,
        "display_num": int(display_num or 0) or None,
        "label": str(label or "").strip() or "Serie",
        "name": str(name or "").strip(),
        "subjects": serialized_subjects,
        "subjects_count": subjects_count,
        "is_empty": subjects_count <= 0,
        "is_programmed": bool(is_programmed),
        "is_out_of_program": not bool(is_programmed),
        "impact_kind": str(impact_kind or "existing").strip() or "existing",
        "incoming_count": int(incoming_count or 0),
        "outgoing_count": int(outgoing_count or 0),
        "remaining_subjects_count": subjects_count,
        "preview_subject_names": samples,
        "preview_subject_names_remaining": remaining,
        "status_label": " | ".join(part for part in status_parts if part),
        "summary_label": " | ".join(
            part
            for part in [
                str(label or "").strip(),
                f"{subjects_count} unitat{'s' if subjects_count != 1 else ''}",
                "programada" if is_programmed else "no programada",
            ]
            if part
        ),
        "will_create": bool(will_create),
        "will_delete": bool(will_delete),
    }


def _preview_card_from_row(row, *, impact_kind="existing", incoming_count=0, outgoing_count=0, status_note=""):
    card = _serialize_series_card(row)
    samples, remaining = _preview_subject_samples(card.get("subjects") or [])
    card.update(
        {
            "impact_kind": str(impact_kind or "existing").strip() or "existing",
            "incoming_count": int(incoming_count or 0),
            "outgoing_count": int(outgoing_count or 0),
            "remaining_subjects_count": int(card.get("subjects_count") or 0),
            "preview_subject_names": samples,
            "preview_subject_names_remaining": remaining,
            "will_create": False,
            "will_delete": False,
        }
    )
    if status_note:
        note = str(status_note).strip()
        card["status_label"] = " | ".join(
            part for part in [str(card.get("status_label") or "").strip(), note] if part
        )
    return card


def _visible_subject_maps(competicio, comp_aparell):
    subjects = _collect_visible_team_subjects(competicio, comp_aparell)
    return subjects, {int(subject["subject_id"]): dict(subject) for subject in subjects}


def _active_series_card_maps(competicio, comp_aparell, subjects):
    rows = get_series_cards_payload(competicio, comp_aparell, list(subjects or []), include_inactive=False)
    return rows, {int(row["id"]): row for row in rows}


def _blocked_reason_message(reason):
    reason = str(reason or "").strip()
    if reason == "serie_programmed":
        return "La serie continua programada a rotacions."
    if reason == "serie_not_empty":
        return "La serie encara te unitats competitives assignades."
    if reason == "preview_required":
        return "Cal generar una nova previsualitzacio abans d'executar l'accio."
    return "La previsualitzacio ha quedat obsoleta."


def _subject_matches_filters(subject, filters):
    filters = filters if isinstance(filters, dict) else {}
    q = str(filters.get("q") or "").strip().lower()
    context_name = str(filters.get("context_name") or "").strip().lower()
    series_state = str(filters.get("series_state") or "all").strip().lower() or "all"
    serie_id = str(filters.get("serie_id") or "").strip()

    if q:
        hay = " ".join([
            str(subject.get("name") or ""),
            str(subject.get("context_name") or ""),
            str(subject.get("members_text") or ""),
            str(subject.get("meta") or ""),
        ]).lower()
        if q not in hay:
            return False
    if context_name and str(subject.get("context_name") or "").strip().lower() != context_name:
        return False
    if series_state == "assigned" and not subject.get("serie_id"):
        return False
    if series_state == "unassigned" and subject.get("serie_id"):
        return False
    if series_state == "invalid" and str(subject.get("series_state") or "") != "invalid":
        return False
    if serie_id and str(subject.get("serie_id") or "") != serie_id:
        return False
    return True


def _build_workspace_payload(competicio, comp_aparell, payload):
    page = max(1, int(payload.get("page") or 1))
    page_size = max(10, min(200, int(payload.get("page_size") or 40)))
    selected_ids = normalize_subject_ids(payload.get("selected_ids") or [])
    filters = payload.get("filters") or {}

    raw_subjects, issues = build_team_subjects_for_comp_aparell(competicio, comp_aparell)
    subjects = [
        dict(subject)
        for subject in raw_subjects
        if int(comp_aparell.id) in (subject.get("allowed_app_ids") or []) or subject.get("invalid_reasons")
    ]
    subjects.sort(key=workspace_subject_order)

    all_context_names = sorted({
        str(subject.get("context_name") or "").strip()
        for subject in subjects
        if str(subject.get("context_name") or "").strip()
    })
    filtered = [subject for subject in subjects if _subject_matches_filters(subject, filters)]
    total = len(filtered)
    start = (page - 1) * page_size
    end = start + page_size
    page_rows = filtered[start:end]
    summary = get_series_summary_payload(competicio, comp_aparell, subjects)
    series_cards = [
        _serialize_series_card(row)
        for row in get_series_cards_payload(competicio, comp_aparell, subjects, include_inactive=False)
    ]
    programmed_ids = set(int(x) for x in get_programmed_series_ids(competicio, comp_aparell))

    return {
        "summary": summary,
        "selected_ids": selected_ids,
        "filters": {
            "q": str(filters.get("q") or "").strip(),
            "context_name": str(filters.get("context_name") or "").strip(),
            "series_state": str(filters.get("series_state") or "all").strip().lower() or "all",
            "serie_id": str(filters.get("serie_id") or "").strip(),
        },
        "filter_options": {
            "context_names": all_context_names,
            "series": [
                {
                    "id": int(row["id"]),
                    "label": str(row["label"]),
                    "subjects_count": int(row["subjects_count"] or 0),
                }
                for row in series_cards
            ],
            "series_states": [
                {"id": "all", "label": "Totes"},
                {"id": "assigned", "label": "Amb serie"},
                {"id": "unassigned", "label": "Sense serie"},
                {"id": "invalid", "label": "Invalides"},
            ],
        },
        "paging": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "pages": max(1, (total + page_size - 1) // page_size) if total else 1,
        },
        "candidates": {
            "items": [_serialize_candidate(subject) for subject in page_rows],
            "total": total,
            "page": page,
            "page_size": page_size,
            "has_more": end < total,
        },
        "series": series_cards,
        "issues": issues,
        "comp_aparell": {
            "id": int(comp_aparell.id),
            "nom": str(getattr(comp_aparell, "display_nom", "") or getattr(comp_aparell.aparell, "nom", "") or "").strip(),
            "codi": str(getattr(comp_aparell, "display_codi", "") or getattr(comp_aparell.aparell, "codi", "") or "").strip(),
        },
        "programmed_series_ids": sorted(programmed_ids),
    }


def _resolve_serie_or_400(competicio, comp_aparell, payload, *, include_inactive=True):
    serie_id = payload.get("serie_id") or payload.get("id")
    if not str(serie_id).isdigit():
        return None
    qs = SerieEquip.objects.filter(competicio=competicio, comp_aparell=comp_aparell, id=int(serie_id))
    if not include_inactive:
        qs = qs.filter(actiu=True)
    return qs.first()


def _preview_message(action, preview):
    counts = preview.get("counts") or {}
    if action == "create":
        selected = int(counts.get("valid", 0))
        if int(counts.get("requested", 0)) > 0 and selected <= 0:
            return "La seleccio actual no aporta cap unitat valida. Es creara una serie buida?"
        if selected <= 0:
            return "Crear una serie buida?"
        return f"Crear una nova serie amb {selected} unitats competitives?"
    if action == "assign":
        selected = int(counts.get("valid", 0))
        if selected <= 0:
            return "No hi ha cap unitat competitiva valida per assignar."
        moved = int(counts.get("already_assigned_elsewhere", 0))
        if moved > 0:
            return f"Assignar {selected} unitats competitives a la serie i reassignar-ne {moved} des d'altres series?"
        return f"Assignar {selected} unitats competitives a la serie?"
    if action == "unassign":
        selected = int(counts.get("assigned", 0))
        if selected <= 0:
            return "No hi ha cap unitat competitiva assignada dins la seleccio actual."
        return f"Treure {selected} unitats competitives de les seves series?"
    if action == "delete":
        if str(preview.get("reason") or "") == "serie_programmed":
            return "Aquesta serie no es pot desactivar mentre continuï programada a rotacions."
        return "Desactivar aquesta serie buida?"
    return "Continuar amb l'accio?"


def _finalize_preview_payload(preview):
    payload = dict(preview or {})
    payload["message"] = _preview_message(payload.get("action"), payload)
    payload["plan_signature"] = series_plan_signature(payload)
    return payload


def _validate_preview_signature(competicio, comp_aparell, payload, action):
    expected_signature = str(payload.get("plan_signature") or "").strip()
    if not expected_signature:
        return False, "preview_required"
    try:
        preview = _build_preview_payload(
            competicio,
            comp_aparell,
            {
                "action": action,
                "serie_id": payload.get("serie_id"),
                "selected_ids": payload.get("selected_ids") or [],
                "name": payload.get("name") or "",
            },
        )
    except (ValueError, LookupError):
        return False, "preview_stale"
    if not preview.get("can_run"):
        return False, str(preview.get("reason") or "preview_stale")
    if str(preview.get("plan_signature") or "") != expected_signature:
        return False, "preview_stale"
    return True, ""


def _build_preview_payload(competicio, comp_aparell, payload):
    action = str(payload.get("action") or "").strip().lower()
    if action not in {"create", "assign", "unassign", "delete"}:
        raise ValueError("action invalid")

    subjects, visible_subject_map = _visible_subject_maps(competicio, comp_aparell)
    current_summary = get_series_summary_payload(competicio, comp_aparell, subjects)
    current_rows, current_row_map = _active_series_card_maps(competicio, comp_aparell, subjects)
    current_order_by_id = {
        int(row["id"]): idx
        for idx, row in enumerate(current_rows)
        if str(row.get("id") or "").isdigit()
    }
    programmed_ids = {int(x) for x in get_programmed_series_ids(competicio, comp_aparell)}

    def ordered_series_ids(values):
        return sorted(
            {int(value) for value in list(values or []) if int(value or 0) > 0},
            key=lambda serie_id: (current_order_by_id.get(int(serie_id), 10**9), int(serie_id)),
        )

    def current_card_for(serie_id, *, impact_kind="existing", incoming_count=0, outgoing_count=0, status_note=""):
        row = current_row_map.get(int(serie_id or 0))
        if row is None:
            serie = (
                SerieEquip.objects
                .filter(competicio=competicio, comp_aparell=comp_aparell, actiu=True, id=int(serie_id or 0))
                .first()
            )
            if serie is None:
                return None
            return _preview_card_from_subjects(
                label=serie_label(serie),
                display_num=getattr(serie, "display_num", None),
                serie_id=getattr(serie, "id", None),
                name=str(getattr(serie, "nom", "") or "").strip(),
                subjects=[],
                is_programmed=int(serie.id) in programmed_ids,
                impact_kind=impact_kind,
                incoming_count=incoming_count,
                outgoing_count=outgoing_count,
                status_note=status_note,
            )
        return _preview_card_from_row(
            row,
            impact_kind=impact_kind,
            incoming_count=incoming_count,
            outgoing_count=outgoing_count,
            status_note=status_note,
        )

    if action == "delete":
        serie = _resolve_serie_or_400(competicio, comp_aparell, payload, include_inactive=False)
        if serie is None:
            raise LookupError("serie invalid")
        is_empty = not serie.items.exists()
        is_programmed = int(serie.id) in programmed_ids
        reason = "" if is_empty and not is_programmed else ("serie_not_empty" if not is_empty else "serie_programmed")
        current_card = current_card_for(
            serie.id,
            impact_kind="existing" if reason else "removed",
            status_note="Serie actual",
        )
        preview = {
            "action": action,
            "serie_id": int(serie.id),
            "serie_label": serie_label(serie),
            "can_run": bool(is_empty and not is_programmed),
            "counts": {"subjects": int(serie.items.count())},
            "requested_ids": [],
            "effective_subject_ids": [],
            "invalid_selection_ids": [],
            "invalid_subject_ids": [],
            "source_series_ids": [int(serie.id)] if is_programmed else [],
            "touched_series_ids": [int(serie.id)],
            "reason": reason,
            "selection": {
                "count": 0,
                "requested_count": 0,
                "valid_count": 0,
                "invalid_selection_count": 0,
                "invalid_subject_count": 0,
                "assigned_count": 0,
                "unassigned_count": 0,
                "label": "Sense seleccio",
                "subject_names_preview": [],
                "subject_names_remaining": 0,
            },
            "summary": {
                **current_summary,
                "affected_subjects_count": 0,
                "existing_series_total": 1,
                "planned_series_total": 1 if not reason else 0,
            },
            "existing_series": [current_card] if current_card else [],
            "planned_series": [
                _preview_card_from_subjects(
                    label=serie_label(serie),
                    display_num=getattr(serie, "display_num", None),
                    serie_id=getattr(serie, "id", None),
                    name=str(getattr(serie, "nom", "") or "").strip(),
                    subjects=[],
                    is_programmed=False,
                    impact_kind="removed",
                    will_delete=True,
                )
            ] if not reason else [],
            "blocked": bool(reason),
            "blocked_reasons": [_blocked_reason_message(reason)] if reason else [],
        }
        return _finalize_preview_payload(preview)

    summary = summarize_subject_selection(competicio, comp_aparell, payload.get("selected_ids") or [])
    counts = {
        "requested": len(summary.get("requested_ids") or []),
        "valid": len(summary.get("valid_ids") or []),
        "invalid_selection": len(summary.get("invalid_ids") or []),
        "invalid_subjects": len(summary.get("invalid_subject_ids") or []),
        "already_assigned_elsewhere": 0,
        "assigned": len(summary.get("assigned_ids") or []),
        "unassigned": len(summary.get("unassigned_ids") or []),
    }
    subject_map = summary.get("subject_map") or {}
    source_series_ids = sorted({
        int(subject.get("serie_id") or 0)
        for subject in subject_map.values()
        if int(subject.get("serie_id") or 0) > 0
    })
    selection = _selection_summary(summary)
    mutated_subjects = [dict(subject) for subject in subjects]
    mutated_subject_map = {int(subject["subject_id"]): subject for subject in mutated_subjects}
    current_series_by_subject = {
        int(subject_id): int(subject.get("serie_id") or 0)
        for subject_id, subject in visible_subject_map.items()
    }
    effective_ids = list(summary.get("valid_ids") or [])
    reason = ""
    target_serie = None
    planned_new_card = None

    if action == "assign":
        serie = _resolve_serie_or_400(competicio, comp_aparell, payload, include_inactive=False)
        if serie is None:
            raise LookupError("serie invalid")
        target_serie = serie
        counts["already_assigned_elsewhere"] = sum(
            1
            for subject_id in summary.get("valid_ids") or []
            if int((summary.get("subject_map") or {}).get(subject_id, {}).get("serie_id") or 0) not in (0, int(serie.id))
        )
        reason = "" if effective_ids else "no_valid_selection"
        for subject_id in effective_ids:
            subject = mutated_subject_map.get(int(subject_id))
            if subject is None:
                continue
            subject["serie_id"] = int(serie.id)
            subject["serie_label"] = serie_label(serie)
            subject["serie_display_num"] = int(getattr(serie, "display_num", 0) or 0) or None
            subject["series_state"] = "assigned"
    elif action == "create":
        next_display_num = next_serie_display_num(competicio, comp_aparell)
        next_label = str(payload.get("name") or "").strip() or f"Serie {next_display_num}"
        for subject_id in effective_ids:
            subject = mutated_subject_map.get(int(subject_id))
            if subject is None:
                continue
            subject["serie_id"] = None
            subject["serie_label"] = ""
            subject["serie_display_num"] = None
            subject["series_state"] = "unassigned"
        planned_subjects = []
        for subject_id in effective_ids:
            subject = dict(visible_subject_map.get(int(subject_id)) or {})
            if not subject:
                continue
            subject["serie_id"] = None
            subject["serie_label"] = next_label
            subject["serie_display_num"] = next_display_num
            subject["series_state"] = "assigned"
            planned_subjects.append(subject)
        planned_new_card = _preview_card_from_subjects(
            label=next_label,
            display_num=next_display_num,
            subjects=planned_subjects,
            is_programmed=False,
            impact_kind="created",
            incoming_count=len(effective_ids),
            will_create=True,
        )
    else:
        effective_ids = list(summary.get("assigned_ids") or [])
        reason = "" if effective_ids else "no_assigned_selection"
        for subject_id in effective_ids:
            subject = mutated_subject_map.get(int(subject_id))
            if subject is None:
                continue
            subject["serie_id"] = None
            subject["serie_label"] = ""
            subject["serie_display_num"] = None
            subject["series_state"] = "unassigned"

    planned_rows, planned_row_map = _active_series_card_maps(competicio, comp_aparell, mutated_subjects)
    planned_series_by_subject = {
        int(subject_id): int(subject.get("serie_id") or 0)
        for subject_id, subject in mutated_subject_map.items()
    }

    touched_series_ids = set(source_series_ids)
    if target_serie is not None:
        touched_series_ids.add(int(target_serie.id))

    existing_series = []
    planned_series = []
    for serie_id in ordered_series_ids(touched_series_ids):
        incoming_count = sum(
            1
            for subject_id in effective_ids
            if int(planned_series_by_subject.get(int(subject_id), 0) or 0) == int(serie_id)
            and int(current_series_by_subject.get(int(subject_id), 0) or 0) != int(serie_id)
        )
        outgoing_count = sum(
            1
            for subject_id in effective_ids
            if int(current_series_by_subject.get(int(subject_id), 0) or 0) == int(serie_id)
            and int(planned_series_by_subject.get(int(subject_id), 0) or 0) != int(serie_id)
        )
        current_card = current_card_for(
            serie_id,
            impact_kind="existing",
            incoming_count=incoming_count,
            outgoing_count=outgoing_count,
            status_note="Estat actual",
        )
        if current_card:
            existing_series.append(current_card)

        planned_row = planned_row_map.get(int(serie_id))
        if planned_row is not None:
            impact_kind = "incoming" if incoming_count > 0 and outgoing_count <= 0 else (
                "reduced" if outgoing_count > 0 else "existing"
            )
            if outgoing_count > 0 and int(planned_row.get("subjects_count") or 0) <= 0:
                impact_kind = "removed"
            planned_series.append(
                _preview_card_from_row(
                    planned_row,
                    impact_kind=impact_kind,
                    incoming_count=incoming_count,
                    outgoing_count=outgoing_count,
                    status_note="Resultat previst",
                )
            )

    if planned_new_card is not None:
        planned_series.insert(0, planned_new_card)

    preview = {
        "action": action,
        "serie_id": int(target_serie.id) if target_serie is not None else None,
        "serie_label": serie_label(target_serie) if target_serie is not None else str(payload.get("name") or "").strip() or "Nova serie",
        "next_display_num": planned_new_card.get("display_num") if planned_new_card else None,
        "can_run": False if reason else (True if action == "create" else bool(effective_ids)),
        "counts": counts,
        "requested_ids": list(summary.get("requested_ids") or []),
        "effective_subject_ids": list(effective_ids),
        "invalid_selection_ids": list(summary.get("invalid_ids") or []),
        "invalid_subject_ids": list(summary.get("invalid_subject_ids") or []),
        "source_series_ids": source_series_ids,
        "touched_series_ids": ordered_series_ids(touched_series_ids),
        "reason": reason,
        "selection": selection,
        "summary": {
            **current_summary,
            "affected_subjects_count": len(effective_ids),
            "existing_series_total": len(existing_series),
            "planned_series_total": len(planned_series),
        },
        "existing_series": existing_series,
        "planned_series": planned_series,
        "blocked": bool(reason),
        "blocked_reasons": [_blocked_reason_message(reason)] if reason else [],
    }
    return _finalize_preview_payload(preview)


def _collect_visible_team_subjects(competicio, comp_aparell):
    subjects = [
        dict(subject)
        for subject in build_team_subjects_for_comp_aparell(competicio, comp_aparell)[0]
        if int(comp_aparell.id) in (subject.get("allowed_app_ids") or []) or subject.get("invalid_reasons")
    ]
    subjects.sort(key=workspace_subject_order)
    return subjects


def _preview_error_response(reason):
    if reason == "preview_required":
        return HttpResponseBadRequest("preview required")
    if reason == "serie_not_empty":
        return HttpResponseBadRequest("serie not empty")
    if reason == "serie_programmed":
        return HttpResponseBadRequest("serie programmed")
    return HttpResponseBadRequest("preview stale")


def _build_series_export_rows(subjects, *, use_full_label=False):
    rows = []
    for idx, subject in enumerate(subjects, start=1):
        full_label = str(subject.get("label") or "").strip()
        short_label = str(subject.get("name") or "").strip()
        rows.append({
            "ordre": int(subject.get("serie_order") or idx),
            "equip": full_label if use_full_label and full_label else short_label,
            "context": str(subject.get("context_name") or "").strip(),
            "membres": str(subject.get("members_text") or "").strip(),
            "meta": str(subject.get("meta") or "").strip(),
            "estat": str(subject.get("series_state") or "").strip(),
        })
        if not use_full_label and short_label:
            rows[-1]["equip"] = short_label
    return rows


def _build_series_workbook(title, rows_by_block):
    wb = Workbook()
    ws = wb.active
    ws.title = "Series"
    ws.append([title])
    ws.append([])
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    row_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    current_row = ws.max_row + 1
    for block in rows_by_block:
        ws.cell(row=current_row, column=1, value=block["title"]).font = Font(bold=True, size=12)
        current_row += 1
        headers = ["Ordre", "Equip", "Context", "Membres", "Meta", "Estat"]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        current_row += 1
        for row in block["rows"]:
            ws.cell(row=current_row, column=1, value=row["ordre"]).alignment = row_alignment
            ws.cell(row=current_row, column=2, value=row["equip"]).alignment = row_alignment
            ws.cell(row=current_row, column=3, value=row["context"]).alignment = row_alignment
            ws.cell(row=current_row, column=4, value=row["membres"]).alignment = row_alignment
            ws.cell(row=current_row, column=5, value=row["meta"]).alignment = row_alignment
            ws.cell(row=current_row, column=6, value=row["estat"]).alignment = row_alignment
            current_row += 1
        current_row += 1
    widths = {1: 10, 2: 24, 3: 18, 4: 44, 5: 18, 6: 12}
    for col, width in widths.items():
        ws.column_dimensions[chr(64 + col)].width = width
    return wb


@require_POST
@csrf_protect
def series_workspace(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")

    workspace = _build_workspace_payload(competicio, comp_aparell, payload)
    return JsonResponse(
        with_inscripcions_history_payload(
            {"ok": True, "workspace": workspace, **workspace},
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def series_detail(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")
    serie = _resolve_serie_or_400(competicio, comp_aparell, payload, include_inactive=False)
    if serie is None:
        return HttpResponseBadRequest("serie invalid")

    subjects = {int(subject["subject_id"]): subject for subject in _collect_visible_team_subjects(competicio, comp_aparell)}
    detail = next(
        (
            _serialize_series_card(row)
            for row in get_series_cards_payload(competicio, comp_aparell, list(subjects.values()), include_inactive=False)
            if int(row["id"]) == int(serie.id)
        ),
        None,
    )
    if detail is None:
        return HttpResponseBadRequest("serie invalid")

    return JsonResponse(
        with_inscripcions_history_payload(
            {"ok": True, "serie": detail, "subjects": detail.get("subjects") or []},
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def series_preview(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")
    try:
        preview = _build_preview_payload(competicio, comp_aparell, payload)
    except ValueError:
        return HttpResponseBadRequest("action invalid")
    except LookupError:
        return HttpResponseBadRequest("serie invalid")

    return JsonResponse(
        with_inscripcions_history_payload(
            {"ok": True, "preview": preview},
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def series_create(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")

    selected_ids = normalize_subject_ids(payload.get("selected_ids") or [])
    name = str(payload.get("name") or "").strip()
    if selected_ids:
        ok, reason = _validate_preview_signature(competicio, comp_aparell, payload, "create")
        if not ok:
            return _preview_error_response(reason)

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    serie = ensure_serie(competicio, comp_aparell, name=name)
    result = assign_subjects_to_serie(serie, selected_ids)
    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="series_equip_create",
        action_label="Crear serie d'equips",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )
    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "serie_id": int(serie.id),
                "updated": len(result.get("updated_ids") or []),
                "skipped_ids": list(result.get("skipped_ids") or []),
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def series_assign(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")
    serie = _resolve_serie_or_400(competicio, comp_aparell, payload, include_inactive=False)
    if serie is None:
        return HttpResponseBadRequest("serie invalid")
    subject_ids = normalize_subject_ids(payload.get("selected_ids") or [])
    if not subject_ids:
        return HttpResponseBadRequest("No hi ha unitats competitives seleccionades")
    ok, reason = _validate_preview_signature(competicio, comp_aparell, payload, "assign")
    if not ok:
        return _preview_error_response(reason)

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    result = assign_subjects_to_serie(serie, subject_ids)
    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="series_equip_assign",
        action_label="Assignar unitats competitives a serie",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )
    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "serie_id": int(serie.id),
                "updated": len(result.get("updated_ids") or []),
                "skipped_ids": list(result.get("skipped_ids") or []),
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def series_unassign(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")
    subject_ids = normalize_subject_ids(payload.get("selected_ids") or [])
    if not subject_ids:
        return HttpResponseBadRequest("No hi ha unitats competitives seleccionades")
    ok, reason = _validate_preview_signature(competicio, comp_aparell, payload, "unassign")
    if not ok:
        return _preview_error_response(reason)

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    result = unassign_subjects_from_series(competicio, comp_aparell, subject_ids)
    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="series_equip_unassign",
        action_label="Treure unitats competitives de la serie",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )
    return JsonResponse(
        with_inscripcions_history_payload(
            {"ok": True, "updated": len(result.get("updated_ids") or [])},
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def series_delete(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")
    serie = _resolve_serie_or_400(competicio, comp_aparell, payload, include_inactive=False)
    if serie is None:
        return HttpResponseBadRequest("serie invalid")
    ok, reason = _validate_preview_signature(competicio, comp_aparell, payload, "delete")
    if not ok:
        return _preview_error_response(reason)

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    ok, reason = safe_deactivate_empty_serie(serie)
    if not ok:
        if reason == "serie_not_empty":
            return HttpResponseBadRequest("serie not empty")
        if reason == "serie_programmed":
            return HttpResponseBadRequest("serie programmed")
        return HttpResponseBadRequest("serie invalid")
    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="series_equip_delete",
        action_label="Desactivar serie d'equips",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )
    return JsonResponse(
        with_inscripcions_history_payload(
            {"ok": True, "deleted": True, "serie_id": int(serie.id)},
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def series_delete_empty(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")

    deleted_ids = []
    skipped_programmed_ids = []
    skipped_not_empty_ids = []
    series = list(
        SerieEquip.objects
        .filter(competicio=competicio, comp_aparell=comp_aparell, actiu=True)
        .order_by("display_num", "id")
    )

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    for serie in series:
        ok, reason = safe_deactivate_empty_serie(serie)
        if ok:
            deleted_ids.append(int(serie.id))
        elif reason == "serie_programmed":
            skipped_programmed_ids.append(int(serie.id))
        elif reason == "serie_not_empty":
            skipped_not_empty_ids.append(int(serie.id))

    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="series_equip_delete_empty",
        action_label="Desactivar series buides",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )
    return JsonResponse(
        with_inscripcions_history_payload(
            {
                "ok": True,
                "deleted": len(deleted_ids),
                "deleted_ids": deleted_ids,
                "skipped_programmed_ids": skipped_programmed_ids,
                "skipped_not_empty_ids": skipped_not_empty_ids,
                "comp_aparell_id": int(comp_aparell.id),
            },
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def series_rename(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")
    serie = _resolve_serie_or_400(competicio, comp_aparell, payload, include_inactive=False)
    if serie is None:
        return HttpResponseBadRequest("serie invalid")

    name = str(payload.get("name") or "").strip()
    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    serie.nom = name
    serie.save(update_fields=["nom", "updated_at"])
    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="series_equip_rename",
        action_label="Renombrar serie d'equips",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )
    return JsonResponse(
        with_inscripcions_history_payload(
            {"ok": True, "serie_id": int(serie.id), "name": name},
            request,
            competicio.id,
        )
    )


@require_POST
@csrf_protect
def series_reorder(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    payload = _parse_payload(request)
    comp_aparell = _resolve_team_comp_aparell(competicio, payload)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")
    serie = _resolve_serie_or_400(competicio, comp_aparell, payload, include_inactive=False)
    if serie is None:
        return HttpResponseBadRequest("serie invalid")
    subject_ids = normalize_subject_ids(payload.get("subject_ids") or [])
    if not subject_ids:
        return HttpResponseBadRequest("subject_ids invalid")

    before_snapshot = capture_inscripcions_history_snapshot(request, competicio)
    ordered_ids = reorder_serie_subjects(serie, subject_ids)
    record_inscripcions_history_entry(
        request,
        competicio,
        action_type="series_equip_reorder",
        action_label="Reordenar serie d'equips",
        before_snapshot=before_snapshot,
        after_snapshot=capture_inscripcions_history_snapshot(request, competicio),
    )
    return JsonResponse(
        with_inscripcions_history_payload(
            {"ok": True, "serie_id": int(serie.id), "subject_ids": ordered_ids},
            request,
            competicio.id,
        )
    )


@require_GET
def series_start_list_export(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    comp_aparell = _resolve_team_comp_aparell(competicio, request.GET)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")

    subjects = _collect_visible_team_subjects(competicio, comp_aparell)
    series = get_series_cards_payload(competicio, comp_aparell, subjects, include_inactive=False)
    rows_by_block = [
        {
            "title": str(serie.get("label") or f"Serie {serie.get('display_num') or ''}"),
            "rows": _build_series_export_rows(serie.get("subjects") or []),
        }
        for serie in series
    ]
    unassigned_rows = _build_series_export_rows([subject for subject in subjects if not subject.get("serie_id")])
    if unassigned_rows:
        rows_by_block.append({"title": "Sense serie", "rows": unassigned_rows})
    wb = _build_series_workbook(
        f"Start list - {competicio.nom} - {getattr(comp_aparell, 'display_nom', '') or getattr(comp_aparell.aparell, 'nom', '')}",
        rows_by_block,
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="series_start_list_{competicio.id}_{comp_aparell.id}.xlsx"'
    )
    wb.save(response)
    return response


@require_GET
def series_work_sheet_export(request, pk):
    competicio = get_object_or_404(Competicio, pk=pk)
    comp_aparell = _resolve_team_comp_aparell(competicio, request.GET)
    if comp_aparell is None:
        return HttpResponseBadRequest("comp_aparell_id invalid")
    serie = _resolve_serie_or_400(competicio, comp_aparell, request.GET, include_inactive=False)
    if serie is None:
        return HttpResponseBadRequest("serie invalid")

    subjects = _collect_visible_team_subjects(competicio, comp_aparell)
    detail = next(
        (
            row
            for row in get_series_cards_payload(competicio, comp_aparell, subjects, include_inactive=False)
            if int(row["id"]) == int(serie.id)
        ),
        None,
    )
    if detail is None:
        return HttpResponseBadRequest("serie invalid")
    wb = _build_series_workbook(
        f"Full de treball - {competicio.nom} - {serie_label(serie)}",
        [{"title": serie_label(serie), "rows": _build_series_export_rows(detail.get("subjects") or [], use_full_label=True)}],
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="serie_work_sheet_{competicio.id}_{serie.id}.xlsx"'
    )
    wb.save(response)
    return response
