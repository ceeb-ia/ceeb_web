from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from functools import lru_cache

from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import Referee
from .colors import color_per_tutor


@dataclass(frozen=True)
class ColumnSpec:
    header: str
    key: str
    width: float
    align: str = "left"
    wrap: bool = False


HEADER_FILL = PatternFill("solid", fgColor="FFE8ECF1")
HEADER_FONT = Font(bold=True, color="FF1F2937")
BODY_FILL_A = PatternFill("solid", fgColor="FFFFFFFF")
BODY_FILL_B = PatternFill("solid", fgColor="FFF7F7F7")
BODY_BORDER = Border(
    left=Side(style="thin", color="FFD1D5DB"),
    right=Side(style="thin", color="FFD1D5DB"),
    top=Side(style="thin", color="FFE5E7EB"),
    bottom=Side(style="thin", color="FFE5E7EB"),
)
GROUP_BORDER = Border(
    left=Side(style="thin", color="FFD1D5DB"),
    right=Side(style="thin", color="FFD1D5DB"),
    top=Side(style="medium", color="FF9CA3AF"),
    bottom=Side(style="thin", color="FFE5E7EB"),
)


@lru_cache(maxsize=None)
def _alignment(horizontal: str, wrap: bool) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


@lru_cache(maxsize=None)
def _soft_tutor_fill(color_hex: str | None) -> PatternFill:
    rgb = (color_hex or "").replace("#", "").strip()
    if len(rgb) != 6:
        return PatternFill("solid", fgColor="FFF3F4F6")
    red = int(rgb[0:2], 16)
    green = int(rgb[2:4], 16)
    blue = int(rgb[4:6], 16)
    mix = 0.88
    red = int(red + (255 - red) * mix)
    green = int(green + (255 - green) * mix)
    blue = int(blue + (255 - blue) * mix)
    return PatternFill("solid", fgColor=f"FF{red:02X}{green:02X}{blue:02X}")


def _parse_time_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    raw = str(value).strip()
    if not raw or raw == "-":
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def _parse_date_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw or raw.lower() == "nat":
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _availability_time(raw: dict | None, key: str):
    if not isinstance(raw, dict):
        return "-"
    parsed = _parse_time_value(raw.get(key))
    return parsed or "-"


def _availability_lookup_by_ref_and_date(run):
    lookup = {}
    for availability in run.availabilities.select_related("referee").all():
        raw = availability.raw or {}
        availability_date = _parse_date_value(raw.get("Data"))
        if availability_date is None:
            continue
        key = (availability.referee_id, availability_date)
        current = lookup.get(key)
        current_score = sum(bool((current or {}).get(field)) for field in ("Hora Inici", "Hora Fi"))
        new_score = sum(bool(raw.get(field)) for field in ("Hora Inici", "Hora Fi"))
        if current is None or new_score >= current_score:
            lookup[key] = raw
    return lookup


def _ordered_referees(run):
    return (
        Referee.objects.filter(active=True, availabilities__run=run)
        .annotate(n=Count("assignments", filter=Q(assignments__run=run), distinct=True))
        .distinct()
        .order_by("name", "code")
    )


def _write_sheet(ws, columns: list[ColumnSpec], rows: list[dict]):
    for col_idx, spec in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=spec.header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BODY_BORDER
        cell.alignment = _alignment(spec.align, spec.wrap)
        ws.column_dimensions[get_column_letter(col_idx)].width = spec.width

    for row_idx, row in enumerate(rows, start=2):
        row_fill = row.get("_fill") or BODY_FILL_A
        border = GROUP_BORDER if row.get("_group_start") else BODY_BORDER
        for col_idx, spec in enumerate(columns, start=1):
            value = row.get(spec.key, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = _alignment(spec.align, spec.wrap)
            if isinstance(value, datetime):
                cell.number_format = "dd/mm/yyyy hh:mm"
            elif isinstance(value, date):
                cell.number_format = "dd/mm/yyyy"
            elif isinstance(value, time):
                cell.number_format = "hh:mm"
            cell.fill = row_fill

        accent_fill = row.get("_accent_fill")
        if accent_fill:
            ws.cell(row=row_idx, column=1).fill = accent_fill
            ws.cell(row=row_idx, column=1).font = Font(bold=True, color="FF111827")

    ws.freeze_panes = "A2"
    end_col = get_column_letter(len(columns))
    end_row = max(len(rows) + 1, 1)
    ws.auto_filter.ref = f"A1:{end_col}{end_row}"
    ws.sheet_view.showGridLines = False


def export_run_to_excel(run, output_path: str):
    assigned_qs = (
        run.assignments.select_related("match", "referee")
        .filter(referee__isnull=False)
        .order_by("referee__code", "match__date", "match__hour_raw", "match__code")
    )
    unassigned_matches = (
        run.assignments.select_related("match")
        .filter(referee__isnull=True)
        .order_by("match__date", "match__hour_raw", "match__code")
    )
    referees_with_counts = list(_ordered_referees(run))
    unassigned_referees = [ref for ref in referees_with_counts if (ref.n or 0) == 0]
    needs_review_referees = [
        ref
        for ref in referees_with_counts
        if not (ref.level or "").strip()
    ]
    availability_by_ref_and_date = _availability_lookup_by_ref_and_date(run)

    assigned_columns = [
        ColumnSpec("Tutor Codi", "tutor_code", 14, "center"),
        ColumnSpec("Tutor", "tutor_name", 24),
        ColumnSpec("Nivell Tutor", "tutor_level", 14, "center"),
        ColumnSpec("Hora Inici Tutor", "tutor_start", 16, "center"),
        ColumnSpec("Hora Fi Tutor", "tutor_end", 14, "center"),
        ColumnSpec("Data Partit", "match_date", 14, "center"),
        ColumnSpec("Hora Partit", "match_time", 12, "center"),
        ColumnSpec("Codi Partit", "match_code", 14, "center"),
        ColumnSpec("Equip local", "home_team", 26),
        ColumnSpec("Equip visitant", "away_team", 26),
        ColumnSpec("Pista", "venue", 28),
        ColumnSpec("Categoria", "category", 14, "center"),
        ColumnSpec("Nota", "note", 26, wrap=True),
        ColumnSpec("Bloquejat", "locked", 12, "center"),
    ]
    assigned_rows = []
    current_referee_id = None
    group_index = 0
    for assignment in assigned_qs:
        referee = assignment.referee
        match = assignment.match
        if referee.id != current_referee_id:
            current_referee_id = referee.id
            group_index += 1
            group_start = True
        else:
            group_start = False
        fill = BODY_FILL_A if group_index % 2 else BODY_FILL_B
        assigned_rows.append(
            {
                "tutor_code": referee.code or "",
                "tutor_name": referee.name or "",
                "tutor_level": referee.level or "",
                "tutor_start": _availability_time(
                    availability_by_ref_and_date.get((referee.id, match.date)),
                    "Hora Inici",
                ),
                "tutor_end": _availability_time(
                    availability_by_ref_and_date.get((referee.id, match.date)),
                    "Hora Fi",
                ),
                "match_date": match.date,
                "match_time": _parse_time_value(match.hour_raw) or (match.hour_raw or ""),
                "match_code": match.code or "",
                "home_team": match.equip_local or "",
                "away_team": match.equip_visitant or "",
                "venue": match.venue or "",
                "category": match.category or "",
                "note": assignment.note or "",
                "locked": "Si" if assignment.locked else "No",
                "_fill": fill,
                "_group_start": group_start,
                "_accent_fill": _soft_tutor_fill(color_per_tutor(referee.code)),
            }
        )

    unassigned_match_columns = [
        ColumnSpec("Codi Partit", "match_code", 14, "center"),
        ColumnSpec("Data Partit", "match_date", 14, "center"),
        ColumnSpec("Hora Partit", "match_time", 12, "center"),
        ColumnSpec("Equip local", "home_team", 26),
        ColumnSpec("Equip visitant", "away_team", 26),
        ColumnSpec("Pista", "venue", 28),
        ColumnSpec("Categoria", "category", 14, "center"),
        ColumnSpec("Municipi", "municipality", 18),
        ColumnSpec("Nota", "note", 24, wrap=True),
    ]
    unassigned_match_rows = [
        {
            "match_code": assignment.match.code or "",
            "match_date": assignment.match.date,
            "match_time": _parse_time_value(assignment.match.hour_raw) or (assignment.match.hour_raw or ""),
            "home_team": assignment.match.equip_local or "",
            "away_team": assignment.match.equip_visitant or "",
            "venue": assignment.match.venue or "",
            "category": assignment.match.category or "",
            "municipality": assignment.match.municipality or "",
            "note": assignment.note or "",
        }
        for assignment in unassigned_matches
    ]

    unassigned_ref_columns = [
        ColumnSpec("Tutor Codi", "tutor_code", 14, "center"),
        ColumnSpec("Tutor", "tutor_name", 24),
        ColumnSpec("Nivell", "level", 14, "center"),
        ColumnSpec("Modalitat", "modality", 18),
        ColumnSpec("Transport", "transport", 18),
    ]
    unassigned_ref_rows = [
        {
            "tutor_code": referee.code or "",
            "tutor_name": referee.name or "",
            "level": referee.level or "",
            "modality": referee.modality or "",
            "transport": referee.transport or "",
        }
        for referee in unassigned_referees
    ]

    needs_review_columns = [
        ColumnSpec("Tutor Codi", "tutor_code", 14, "center"),
        ColumnSpec("Tutor", "tutor_name", 24),
        ColumnSpec("Modalitat", "modality", 18),
        ColumnSpec("Assignacions", "assignments_count", 14, "center"),
    ]
    needs_review_rows = [
        {
            "tutor_code": referee.code or "",
            "tutor_name": referee.name or "",
            "modality": referee.modality or "",
            "assignments_count": referee.n or 0,
        }
        for referee in needs_review_referees
    ]

    wb = Workbook()
    ws_assigned = wb.active
    ws_assigned.title = "Assignacions"
    _write_sheet(ws_assigned, assigned_columns, assigned_rows)

    ws_unassigned_matches = wb.create_sheet("Partits sense assignar")
    _write_sheet(ws_unassigned_matches, unassigned_match_columns, unassigned_match_rows)

    ws_unassigned_refs = wb.create_sheet("Tutors sense assignar")
    _write_sheet(ws_unassigned_refs, unassigned_ref_columns, unassigned_ref_rows)

    ws_needs_review = wb.create_sheet("Tutors sense nivell")
    _write_sheet(ws_needs_review, needs_review_columns, needs_review_rows)

    wb.save(output_path)
    return output_path
