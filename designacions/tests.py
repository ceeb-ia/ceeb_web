import json
import os
import tempfile
import asyncio
import xml.etree.ElementTree as ET
from datetime import date, time
from types import SimpleNamespace
from unittest.mock import patch

import httpx
import pandas as pd
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from logs import _read_job, _write_job, push_log
from .models import Address, AddressCluster, Assignment, Availability, DesignationRun, Match, Referee
from .main_fixed import (
    _availability_penalty_for_subgroup,
    _build_daily_subgroups,
    _build_tutor_working_id,
    _run_rescue_assignment,
    _safe_position_int,
    _segment_failed_subgroup,
)
from .consulta_resultats import fetch_ceeb_classification_async, xml_to_dataframe
from .services.excel_export import export_run_to_excel
from .services.assignment_explainer import (
    explain_candidate_for_assignment,
    explain_current_assignment,
    find_better_alternatives,
)
from .services.assignment_feasibility import build_match_descriptor, has_vehicle, inspect_mobility_transitions
from .services.manual_assignment import (
    build_manual_assignment_context,
    build_run_mobility_summary,
    build_run_scoped_referee_summaries,
    build_top_proposals_for_assignments,
    diagnose_assignment_for_referee,
)
from .services.run_scope import load_scoped_run_data
from .tasks import rebuild_run_map_task


class DesignacionsDateAwareHelpersTests(SimpleTestCase):
    def test_tutor_working_id_includes_date(self):
        row_day_1 = {
            "Codi Tutor de Joc": "5002 F5",
            "Modalitat": "Futbol Sala",
            "Nivell": "NIVELLA1",
            "Data": "2026-02-24",
        }
        row_day_2 = dict(row_day_1, Data="2026-02-25")

        self.assertNotEqual(_build_tutor_working_id(row_day_1), _build_tutor_working_id(row_day_2))

    def test_daily_subgroups_do_not_mix_same_hour_from_different_days(self):
        df = pd.DataFrame(
            [
                {
                    "ID": "P1",
                    "Data": "2026-02-24",
                    "Hora": time(18, 0),
                    "Pista joc": "Pista 1",
                    "cluster": 7,
                },
                {
                    "ID": "P2",
                    "Data": "2026-02-25",
                    "Hora": time(18, 0),
                    "Pista joc": "Pista 1",
                    "cluster": 7,
                },
            ]
        )

        subgrups = _build_daily_subgroups(df, gap_same_pitch_min=60, gap_diff_pitch_min=75, max_partits_subgrup=3)

        self.assertEqual(len(subgrups), 2)
        self.assertEqual({subgrup[0]["ID"] for subgrup in subgrups}, {"P1", "P2"})

    def test_daily_subgroups_keep_same_pitch_matches_together_when_other_pitch_interleaves(self):
        df = pd.DataFrame(
            [
                {
                    "ID": "A1",
                    "Data": "2026-02-24",
                    "Hora": time(10, 0),
                    "Pista joc": "Pista A",
                    "Modalitat": "Futbol Sala",
                    "cluster": 7,
                },
                {
                    "ID": "B1",
                    "Data": "2026-02-24",
                    "Hora": time(10, 30),
                    "Pista joc": "Pista B",
                    "Modalitat": "Futbol Sala",
                    "cluster": 8,
                },
                {
                    "ID": "A2",
                    "Data": "2026-02-24",
                    "Hora": time(11, 0),
                    "Pista joc": "Pista A",
                    "Modalitat": "Futbol Sala",
                    "cluster": 7,
                },
            ]
        )

        subgrups = _build_daily_subgroups(df, gap_same_pitch_min=60, gap_diff_pitch_min=75, max_partits_subgrup=3)

        self.assertEqual([[item["ID"] for item in subgrup] for subgrup in subgrups], [["A1", "A2"], ["B1"]])

    def test_availability_penalty_blocks_other_day_reuse(self):
        tutor_row = {
            "Data": "2026-02-24",
            "Hora Inici": time(17, 0),
            "Hora Fi": time(21, 0),
        }
        subgrup = [
            {
                "Data": "2026-02-25",
                "__match_datetime": pd.Timestamp("2026-02-25 18:00:00"),
            }
        ]

        penalty = _availability_penalty_for_subgroup(tutor_row, subgrup, availability_end_buffer_min=60)

        self.assertGreaterEqual(penalty, 1e6)

    def test_availability_penalty_respects_same_day_window(self):
        tutor_row = {
            "Data": "2026-02-24",
            "Hora Inici": time(17, 0),
            "Hora Fi": time(21, 0),
        }
        subgrup = [
            {
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 18:00:00"),
            },
            {
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 19:00:00"),
            },
        ]

        penalty = _availability_penalty_for_subgroup(tutor_row, subgrup, availability_end_buffer_min=60)

        self.assertEqual(penalty, 0.0)

    def test_has_vehicle_normalizes_transport_values(self):
        self.assertTrue(has_vehicle("Patinet elèctric"))
        self.assertTrue(has_vehicle("Bicicleta"))
        self.assertFalse(has_vehicle("Bus"))

    def test_mobility_blocks_cross_cluster_without_vehicle(self):
        descriptors = [
            build_match_descriptor(
                identifier="M1",
                date_value="2026-02-24",
                time_value="18:00:00",
                venue="Pista A",
                modality="Futbol Sala",
                cluster_id=7,
            ),
            build_match_descriptor(
                identifier="M2",
                date_value="2026-02-24",
                time_value="20:00:00",
                venue="Pista B",
                modality="Futbol Sala",
                cluster_id=8,
            ),
        ]

        issues = inspect_mobility_transitions(
            descriptors,
            transport="Bus",
            gap_same_pitch_min=60,
            gap_diff_pitch_min=75,
            gap_diff_cluster_min=100,
        )

        self.assertEqual([issue.reason_code for issue in issues], ["cross_cluster_without_vehicle"])

    def test_mobility_requires_cluster_gap_for_vehicle(self):
        descriptors = [
            build_match_descriptor(
                identifier="M1",
                date_value="2026-02-24",
                time_value="18:00:00",
                venue="Pista A",
                modality="Futbol Sala",
                cluster_id=7,
            ),
            build_match_descriptor(
                identifier="M2",
                date_value="2026-02-24",
                time_value="19:15:00",
                venue="Pista B",
                modality="Futbol Sala",
                cluster_id=8,
            ),
        ]

        issues = inspect_mobility_transitions(
            descriptors,
            transport="Cotxe",
            gap_same_pitch_min=60,
            gap_diff_pitch_min=75,
            gap_diff_cluster_min=100,
        )

        self.assertEqual([issue.reason_code for issue in issues], ["cross_cluster_gap_violation"])
        self.assertEqual(issues[0].required_gap_min, 100)

    def test_mobility_detects_same_cluster_gap_violation(self):
        descriptors = [
            build_match_descriptor(
                identifier="M1",
                date_value="2026-02-24",
                time_value="18:00:00",
                venue="Pista A",
                modality="Futbol Sala",
                cluster_id=7,
            ),
            build_match_descriptor(
                identifier="M2",
                date_value="2026-02-24",
                time_value="19:00:00",
                venue="Pista B",
                modality="Futbol Sala",
                cluster_id=7,
            ),
        ]

        issues = inspect_mobility_transitions(
            descriptors,
            transport="Cotxe",
            gap_same_pitch_min=60,
            gap_diff_pitch_min=75,
            gap_diff_cluster_min=100,
        )

        self.assertEqual([issue.reason_code for issue in issues], ["same_cluster_gap_violation"])
        self.assertFalse(issues[0].same_pitch)

    def test_xml_to_dataframe_returns_empty_dataframe_when_group_is_missing(self):
        parsed = {
            "grups": [
                {
                    "info": {"nomGrup": "GRUP 01"},
                    "equips_all": [{"NomEquipMostrar": "Equip A"}],
                }
            ]
        }

        df = xml_to_dataframe(parsed, grup="GRUP INEXISTENT")

        self.assertTrue(df.empty)

    def test_safe_position_int_returns_default_for_nan(self):
        self.assertEqual(_safe_position_int(float("nan")), -1)
        self.assertEqual(_safe_position_int(pd.NA), -1)
        self.assertEqual(_safe_position_int(None), -1)

    def test_load_scoped_run_data_filters_by_modalitat_and_dates(self):
        df_disp = pd.DataFrame(
            [
                {
                    "Codi Tutor de Joc": "5001 F5",
                    "Nom": "Tutor F5",
                    "Categoria": "TUTOR/TUTORA DE JOC",
                    "Modalitat": "Futbol Sala",
                    "Nivell": "NIVELLA1",
                    "Data": "2026-02-24",
                },
                {
                    "Codi Tutor de Joc": "5002 BQ",
                    "Nom": "Tutor BQ",
                    "Categoria": "TUTOR/TUTORA DE JOC",
                    "Modalitat": "Basquet",
                    "Nivell": "NIVELLA1",
                    "Data": "2026-02-24",
                },
                {
                    "Codi Tutor de Joc": "5004 F5",
                    "Nom": "Tutor Sense Nivell",
                    "Categoria": "TUTOR/TUTORA DE JOC",
                    "Modalitat": "Futbol Sala",
                    "Nivell": None,
                    "Data": "2026-02-24",
                },
                {
                    "Codi Tutor de Joc": "5003 F5",
                    "Nom": "Tutor Fora Rang",
                    "Categoria": "TUTOR/TUTORA DE JOC",
                    "Modalitat": "Futbol Sala",
                    "Nivell": "NIVELLA1",
                    "Data": "2026-02-25",
                },
            ]
        )
        df_partits = pd.DataFrame(
            [
                {
                    "Codi": "M-F5",
                    "Modalitat": "Futbol Sala",
                    "Data": "2026-02-24",
                    "Grup": "GRUP 01",
                },
                {
                    "Codi": "M-BQ",
                    "Modalitat": "Basquet",
                    "Data": "2026-02-24",
                    "Grup": "GRUP 01",
                },
                {
                    "Codi": "M-F5-LATE",
                    "Modalitat": "Futbol Sala",
                    "Data": "2026-02-25",
                    "Grup": "GRUP 01",
                },
            ]
        )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as disp_tmp:
            disp_path = disp_tmp.name
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as partits_tmp:
            partits_path = partits_tmp.name
        self.addCleanup(lambda: os.path.exists(disp_path) and os.unlink(disp_path))
        self.addCleanup(lambda: os.path.exists(partits_path) and os.unlink(partits_path))

        df_disp.to_excel(disp_path, index=False)
        df_partits.to_excel(partits_path, index=False)

        scoped_disp, scoped_partits = load_scoped_run_data(
            disp_path,
            partits_path,
            params={
                "modalitats": ["Futbol Sala"],
                "date_from": "2026-02-24",
                "date_to": "2026-02-24",
            },
        )

        self.assertEqual(scoped_partits["Codi"].tolist(), ["M-F5"])
        self.assertEqual(scoped_disp["Codi Tutor de Joc"].tolist(), ["5001 F5", "5004 F5"])

    def test_segment_failed_subgroup_prefers_contiguous_two_plus_one_split(self):
        subgrup = [
            {
                "ID": "M1",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 18:00:00"),
            },
            {
                "ID": "M2",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 19:00:00"),
            },
            {
                "ID": "M3",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 20:00:00"),
            },
        ]
        candidate_referees = pd.DataFrame([{"ID": "R1"}, {"ID": "R2"}])

        def subgroup_cost_fn(referee_row, segment):
            key = (referee_row["ID"], tuple(item["ID"] for item in segment))
            viable = {
                ("R1", ("M1", "M2")): 10,
                ("R2", ("M3",)): 10,
            }
            return viable.get(key, 1e6)

        segments = _segment_failed_subgroup(subgrup, candidate_referees, subgroup_cost_fn)

        self.assertEqual([[item["ID"] for item in segment] for segment in segments], [["M1", "M2"], ["M3"]])

    def test_run_rescue_assignment_splits_pair_into_individuals_and_recovers_one(self):
        failed_subgroups = [[
            {
                "ID": "M1",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 18:00:00"),
            },
            {
                "ID": "M2",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 19:00:00"),
            },
        ]]
        candidate_referees = pd.DataFrame([{"ID": "R1"}])

        def subgroup_cost_fn(referee_row, segment):
            key = (referee_row["ID"], tuple(item["ID"] for item in segment))
            viable = {
                ("R1", ("M2",)): 10,
            }
            return viable.get(key, 1e6)

        rescue = _run_rescue_assignment(candidate_referees, failed_subgroups, subgroup_cost_fn)

        self.assertEqual(len(rescue["rounds"]), 1)
        self.assertEqual(len(rescue["assigned_segments"]), 1)
        self.assertEqual([item["ID"] for item in rescue["assigned_segments"][0]["segment"]], ["M2"])
        self.assertEqual([[item["ID"] for item in segment] for segment in rescue["remaining_subgroups"]], [["M1"]])

    def test_run_rescue_assignment_without_idle_referees_returns_no_pairs(self):
        failed_subgroups = [[
            {
                "ID": "M1",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 18:00:00"),
            }
        ]]
        candidate_referees = pd.DataFrame(columns=["ID"])

        rescue = _run_rescue_assignment(candidate_referees, failed_subgroups, lambda *_args, **_kwargs: 1e6)

        self.assertEqual(len(rescue["assigned_segments"]), 0)
        self.assertEqual([[item["ID"] for item in segment] for segment in rescue["remaining_subgroups"]], [["M1"]])

    def test_run_rescue_assignment_keeps_one_segment_per_tutor(self):
        failed_subgroups = [[
            {
                "ID": "M1",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 18:00:00"),
            },
            {
                "ID": "M2",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 19:00:00"),
            },
        ]]
        candidate_referees = pd.DataFrame([{"ID": "R1"}])

        def subgroup_cost_fn(referee_row, segment):
            key = (referee_row["ID"], tuple(item["ID"] for item in segment))
            viable = {
                ("R1", ("M1",)): 10,
                ("R1", ("M2",)): 15,
            }
            return viable.get(key, 1e6)

        rescue = _run_rescue_assignment(candidate_referees, failed_subgroups, subgroup_cost_fn)

        self.assertEqual(len(rescue["rounds"][0]["pairs"]), 1)

    def test_run_rescue_assignment_reuses_same_tutor_across_rounds_without_conflicts(self):
        failed_subgroups = [[
            {
                "ID": "M1",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 18:00:00"),
            },
            {
                "ID": "M2",
                "Data": "2026-02-24",
                "__match_datetime": pd.Timestamp("2026-02-24 20:00:00"),
            },
        ]]
        candidate_referees = pd.DataFrame([{"ID": "R1"}])
        committed = []

        def subgroup_cost_fn(referee_row, segment):
            taken_match_ids = {item["ID"] for assigned_segment in committed for item in assigned_segment}
            key = (referee_row["ID"], tuple(item["ID"] for item in segment))
            viable = {
                ("R1", ("M1",)): 10,
                ("R1", ("M2",)): 10 if "M1" in taken_match_ids else 1e6,
            }
            return viable.get(key, 1e6)

        def commit(_row, segment):
            committed.append(segment)

        rescue = _run_rescue_assignment(
            candidate_referees,
            failed_subgroups,
            subgroup_cost_fn,
            assignment_committer=commit,
        )

        self.assertEqual(len(rescue["rounds"]), 2)
        self.assertEqual(
            [[item["ID"] for item in assigned["segment"]] for assigned in rescue["assigned_segments"]],
            [["M1"], ["M2"]],
        )
        self.assertEqual(rescue["remaining_subgroups"], [])


class DesignacionsExcelExportTests(TestCase):
    def setUp(self):
        self.run = DesignationRun.objects.create(task_id="task-export-tests")

        self.ref_blank_level = Referee.objects.create(
            code="5001 F5",
            name="Tutor Sense Nivell",
            level="NIVELLA1",
            modality="Basquet",
            transport="Metro",
            active=True,
        )
        self.ref_grouped = Referee.objects.create(
            code="5002 F5",
            name="Tutor Agrupat",
            level="NIVELLA1",
            modality="Futbol Sala",
            transport="Cotxe",
            active=True,
        )
        self.ref_unassigned = Referee.objects.create(
            code="5003 F5",
            name="Tutor Sense Partits",
            level="",
            modality="Basquet",
            transport="Metro",
            active=True,
        )

        Availability.objects.create(
            run=self.run,
            referee=self.ref_grouped,
            raw={"Data": "2026-02-24", "Hora Inici": "17:00:00", "Hora Fi": "21:00:00"},
        )
        Availability.objects.create(
            run=self.run,
            referee=self.ref_grouped,
            raw={"Data": "2026-02-25", "Hora Inici": "18:00:00", "Hora Fi": "22:00:00"},
        )
        Availability.objects.create(
            run=self.run,
            referee=self.ref_blank_level,
            raw={
                "Data": "2026-02-24",
                "Modalitat": "Futbol Sala",
                "Nivell": "",
                "Mitjà de Transport": "Bus",
            },
        )
        Availability.objects.create(
            run=self.run,
            referee=self.ref_unassigned,
            raw={
                "Data": "2026-02-26",
                "Hora Inici": "19:00:00",
                "Hora Fi": "22:00:00",
                "Modalitat": "Futbol Sala",
                "Nivell": "NIVELLB1",
                "Mitjà de Transport": "Bus",
            },
        )

        match_blank_level = Match.objects.create(
            run=self.run,
            code="M-10",
            equip_local="Equip Local 1",
            equip_visitant="Equip Visitant 1",
            category="JUNIOR",
            date=date(2026, 2, 25),
            hour_raw="18:00:00",
            venue="Pista 1",
            municipality="Barcelona",
        )
        match_grouped_early = Match.objects.create(
            run=self.run,
            code="M-25",
            equip_local="Equip Local 2",
            equip_visitant="Equip Visitant 2",
            category="CADET",
            date=date(2026, 2, 24),
            hour_raw="18:30:00",
            venue="Pista 2",
            municipality="Barcelona",
        )
        match_grouped_late = Match.objects.create(
            run=self.run,
            code="M-30",
            equip_local="Equip Local 3",
            equip_visitant="Equip Visitant 3",
            category="INFANTIL",
            date=date(2026, 2, 25),
            hour_raw="19:00:00",
            venue="Pista 3",
            municipality="Barcelona",
        )
        match_unassigned = Match.objects.create(
            run=self.run,
            code="M-99",
            equip_local="Equip Local 4",
            equip_visitant="Equip Visitant 4",
            category="ALEVI",
            date=date(2026, 2, 26),
            hour_raw="20:15:00",
            venue="Pista 4",
            municipality="Badalona",
        )

        Assignment.objects.create(
            run=self.run,
            match=match_blank_level,
            referee=self.ref_blank_level,
            note="Revisar nivell",
            locked=True,
        )
        Assignment.objects.create(
            run=self.run,
            match=match_grouped_early,
            referee=self.ref_grouped,
            note="Primer partit",
            locked=False,
        )
        Assignment.objects.create(
            run=self.run,
            match=match_grouped_late,
            referee=self.ref_grouped,
            note="Segon partit",
            locked=False,
        )
        Assignment.objects.create(
            run=self.run,
            match=match_unassigned,
            referee=None,
            note="Pendent",
            locked=False,
        )

    def _export_workbook(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        self.addCleanup(lambda: os.path.exists(path) and os.unlink(path))
        export_run_to_excel(self.run, path)
        return load_workbook(path)

    def test_export_creates_expected_sheets_and_headers(self):
        wb = self._export_workbook()

        self.assertEqual(
            wb.sheetnames,
            [
                "Assignacions",
                "Partits sense assignar",
                "Tutors sense assignar",
                "Tutors sense nivell",
                "Validacions mobilitat",
            ],
        )

        ws = wb["Assignacions"]
        headers = [ws.cell(1, idx).value for idx in range(1, 15)]
        self.assertEqual(
            headers,
            [
                "Tutor Codi",
                "Tutor",
                "Nivell Tutor",
                "Hora Inici Tutor",
                "Hora Fi Tutor",
                "Data Partit",
                "Hora Partit",
                "Codi Partit",
                "Equip local",
                "Equip visitant",
                "Pista",
                "Categoria",
                "Nota",
                "Bloquejat",
            ],
        )
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertEqual(ws.auto_filter.ref, "A1:N4")

        ws_mobility = wb["Validacions mobilitat"]
        mobility_headers = [ws_mobility.cell(1, idx).value for idx in range(1, 8)]
        self.assertEqual(
            mobility_headers,
            [
                "Tipus",
                "Tutor Codi",
                "Tutor",
                "Motiu",
                "Clusters",
                "Partits",
                "Override manual",
            ],
        )

    def test_assignacions_sheet_is_sorted_grouped_and_styled(self):
        wb = self._export_workbook()
        ws = wb["Assignacions"]

        exported_rows = [
            (ws["A2"].value, ws["H2"].value),
            (ws["A3"].value, ws["H3"].value),
            (ws["A4"].value, ws["H4"].value),
        ]
        self.assertEqual(
            exported_rows,
            [
                ("5001 F5", "M-10"),
                ("5002 F5", "M-25"),
                ("5002 F5", "M-30"),
            ],
        )

        self.assertEqual(ws["F2"].number_format, "dd/mm/yyyy")
        self.assertEqual(ws["G2"].number_format, "hh:mm")
        self.assertEqual(ws["D2"].value, "-")
        self.assertEqual(ws["D3"].number_format, "hh:mm")
        self.assertEqual(ws["D3"].value, time(17, 0))
        self.assertEqual(ws["D4"].value, time(18, 0))
        self.assertEqual(ws["E3"].value, time(21, 0))
        self.assertEqual(ws["E4"].value, time(22, 0))
        self.assertEqual(ws["N2"].value, "Si")
        self.assertNotEqual(ws["A2"].fill.fgColor.rgb, ws["B2"].fill.fgColor.rgb)
        self.assertGreater(ws.column_dimensions["B"].width, 20)

    def test_secondary_sheets_include_unassigned_and_needs_review_data(self):
        wb = self._export_workbook()

        ws_unassigned_matches = wb["Partits sense assignar"]
        self.assertEqual(ws_unassigned_matches.max_row, 2)
        self.assertEqual(ws_unassigned_matches["A2"].value, "M-99")
        self.assertEqual(ws_unassigned_matches["I2"].value, "Pendent")

        ws_unassigned_refs = wb["Tutors sense assignar"]
        self.assertEqual(ws_unassigned_refs.max_row, 2)
        self.assertEqual(ws_unassigned_refs["A2"].value, "5003 F5")
        self.assertEqual(ws_unassigned_refs["B2"].value, "Tutor Sense Partits")
        self.assertEqual(ws_unassigned_refs["C2"].value, "NIVELLB1")
        self.assertEqual(ws_unassigned_refs["D2"].value, "Futbol Sala")
        self.assertEqual(ws_unassigned_refs["E2"].value, "Bus")

        ws_needs_review = wb["Tutors sense nivell"]
        self.assertEqual(ws_needs_review.max_row, 2)
        self.assertEqual(ws_needs_review["A2"].value, "5001 F5")
        self.assertEqual(ws_needs_review["C2"].value, "Futbol Sala")
        self.assertEqual(ws_needs_review["D2"].value, 1)

    def test_export_handles_empty_run(self):
        empty_run = DesignationRun.objects.create(task_id="task-export-empty")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        self.addCleanup(lambda: os.path.exists(path) and os.unlink(path))

        export_run_to_excel(empty_run, path)
        wb = load_workbook(path)

        self.assertEqual(wb["Assignacions"].max_row, 1)
        self.assertEqual(wb["Assignacions"].auto_filter.ref, "A1:N1")
        self.assertEqual(wb["Partits sense assignar"].max_row, 1)
        self.assertEqual(wb["Validacions mobilitat"].max_row, 1)


class DesignacionsManualAssignmentsTests(TestCase):
    def setUp(self):
        self.run = DesignationRun.objects.create(
            task_id="task-manual-assignments",
            params={
                "gap_same_pitch_min": 60,
                "gap_diff_pitch_min": 75,
                "gap_diff_cluster_min": 100,
                "availability_end_buffer_min": 60,
            },
        )

        self.ref_best = self._create_referee("5001 F5", "Tutor Millor", "NIVELLA1")
        self.ref_conflict_same = self._create_referee("5002 F5", "Tutor Mateixa Pista", "NIVELLA1")
        self.ref_modality = self._create_referee("5003 F5", "Tutor Modalitat", "NIVELLA1", modality="Basquet")
        self.ref_conflict_diff = self._create_referee("5004 F5", "Tutor Altra Pista", "NIVELLA1")
        self.ref_display = self._create_referee("5005 F5", "Tutor Amb Dues Disponibilitats", "NIVELLB1")
        self.ref_second = self._create_referee("5006 F5", "Tutor Segon", "NIVELLB1")
        self.ref_third = self._create_referee("5007 F5", "Tutor Tercer", "NIVELLC1")
        self.ref_fourth = self._create_referee("5008 F5", "Tutor Quart", "D")

        self._add_availability(self.ref_best, "2026-03-01", "17:00:00", "22:00:00")
        self._add_availability(self.ref_conflict_same, "2026-03-01", "17:00:00", "22:00:00")
        self._add_availability(self.ref_modality, "2026-03-01", "17:00:00", "22:00:00")
        self._add_availability(self.ref_conflict_diff, "2026-03-01", "17:00:00", "22:00:00")
        self._add_availability(self.ref_display, "2026-03-01", "16:00:00", "21:00:00")
        self._add_availability(self.ref_display, "2026-03-02", "18:00:00", "23:00:00")
        self._add_availability(self.ref_second, "2026-03-01", "17:00:00", "22:00:00")
        self._add_availability(self.ref_third, "2026-03-01", "17:00:00", "22:00:00")
        self._add_availability(self.ref_fourth, "2026-03-01", "17:00:00", "22:00:00")

        self.target_match = self._create_match("M-100", date(2026, 3, 1), "18:00:00", "Pista A", category="SÈNIOR")
        self.target_assignment = Assignment.objects.create(run=self.run, match=self.target_match, referee=None, note="Pendent")

        self.same_pitch_match = self._create_match("M-200", date(2026, 3, 1), "17:15:00", "Pista A", category="SÈNIOR")
        Assignment.objects.create(run=self.run, match=self.same_pitch_match, referee=self.ref_conflict_same)

        self.diff_pitch_match = self._create_match("M-210", date(2026, 3, 1), "17:00:00", "Pista B", category="SÈNIOR")
        Assignment.objects.create(run=self.run, match=self.diff_pitch_match, referee=self.ref_conflict_diff)

        self.display_match = self._create_match("M-300", date(2026, 3, 2), "19:00:00", "Pista C", category="INFANTIL")
        self.display_assignment = Assignment.objects.create(run=self.run, match=self.display_match, referee=self.ref_display)
        self.second_target_match = self._create_match("M-110", date(2026, 3, 1), "20:30:00", "Pista D", category="INFANTIL")
        self.second_target_assignment = Assignment.objects.create(run=self.run, match=self.second_target_match, referee=None, note="Pendent 2")
        self._set_match_cluster(self.target_match, 7)
        self._set_match_cluster(self.same_pitch_match, 7)
        self._set_match_cluster(self.diff_pitch_match, 7)
        self._set_match_cluster(self.display_match, 9)
        self._set_match_cluster(self.second_target_match, 7)

    def _create_referee(self, code, name, level, modality="Futbol Sala"):
        return Referee.objects.create(
            code=code,
            name=name,
            level=level,
            modality=modality,
            transport="Cotxe",
            active=True,
        )

    def _add_availability(self, referee, availability_date, start, end, *, level=None, modality=None, transport=None):
        Availability.objects.create(
            run=self.run,
            referee=referee,
            raw={
                "Data": availability_date,
                "Hora Inici": start,
                "Hora Fi": end,
                "Nivell": referee.level if level is None else level,
                "Modalitat": referee.modality if modality is None else modality,
                "Mitjà de Transport": referee.transport if transport is None else transport,
            },
        )

    def _create_match(self, code, match_date, hour_raw, venue, category="CADET", modality="Futbol Sala"):
        return Match.objects.create(
            run=self.run,
            code=code,
            equip_local=f"{code} Local",
            equip_visitant=f"{code} Visitant",
            category=category,
            modality=modality,
            date=match_date,
            hour_raw=hour_raw,
            domicile=venue,
            venue=venue,
            municipality="Barcelona",
        )

    def _set_match_cluster(self, match, cluster_id):
        address, _ = Address.objects.get_or_create(
            text=f"{match.domicile}, {match.municipality}",
            defaults={"municipality": match.municipality},
        )
        AddressCluster.objects.update_or_create(
            run=self.run,
            address=address,
            defaults={"cluster_id": cluster_id},
        )

    def test_diagnosis_marks_valid_candidate_with_cost(self):
        diagnosis = diagnose_assignment_for_referee(self.run, self.target_assignment, self.ref_best)

        self.assertTrue(diagnosis["is_valid"])
        self.assertIsNotNone(diagnosis["cost"])
        self.assertEqual(diagnosis["warning_reasons"], [])

    def test_diagnosis_marks_modality_mismatch_as_warning(self):
        diagnosis = diagnose_assignment_for_referee(self.run, self.target_assignment, self.ref_modality)

        self.assertFalse(diagnosis["is_valid"])
        self.assertIn("modality_mismatch", diagnosis["warning_reasons"])

    def test_diagnosis_detects_same_pitch_conflict(self):
        diagnosis = diagnose_assignment_for_referee(self.run, self.target_assignment, self.ref_conflict_same)

        self.assertFalse(diagnosis["is_valid"])
        self.assertIn("same_cluster_gap_violation", diagnosis["warning_reasons"])
        self.assertTrue(diagnosis["mobility_issues"][0].same_pitch)

    def test_diagnosis_detects_diff_pitch_conflict(self):
        diagnosis = diagnose_assignment_for_referee(self.run, self.target_assignment, self.ref_conflict_diff)

        self.assertFalse(diagnosis["is_valid"])
        self.assertIn("same_cluster_gap_violation", diagnosis["warning_reasons"])
        self.assertFalse(diagnosis["mobility_issues"][0].same_pitch)

    def test_explain_candidate_marks_ideal_assignment_as_recommended(self):
        explanation = explain_candidate_for_assignment(
            self.run,
            self.target_assignment,
            self.ref_best,
            context=build_manual_assignment_context(self.run),
        )

        self.assertEqual(explanation["feasibility"]["status"], "valid")
        self.assertEqual(explanation["level_fit"], "ideal")
        self.assertEqual(explanation["quality_label"], "recommended")
        self.assertIsNotNone(explanation["score_breakdown"]["base_level_cost"])
        self.assertIn("encaix de nivell ideal", explanation["selection_reason_summary"])

    def test_explain_candidate_marks_forced_when_underleveled_without_better_alternatives(self):
        forced_match = self._create_match("M-120", date(2026, 3, 1), "18:30:00", "Pista Forced", category="INFANTIL")
        forced_assignment = Assignment.objects.create(run=self.run, match=forced_match, referee=None)

        for referee, code in (
            (self.ref_best, "B-1"),
            (self.ref_second, "B-2"),
            (self.ref_display, "B-3"),
            (self.ref_third, "B-4"),
            (self.ref_conflict_same, "B-5"),
            (self.ref_conflict_diff, "B-6"),
        ):
            blocker_match = self._create_match(f"M-{code}", date(2026, 3, 1), "18:10:00", f"Pista {code}", category="CADET")
            Assignment.objects.create(run=self.run, match=blocker_match, referee=referee)

        context = build_manual_assignment_context(self.run)
        explanation = explain_candidate_for_assignment(self.run, forced_assignment, self.ref_fourth, context=context)
        better_alternatives = find_better_alternatives(self.run, forced_assignment, self.ref_fourth, context=context)

        self.assertEqual(explanation["level_fit"], "clearly_underleveled")
        self.assertEqual(explanation["quality_label"], "forced_by_constraints")
        self.assertEqual(better_alternatives, [])

    def test_explain_current_assignment_marks_suspicious_when_better_alternatives_exist(self):
        self.target_assignment.referee = self.ref_fourth
        self.target_assignment.manual_override_warning = False
        self.target_assignment.manual_override_reason = ""
        self.target_assignment.save(update_fields=["referee", "manual_override_warning", "manual_override_reason", "updated_at"])

        explanation = explain_current_assignment(
            self.run,
            self.target_assignment,
            context=build_manual_assignment_context(self.run),
        )

        self.assertEqual(explanation["explanation"]["level_fit"], "clearly_underleveled")
        self.assertEqual(explanation["explanation"]["quality_label"], "suspicious")
        self.assertGreaterEqual(len(explanation["better_alternatives"]), 1)
        self.assertEqual(explanation["better_alternatives"][0]["code"], "5001 F5")

    def test_explain_candidate_marks_unscorable_when_level_inputs_are_missing(self):
        unscorable_ref = self._create_referee("5016 F5", "Tutor Sense Nivell", "", modality="Futbol Sala")
        self._add_availability(unscorable_ref, "2026-03-01", "17:00:00", "22:00:00", level="")
        unscorable_match = self._create_match("M-130", date(2026, 3, 1), "19:00:00", "Pista U", category="DESCONEGUDA")
        unscorable_assignment = Assignment.objects.create(run=self.run, match=unscorable_match, referee=None)

        explanation = explain_candidate_for_assignment(
            self.run,
            unscorable_assignment,
            unscorable_ref,
            context=build_manual_assignment_context(self.run),
        )

        self.assertEqual(explanation["level_fit"], "unscorable")
        self.assertIsNone(explanation["score_breakdown"]["base_level_cost"])
        self.assertIn("no puntuable", explanation["selection_reason_summary"])

    def test_top_proposals_include_compatible_assigned_tutors_with_soft_penalty(self):
        proposals = build_top_proposals_for_assignments(self.run, [self.target_assignment])

        proposal_codes = [item["referee"].code for item in proposals[self.target_assignment.id]]
        self.assertEqual(proposal_codes, ["5001 F5", "5006 F5", "5005 F5"])
        self.assertNotIn("5002 F5", proposal_codes)
        self.assertNotIn("5004 F5", proposal_codes)
        self.assertNotIn("5003 F5", proposal_codes)
        self.assertEqual(len(proposal_codes), 3)
        self.assertGreater(proposals[self.target_assignment.id][2]["effective_cost"], proposals[self.target_assignment.id][1]["effective_cost"])

    def test_assignments_view_renders_placeholders_for_manual_loading(self):
        response = self.client.get(reverse("designacions_assignments", args=[self.run.id]))

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("eligible_proposals_by_assignment", response.context)
        self.assertNotIn("referee_options_by_assignment", response.context)
        self.assertEqual(
            response.context["availability_by_assignment"][self.display_assignment.id]["Hora Inici"],
            "18:00:00",
        )
        content = response.content.decode("utf-8")
        self.assertIn("Millors Propostes", content)
        self.assertIn("Carrega les propostes en obrir el selector.", content)
        self.assertIn("countMobilityWarnings", content)
        self.assertIn("countMobilityErrors", content)
        self.assertIn("mobilityErrorBanner", content)
        self.assertIn("data-manual-options-url", content)
        self.assertIn("data-manual-suggestions-bulk-url", content)
        self.assertIn("data-assignment-explanation-url", content)
        self.assertIn('data-suggestions-state="idle"', content)
        self.assertIn('data-rank1-referee-id=""', content)
        self.assertIn('data-last-suggestions-revision="0"', content)
        self.assertIn("data-update-async-url", content)
        self.assertIn("assignment-explain-btn", content)
        self.assertIn("quality-badge--recommended", content)
        self.assertIn("scheduleInitialSuggestionPreload", content)
        self.assertIn("markAllUnassignedRowsStaleAndRefresh", content)
        self.assertIn("row.dataset.rank1RefereeId === String(prioritizedRefereeId)", content)
        self.assertIn("loadAssignmentExplanation", content)
        self.assertIn("SUGGESTION_BATCH_SIZE = 10", content)
        self.assertIn("SUGGESTION_MAX_CONCURRENCY = 2", content)
        self.assertIn("proposal-rank--1", content)
        self.assertIn("proposal-rank--2", content)
        self.assertIn("proposal-rank--3", content)

    def test_manual_assignment_options_view_returns_top_proposals_and_compatible_referees(self):
        response = self.client.get(
            reverse("designacions_manual_assignment_options", args=[self.run.id, self.target_assignment.id])
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            [item["code"] for item in payload["top_proposals"]],
            ["5001 F5", "5006 F5", "5005 F5"],
        )
        self.assertEqual(payload["top_proposals"][0]["level_fit"], "ideal")
        self.assertEqual(payload["top_proposals"][0]["quality_label"], "recommended")
        self.assertTrue(payload["top_proposals"][0]["selection_reason_summary"])
        self.assertEqual(
            [item["code"] for item in payload["compatible_referees"]],
            ["5001 F5", "5002 F5", "5004 F5", "5005 F5", "5006 F5", "5007 F5", "5008 F5"],
        )
        by_code = {item["code"]: item for item in payload["compatible_referees"]}
        self.assertTrue(by_code["5001 F5"]["is_valid"])
        self.assertFalse(by_code["5002 F5"]["is_valid"])
        self.assertIn("mateix cluster", by_code["5002 F5"]["warning_text"].lower())
        self.assertFalse(by_code["5004 F5"]["is_valid"])
        self.assertIn("mateix cluster", by_code["5004 F5"]["warning_text"].lower())

    def test_manual_assignment_options_view_filters_compatible_referees_by_query(self):
        response = self.client.get(
            reverse("designacions_manual_assignment_options", args=[self.run.id, self.target_assignment.id]),
            {"q": "5006"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([item["code"] for item in payload["compatible_referees"]], ["5006 F5"])

    def test_manual_assignment_options_view_marks_cross_cluster_without_vehicle_as_warning(self):
        no_vehicle_ref = self._create_referee("5017 F5", "Tutor Sense Vehicle", "NIVELLA1")
        no_vehicle_ref.transport = "Bus"
        no_vehicle_ref.save(update_fields=["transport"])
        self._add_availability(no_vehicle_ref, "2026-03-01", "17:00:00", "22:00:00", transport="Bus")
        remote_match = self._create_match("M-401", date(2026, 3, 1), "16:00:00", "Pista Remota", category="CADET")
        self._set_match_cluster(remote_match, 8)
        Assignment.objects.create(run=self.run, match=remote_match, referee=no_vehicle_ref)

        response = self.client.get(
            reverse("designacions_manual_assignment_options", args=[self.run.id, self.target_assignment.id])
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        by_code = {item["code"]: item for item in payload["compatible_referees"]}
        self.assertFalse(by_code["5017 F5"]["is_valid"])
        self.assertIn("sense vehicle", by_code["5017 F5"]["warning_text"].lower())

    def test_build_run_mobility_summary_marks_valid_multi_cluster_assignment_as_warning(self):
        vehicle_ref = self._create_referee("5019 F5", "Tutor Multi Cluster", "NIVELLA1")
        self._add_availability(vehicle_ref, "2026-03-01", "15:00:00", "22:00:00", transport="Cotxe")
        early_match = self._create_match("M-410", date(2026, 3, 1), "16:00:00", "Pista Cl1", category="CADET")
        late_match = self._create_match("M-411", date(2026, 3, 1), "19:00:00", "Pista Cl2", category="CADET")
        self._set_match_cluster(early_match, 7)
        self._set_match_cluster(late_match, 8)
        Assignment.objects.create(run=self.run, match=early_match, referee=vehicle_ref)
        Assignment.objects.create(run=self.run, match=late_match, referee=vehicle_ref)

        summary = build_run_mobility_summary(self.run)

        self.assertEqual(summary["mobility_warning_count"], 1)
        self.assertEqual(summary["mobility_error_count"], 0)
        self.assertEqual(summary["mobility_warnings"][0]["referee_code"], "5019 F5")

    def test_assignment_explanation_view_returns_expected_shape_and_better_alternatives(self):
        self.target_assignment.referee = self.ref_fourth
        self.target_assignment.manual_override_warning = False
        self.target_assignment.manual_override_reason = ""
        self.target_assignment.save(update_fields=["referee", "manual_override_warning", "manual_override_reason", "updated_at"])

        response = self.client.get(
            reverse("designacions_assignment_explanation", args=[self.run.id, self.target_assignment.id])
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["assignment_id"], self.target_assignment.id)
        self.assertEqual(payload["assigned_referee"]["code"], "5008 F5")
        self.assertEqual(payload["explanation"]["quality_label"], "suspicious")
        self.assertEqual(payload["better_alternatives"][0]["code"], "5001 F5")
        self.assertNotIn("5008 F5", [item["code"] for item in payload["better_alternatives"]])

    @patch("designacions.views.build_manual_assignment_context", wraps=build_manual_assignment_context)
    def test_manual_assignment_suggestions_bulk_view_returns_sorted_items_and_builds_context_once(self, build_context_mock):
        other_run = DesignationRun.objects.create(task_id="task-other-run")
        other_match = Match.objects.create(
            run=other_run,
            code="M-999",
            equip_local="Altres Local",
            equip_visitant="Altres Visitant",
            category="INFANTIL",
            modality="Futbol Sala",
            date=date(2026, 3, 1),
            hour_raw="18:00:00",
            venue="Pista Z",
            municipality="Barcelona",
        )
        other_assignment = Assignment.objects.create(run=other_run, match=other_match, referee=None)

        response = self.client.post(
            reverse("designacions_manual_assignment_suggestions_bulk", args=[self.run.id]),
            data=json.dumps({
                "assignment_ids": [self.second_target_assignment.id, 999999, self.target_assignment.id, other_assignment.id],
                "limit": 3,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            [item["assignment_id"] for item in payload["items"]],
            sorted([self.second_target_assignment.id, 999999, self.target_assignment.id, other_assignment.id]),
        )
        item_by_assignment = {item["assignment_id"]: item for item in payload["items"]}
        self.assertEqual(item_by_assignment[self.target_assignment.id]["status"], "ok")
        self.assertEqual(item_by_assignment[self.second_target_assignment.id]["status"], "ok")
        self.assertEqual(item_by_assignment[other_assignment.id]["status"], "missing_assignment")
        self.assertEqual(item_by_assignment[999999]["status"], "missing_assignment")
        self.assertEqual(item_by_assignment[self.target_assignment.id]["rank_1_referee_id"], self.ref_best.id)
        self.assertEqual(
            [item["code"] for item in item_by_assignment[self.target_assignment.id]["top_proposals"]],
            ["5001 F5", "5006 F5", "5005 F5"],
        )
        self.assertEqual(build_context_mock.call_count, 1)

    def test_manual_assignment_suggestions_bulk_view_matches_single_assignment_endpoint_top_proposals(self):
        bulk_response = self.client.post(
            reverse("designacions_manual_assignment_suggestions_bulk", args=[self.run.id]),
            data=json.dumps({"assignment_ids": [self.target_assignment.id], "limit": 3}),
            content_type="application/json",
        )
        single_response = self.client.get(
            reverse("designacions_manual_assignment_options", args=[self.run.id, self.target_assignment.id])
        )

        self.assertEqual(bulk_response.status_code, 200)
        self.assertEqual(single_response.status_code, 200)
        bulk_payload = bulk_response.json()
        single_payload = single_response.json()
        self.assertEqual(len(bulk_payload["items"]), 1)
        self.assertEqual(bulk_payload["items"][0]["status"], "ok")
        self.assertEqual(bulk_payload["items"][0]["top_proposals"], single_payload["top_proposals"])

    def test_run_scoped_summary_prefers_run_availability_fields(self):
        scoped_ref = self._create_referee("5011 F5", "Tutor Scoped", "", modality="Basquet")
        self._add_availability(
            scoped_ref,
            "2026-03-01",
            "17:00:00",
            "22:00:00",
            level="NIVELLC1",
            modality="Futbol Sala",
            transport="Bus",
        )
        self._add_availability(
            scoped_ref,
            "2026-03-02",
            "17:00:00",
            "22:00:00",
            level="NIVELLC1",
            modality="Futbol Sala",
            transport="Bus",
        )

        summary = next(
            item
            for item in build_run_scoped_referee_summaries(self.run)
            if item.id == scoped_ref.id
        )

        self.assertEqual(summary.level, "NIVELLC1")
        self.assertEqual(summary.modality, "Futbol Sala")
        self.assertEqual(summary.transport, "Bus")

    def test_assignments_view_uses_run_scoped_tutor_metadata(self):
        scoped_unassigned = self._create_referee("5012 F5", "Tutor Llista Scoped", "", modality="Basquet")
        self._add_availability(
            scoped_unassigned,
            "2026-03-01",
            "17:00:00",
            "22:00:00",
            level="NIVELLC1",
            modality="Futbol Sala",
            transport="Bus",
        )
        scoped_blank = self._create_referee("5013 F5", "Tutor Blank Scoped", "NIVELLA1")
        self._add_availability(
            scoped_blank,
            "2026-03-01",
            "17:00:00",
            "22:00:00",
            level="",
            modality="Futbol Sala",
        )
        scoped_assigned = self._create_referee("5014 F5", "Tutor Assignat Scoped", "", modality="Basquet")
        self._add_availability(
            scoped_assigned,
            "2026-03-02",
            "17:00:00",
            "22:00:00",
            level="NIVELLB1",
            modality="Futbol Sala",
            transport="Bus",
        )
        scoped_match = self._create_match("M-350", date(2026, 3, 2), "18:30:00", "Pista D")
        Assignment.objects.create(run=self.run, match=scoped_match, referee=scoped_assigned)

        response = self.client.get(reverse("designacions_assignments", args=[self.run.id]))

        self.assertEqual(response.status_code, 200)
        unassigned_by_code = {ref.code: ref for ref in response.context["unassigned_referees"]}
        needs_review_codes = {ref.code for ref in response.context["needs_review_referees"]}
        group_levels = {group["referee"].code: group["referee"].level for group in response.context["groups"]}

        self.assertEqual(unassigned_by_code["5012 F5"].modality, "Futbol Sala")
        self.assertEqual(unassigned_by_code["5012 F5"].level, "NIVELLC1")
        self.assertIn("5013 F5", needs_review_codes)
        self.assertNotIn("5012 F5", needs_review_codes)
        self.assertEqual(group_levels["5014 F5"], "NIVELLB1")

    @patch("designacions.views.rebuild_run_map_task.delay")
    def test_update_assignment_async_returns_run_scoped_level_in_payload(self, rebuild_map_delay_mock):
        scoped_ref = self._create_referee("5015 F5", "Tutor Payload Scoped", "", modality="Basquet")
        self._add_availability(
            scoped_ref,
            "2026-03-01",
            "17:00:00",
            "22:00:00",
            level="NIVELLC1",
            modality="Futbol Sala",
        )

        response = self.client.post(
            reverse("designacions_update_assignment_async", args=[self.run.id, self.target_assignment.id]),
            data=json.dumps({"referee_id": scoped_ref.id, "note": "Scoped level", "locked": False}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["referee"]["level"], "NIVELLC1")
        rebuild_map_delay_mock.assert_called_once_with(self.run.id)

    @patch("designacions.views.rebuild_run_map_task.delay")
    def test_invalid_manual_assignment_for_other_modality_is_rejected(self, rebuild_map_delay_mock):
        response = self.client.post(
            reverse("designacions_update_assignment", args=[self.run.id, self.target_assignment.id]),
            {"referee_id": str(self.ref_modality.id), "note": "Override"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.target_assignment.refresh_from_db()
        self.assertIsNone(self.target_assignment.referee_id)
        self.assertFalse(self.target_assignment.manual_override_warning)
        self.assertEqual(self.target_assignment.manual_override_reason, "")
        rebuild_map_delay_mock.assert_not_called()

    @patch("designacions.views.rebuild_run_map_task.delay")
    def test_valid_manual_assignment_clears_previous_warning(self, rebuild_map_delay_mock):
        self.target_assignment.referee = self.ref_modality
        self.target_assignment.manual_override_warning = True
        self.target_assignment.manual_override_reason = "Modalitat diferent de la del partit."
        self.target_assignment.save(update_fields=["referee", "manual_override_warning", "manual_override_reason", "updated_at"])

        response = self.client.post(
            reverse("designacions_update_assignment", args=[self.run.id, self.target_assignment.id]),
            {"referee_id": str(self.ref_best.id), "note": "Corregit"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.target_assignment.refresh_from_db()
        self.assertEqual(self.target_assignment.referee_id, self.ref_best.id)
        self.assertFalse(self.target_assignment.manual_override_warning)
        self.assertEqual(self.target_assignment.manual_override_reason, "")
        self.run.refresh_from_db()
        self.assertEqual(self.run.map_status, "queued")
        rebuild_map_delay_mock.assert_called_once_with(self.run.id)

    @patch("designacions.views.rebuild_run_map_task.delay")
    def test_update_assignment_async_returns_json_and_queues_map(self, rebuild_map_delay_mock):
        response = self.client.post(
            reverse("designacions_update_assignment_async", args=[self.run.id, self.target_assignment.id]),
            data=json.dumps({"referee_id": self.ref_best.id, "note": "Async ok", "locked": False}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.target_assignment.refresh_from_db()
        self.run.refresh_from_db()
        self.assertEqual(self.target_assignment.referee_id, self.ref_best.id)
        self.assertEqual(payload["row_state"], "assigned")
        self.assertEqual(payload["assigned_referee_id"], self.ref_best.id)
        self.assertEqual(payload["assigned_referee_code"], self.ref_best.code)
        self.assertTrue(payload["refresh_suggestions"])
        self.assertFalse(payload["warning"]["active"])
        self.assertEqual(payload["counts"]["assigned"], 4)
        self.assertEqual(self.run.map_status, "queued")
        rebuild_map_delay_mock.assert_called_once_with(self.run.id)

    @patch("designacions.views.rebuild_run_map_task.delay")
    def test_update_assignment_async_returns_warning_payload_when_candidate_is_allowed_but_costly(self, rebuild_map_delay_mock):
        warning_ref = self._create_referee("5010 F5", "Tutor Warning", "NIVELLA1")
        self._add_availability(warning_ref, "2026-03-02", "18:00:00", "22:00:00")

        response = self.client.post(
            reverse("designacions_update_assignment_async", args=[self.run.id, self.target_assignment.id]),
            data=json.dumps({"referee_id": warning_ref.id, "note": "Warning", "locked": False}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.target_assignment.refresh_from_db()
        self.assertEqual(self.target_assignment.referee_id, warning_ref.id)
        self.assertTrue(payload["warning"]["active"])
        self.assertIn("Sense disponibilitat registrada", payload["warning"]["text"])
        rebuild_map_delay_mock.assert_called_once_with(self.run.id)

    @patch("designacions.views.rebuild_run_map_task.delay")
    def test_update_assignment_async_allows_manual_mobility_override_and_records_summary_error(self, rebuild_map_delay_mock):
        no_vehicle_ref = self._create_referee("5018 F5", "Tutor Override Mobilitat", "NIVELLA1")
        no_vehicle_ref.transport = "Bus"
        no_vehicle_ref.save(update_fields=["transport"])
        self._add_availability(no_vehicle_ref, "2026-03-01", "17:00:00", "22:00:00", transport="Bus")
        remote_match = self._create_match("M-402", date(2026, 3, 1), "16:00:00", "Pista Remota 2", category="CADET")
        self._set_match_cluster(remote_match, 8)
        Assignment.objects.create(run=self.run, match=remote_match, referee=no_vehicle_ref)

        response = self.client.post(
            reverse("designacions_update_assignment_async", args=[self.run.id, self.target_assignment.id]),
            data=json.dumps({"referee_id": no_vehicle_ref.id, "note": "Override mobilitat", "locked": False}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.target_assignment.refresh_from_db()
        self.assertEqual(self.target_assignment.referee_id, no_vehicle_ref.id)
        self.assertTrue(self.target_assignment.manual_override_warning)
        self.assertIn("sense vehicle", self.target_assignment.manual_override_reason.lower())
        self.assertTrue(payload["warning"]["active"])
        self.assertEqual(payload["counts"]["mobility_errors"], 1)
        self.assertEqual(payload["mobility_summary"]["mobility_error_count"], 1)
        self.assertTrue(payload["mobility_summary"]["mobility_errors"][0]["manual_override"])
        rebuild_map_delay_mock.assert_called_once_with(self.run.id)

    @patch("designacions.views.rebuild_run_map_task.delay")
    def test_update_assignment_async_rejects_incompatible_referee(self, rebuild_map_delay_mock):
        response = self.client.post(
            reverse("designacions_update_assignment_async", args=[self.run.id, self.target_assignment.id]),
            data=json.dumps({"referee_id": self.ref_modality.id, "note": "Invalid", "locked": False}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.target_assignment.refresh_from_db()
        self.assertIsNone(self.target_assignment.referee_id)
        rebuild_map_delay_mock.assert_not_called()

    @patch("designacions.views.rebuild_run_map_task.delay")
    def test_update_assignment_async_can_unassign(self, rebuild_map_delay_mock):
        self.target_assignment.referee = self.ref_best
        self.target_assignment.manual_override_warning = True
        self.target_assignment.manual_override_reason = "Old warning"
        self.target_assignment.locked = True
        self.target_assignment.save(update_fields=["referee", "manual_override_warning", "manual_override_reason", "locked", "updated_at"])

        response = self.client.post(
            reverse("designacions_update_assignment_async", args=[self.run.id, self.target_assignment.id]),
            data=json.dumps({"referee_id": None, "note": "Sense assignar", "locked": False}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.target_assignment.refresh_from_db()
        self.assertIsNone(self.target_assignment.referee_id)
        self.assertFalse(self.target_assignment.locked)
        self.assertFalse(self.target_assignment.manual_override_warning)
        self.assertEqual(self.target_assignment.manual_override_reason, "")
        self.assertEqual(payload["row_state"], "unassigned")
        self.assertIsNone(payload["assigned_referee_id"])
        self.assertEqual(payload["assigned_referee_code"], "")
        self.assertTrue(payload["refresh_suggestions"])
        rebuild_map_delay_mock.assert_called_once_with(self.run.id)

    @patch("designacions.tasks.rebuild_run_map", return_value="designacions/maps/run_test.html")
    def test_rebuild_run_map_task_updates_map_status(self, rebuild_run_map_mock):
        self.run.map_status = "queued"
        self.run.save(update_fields=["map_status"])

        result = rebuild_run_map_task.run(self.run.id)

        self.run.refresh_from_db()
        self.assertEqual(self.run.map_status, "ready")
        self.assertEqual(result["run_id"], self.run.id)
        rebuild_run_map_mock.assert_called_once()


class _FakeAsyncRedis:
    def __init__(self, store, lists, published):
        self.store = store
        self.lists = lists
        self.published = published

    async def get(self, key):
        return self.store.get(key)

    async def set(self, key, value):
        self.store[key] = value

    async def rpush(self, key, value):
        self.lists.setdefault(key, []).append(value)

    async def publish(self, channel, value):
        self.published.append((channel, value))

    async def aclose(self):
        return None


class DesignacionsJobStoreTests(SimpleTestCase):
    def setUp(self):
        self.redis_store = {}
        self.redis_lists = {}
        self.redis_published = []

    def _fake_redis(self):
        return _FakeAsyncRedis(self.redis_store, self.redis_lists, self.redis_published)

    @patch("logs._redis_async")
    def test_job_store_merges_status_and_keeps_progress_monotonic(self, redis_async_mock):
        redis_async_mock.side_effect = self._fake_redis

        asyncio.run(_write_job("task-1", {"status": "processing", "progress": 55}))
        asyncio.run(_write_job("task-1", {"message": "Seguim"}))
        asyncio.run(push_log("task-1", "Intent de regressio", 10))
        asyncio.run(push_log("task-1", "Error final", status="failed"))

        job = asyncio.run(_read_job("task-1"))

        self.assertEqual(job["status"], "failed")
        self.assertEqual(job["progress"], 55)
        self.assertEqual(job["message"], "Error final")
        self.assertIn("job:task-1:logs", self.redis_lists)
        self.assertEqual(len(self.redis_published), 2)


class _AsyncClientContext:
    async def __aenter__(self):
        return object()

    async def __aexit__(self, exc_type, exc, tb):
        return False


class DesignacionsClassificacionsFetchTests(SimpleTestCase):
    @patch("designacions.consulta_resultats.httpx.AsyncClient", side_effect=lambda *args, **kwargs: _AsyncClientContext())
    @patch("designacions.consulta_resultats._fetch_ceeb_once")
    def test_fetch_classificacions_uses_cache_for_same_key(self, fetch_once_mock, async_client_mock):
        root = ET.fromstring(
            "<root><grup_classificacions><info_lliga><nomGrup>GRUP 01</nomGrup></info_lliga></grup_classificacions></root>"
        )
        fetch_once_mock.return_value = root
        cache = {}

        first = asyncio.run(fetch_ceeb_classification_async("9.0", "SXMIX", fase="FS1", cache=cache))
        second = asyncio.run(fetch_ceeb_classification_async("9.0", "SXMIX", fase="FS1", cache=cache))

        self.assertIs(first.root, root)
        self.assertFalse(first.from_cache)
        self.assertTrue(second.from_cache)
        self.assertIs(second.root, root)
        self.assertEqual(fetch_once_mock.call_count, 1)
        self.assertEqual(async_client_mock.call_count, 1)

    @patch("designacions.consulta_resultats.httpx.AsyncClient", side_effect=lambda *args, **kwargs: _AsyncClientContext())
    @patch("designacions.consulta_resultats._fetch_ceeb_once")
    def test_fetch_classificacions_retries_transient_timeout(self, fetch_once_mock, _async_client_mock):
        root = ET.fromstring(
            "<root><grup_classificacions><info_lliga><nomGrup>GRUP 01</nomGrup></info_lliga></grup_classificacions></root>"
        )
        fetch_once_mock.side_effect = [httpx.TimeoutException("boom"), root]

        result = asyncio.run(
            fetch_ceeb_classification_async("9.0", "SXMIX", fase="FS1", max_retries=2, backoff_seconds=0)
        )

        self.assertIs(result.root, root)
        self.assertEqual(result.attempts, 2)
        self.assertIsNone(result.error)

    @patch("designacions.consulta_resultats.httpx.AsyncClient", side_effect=lambda *args, **kwargs: _AsyncClientContext())
    @patch("designacions.consulta_resultats._fetch_ceeb_once")
    def test_fetch_classificacions_returns_warning_after_persistent_error(self, fetch_once_mock, _async_client_mock):
        fetch_once_mock.side_effect = httpx.TimeoutException("boom")

        result = asyncio.run(
            fetch_ceeb_classification_async("9.0", "SXMIX", fase="FS1", max_retries=3, backoff_seconds=0)
        )

        self.assertIsNone(result.root)
        self.assertEqual(result.attempts, 3)
        self.assertIn("Timeout", result.error)


class DesignacionsProgressFlowTests(SimpleTestCase):
    @patch("designacions.main_fixed.mapa_assignacions_interactiu", side_effect=lambda **kwargs: kwargs["out_html"])
    @patch("designacions.main_fixed.persist_assignacions_to_db")
    @patch("designacions.main_fixed.AddressCluster.objects.update_or_create")
    @patch("designacions.main_fixed.Address.objects.filter")
    @patch("designacions.main_fixed.clusteritza_i_plota")
    @patch("designacions.main_fixed.addresses_to_df")
    @patch("designacions.main_fixed.geocodifica_adreces")
    @patch("designacions.main_fixed.load_modalitat_map_df")
    @patch("designacions.main_fixed.fetch_ceeb_classification_async")
    @patch("designacions.main_fixed.push_log")
    def test_main_emits_non_decreasing_progress_with_multiple_modalitats(
        self,
        push_log_mock,
        fetch_classification_mock,
        load_modalitat_map_df_mock,
        geocodifica_adreces_mock,
        addresses_to_df_mock,
        clusteritza_i_plota_mock,
        address_filter_mock,
        _address_cluster_update_mock,
        _persist_mock,
        _map_mock,
    ):
        captured_progress = []

        async def fake_push_log(_task_id, _message, progress=None, status=None):
            if isinstance(progress, int):
                captured_progress.append(progress)

        async def fake_fetch_classification(id_categoria, p5, **kwargs):
            xml_by_key = {
                ("9.0", "SXMIX"): ET.fromstring(
                    """
                    <root>
                      <grup_classificacions>
                        <info_lliga><nomGrup>GRUP A</nomGrup></info_lliga>
                        <prt_class_all>
                          <equip><NomEquipMostrar>Equip A Local</NomEquipMostrar></equip>
                          <equip><NomEquipMostrar>Equip A Visitant</NomEquipMostrar></equip>
                        </prt_class_all>
                      </grup_classificacions>
                    </root>
                    """
                ),
                ("10.0", "SXFEM"): ET.fromstring(
                    """
                    <root>
                      <grup_classificacions>
                        <info_lliga><nomGrup>GRUP B</nomGrup></info_lliga>
                        <prt_class_all>
                          <equip><NomEquipMostrar>Equip B Local</NomEquipMostrar></equip>
                          <equip><NomEquipMostrar>Equip B Visitant</NomEquipMostrar></equip>
                        </prt_class_all>
                      </grup_classificacions>
                    </root>
                    """
                ),
            }
            return SimpleNamespace(
                root=xml_by_key[(id_categoria, p5)],
                from_cache=False,
                attempts=1,
                error=None,
            )

        push_log_mock.side_effect = fake_push_log
        fetch_classification_mock.side_effect = fake_fetch_classification
        load_modalitat_map_df_mock.return_value = pd.DataFrame(
            [
                {"Id Categoria": "9.0", "Modalitat": "Futbol Sala", "Nom": "CADET"},
                {"Id Categoria": "10.0", "Modalitat": "Basquet", "Nom": "INFANTIL"},
            ]
        )
        geocodifica_adreces_mock.return_value = ["ok"]
        addresses_to_df_mock.return_value = pd.DataFrame(
            [
                {"adreca": "Carrer 1, Barcelona", "lat": 41.0, "lon": 2.0},
                {"adreca": "Carrer 2, Barcelona", "lat": 41.1, "lon": 2.1},
            ]
        )
        clusteritza_i_plota_mock.return_value = (
            pd.DataFrame(
                [
                    {"adreca": "Carrer 1, Barcelona", "cluster": 1, "lat": 41.0, "lon": 2.0},
                    {"adreca": "Carrer 2, Barcelona", "cluster": 2, "lat": 41.1, "lon": 2.1},
                ]
            ),
            None,
            None,
            None,
        )
        address_filter_mock.return_value.first.return_value = None

        df_dispos = pd.DataFrame(
            [
                {
                    "Codi Tutor de Joc": "5001 F5",
                    "Nom": "Tutor A",
                    "Cognoms": "Alpha",
                    "Categoria": "TUTOR/TUTORA DE JOC",
                    "Modalitat": "Futbol Sala",
                    "Nivell": "NIVELLA1",
                    "Data": "2026-03-01",
                    "Hora Inici": "17:00:00",
                    "Hora Fi": "21:00:00",
                },
                {
                    "Codi Tutor de Joc": "5002 BQ",
                    "Nom": "Tutor B",
                    "Cognoms": "Beta",
                    "Categoria": "TUTOR/TUTORA DE JOC",
                    "Modalitat": "Basquet",
                    "Nivell": "NIVELLA1",
                    "Data": "2026-03-01",
                    "Hora Inici": "17:00:00",
                    "Hora Fi": "21:00:00",
                },
            ]
        )
        df_partits = pd.DataFrame(
            [
                {
                    "Codi": "M-A",
                    "Codi Extern Local": "EXT-A",
                    "Lliga": "Lliga A",
                    "Modalitat": "Futbol Sala",
                    "Categoria": "CADET",
                    "Subcategoria": "MIXT",
                    "Data": "2026-03-01",
                    "Hora": "18:00:00",
                    "Domicili": "Carrer 1",
                    "Municipi": "Barcelona",
                    "Pista joc": "Pista 1",
                    "Grup": "GRUP A",
                    "Equip local": "Equip A Local",
                    "Equip visitant": "Equip A Visitant",
                    "Club Local": "Club A",
                },
                {
                    "Codi": "M-B",
                    "Codi Extern Local": "EXT-B",
                    "Lliga": "Lliga B",
                    "Modalitat": "Basquet",
                    "Categoria": "INFANTIL",
                    "Subcategoria": "FEMENÍ",
                    "Data": "2026-03-01",
                    "Hora": "19:00:00",
                    "Domicili": "Carrer 2",
                    "Municipi": "Barcelona",
                    "Pista joc": "Pista 2",
                    "Grup": "GRUP B",
                    "Equip local": "Equip B Local",
                    "Equip visitant": "Equip B Visitant",
                    "Club Local": "Club B",
                },
            ]
        )

        from .main_fixed import main as engine_main

        result = engine_main(
            "disp.xlsx",
            "partits.xlsx",
            task_id="task-progress",
            run_id=1,
            config={"fase": "FS1"},
            df_dispos=df_dispos,
            df_partits=df_partits,
        )

        self.assertEqual(captured_progress, sorted(captured_progress))
        self.assertEqual(result["classification_failed_requests"], 0)
        self.assertEqual(result["classification_cache_hits"], 0)
        self.assertIn("initial_subgroups", result)
        self.assertIn("rescue_rounds", result)
        self.assertIn("remaining_unassigned_breakdown", result)

    @patch("designacions.main_fixed.mapa_assignacions_interactiu", side_effect=lambda **kwargs: kwargs["out_html"])
    @patch("designacions.main_fixed.persist_assignacions_to_db")
    @patch("designacions.main_fixed.AddressCluster.objects.update_or_create")
    @patch("designacions.main_fixed.Address.objects.filter")
    @patch("designacions.main_fixed.clusteritza_i_plota")
    @patch("designacions.main_fixed.addresses_to_df")
    @patch("designacions.main_fixed.geocodifica_adreces")
    @patch("designacions.main_fixed.load_modalitat_map_df")
    @patch("designacions.main_fixed.fetch_ceeb_classification_async")
    def test_main_reuses_assigned_referee_in_rescue_to_improve_coverage(
        self,
        fetch_classification_mock,
        load_modalitat_map_df_mock,
        geocodifica_adreces_mock,
        addresses_to_df_mock,
        clusteritza_i_plota_mock,
        address_filter_mock,
        _address_cluster_update_mock,
        _persist_mock,
        _map_mock,
    ):
        fetch_classification_mock.return_value = SimpleNamespace(
            root=ET.fromstring(
                """
                <root>
                  <grup_classificacions>
                    <info_lliga><nomGrup>GRUP A</nomGrup></info_lliga>
                    <prt_class_all>
                      <equip><NomEquipMostrar>Equip A Local</NomEquipMostrar></equip>
                      <equip><NomEquipMostrar>Equip A Visitant</NomEquipMostrar></equip>
                    </prt_class_all>
                  </grup_classificacions>
                </root>
                """
            ),
            from_cache=False,
            attempts=1,
            error=None,
        )
        load_modalitat_map_df_mock.return_value = pd.DataFrame(
            [{"Id Categoria": "9.0", "Modalitat": "Futbol Sala", "Nom": "CADET"}]
        )
        geocodifica_adreces_mock.return_value = ["ok"]
        addresses_to_df_mock.return_value = pd.DataFrame(
            [
                {"adreca": "Carrer 1, Barcelona", "lat": 41.0, "lon": 2.0},
                {"adreca": "Carrer 2, Barcelona", "lat": 41.1, "lon": 2.1},
            ]
        )
        clusteritza_i_plota_mock.return_value = (
            pd.DataFrame(
                [
                    {"adreca": "Carrer 1, Barcelona", "cluster": 1, "lat": 41.0, "lon": 2.0},
                    {"adreca": "Carrer 2, Barcelona", "cluster": 2, "lat": 41.1, "lon": 2.1},
                ]
            ),
            None,
            None,
            None,
        )
        address_filter_mock.return_value.first.return_value = None

        df_dispos = pd.DataFrame(
            [
                {
                    "Codi Tutor de Joc": "5001 F5",
                    "Nom": "Tutor A",
                    "Cognoms": "Alpha",
                    "Categoria": "TUTOR/TUTORA DE JOC",
                    "Modalitat": "Futbol Sala",
                    "Nivell": "NIVELLA1",
                    "Data": "2026-03-01",
                    "Hora Inici": "17:00:00",
                    "Hora Fi": "23:00:00",
                    "Mitjà de Transport": "Cotxe",
                },
            ]
        )
        df_partits = pd.DataFrame(
            [
                {
                    "Codi": "M-A",
                    "Codi Extern Local": "EXT-A",
                    "Lliga": "Lliga A",
                    "Modalitat": "Futbol Sala",
                    "Categoria": "CADET",
                    "Subcategoria": "MIXT",
                    "Data": "2026-03-01",
                    "Hora": "18:00:00",
                    "Domicili": "Carrer 1",
                    "Municipi": "Barcelona",
                    "Pista joc": "Pista 1",
                    "Grup": "GRUP A",
                    "Equip local": "Equip A Local",
                    "Equip visitant": "Equip A Visitant",
                    "Club Local": "Club A",
                },
                {
                    "Codi": "M-B",
                    "Codi Extern Local": "EXT-B",
                    "Lliga": "Lliga A",
                    "Modalitat": "Futbol Sala",
                    "Categoria": "CADET",
                    "Subcategoria": "MIXT",
                    "Data": "2026-03-01",
                    "Hora": "20:30:00",
                    "Domicili": "Carrer 2",
                    "Municipi": "Barcelona",
                    "Pista joc": "Pista 2",
                    "Grup": "GRUP A",
                    "Equip local": "Equip A Local",
                    "Equip visitant": "Equip A Visitant",
                    "Club Local": "Club B",
                },
            ]
        )

        from .main_fixed import main as engine_main

        result = engine_main(
            "disp.xlsx",
            "partits.xlsx",
            run_id=1,
            config={"fase": "FS1"},
            df_dispos=df_dispos,
            df_partits=df_partits,
        )

        self.assertEqual(result["initial_assigned_matches"], 1)
        self.assertEqual(result["rescue_matches_recovered_idle"], 0)
        self.assertEqual(result["rescue_matches_recovered_reused_referees"], 1)
        self.assertEqual(result["unassigned_matches"], 0)

    @patch("designacions.main_fixed.mapa_assignacions_interactiu", side_effect=lambda **kwargs: kwargs["out_html"])
    @patch("designacions.main_fixed.persist_assignacions_to_db")
    @patch("designacions.main_fixed.AddressCluster.objects.update_or_create")
    @patch("designacions.main_fixed.Address.objects.filter")
    @patch("designacions.main_fixed.clusteritza_i_plota")
    @patch("designacions.main_fixed.addresses_to_df")
    @patch("designacions.main_fixed.geocodifica_adreces")
    @patch("designacions.main_fixed.load_modalitat_map_df")
    @patch("designacions.main_fixed.fetch_ceeb_classification_async")
    def test_main_reports_reason_for_remaining_unassigned_matches(
        self,
        fetch_classification_mock,
        load_modalitat_map_df_mock,
        geocodifica_adreces_mock,
        addresses_to_df_mock,
        clusteritza_i_plota_mock,
        address_filter_mock,
        _address_cluster_update_mock,
        _persist_mock,
        _map_mock,
    ):
        fetch_classification_mock.return_value = SimpleNamespace(
            root=ET.fromstring(
                """
                <root>
                  <grup_classificacions>
                    <info_lliga><nomGrup>GRUP A</nomGrup></info_lliga>
                    <prt_class_all>
                      <equip><NomEquipMostrar>Equip A Local</NomEquipMostrar></equip>
                      <equip><NomEquipMostrar>Equip A Visitant</NomEquipMostrar></equip>
                    </prt_class_all>
                  </grup_classificacions>
                </root>
                """
            ),
            from_cache=False,
            attempts=1,
            error=None,
        )
        load_modalitat_map_df_mock.return_value = pd.DataFrame(
            [{"Id Categoria": "9.0", "Modalitat": "Futbol Sala", "Nom": "CADET"}]
        )
        geocodifica_adreces_mock.return_value = ["ok"]
        addresses_to_df_mock.return_value = pd.DataFrame(
            [{"adreca": "Carrer 1, Barcelona", "lat": 41.0, "lon": 2.0}]
        )
        clusteritza_i_plota_mock.return_value = (
            pd.DataFrame(
                [{"adreca": "Carrer 1, Barcelona", "cluster": 1, "lat": 41.0, "lon": 2.0}]
            ),
            None,
            None,
            None,
        )
        address_filter_mock.return_value.first.return_value = None

        df_dispos = pd.DataFrame(
            [
                {
                    "Codi Tutor de Joc": "5001 F5",
                    "Nom": "Tutor A",
                    "Cognoms": "Alpha",
                    "Categoria": "TUTOR/TUTORA DE JOC",
                    "Modalitat": "Futbol Sala",
                    "Nivell": "NIVELLA1",
                    "Data": "2026-03-01",
                    "Hora Inici": "17:00:00",
                    "Hora Fi": "18:30:00",
                },
            ]
        )
        df_partits = pd.DataFrame(
            [
                {
                    "Codi": "M-A",
                    "Codi Extern Local": "EXT-A",
                    "Lliga": "Lliga A",
                    "Modalitat": "Futbol Sala",
                    "Categoria": "CADET",
                    "Subcategoria": "MIXT",
                    "Data": "2026-03-01",
                    "Hora": "20:00:00",
                    "Domicili": "Carrer 1",
                    "Municipi": "Barcelona",
                    "Pista joc": "Pista 1",
                    "Grup": "GRUP A",
                    "Equip local": "Equip A Local",
                    "Equip visitant": "Equip A Visitant",
                    "Club Local": "Club A",
                },
            ]
        )

        from .main_fixed import main as engine_main

        result = engine_main(
            "disp.xlsx",
            "partits.xlsx",
            run_id=1,
            config={"fase": "FS1"},
            df_dispos=df_dispos,
            df_partits=df_partits,
        )

        self.assertEqual(result["unassigned_matches"], 1)
        self.assertEqual(result["remaining_unassigned_matches"], 1)
        self.assertEqual(result["remaining_unassigned_breakdown"], {"outside_availability_window": 1})
        self.assertEqual(result["remaining_unassigned_details"][0]["match_code"], "M-A")


class DesignacionsTaskStatusViewTests(TestCase):
    @patch("designacions.views.read_job_sync")
    def test_task_status_view_returns_current_monotonic_progress(self, read_job_sync_mock):
        read_job_sync_mock.return_value = {
            "status": "processing",
            "progress": 72,
            "message": "Assignant tutors",
        }

        response = self.client.get(reverse("designacions_task_status", args=["task-123"]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json(),
            {
                "status": "processing",
                "progress": 72,
                "message": "Assignant tutors",
                "error": None,
            },
        )

    @patch("designacions.views.read_job_sync")
    def test_task_status_view_falls_back_to_done_run_when_redis_is_empty(self, read_job_sync_mock):
        run = DesignationRun.objects.create(task_id="task-done-fallback", status="done")
        read_job_sync_mock.return_value = {}

        response = self.client.get(reverse("designacions_task_status", args=[run.task_id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json(),
            {
                "status": "SUCCESS",
                "progress": 100,
                "message": "Procés finalitzat.",
                "error": None,
            },
        )

    @patch("designacions.views.read_job_sync")
    def test_task_status_view_falls_back_to_failed_run_when_redis_is_empty(self, read_job_sync_mock):
        run = DesignationRun.objects.create(task_id="task-failed-fallback", status="failed", error="boom")
        read_job_sync_mock.return_value = {}

        response = self.client.get(reverse("designacions_task_status", args=[run.task_id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json(),
            {
                "status": "FAILURE",
                "progress": None,
                "message": None,
                "error": "boom",
            },
        )

    @patch("designacions.views.read_job_sync")
    def test_task_status_view_prefers_terminal_run_over_stale_processing_job(self, read_job_sync_mock):
        run = DesignationRun.objects.create(task_id="task-stale-job", status="done")
        read_job_sync_mock.return_value = {
            "status": "processing",
            "progress": 72,
            "message": "Encara corrent",
        }

        response = self.client.get(reverse("designacions_task_status", args=[run.task_id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "SUCCESS")
        self.assertEqual(response.json()["progress"], 100)


class DesignacionsRunDetailViewTests(TestCase):
    def test_run_detail_view_renders_assignments_button_for_completed_run(self):
        run = DesignationRun.objects.create(task_id="task-run-detail-done", status="done")

        response = self.client.get(reverse("designacions_run_detail", args=[run.id]))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("Veure i editar assignacions", content)
        self.assertIn(f"/designacions/run/{run.id}/assignments/", content)

    def test_run_detail_view_renders_error_notice_for_failed_run(self):
        run = DesignationRun.objects.create(task_id="task-run-detail-failed", status="failed", error="boom")

        response = self.client.get(reverse("designacions_run_detail", args=[run.id]))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("El run ha finalitzat amb error.", content)
